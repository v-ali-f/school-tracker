from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.extensions import db
from app.models import (
    AcademicYear,
    Building,
    Child,
    ChildEnrollment,
    EducationActivity,
    EducationPlan,
    EducationPlanBinding,
    EducationPlanLine,
    EducationPlanLinePeriod,
    EducationPlanLineScope,
    PopulationSnapshot,
    PopulationSnapshotClass,
    PopulationSnapshotEnrollment,
    SchoolClass,
    TariffVersion,
    TeachingGroup,
    TeachingGroupClass,
    TeachingGroupCompositionApproval,
    TeachingGroupHistory,
    TeachingGroupMember,
    TeachingMetagroupSource,
    User,
    WorkloadAssignment,
    WorkloadAssignmentChange,
    WorkloadNeed,
)
from app.services.education_plan_service import (
    create_plan_bundle,
    ensure_draft_tariff_version,
    line_scope_key,
    plan_scope_code,
)
from app.services.education_plan_binding_service import (
    assign_class_plan,
    class_level_plan_ids,
    class_plan_allocations,
    replace_plan_binding_members,
)
from app.services.class_plan_matrix_service import (
    build_class_plan_matrix,
    class_period_label,
    effective_line_weekly_hours,
)
from app.services.teaching_group_service import (
    GroupValidationError,
    build_population_snapshot,
    change_group_status,
    group_coverage,
    population_registry_status,
    validate_group_sources,
)
from app.services.teaching_metagroup_service import (
    build_metagroup_workspace,
    create_metagroup,
)
from app.services.teaching_group_matrix_service import (
    build_group_composition_workspace,
    build_teaching_group_matrix,
    materialize_default_teaching_groups,
    replace_group_composition_assignments,
    replace_teaching_group_count,
)
from app.services.workload_assignment_matrix_service import (
    build_workload_assignment_matrix,
)
from app.services.workload_distribution_service import generate_plan_needs
from app.services.workload_snapshot_service import (
    relink_assignments_to_population_snapshot,
)
from app.services.workload_editing_workflow_service import (
    WorkloadEditingWorkflowError,
    change_groups_editing_status,
    require_groups_editable,
)


def _child(last_name, first_name):
    child = Child(last_name=last_name, first_name=first_name)
    db.session.add(child)
    db.session.flush()
    return child


def test_group_changes_can_be_saved_and_reopened(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        version = db.session.get(TariffVersion, context["version_id"])

        change_groups_editing_status(
            version,
            "SAVE",
            user_id=user_id,
        )
        assert version.groups_editing_status == "SAVED"
        with pytest.raises(WorkloadEditingWorkflowError):
            require_groups_editable(version)

        change_groups_editing_status(
            version,
            "EDIT",
            user_id=user_id,
        )
        assert version.groups_editing_status == "EDITING"
        require_groups_editable(version)


def test_building_registry_updates_matrix_tone(
    app,
    client,
    make_user,
    login,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        building = Building(
            name="Новый корпус",
            address="Тестовый адрес",
        )
        db.session.add(building)
        db.session.commit()
        building_id = building.id

    login(user_id)
    registry = client.get("/buildings")
    assert registry.status_code == 200
    assert "Цвет в таблицах".encode() in registry.data
    assert "Голубой".encode() in registry.data

    response = client.post(
        f"/buildings/{building_id}/update",
        data={
            "name": "Новый корпус",
            "address": "Тестовый адрес",
            "matrix_tone": "1",
        },
    )
    assert response.status_code == 302
    with app.app_context():
        assert db.session.get(Building, building_id).matrix_tone == 1


def _group_context(user_id):
    year = AcademicYear(
        name="2026/2027",
        is_current=True,
        start_date=date(2026, 9, 1),
        end_date=date(2027, 8, 31),
    )
    building = Building(name="Главное здание", short_name="ГЗ")
    db.session.add_all([year, building])
    db.session.flush()

    class_a = SchoolClass(
        academic_year_id=year.id,
        building_id=building.id,
        name="5А",
        grade=5,
        letter="А",
    )
    class_b = SchoolClass(
        academic_year_id=year.id,
        building_id=building.id,
        name="5Б",
        grade=5,
        letter="Б",
    )
    db.session.add_all([class_a, class_b])
    db.session.flush()

    children = [
        _child("Иванов", "Иван"),
        _child("Петрова", "Анна"),
        _child("Сидоров", "Олег"),
    ]
    for child, school_class in (
        (children[0], class_a),
        (children[1], class_a),
        (children[2], class_b),
    ):
        db.session.add(ChildEnrollment(
            child_id=child.id,
            academic_year_id=year.id,
            school_class_id=school_class.id,
            status="ACTIVE",
            enrolled_at=datetime(2026, 9, 1),
        ))

    activity = EducationActivity(
        code="MATH",
        name="Математика",
        activity_kind="SUBJECT",
        is_global=True,
        is_tariffable=True,
        is_active=True,
    )
    db.session.add(activity)
    db.session.flush()
    _, version = ensure_draft_tariff_version(year, user_id=user_id)
    plan = EducationPlan(
        tariff_version_id=version.id,
        plan_kind="CURRICULUM",
        name="Основной учебный план",
        education_level="OOO",
        building_id=building.id,
        scope_code=plan_scope_code("OOO", building.id),
        status="DRAFT",
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(plan)
    db.session.flush()
    line = EducationPlanLine(
        education_plan_id=plan.id,
        education_activity_id=activity.id,
        component_kind="MANDATORY",
        weekly_hours=Decimal("5"),
        weeks_count=Decimal("34"),
        annual_hours=Decimal("170"),
        requires_division=True,
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(line)
    db.session.flush()
    db.session.add(EducationPlanLineScope(
        education_plan_line_id=line.id,
        scope_kind="GRADE",
        grade=5,
        building_id=building.id,
        scope_key=line_scope_key(
            "GRADE",
            grade=5,
            building_id=building.id,
        ),
    ))
    db.session.commit()
    return {
        "year_id": year.id,
        "building_id": building.id,
        "version_id": version.id,
        "plan_line_id": line.id,
    }


def _snapshot(user_id, version_id):
    version = db.session.get(TariffVersion, version_id)
    snapshot = build_population_snapshot(
        version,
        user_id=user_id,
        snapshot_date=date(2026, 9, 1),
    )
    db.session.commit()
    return snapshot.id


def _group_form(context, class_ids, member_ids=(), **overrides):
    data = {
        "source_plan_line_id": str(context["plan_line_id"]),
        "group_type": "SUBGROUP",
        "code": "MATH_5A_1",
        "name": "Математика 5А, группа 1",
        "composition_mode": "PERSONAL",
        "planned_size": "1",
        "valid_from": "2026-09-01",
        "valid_to": "2027-05-31",
        "building_id": str(context["building_id"]),
        "snapshot_class_ids": [str(item) for item in class_ids],
        "snapshot_enrollment_ids": [str(item) for item in member_ids],
    }
    data.update(overrides)
    return data


def _ready_source_group(
    context,
    snapshot_class,
    *,
    user_id,
    suffix,
    plan_line_id=None,
    member_ids=None,
):
    line = db.session.get(
        EducationPlanLine,
        plan_line_id or context["plan_line_id"],
    )
    member_ids = (
        [item.id for item in snapshot_class.enrollments]
        if member_ids is None else list(member_ids)
    )
    group = TeachingGroup(
        tariff_version_id=context["version_id"],
        education_activity_id=line.education_activity_id,
        group_type="CLASS",
        code=f"META_SOURCE_{suffix}",
        name=f"{snapshot_class.name_snapshot} · Математика",
        composition_mode="PERSONAL",
        building_id=context["building_id"],
        planned_size=len(member_ids),
        actual_size=len(member_ids),
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 8, 31),
        source_plan_line_id=line.id,
        status="READY",
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(group)
    db.session.flush()
    db.session.add(TeachingGroupClass(
        teaching_group_id=group.id,
        population_snapshot_class_id=snapshot_class.id,
        relation_kind="FULL",
        student_count=len(member_ids),
    ))
    for member_id in member_ids:
        db.session.add(TeachingGroupMember(
            teaching_group_id=group.id,
            snapshot_enrollment_id=member_id,
            valid_from=group.valid_from,
            valid_to=group.valid_to,
            source_kind="AUTO",
        ))
    db.session.flush()
    return group


def test_population_snapshot_preserves_classes_and_students(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])

        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        assert snapshot.revision_no == 1
        assert len(snapshot.classes) == 2
        assert PopulationSnapshotEnrollment.query.count() == 3
        assert sorted(item.student_count for item in snapshot.classes) == [1, 2]


def test_snapshot_refresh_reuses_current_revision_when_registry_is_unchanged(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        first_id = _snapshot(user_id, context["version_id"])
        second_id = _snapshot(user_id, context["version_id"])

        first = db.session.get(PopulationSnapshot, first_id)
        second = db.session.get(PopulationSnapshot, second_id)
        assert first_id == second_id
        assert first.status == "CURRENT"
        assert second.status == "CURRENT"
        assert second.revision_no == 1


def test_class_students_can_be_split_between_two_curricula(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        snapshot_class = next(
            item
            for item in snapshot.classes
            if item.name_snapshot == "5А"
        )
        first_plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="Профильный учебный план",
            education_level="OOO",
            building_id=context["building_id"],
            scope_code="OOO_PROFILE",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_plan)
        db.session.flush()
        enrollment_ids = {
            item.id for item in snapshot_class.enrollments
        }

        replace_plan_binding_members(
            first_plan,
            snapshot_class,
            enrollment_ids,
            user_id=user_id,
        )
        moved_enrollment_id = next(iter(enrollment_ids))
        replace_plan_binding_members(
            second_plan,
            snapshot_class,
            {moved_enrollment_id},
            user_id=user_id,
        )
        db.session.commit()

        allocations, student_plan_ids = class_plan_allocations(
            snapshot_class,
            [first_plan, second_plan],
        )
        assert allocations[second_plan.id] == {moved_enrollment_id}
        assert allocations[first_plan.id] == (
            enrollment_ids - {moved_enrollment_id}
        )
        assert student_plan_ids[moved_enrollment_id] == second_plan.id
        assert EducationPlanBinding.query.count() == 2


def test_empty_class_keeps_plan_binding_and_enters_planning_matrix(
    app,
    client,
    make_user,
    login,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        db.session.get(Building, context["building_id"]).matrix_tone = 1
        version = db.session.get(TariffVersion, context["version_id"])
        year = version.tariff_cycle.academic_year
        empty_class = SchoolClass(
            academic_year_id=year.id,
            building_id=context["building_id"],
            name="5В",
            grade=5,
            letter="В",
        )
        db.session.add(empty_class)
        db.session.commit()

        first_snapshot_id = _snapshot(user_id, version.id)
        first_snapshot = db.session.get(
            PopulationSnapshot,
            first_snapshot_id,
        )
        snapshot_class = next(
            item
            for item in first_snapshot.classes
            if item.source_school_class_id == empty_class.id
        )
        plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        plan.profile_name = "Математический"
        assign_class_plan(
            snapshot_class,
            [plan],
            plan.id,
            user_id=user_id,
        )
        db.session.commit()

        assert class_level_plan_ids(snapshot_class, [plan]) == {plan.id}
        allocations, student_plan_ids = class_plan_allocations(
            snapshot_class,
            [plan],
        )
        assert allocations[plan.id] == set()
        assert student_plan_ids == {}
        matrix = build_class_plan_matrix(
            first_snapshot,
            [plan],
            "OOO",
            grade=5,
        )
        empty_group = next(
            item
            for item in matrix["class_groups"]
            if item["snapshot_class"].id == snapshot_class.id
        )
        assert len(empty_group["columns"]) == 1
        assert empty_group["columns"][0]["plan"].id == plan.id
        assert empty_group["columns"][0]["student_count"] == 0
        assert not empty_group["columns"][0]["is_unassigned"]

        child = _child("Новый", "Ученик")
        db.session.add(ChildEnrollment(
            child_id=child.id,
            academic_year_id=year.id,
            school_class_id=empty_class.id,
            status="ACTIVE",
            enrolled_at=datetime(2026, 9, 2),
        ))
        db.session.commit()
        second_snapshot = build_population_snapshot(
            version,
            user_id=user_id,
            snapshot_date=date(2026, 9, 2),
        )
        db.session.commit()
        new_snapshot_class = next(
            item
            for item in second_snapshot.classes
            if item.source_school_class_id == empty_class.id
        )
        allocations, student_plan_ids = class_plan_allocations(
            new_snapshot_class,
            [plan],
        )
        new_enrollment_id = new_snapshot_class.enrollments[0].id
        assert allocations[plan.id] == {new_enrollment_id}
        assert student_plan_ids[new_enrollment_id] == plan.id
        assert (
            EducationPlanBinding.query
            .filter_by(
                population_snapshot_class_id=new_snapshot_class.id,
            )
            .one()
            .binding_mode
            == "CLASS"
        )
        year_id = year.id

    login(user_id)
    response = client.get(f"/contingent?year_id={year_id}")
    assert response.status_code == 200
    assert "Профиль".encode() in response.data
    assert "Математический".encode() in response.data
    assert "Заяв.".encode() in response.data
    assert "В пар.".encode() in response.data
    assert "Основное общее образование".encode() in response.data
    assert "5–9 ·".encode() in response.data
    assert b'value=""' in response.data
    assert b'data-saved-value="0"' in response.data
    assert (
        b".contingent-classes-table.table tbody td"
        in response.data
    )
    assert b"height: 31px !important" in response.data
    assert b"font-size: 15px !important" in response.data
    assert b"padding: 3px 5px !important" in response.data
    assert b"building-tone-1" in response.data
    assert b"contingent-classes-scroll" in response.data
    assert b"contingent-home-action" in response.data
    assert b".contingent-page .contingent-home-action" in response.data
    assert b"col.col-teacher" in response.data
    assert b"width: 1120px !important" in response.data
    assert b"visibility: collapse" not in response.data


def test_plan_bindings_page_reads_snapshot_classes(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class_id = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
            .id
        )

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/?version_id={context['version_id']}"
    )

    assert response.status_code == 200
    assert "Привязка учебных планов".encode() in response.data
    assert b"registry-filter-panel" in response.data
    assert b"registry-matrix registry-matrix-wide" in response.data
    assert b"registry-matrix-card" in response.data
    assert "5А".encode() in response.data
    assert "Иванов Иван".encode() not in response.data

    response = client.get(
        f"/workload/plan-bindings/?version_id={context['version_id']}"
        f"&class_id={snapshot_class_id}"
    )
    assert response.status_code == 200
    assert "Иванов Иван".encode() in response.data


def test_plan_bindings_page_filters_classes_by_level_and_grade(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        _snapshot(user_id, context["version_id"])
        version_id = context["version_id"]

    login(user_id)
    grade_response = client.get(
        "/workload/plan-bindings/",
        query_string={
            "version_id": version_id,
            "level": "OOO",
            "grade": 5,
        },
    )
    grade_html = grade_response.get_data(as_text=True)
    assert grade_response.status_code == 200
    assert 'name="level"' in grade_html
    assert 'value="OOO" selected' in grade_html
    assert 'name="grade"' in grade_html
    assert 'value="5" selected' in grade_html
    assert "5А" in grade_html

    other_level = client.get(
        "/workload/plan-bindings/",
        query_string={
            "version_id": version_id,
            "level": "NOO",
        },
    )
    other_html = other_level.get_data(as_text=True)
    assert other_level.status_code == 200
    assert 'value="NOO" selected' in other_html
    assert "5А" not in other_html


def test_senior_class_can_select_two_plans_before_students_are_added(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        source_class = SchoolClass.query.filter_by(name="5А").one()
        ChildEnrollment.query.filter_by(
            school_class_id=source_class.id,
        ).delete(synchronize_session=False)
        source_class.name = "10А"
        source_class.grade = 10
        source_class.letter = "А"

        first_plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        first_plan.education_level = "SOO"
        first_plan.scope_code = "SOO_MAIN"
        first_plan.profile_name = "Инженерный"
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="Технологический профиль",
            education_level="SOO",
            building_id=context["building_id"],
            scope_code="SOO_TECH",
            profile_name="Предпринимательский",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_plan)
        db.session.commit()

        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="10А",
            )
            .one()
        )
        class_id = snapshot_class.id
        first_plan_id = first_plan.id
        second_plan_id = second_plan.id
        version_id = context["version_id"]

    login(user_id)
    response = client.post(
        "/workload/plan-bindings/class-plans",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "plan_ids": [first_plan_id, second_plan_id],
            "level": "SOO",
            "grade": "10",
        },
    )
    assert response.status_code == 302

    with app.app_context():
        bindings = (
            EducationPlanBinding.query
            .filter_by(population_snapshot_class_id=class_id)
            .order_by(EducationPlanBinding.education_plan_id)
            .all()
        )
        assert len(bindings) == 2
        assert {item.binding_mode for item in bindings} == {"PLAN_SET"}
        assert all(not item.members for item in bindings)

        matrix = build_class_plan_matrix(
            db.session.get(
                PopulationSnapshotClass,
                class_id,
            ).population_snapshot,
            [
                db.session.get(EducationPlan, first_plan_id),
                db.session.get(EducationPlan, second_plan_id),
            ],
            "SOO",
            grade=10,
        )
        profile_group = next(
            item
            for item in matrix["class_groups"]
            if item["snapshot_class"].id == class_id
        )
        assert profile_group["split_profile_columns"]
        assert {
            item["class_display_name"]
            for item in profile_group["columns"]
        } == {"10А Инженерный", "10А Предпринимательский"}
        assert all(
            not item["is_unassigned"]
            for item in profile_group["columns"]
        )
        workload_matrix = build_workload_assignment_matrix(
            [],
            [],
            plan_matrices=[matrix],
        )
        assert len(workload_matrix["class_groups"]) == 1
        assert workload_matrix["class_groups"][0]["label"] == "10А"
        assert {
            item["subheader_label"]
            for item in workload_matrix["class_groups"][0]["columns"]
        } == {"Инженерный", "Предпринимательский"}

        child = Child.query.order_by(Child.id.asc()).first()
        enrollment = PopulationSnapshotEnrollment(
            population_snapshot_class_id=class_id,
            source_child_id=child.id,
            fio_snapshot="Иванов Иван",
            status_snapshot="ACTIVE",
        )
        db.session.add(enrollment)
        db.session.get(
            PopulationSnapshotClass,
            class_id,
        ).student_count = 1
        db.session.commit()
        enrollment_id = enrollment.id

    page = client.get(response.headers["Location"])
    assert page.status_code == 200
    assert "Основной учебный план".encode() in page.data
    assert "Технологический профиль".encode() in page.data

    matrix_page = client.get(
        "/workload/plan-bindings/matrix",
        query_string={
            "version_id": version_id,
            "level": "SOO",
            "grade": 10,
        },
    )
    assert matrix_page.status_code == 200
    assert 'data-class-name="10А Инженерный"'.encode() in matrix_page.data
    assert (
        'data-class-name="10А Предпринимательский"'.encode()
        in matrix_page.data
    )

    assigned = client.post(
        "/workload/plan-bindings/student",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "enrollment_id": enrollment_id,
            "plan_id": second_plan_id,
            "level": "SOO",
            "grade": "10",
        },
    )
    assert assigned.status_code == 302

    with app.app_context():
        bindings = (
            EducationPlanBinding.query
            .filter_by(population_snapshot_class_id=class_id)
            .all()
        )
        assert len(bindings) == 2
        assert {item.binding_mode for item in bindings} == {"PLAN_SET"}
        allocations, student_plan_ids = class_plan_allocations(
            db.session.get(PopulationSnapshotClass, class_id),
            [
                db.session.get(EducationPlan, first_plan_id),
                db.session.get(EducationPlan, second_plan_id),
            ],
        )
        assert allocations[first_plan_id] == set()
        assert allocations[second_plan_id] == {enrollment_id}
        assert student_plan_ids[enrollment_id] == second_plan_id

    contingent_page = client.get(
        f"/contingent?year_id={context['year_id']}"
    )
    assert contingent_page.status_code == 200
    assert "Инженерный / Предпринимательский".encode() in (
        contingent_page.data
    )


def test_class_plan_matrix_uses_assigned_plan_hours(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        db.session.get(Building, context["building_id"]).matrix_tone = 1
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        replace_plan_binding_members(
            plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        version_id = context["version_id"]

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO"
    )

    assert response.status_code == 200
    assert "Свод учебных планов по классам".encode() in response.data
    assert 'data-class-name="5А"'.encode() in response.data
    assert 'data-plan-name="Основной учебный план"'.encode() in response.data
    assert 'class="class-plan-matrix class-plan-summary-matrix"'.encode() in response.data
    assert 'data-activity-name="Математика"'.encode() in response.data
    assert "ч/нед.".encode() in response.data
    assert "Без УП".encode() in response.data
    assert "Свод по классам".encode() in response.data
    assert "Excel".encode() in response.data
    assert "PDF".encode() in response.data
    assert 'aria-label="Параллель"'.encode() in response.data
    assert "Всего по учебному плану".encode() in response.data
    assert 'class="class-plan-matrix__section-label"'.encode() in response.data
    assert 'class="class-plan-matrix__section-band"'.encode() in response.data
    assert "building-tone-1".encode() in response.data
    assert "class-plan-matrix__empty-value".encode() not in response.data

    building_response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO"
        f"&building_id={context['building_id']}"
    )
    assert building_response.status_code == 200
    assert "building-tone-".encode() not in building_response.data

    grade_response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO&grade=5"
    )
    grade_html = grade_response.get_data(as_text=True)
    assert grade_response.status_code == 200
    assert 'data-class-name="5А"' in grade_html
    assert "Математика" in grade_html

    empty_grade_response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO&grade=6"
    )
    empty_grade_html = empty_grade_response.get_data(as_text=True)
    assert empty_grade_response.status_code == 200
    assert 'data-class-name="5А"' not in empty_grade_html
    assert "Для выбранных фильтров" in empty_grade_html

    xlsx_response = client.get(
        f"/workload/plan-bindings/matrix/export.xlsx"
        f"?version_id={version_id}&level=OOO"
    )
    assert xlsx_response.status_code == 200
    assert xlsx_response.data.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx_response.data), data_only=True)
    sheet = workbook["OOO"]
    values = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "Свод учебных планов по классам" in values
    assert "Математика" in values
    assert "Всего по учебному плану" in values
    subtotal_row = next(
        row[0].row
        for row in sheet.iter_rows()
        if row[0].value == "Итого по разделу"
    )
    assert any(
        cell.value is None
        for cell in sheet[subtotal_row][1:]
    )

    pdf_response = client.get(
        f"/workload/plan-bindings/matrix/export.pdf"
        f"?version_id={version_id}&level=OOO"
    )
    assert pdf_response.status_code == 200
    assert pdf_response.data.startswith(b"%PDF")


def test_class_plan_matrix_splits_class_between_two_plans(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        first_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        first_plan = first_line.education_plan
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="Профильный учебный план",
            education_level="OOO",
            building_id=context["building_id"],
            scope_code="OOO_PROFILE_MATRIX",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_plan)
        db.session.flush()
        second_line = EducationPlanLine(
            education_plan_id=second_plan.id,
            education_activity_id=first_line.education_activity_id,
            component_kind="MANDATORY",
            weekly_hours=Decimal("7"),
            weeks_count=Decimal("34"),
            annual_hours=Decimal("238"),
            sort_order=10,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_line)
        db.session.flush()
        db.session.add(EducationPlanLineScope(
            education_plan_line_id=second_line.id,
            scope_kind="GRADE",
            grade=5,
            building_id=context["building_id"],
            scope_key=line_scope_key(
                "GRADE",
                grade=5,
                building_id=context["building_id"],
            ),
        ))
        enrollment_ids = [
            item.id for item in snapshot_class.enrollments
        ]
        replace_plan_binding_members(
            first_plan,
            snapshot_class,
            set(enrollment_ids),
            user_id=user_id,
        )
        replace_plan_binding_members(
            second_plan,
            snapshot_class,
            {enrollment_ids[0]},
            user_id=user_id,
        )
        db.session.commit()
        version_id = context["version_id"]

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO"
    )

    assert response.status_code == 200
    assert 'data-class-name="5А" colspan="2"'.encode() in response.data
    assert 'data-plan-name="Основной учебный план"'.encode() in response.data
    assert 'data-plan-name="Профильный учебный план"'.encode() in response.data
    assert response.data.count("1 уч. · ч/нед.".encode()) >= 2


def test_planning_matrices_filter_snapshot_classes_by_building(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        second_building = Building(
            name="Второй учебный корпус",
            short_name="Корпус Б",
        )
        db.session.add(second_building)
        db.session.flush()
        class_b = SchoolClass.query.filter_by(
            academic_year_id=context["year_id"],
            name="5Б",
        ).one()
        class_b.building_id = second_building.id
        db.session.commit()

        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class_a = PopulationSnapshotClass.query.filter_by(
            population_snapshot_id=snapshot_id,
            name_snapshot="5А",
        ).one()
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class_a,
            {item.id for item in snapshot_class_a.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        version_id = context["version_id"]
        main_building_id = context["building_id"]
        second_building_id = second_building.id

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO"
        f"&building_id={main_building_id}"
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert 'aria-label="Здание"' in html
    assert "ГЗ" in html
    assert "Корпус Б" in html
    assert 'data-class-name="5А"' in html
    assert 'data-class-name="5Б"' not in html

    response = client.get(
        f"/workload/groups/"
        f"?version_id={version_id}&level=OOO"
        f"&building_id={second_building_id}"
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert 'value="{}" selected'.format(second_building_id) in html
    assert "5Б" in html
    assert "5А" not in html

    for path in ("composition/", "metagroups/"):
        response = client.get(
            f"/workload/groups/{path}"
            f"?version_id={version_id}&level=OOO"
            f"&building_id={second_building_id}"
        )
        assert response.status_code == 200
        assert (
            'value="{}" selected'.format(second_building_id)
            in response.get_data(as_text=True)
        )


def test_group_matrix_uses_one_as_default_for_existing_plan_cells(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        db.session.get(Building, context["building_id"]).matrix_tone = 1
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        version_id = context["version_id"]

    login(user_id)
    response = client.get(
        f"/workload/groups/?version_id={version_id}&level=OOO"
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Количество учебных групп" in html
    assert 'data-activity-name="Математика"' in html
    assert 'data-original-value="1"' in html
    assert "Без УП" in html
    assert 'aria-label="Параллель"' in html
    assert 'class="workload-indicators"' not in html
    assert 'class="class-plan-matrix__section-label"' in html
    assert 'class="class-plan-matrix__section-band"' in html
    assert "group-matrix__cell--no-plan" in html
    assert "Учебный план не назначен классу" in html
    assert "building-tone-1" in html
    assert html.count("data-group-count") == 1
    assert "Основной учебный план" in html
    assert 'class="group-matrix__plan-short"' in html
    assert "Excel" in html
    assert "PDF" in html

    xlsx_response = client.get(
        f"/workload/groups/export.xlsx"
        f"?version_id={version_id}&level=OOO"
    )
    assert xlsx_response.status_code == 200
    assert xlsx_response.data.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx_response.data), data_only=True)
    sheet = workbook["OOO"]
    exported_values = {
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Количество учебных групп" in exported_values
    assert "Математика" in exported_values
    assert 1 in exported_values

    pdf_response = client.get(
        f"/workload/groups/export.pdf"
        f"?version_id={version_id}&level=OOO"
    )
    assert pdf_response.status_code == 200
    assert pdf_response.data.startswith(b"%PDF")

    grade_response = client.get(
        f"/workload/groups/"
        f"?version_id={version_id}&level=OOO&grade=5"
    )
    grade_html = grade_response.get_data(as_text=True)
    assert grade_response.status_code == 200
    assert "5А" in grade_html
    assert 'data-activity-name="Математика"' in grade_html

    empty_grade_response = client.get(
        f"/workload/groups/"
        f"?version_id={version_id}&level=OOO&grade=6"
    )
    empty_grade_html = empty_grade_response.get_data(as_text=True)
    assert empty_grade_response.status_code == 200
    assert 'data-activity-name="Математика"' not in empty_grade_html
    assert "Для выбранных фильтров нет классов." in empty_grade_html
    with app.app_context():
        assert TeachingGroup.query.count() == 0


def test_class_plan_matrix_section_and_total_labels_are_sticky(
    app,
    client,
):
    response = client.get(
        "/static/css/workload_class_plan_matrix.css",
    )
    css = response.get_data(as_text=True)

    assert response.status_code == 200
    sticky_rule_start = css.index(
        ".class-plan-matrix__subject,"
    )
    sticky_rule_end = css.index("}", sticky_rule_start)
    sticky_rule = css[sticky_rule_start:sticky_rule_end]
    assert ".class-plan-matrix__section-label" in sticky_rule
    assert ".class-plan-matrix__curriculum-total th" in sticky_rule
    assert "position: sticky" in sticky_rule
    assert "left: 0" in sticky_rule


def test_group_matrix_locks_zero_hour_subject_cells(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = {
            item.name_snapshot: item
            for item in snapshot.classes
        }
        first_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        first_plan = first_line.education_plan
        english = EducationActivity(
            code="ENGLISH_ZERO_MATRIX",
            name="Английский язык",
            activity_kind="SUBJECT",
            is_global=True,
            is_tariffable=True,
            is_active=True,
        )
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="План 5Б",
            education_level="OOO",
            building_id=context["building_id"],
            scope_code="OOO_5B_ZERO_MATRIX",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add_all([english, second_plan])
        db.session.flush()

        zero_line = EducationPlanLine(
            education_plan_id=first_plan.id,
            education_activity_id=english.id,
            component_kind="MANDATORY",
            weekly_hours=Decimal("0"),
            weeks_count=Decimal("34"),
            annual_hours=Decimal("0"),
            sort_order=20,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        positive_line = EducationPlanLine(
            education_plan_id=second_plan.id,
            education_activity_id=english.id,
            component_kind="MANDATORY",
            weekly_hours=Decimal("3"),
            weeks_count=Decimal("34"),
            annual_hours=Decimal("102"),
            sort_order=20,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add_all([zero_line, positive_line])
        db.session.flush()
        for line in (zero_line, positive_line):
            db.session.add(EducationPlanLineScope(
                education_plan_line_id=line.id,
                scope_kind="GRADE",
                grade=5,
                building_id=context["building_id"],
                scope_key=line_scope_key(
                    "GRADE",
                    grade=5,
                    building_id=context["building_id"],
                ),
            ))

        replace_plan_binding_members(
            first_plan,
            classes["5А"],
            {
                enrollment.id
                for enrollment in classes["5А"].enrollments
            },
            user_id=user_id,
        )
        replace_plan_binding_members(
            second_plan,
            classes["5Б"],
            {
                enrollment.id
                for enrollment in classes["5Б"].enrollments
            },
            user_id=user_id,
        )
        db.session.flush()
        created_count = materialize_default_teaching_groups(
            version=db.session.get(TariffVersion, context["version_id"]),
            snapshot=snapshot,
            plans=[first_plan, second_plan],
            user_id=user_id,
        )
        db.session.commit()
        zero_line_id = zero_line.id
        positive_line_id = positive_line.id
        class_5a_id = classes["5А"].id
        first_plan_id = first_plan.id
        version_id = context["version_id"]

        assert created_count == 2
        assert TeachingGroup.query.filter_by(
            source_plan_line_id=zero_line_id,
        ).count() == 0
        assert TeachingGroup.query.filter_by(
            source_plan_line_id=positive_line_id,
        ).count() == 1

    login(user_id)
    response = client.get(
        f"/workload/groups/?version_id={version_id}&level=OOO&grade=5"
    )
    html = response.get_data(as_text=True)
    row_start = html.index('data-activity-name="Английский язык"')
    row_end = html.index("</tr>", row_start)
    english_row = html[row_start:row_end]

    assert response.status_code == 200
    assert english_row.count("data-group-count") == 1
    assert "group-matrix__cell--not-in-plan" in english_row
    assert "Предмет отсутствует в учебном плане класса" in english_row
    assert f'data-plan-line-id="{positive_line_id}"' in english_row
    assert f'data-plan-line-id="{zero_line_id}"' not in english_row
    assert ">—<" not in english_row

    rejected = client.post(
        "/workload/groups/matrix/cell",
        data={
            "version_id": version_id,
            "plan_line_id": zero_line_id,
            "snapshot_class_id": class_5a_id,
            "plan_id": first_plan_id,
            "group_count": 1,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert rejected.status_code == 422
    assert "Предмет отсутствует" in rejected.get_json()["message"]


def test_group_matrix_creates_split_groups_and_restores_whole_class(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": snapshot_class.id,
            "plan_id": line.education_plan_id,
        }

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data={**payload, "group_count": 2},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    assert response.get_json()["needs_composition"] is True
    with app.app_context():
        groups = TeachingGroup.query.order_by(TeachingGroup.id.asc()).all()
        assert len(groups) == 2
        assert {group.group_type for group in groups} == {"SUBGROUP"}
        assert {group.status for group in groups} == {"DRAFT"}
        assert all(group.actual_size == 0 for group in groups)
        assert all(not group.members for group in groups)

    matrix_response = client.get(
        f"/workload/groups/"
        f"?version_id={context['version_id']}&level=OOO&grade=5"
    )
    matrix_html = matrix_response.get_data(as_text=True)
    assert matrix_response.status_code == 200
    assert "is-divided" in matrix_html
    assert "needs-composition" in matrix_html

    response = client.post(
        "/workload/groups/matrix/cell",
        data={**payload, "group_count": 1},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    assert response.get_json()["needs_composition"] is False
    with app.app_context():
        group = TeachingGroup.query.one()
        assert group.group_type == "CLASS"
        assert group.status == "READY"
        assert group.actual_size == 2
        assert len(group.members) == 2


def test_group_matrix_saves_multiple_cells_in_one_request(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        classes = sorted(snapshot.classes, key=lambda item: item.name_snapshot)
        for snapshot_class in classes:
            replace_plan_binding_members(
                line.education_plan,
                snapshot_class,
                {item.id for item in snapshot_class.enrollments},
                user_id=user_id,
            )
        db.session.commit()
        changes = [
            {
                "plan_line_id": line.id,
                "snapshot_class_id": snapshot_class.id,
                "plan_id": line.education_plan_id,
                "group_count": 2,
            }
            for snapshot_class in classes
        ]

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cells",
        json={
            "version_id": context["version_id"],
            "changes": changes,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["ok"] is True
    assert payload["message"] == "Сохранено изменений: 2."
    assert len(payload["results"]) == 2
    assert {item["group_count"] for item in payload["results"]} == {2}
    with app.app_context():
        assert TeachingGroup.query.count() == 4


def test_unchanged_population_refresh_preserves_split_group_count(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        first_snapshot_id = _snapshot(user_id, context["version_id"])
        version = db.session.get(TariffVersion, context["version_id"])
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        first_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=first_snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        replace_plan_binding_members(
            line.education_plan,
            first_class,
            {item.id for item in first_class.enrollments},
            user_id=user_id,
        )
        replace_teaching_group_count(
            version=version,
            snapshot=db.session.get(PopulationSnapshot, first_snapshot_id),
            plans=[line.education_plan],
            plan_line_id=line.id,
            snapshot_class_id=first_class.id,
            plan_id=line.education_plan_id,
            group_count=2,
            user_id=user_id,
        )
        db.session.commit()

        second_snapshot = build_population_snapshot(
            version,
            user_id=user_id,
            snapshot_date=date(2026, 9, 2),
        )
        materialize_default_teaching_groups(
            version=version,
            snapshot=second_snapshot,
            plans=[line.education_plan],
            user_id=user_id,
        )
        db.session.commit()

        second_class = next(
            item
            for item in second_snapshot.classes
            if item.name_snapshot == "5А"
        )
        current_groups = (
            TeachingGroup.query
            .join(TeachingGroupClass)
            .filter(
                TeachingGroup.source_plan_line_id == line.id,
                TeachingGroup.status != "CLOSED",
                TeachingGroupClass.population_snapshot_class_id
                == second_class.id,
            )
            .all()
        )
        assert len(current_groups) == 2


def test_adding_plan_subject_only_opens_its_workload_cells(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        version = db.session.get(TariffVersion, context["version_id"])
        original_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        plan = original_line.education_plan
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        snapshot_class = next(
            item
            for item in snapshot.classes
            if item.name_snapshot == "5А"
        )
        replace_plan_binding_members(
            plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        original_groups = replace_teaching_group_count(
            version=version,
            snapshot=snapshot,
            plans=[plan],
            plan_line_id=original_line.id,
            snapshot_class_id=snapshot_class.id,
            plan_id=plan.id,
            group_count=2,
            user_id=user_id,
        )
        original_group_ids = {item.id for item in original_groups}

        activity = EducationActivity(
            code="NEW-PLAN-SUBJECT",
            name="Новый предмет",
            activity_kind="SUBJECT",
            is_global=True,
            is_tariffable=True,
            is_active=True,
        )
        db.session.add(activity)
        db.session.flush()
        new_line = EducationPlanLine(
            education_plan_id=plan.id,
            education_activity_id=activity.id,
            component_kind="MANDATORY",
            weekly_hours=Decimal("2"),
            weeks_count=Decimal("34"),
            annual_hours=Decimal("68"),
            requires_division=False,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(new_line)
        db.session.flush()
        db.session.add(EducationPlanLineScope(
            education_plan_line_id=new_line.id,
            scope_kind="GRADE",
            grade=5,
            building_id=context["building_id"],
            scope_key=line_scope_key(
                "GRADE",
                grade=5,
                building_id=context["building_id"],
            ),
        ))
        db.session.flush()
        db.session.expire(plan, ["lines"])

        generate_plan_needs(
            version,
            user_id=user_id,
            source_plan_line_ids={new_line.id},
        )
        db.session.flush()

        preserved_group_ids = {
            item.id
            for item in TeachingGroup.query.filter(
                TeachingGroup.source_plan_line_id == original_line.id,
                TeachingGroup.status != "CLOSED",
            ).all()
        }
        assert preserved_group_ids == original_group_ids

        new_groups = TeachingGroup.query.filter(
            TeachingGroup.source_plan_line_id == new_line.id,
            TeachingGroup.status != "CLOSED",
        ).all()
        assert len(new_groups) == 1
        new_need = WorkloadNeed.query.filter_by(
            teaching_group_id=new_groups[0].id,
        ).one()

        plan_matrix = build_teaching_group_matrix(
            snapshot,
            [plan],
            "OOO",
            version.id,
            grade=5,
        )
        teacher = db.session.get(User, user_id)
        workload_matrix = build_workload_assignment_matrix(
            [new_need],
            [],
            plan_matrices=[plan_matrix],
            extra_teachers=[teacher],
            draft_rows=[(teacher, activity, "CURRICULUM")],
        )
        row = workload_matrix["blocks"][0]["rows"][0]
        cells = list(row["matrix_cells"].values())
        assert len(cells) == 1
        assert cells[0]["planned"] is True
        assert [slot["need"].id for slot in cells[0]["slots"]] == [
            new_need.id,
        ]


def test_group_matrix_can_plan_split_before_students_are_enrolled(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        version = db.session.get(TariffVersion, context["version_id"])
        empty_class = SchoolClass(
            academic_year_id=version.tariff_cycle.academic_year_id,
            building_id=context["building_id"],
            name="5В",
            grade=5,
            letter="В",
        )
        db.session.add(empty_class)
        db.session.commit()
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                source_school_class_id=empty_class.id,
            )
            .one()
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        assign_class_plan(
            snapshot_class,
            [line.education_plan],
            line.education_plan_id,
            user_id=user_id,
        )
        db.session.commit()
        payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": snapshot_class.id,
            "plan_id": line.education_plan_id,
            "group_count": 2,
        }

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data=payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    assert response.get_json()["group_count"] == 2
    assert response.get_json()["needs_composition"] is True
    with app.app_context():
        groups = TeachingGroup.query.order_by(TeachingGroup.id.asc()).all()
        assert len(groups) == 2
        assert {group.group_type for group in groups} == {"SUBGROUP"}
        assert {group.status for group in groups} == {"DRAFT"}
        assert all(group.planned_size == 0 for group in groups)
        assert all(group.actual_size == 0 for group in groups)
        assert all(not group.members for group in groups)


def test_group_count_change_is_blocked_when_target_has_assignment(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        classes = {
            item.name_snapshot: item
            for item in snapshot.classes
        }
        for snapshot_class in classes.values():
            replace_plan_binding_members(
                line.education_plan,
                snapshot_class,
                {item.id for item in snapshot_class.enrollments},
                user_id=user_id,
            )
        generate_plan_needs(
            db.session.get(TariffVersion, context["version_id"]),
            user_id=user_id,
        )
        db.session.flush()
        groups = (
            TeachingGroup.query
            .join(TeachingGroupClass)
            .order_by(TeachingGroup.id.asc())
            .all()
        )
        group_by_class_id = {
            group.source_classes[0].population_snapshot_class_id: group
            for group in groups
        }
        for need in WorkloadNeed.query.all():
            db.session.add(WorkloadAssignment(
                organization_id=need.organization_id,
                tariff_version_id=need.tariff_version_id,
                workload_need_id=need.id,
                employee_user_id=teacher_id,
                position_code="TEACHER",
                position_title="Учитель",
                department_id=need.department_id,
                building_id=need.building_id,
                assignment_kind="MAIN",
                date_from=need.date_from,
                date_to=need.date_to,
                weekly_hours=need.weekly_hours,
                annual_hours=need.annual_hours,
                status="DRAFT",
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            ))
        db.session.commit()
        target_class = classes["5А"]
        target_class_id = target_class.id
        untouched_group_id = group_by_class_id[classes["5Б"].id].id
        target_group_id = group_by_class_id[target_class_id].id
        payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": target_class_id,
            "plan_id": line.education_plan_id,
            "group_count": 2,
        }

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data=payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 422
    assert "уже назначена нагрузка" in response.get_json()["message"]
    with app.app_context():
        assert TeachingGroup.query.filter_by(id=target_group_id).one()
        assert TeachingGroup.query.filter_by(id=untouched_group_id).one()
        assert WorkloadNeed.query.count() == 2
        assert WorkloadAssignment.query.count() == 2


def test_group_count_change_removes_only_unassigned_target_need(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        classes = {
            item.name_snapshot: item
            for item in snapshot.classes
        }
        for snapshot_class in classes.values():
            replace_plan_binding_members(
                line.education_plan,
                snapshot_class,
                {item.id for item in snapshot_class.enrollments},
                user_id=user_id,
            )
        generate_plan_needs(
            db.session.get(TariffVersion, context["version_id"]),
            user_id=user_id,
        )
        db.session.commit()
        target_class = classes["5А"]
        target_group = (
            TeachingGroup.query
            .join(TeachingGroupClass)
            .filter(
                TeachingGroupClass.population_snapshot_class_id
                == target_class.id,
            )
            .one()
        )
        untouched_class = classes["5Б"]
        untouched_group = (
            TeachingGroup.query
            .join(TeachingGroupClass)
            .filter(
                TeachingGroupClass.population_snapshot_class_id
                == untouched_class.id,
            )
            .one()
        )
        target_group_id = target_group.id
        untouched_group_id = untouched_group.id
        payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": target_class.id,
            "plan_id": line.education_plan_id,
            "group_count": 2,
        }

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data=payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    assert response.get_json()["group_count"] == 2
    with app.app_context():
        assert TeachingGroup.query.filter_by(id=target_group_id).first() is None
        replacement_groups = (
            TeachingGroup.query
            .join(TeachingGroupClass)
            .filter(
                TeachingGroupClass.population_snapshot_class_id
                == target_class.id,
            )
            .all()
        )
        assert len(replacement_groups) == 2
        active_need_group_ids = {
            need.teaching_group_id
            for need in WorkloadNeed.query.filter(
                WorkloadNeed.status != "CANCELLED"
            ).all()
        }
        assert active_need_group_ids == {
            untouched_group_id,
            *(group.id for group in replacement_groups),
        }


def test_population_refresh_keeps_workload_assignment_identity(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _group_context(user_id)
        version = db.session.get(TariffVersion, context["version_id"])
        first_snapshot_id = _snapshot(user_id, version.id)
        first_snapshot = db.session.get(PopulationSnapshot, first_snapshot_id)
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        for snapshot_class in first_snapshot.classes:
            replace_plan_binding_members(
                line.education_plan,
                snapshot_class,
                {item.id for item in snapshot_class.enrollments},
                user_id=user_id,
            )
        generate_plan_needs(version, user_id=user_id)
        need = (
            WorkloadNeed.query
            .join(TeachingGroup)
            .join(TeachingGroupClass)
            .join(PopulationSnapshotClass)
            .filter(PopulationSnapshotClass.name_snapshot == "5А")
            .one()
        )
        assignment = WorkloadAssignment(
            organization_id=need.organization_id,
            tariff_version_id=version.id,
            workload_need_id=need.id,
            employee_user_id=teacher_id,
            position_code="TEACHER",
            position_title="Учитель",
            department_id=need.department_id,
            building_id=need.building_id,
            assignment_kind="MAIN",
            date_from=need.date_from,
            date_to=need.date_to,
            weekly_hours=need.weekly_hours,
            annual_hours=need.annual_hours,
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(assignment)
        db.session.commit()
        old_need_id = need.id

        school_class = SchoolClass.query.filter_by(
            academic_year_id=context["year_id"],
            name="5А",
        ).one()
        added_child = _child("Новый", "Ученик")
        db.session.add(ChildEnrollment(
            child_id=added_child.id,
            academic_year_id=context["year_id"],
            school_class_id=school_class.id,
            status="ACTIVE",
            enrolled_at=datetime(2026, 9, 2),
        ))
        db.session.commit()

        current_snapshot = build_population_snapshot(
            version,
            user_id=user_id,
            snapshot_date=date(2026, 9, 2),
        )
        materialize_default_teaching_groups(
            version=version,
            snapshot=current_snapshot,
            plans=[line.education_plan],
            user_id=user_id,
        )
        generate_plan_needs(version, user_id=user_id)

        changed = relink_assignments_to_population_snapshot(
            version,
            current_snapshot,
            user_id=user_id,
        )
        db.session.commit()

        assert changed == 0
        db.session.refresh(assignment)
        assert assignment.workload_need_id == old_need_id
        target_class = (
            assignment.workload_need.teaching_group.source_classes[0]
            .population_snapshot_class
        )
        assert target_class.population_snapshot_id == current_snapshot.id
        assert WorkloadAssignmentChange.query.count() == 0
        assert {
            member.snapshot_enrollment.source_child_id
            for member in assignment.workload_need.teaching_group.members
        } == {
            item.source_child_id
            for item in target_class.enrollments
        }


def test_population_refresh_relinks_assignment_after_bound_plan_change(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _group_context(user_id)
        version = db.session.get(TariffVersion, context["version_id"])
        first_snapshot = db.session.get(
            PopulationSnapshot,
            _snapshot(user_id, version.id),
        )
        first_class = next(
            item for item in first_snapshot.classes
            if item.name_snapshot == "5А"
        )
        old_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        old_group = _ready_source_group(
            context,
            first_class,
            user_id=user_id,
            suffix="OLD_PLAN",
            member_ids=[],
        )
        old_need = WorkloadNeed(
            tariff_version_id=version.id,
            teaching_group_id=old_group.id,
            education_activity_id=old_line.education_activity_id,
            building_id=context["building_id"],
            date_from=date(2026, 9, 1),
            date_to=date(2027, 5, 28),
            weekly_hours=Decimal("5"),
            annual_hours=Decimal("170"),
            need_kind="PLAN",
            status="COVERED",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(old_need)
        db.session.flush()
        assignment = WorkloadAssignment(
            tariff_version_id=version.id,
            workload_need_id=old_need.id,
            employee_user_id=teacher_id,
            position_code="TEACHER",
            position_title="Учитель",
            building_id=context["building_id"],
            assignment_kind="MAIN",
            date_from=old_need.date_from,
            date_to=old_need.date_to,
            weekly_hours=old_need.weekly_hours,
            annual_hours=old_need.annual_hours,
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(assignment)

        replacement_line = EducationPlanLine(
            education_plan_id=old_line.education_plan_id,
            education_activity_id=old_line.education_activity_id,
            component_kind=old_line.component_kind,
            weekly_hours=old_line.weekly_hours,
            weeks_count=old_line.weeks_count,
            annual_hours=old_line.annual_hours,
            requires_division=old_line.requires_division,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(replacement_line)
        first_snapshot.status = "SUPERSEDED"
        db.session.flush()
        current_snapshot = PopulationSnapshot(
            tariff_version_id=version.id,
            revision_no=2,
            snapshot_date=date(2026, 9, 2),
            status="CURRENT",
            source_kind="REGISTRY",
            checksum="replacement-plan",
            created_by_user_id=user_id,
        )
        db.session.add(current_snapshot)
        db.session.flush()
        current_class = PopulationSnapshotClass(
            population_snapshot_id=current_snapshot.id,
            source_school_class_id=first_class.source_school_class_id,
            name_snapshot=first_class.name_snapshot,
            grade_snapshot=first_class.grade_snapshot,
            building_id=first_class.building_id,
            building_name_snapshot=first_class.building_name_snapshot,
            student_count=first_class.student_count,
        )
        db.session.add(current_class)
        db.session.flush()
        new_group = _ready_source_group(
            context,
            current_class,
            user_id=user_id,
            suffix="NEW_PLAN",
            plan_line_id=replacement_line.id,
            member_ids=[],
        )
        new_need = WorkloadNeed(
            tariff_version_id=version.id,
            teaching_group_id=new_group.id,
            education_activity_id=replacement_line.education_activity_id,
            building_id=context["building_id"],
            date_from=old_need.date_from,
            date_to=old_need.date_to,
            weekly_hours=old_need.weekly_hours,
            annual_hours=old_need.annual_hours,
            need_kind=old_need.need_kind,
            status="OPEN",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(new_need)
        db.session.commit()

        changed = relink_assignments_to_population_snapshot(
            version,
            current_snapshot,
            user_id=user_id,
        )
        db.session.commit()

        assert changed == 1
        db.session.refresh(assignment)
        assert assignment.workload_need_id == new_need.id
        assert WorkloadAssignmentChange.query.filter_by(
            workload_assignment_id=assignment.id,
            change_kind="TRANSFER",
        ).count() == 1


def test_population_refresh_updates_only_split_group_rosters(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _group_context(user_id)
        version = db.session.get(TariffVersion, context["version_id"])
        first_snapshot_id = _snapshot(user_id, version.id)
        first_snapshot = db.session.get(PopulationSnapshot, first_snapshot_id)
        snapshot_class = next(
            item
            for item in first_snapshot.classes
            if item.name_snapshot == "5А"
        )
        line = db.session.get(EducationPlanLine, context["plan_line_id"])
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        groups = replace_teaching_group_count(
            version=version,
            snapshot=first_snapshot,
            plans=[line.education_plan],
            plan_line_id=line.id,
            snapshot_class_id=snapshot_class.id,
            plan_id=line.education_plan_id,
            group_count=2,
            user_id=user_id,
        )
        enrollments = list(snapshot_class.enrollments)
        original_child_ids = {
            item.source_child_id for item in enrollments
        }
        for group, enrollment in zip(groups, enrollments, strict=True):
            db.session.add(TeachingGroupMember(
                teaching_group_id=group.id,
                snapshot_enrollment_id=enrollment.id,
                valid_from=group.valid_from,
                valid_to=group.valid_to,
                source_kind="MANUAL",
            ))
            group.actual_size = 1
            group.status = "READY"
            group.source_classes[0].student_count = 1
        db.session.add(TeachingGroupCompositionApproval(
            tariff_version_id=version.id,
            education_plan_line_id=line.id,
            population_snapshot_class_id=snapshot_class.id,
            approved_by_user_id=user_id,
        ))
        generate_plan_needs(version, user_id=user_id)
        assigned_need = WorkloadNeed.query.filter_by(
            teaching_group_id=groups[0].id,
        ).one()
        assignment = WorkloadAssignment(
            organization_id=assigned_need.organization_id,
            tariff_version_id=version.id,
            workload_need_id=assigned_need.id,
            employee_user_id=teacher_id,
            position_code="TEACHER",
            position_title="Учитель",
            department_id=assigned_need.department_id,
            building_id=assigned_need.building_id,
            assignment_kind="MAIN",
            date_from=assigned_need.date_from,
            date_to=assigned_need.date_to,
            weekly_hours=assigned_need.weekly_hours,
            annual_hours=assigned_need.annual_hours,
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(assignment)
        db.session.commit()
        group_ids = {item.id for item in groups}
        need_id = assigned_need.id
        assignment_id = assignment.id

        school_class = SchoolClass.query.filter_by(
            academic_year_id=context["year_id"],
            name="5А",
        ).one()
        departed_enrollment = db.session.get(
            ChildEnrollment,
            enrollments[1].source_enrollment_id,
        )
        departed_child_id = departed_enrollment.child_id
        departed_enrollment.status = "EXPELLED"
        departed_enrollment.ended_at = datetime(2026, 9, 2)
        newcomer = _child("Новая", "Ученица")
        db.session.add(ChildEnrollment(
            child_id=newcomer.id,
            academic_year_id=context["year_id"],
            school_class_id=school_class.id,
            status="ACTIVE",
            enrolled_at=datetime(2026, 9, 2),
        ))
        db.session.commit()
        newcomer_id = newcomer.id

        current_snapshot = build_population_snapshot(
            version,
            user_id=user_id,
            snapshot_date=date(2026, 9, 2),
        )
        db.session.commit()

        current_groups = TeachingGroup.query.filter(
            TeachingGroup.id.in_(group_ids)
        ).all()
        assert {item.id for item in current_groups} == group_ids
        active_member_child_ids = {
            member.snapshot_enrollment.source_child_id
            for group in current_groups
            for member in group.members
        }
        assert active_member_child_ids == (
            original_child_ids - {departed_child_id}
        )
        assert newcomer_id not in active_member_child_ids
        assert TeachingGroupCompositionApproval.query.count() == 0
        db.session.refresh(assignment)
        assert assignment.id == assignment_id
        assert assignment.workload_need_id == need_id
        assert all(
            source.population_snapshot_class.population_snapshot_id
            == current_snapshot.id
            for group in current_groups
            for source in group.source_classes
        )


def test_group_composition_assigns_students_to_split_groups(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        enrollment_ids = [
            item.id for item in snapshot_class.enrollments
        ]
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            set(enrollment_ids),
            user_id=user_id,
        )
        db.session.commit()
        matrix_payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": snapshot_class.id,
            "plan_id": line.education_plan_id,
            "group_count": 2,
        }
        item_key = f"line-{line.id}-class-{snapshot_class.id}"

    login(user_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data=matrix_payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 200

    response = client.get(
        f"/workload/groups/composition/"
        f"?version_id={context['version_id']}&level=OOO"
        f"&grade=5&item={item_key}"
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert "Состав учебных групп" in html
    assert "Иванов Иван" in html
    assert "Петрова Анна" in html
    assert "Распределить по порядку" in html
    assert "Распределить пополам" in html
    assert "Сбросить" in html
    assert "data-distribute-in-order" in html
    assert "data-distribute-in-halves" in html
    assert "data-reset-composition" in html
    assert html.count("data-unassigned-option") == 2
    assert 'aria-label="Параллель"' in html
    assert 'name="grade" value="5"' in html
    assert 'class="workload-indicators"' not in html

    with app.app_context():
        group_ids = [
            group.id
            for group in TeachingGroup.query.order_by(
                TeachingGroup.id.asc()
            )
        ]
    response = client.post(
        "/workload/groups/composition/",
        data={
            "version_id": context["version_id"],
            "level": "OOO",
            "grade": "5",
            "item_key": item_key,
            f"member_{enrollment_ids[0]}": group_ids[0],
            f"member_{enrollment_ids[1]}": group_ids[1],
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["complete"] is True
    assert payload["assigned_count"] == 2
    with app.app_context():
        groups = TeachingGroup.query.order_by(
            TeachingGroup.id.asc()
        ).all()
        assert {group.status for group in groups} == {"READY"}
        assert [group.actual_size for group in groups] == [1, 1]
        assert sum(len(group.members) for group in groups) == 2

    response = client.post(
        "/workload/groups/composition/",
        data={
            "version_id": context["version_id"],
            "level": "OOO",
            "grade": "5",
            "item_key": item_key,
            f"member_{enrollment_ids[0]}": "",
            f"member_{enrollment_ids[1]}": "",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["complete"] is False
    assert payload["assigned_count"] == 0
    assert payload["student_count"] == 2
    with app.app_context():
        groups = TeachingGroup.query.order_by(
            TeachingGroup.id.asc()
        ).all()
        assert {group.status for group in groups} == {"DRAFT"}
        assert [group.actual_size for group in groups] == [0, 0]
        assert all(not group.members for group in groups)


def test_class_teacher_distributes_approves_and_exports_groups(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    class_teacher_id = make_user("CLASS_TEACHER")
    subject_teacher_id = make_user("TEACHER")
    with app.app_context():
        class_teacher = db.session.get(User, class_teacher_id)
        class_teacher.last_name = "Иванова"
        class_teacher.first_name = "Мария"
        subject_teacher = db.session.get(User, subject_teacher_id)
        subject_teacher.last_name = "Смирнова"
        subject_teacher.first_name = "Елена"

        context = _group_context(admin_id)
        snapshot_id = _snapshot(admin_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        snapshot_class.source_school_class.teacher_user_id = (
            class_teacher_id
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        enrollment_ids = [
            item.id
            for item in snapshot_class.enrollments
        ]
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            set(enrollment_ids),
            user_id=admin_id,
        )
        db.session.commit()
        item_key = f"line-{line.id}-class-{snapshot_class.id}"
        matrix_payload = {
            "version_id": context["version_id"],
            "plan_line_id": line.id,
            "snapshot_class_id": snapshot_class.id,
            "plan_id": line.education_plan_id,
            "group_count": 2,
        }
        class_id = snapshot_class.source_school_class_id

    login(admin_id)
    response = client.post(
        "/workload/groups/matrix/cell",
        data=matrix_payload,
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert response.status_code == 200

    with app.app_context():
        groups = (
            TeachingGroup.query
            .order_by(TeachingGroup.id.asc())
            .all()
        )
        for group in groups:
            need = WorkloadNeed.query.filter_by(
                teaching_group_id=group.id,
                status="OPEN",
            ).one()
            need.weekly_hours = Decimal("2.5")
            need.annual_hours = Decimal("85")
            need.status = "COVERED"
            db.session.flush()
            db.session.add(WorkloadAssignment(
                tariff_version_id=context["version_id"],
                workload_need_id=need.id,
                employee_user_id=subject_teacher_id,
                position_code="TEACHER",
                building_id=context["building_id"],
                assignment_kind="MAIN",
                date_from=date(2026, 9, 1),
                date_to=date(2027, 8, 31),
                weekly_hours=Decimal("2.5"),
                annual_hours=Decimal("85"),
                status="CONFIRMED",
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            ))
        db.session.commit()
        group_ids = [group.id for group in groups]

    login(class_teacher_id)
    hub_response = client.get("/hub/classroom")
    hub_html = hub_response.get_data(as_text=True)
    assert hub_response.status_code == 200
    assert "Учебный план класса" in hub_html
    assert "/hub/classroom/curriculum" in hub_html
    assert "Распределение по учебным группам" in hub_html
    assert "/hub/classroom/groups" in hub_html

    curriculum_response = client.get(
        f"/hub/classroom/curriculum?class_id={class_id}"
    )
    curriculum_html = curriculum_response.get_data(as_text=True)
    assert curriculum_response.status_code == 200
    assert "Учебный план класса" in curriculum_html
    assert "Часов в неделю" in curriculum_html
    assert "Преподаватели из нагрузки" in curriculum_html
    assert "Математика" in curriculum_html
    assert "5 ч." in curriculum_html
    assert "Смирнова Елена" in curriculum_html

    page_response = client.get(
        f"/hub/classroom/groups?class_id={class_id}&item={item_key}"
    )
    page_html = page_response.get_data(as_text=True)
    assert page_response.status_code == 200
    assert "5А" in page_html
    assert "5Б" not in page_html
    assert "Математика" in page_html
    assert page_html.count("Смирнова Елена") >= 2
    assert "Скачать Excel" in page_html
    assert "Согласовано классным руководителем" in page_html
    assert "group-composition-choice" in page_html
    assert (
        '<span class="group-composition-choice" '
        'aria-hidden="true">✓</span>'
    ) in page_html

    update_response = client.post(
        "/hub/classroom/groups",
        data={
            "class_id": class_id,
            "item_key": item_key,
            f"member_{enrollment_ids[0]}": group_ids[0],
            f"member_{enrollment_ids[1]}": group_ids[1],
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert update_response.status_code == 200
    assert update_response.get_json()["complete"] is True

    approve_response = client.post(
        "/hub/classroom/groups/approve",
        data={
            "class_id": class_id,
            "item_key": item_key,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert approve_response.status_code == 200
    assert approve_response.get_json()["ok"] is True
    with app.app_context():
        approval = TeachingGroupCompositionApproval.query.one()
        assert approval.approved_by_user_id == class_teacher_id
        assert approval.education_plan_line_id == context["plan_line_id"]
        assert (
            approval.population_snapshot_class_id
            == snapshot_class.id
        )

    approved_page = client.get(
        f"/hub/classroom/groups?class_id={class_id}&item={item_key}"
    ).get_data(as_text=True)
    assert "Согласовано" in approved_page
    assert "Иванова Мария" in approved_page

    export_response = client.get(
        f"/hub/classroom/groups/export.xlsx?class_id={class_id}"
    )
    assert export_response.status_code == 200
    assert export_response.mimetype == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    workbook = load_workbook(
        BytesIO(export_response.data),
        data_only=True,
    )
    sheet = workbook["Распределение"]
    exported_rows = list(sheet.iter_rows(values_only=True))
    assert any("Смирнова Елена" in row for row in exported_rows)
    assert any(
        "Согласовано классным руководителем" in row
        for row in exported_rows
    )

    login(admin_id)
    matrix_response = client.get(
        f"/workload/groups/"
        f"?version_id={context['version_id']}&level=OOO&grade=5"
    )
    matrix_html = matrix_response.get_data(as_text=True)
    assert matrix_response.status_code == 200
    assert "is-approved" in matrix_html
    assert "Деление согласовано классным руководителем" in matrix_html

    with app.app_context():
        new_child = Child(
            last_name="Новикова",
            first_name="Анна",
        )
        db.session.add(new_child)
        db.session.flush()
        db.session.add(ChildEnrollment(
            child_id=new_child.id,
            academic_year_id=db.session.get(
                SchoolClass,
                class_id,
            ).academic_year_id,
            school_class_id=class_id,
            status="ACTIVE",
        ))
        db.session.commit()

        refreshed_snapshot_class = db.session.get(
            PopulationSnapshotClass,
            snapshot_class.id,
        )
        new_snapshot_enrollment = (
            PopulationSnapshotEnrollment.query
            .filter_by(
                population_snapshot_class_id=snapshot_class.id,
                source_child_id=new_child.id,
            )
            .one()
        )
        assert refreshed_snapshot_class.student_count == 3
        assert TeachingGroupCompositionApproval.query.count() == 0
        assert all(
            new_snapshot_enrollment.id not in {
                member.snapshot_enrollment_id
                for member in group.members
            }
            for group in TeachingGroup.query.all()
        )

    login(class_teacher_id)
    refreshed_page = client.get(
        f"/hub/classroom/groups?class_id={class_id}&item={item_key}"
    )
    refreshed_html = refreshed_page.get_data(as_text=True)
    assert refreshed_page.status_code == 200
    assert "Новикова Анна" in refreshed_html
    assert "2 из 3" in refreshed_html
    assert "Требуют согласования: <strong>1</strong>" in refreshed_html
    assert "Матрица распределения класса" not in refreshed_html

    reapprove_response = client.post(
        "/hub/classroom/groups/approve",
        data={
            "class_id": class_id,
            "item_key": item_key,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert reapprove_response.status_code == 422
    assert "Сначала распределите всех учеников" in (
        reapprove_response.get_json()["message"]
    )


def test_generating_needs_materializes_default_one_group(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        replace_plan_binding_members(
            line.education_plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()

        result = generate_plan_needs(
            db.session.get(TariffVersion, context["version_id"]),
            user_id=user_id,
        )
        db.session.commit()

        assert result["created"] == 1
        group = TeachingGroup.query.one()
        assert group.group_type == "CLASS"
        assert group.status == "READY"
        assert len(group.members) == 2
        assert WorkloadNeed.query.one().teaching_group_id == group.id


def test_generating_needs_includes_all_plan_subjects_for_empty_class(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        source_class = SchoolClass.query.filter_by(name="5А").one()
        ChildEnrollment.query.filter_by(
            school_class_id=source_class.id,
        ).delete(synchronize_session=False)
        first_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        plan = first_line.education_plan
        literature = EducationActivity(
            code="LITERATURE_EMPTY_CLASS",
            name="Литература",
            activity_kind="SUBJECT",
            is_global=True,
            is_tariffable=True,
            is_active=True,
        )
        db.session.add(literature)
        db.session.flush()
        literature_line = EducationPlanLine(
            education_plan_id=plan.id,
            education_activity_id=literature.id,
            component_kind="MANDATORY",
            weekly_hours=Decimal("3"),
            weeks_count=Decimal("34"),
            annual_hours=Decimal("102"),
            sort_order=20,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(literature_line)
        db.session.flush()
        db.session.add(EducationPlanLineScope(
            education_plan_line_id=literature_line.id,
            scope_kind="GRADE",
            grade=5,
            building_id=context["building_id"],
            scope_key=line_scope_key(
                "GRADE",
                grade=5,
                building_id=context["building_id"],
            ),
        ))
        db.session.commit()

        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        assign_class_plan(
            snapshot_class,
            [plan],
            plan.id,
            user_id=user_id,
        )
        db.session.commit()

        result = generate_plan_needs(
            db.session.get(TariffVersion, context["version_id"]),
            user_id=user_id,
        )
        db.session.commit()

        assert result["created"] == 2
        groups = TeachingGroup.query.order_by(TeachingGroup.id).all()
        assert len(groups) == 2
        assert all(group.status == "READY" for group in groups)
        assert all(not group.members for group in groups)
        needs = WorkloadNeed.query.order_by(
            WorkloadNeed.education_activity_id
        ).all()
        assert len(needs) == 2
        literature_need = next(
            item
            for item in needs
            if item.education_activity_id == literature.id
        )
        assert literature_need.weekly_hours == Decimal("3.000")

        child = Child.query.order_by(Child.id.asc()).first()
        enrollment = PopulationSnapshotEnrollment(
            population_snapshot_class_id=snapshot_class.id,
            source_child_id=child.id,
            fio_snapshot="Новый Ученик",
            status_snapshot="ACTIVE",
        )
        db.session.add(enrollment)
        snapshot_class.student_count = 1
        db.session.commit()

        created = materialize_default_teaching_groups(
            version=db.session.get(
                TariffVersion,
                context["version_id"],
            ),
            snapshot=snapshot,
            plans=[plan],
            user_id=user_id,
        )
        db.session.commit()

        assert created == 0
        assert all(
            {member.snapshot_enrollment_id for member in group.members}
            == {enrollment.id}
            for group in TeachingGroup.query.all()
        )


def test_class_plan_matrix_places_all_bundle_parts_on_one_sheet(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        root_plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        extracurricular = EducationActivity(
            code="FUNCTIONAL_LITERACY",
            name="Функциональная грамотность",
            activity_kind="EXTRACURRICULAR_COURSE",
            is_global=True,
            is_tariffable=True,
            is_active=True,
        )
        additional = EducationActivity(
            code="ROBOTICS",
            name="Робототехника",
            activity_kind="ADDITIONAL_PROGRAM",
            is_global=True,
            is_tariffable=True,
            is_active=True,
        )
        db.session.add_all([extracurricular, additional])
        db.session.flush()
        for plan_kind, component, activity, hours in (
            (
                "EXTRACURRICULAR",
                "EXTRACURRICULAR",
                extracurricular,
                Decimal("1"),
            ),
            (
                "ADDITIONAL_EDUCATION",
                "ADDITIONAL",
                additional,
                Decimal("2"),
            ),
        ):
            companion = EducationPlan(
                tariff_version_id=context["version_id"],
                root_plan_id=root_plan.id,
                plan_kind=plan_kind,
                name=f"{root_plan.name} · {plan_kind}",
                education_level="OOO",
                building_id=context["building_id"],
                scope_code=root_plan.scope_code,
                status="DRAFT",
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            db.session.add(companion)
            db.session.flush()
            line = EducationPlanLine(
                education_plan_id=companion.id,
                education_activity_id=activity.id,
                component_kind=component,
                weekly_hours=hours,
                weeks_count=Decimal("34"),
                annual_hours=hours * Decimal("34"),
                sort_order=10,
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            db.session.add(line)
            db.session.flush()
            db.session.add(EducationPlanLineScope(
                education_plan_line_id=line.id,
                scope_kind="GRADE",
                grade=5,
                building_id=context["building_id"],
                scope_key=line_scope_key(
                    "GRADE",
                    grade=5,
                    building_id=context["building_id"],
                ),
            ))
        replace_plan_binding_members(
            root_plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        version_id = context["version_id"]

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/matrix"
        f"?version_id={version_id}&level=OOO"
    )

    assert response.status_code == 200
    mandatory_index = response.data.index("Обязательная часть".encode())
    extracurricular_index = response.data.index(
        "Внеурочная деятельность".encode()
    )
    additional_index = response.data.index(
        "Дополнительное образование".encode()
    )
    assert mandatory_index < extracurricular_index < additional_index
    curriculum_total_index = response.data.index(
        "class-plan-matrix__curriculum-total".encode()
    )
    extracurricular_total_index = response.data.index(
        'data-plan-kind-total="EXTRACURRICULAR"'.encode()
    )
    additional_total_index = response.data.index(
        'data-plan-kind-total="ADDITIONAL_EDUCATION"'.encode()
    )
    assert (
        curriculum_total_index
        < extracurricular_total_index
        < additional_total_index
    )
    assert 'data-activity-name="Математика"'.encode() in response.data
    assert (
        'data-activity-name="Функциональная грамотность"'.encode()
        in response.data
    )
    assert 'data-activity-name="Робототехника"'.encode() in response.data


def test_class_plan_matrix_uses_first_period_on_september_first():
    line = EducationPlanLine(weekly_hours=Decimal("5"))
    line.periods = [
        EducationPlanLinePeriod(
            date_from=date(2026, 9, 1),
            date_to=date(2026, 10, 31),
            weekly_hours=Decimal("3"),
        ),
        EducationPlanLinePeriod(
            date_from=date(2026, 11, 1),
            date_to=date(2027, 5, 31),
            weekly_hours=Decimal("5"),
        ),
    ]

    assert effective_line_weekly_hours(line, 1) == Decimal("3")
    assert effective_line_weekly_hours(line, 11) == Decimal("3")
    assert effective_line_weekly_hours(line, 10) == Decimal("5")
    assert class_period_label("NOO", 1) == "сент.–окт."
    assert class_period_label("SOO", 11) == "I период"
    assert class_period_label("OOO", 5) == "ч/нед."


def test_plan_bindings_page_warns_when_population_registry_changed(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        db.session.add(SchoolClass(
            academic_year_id=context["year_id"],
            building_id=context["building_id"],
            name="5В",
            grade=5,
            letter="В",
        ))
        db.session.commit()

        version = db.session.get(TariffVersion, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        registry_status = population_registry_status(version, snapshot)
        assert registry_status["is_stale"] is True
        assert registry_status["class_count"] == 3
        assert registry_status["snapshot_class_count"] == 2

    login(user_id)
    response = client.get(
        f"/workload/plan-bindings/?version_id={context['version_id']}"
    )

    assert response.status_code == 200
    assert "Сводный контингент изменён".encode() in response.data
    assert "Обновить данные".encode() in response.data


def test_population_registry_detects_building_change_without_count_change(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        version = db.session.get(TariffVersion, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)

        initial_status = population_registry_status(version, snapshot)
        assert initial_status["is_stale"] is False
        assert initial_status["structure_changed"] is False

        second_building = Building(
            name="Второе здание",
            short_name="ВЗ",
        )
        db.session.add(second_building)
        db.session.flush()
        school_class = SchoolClass.query.filter_by(
            academic_year_id=context["year_id"],
            name="5А",
        ).one()
        school_class.building_id = second_building.id
        db.session.commit()

        changed_status = population_registry_status(version, snapshot)
        assert changed_status["is_stale"] is True
        assert changed_status["structure_changed"] is True
        assert changed_status["class_count"] == changed_status[
            "snapshot_class_count"
        ]
        assert changed_status["student_count"] == changed_status[
            "snapshot_student_count"
        ]

        refreshed_snapshot = build_population_snapshot(
            version,
            user_id=user_id,
            snapshot_date=date(2026, 9, 2),
        )
        db.session.commit()
        refreshed_class = next(
            item
            for item in refreshed_snapshot.classes
            if item.source_school_class_id == school_class.id
        )
        assert refreshed_class.building_id == second_building.id
        assert refreshed_class.building_name_snapshot == "ВЗ"
        assert population_registry_status(
            version,
            refreshed_snapshot,
        )["is_stale"] is False


def test_class_plan_can_be_overridden_for_one_student(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        first_plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="Профильный учебный план",
            education_level="OOO",
            building_id=context["building_id"],
            scope_code="OOO_PROFILE",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_plan)
        db.session.commit()
        class_id = snapshot_class.id
        first_plan_id = first_plan.id
        second_plan_id = second_plan.id
        enrollment_ids = [
            item.id for item in snapshot_class.enrollments
        ]
        version_id = context["version_id"]

    login(user_id)
    response = client.post(
        "/workload/plan-bindings/class",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "plan_id": first_plan_id,
            "level": "OOO",
            "grade": "5",
        },
    )
    assert response.status_code == 302
    assert "class_id=" not in response.headers["Location"]
    assert "level=OOO" in response.headers["Location"]
    assert "grade=5" in response.headers["Location"]

    summary_page = client.get(response.headers["Location"])
    assert summary_page.status_code == 200
    assert (
        f'value="{first_plan_id}" selected'.encode()
        in summary_page.data
    )
    assert "Индивидуальные назначения".encode() not in summary_page.data

    with app.app_context():
        snapshot_class = db.session.get(
            PopulationSnapshotClass,
            class_id,
        )
        plans = [
            db.session.get(EducationPlan, first_plan_id),
            db.session.get(EducationPlan, second_plan_id),
        ]
        allocations, _ = class_plan_allocations(
            snapshot_class,
            plans,
        )
        assert allocations[first_plan_id] == set(enrollment_ids)
        assert EducationPlanBinding.query.one().binding_mode == "CLASS"

    response = client.post(
        "/workload/plan-bindings/student",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "enrollment_id": enrollment_ids[0],
            "plan_id": second_plan_id,
            "level": "OOO",
            "grade": "5",
        },
    )
    assert response.status_code == 302
    assert "level=OOO" in response.headers["Location"]
    assert "grade=5" in response.headers["Location"]

    with app.app_context():
        snapshot_class = db.session.get(
            PopulationSnapshotClass,
            class_id,
        )
        plans = [
            db.session.get(EducationPlan, first_plan_id),
            db.session.get(EducationPlan, second_plan_id),
        ]
        allocations, student_plan_ids = class_plan_allocations(
            snapshot_class,
            plans,
        )
        assert allocations[second_plan_id] == {enrollment_ids[0]}
        assert allocations[first_plan_id] == {enrollment_ids[1]}
        assert student_plan_ids[enrollment_ids[0]] == second_plan_id
        assert {
            item.binding_mode
            for item in EducationPlanBinding.query.all()
        } == {"STUDENTS"}

    page = client.get(
        f"/workload/plan-bindings/?version_id={version_id}"
        f"&class_id={class_id}"
    )
    assert page.status_code == 200
    assert "Индивидуальные назначения".encode() in page.data
    assert "Профильный учебный план".encode() in page.data


def test_class_plan_assignment_returns_json_without_page_reload(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        plan_id = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        ).education_plan_id
        class_id = snapshot_class.id
        student_count = len(snapshot_class.enrollments)
        version_id = context["version_id"]

    login(user_id)
    page = client.get(
        f"/workload/plan-bindings/?version_id={version_id}"
    )
    assert page.status_code == 200
    assert "Сохранить учебный план класса".encode() in page.data

    response = client.post(
        "/workload/plan-bindings/class",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "plan_id": plan_id,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    assert response.get_json() == {
        "ok": True,
        "message": "Учебный план для 5А сохранён.",
        "assigned_count": student_count,
        "student_count": student_count,
    }


def test_missing_plan_id_does_not_clear_class_binding(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        class_id = snapshot_class.id
        plan_id = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        ).education_plan_id
        version_id = context["version_id"]

    login(user_id)
    saved = client.post(
        "/workload/plan-bindings/class",
        data={
            "version_id": version_id,
            "class_id": class_id,
            "plan_id": plan_id,
        },
    )
    assert saved.status_code == 302

    missing = client.post(
        "/workload/plan-bindings/class",
        data={
            "version_id": version_id,
            "class_id": class_id,
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert missing.status_code == 422
    assert missing.get_json()["ok"] is False
    with app.app_context():
        binding = EducationPlanBinding.query.one()
        assert binding.education_plan_id == plan_id
        assert binding.binding_mode == "CLASS"


def test_deleting_plan_removes_bindings_but_keeps_class_and_students(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        create_plan_bundle(plan, user_id=user_id)
        replace_plan_binding_members(
            plan,
            snapshot_class,
            {item.id for item in snapshot_class.enrollments},
            user_id=user_id,
        )
        db.session.commit()
        plan_id = plan.id
        revision = plan.revision
        snapshot_class_id = snapshot_class.id
        enrollment_ids = {
            item.id for item in snapshot_class.enrollments
        }

    login(user_id)
    registry = client.get("/workload/plans/")
    assert registry.status_code == 200
    assert "К плану привязано классов:".encode() in registry.data

    response = client.post(
        f"/workload/plans/{plan_id}/delete",
        data={"revision": revision},
    )

    assert response.status_code == 302
    with app.app_context():
        assert db.session.get(EducationPlan, plan_id) is None
        assert EducationPlan.query.count() == 0
        assert EducationPlanBinding.query.count() == 0
        assert (
            db.session.get(PopulationSnapshotClass, snapshot_class_id)
            is not None
        )
        assert {
            item.id
            for item in PopulationSnapshotEnrollment.query.filter(
                PopulationSnapshotEnrollment.id.in_(enrollment_ids)
            ).all()
        } == enrollment_ids


def test_metagroup_requires_two_source_classes():
    with pytest.raises(GroupValidationError):
        validate_group_sources("METAGROUP", [object()])


def test_administrator_creates_personal_subgroup(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        member_id = snapshot_class.enrollments[0].id
        class_id = snapshot_class.id
    login(user_id)

    response = client.post(
        "/workload/groups/new",
        data=_group_form(context, [class_id], [member_id]),
    )

    assert response.status_code == 302
    with app.app_context():
        group = TeachingGroup.query.one()
        assert group.group_type == "SUBGROUP"
        assert group.actual_size == 1
        assert len(group.members) == 1
        assert TeachingGroupHistory.query.count() == 1

    matrix = client.get(
        f"/workload/groups/?version_id={context['version_id']}&level=OOO"
    )
    matrix_html = matrix.get_data(as_text=True)
    assert matrix.status_code == 200
    assert "workload_distribution.css" in matrix_html
    assert "data-group-matrix" in matrix_html
    assert "Количество учебных групп" in matrix_html


def test_whole_class_automatically_gets_all_members(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        class_id = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
            .id
        )
    login(user_id)

    response = client.post(
        "/workload/groups/new",
        data=_group_form(
            context,
            [class_id],
            group_type="CLASS",
            code="MATH_5A",
            name="Математика 5А",
            snapshot_enrollment_ids=[],
        ),
    )

    assert response.status_code == 302
    with app.app_context():
        group = TeachingGroup.query.one()
        assert group.actual_size == 2
        assert len(group.members) == 2
        assert group.source_classes[0].relation_kind == "FULL"


def test_duplicate_parallel_membership_is_rejected(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        class_id = snapshot_class.id
        member_id = snapshot_class.enrollments[0].id
    login(user_id)
    first = client.post(
        "/workload/groups/new",
        data=_group_form(context, [class_id], [member_id]),
    )
    assert first.status_code == 302

    duplicate = client.post(
        "/workload/groups/new",
        data=_group_form(
            context,
            [class_id],
            [member_id],
            code="MATH_5A_2",
            name="Математика 5А, группа 2",
        ),
    )

    assert duplicate.status_code == 200
    assert "уже включён".encode() in duplicate.data
    with app.app_context():
        assert TeachingGroup.query.count() == 1


def test_count_only_group_uses_manual_actual_size(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        class_id = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
            .id
        )
    login(user_id)

    response = client.post(
        "/workload/groups/new",
        data=_group_form(
            context,
            [class_id],
            member_ids=[],
            composition_mode="COUNT_ONLY",
            actual_size="14",
        ),
    )

    assert response.status_code == 302
    with app.app_context():
        group = TeachingGroup.query.one()
        assert group.actual_size == 14
        assert group.members == []


def test_subgroup_coverage_reports_missing_students(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        class_id = snapshot_class.id
        member_id = snapshot_class.enrollments[0].id
    login(user_id)
    client.post(
        "/workload/groups/new",
        data=_group_form(context, [class_id], [member_id]),
    )

    with app.app_context():
        coverage = group_coverage(context["plan_line_id"])
        assert coverage[0]["expected_count"] == 2
        assert coverage[0]["assigned_count"] == 1
        assert coverage[0]["missing_count"] == 1
        assert coverage[0]["complete"] is False


def test_ready_group_is_read_only(app, client, make_user, login):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        class_id = snapshot_class.id
        member_id = snapshot_class.enrollments[0].id
    login(user_id)
    client.post(
        "/workload/groups/new",
        data=_group_form(context, [class_id], [member_id]),
    )
    with app.app_context():
        group_id = TeachingGroup.query.one().id

    response = client.post(
        f"/workload/groups/{group_id}/status",
        data={"status": "READY", "revision": "1"},
    )

    assert response.status_code == 302
    with app.app_context():
        group = db.session.get(TeachingGroup, group_id)
        assert group.status == "READY"
        assert group.revision == 2


def test_group_close_records_date_and_truncates_members(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot_class = (
            PopulationSnapshotClass.query
            .filter_by(
                population_snapshot_id=snapshot_id,
                name_snapshot="5А",
            )
            .one()
        )
        class_id = snapshot_class.id
        member_id = snapshot_class.enrollments[0].id
    login(user_id)
    client.post(
        "/workload/groups/new",
        data=_group_form(context, [class_id], [member_id]),
    )
    with app.app_context():
        group_id = TeachingGroup.query.one().id

    response = client.post(
        f"/workload/groups/{group_id}/status",
        data={
            "status": "CLOSED",
            "revision": "1",
            "close_date": "2027-01-31",
            "close_reason": "Изменение учебного плана",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        group = db.session.get(TeachingGroup, group_id)
        assert group.status == "CLOSED"
        assert group.valid_to == date(2027, 1, 31)
        assert group.members[0].valid_to == date(2027, 1, 31)
        assert group.close_reason == "Изменение учебного плана"


def test_metagroup_inherits_sources_and_replaces_them_in_workload(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        sources = [
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=index,
            )
            for index, snapshot_class in enumerate(classes, start=1)
        ]
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        plan = line.education_plan
        metagroup = create_metagroup(
            version=db.session.get(
                TariffVersion,
                context["version_id"],
            ),
            snapshot=snapshot,
            plans=[plan],
            source_tokens=[
                f"group:{group.id}" for group in sources
            ],
            name="Математика · метагруппа 5-х классов",
            user_id=user_id,
        )
        db.session.commit()

        assert metagroup.group_type == "METAGROUP"
        assert metagroup.status == "READY"
        assert metagroup.actual_size == 3
        assert len(metagroup.members) == 3
        assert len(metagroup.source_classes) == 2
        assert len(metagroup.metagroup_sources) == 2
        assert TeachingMetagroupSource.query.count() == 2

        result = generate_plan_needs(
            metagroup.tariff_version,
            user_id=user_id,
        )
        db.session.commit()

        needs = WorkloadNeed.query.filter_by(
            tariff_version_id=context["version_id"],
            status="OPEN",
        ).all()
        assert result["created"] == 1
        assert len(needs) == 1
        assert needs[0].teaching_group_id == metagroup.id
        assert [source.source_kind for source in needs[0].sources] == [
            "MERGE"
        ]


def test_metagroup_replaces_unassigned_source_needs(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        sources = [
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=index,
            )
            for index, snapshot_class in enumerate(classes, start=1)
        ]
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        version = line.education_plan.tariff_version

        generated = generate_plan_needs(version, user_id=user_id)
        db.session.flush()
        source_ids = {group.id for group in sources}
        source_needs = WorkloadNeed.query.filter(
            WorkloadNeed.teaching_group_id.in_(source_ids),
            WorkloadNeed.status != "CANCELLED",
        ).all()
        assert generated["created"] == 2
        assert len(source_needs) == 2
        assert WorkloadAssignment.query.count() == 0

        metagroup = create_metagroup(
            version=version,
            snapshot=snapshot,
            plans=[line.education_plan],
            source_tokens=[
                f"group:{group.id}" for group in sources
            ],
            name="Математика · объединённая группа",
            user_id=user_id,
        )
        db.session.flush()

        assert all(need.status == "CANCELLED" for need in source_needs)
        active_needs = WorkloadNeed.query.filter(
            WorkloadNeed.status != "CANCELLED",
        ).all()
        assert len(active_needs) == 1
        assert active_needs[0].teaching_group_id == metagroup.id


def test_metagroup_combines_one_class_across_two_curricula(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        first_line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        first_plan = first_line.education_plan
        first_plan.education_level = "SOO"
        first_scope = first_line.scopes[0]
        first_scope.grade = 10
        first_scope.scope_key = line_scope_key(
            "GRADE",
            grade=10,
            building_id=context["building_id"],
        )
        school_classes = (
            SchoolClass.query
            .filter_by(academic_year_id=context["year_id"])
            .order_by(SchoolClass.id.asc())
            .all()
        )
        for school_class, name, letter in zip(
            school_classes,
            ("10З", "10Б"),
            ("З", "Б"),
        ):
            school_class.grade = 10
            school_class.name = name
            school_class.letter = letter
        db.session.commit()
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        snapshot_class = next(
            item
            for item in snapshot.classes
            if item.name_snapshot == "10З"
        )
        second_plan = EducationPlan(
            tariff_version_id=context["version_id"],
            plan_kind="CURRICULUM",
            name="Второй профильный учебный план",
            education_level="SOO",
            building_id=context["building_id"],
            scope_code="OOO_SECOND_PROFILE",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_plan)
        db.session.flush()
        second_line = EducationPlanLine(
            education_plan_id=second_plan.id,
            education_activity_id=first_line.education_activity_id,
            component_kind="MANDATORY",
            weekly_hours=first_line.weekly_hours,
            weeks_count=first_line.weeks_count,
            annual_hours=first_line.annual_hours,
            requires_division=True,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(second_line)
        db.session.flush()
        db.session.add(EducationPlanLineScope(
            education_plan_line_id=second_line.id,
            scope_kind="GRADE",
            grade=10,
            building_id=context["building_id"],
            scope_key=line_scope_key(
                "GRADE",
                grade=10,
                building_id=context["building_id"],
            ),
        ))
        enrollment_ids = sorted(
            item.id for item in snapshot_class.enrollments
        )
        replace_plan_binding_members(
            first_plan,
            snapshot_class,
            {enrollment_ids[0]},
            user_id=user_id,
        )
        replace_plan_binding_members(
            second_plan,
            snapshot_class,
            {enrollment_ids[1]},
            user_id=user_id,
        )
        sources = [
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=1,
                plan_line_id=first_line.id,
                member_ids={enrollment_ids[0]},
            ),
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=2,
                plan_line_id=second_line.id,
                member_ids={enrollment_ids[1]},
            ),
        ]
        plan_matrix = build_teaching_group_matrix(
            snapshot,
            [first_plan, second_plan],
            "SOO",
            context["version_id"],
            grade=10,
        )
        workspace = build_metagroup_workspace(
            plan_matrix,
            context["version_id"],
        )
        assert [
            option["activity"].name
            for option in workspace["activity_options"]
        ] == ["Математика"]

        metagroup = create_metagroup(
            version=first_plan.tariff_version,
            snapshot=snapshot,
            plans=[first_plan, second_plan],
            source_tokens=[
                f"group:{group.id}" for group in sources
            ],
            name="Математика · два профиля 10З",
            user_id=user_id,
        )
        db.session.flush()
        assert len(metagroup.source_classes) == 1
        assert len(metagroup.metagroup_sources) == 2

        generate_plan_needs(
            first_plan.tariff_version,
            user_id=user_id,
        )
        need = WorkloadNeed.query.filter(
            WorkloadNeed.status != "CANCELLED",
        ).one()
        matrix = build_workload_assignment_matrix(
            [need],
            [],
            plan_matrices=[plan_matrix],
            extra_teachers=[db.session.get(User, user_id)],
            draft_rows=[(
                db.session.get(User, user_id),
                first_line.education_activity,
                "CURRICULUM",
            )],
        )
        row = matrix["blocks"][0]["rows"][0]
        mirrored_slots = [
            slot
            for cell in row["matrix_cells"].values()
            for slot in cell["slots"]
        ]
        assert len(mirrored_slots) == 2
        assert {slot["need"].id for slot in mirrored_slots} == {need.id}


def test_metagroup_still_blocks_active_teacher_assignments(app, make_user):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        sources = [
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=index,
            )
            for index, snapshot_class in enumerate(classes, start=1)
        ]
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        version = line.education_plan.tariff_version
        generate_plan_needs(version, user_id=user_id)
        need = WorkloadNeed.query.filter_by(
            teaching_group_id=sources[0].id,
        ).one()
        db.session.add(WorkloadAssignment(
            organization_id=need.organization_id,
            tariff_version_id=version.id,
            workload_need_id=need.id,
            employee_user_id=teacher_id,
            position_code="TEACHER",
            position_title="Учитель",
            department_id=need.department_id,
            building_id=need.building_id,
            assignment_kind="MAIN",
            date_from=need.date_from,
            date_to=need.date_to,
            weekly_hours=need.weekly_hours,
            annual_hours=need.annual_hours,
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        ))
        db.session.flush()

        with pytest.raises(GroupValidationError, match="уже назначена"):
            create_metagroup(
                version=version,
                snapshot=snapshot,
                plans=[line.education_plan],
                source_tokens=[
                    f"group:{group.id}" for group in sources
                ],
                name="Математика · объединённая группа",
                user_id=user_id,
            )


def test_metagroup_constructor_filters_by_grade_and_activity(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        plan = db.session.get(
            EducationPlan,
            db.session.get(
                EducationPlanLine,
                context["plan_line_id"],
            ).education_plan_id,
        )
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        for index, snapshot_class in enumerate(classes, start=1):
            replace_plan_binding_members(
                plan,
                snapshot_class,
                {item.id for item in snapshot_class.enrollments},
                user_id=user_id,
            )
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=index,
            )
        activity_id = plan.lines[0].education_activity_id
        db.session.commit()

    login(user_id)
    base_query = {
        "version_id": context["version_id"],
        "level": "OOO",
        "grade": "5",
    }
    response = client.get(
        "/workload/groups/metagroups/",
        query_string=base_query,
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert f'value="{activity_id}:CURRICULUM"' in html
    assert 'class="metagroup-cluster"' not in html
    assert "После выбора предмета" in html

    response = client.get(
        "/workload/groups/metagroups/",
        query_string={
            **base_query,
            "activity": f"{activity_id}:CURRICULUM",
        },
    )
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert 'class="metagroup-cluster"' in html
    assert "5А" in html
    assert "5Б" in html
    assert "Excel" in html
    assert "PDF" in html

    xlsx_response = client.get(
        "/workload/groups/metagroups/export.xlsx",
        query_string=base_query,
    )
    assert xlsx_response.status_code == 200
    assert xlsx_response.data.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx_response.data), data_only=True)
    sheet = workbook["Метагруппы"]
    assert sheet["A1"].value == "Реестр метагрупп"

    pdf_response = client.get(
        "/workload/groups/metagroups/export.pdf",
        query_string=base_query,
    )
    assert pdf_response.status_code == 200
    assert pdf_response.data.startswith(b"%PDF")


def test_metagroup_can_be_planned_before_children_are_distributed(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        class_b = SchoolClass.query.filter_by(name="5Б").one()
        child = _child("Кузнецова", "Мария")
        db.session.add(ChildEnrollment(
            child_id=child.id,
            academic_year_id=class_b.academic_year_id,
            school_class_id=class_b.id,
            status="ACTIVE",
            enrolled_at=datetime(2026, 9, 1),
        ))
        db.session.commit()

        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        plan = line.education_plan
        groups_by_class = {}
        for snapshot_class in classes:
            replace_plan_binding_members(
                plan,
                snapshot_class,
                {
                    enrollment.id
                    for enrollment in snapshot_class.enrollments
                },
                user_id=user_id,
            )
            groups_by_class[snapshot_class.id] = (
                replace_teaching_group_count(
                    version=line.education_plan.tariff_version,
                    snapshot=snapshot,
                    plans=[plan],
                    plan_line_id=line.id,
                    snapshot_class_id=snapshot_class.id,
                    plan_id=plan.id,
                    group_count=2,
                    user_id=user_id,
                )
            )
        db.session.flush()

        matrix = build_teaching_group_matrix(
            snapshot,
            [plan],
            "OOO",
            context["version_id"],
            grade=5,
        )
        workspace = build_metagroup_workspace(
            matrix,
            context["version_id"],
        )
        assert [
            option["activity"].name
            for option in workspace["activity_options"]
        ] == ["Математика"]

        metagroup = create_metagroup(
            version=line.education_plan.tariff_version,
            snapshot=snapshot,
            plans=[plan],
            source_tokens=[
                f"group:{groups_by_class[snapshot_class.id][0].id}"
                for snapshot_class in classes
            ],
            name="Математика · объединённая группа",
            user_id=user_id,
        )
        db.session.commit()
        assert metagroup.status == "DRAFT"
        assert metagroup.actual_size == 0

        for index, snapshot_class in enumerate(classes):
            matrix = build_teaching_group_matrix(
                snapshot,
                [plan],
                "OOO",
                context["version_id"],
                grade=5,
            )
            composition = build_group_composition_workspace(matrix)
            item = next(
                item
                for item in composition["items"]
                if item["snapshot_class"].id == snapshot_class.id
            )
            assignments = {
                enrollment.id: item["groups"][member_index].id
                for member_index, enrollment in enumerate(
                    item["enrollments"]
                )
            }
            replace_group_composition_assignments(
                item,
                assignments,
                user_id=user_id,
            )
            db.session.commit()
            db.session.refresh(metagroup)
            assert metagroup.status == (
                "READY" if index == len(classes) - 1 else "DRAFT"
            )

        assert metagroup.actual_size == 2
        assert len(metagroup.members) == 2
        assert sorted(
            item.student_count for item in metagroup.source_classes
        ) == [1, 1]


def test_metagroup_need_is_mirrored_across_source_class_columns(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        classes = sorted(
            snapshot.classes,
            key=lambda item: item.name_snapshot,
        )
        line = db.session.get(
            EducationPlanLine,
            context["plan_line_id"],
        )
        plan = line.education_plan
        for snapshot_class in classes:
            replace_plan_binding_members(
                plan,
                snapshot_class,
                {
                    enrollment.id
                    for enrollment in snapshot_class.enrollments
                },
                user_id=user_id,
            )
        db.session.flush()
        sources = [
            _ready_source_group(
                context,
                snapshot_class,
                user_id=user_id,
                suffix=index,
            )
            for index, snapshot_class in enumerate(classes, start=1)
        ]
        metagroup = create_metagroup(
            version=db.session.get(
                TariffVersion,
                context["version_id"],
            ),
            snapshot=snapshot,
            plans=[plan],
            source_tokens=[
                f"group:{group.id}" for group in sources
            ],
            name="Математика · метагруппа 5-х классов",
            user_id=user_id,
        )
        db.session.commit()
        generate_plan_needs(
            metagroup.tariff_version,
            user_id=user_id,
        )
        db.session.commit()

        need = WorkloadNeed.query.one()
        teacher = db.session.get(User, user_id)
        plan_matrix = build_teaching_group_matrix(
            snapshot,
            [plan],
            "OOO",
            context["version_id"],
        )
        matrix = build_workload_assignment_matrix(
            [need],
            [],
            plan_matrices=[plan_matrix],
            extra_teachers=[teacher],
            draft_rows=[(
                teacher,
                line.education_activity,
                "CURRICULUM",
            )],
        )

        row = matrix["blocks"][0]["rows"][0]
        mirrored_slots = [
            slot
            for cell in row["matrix_cells"].values()
            for slot in cell["slots"]
        ]
        assert len(matrix["columns"]) == 2
        assert [slot["need"].id for slot in mirrored_slots] == [
            need.id,
            need.id,
        ]
        assert all(slot["is_metagroup"] for slot in mirrored_slots)

        lightweight_matrix = build_workload_assignment_matrix(
            [need],
            [],
            extra_teachers=[teacher],
            draft_rows=[(
                teacher,
                line.education_activity,
                "CURRICULUM",
            )],
        )
        lightweight_row = lightweight_matrix["blocks"][0]["rows"][0]
        lightweight_slots = [
            slot
            for cell in lightweight_row["matrix_cells"].values()
            for slot in cell["slots"]
        ]
        assert len(lightweight_matrix["columns"]) == 2
        assert [slot["need"].id for slot in lightweight_slots] == [
            need.id,
            need.id,
        ]


def test_metagroup_rejects_sources_from_one_class(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _group_context(user_id)
        snapshot_id = _snapshot(user_id, context["version_id"])
        snapshot = db.session.get(PopulationSnapshot, snapshot_id)
        snapshot_class = snapshot.classes[0]
        source = _ready_source_group(
            context,
            snapshot_class,
            user_id=user_id,
            suffix=1,
        )
        db.session.flush()

        with pytest.raises(
            GroupValidationError,
            match="не менее двух",
        ):
            create_metagroup(
                version=db.session.get(
                    TariffVersion,
                    context["version_id"],
                ),
                snapshot=snapshot,
                plans=[source.source_plan_line.education_plan],
                source_tokens=[
                    f"group:{source.id}",
                    f"group:{source.id}",
                ],
                name="Некорректная метагруппа",
                user_id=user_id,
            )


def test_teacher_cannot_open_group_registry(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    login(make_user("TEACHER"))

    response = client.get("/workload/groups/")

    assert response.status_code == 403
