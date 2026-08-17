from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from urllib.parse import parse_qs, urlparse

import pytest
from openpyxl import load_workbook
from werkzeug.datastructures import MultiDict

from app.core.extensions import db
from app.modules.workload.assignment_routes import (
    _workspace_department_holder_keys,
    _workspace_state_department_holder_keys,
    _workspace_teacher_metadata,
    _workspace_selected_subject_ids,
)
from app.models import (
    AcademicYear,
    Building,
    CalculationParameterSet,
    Department,
    EducationActivity,
    EducationActivityDepartment,
    EducationPlan,
    EducationPlanLine,
    EducationPlanLinePeriod,
    TariffVersion,
    TariffAllowanceRule,
    TariffCalculationRun,
    TariffCoefficientValue,
    TariffLine,
    TariffLineComponent,
    TariffDocumentArtifact,
    TariffReviewComment,
    TariffReviewDecision,
    TariffValidationRun,
    TariffRateNorm,
    TeachingGroup,
    TeacherLoad,
    TeacherAttestation,
    TeacherMckoResult,
    User,
    WorkloadAssignment,
    WorkloadAssignmentChange,
    WorkloadEditorAccess,
    WorkloadNeed,
    WorkloadNeedSource,
    WorkloadReconciliationRun,
    WorkloadSourceSetting,
    WorkloadSourceTransition,
    Subject,
    SchoolClass,
)
from app.services.education_plan_service import (
    ensure_draft_tariff_version,
    plan_scope_code,
)
from app.services.workload_distribution_service import (
    WorkloadDistributionError,
    calculate_assignment_annual_hours,
    cancel_assignment,
    delete_plan_lines_with_dependencies,
    generate_plan_needs,
    refresh_need_status,
    resolve_line_hours,
    validate_assignment,
)
from app.services.tariff_calculation_service import (
    calculate_tariff_version,
    ensure_standard_tariff_types,
)
from app.services.tariff_document_service import (
    generate_tariff_document,
    resolve_artifact_path,
)
from app.services.tariff_workflow_service import (
    TariffWorkflowError,
    answer_review_comment,
    approve_version,
    clone_correction_version,
    close_review_comment,
    record_review_decision,
    run_full_validation,
    start_review,
)
from app.services.workload_integration_service import (
    WorkloadIntegrationError,
    internal_department_load_rows,
    reconcile_workload_sources,
    source_state,
    switch_workload_source,
)
from app.services.teaching_group_display_service import (
    teaching_group_assignment_label,
    teaching_group_class_label,
)
from app.services.workload_editing_workflow_service import (
    WorkloadEditingWorkflowError,
    change_workload_approval_status,
    require_workload_editable,
)
from app.services.workload_assignment_matrix_service import (
    build_workload_assignment_matrix,
)
from app.modules.workload.assignment_routes import (
    _filter_workspace_needs,
    _workspace_matrix_specs,
)


def _distribution_context(user_id):
    year = AcademicYear(
        name="2026/2027",
        is_current=True,
        start_date=date(2026, 9, 1),
        end_date=date(2027, 8, 31),
    )
    building = Building(name="Главное здание", short_name="ГЗ")
    department = Department(
        name="Математика и информатика",
        code="math",
    )
    db.session.add_all([year, building, department])
    db.session.flush()
    activity = EducationActivity(
        code="MATH-DISTRIBUTION",
        name="Математика",
        activity_kind="SUBJECT",
        is_global=True,
        is_tariffable=True,
        is_active=True,
    )
    db.session.add(activity)
    db.session.flush()
    db.session.add(EducationActivityDepartment(
        education_activity_id=activity.id,
        department_id=department.id,
        is_primary=True,
        is_active=True,
    ))
    _, version = ensure_draft_tariff_version(year, user_id=user_id)
    plan = EducationPlan(
        tariff_version_id=version.id,
        plan_kind="CURRICULUM",
        name="Основной учебный план",
        education_level="OOO",
        building_id=building.id,
        scope_code=plan_scope_code("OOO", building.id),
        status="DRAFT",
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(plan)
    db.session.flush()
    line = EducationPlanLine(
        education_plan_id=plan.id,
        education_activity_id=activity.id,
        component_kind="MANDATORY",
        weekly_hours=Decimal("5"),
        weeks_count=Decimal("34"),
        annual_hours=Decimal("170"),
        requires_division=True,
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(line)
    db.session.flush()
    group = TeachingGroup(
        tariff_version_id=version.id,
        education_activity_id=activity.id,
        group_type="SUBGROUP",
        code="MATH_5A_1",
        name="Математика 5А, группа 1",
        composition_mode="COUNT_ONLY",
        building_id=building.id,
        planned_size=15,
        actual_size=15,
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 5, 31),
        source_plan_line_id=line.id,
        status="READY",
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(group)
    db.session.commit()
    return {
        "year_id": year.id,
        "building_id": building.id,
        "department_id": department.id,
        "version_id": version.id,
        "plan_id": plan.id,
        "group_id": group.id,
    }


def _generate(context, user_id):
    version = db.session.get(TariffVersion, context["version_id"])
    result = generate_plan_needs(version, user_id=user_id)
    db.session.commit()
    return result


def test_workspace_class_teacher_metadata_uses_class_registry_assignment(
    app,
    make_user,
):
    admin_id = make_user("ADMIN")
    workload_teacher_id = make_user("TEACHER")
    registry_teacher_id = make_user("VIEWER")
    with app.app_context():
        context = _distribution_context(admin_id)
        workload_teacher = db.session.get(User, workload_teacher_id)
        registry_teacher = db.session.get(User, registry_teacher_id)
        for teacher in (workload_teacher, registry_teacher):
            teacher.last_name = "Иванова"
            teacher.first_name = "Мария"
            teacher.middle_name = "Петровна"
        db.session.add(SchoolClass(
            academic_year_id=context["year_id"],
            building_id=context["building_id"],
            name="5А",
            grade=5,
            letter="А",
            teacher_user_id=registry_teacher_id,
            is_active=True,
            is_archived=False,
        ))
        db.session.commit()

        version = db.session.get(TariffVersion, context["version_id"])
        metadata = _workspace_teacher_metadata(
            [workload_teacher],
            version,
        )

        assert metadata[workload_teacher_id]["class_teacher"] == "5А"


def _assignment(need, employee_id, weekly, annual=None, kind="MAIN"):
    weekly = Decimal(weekly)
    annual = (
        Decimal(annual)
        if annual is not None
        else calculate_assignment_annual_hours(need, weekly, None)
    )
    return WorkloadAssignment(
        organization_id=need.organization_id,
        tariff_version_id=need.tariff_version_id,
        workload_need_id=need.id,
        employee_user_id=employee_id if kind != "VACANCY" else None,
        position_code="TEACHER",
        position_title="Учитель" if kind != "VACANCY" else "Вакансия",
        department_id=need.department_id,
        building_id=need.building_id,
        assignment_kind=kind,
        date_from=need.date_from,
        date_to=need.date_to,
        weekly_hours=weekly,
        annual_hours=annual,
        status="DRAFT",
    )


def test_resolve_line_hours_uses_period_active_at_start_of_full_year():
    line = EducationPlanLine(
        weekly_hours=Decimal("5"),
        weeks_count=Decimal("34"),
        annual_hours=Decimal("161"),
    )
    line.periods = [
        EducationPlanLinePeriod(
            date_from=date(2026, 9, 1),
            date_to=date(2026, 10, 31),
            weekly_hours=Decimal("4"),
            weeks_count=Decimal("9"),
            annual_hours=Decimal("36"),
        ),
        EducationPlanLinePeriod(
            date_from=date(2026, 11, 1),
            date_to=date(2027, 5, 31),
            weekly_hours=Decimal("5"),
            weeks_count=Decimal("25"),
            annual_hours=Decimal("125"),
        ),
    ]

    weekly, annual = resolve_line_hours(
        line,
        date(2026, 9, 1),
        date(2027, 5, 31),
    )

    assert weekly == Decimal("4.000")
    assert annual == Decimal("161.000")


def test_workspace_repairs_existing_need_with_wrong_start_period_hours(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        line = EducationPlanLine.query.one()
        line.weeks_count = Decimal("34")
        line.annual_hours = Decimal("161")
        line.periods = [
            EducationPlanLinePeriod(
                date_from=date(2026, 9, 1),
                date_to=date(2026, 10, 31),
                weekly_hours=Decimal("4"),
                weeks_count=Decimal("9"),
                annual_hours=Decimal("36"),
            ),
            EducationPlanLinePeriod(
                date_from=date(2026, 11, 1),
                date_to=date(2027, 5, 31),
                weekly_hours=Decimal("5"),
                weeks_count=Decimal("25"),
                annual_hours=Decimal("125"),
            ),
        ]
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        need.weekly_hours = Decimal("5")
        # Production needs are commonly newer than their source plans because
        # teachers are assigned after plan generation.  A newer timestamp must
        # not hide a mismatch between the stored and start-period hours.
        need.updated_at = datetime(2030, 1, 1)
        db.session.commit()
        need_id = need.id
    login(admin_id)

    response = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    )

    assert response.status_code == 200
    with app.app_context():
        repaired = db.session.get(WorkloadNeed, need_id)
        assert repaired.weekly_hours == Decimal("4.000")
        assert repaired.annual_hours == Decimal("161.000")


def test_generate_plan_needs_removes_line_inactive_at_year_start(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        line = EducationPlanLine.query.one()
        line.weekly_hours = Decimal("1")
        line.weeks_count = Decimal("26")
        line.annual_hours = Decimal("26")
        _generate(context, admin_id)
        need_id = WorkloadNeed.query.one().id

        line.periods = [
            EducationPlanLinePeriod(
                date_from=date(2026, 9, 1),
                date_to=date(2026, 10, 31),
                weekly_hours=Decimal("0"),
                weeks_count=Decimal("9"),
                annual_hours=Decimal("0"),
            ),
            EducationPlanLinePeriod(
                date_from=date(2026, 11, 1),
                date_to=date(2027, 5, 31),
                weekly_hours=Decimal("1"),
                weeks_count=Decimal("26"),
                annual_hours=Decimal("26"),
            ),
        ]
        db.session.commit()

    login(admin_id)
    response = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    )

    assert response.status_code == 200
    with app.app_context():
        need = db.session.get(WorkloadNeed, need_id)
        assert need.status == "CANCELLED"
        assert WorkloadNeed.query.filter(
            WorkloadNeed.status != "CANCELLED"
        ).count() == 0


def test_workspace_filter_builds_matrices_for_multiple_levels_and_grades():
    assert _workspace_matrix_specs(set(), {3, 9, 10}) == [
        ("NOO", 3),
        ("OOO", 9),
        ("SOO", 10),
    ]
    assert _workspace_matrix_specs({"OOO", "SOO"}, set()) == [
        ("OOO", None),
        ("SOO", None),
    ]
    assert _workspace_matrix_specs({"OOO"}, {3, 9, 10}) == [
        ("OOO", 9),
    ]


def test_workspace_uses_only_needs_from_current_population_snapshot():
    def need(need_id, snapshot_id):
        snapshot_class = SimpleNamespace(
            population_snapshot_id=snapshot_id,
        )
        source_link = SimpleNamespace(
            population_snapshot_class=snapshot_class,
        )
        group = SimpleNamespace(
            group_type="CLASS",
            source_classes=[source_link],
        )
        return SimpleNamespace(
            id=need_id,
            teaching_group=group,
            building_id=None,
        )

    current_need = need(1, 20)
    stale_need = need(2, 19)

    assert _filter_workspace_needs(
        [stale_need, current_need],
        population_snapshot_id=20,
    ) == [current_need]


def test_workspace_hides_needs_from_plan_no_longer_bound_to_class():
    snapshot_class = SimpleNamespace(population_snapshot_id=20)
    source_link = SimpleNamespace(
        population_snapshot_class_id=101,
        population_snapshot_class=snapshot_class,
    )
    plan = SimpleNamespace(id=301, root_plan_id=None)
    line = SimpleNamespace(education_plan=plan)
    group = SimpleNamespace(
        group_type="CLASS",
        source_plan_line=line,
        source_classes=[source_link],
    )
    need = SimpleNamespace(
        id=1,
        teaching_group=group,
        building_id=None,
    )

    assert _filter_workspace_needs(
        [need],
        population_snapshot_id=20,
        bound_plan_ids_by_class={101: {301}},
    ) == [need]
    assert _filter_workspace_needs(
        [need],
        population_snapshot_id=20,
        bound_plan_ids_by_class={101: {302}},
    ) == []


def test_department_membership_uses_only_curriculum_then_keeps_full_load():
    department_id = 7

    def need(
        need_id,
        plan_kind,
        linked_department_id,
        activity_name="Обычный предмет",
    ):
        plan = SimpleNamespace(plan_kind=plan_kind, root_plan=None)
        line = SimpleNamespace(education_plan=plan)
        activity = SimpleNamespace(
            name=activity_name,
            department_links=[SimpleNamespace(
                department_id=linked_department_id,
                is_active=True,
            )],
        )
        return SimpleNamespace(
            id=need_id,
            department_id=None,
            education_activity=activity,
            teaching_group=SimpleNamespace(source_plan_line=line),
        )

    curriculum_need = need(1, "CURRICULUM", department_id)
    other_department_need = need(2, "CURRICULUM", 8)
    extracurricular_need = need(3, "EXTRACURRICULAR", department_id)
    additional_need = need(4, "ADDITIONAL_EDUCATION", department_id)
    individual_project_need = need(
        5,
        "CURRICULUM",
        department_id,
        "Индивидуальный проект",
    )
    assignments = [
        SimpleNamespace(
            workload_need_id=1,
            assignment_kind="MAIN",
            employee_user_id=101,
            position_code="TEACHER",
        ),
        SimpleNamespace(
            workload_need_id=2,
            assignment_kind="MAIN",
            employee_user_id=101,
            position_code="TEACHER",
        ),
        SimpleNamespace(
            workload_need_id=3,
            assignment_kind="MAIN",
            employee_user_id=102,
            position_code="TEACHER",
        ),
        SimpleNamespace(
            workload_need_id=4,
            assignment_kind="MAIN",
            employee_user_id=103,
            position_code="TEACHER",
        ),
        SimpleNamespace(
            workload_need_id=5,
            assignment_kind="MAIN",
            employee_user_id=104,
            position_code="TEACHER",
        ),
    ]
    needs_by_id = {
        item.id: item
        for item in (
            curriculum_need,
            other_department_need,
            extracurricular_need,
            additional_need,
            individual_project_need,
        )
    }

    holder_keys = _workspace_department_holder_keys(
        assignments,
        needs_by_id,
        department_id,
    )
    visible_assignments = [
        assignment
        for assignment in assignments
        if f"teacher:{assignment.employee_user_id}" in holder_keys
    ]

    assert holder_keys == {"teacher:101"}
    assert [item.workload_need_id for item in visible_assignments] == [1, 2]


def test_draft_department_membership_ignores_non_curriculum_rows():
    department_id = 7
    activity = SimpleNamespace(
        name="Обычный предмет",
        department_links=[SimpleNamespace(
            department_id=department_id,
            is_active=True,
        )],
    )
    individual_project = SimpleNamespace(
        name="Индивидуальный проект",
        department_links=activity.department_links,
    )
    rows = [
        {
            "holder_type": "teacher",
            "teacher_id": 101,
            "activity_id": 1,
            "plan_kind": "CURRICULUM",
        },
        {
            "holder_type": "teacher",
            "teacher_id": 102,
            "activity_id": 1,
            "plan_kind": "EXTRACURRICULAR",
        },
        {
            "holder_type": "teacher",
            "teacher_id": 103,
            "activity_id": 1,
            "plan_kind": "ADDITIONAL_EDUCATION",
        },
        {
            "holder_type": "teacher",
            "teacher_id": 104,
            "activity_id": 2,
            "plan_kind": "CURRICULUM",
        },
    ]

    assert _workspace_state_department_holder_keys(
        rows,
        {1: activity, 2: individual_project},
        department_id,
    ) == {"teacher:101"}


def test_workspace_merges_non_profile_plan_columns_and_keeps_global_totals():
    building = SimpleNamespace(
        id=1,
        name="Главное здание",
        short_name="ГЗ",
        matrix_tone=0,
    )
    snapshot_class = SimpleNamespace(
        id=101,
        name_snapshot="9А",
        grade_snapshot=9,
        building_id=building.id,
        building_name_snapshot=building.name,
        building=building,
    )
    activity = SimpleNamespace(
        id=201,
        name="Разговоры о важном",
        department_links=[],
    )
    root_plan_a = SimpleNamespace(
        id=301,
        name="5–9 ООО",
        education_level="OOO",
        plan_kind="CURRICULUM",
        root_plan=None,
    )
    root_plan_b = SimpleNamespace(
        id=302,
        name="5–9 ООО",
        education_level="OOO",
        plan_kind="CURRICULUM",
        root_plan=None,
    )
    extra_plan_a = SimpleNamespace(
        id=303,
        name="5–9 ООО · Внеурочная деятельность",
        plan_kind="EXTRACURRICULAR",
        root_plan=root_plan_a,
    )
    line = SimpleNamespace(education_plan=extra_plan_a)
    class_link = SimpleNamespace(
        population_snapshot_class=snapshot_class,
        population_snapshot_class_id=snapshot_class.id,
    )
    group = SimpleNamespace(
        id=401,
        name="9А · внеурочная деятельность",
        group_type="EXTRACURRICULAR_GROUP",
        source_plan_line=line,
        source_classes=[class_link],
        metagroup_sources=[],
        metagroup_membership=None,
        building=building,
        building_id=building.id,
    )
    need = SimpleNamespace(
        id=501,
        education_activity_id=activity.id,
        education_activity=activity,
        teaching_group=group,
        teaching_group_id=group.id,
        weekly_hours=Decimal("1"),
        allocated_weekly_hours=Decimal("1"),
        remaining_weekly_hours=Decimal("0"),
    )
    other_group = SimpleNamespace(
        **{
            **group.__dict__,
            "id": 402,
            "name": "8А · внеурочная деятельность",
        }
    )
    other_need = SimpleNamespace(
        **{
            **need.__dict__,
            "id": 502,
            "teaching_group": other_group,
            "teaching_group_id": other_group.id,
        }
    )
    teacher = SimpleNamespace(id=601, fio="Учитель Тестовый")

    def assignment(assignment_id, target_need, hours):
        return SimpleNamespace(
            id=assignment_id,
            status="DRAFT",
            assignment_kind="MAIN",
            position_code="TEACHER",
            employee_user_id=teacher.id,
            employee=teacher,
            workload_need_id=target_need.id,
            workload_need=target_need,
            weekly_hours=Decimal(hours),
        )

    source_columns = [
        {
            "key": f"class-{snapshot_class.id}-plan-{plan.id}",
            "is_unassigned": False,
            "is_profile_column": False,
            "plan": plan,
            "class_display_name": snapshot_class.name_snapshot,
        }
        for plan in (root_plan_a, root_plan_b)
    ]
    plan_matrix = {
        "class_groups": [{
            "snapshot_class": snapshot_class,
            "columns": source_columns,
        }],
        "sections": [{
            "rows": [{
                "activity": activity,
                "plan_kind": "EXTRACURRICULAR",
                "cells": {
                    source_columns[0]["key"]: {"groups": [group]},
                    source_columns[1]["key"]: {"groups": [group]},
                },
            }],
        }],
    }
    visible_assignment = assignment(701, need, "1")
    matrix = build_workload_assignment_matrix(
        [need],
        [visible_assignment],
        plan_matrices=[plan_matrix],
        total_assignments=[
            visible_assignment,
            assignment(702, other_need, "2"),
        ],
    )

    assert len(matrix["columns"]) == 1
    assert len(matrix["class_groups"]) == 1
    assert len(matrix["class_groups"][0]["columns"]) == 1
    assert matrix["columns"][0]["subheader_label"] == "5–9 ООО"
    assert matrix["total_allocated"] == Decimal("1")
    assert matrix["total_remaining"] == Decimal("0")
    assert matrix["blocks"][0]["total"] == Decimal("3")
    assert matrix["blocks"][0]["rows"][0]["total"] == Decimal("3")

    paged_matrix = build_workload_assignment_matrix(
        [need],
        [visible_assignment],
        plan_matrices=[plan_matrix],
        total_assignments=[visible_assignment],
        visible_holder_keys=set(),
    )
    assert paged_matrix["blocks"] == []
    assert paged_matrix["total_allocated"] == Decimal("1")


def test_workspace_subject_filter_accepts_multiple_values():
    selected = _workspace_selected_subject_ids(MultiDict([
        ("subject_id", "12"),
        ("subject_id", "35"),
        ("subject_id", "12"),
        ("subject_id", "bad"),
    ]))

    assert selected == {12, 35}


def test_workload_save_submit_approve_and_return_cycle(
    app,
    make_user,
):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        version = db.session.get(TariffVersion, context["version_id"])
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "5")
        db.session.add(assignment)
        db.session.flush()

        change_workload_approval_status(
            version,
            "SAVE",
            user_id=admin_id,
        )
        with pytest.raises(WorkloadEditingWorkflowError):
            require_workload_editable(version)
        change_workload_approval_status(
            version,
            "SUBMIT",
            user_id=admin_id,
        )
        change_workload_approval_status(
            version,
            "APPROVE",
            user_id=admin_id,
        )
        assert version.workload_approval_status == "APPROVED"
        assert assignment.status == "CONFIRMED"

        change_workload_approval_status(
            version,
            "REQUEST_CHANGES",
            user_id=admin_id,
            comment="Уточнить распределение.",
        )
        assert version.workload_approval_status == "CHANGES_REQUESTED"
        assert version.workload_review_comment == "Уточнить распределение."
        assert assignment.status == "DRAFT"
        change_workload_approval_status(
            version,
            "EDIT",
            user_id=admin_id,
        )
        require_workload_editable(version)


def test_director_approves_workload_through_workspace_route(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    director_id = make_user("DIRECTOR")
    with app.app_context():
        context = _distribution_context(director_id)

    login(director_id)
    for action, expected in (
        ("SAVE", "SAVED"),
        ("SUBMIT", "PENDING_APPROVAL"),
        ("APPROVE", "APPROVED"),
    ):
        response = client.post(
            "/workload/assignments/workspace/status",
            data={
                "version_id": context["version_id"],
                "action": action,
                "view": "all",
            },
        )
        assert response.status_code == 302
        with app.app_context():
            version = db.session.get(
                TariffVersion,
                context["version_id"],
            )
            assert version.workload_approval_status == expected


def test_teacher_profile_marks_workload_preliminary_or_approved(
    app,
    client,
    make_user,
    login,
):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        db.session.add(_assignment(need, teacher_id, "5"))
        version = db.session.get(TariffVersion, context["version_id"])
        version.workload_approval_status = "SAVED"
        db.session.commit()

    login(admin_id)
    response = client.get(
        f"/departments/teachers/{teacher_id}",
        query_string={"academic_year_id": context["year_id"]},
    )
    assert response.status_code == 200
    assert "Предварительная" in response.get_data(as_text=True)
    assert "registry-matrix" in response.get_data(as_text=True)
    assert "<th>Класс</th>" in response.get_data(as_text=True)
    assert "Весь класс / группа" in response.get_data(as_text=True)
    assert "5А" in response.get_data(as_text=True)
    assert "Группа 1" in response.get_data(as_text=True)

    with app.app_context():
        version = db.session.get(TariffVersion, context["version_id"])
        version.workload_approval_status = "APPROVED"
        db.session.commit()

    response = client.get(
        f"/departments/teachers/{teacher_id}",
        query_string={"academic_year_id": context["year_id"]},
    )
    assert response.status_code == 200
    assert "Согласованная" in response.get_data(as_text=True)


def test_generate_need_from_ready_group(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(user_id)

        result = _generate(context, user_id)

        need = WorkloadNeed.query.one()
        assert result["created"] == 1
        assert need.weekly_hours == Decimal("5.000")
        assert need.annual_hours == Decimal("170.000")
        assert need.department_id == context["department_id"]
        assert need.status == "OPEN"
        source = WorkloadNeedSource.query.one()
        assert source.source_kind == "DIVISION"


def test_need_generation_is_idempotent(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)

        second = _generate(context, user_id)

        assert second["created"] == 0
        assert WorkloadNeed.query.count() == 1
        assert WorkloadNeedSource.query.count() == 1


def test_need_generation_updates_full_teacher_assignment_from_plan(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "5")
        validate_assignment(need, assignment)
        db.session.add(assignment)
        db.session.flush()
        refresh_need_status(need)
        db.session.commit()

        line = EducationPlanLine.query.one()
        line.weekly_hours = Decimal("6")
        line.annual_hours = Decimal("204")
        db.session.commit()
        result = _generate(context, user_id)

        assignment = WorkloadAssignment.query.one()
        need = WorkloadNeed.query.one()
        assert result["updated"] == 1
        assert need.weekly_hours == Decimal("6.000")
        assert need.status == "COVERED"
        assert assignment.weekly_hours == Decimal("6.000")
        assert assignment.annual_hours == Decimal("204.000")
        assert assignment.revision == 2
        change = WorkloadAssignmentChange.query.one()
        assert change.change_kind == "UPDATE"
        assert "учебным планом" in change.reason


def test_delete_plan_line_removes_groups_needs_and_assignments(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "5")
        validate_assignment(need, assignment)
        db.session.add(assignment)
        db.session.commit()

        line = EducationPlanLine.query.one()
        deleted = delete_plan_lines_with_dependencies([line])
        db.session.commit()

        assert deleted == {"groups": 1, "needs": 1, "assignments": 1}
        assert EducationPlanLine.query.count() == 0
        assert TeachingGroup.query.count() == 0
        assert WorkloadNeed.query.count() == 0
        assert WorkloadNeedSource.query.count() == 0
        assert WorkloadAssignment.query.count() == 0


def test_delete_plan_line_detaches_independent_copied_line(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(user_id)
        source_line = EducationPlanLine.query.one()
        source_plan = source_line.education_plan
        copied_plan = EducationPlan(
            tariff_version_id=source_plan.tariff_version_id,
            plan_kind=source_plan.plan_kind,
            name="Независимая копия плана",
            education_level=source_plan.education_level,
            scope_code="COPY",
            status="DRAFT",
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(copied_plan)
        db.session.flush()
        copied_line = EducationPlanLine(
            education_plan_id=copied_plan.id,
            education_activity_id=source_line.education_activity_id,
            component_kind=source_line.component_kind,
            weekly_hours=0,
            annual_hours=0,
            source_line_id=source_line.id,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(copied_line)
        db.session.commit()
        copied_line_id = copied_line.id

        deleted = delete_plan_lines_with_dependencies([source_line])
        db.session.commit()

        assert deleted == {"groups": 1, "needs": 0, "assignments": 0}
        assert db.session.get(EducationPlanLine, source_line.id) is None
        preserved_copy = db.session.get(EducationPlanLine, copied_line_id)
        assert preserved_copy is not None
        assert preserved_copy.source_line_id is None


def test_plan_matrix_delete_route_cascades_to_workload(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "5")
        validate_assignment(need, assignment)
        db.session.add(assignment)
        activity_id = need.education_activity_id
        db.session.commit()
    login(user_id)

    response = client.post(
        f"/workload/plans/{context['plan_id']}/matrix/rows/delete",
        data={
            "revision": "1",
            "education_activity_id": str(activity_id),
            "component_kind": "MANDATORY",
            "profile_code": "",
        },
    )

    assert response.status_code == 302
    with app.app_context():
        assert EducationPlanLine.query.count() == 0
        assert TeachingGroup.query.count() == 0
        assert WorkloadNeed.query.count() == 0
        assert WorkloadAssignment.query.count() == 0


def test_need_generation_skips_plan_line_without_hours(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(user_id)
        line = EducationPlanLine.query.one()
        line.weekly_hours = Decimal("0")
        line.annual_hours = Decimal("0")
        db.session.commit()

        result = _generate(context, user_id)

        assert result["created"] == 0
        assert result["skipped_empty"] == 1
        assert WorkloadNeed.query.count() == 0


def test_partial_assignment_updates_need_status(app, make_user):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "3")
        validate_assignment(need, assignment)
        db.session.add(assignment)
        db.session.flush()
        refresh_need_status(need)
        db.session.commit()

        assert assignment.annual_hours == Decimal("102.000")
        assert need.status == "PARTIAL"
        assert need.remaining_weekly_hours == Decimal("2.000")


def test_overallocation_is_rejected(app, make_user):
    user_id = make_user("ADMIN")
    first_teacher = make_user("TEACHER")
    second_teacher = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        first = _assignment(need, first_teacher, "4")
        validate_assignment(need, first)
        db.session.add(first)
        db.session.flush()

        with pytest.raises(
            WorkloadDistributionError,
            match="превышают потребность",
        ):
            validate_assignment(
                need,
                _assignment(need, second_teacher, "2"),
            )


def test_vacancy_does_not_cover_need(app, make_user):
    user_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        vacancy = _assignment(
            need,
            None,
            "5",
            annual="170",
            kind="VACANCY",
        )
        validate_assignment(need, vacancy)
        db.session.add(vacancy)
        db.session.flush()
        refresh_need_status(need)

        assert need.status == "OPEN"
        assert need.remaining_weekly_hours == Decimal("5.000")


def test_cancel_assignment_reopens_need_and_keeps_history(
    app,
    make_user,
):
    user_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(user_id)
        _generate(context, user_id)
        need = WorkloadNeed.query.one()
        assignment = _assignment(need, teacher_id, "5")
        validate_assignment(need, assignment)
        db.session.add(assignment)
        db.session.flush()
        refresh_need_status(need)
        assert need.status == "COVERED"

        cancel_assignment(
            assignment,
            user_id=user_id,
            expected_revision=1,
            reason="Передача нагрузки",
        )
        db.session.commit()

        assert assignment.status == "CANCELLED"
        assert need.status == "OPEN"
        change = WorkloadAssignmentChange.query.one()
        assert change.change_kind == "CANCEL"
        assert change.reason == "Передача нагрузки"


def test_admin_creates_assignment_through_route(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        db.session.get(Building, context["building_id"]).matrix_tone = 1
        _generate(context, admin_id)
        need_id = WorkloadNeed.query.one().id
    login(admin_id)

    response = client.post(
        f"/workload/needs/{need_id}/assignments/new",
        data={
            "assignment_kind": "MAIN",
            "employee_user_id": str(teacher_id),
            "position_title": "Учитель математики",
            "weekly_hours": "5",
            "annual_hours": "",
            "date_from": "2026-09-01",
            "date_to": "2027-05-31",
            "department_id": str(context["department_id"]),
            "building_id": str(context["building_id"]),
        },
    )

    assert response.status_code == 302
    with app.app_context():
        assignment = WorkloadAssignment.query.one()
        assert assignment.employee_user_id == teacher_id
        assert assignment.annual_hours == Decimal("170.000")
        assert assignment.workload_need.status == "COVERED"
        assert WorkloadAssignmentChange.query.one().change_kind == "CREATE"

    assert client.get("/workload/assignments/").status_code == 200
    workspace_response = client.get("/workload/assignments/workspace")
    workspace_html = workspace_response.get_data(as_text=True)
    assert workspace_response.status_code == 200
    assert "workload_distribution.css" in workspace_html
    assert "data-workload-matrix" in workspace_html
    assert "data-matrix-subject-column" in workspace_html
    assert "data-matrix-total-column" in workspace_html
    assert "data-matrix-class-column" in workspace_html
    assert "ФИО преподавателя" in workspace_html
    assert "По предмету" in workspace_html
    assert "data-workload-secondary-headers=" in workspace_html
    assert "building-tone-1" in workspace_html
    assert workspace_html.index(
        "data-matrix-subject-column"
    ) < workspace_html.index(
        "data-matrix-total-column"
    ) < workspace_html.index(
        "data-matrix-class-column"
    )
    assert "data-need-context" not in workspace_html
    assert "Назначено" in workspace_html
    assert 'name="version_id"' in workspace_html
    assert "data-workload-filter-auto" in workspace_html
    assert "workload-filterbar__primary" in workspace_html
    assert 'aria-label="Вид нагрузки"' in workspace_html
    assert "class-plan-levels" in workspace_html
    assert "class-plan-grades" in workspace_html
    assert 'id="workspace-view"' not in workspace_html
    assert ">Показать<" not in workspace_html
    assert "2026/2027" in workspace_html
    assert "версия 1" not in workspace_html.lower()
    assert 'data-active-mode="workload"' in workspace_html
    assert client.get(f"/workload/needs/{need_id}").status_code == 200
    assert client.get("/workload/teachers/").status_code == 200
    assert client.get(
        f"/workload/teachers/{teacher_id}"
    ).status_code == 200
    assert client.get("/workload/departments/").status_code == 200


def test_assignment_workspace_hides_unassigned_pseudo_teacher(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
    login(admin_id)

    response = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "data-workload-matrix" in html
    assert "Не распределено" not in html
    assert "не назначено" not in html
    assert "workload-add-teacher-row" in html
    assert "Добавить преподавателя" in html
    assert "workload-matrix-head-add" not in html
    assert "Какие часы не назначены" in html
    assert "Неназначенные часы" in html
    assert "Математика" in html
    assert "Учебный план" in html


def test_department_view_selects_by_curriculum_and_shows_full_teacher_load(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = False
    admin_id = make_user("ADMIN")
    matching_teacher_id = make_user("TEACHER")
    extracurricular_only_teacher_id = make_user("TEACHER")
    additional_only_teacher_id = make_user("TEACHER")
    project_only_teacher_id = make_user("TEACHER")

    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        version = db.session.get(TariffVersion, context["version_id"])
        building = db.session.get(Building, context["building_id"])
        math_department = db.session.get(
            Department,
            context["department_id"],
        )
        language_department = Department(
            name="Кафедра словесности",
            code="language-department-filter",
        )
        db.session.add(language_department)
        db.session.flush()

        matching_teacher = db.session.get(User, matching_teacher_id)
        matching_teacher.last_name = "Кафедральный"
        matching_teacher.first_name = "Подходящий"
        extracurricular_only_teacher = db.session.get(
            User,
            extracurricular_only_teacher_id,
        )
        extracurricular_only_teacher.last_name = "Внеурочный"
        extracurricular_only_teacher.first_name = "Только"
        additional_only_teacher = db.session.get(
            User,
            additional_only_teacher_id,
        )
        additional_only_teacher.last_name = "Дополнительный"
        additional_only_teacher.first_name = "Только"
        project_only_teacher = db.session.get(User, project_only_teacher_id)
        project_only_teacher.last_name = "Проектный"
        project_only_teacher.first_name = "Только"

        def add_need(
            code,
            name,
            plan_kind,
            group_type,
            department,
            component_kind,
        ):
            activity = EducationActivity(
                code=code,
                name=name,
                activity_kind="COURSE",
                is_global=True,
                is_tariffable=True,
                is_active=True,
            )
            db.session.add(activity)
            db.session.flush()
            db.session.add(EducationActivityDepartment(
                education_activity_id=activity.id,
                department_id=department.id,
                is_primary=True,
                is_active=True,
            ))
            plan = EducationPlan(
                tariff_version_id=version.id,
                root_plan_id=(
                    context["plan_id"]
                    if plan_kind != "CURRICULUM" else None
                ),
                plan_kind=plan_kind,
                name=f"{name} — план",
                education_level="OOO",
                building_id=building.id,
                scope_code=f"TEST-{code}",
                status="DRAFT",
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            )
            db.session.add(plan)
            db.session.flush()
            line = EducationPlanLine(
                education_plan_id=plan.id,
                education_activity_id=activity.id,
                component_kind=component_kind,
                weekly_hours=Decimal("1"),
                weeks_count=Decimal("34"),
                annual_hours=Decimal("34"),
                requires_division=False,
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            )
            db.session.add(line)
            db.session.flush()
            group = TeachingGroup(
                tariff_version_id=version.id,
                education_activity_id=activity.id,
                group_type=group_type,
                code=f"GROUP-{code}",
                name=f"Группа {name}",
                composition_mode="COUNT_ONLY",
                building_id=building.id,
                department_id=department.id,
                planned_size=10,
                actual_size=10,
                valid_from=date(2026, 9, 1),
                valid_to=date(2027, 5, 31),
                source_plan_line_id=line.id,
                status="READY",
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            )
            db.session.add(group)
            db.session.flush()
            need = WorkloadNeed(
                tariff_version_id=version.id,
                teaching_group_id=group.id,
                education_activity_id=activity.id,
                department_id=department.id,
                building_id=building.id,
                date_from=date(2026, 9, 1),
                date_to=date(2027, 5, 31),
                weekly_hours=Decimal("1"),
                annual_hours=Decimal("34"),
                need_kind="PLAN",
                status="COVERED",
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            )
            db.session.add(need)
            db.session.flush()
            return need

        language_need = add_need(
            "LANGUAGE-FULL-LOAD",
            "Литература другой кафедры",
            "CURRICULUM",
            "CLASS",
            language_department,
            "MANDATORY",
        )
        extracurricular_need = add_need(
            "EXTRA-DEPARTMENT-FILTER",
            "Общий внеурочный курс",
            "EXTRACURRICULAR",
            "EXTRACURRICULAR_GROUP",
            math_department,
            "EXTRACURRICULAR",
        )
        extracurricular_line = (
            extracurricular_need.teaching_group.source_plan_line
        )
        other_class_extracurricular_group = TeachingGroup(
            tariff_version_id=version.id,
            education_activity_id=(
                extracurricular_need.education_activity_id
            ),
            group_type="EXTRACURRICULAR_GROUP",
            code="GROUP-EXTRA-OTHER-CLASS",
            name="Другой класс · общий внеурочный курс",
            composition_mode="COUNT_ONLY",
            building_id=building.id,
            department_id=math_department.id,
            planned_size=10,
            actual_size=10,
            valid_from=date(2026, 9, 1),
            valid_to=date(2027, 5, 31),
            source_plan_line_id=extracurricular_line.id,
            status="READY",
            created_by_user_id=admin_id,
            updated_by_user_id=admin_id,
        )
        db.session.add(other_class_extracurricular_group)
        db.session.flush()
        other_class_extracurricular_need = WorkloadNeed(
            tariff_version_id=version.id,
            teaching_group_id=other_class_extracurricular_group.id,
            education_activity_id=(
                extracurricular_need.education_activity_id
            ),
            department_id=math_department.id,
            building_id=building.id,
            date_from=date(2026, 9, 1),
            date_to=date(2027, 5, 31),
            weekly_hours=Decimal("1"),
            annual_hours=Decimal("34"),
            need_kind="PLAN",
            status="COVERED",
            created_by_user_id=admin_id,
            updated_by_user_id=admin_id,
        )
        db.session.add(other_class_extracurricular_need)
        db.session.flush()
        additional_need = add_need(
            "ADDITIONAL-DEPARTMENT-FILTER",
            "Общий дополнительный курс",
            "ADDITIONAL_EDUCATION",
            "ADDITIONAL_GROUP",
            math_department,
            "ADDITIONAL",
        )
        individual_project_need = add_need(
            "INDIVIDUAL-PROJECT-DEPARTMENT-FILTER",
            "Индивидуальный проект",
            "CURRICULUM",
            "CLASS",
            math_department,
            "MANDATORY",
        )
        math_need = WorkloadNeed.query.filter_by(
            education_activity_id=(
                EducationActivity.query.filter_by(
                    code="MATH-DISTRIBUTION"
                ).one().id
            )
        ).one()
        math_need.status = "COVERED"

        def assign(need, teacher_id, department):
            db.session.add(WorkloadAssignment(
                tariff_version_id=version.id,
                workload_need_id=need.id,
                employee_user_id=teacher_id,
                position_code="TEACHER",
                department_id=department.id,
                building_id=building.id,
                assignment_kind="MAIN",
                date_from=need.date_from,
                date_to=need.date_to,
                weekly_hours=need.weekly_hours,
                annual_hours=need.annual_hours,
                status="DRAFT",
                created_by_user_id=admin_id,
                updated_by_user_id=admin_id,
            ))

        assign(math_need, matching_teacher_id, math_department)
        assign(language_need, matching_teacher_id, language_department)
        assign(extracurricular_need, matching_teacher_id, math_department)
        assign(additional_need, matching_teacher_id, math_department)
        assign(
            individual_project_need,
            matching_teacher_id,
            math_department,
        )
        assign(
            extracurricular_need,
            extracurricular_only_teacher_id,
            math_department,
        )
        assign(
            other_class_extracurricular_need,
            extracurricular_only_teacher_id,
            math_department,
        )
        assign(
            additional_need,
            additional_only_teacher_id,
            math_department,
        )
        assign(
            individual_project_need,
            project_only_teacher_id,
            math_department,
        )
        db.session.commit()
        other_class_extracurricular_need_id = (
            other_class_extracurricular_need.id
        )

    login(admin_id)
    response = client.get(
        "/workload/assignments/workspace",
        query_string={
            "version_id": context["version_id"],
            "view": "department",
            "department_id": context["department_id"],
            "presentation": "list",
        },
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Кафедральный" in html
    assert "Литература другой кафедры" in html
    assert "Общий внеурочный курс" in html
    assert "Общий дополнительный курс" in html
    assert "Индивидуальный проект" in html
    assert "Внеурочный" not in html
    assert "Дополнительный" not in html
    assert "Проектный" not in html

    matrix_response = client.get(
        "/workload/assignments/workspace",
        query_string={
            "version_id": context["version_id"],
            "view": "department",
            "department_id": context["department_id"],
            "presentation": "matrix",
        },
    )
    matrix_html = matrix_response.get_data(as_text=True)
    need_marker = (
        f'data-need-id="{other_class_extracurricular_need_id}"'
    )
    need_index = matrix_html.index(need_marker)
    cell_start = matrix_html.rfind("<div", 0, need_index)
    cell_end = matrix_html.index(">", need_index)

    assert matrix_response.status_code == 200
    assert "is-locked" in matrix_html[cell_start:cell_end]
    assert "Назначено другому преподавателю." in matrix_html[
        cell_start:cell_start + 900
    ]


def test_generate_from_workspace_returns_to_matrix(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        line = EducationPlanLine.query.one()
        line.weekly_hours = Decimal("0")
        line.annual_hours = Decimal("0")
        db.session.commit()
    login(admin_id)

    response = client.post(
        "/workload/assignments/generate",
        data={
            "version_id": context["version_id"],
            "return_to": "workspace",
        },
    )

    assert response.status_code == 302
    assert response.location.endswith(
        f"/workload/assignments/workspace?version_id={context['version_id']}"
    )


def test_workspace_adds_teacher_subject_and_assigns_full_need(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    unused_teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        need_id = need.id
        activity_id = need.education_activity_id
        teacher = db.session.get(User, teacher_id)
        teacher.last_name = "Иванова"
        teacher.first_name = "Анна"
        teacher_name = teacher.fio
        unused_teacher = db.session.get(User, unused_teacher_id)
        unused_teacher.last_name = "Ивановский"
        unused_teacher.first_name = "Незагруженный"
        db.session.get(User, admin_id).last_name = "Администраторов"
        db.session.commit()
    login(admin_id)
    filters = {
        "version_id": str(context["version_id"]),
        "view": "all",
    }

    add_teacher = client.post(
        "/workload/assignments/workspace/teachers",
        data={**filters, "teacher_id": str(teacher_id)},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert add_teacher.status_code == 200
    assert add_teacher.get_json() == {
        "ok": True,
        "holder_key": f"teacher:{teacher_id}",
        "teacher_id": teacher_id,
        "teacher_name": teacher_name,
    }
    empty_holder_fragment = client.get(
        "/workload/assignments/workspace",
        query_string={
            **filters,
            "fragment_holder_key": f"teacher:{teacher_id}",
        },
    )
    empty_fragment_html = empty_holder_fragment.get_data(as_text=True)
    assert empty_holder_fragment.status_code == 200
    assert "data-workload-holder-fragment" in empty_fragment_html
    assert "workload-add-subject-row" in empty_fragment_html
    after_teacher_add = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "workload-add-teacher-row" in after_teacher_add
    assert "Добавить преподавателя" in after_teacher_add

    add_subject = client.post(
        "/workload/assignments/workspace/subjects",
        data={
            **filters,
            "teacher_id": str(teacher_id),
            "activity_plan_kind": f"{activity_id}:CURRICULUM",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert add_subject.status_code == 200
    assert add_subject.get_json() == {
        "ok": True,
        "holder_key": f"teacher:{teacher_id}",
    }

    matrix = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    )
    html = matrix.get_data(as_text=True)
    assert matrix.status_code == 200
    assert "Предмет не выбран" not in html
    assert "data-workload-cell-toggle" in html
    assert "workload-cell-entry is-available" in html
    assert "data-workload-cell-value></span>" in html
    assert "data-cell-update-url=" in html
    assert 'data-hours="5"' in html
    assert 'name="hours"' not in html
    assert f'data-workload-holder-row="teacher:{teacher_id}"' in html
    assert "data-holder-sort-key=" in html
    assert "sortHolderRows(current);" in html
    assert "workload-matrix-sticky-class" in html
    assert "workload-matrix-sticky-plan" in html
    assert "workload-assignment-matrix__class-group" not in html
    assert "data-matrix-column-key=" not in html
    assert "columnCell?.cellIndex" not in html
    assert 'data-matrix-column-index="0"' in html
    assert "columnCell?.dataset.matrixColumnIndex" in html
    assert 'id="workspace-holder-page-size"' in html
    assert '<option value="5" selected>5</option>' in html
    assert "is-row-hovered" in html
    assert "is-column-hovered" in html
    assert html.count('class="workload-subject-add"') == 1
    assert "Добавить предмет" in html
    assert "<small>Учебный план</small>" not in html
    teacher_dialog_html = html.split('data-teacher-dialog', 1)[1].split(
        'data-vacancy-dialog',
        1,
    )[0]
    assert 'name="vacancy_note"' not in teacher_dialog_html
    assert "data-teacher-picker-search" in teacher_dialog_html
    assert "Введите фамилию или её часть" in teacher_dialog_html
    assert f'data-teacher-id="{teacher_id}"' not in teacher_dialog_html
    assert f'data-teacher-id="{unused_teacher_id}"' in teacher_dialog_html
    assert "normalizeTeacherSearch(option.dataset.search).includes(query)" in html
    assert "teacherPickerOptions().forEach((option) => option.hidden = true)" in html
    assert "removeTeacherPickerOption(payload.teacher_id)" in html

    holder_fragment = client.get(
        "/workload/assignments/workspace",
        query_string={
            **filters,
            "fragment_holder_key": f"teacher:{teacher_id}",
        },
    )
    fragment_html = holder_fragment.get_data(as_text=True)
    assert holder_fragment.status_code == 200
    assert "data-workload-holder-fragment" in fragment_html
    assert "data-workload-matrix" not in fragment_html
    assert f'data-workload-holder-row="teacher:{teacher_id}"' in fragment_html
    subject_select = fragment_html.split(
        'name="activity_plan_kind"',
        1,
    )[1].split("</select>", 1)[0]
    assert f'value="{activity_id}:CURRICULUM"' not in subject_select

    assign = client.post(
        "/workload/assignments/workspace/cell",
        data={
            **filters,
            "need_id": str(need_id),
            "teacher_id": str(teacher_id),
            "hours": "5",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert assign.status_code == 200
    assert assign.get_json()["allocated_delta"] == 5.0
    assigned_matrix = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "data-workload-cell-value>5</span>" in assigned_matrix
    with app.app_context():
        assignment = WorkloadAssignment.query.one()
        assert assignment.employee_user_id == teacher_id
        assert assignment.weekly_hours == Decimal("5.000")
        assert assignment.workload_need.status == "COVERED"

    locked = client.post(
        "/workload/assignments/workspace/cell",
        data={
            **filters,
            "need_id": str(need_id),
            "teacher_id": str(admin_id),
            "hours": "5",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert locked.status_code == 422
    assert "другому преподавателю" in locked.get_json()["message"]

    client.post(
        "/workload/assignments/workspace/teachers",
        data={**filters, "teacher_id": str(admin_id)},
    )
    client.post(
        "/workload/assignments/workspace/subjects",
        data={
            **filters,
            "teacher_id": str(admin_id),
            "activity_id": str(activity_id),
            "plan_kind": "CURRICULUM",
        },
    )
    locked_matrix = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "is-locked" in locked_matrix
    assert "Назначено другому преподавателю." in locked_matrix
    assert "disabled" in locked_matrix

    filtered_matrix = client.get(
        "/workload/assignments/workspace",
        query_string={
            **filters,
            "teacher_query": "Иван",
        },
    )
    filtered_html = filtered_matrix.get_data(as_text=True)
    assert filtered_matrix.status_code == 200
    assert f'data-workload-holder-row="teacher:{teacher_id}"' in filtered_html
    assert f'data-workload-holder-row="teacher:{admin_id}"' not in filtered_html
    assert f'data-workload-holder-row="teacher:{unused_teacher_id}"' not in filtered_html
    filter_options = filtered_html.split(
        'id="workspace-holder-options"',
        1,
    )[1].split('data-workload-holder-filter-empty', 1)[0]
    assert "Незагруженный" not in filter_options
    assert teacher_name in filter_options
    assert 'name="teacher_query"' in filtered_html
    assert 'placeholder="Введите фамилию или название вакансии"' in filtered_html
    assert "data-workload-holder-filter-option" in filtered_html
    assert 'class="workload-matrix-toolbar__guide"' in filtered_html
    assert filtered_html.index("workload-matrix-totals") < (
        filtered_html.index("workload-matrix-legend")
    )

    subject_filtered = client.get(
        "/workload/assignments/workspace",
        query_string={
            **filters,
            "subject_id": str(activity_id),
        },
    )
    subject_filtered_html = subject_filtered.get_data(as_text=True)
    assert subject_filtered.status_code == 200
    assert "is-filter-compact" in subject_filtered_html
    assert "Кл. руководство:" not in subject_filtered_html
    assert "МЦКО:" not in subject_filtered_html
    assert "workload-add-subject-row" not in subject_filtered_html


def test_workspace_preserves_multiple_level_and_grade_filters(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
    login(admin_id)

    response = client.get(
        "/workload/assignments/workspace",
        query_string=MultiDict([
            ("version_id", str(context["version_id"])),
            ("education_level", "OOO"),
            ("education_level", "SOO"),
            ("grade", "5"),
            ("grade", "10"),
        ]),
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert (
        'name="education_level" value="OOO" '
        'data-workload-filter-auto checked'
    ) in html
    assert (
        'name="education_level" value="SOO" '
        'data-workload-filter-auto checked'
    ) in html
    assert (
        'name="grade" value="5" data-workload-filter-auto checked'
    ) in html
    assert (
        'name="grade" value="10" data-workload-filter-auto checked'
    ) in html
    assert 'data-workload-filter-clear="education_level"' in html
    assert 'data-workload-filter-clear="grade"' in html
    assert "education_level=OOO&amp;education_level=SOO" in html
    assert "grade=5&amp;grade=10" in html


def test_workspace_can_delete_teacher_subject_row(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        teacher = db.session.get(User, teacher_id)
        teacher.last_name = "Удаляемов"
        teacher.first_name = "Предмет"
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        need_id = need.id
        activity_id = need.education_activity_id
        db.session.commit()
    login(admin_id)
    filters = {
        "version_id": str(context["version_id"]),
        "view": "all",
    }

    client.post(
        "/workload/assignments/workspace/teachers",
        data={**filters, "teacher_id": str(teacher_id)},
    )
    client.post(
        "/workload/assignments/workspace/subjects",
        data={
            **filters,
            "teacher_id": str(teacher_id),
            "activity_plan_kind": f"{activity_id}:CURRICULUM",
        },
    )
    client.post(
        "/workload/assignments/workspace/cell",
        data={
            **filters,
            "need_id": str(need_id),
            "teacher_id": str(teacher_id),
            "hours": "5",
        },
    )

    before = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "Удалить строку предмета" in before
    assert "/workload/assignments/workspace/subjects/delete" in before
    assert "data-workload-async-subject-delete" in before
    assert '"X-Requested-With": "XMLHttpRequest"' in before
    assert "Удалить всю строку" in before

    deleted = client.post(
        "/workload/assignments/workspace/subjects/delete",
        data={
            **filters,
            "holder_type": "teacher",
            "teacher_id": str(teacher_id),
            "activity_id": str(activity_id),
            "plan_kind": "CURRICULUM",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert deleted.status_code == 200
    assert deleted.get_json()["holder_key"] == f"teacher:{teacher_id}"
    assert "строка удалена" in deleted.get_json()["message"]
    after_html = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "Предмет не выбран" in after_html

    with app.app_context():
        assignment = WorkloadAssignment.query.one()
        assert assignment.status == "CANCELLED"
        assert assignment.workload_need.status == "OPEN"
        change = (
            WorkloadAssignmentChange.query
            .filter_by(
                workload_assignment_id=assignment.id,
                change_kind="CANCEL",
            )
            .one()
        )
        assert "предметная строка" in change.reason


def test_workspace_can_delete_entire_teacher_row(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        teacher = db.session.get(User, teacher_id)
        teacher.last_name = "Удаляемов"
        teacher.first_name = "Полностью"
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        db.session.add(_assignment(need, teacher_id, "5"))
        db.session.commit()
    login(admin_id)
    filters = {
        "version_id": str(context["version_id"]),
        "view": "all",
    }

    before = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert f'data-source-teacher-id="{teacher_id}"' in before
    assert "/workload/assignments/workspace/holders/delete" in before
    assert "data-workload-async-holder-delete" in before

    deleted = client.post(
        "/workload/assignments/workspace/holders/delete",
        data={
            **filters,
            "holder_type": "teacher",
            "teacher_id": str(teacher_id),
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert deleted.status_code == 200
    payload = deleted.get_json()
    assert payload["ok"] is True
    assert payload["holder_key"] == f"teacher:{teacher_id}"
    assert payload["released_weekly_hours"] == 5.0
    assert payload["teacher"] == {
        "id": teacher_id,
        "name": "Удаляемов Полностью",
    }
    assert "строка удалена" in payload["message"]

    deleted_html = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert f'data-source-teacher-id="{teacher_id}"' not in deleted_html
    teacher_dialog_html = deleted_html.split(
        'data-teacher-dialog',
        1,
    )[1].split('data-vacancy-dialog', 1)[0]
    assert f'data-teacher-id="{teacher_id}"' in teacher_dialog_html

    with app.app_context():
        assignment = WorkloadAssignment.query.one()
        assert assignment.status == "CANCELLED"
        assert assignment.workload_need.status == "OPEN"
        change = (
            WorkloadAssignmentChange.query
            .filter_by(
                workload_assignment_id=assignment.id,
                change_kind="CANCEL",
            )
            .one()
        )
        assert "строка преподавателя" in change.reason


def test_workspace_export_and_compact_hours(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
    login(admin_id)

    matrix = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    )
    html = matrix.get_data(as_text=True)
    assert "5,000" not in html
    assert "5.000" not in html

    export = client.get(
        "/workload/assignments/workspace/export.xlsx",
        query_string={"version_id": context["version_id"]},
    )
    assert export.status_code == 200
    assert export.data.startswith(b"PK")
    assert "spreadsheetml.sheet" in export.content_type


def test_workspace_paginates_five_teachers_by_default(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_ids = [make_user("TEACHER") for _ in range(6)]
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
    login(admin_id)
    with client.session_transaction() as workspace_session:
        workspace_session[
            f"workload_matrix_state_{context['version_id']}"
        ] = {
            "teacher_ids": teacher_ids,
            "vacancies": [],
            "rows": [],
        }

    default_response = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    )
    default_html = default_response.get_data(as_text=True)

    assert default_response.status_code == 200
    assert default_html.count("<!-- workload-holder-start:") == 5
    assert "Педагоги 1–5 из 6" in default_html
    assert default_html.count('aria-label="Страницы педагогов"') == 2
    assert "workload-holder-pagination--top" in default_html
    assert "workload-holder-pagination--bottom" in default_html
    assert '<option value="5" selected>5</option>' in default_html
    assert "loadHolderPage" in default_html
    assert 'searchParams.set("page_fragment", "1")' in default_html
    assert "window.history.pushState" in default_html
    assert 'window.addEventListener("popstate"' in default_html

    fragment_response = client.get(
        "/workload/assignments/workspace",
        query_string={
            "version_id": context["version_id"],
            "holder_page": 2,
            "page_fragment": 1,
        },
    )
    fragment_html = fragment_response.get_data(as_text=True)

    assert fragment_response.status_code == 200
    assert 'data-workload-page-region' in fragment_html
    assert "Педагоги 6–6 из 6" in fragment_html
    assert fragment_html.count('aria-label="Страницы педагогов"') == 2
    assert fragment_html.count("<!-- workload-holder-start:") == 1
    assert "data-workload-filterbar" not in fragment_html
    assert "data-copy-subjects-dialog" not in fragment_html
    assert "<script>" not in fragment_html
    assert len(fragment_html) < len(default_html)

    expanded_response = client.get(
        "/workload/assignments/workspace",
        query_string={
            "version_id": context["version_id"],
            "holder_page_size": 10,
        },
    )
    expanded_html = expanded_response.get_data(as_text=True)

    assert expanded_response.status_code == 200
    assert expanded_html.count("<!-- workload-holder-start:") == 6
    assert '<option value="10" selected>10</option>' in expanded_html


def test_workspace_list_view_and_matching_excel_export(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        db.session.add(_assignment(need, teacher_id, "5"))
        db.session.commit()
        teacher_name = db.session.get(User, teacher_id).fio
    login(admin_id)

    response = client.get(
        "/workload/assignments/workspace",
        query_string={
            "version_id": context["version_id"],
            "presentation": "list",
        },
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "workload-teacher-list" in html
    assert "workload-assignment-matrix-wrap" not in html
    assert teacher_name in html
    assert "Урочная деятельность" in html
    assert "Математика" in html
    assert "Группа / весь класс" in html
    assert "Группа 1" in html
    assert ">5А<" in html
    assert "Главное здание" in html
    assert "Всего у преподавателя" in html

    export = client.get(
        "/workload/assignments/workspace/export.xlsx",
        query_string={
            "version_id": context["version_id"],
            "presentation": "list",
        },
    )

    assert export.status_code == 200
    workbook = load_workbook(BytesIO(export.data), data_only=True)
    sheet = workbook["Нагрузка списком"]
    values = {
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert teacher_name in values
    assert "Предмет" in values
    assert "Класс" in values
    assert "Часы" in values
    assert "Группа / весь класс" in values
    assert "Здание" in values
    assert "Математика" in values
    assert "5А" in values
    assert "Группа 1" in values
    assert "Урочная деятельность" in values
    assert "Итоги по всей выборке" in values


def test_metagroup_display_lists_its_combined_classes():
    def source_group(class_name):
        return SimpleNamespace(source_classes=[SimpleNamespace(
            population_snapshot_class=SimpleNamespace(
                name_snapshot=class_name,
            ),
        )])

    group = SimpleNamespace(
        group_type="METAGROUP",
        name="Метагруппа английского языка",
        metagroup_sources=[
            SimpleNamespace(source_group=source_group("5А")),
            SimpleNamespace(source_group=source_group("5Б")),
        ],
    )

    assert teaching_group_class_label(group) == "5А, 5Б"
    assert teaching_group_assignment_label(group) == "Метагруппа: 5А + 5Б"


def test_extracurricular_group_display_uses_composition_not_subject_name():
    snapshot_class = SimpleNamespace(name_snapshot="2И")
    full_class_group = SimpleNamespace(
        group_type="EXTRACURRICULAR_GROUP",
        name="2И · Разговоры о важном",
        source_classes=[SimpleNamespace(
            relation_kind="FULL",
            population_snapshot_class=snapshot_class,
        )],
    )
    numbered_group = SimpleNamespace(
        group_type="EXTRACURRICULAR_GROUP",
        name="2И · Робототехника · группа 2",
        source_classes=[SimpleNamespace(
            relation_kind="SOURCE",
            population_snapshot_class=snapshot_class,
        )],
    )

    assert teaching_group_class_label(full_class_group) == "2И"
    assert teaching_group_assignment_label(full_class_group) == "Весь класс"
    assert teaching_group_assignment_label(numbered_group) == "Группа 2"


def test_workspace_vacancy_can_be_filled_by_teacher(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        need_id = need.id
        activity_id = need.education_activity_id
    login(admin_id)
    filters = {
        "version_id": str(context["version_id"]),
        "view": "all",
    }

    assert client.post(
        "/workload/assignments/workspace/teachers",
        data={
            **filters,
            "holder_type": "vacancy",
            "vacancy_note": "  (Русский язык)  ",
        },
    ).status_code == 302
    assert client.post(
        "/workload/assignments/workspace/subjects",
        data={
            **filters,
            "holder_type": "vacancy",
            "vacancy_key": "VACANCY_1",
            "activity_plan_kind": f"{activity_id}:CURRICULUM",
        },
    ).status_code == 302
    assigned = client.post(
        "/workload/assignments/workspace/cell",
        data={
            **filters,
            "holder_type": "vacancy",
            "vacancy_key": "VACANCY_1",
            "need_id": str(need_id),
            "hours": "5",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert assigned.status_code == 200
    assert assigned.get_json()["allocated_delta"] == 5.0
    with app.app_context():
        vacancy = WorkloadAssignment.query.one()
        assert vacancy.assignment_kind == "VACANCY"
        assert vacancy.employee_user_id is None
        assert vacancy.position_code == "VACANCY_1"
        assert vacancy.position_title == "Вакансия 1 (Русский язык)"

    renamed = client.post(
        "/workload/assignments/workspace/vacancies/label",
        data={
            **filters,
            "vacancy_key": "VACANCY_1",
            "vacancy_note": "Иванова А.А.",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert renamed.status_code == 200
    assert renamed.get_json() == {
        "ok": True,
        "holder_key": "vacancy:VACANCY_1",
        "holder_name": "Вакансия 1 (Иванова А.А.)",
    }
    with app.app_context():
        assert (
            WorkloadAssignment.query.one().position_title
            == "Вакансия 1 (Иванова А.А.)"
        )

    vacancy_view = client.get(
        "/workload/assignments/workspace",
        query_string={
            **filters,
            "view": "vacancies",
            "teacher_query": "Иванов",
        },
    ).get_data(as_text=True)
    assert "Вакансия 1 (Иванова А.А.)" in vacancy_view
    assert 'data-workload-total-allocated>5</strong>' in vacancy_view
    assert 'data-workload-total-remaining>0</strong>' in vacancy_view
    holder_filter = vacancy_view.split(
        'id="workspace-holder-options"',
        1,
    )[1].split('data-workload-holder-filter-empty', 1)[0]
    assert 'data-holder-key="vacancy:VACANCY_1"' in holder_filter
    assert "Вакансия 1 (Иванова А.А.)" in holder_filter
    assert "Преподаватель не назначен" in vacancy_view
    assert 'data-open-vacancy-dialog' in vacancy_view
    assert 'data-vacancy-key="VACANCY_1"' in vacancy_view

    replaced = client.post(
        "/workload/assignments/workspace/holder/replace",
        data={
            **filters,
            "source_type": "vacancy",
            "source_vacancy_key": "VACANCY_1",
            "target_holder": f"teacher:{teacher_id}",
        },
    )
    assert replaced.status_code == 302
    with app.app_context():
        assignment = WorkloadAssignment.query.one()
        assert assignment.assignment_kind == "MAIN"
        assert assignment.employee_user_id == teacher_id
        assert assignment.position_code == "TEACHER"


def test_workspace_rejects_transfer_to_teacher_with_existing_load(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    source_teacher_id = make_user("TEACHER")
    occupied_teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        source = _assignment(need, source_teacher_id, "5")
        occupied = _assignment(need, occupied_teacher_id, "1")
        occupied.workload_need_id = need.id
        db.session.add_all([source, occupied])
        db.session.commit()
    login(admin_id)

    response = client.post(
        "/workload/assignments/workspace/holder/replace",
        data={
            "version_id": str(context["version_id"]),
            "view": "all",
            "source_type": "teacher",
            "source_teacher_id": str(source_teacher_id),
            "target_holder": f"teacher:{occupied_teacher_id}",
        },
        follow_redirects=True,
    )
    assert "у Тестов Пользователь уже есть нагрузка" in response.get_data(as_text=True)
    with app.app_context():
        assert WorkloadAssignment.query.filter_by(
            employee_user_id=source_teacher_id,
        ).count() == 1


def test_workspace_copies_subject_set_only_to_teacher_without_load(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    source_teacher_id = make_user("TEACHER")
    target_teacher_id = make_user("TEACHER")
    occupied_teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        source_teacher = db.session.get(User, source_teacher_id)
        source_teacher.last_name = "Исходный"
        target_teacher = db.session.get(User, target_teacher_id)
        target_teacher.last_name = "Свободный"
        occupied_teacher = db.session.get(User, occupied_teacher_id)
        occupied_teacher.last_name = "Занятый"
        db.session.add_all([
            _assignment(need, source_teacher_id, "5"),
            _assignment(need, occupied_teacher_id, "1"),
        ])
        db.session.commit()
        activity_id = need.education_activity_id
        source_name = source_teacher.fio
        target_name = target_teacher.fio
        occupied_name = occupied_teacher.fio
    login(admin_id)
    filters = {
        "version_id": str(context["version_id"]),
        "view": "all",
    }

    before = client.get(
        "/workload/assignments/workspace",
        query_string=filters,
    ).get_data(as_text=True)
    assert "Скопировать предметы" in before
    copy_dialog = before.split(
        'data-copy-subjects-dialog',
        1,
    )[1].split('data-holder-dialog', 1)[0]
    assert target_name in copy_dialog
    assert occupied_name not in copy_dialog
    assert source_name not in copy_dialog

    copied = client.post(
        "/workload/assignments/workspace/holder/copy-subjects",
        data={
            **filters,
            "source_teacher_id": str(source_teacher_id),
            "target_teacher_id": str(target_teacher_id),
        },
        follow_redirects=True,
    )
    copied_html = copied.get_data(as_text=True)
    assert copied.status_code == 200
    assert "Набор из 1 предметов скопирован" in copied_html
    redirect_query = parse_qs(urlparse(
        copied.history[0].headers["Location"]
    ).query)
    assert redirect_query["teacher_query"] == [target_name]
    assert redirect_query["focus_holder"] == [f"teacher:{target_teacher_id}"]
    assert "scrollIntoView" in copied_html
    assert f'data-workload-holder-row="teacher:{target_teacher_id}"' in copied_html
    target_fragment = copied_html.split(
        f'workload-holder-start:teacher:{target_teacher_id}',
        1,
    )[1].split(
        f'workload-holder-end:teacher:{target_teacher_id}',
        1,
    )[0]
    assert f'name="activity_id" value="{activity_id}"' in target_fragment
    with app.app_context():
        assert WorkloadAssignment.query.filter_by(
            employee_user_id=target_teacher_id,
        ).count() == 0
        assert WorkloadAssignment.query.filter_by(
            employee_user_id=source_teacher_id,
        ).count() == 1

    rejected = client.post(
        "/workload/assignments/workspace/holder/copy-subjects",
        data={
            **filters,
            "source_teacher_id": str(source_teacher_id),
            "target_teacher_id": str(occupied_teacher_id),
        },
        follow_redirects=True,
    )
    assert "уже есть нагрузка" in rejected.get_data(as_text=True)


def test_workload_editor_settings_and_workspace_hide_draft_badge(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    methodist_id = make_user("METHODIST")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
    login(admin_id)

    settings = client.post(
        "/workload/settings/editors",
        data={"editor_ids": [str(admin_id), str(methodist_id)]},
        follow_redirects=True,
    )
    assert settings.status_code == 200
    settings_html = settings.get_data(as_text=True)
    assert "Ответственные за учебные планы и нагрузку сохранены" in settings_html
    assert f'data-editor-user data-user-id="{methodist_id}"' in settings_html
    assert f'data-editor-user data-user-id="{teacher_id}"' not in settings_html
    assert f'data-editor-option' in settings_html
    assert f'data-user-id="{teacher_id}"' in settings_html
    with app.app_context():
        assert WorkloadEditorAccess.query.filter_by(
            user_id=methodist_id,
            is_active=True,
        ).count() == 1

    workspace = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    ).get_data(as_text=True)
    assert "2026/2027" in workspace
    assert "Версия 1" not in workspace
    assert ">Черновик<" not in workspace


def test_department_and_workspace_read_current_load_and_mcko(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        db.session.add(_assignment(need, teacher_id, "5"))
        db.session.add(TeacherMckoResult(
            teacher_id=teacher_id,
            education_activity_id=need.education_activity_id,
            passed_at=date(2026, 5, 26),
            expires_at=date(2029, 5, 26),
            level="HIGH",
        ))
        teacher = db.session.get(User, teacher_id)
        teacher.employment_start_date = date(2020, 9, 1)
        db.session.add(TeacherAttestation(
            teacher_id=teacher_id,
            category="HIGHEST",
            decision_date=date(2026, 4, 15),
            is_indefinite=True,
            entry_source="ADMINISTRATION",
            created_by_user_id=admin_id,
        ))
        db.session.commit()
        year_id = context["year_id"]

    login(admin_id)
    workspace = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    ).get_data(as_text=True)
    assert f"/departments/teachers/{teacher_id}" in workspace
    assert "МЦКО: Математика — Высокий" in workspace
    assert "workload-teacher-mcko is-active" in workspace
    assert "до 26.05.2029" in workspace

    load_page = client.get(
        "/departments/loads",
        query_string={"academic_year_id": year_id},
    ).get_data(as_text=True)
    assert "Данные поступают из матрицы распределения нагрузки" in load_page
    assert "Импортировать Excel" not in load_page
    assert "Математика" in load_page

    profile = client.get(
        f"/departments/teachers/{teacher_id}",
        query_string={"academic_year_id": year_id},
    )
    assert profile.status_code == 200
    profile_html = profile.get_data(as_text=True)
    assert "Профиль преподавателя" in profile_html
    assert "5 ч/нед." in profile_html
    assert "Высокий" in profile_html
    assert "professional-level-HIGH" in profile_html
    assert "Высшая квалификационная категория" in profile_html
    assert "Бессрочно" in profile_html
    assert "Действующая квалификационная категория" in profile_html

    summary = client.get(
        "/departments/summary",
        query_string={
            "academic_year_id": year_id,
            "department_id": context["department_id"],
        },
    )
    summary_html = summary.get_data(as_text=True)
    assert summary.status_code == 200
    assert f"/departments/teachers/{teacher_id}" in summary_html
    assert "Диагностика действует" in summary_html
    assert "professional-level-HIGH" in summary_html
    assert "Высшая квалификационная категория" in summary_html
    assert "Бессрочно" in summary_html
    assert "5" in summary_html


def test_workspace_and_department_summary_highlight_missing_mcko(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        need = WorkloadNeed.query.one()
        db.session.add(_assignment(need, teacher_id, "5"))
        db.session.commit()

    login(admin_id)
    workspace = client.get(
        "/workload/assignments/workspace",
        query_string={"version_id": context["version_id"]},
    ).get_data(as_text=True)
    summary = client.get(
        "/departments/summary",
        query_string={
            "academic_year_id": context["year_id"],
            "department_id": context["department_id"],
        },
    ).get_data(as_text=True)

    assert "МЦКО: Диагностика отсутствует" in workspace
    assert "workload-teacher-mcko is-danger" in workspace
    assert "Диагностика отсутствует" in summary
    assert "professional-status-MISSING" in summary
    assert "Не указана дата приёма" in summary


def test_teacher_can_view_only_own_workload(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    teacher_id = make_user("TEACHER")
    other_teacher_id = make_user("TEACHER")
    login(teacher_id)

    own = client.get(f"/workload/teachers/{teacher_id}")
    other = client.get(f"/workload/teachers/{other_teacher_id}")
    registry = client.get("/workload/assignments/")

    assert own.status_code == 200
    assert other.status_code == 403
    assert registry.status_code == 403


def test_inactive_employee_is_rejected(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context = _distribution_context(admin_id)
        _generate(context, admin_id)
        employee = db.session.get(User, teacher_id)
        employee.employment_status = "DISMISSED"
        need = WorkloadNeed.query.one()

        with pytest.raises(
            WorkloadDistributionError,
            match="неработающему",
        ):
            validate_assignment(
                need,
                _assignment(need, teacher_id, "5"),
            )


def _calculation_parameter_set(
    context,
    user_id,
    teacher_id,
    *,
    include_norm=True,
):
    version = db.session.get(TariffVersion, context["version_id"])
    parameter_set = CalculationParameterSet(
        tariff_version=version,
        code="BASE_2026",
        name="Основные параметры 2026/2027",
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 8, 31),
        student_hour_rate=Decimal("37"),
        periods_per_year=Decimal("12"),
        rounding_rule="HALF_UP",
        currency_code="RUB",
        status="DRAFT",
        created_by_user_id=user_id,
        updated_by_user_id=user_id,
    )
    db.session.add(parameter_set)
    db.session.flush()
    coefficient_types, allowance_types = ensure_standard_tariff_types()
    if include_norm:
        db.session.add(TariffRateNorm(
            parameter_set_id=parameter_set.id,
            position_code="TEACHER",
            position_name="Учитель",
            activity_kind="SUBJECT",
            weekly_norm_hours=Decimal("18"),
            valid_from=date(2026, 9, 1),
            valid_to=date(2027, 8, 31),
            source_text="Тестовая норма",
            created_by_user_id=user_id,
        ))
    db.session.add(TariffCoefficientValue(
        parameter_set_id=parameter_set.id,
        coefficient_type_id=coefficient_types["DIVISION"].id,
        value=Decimal("25"),
        condition_kind="ALWAYS",
        condition_data={"only_above_one": True},
        priority=10,
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 8, 31),
        created_by_user_id=user_id,
    ))
    db.session.add(TariffCoefficientValue(
        parameter_set_id=parameter_set.id,
        coefficient_type_id=coefficient_types["COMPLEXITY"].id,
        value=Decimal("1.4"),
        condition_kind="ACTIVITY",
        condition_data={
            "activity_ids": [
                db.session.get(
                    TeachingGroup,
                    context["group_id"],
                ).education_activity_id,
            ],
        },
        priority=20,
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 8, 31),
        created_by_user_id=user_id,
    ))
    db.session.add(TariffAllowanceRule(
        parameter_set_id=parameter_set.id,
        allowance_type_id=allowance_types["CLASSROOM_CITY"].id,
        fixed_amount=Decimal("16000"),
        percent_value=None,
        base_kind="BASE",
        condition_data={"employee_user_ids": [teacher_id]},
        priority=30,
        valid_from=date(2026, 9, 1),
        valid_to=date(2027, 8, 31),
        source_text="Тестовая доплата",
        created_by_user_id=user_id,
    ))
    db.session.commit()
    return parameter_set.id


def _calculation_context(user_id, teacher_id, *, include_norm=True):
    context = _distribution_context(user_id)
    _generate(context, user_id)
    need = WorkloadNeed.query.one()
    assignment = _assignment(need, teacher_id, "5")
    validate_assignment(need, assignment)
    db.session.add(assignment)
    db.session.flush()
    refresh_need_status(need)
    db.session.commit()
    parameter_set_id = _calculation_parameter_set(
        context,
        user_id,
        teacher_id,
        include_norm=include_norm,
    )
    return context, assignment.id, parameter_set_id


def _workflow_context(user_id, teacher_id):
    context, assignment_id, parameter_set_id = _calculation_context(
        user_id,
        teacher_id,
    )
    version = db.session.get(TariffVersion, context["version_id"])
    version.effective_from = date(2026, 9, 1)
    version.effective_to = date(2027, 8, 31)
    version.plans[0].status = "READY"
    run, reused = calculate_tariff_version(
        version,
        db.session.get(CalculationParameterSet, parameter_set_id),
        user_id=user_id,
    )
    assert reused is False
    assert run.status == "SUCCEEDED"
    db.session.commit()
    return context, assignment_id, parameter_set_id


def test_tariff_calculation_matches_reference_formula(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, parameter_set_id = _calculation_context(
            admin_id,
            teacher_id,
        )
        run, reused = calculate_tariff_version(
            db.session.get(TariffVersion, context["version_id"]),
            db.session.get(CalculationParameterSet, parameter_set_id),
            user_id=admin_id,
        )
        db.session.commit()

        assert reused is False
        assert run.status == "SUCCEEDED"
        line = TariffLine.query.one()
        assert line.population_value == Decimal("15.000")
        assert line.rate_norm_hours == Decimal("18.000")
        assert line.fte_value == Decimal("0.277778")
        assert line.base_amount == Decimal("7862.50")
        assert line.total_amount == Decimal("32249.17")
        components = {
            item.component_code: item
            for item in TariffLineComponent.query.all()
        }
        assert components["DIVISION"].amount_value == Decimal("5241.67")
        assert components["COMPLEXITY"].amount_value == Decimal("3145.00")
        assert (
            components["CLASSROOM_CITY"].amount_value
            == Decimal("16000.00")
        )
        assert line.formula_snapshot["algorithm_version"] == "ALT-TARIFF-1.0"


def test_tariff_calculation_is_idempotent(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, parameter_set_id = _calculation_context(
            admin_id,
            teacher_id,
        )
        version = db.session.get(TariffVersion, context["version_id"])
        parameter_set = db.session.get(
            CalculationParameterSet,
            parameter_set_id,
        )
        first, first_reused = calculate_tariff_version(
            version,
            parameter_set,
            user_id=admin_id,
        )
        db.session.commit()
        second, second_reused = calculate_tariff_version(
            version,
            parameter_set,
            user_id=admin_id,
        )

        assert first_reused is False
        assert second_reused is True
        assert second.id == first.id
        assert TariffCalculationRun.query.count() == 1
        assert TariffLine.query.count() == 1


def test_missing_rate_norm_produces_failed_diagnostic_run(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, parameter_set_id = _calculation_context(
            admin_id,
            teacher_id,
            include_norm=False,
        )
        run, reused = calculate_tariff_version(
            db.session.get(TariffVersion, context["version_id"]),
            db.session.get(CalculationParameterSet, parameter_set_id),
            user_id=admin_id,
        )
        db.session.commit()

        assert reused is False
        assert run.status == "FAILED"
        assert run.summary_data["error_count"] == 1
        assert "норма ставки" in run.error_text
        assert TariffLine.query.count() == 0


def test_retired_tariffication_and_integration_routes_return_404(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    login(admin_id)

    for path in (
        "/workload/tariffication/",
        "/workload/tariffication/runs/1",
        "/workload/tariffication/lines/1",
        "/workload/tariffication/settings/",
        "/workload/tariffication/settings/1",
        "/workload/integration/",
    ):
        assert client.get(path).status_code == 404
    for path in (
        "/workload/tariffication/calculate",
        "/workload/tariffication/settings/new",
        "/workload/integration/reconcile",
        "/workload/integration/source",
    ):
        assert client.post(path).status_code == 404


def test_workflow_validation_blocks_incomplete_draft(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, parameter_set_id = _calculation_context(
            admin_id,
            teacher_id,
        )
        version = db.session.get(TariffVersion, context["version_id"])
        version.effective_from = date(2026, 9, 1)
        calculate_tariff_version(
            version,
            db.session.get(CalculationParameterSet, parameter_set_id),
            user_id=admin_id,
        )
        validation = run_full_validation(version, user_id=admin_id)
        db.session.commit()

        assert validation.status == "FAILED"
        assert any(
            issue.rule_code == "WF-PLAN-NOT-READY"
            for issue in validation.issues
        )
        run, cycle = start_review(version, user_id=admin_id)
        assert run.status == "FAILED"
        assert cycle is None
        assert version.status == "DRAFT"


def test_sequential_reviews_approve_and_activate_version(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        validation, cycle = start_review(version, user_id=admin_id)
        assert validation.status == "PASSED"
        assert version.status == "VALIDATION"

        for stage in ("ACADEMIC", "HR", "FINANCE"):
            record_review_decision(
                version,
                cycle,
                review_stage=stage,
                decision="APPROVED",
                comment=f"Этап {stage} проверен",
                user_id=admin_id,
            )
        assert version.status == "APPROVAL"
        assert TariffReviewDecision.query.count() == 3

        approval = approve_version(
            version,
            decision="APPROVED",
            comment="Утверждаю",
            effective_from=date(2026, 9, 1),
            user_id=admin_id,
            today=date(2026, 9, 1),
        )
        db.session.commit()

        assert approval.decision == "APPROVED"
        assert version.status == "EFFECTIVE"
        assert version.checksum
        assert version.plans[0].status == "LOCKED"
        assert version.calculation_parameter_sets[0].status == "LOCKED"
        assignment = WorkloadAssignment.query.one()
        assert assignment.status == "CONFIRMED"
        assert [item.to_status for item in version.status_history] == [
            "DRAFT",
            "VALIDATION",
            "APPROVAL",
            "APPROVED",
            "EFFECTIVE",
        ]


def test_reviews_must_follow_required_order(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        _, cycle = start_review(version, user_id=admin_id)

        with pytest.raises(
            TariffWorkflowError,
            match="последовательно",
        ):
            record_review_decision(
                version,
                cycle,
                review_stage="HR",
                decision="APPROVED",
                comment=None,
                user_id=admin_id,
            )


def test_returned_review_comment_requires_answer_and_author_close(
    app,
    make_user,
):
    admin_id = make_user("ADMIN")
    deputy_id = make_user("DEPUTY_DIRECTOR")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        _, cycle = start_review(version, user_id=admin_id)
        record_review_decision(
            version,
            cycle,
            review_stage="ACADEMIC",
            decision="CHANGES_REQUESTED",
            comment="Уточнить период нагрузки",
            user_id=admin_id,
        )
        comment = TariffReviewComment.query.one()

        assert version.status == "DRAFT"
        assert cycle.status == "RETURNED"
        assert comment.status == "OPEN"
        answer_review_comment(
            comment,
            response_text="Период проверен и исправлен",
            user_id=deputy_id,
        )
        with pytest.raises(TariffWorkflowError, match="только его автор"):
            close_review_comment(comment, user_id=deputy_id)
        close_review_comment(comment, user_id=admin_id)
        assert comment.status == "CLOSED"


def test_stale_calculation_blocks_review(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, assignment_id, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        assignment = db.session.get(WorkloadAssignment, assignment_id)
        assignment.weekly_hours = Decimal("4")
        assignment.annual_hours = Decimal("136")
        assignment.revision += 1
        validation = run_full_validation(version, user_id=admin_id)

        assert validation.status == "FAILED"
        assert any(
            issue.rule_code == "WF-CALCULATION-STALE"
            for issue in validation.issues
        )


def test_correction_clones_source_without_calculation(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        source = db.session.get(TariffVersion, context["version_id"])
        _, cycle = start_review(source, user_id=admin_id)
        for stage in ("ACADEMIC", "HR", "FINANCE"):
            record_review_decision(
                source,
                cycle,
                review_stage=stage,
                decision="APPROVED",
                comment=None,
                user_id=admin_id,
            )
        approve_version(
            source,
            decision="APPROVED",
            comment=None,
            effective_from=date(2026, 9, 1),
            user_id=admin_id,
            today=date(2026, 9, 1),
        )
        target = clone_correction_version(
            source,
            effective_from=date(2027, 1, 15),
            reason_text="Изменение нагрузки с 15 января",
            user_id=admin_id,
        )
        db.session.commit()

        assert target.status == "DRAFT"
        assert target.origin_version_id == source.id
        assert target.calculation_runs == []
        assert len(target.plans) == len(source.plans)
        assert target.plans[0].status == "DRAFT"
        copied = WorkloadAssignment.query.filter_by(
            tariff_version_id=target.id,
        ).one()
        assert copied.origin_assignment_id is not None
        assert copied.date_from == date(2027, 1, 15)
        assert copied.revision == 1


def test_project_and_official_documents_are_versioned(app, make_user):
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        project = generate_tariff_document(
            version,
            document_type="PERSONAL_TARIFF",
            employee_user_id=teacher_id,
            user_id=admin_id,
        )
        db.session.commit()
        assert project.status == "PROJECT"
        assert resolve_artifact_path(project).suffix == ".docx"

        _, cycle = start_review(version, user_id=admin_id)
        for stage in ("ACADEMIC", "HR", "FINANCE"):
            record_review_decision(
                version,
                cycle,
                review_stage=stage,
                decision="APPROVED",
                comment=None,
                user_id=admin_id,
            )
        approve_version(
            version,
            decision="APPROVED",
            comment=None,
            effective_from=date(2026, 9, 1),
            user_id=admin_id,
            today=date(2026, 9, 1),
        )
        official = generate_tariff_document(
            version,
            document_type="PERSONAL_TARIFF",
            employee_user_id=teacher_id,
            user_id=admin_id,
        )
        db.session.commit()

        assert official.status == "OFFICIAL"
        assert official.revision_no == 2
        assert official.version_checksum == version.checksum
        assert TariffDocumentArtifact.query.count() == 2
        assert resolve_artifact_path(official).is_file()


def test_workflow_routes_and_permissions(app, client, make_user, login):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    deputy_id = make_user("DEPUTY_DIRECTOR")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, _, _ = _workflow_context(admin_id, teacher_id)
        version = db.session.get(TariffVersion, context["version_id"])
        _, cycle = start_review(version, user_id=admin_id)
        db.session.commit()
        cycle_id = cycle.id

    login(admin_id)
    workflow_page = client.get(
        f"/workload/workflow/?version_id={context['version_id']}"
    )
    assert workflow_page.status_code == 200
    workflow_html = workflow_page.get_data(as_text=True)
    assert "workload-control-nav" in workflow_html
    assert "workload-control-counts" in workflow_html
    assert "Тарификация" not in workflow_html
    assert "Источник кафедр" not in workflow_html
    assert "/workload/integration" not in workflow_html
    documents_page = client.get(
        f"/workload/documents/?version_id={context['version_id']}"
    )
    assert documents_page.status_code == 200
    documents_html = documents_page.get_data(as_text=True)
    assert "workload-control-nav" in documents_html
    assert "workload-control-summary" in documents_html
    assert "Документы учебных планов и нагрузки" in documents_html
    assert "Тарификация" not in documents_html
    assert "Источник кафедр" not in documents_html
    assert "Пакет версии" not in documents_html

    login(deputy_id)
    forbidden = client.post(
        f"/workload/workflow/reviews/{cycle_id}/decision",
        data={"review_stage": "FINANCE", "decision": "APPROVED"},
    )
    assert forbidden.status_code == 403


def _effective_integration_context(app, admin_id, teacher_id, legacy_hours):
    context, assignment_id, _ = _workflow_context(admin_id, teacher_id)
    version = db.session.get(TariffVersion, context["version_id"])
    version.status = "EFFECTIVE"
    version.effective_at = datetime.utcnow()
    assignment = db.session.get(WorkloadAssignment, assignment_id)
    assignment.status = "CONFIRMED"
    activity = assignment.workload_need.education_activity
    subject = Subject(
        name=f"{activity.name} {context['version_id']}",
        education_activity_id=activity.id,
    )
    db.session.add(subject)
    db.session.flush()
    group = assignment.workload_need.teaching_group
    year = version.tariff_cycle.academic_year
    db.session.add(TeacherLoad(
        teacher_id=teacher_id,
        subject_id=subject.id,
        academic_year_id=year.id,
        department_id=assignment.department_id,
        building_id=assignment.building_id,
        class_name=group.name,
        group_name=group.name,
        hours=float(legacy_hours),
        subject_name=subject.name,
        building_name=assignment.building.name,
    ))
    db.session.commit()
    return context, version, year


def test_stage_eight_reconciles_switches_and_rolls_back(
    app,
    make_user,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_NEW_SOURCE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        context, version, year = _effective_integration_context(
            app,
            admin_id,
            teacher_id,
            "5",
        )

        run = reconcile_workload_sources(version, user_id=admin_id)
        assert run.status == "PASSED"
        assert run.blocking_count == 0
        assert run.internal_row_count == 1

        switch_workload_source(
            year,
            mode="COMPARE",
            tariff_version=version,
            user_id=admin_id,
            reason="Параллельная проверка",
        )
        switch_workload_source(
            year,
            mode="INTERNAL",
            tariff_version=version,
            user_id=admin_id,
            reason="Сверка завершена",
        )
        db.session.commit()

        state = source_state(year.id)
        assert state.configured_mode == "INTERNAL"
        assert state.effective_mode == "INTERNAL"
        rows = internal_department_load_rows(
            version,
            department_id=context["department_id"],
        )
        assert len(rows) == 1
        assert rows[0].hours == 5.0
        assert rows[0].teacher.id == teacher_id

        switch_workload_source(
            year,
            mode="LEGACY",
            tariff_version=version,
            user_id=admin_id,
            reason="Проверка аварийного отката",
        )
        db.session.commit()
        assert source_state(year.id).effective_mode == "LEGACY"
        assert WorkloadSourceTransition.query.count() == 3


def test_stage_eight_blocks_cutover_when_hours_differ(app, make_user):
    app.config["FEATURE_WORKLOAD_NEW_SOURCE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        _, version, year = _effective_integration_context(
            app,
            admin_id,
            teacher_id,
            "4",
        )
        run = reconcile_workload_sources(version, user_id=admin_id)
        assert run.status == "FAILED"
        assert run.blocking_count == 1

        with pytest.raises(
            WorkloadIntegrationError,
            match="успешная сверка",
        ):
            switch_workload_source(
                year,
                mode="INTERNAL",
                tariff_version=version,
                user_id=admin_id,
                reason="Нельзя применять",
            )


def test_stage_eight_feature_flag_forces_legacy_fallback(app, make_user):
    admin_id = make_user("ADMIN")
    with app.app_context():
        year = AcademicYear(
            name="2030/2031",
            is_current=True,
            start_date=date(2030, 9, 1),
            end_date=date(2031, 8, 31),
        )
        db.session.add(year)
        db.session.flush()
        db.session.add(WorkloadSourceSetting(
            academic_year_id=year.id,
            source_mode="INTERNAL",
            revision=2,
            change_reason="Тест",
            changed_by_user_id=admin_id,
        ))
        db.session.commit()

        app.config["FEATURE_WORKLOAD_NEW_SOURCE_ENABLED"] = False
        state = source_state(year.id)
        assert state.configured_mode == "INTERNAL"
        assert state.effective_mode == "LEGACY"


def test_internal_department_load_remains_available_without_integration_page(
    app,
    client,
    make_user,
    login,
):
    app.config["FEATURE_WORKLOAD_MODULE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_WRITE_ENABLED"] = True
    app.config["FEATURE_WORKLOAD_NEW_SOURCE_ENABLED"] = True
    admin_id = make_user("ADMIN")
    teacher_id = make_user("TEACHER")
    with app.app_context():
        _, version, year = _effective_integration_context(
            app,
            admin_id,
            teacher_id,
            "5",
        )
        run = reconcile_workload_sources(version, user_id=admin_id)
        switch_workload_source(
            year,
            mode="INTERNAL",
            tariff_version=version,
            user_id=admin_id,
            reason="Тест маршрута",
        )
        db.session.commit()
        run_id = run.id
        year_id = year.id

    login(admin_id)
    integration_page = client.get(
        f"/workload/integration/?academic_year_id={year_id}&run_id={run_id}"
    )
    assert integration_page.status_code == 404

    load_page = client.get(
        f"/departments/loads?academic_year_id={year_id}"
    )
    assert load_page.status_code == 200
    html = load_page.get_data(as_text=True)
    assert "данные поступают из матрицы распределения нагрузки" in html.lower()
    assert "Импортировать Excel" not in html
    assert "Добавить нагрузку вручную" not in html
