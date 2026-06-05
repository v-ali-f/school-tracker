
import argparse

from datetime import datetime, date

from pathlib import Path

from openpyxl import load_workbook

from openpyxl.utils.datetime import from_excel

from app import create_app

from app.core.extensions import db

from app.models import DocumentRegistryRecord

def parse_date(value):

    if value is None or value == "":

        return None

    if isinstance(value, datetime):

        return value.date()

    if isinstance(value, date):

        return value

    if isinstance(value, (int, float)):

        try:

            return from_excel(value).date()

        except Exception:

            return None

    raw = str(value).strip()

    if not raw:

        return None

    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):

        try:

            return datetime.strptime(raw, fmt).date()

        except ValueError:

            pass

    return None

def cell_text(value):

    if value is None:

        return None

    raw = str(value).strip()

    return raw or None

def empty_row(values):

    return all(value is None or str(value).strip() == "" for value in values)

def registry_type_by_sheet(sheet_name):

    name = str(sheet_name or "").lower().replace("ё", "е")

    if "вход" in name:

        return "incoming"

    if "исход" in name:

        return "outgoing"

    return None

def find_header_row(ws):

    for row in range(1, min(ws.max_row, 25) + 1):

        values = [cell_text(ws.cell(row=row, column=col).value) for col in range(1, min(ws.max_column, 12) + 1)]

        joined = " ".join(v.lower() for v in values if v)

        if "дата" in joined and ("№" in joined or "номер" in joined or "содержание" in joined):

            return row

    return 1

def exists_record(registry_type, doc_date, number, subject):

    return DocumentRegistryRecord.query.filter(

        DocumentRegistryRecord.registry_type == registry_type,

        DocumentRegistryRecord.doc_date == doc_date,

        DocumentRegistryRecord.number == number,

        DocumentRegistryRecord.subject == subject,

    ).first() is not None

def add_note(parts, label, value):

    value = cell_text(value)

    if value:

        parts.append(f"{label}: {value}")

def import_incoming(ws):

    added = 0

    skipped = 0

    warnings = []

    header_row = find_header_row(ws)

    for row in range(header_row + 1, ws.max_row + 1):

        values = [ws.cell(row=row, column=col).value for col in range(1, 9)]

        if empty_row(values):

            continue

        doc_date = parse_date(values[0])

        number = cell_text(values[1])

        subject = cell_text(values[2])

        executor = cell_text(values[3])

        order_book = cell_text(values[4])

        execution_note = cell_text(values[5])

        related_order_number = cell_text(values[6])

        related_order_date = parse_date(values[7])

        if not doc_date and not number and not subject:

            skipped += 1

            continue

        if not doc_date or not number or not subject:

            warnings.append(f"{ws.title}, строка {row}: пропущено — нет даты, номера или содержания")

            skipped += 1

            continue

        if exists_record("incoming", doc_date, number, subject):

            skipped += 1

            continue

        notes = []

        add_note(notes, "Исполнитель", executor)

        add_note(notes, "Книга приказов", order_book)

        add_note(notes, "Отметка о выполнении / примечание", execution_note)

        add_note(notes, "Номер приказа", related_order_number)

        if related_order_date:

            notes.append(f"Дата приказа: {related_order_date.strftime('%d.%m.%Y')}")

        notes.append(f"Источник: импорт из листа «{ws.title}», строка {row}")

        item = DocumentRegistryRecord(

            registry_type="incoming",

            number=number,

            doc_date=doc_date,

            subject=subject,

            correspondent=None,

            delivery_method="Импорт из книги входящих",

            status="registered",

            notes="\n".join(notes),

            created_by_id=None,

        )

        db.session.add(item)

        added += 1

    return added, skipped, warnings

def import_outgoing(ws):

    added = 0

    skipped = 0

    warnings = []

    header_row = find_header_row(ws)

    for row in range(header_row + 1, ws.max_row + 1):

        values = [ws.cell(row=row, column=col).value for col in range(1, 8)]

        if empty_row(values):

            continue

        doc_date = parse_date(values[0])

        number = cell_text(values[1])

        subject = cell_text(values[2])

        executor = cell_text(values[3])

        recipient = cell_text(values[4])

        note = cell_text(values[5])

        if not doc_date and not number and not subject:

            skipped += 1

            continue

        if not doc_date or not number or not subject:

            warnings.append(f"{ws.title}, строка {row}: пропущено — нет даты, номера или содержания")

            skipped += 1

            continue

        if exists_record("outgoing", doc_date, number, subject):

            skipped += 1

            continue

        notes = []

        add_note(notes, "Исполнитель", executor)

        add_note(notes, "Примечание", note)

        notes.append(f"Источник: импорт из листа «{ws.title}», строка {row}")

        item = DocumentRegistryRecord(

            registry_type="outgoing",

            number=number,

            doc_date=doc_date,

            subject=subject,

            correspondent=recipient,

            delivery_method="Импорт из книги исходящих",

            status="registered",

            notes="\n".join(notes),

            created_by_id=None,

        )

        db.session.add(item)

        added += 1

    return added, skipped, warnings

def main():

    parser = argparse.ArgumentParser()

    parser.add_argument("xlsx_path")

    parser.add_argument("--dry-run", action="store_true")

    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)

    if not xlsx_path.exists():

        raise FileNotFoundError(f"Файл не найден: {xlsx_path}")

    app = create_app()

    with app.app_context():

        wb = load_workbook(xlsx_path, data_only=True)

        total_added = 0

        total_skipped = 0

        all_warnings = []

        for ws in wb.worksheets:

            registry_type = registry_type_by_sheet(ws.title)

            if registry_type == "incoming":

                added, skipped, warnings = import_incoming(ws)

            elif registry_type == "outgoing":

                added, skipped, warnings = import_outgoing(ws)

            else:

                print(f"Лист пропущен: {ws.title}")

                continue

            total_added += added

            total_skipped += skipped

            all_warnings.extend(warnings)

            print(f"{ws.title}: добавлено {added}, пропущено {skipped}")

        if args.dry_run:

            db.session.rollback()

            print("DRY RUN: изменения не сохранены.")

        else:

            db.session.commit()

            print("Импорт сохранён в базе.")

        print(f"Итого добавлено: {total_added}")

        print(f"Итого пропущено: {total_skipped}")

        if all_warnings:

            print("Предупреждения:")

            for warning in all_warnings[:50]:

                print(f"- {warning}")

            if len(all_warnings) > 50:

                print(f"... и ещё {len(all_warnings) - 50}")

if __name__ == "__main__":

    main()

