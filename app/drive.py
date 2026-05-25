"""Диск: личное / общее хранилище + сборы файлов.

См. модели в app/models/drive.py. MVP.
"""
import io
import os
import uuid
import zipfile
import mimetypes
from datetime import datetime, timedelta
from pathlib import Path

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, abort, send_file, current_app, jsonify,
)
from flask_login import login_required, current_user
from sqlalchemy import func, or_, and_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.core.extensions import db
from app.models.drive import (
    DriveItem, FileCollection, FileCollectionTarget, FileCollectionSubmission,
)
from app.models.users import User


drive_bp = Blueprint("drive", __name__, url_prefix="/drive")


# ── Конфигурация ──────────────────────────────────────────────────────
MAX_FILE_BYTES = 100 * 1024 * 1024          # 100 МБ на один файл
PER_USER_QUOTA_BYTES = 1024 * 1024 * 1024   # 1 ГБ на «Мои файлы»

ALLOWED_EXTS = {
    # документы
    "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx", "odt", "ods", "odp",
    "txt", "rtf", "csv",
    # картинки
    "png", "jpg", "jpeg", "gif", "webp", "bmp", "svg", "heic",
    # архивы
    "zip", "rar", "7z", "tar", "gz",
}

ROLE_LABELS = {
    "ADMIN": "Администратор",
    "DEPUTY_DIRECTOR": "Заместитель директора",
    "CLASS_TEACHER": "Классный руководитель",
    "TEACHER": "Учитель",
    "PSYCHOLOGIST": "Психолог",
    "SOCIAL_PEDAGOG": "Соц. педагог",
    "METHODIST": "Методист",
    "SPECIALIST": "Специалист",
}


def _upload_root() -> Path:
    base = Path(current_app.config["UPLOAD_FOLDER"]) / "drive"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _ext_of(filename: str) -> str:
    name = (filename or "").lower()
    if "." not in name:
        return ""
    return name.rsplit(".", 1)[1][:20]


def _is_allowed_ext(ext: str) -> bool:
    return ext in ALLOWED_EXTS


def _user_role_codes(user=None) -> set:
    from app.permissions import _user_role_codes as _u
    return _u(user)


def _user_label(u: User) -> str:
    parts = [getattr(u, "last_name", None), getattr(u, "first_name", None), getattr(u, "middle_name", None)]
    fio = " ".join([p for p in parts if p]).strip()
    return fio or u.username or f"user{u.id}"


def _bytes_used_private(user_id: int) -> int:
    total = (
        db.session.query(func.coalesce(func.sum(DriveItem.size_bytes), 0))
        .filter(
            DriveItem.owner_user_id == user_id,
            DriveItem.scope == "private",
            DriveItem.kind == "file",
            DriveItem.deleted_at.is_(None),
        )
        .scalar()
    )
    return int(total or 0)


def _can_view_item(item: DriveItem) -> bool:
    if item.scope == "public":
        return True
    return item.owner_user_id == current_user.id


def _can_delete_item(item: DriveItem) -> bool:
    # Свои — да. Общие — только владелец, который их загрузил.
    return item.owner_user_id == current_user.id


def _ensure_parent_ok(parent_id, scope: str) -> DriveItem | None:
    if not parent_id:
        return None
    parent = DriveItem.query.get(int(parent_id))
    if not parent or parent.kind != "folder" or parent.deleted_at is not None:
        abort(404)
    if parent.scope != scope:
        abort(400)
    if scope == "private" and parent.owner_user_id != current_user.id:
        abort(403)
    return parent


def _breadcrumb(item: DriveItem | None) -> list[DriveItem]:
    out = []
    cur = item
    # ограничим глубину чтобы не зациклиться на битых данных
    for _ in range(30):
        if not cur:
            break
        out.append(cur)
        cur = cur.parent
    return list(reversed(out))


# ──────────────────────────────────────────────────────────────────────
# Главная страница диска
# ──────────────────────────────────────────────────────────────────────
@drive_bp.route("/", methods=["GET"])
@login_required
def index():
    tab = (request.args.get("tab") or "mine").strip()
    if tab not in ("mine", "public"):
        tab = "mine"
    scope = "private" if tab == "mine" else "public"

    parent_id = request.args.get("parent", type=int)
    parent = None
    if parent_id:
        parent = DriveItem.query.get_or_404(parent_id)
        if parent.deleted_at is not None or parent.scope != scope:
            abort(404)
        if scope == "private" and parent.owner_user_id != current_user.id:
            abort(403)

    search_q = (request.args.get("q") or "").strip()
    kind_f = (request.args.get("kind") or "").strip()
    if kind_f not in ("file", "folder"):
        kind_f = ""
    owner_f = request.args.get("owner_id", type=int)
    sort_f = (request.args.get("sort") or "name_asc").strip()
    if sort_f not in ("name_asc", "name_desc", "size_desc", "size_asc",
                      "updated_desc", "updated_asc"):
        sort_f = "name_asc"

    # При активном поиске показываем плоский список по всему scope (не ограничивая
    # текущей папкой), иначе — содержимое текущей папки.
    base_q = DriveItem.query.filter(
        DriveItem.scope == scope,
        DriveItem.deleted_at.is_(None),
    )
    if scope == "private":
        base_q = base_q.filter(DriveItem.owner_user_id == current_user.id)

    if search_q:
        base_q = base_q.filter(DriveItem.name.ilike(f"%{search_q}%"))
    else:
        base_q = base_q.filter(DriveItem.parent_id == (parent.id if parent else None))

    if kind_f:
        base_q = base_q.filter(DriveItem.kind == kind_f)
    if owner_f and scope == "public":
        base_q = base_q.filter(DriveItem.owner_user_id == owner_f)

    if scope == "public":
        base_q = base_q.options(joinedload(DriveItem.owner))

    sort_map = {
        "name_asc":     (DriveItem.kind.desc(), DriveItem.name.asc()),
        "name_desc":    (DriveItem.kind.desc(), DriveItem.name.desc()),
        "size_desc":    (DriveItem.kind.desc(), DriveItem.size_bytes.desc().nullslast(), DriveItem.name.asc()),
        "size_asc":     (DriveItem.kind.desc(), DriveItem.size_bytes.asc().nullsfirst(), DriveItem.name.asc()),
        "updated_desc": (DriveItem.kind.desc(), DriveItem.updated_at.desc(), DriveItem.name.asc()),
        "updated_asc":  (DriveItem.kind.desc(), DriveItem.updated_at.asc(), DriveItem.name.asc()),
    }
    items = base_q.order_by(*sort_map[sort_f]).limit(500).all()

    # Owners list для фильтра — только public
    public_owners = []
    if scope == "public":
        owner_ids = (
            db.session.query(DriveItem.owner_user_id)
            .filter(DriveItem.scope == "public", DriveItem.deleted_at.is_(None))
            .distinct().all()
        )
        owner_ids = [r[0] for r in owner_ids if r[0]]
        if owner_ids:
            public_owners = (
                User.query.filter(User.id.in_(owner_ids))
                .order_by(User.last_name.asc().nullslast(), User.first_name.asc().nullslast())
                .all()
            )

    # Список папок-целей для «переместить» (только текущий scope, владельца).
    move_targets_q = DriveItem.query.filter(
        DriveItem.scope == scope,
        DriveItem.deleted_at.is_(None),
        DriveItem.kind == "folder",
    )
    if scope == "private":
        move_targets_q = move_targets_q.filter(DriveItem.owner_user_id == current_user.id)
    move_targets = move_targets_q.order_by(DriveItem.name.asc()).limit(500).all()

    used = _bytes_used_private(current_user.id)

    # счётчик «нужно сдать»
    pending_count = _pending_collections_count(current_user.id)

    return render_template(
        "drive_index.html",
        tab=tab,
        scope=scope,
        parent=parent,
        crumbs=_breadcrumb(parent),
        items=items,
        used_bytes=used,
        quota_bytes=PER_USER_QUOTA_BYTES,
        max_file_bytes=MAX_FILE_BYTES,
        allowed_exts=sorted(ALLOWED_EXTS),
        user_label=_user_label,
        pending_count=pending_count,
        search_q=search_q,
        move_targets=move_targets,
        kind_f=kind_f,
        owner_f=owner_f,
        sort_f=sort_f,
        public_owners=public_owners,
    )


# ──────────────────────────────────────────────────────────────────────
# Создание папки
# ──────────────────────────────────────────────────────────────────────
@drive_bp.route("/folder/new", methods=["POST"])
@login_required
def folder_new():
    tab = request.form.get("tab", "mine")
    scope = "private" if tab == "mine" else "public"
    name = (request.form.get("name") or "").strip()[:255]
    parent_id = request.form.get("parent_id", type=int)

    if not name:
        flash("Введите название папки.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=parent_id))

    _ensure_parent_ok(parent_id, scope)

    item = DriveItem(
        owner_user_id=current_user.id,
        parent_id=parent_id or None,
        kind="folder",
        scope=scope,
        name=name,
        size_bytes=0,
    )
    db.session.add(item)
    db.session.commit()
    flash(f"Папка «{name}» создана.", "success")
    return redirect(url_for("drive.index", tab=tab, parent=parent_id))


# ──────────────────────────────────────────────────────────────────────
# Загрузка файла
# ──────────────────────────────────────────────────────────────────────
@drive_bp.route("/upload", methods=["POST"])
@login_required
def upload():
    tab = request.form.get("tab", "mine")
    scope = "private" if tab == "mine" else "public"
    parent_id = request.form.get("parent_id", type=int)
    _ensure_parent_ok(parent_id, scope)

    files = request.files.getlist("files")
    if not files:
        flash("Файл не выбран.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=parent_id))

    saved = 0
    used_now = _bytes_used_private(current_user.id) if scope == "private" else 0

    for fs in files:
        if not fs or not fs.filename:
            continue
        original = fs.filename
        ext = _ext_of(original)
        if not _is_allowed_ext(ext):
            flash(f"Файл «{original}»: расширение .{ext} запрещено.", "danger")
            continue

        # Узнаём размер
        fs.stream.seek(0, os.SEEK_END)
        size = fs.stream.tell()
        fs.stream.seek(0)

        if size > MAX_FILE_BYTES:
            flash(f"Файл «{original}» больше 100 МБ.", "danger")
            continue

        if scope == "private" and used_now + size > PER_USER_QUOTA_BYTES:
            flash("Превышена квота 1 ГБ на «Мои файлы».", "danger")
            break

        # Папка под scope
        if scope == "private":
            target_dir = _upload_root() / "private" / str(current_user.id)
        else:
            target_dir = _upload_root() / "public"
        target_dir.mkdir(parents=True, exist_ok=True)

        unique = uuid.uuid4().hex
        stored_name = f"{unique}.{ext}" if ext else unique
        fs.save(target_dir / stored_name)

        rel_path = str((target_dir / stored_name).relative_to(Path(current_app.config["UPLOAD_FOLDER"])))
        mime = fs.mimetype or mimetypes.guess_type(original)[0]

        item = DriveItem(
            owner_user_id=current_user.id,
            parent_id=parent_id or None,
            kind="file",
            scope=scope,
            name=secure_filename(original) or original,
            mime=mime,
            size_bytes=size,
            storage_path=rel_path,
            ext=ext,
        )
        db.session.add(item)
        used_now += size
        saved += 1

    if saved:
        db.session.commit()
        flash(f"Загружено файлов: {saved}.", "success")
    return redirect(url_for("drive.index", tab=tab, parent=parent_id))


# ──────────────────────────────────────────────────────────────────────
# Скачивание / удаление / переименование
# ──────────────────────────────────────────────────────────────────────
@drive_bp.route("/file/<int:item_id>/download", methods=["GET"])
@login_required
def file_download(item_id):
    item = DriveItem.query.get_or_404(item_id)
    if item.kind != "file" or item.deleted_at is not None:
        abort(404)
    if not _can_view_item(item):
        abort(403)
    full = Path(current_app.config["UPLOAD_FOLDER"]) / item.storage_path
    if not full.exists():
        abort(404)
    return send_file(full, as_attachment=True, download_name=item.name)


@drive_bp.route("/item/<int:item_id>/delete", methods=["POST"])
@login_required
def item_delete(item_id):
    item = DriveItem.query.get_or_404(item_id)
    if not _can_delete_item(item):
        abort(403)
    tab = "mine" if item.scope == "private" else "public"
    parent_id = item.parent_id

    # Soft-delete: помечаем самого и всех потомков
    _soft_delete_subtree(item)
    db.session.commit()
    flash("Удалено.", "success")
    return redirect(url_for("drive.index", tab=tab, parent=parent_id))


def _soft_delete_subtree(item: DriveItem):
    now = datetime.utcnow()
    item.deleted_at = now
    if item.kind == "folder":
        stack = [item]
        while stack:
            cur = stack.pop()
            for child in cur.children.filter(DriveItem.deleted_at.is_(None)).all():
                child.deleted_at = now
                if child.kind == "folder":
                    stack.append(child)


@drive_bp.route("/item/<int:item_id>/rename", methods=["POST"])
@login_required
def item_rename(item_id):
    item = DriveItem.query.get_or_404(item_id)
    if item.owner_user_id != current_user.id:
        abort(403)
    new_name = (request.form.get("name") or "").strip()[:255]
    if new_name:
        item.name = new_name
        db.session.commit()
        flash("Имя обновлено.", "success")
    tab = "mine" if item.scope == "private" else "public"
    return redirect(url_for("drive.index", tab=tab, parent=item.parent_id))


@drive_bp.route("/item/<int:item_id>/move", methods=["POST"])
@login_required
def item_move(item_id):
    item = DriveItem.query.get_or_404(item_id)
    if item.owner_user_id != current_user.id:
        abort(403)
    raw = request.form.get("target_parent_id")
    target_id = None
    if raw and raw.strip():
        try:
            target_id = int(raw)
        except ValueError:
            target_id = None

    tab = "mine" if item.scope == "private" else "public"

    if target_id == item.id:
        flash("Нельзя переместить папку саму в себя.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=item.parent_id))

    new_parent = None
    if target_id:
        new_parent = DriveItem.query.get(target_id)
        if (not new_parent or new_parent.kind != "folder"
                or new_parent.deleted_at is not None
                or new_parent.scope != item.scope):
            flash("Папка-назначение недоступна.", "warning")
            return redirect(url_for("drive.index", tab=tab, parent=item.parent_id))
        if item.scope == "private" and new_parent.owner_user_id != current_user.id:
            abort(403)
        # Нельзя переместить папку в её собственного потомка.
        if item.kind == "folder":
            cur = new_parent
            for _ in range(40):
                if not cur:
                    break
                if cur.id == item.id:
                    flash("Нельзя переместить папку внутрь себя самой.", "warning")
                    return redirect(url_for("drive.index", tab=tab, parent=item.parent_id))
                cur = cur.parent

    item.parent_id = new_parent.id if new_parent else None
    db.session.commit()
    flash("Перемещено.", "success")
    return redirect(url_for("drive.index", tab=tab, parent=item.parent_id))


def _parse_id_list(raw_list) -> list[int]:
    out = []
    for r in raw_list or []:
        try:
            out.append(int(r))
        except (TypeError, ValueError):
            pass
    return out


@drive_bp.route("/items/bulk_delete", methods=["POST"])
@login_required
def items_bulk_delete():
    ids = _parse_id_list(request.form.getlist("ids"))
    tab = request.form.get("tab", "mine")
    parent_id = request.form.get("parent_id") or None
    if not ids:
        flash("Ничего не выбрано.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=parent_id))
    items = DriveItem.query.filter(
        DriveItem.id.in_(ids),
        DriveItem.deleted_at.is_(None),
    ).all()
    deleted = 0
    for it in items:
        if it.owner_user_id != current_user.id:
            continue
        _soft_delete_subtree(it)
        deleted += 1
    if deleted:
        db.session.commit()
        flash(f"Удалено: {deleted}.", "success")
    else:
        flash("Нечего удалять (нет прав).", "warning")
    return redirect(url_for("drive.index", tab=tab, parent=parent_id))


@drive_bp.route("/items/bulk_move", methods=["POST"])
@login_required
def items_bulk_move():
    ids = _parse_id_list(request.form.getlist("ids"))
    tab = request.form.get("tab", "mine")
    parent_id = request.form.get("parent_id") or None
    raw_target = request.form.get("target_parent_id")
    target_id = None
    if raw_target and raw_target.strip():
        try:
            target_id = int(raw_target)
        except ValueError:
            target_id = None
    if not ids:
        flash("Ничего не выбрано.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=parent_id))

    items = DriveItem.query.filter(
        DriveItem.id.in_(ids),
        DriveItem.deleted_at.is_(None),
    ).all()
    if not items:
        flash("Не нашёл выбранные элементы.", "warning")
        return redirect(url_for("drive.index", tab=tab, parent=parent_id))

    scope = items[0].scope
    new_parent = None
    if target_id:
        new_parent = DriveItem.query.get(target_id)
        if (not new_parent or new_parent.kind != "folder"
                or new_parent.deleted_at is not None
                or new_parent.scope != scope):
            flash("Папка-назначение недоступна.", "warning")
            return redirect(url_for("drive.index", tab=tab, parent=parent_id))
        if scope == "private" and new_parent.owner_user_id != current_user.id:
            abort(403)

    moved = 0
    for it in items:
        if it.owner_user_id != current_user.id:
            continue
        if it.scope != scope:
            continue
        # Защита от перемещения папки в саму себя/потомка
        if new_parent and it.kind == "folder":
            cur = new_parent
            cycle = False
            for _ in range(40):
                if not cur:
                    break
                if cur.id == it.id:
                    cycle = True
                    break
                cur = cur.parent
            if cycle:
                continue
        if new_parent and target_id == it.id:
            continue
        it.parent_id = new_parent.id if new_parent else None
        moved += 1
    if moved:
        db.session.commit()
        flash(f"Перемещено: {moved}.", "success")
    else:
        flash("Ничего не перемещено.", "warning")
    return redirect(url_for("drive.index", tab=tab, parent=parent_id))


# ──────────────────────────────────────────────────────────────────────
# Сборы файлов
# ──────────────────────────────────────────────────────────────────────
def _user_in_collection_targets(collection: FileCollection, user_id: int) -> bool:
    user = User.query.get(user_id)
    if not user:
        return False
    user_roles = _user_role_codes(user)
    targets = collection.targets.all()
    for t in targets:
        if t.user_id and t.user_id == user_id:
            return True
        if t.role_code and t.role_code in user_roles:
            return True
    return False


def _collection_addressee_ids(collection: FileCollection) -> set[int]:
    targets = collection.targets.all()
    user_ids = set()
    roles = set()
    for t in targets:
        if t.user_id:
            user_ids.add(t.user_id)
        if t.role_code:
            roles.add(str(t.role_code).upper())
    if roles:
        # Учитываем И legacy `user.role`, И M2M `user.roles`. Без второго юзеры,
        # которым роль выдана только через новую схему, не попадали в адресаты.
        # Уволенных/архивных не включаем: is_active_user=False либо employment_status != ACTIVE.
        from app.models.users import User as U
        all_users = U.query.options(joinedload(U.roles)).all()
        for u in all_users:
            if not getattr(u, "is_active_user", True):
                continue
            emp = (getattr(u, "employment_status", "") or "").upper()
            if emp and emp != "ACTIVE":
                continue
            codes = set()
            legacy = getattr(u, "role", None)
            if legacy:
                codes.add(str(legacy).upper())
            for r in (getattr(u, "roles", None) or []):
                code = getattr(r, "code", None)
                if code:
                    codes.add(str(code).upper())
            if codes & roles:
                user_ids.add(u.id)
    return user_ids


def _notify_new_collection(collection: FileCollection) -> None:
    """Уведомить адресатов о новом сборе через MAX-бот и e-mail (в фоне).

    Канал bell сейчас завязан на инциденты (IncidentNotification.incident_id NOT NULL),
    поэтому для drive используем только MAX и email. Настройки notify_inc/task не
    применяем — это разовое уведомление о новой ответственности; пользователь увидит
    его в любом случае, если канал в принципе подключён.
    """
    try:
        addressee_ids = _collection_addressee_ids(collection)
    except Exception:
        current_app.logger.exception("collection notify: addressee resolve failed")
        return
    if not addressee_ids:
        return

    users = User.query.filter(User.id.in_(addressee_ids)).all()
    owner_name = _user_label(collection.owner) if collection.owner else ""
    deadline = collection.deadline_at.strftime('%d.%m.%Y %H:%M') if collection.deadline_at else ""
    base_url = (current_app.config.get("APP_BASE_URL") or "").rstrip("/")
    link = f"{base_url}/drive/collections/{collection.id}" if base_url else ""

    text_lines = [f"📂 Новый сбор файлов: «{collection.title}»."]
    if owner_name:
        text_lines.append(f"От: {owner_name}")
    if deadline:
        text_lines.append(f"Дедлайн: {deadline}")
    if link:
        text_lines.append(f"Открыть: {link}")
    text = "\n".join(text_lines)

    subject = f"[Диск] Новый сбор файлов: {collection.title}"

    try:
        from app.children import _send_max_notification
    except Exception:
        _send_max_notification = None
    try:
        from app.services.mail_settings_service import get_mail_config, send_mail_via_config
    except Exception:
        get_mail_config = send_mail_via_config = None

    _app = current_app._get_current_object()

    def _bg():
        with _app.app_context():
            cfg = None
            try:
                if get_mail_config:
                    cfg = get_mail_config()
            except Exception:
                cfg = None
            for u in users:
                # MAX
                if _send_max_notification:
                    try:
                        _send_max_notification(u, text)
                    except Exception:
                        _app.logger.exception("collection MAX notify failed uid=%s", u.id)
                # email
                email_to = (getattr(u, "email", "") or "").strip()
                if cfg and send_mail_via_config and email_to and cfg.get("smtp_host"):
                    try:
                        send_mail_via_config(recipient=email_to, subject=subject, body=text)
                    except Exception:
                        _app.logger.exception("collection email notify failed uid=%s", u.id)

    import threading
    threading.Thread(target=_bg, name="drive-col-notify", daemon=True).start()


def _pending_collections_count(user_id: int) -> int:
    """Сколько открытых сборов, где user адресат и ещё не сдал ни одного файла."""
    # Открытые сборы где user адресат — через подзапросы по user_id / role
    user = User.query.get(user_id)
    if not user:
        return 0
    user_roles = _user_role_codes(user)
    role_match = FileCollectionTarget.role_code.in_(user_roles) if user_roles else None
    user_match = FileCollectionTarget.user_id == user_id
    cond = user_match if role_match is None else or_(user_match, role_match)
    target_subq = (
        db.session.query(FileCollectionTarget.collection_id)
        .filter(cond)
        .subquery()
    )
    submitted_subq = (
        db.session.query(FileCollectionSubmission.collection_id)
        .filter(FileCollectionSubmission.user_id == user_id)
        .subquery()
    )
    return (
        FileCollection.query
        .filter(FileCollection.status == "open")
        .filter(FileCollection.id.in_(target_subq))
        .filter(~FileCollection.id.in_(submitted_subq))
        .count()
    )


@drive_bp.route("/collections", methods=["GET"])
@login_required
def collections_list():
    tab = (request.args.get("tab") or "incoming").strip()
    if tab not in ("incoming", "owned"):
        tab = "incoming"

    user_roles = _user_role_codes()

    if tab == "owned":
        items = (
            FileCollection.query
            .filter(FileCollection.owner_user_id == current_user.id)
            .order_by(FileCollection.status.asc(), FileCollection.created_at.desc())
            .all()
        )
        # для каждого считаем submissions count
        rows = []
        for c in items:
            submitted = (
                db.session.query(func.count(func.distinct(FileCollectionSubmission.user_id)))
                .filter(FileCollectionSubmission.collection_id == c.id)
                .scalar() or 0
            )
            total = len(_collection_addressee_ids(c))
            rows.append({"col": c, "submitted_users": submitted, "total_users": total})
    else:
        # «Мне нужно сдать» — все открытые где я адресат
        role_match = FileCollectionTarget.role_code.in_(user_roles) if user_roles else None
        user_match = FileCollectionTarget.user_id == current_user.id
        cond = user_match if role_match is None else or_(user_match, role_match)
        target_subq = (
            db.session.query(FileCollectionTarget.collection_id).filter(cond).subquery()
        )
        items = (
            FileCollection.query
            .filter(FileCollection.id.in_(target_subq))
            .order_by(FileCollection.status.asc(), FileCollection.deadline_at.asc())
            .all()
        )
        rows = []
        for c in items:
            my_subs = (
                FileCollectionSubmission.query
                .filter_by(collection_id=c.id, user_id=current_user.id)
                .count()
            )
            rows.append({"col": c, "my_subs": my_subs})

    return render_template(
        "drive_collections.html",
        tab=tab,
        rows=rows,
        now=datetime.utcnow(),
    )


@drive_bp.route("/collections/new", methods=["GET", "POST"])
@login_required
def collections_new():
    if request.method == "POST":
        title = (request.form.get("title") or "").strip()[:200]
        description = (request.form.get("description") or "").strip() or None
        max_files = max(1, min(50, int(request.form.get("max_files_per_user") or 1)))
        deadline_raw = (request.form.get("deadline_at") or "").strip()
        allow_late = request.form.get("allow_late") == "on"

        role_codes = request.form.getlist("target_roles")
        user_ids_raw = request.form.getlist("target_user_ids")
        user_ids = []
        for u in user_ids_raw:
            try:
                user_ids.append(int(u))
            except (TypeError, ValueError):
                pass

        if not title:
            flash("Укажите название сбора.", "warning")
            return redirect(url_for("drive.collections_new"))
        if not deadline_raw:
            flash("Укажите дедлайн.", "warning")
            return redirect(url_for("drive.collections_new"))
        try:
            deadline_at = datetime.fromisoformat(deadline_raw)
        except ValueError:
            flash("Некорректная дата дедлайна.", "warning")
            return redirect(url_for("drive.collections_new"))
        if not role_codes and not user_ids:
            flash("Выберите хотя бы одного адресата (роль или пользователя).", "warning")
            return redirect(url_for("drive.collections_new"))

        c = FileCollection(
            owner_user_id=current_user.id,
            title=title,
            description=description,
            max_files_per_user=max_files,
            deadline_at=deadline_at,
            allow_late=allow_late,
            status="open",
        )
        db.session.add(c)
        db.session.flush()
        for r in role_codes:
            db.session.add(FileCollectionTarget(collection_id=c.id, role_code=r))
        for uid in user_ids:
            db.session.add(FileCollectionTarget(collection_id=c.id, user_id=uid))
        db.session.commit()
        try:
            _notify_new_collection(c)
        except Exception:
            current_app.logger.exception("collection notify dispatch failed")
        flash("Сбор создан. Адресаты увидят его в разделе «Сборы», уведомления отправлены.", "success")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    # GET
    users = (
        User.query
        .order_by(User.last_name.asc().nullslast(), User.first_name.asc().nullslast(), User.username.asc())
        .limit(2000)
        .all()
    )
    return render_template(
        "drive_collection_new.html",
        users=users,
        role_options=list(ROLE_LABELS.items()),
        user_label=_user_label,
    )


@drive_bp.route("/collections/<int:col_id>", methods=["GET"])
@login_required
def collection_detail(col_id):
    c = FileCollection.query.options(joinedload(FileCollection.owner)).get_or_404(col_id)
    is_owner = c.owner_user_id == current_user.id
    is_addressee = _user_in_collection_targets(c, current_user.id)
    if not (is_owner or is_addressee):
        abort(403)

    my_subs = (
        FileCollectionSubmission.query
        .filter_by(collection_id=c.id, user_id=current_user.id)
        .order_by(FileCollectionSubmission.created_at.desc())
        .all()
    )

    by_user_rows = []
    pending_users = []
    if is_owner:
        addressee_ids = _collection_addressee_ids(c)
        # Submissions group by user
        subs = (
            FileCollectionSubmission.query.options(joinedload(FileCollectionSubmission.user))
            .filter_by(collection_id=c.id)
            .order_by(FileCollectionSubmission.created_at.desc())
            .all()
        )
        by_user = {}
        for s in subs:
            by_user.setdefault(s.user_id, []).append(s)
        for uid in addressee_ids:
            u = User.query.get(uid)
            if not u:
                continue
            by_user_rows.append({"user": u, "subs": by_user.get(uid, [])})
            if not by_user.get(uid):
                pending_users.append(u)
        # сортировка: сначала несдавшие, потом сдавшие, по фамилии
        by_user_rows.sort(key=lambda r: (1 if r["subs"] else 0, _user_label(r["user"]).lower()))

    targets = c.targets.all()

    # s123: для адресатов-ролей разворачиваем фактический список пользователей,
    # чтобы создатель видел, кто именно получит сбор. Только для is_owner — другим не нужно.
    targets_resolved = []
    if is_owner:
        from app.models.users import User as _U
        for t in targets:
            if t.role_code:
                role_users = (
                    _U.query.filter(_U.role == t.role_code)
                    .order_by(_U.last_name.asc().nullslast(), _U.first_name.asc().nullslast())
                    .all()
                )
                targets_resolved.append({
                    "kind": "role",
                    "code": t.role_code,
                    "label": ROLE_LABELS.get(t.role_code, t.role_code),
                    "users": role_users,
                })
            else:
                targets_resolved.append({"kind": "user", "user": t.user})

    return render_template(
        "drive_collection_detail.html",
        c=c,
        is_owner=is_owner,
        is_addressee=is_addressee,
        my_subs=my_subs,
        by_user_rows=by_user_rows,
        pending_users=pending_users,
        targets=targets,
        targets_resolved=targets_resolved,
        role_labels=ROLE_LABELS,
        user_label=_user_label,
        max_file_bytes=MAX_FILE_BYTES,
        allowed_exts=sorted(ALLOWED_EXTS),
        now=datetime.utcnow(),
    )


@drive_bp.route("/collections/<int:col_id>/submit", methods=["POST"])
@login_required
def collection_submit(col_id):
    c = FileCollection.query.get_or_404(col_id)
    if not _user_in_collection_targets(c, current_user.id):
        abort(403)
    if c.status != "open":
        flash("Сбор закрыт.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))
    if (not c.allow_late) and datetime.utcnow() > c.deadline_at:
        flash("Дедлайн прошёл, поздняя сдача запрещена.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    files = request.files.getlist("files")
    if not files:
        flash("Файл не выбран.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    have = (
        FileCollectionSubmission.query
        .filter_by(collection_id=c.id, user_id=current_user.id).count()
    )
    slots_left = max(0, c.max_files_per_user - have)
    if slots_left <= 0:
        flash(f"Вы уже сдали максимум ({c.max_files_per_user}). Удалите старый файл, чтобы заменить.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    target_dir = _upload_root() / "collections" / str(c.id) / str(current_user.id)
    target_dir.mkdir(parents=True, exist_ok=True)

    saved = 0
    for fs in files:
        if saved >= slots_left:
            flash(f"Загружено {saved} файлов — достигнут лимит {c.max_files_per_user}.", "info")
            break
        if not fs or not fs.filename:
            continue
        original = fs.filename
        ext = _ext_of(original)
        if not _is_allowed_ext(ext):
            flash(f"Файл «{original}»: расширение .{ext} запрещено.", "danger")
            continue
        fs.stream.seek(0, os.SEEK_END)
        size = fs.stream.tell()
        fs.stream.seek(0)
        if size > MAX_FILE_BYTES:
            flash(f"Файл «{original}» больше 100 МБ.", "danger")
            continue

        unique = uuid.uuid4().hex
        stored_name = f"{unique}.{ext}" if ext else unique
        fs.save(target_dir / stored_name)
        rel_path = str((target_dir / stored_name).relative_to(Path(current_app.config["UPLOAD_FOLDER"])))
        mime = fs.mimetype or mimetypes.guess_type(original)[0]

        sub = FileCollectionSubmission(
            collection_id=c.id,
            user_id=current_user.id,
            file_name=secure_filename(original) or original,
            storage_path=rel_path,
            mime=mime,
            size_bytes=size,
            ext=ext,
        )
        db.session.add(sub)
        saved += 1

    if saved:
        db.session.commit()
        flash(f"Сдано файлов: {saved}.", "success")
    return redirect(url_for("drive.collection_detail", col_id=c.id))


@drive_bp.route("/collections/<int:col_id>/submission/<int:sub_id>/download", methods=["GET"])
@login_required
def collection_submission_download(col_id, sub_id):
    c = FileCollection.query.get_or_404(col_id)
    sub = FileCollectionSubmission.query.get_or_404(sub_id)
    if sub.collection_id != c.id:
        abort(404)
    # Скачать может: владелец сбора или сам автор сдачи.
    if not (c.owner_user_id == current_user.id or sub.user_id == current_user.id):
        abort(403)
    full = Path(current_app.config["UPLOAD_FOLDER"]) / sub.storage_path
    if not full.exists():
        abort(404)
    return send_file(full, as_attachment=True, download_name=sub.file_name)


@drive_bp.route("/collections/<int:col_id>/submission/<int:sub_id>/delete", methods=["POST"])
@login_required
def collection_submission_delete(col_id, sub_id):
    c = FileCollection.query.get_or_404(col_id)
    sub = FileCollectionSubmission.query.get_or_404(sub_id)
    if sub.collection_id != c.id:
        abort(404)
    # Удалять может: владелец сбора (модерация) или сам автор (своя сдача и сбор ещё open).
    if not (c.owner_user_id == current_user.id or (sub.user_id == current_user.id and c.status == "open")):
        abort(403)
    try:
        full = Path(current_app.config["UPLOAD_FOLDER"]) / sub.storage_path
        if full.exists():
            full.unlink()
    except Exception:
        current_app.logger.exception("Failed to remove submission file")
    db.session.delete(sub)
    db.session.commit()
    flash("Файл удалён.", "success")
    return redirect(url_for("drive.collection_detail", col_id=c.id))


@drive_bp.route("/collections/<int:col_id>/zip", methods=["GET"])
@login_required
def collection_zip(col_id):
    """Скачать все сданные файлы одним архивом — только владельцу сбора."""
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    subs = (
        FileCollectionSubmission.query.options(joinedload(FileCollectionSubmission.user))
        .filter_by(collection_id=c.id)
        .order_by(FileCollectionSubmission.user_id.asc(),
                  FileCollectionSubmission.created_at.asc())
        .all()
    )
    if not subs:
        flash("Пока никто ничего не сдал.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    upload_root = Path(current_app.config["UPLOAD_FOLDER"])
    buf = io.BytesIO()
    seen_in_user: dict[int, set] = {}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in subs:
            full = upload_root / s.storage_path
            if not full.exists():
                continue
            user_label = _user_label(s.user) if s.user else f"user{s.user_id}"
            safe_user = "".join(ch if ch.isalnum() or ch in " -_." else "_" for ch in user_label).strip("_ ")
            safe_user = safe_user or f"user{s.user_id}"
            # Чтобы избежать коллизий имён внутри одного пользователя.
            used = seen_in_user.setdefault(s.user_id, set())
            base = s.file_name or f"file_{s.id}"
            arc_name = f"{safe_user}/{base}"
            i = 1
            while arc_name in used:
                stem, dot, ext = base.rpartition(".")
                if dot:
                    arc_name = f"{safe_user}/{stem} ({i}).{ext}"
                else:
                    arc_name = f"{safe_user}/{base} ({i})"
                i += 1
            used.add(arc_name)
            zf.write(full, arcname=arc_name)
    buf.seek(0)

    safe_title = "".join(ch if ch.isalnum() or ch in " -_." else "_" for ch in (c.title or f"col_{c.id}")).strip("_ ")[:60]
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Сбор_{c.id}_{safe_title}.zip",
        mimetype="application/zip",
    )


@drive_bp.route("/collections/<int:col_id>/export.xlsx", methods=["GET"])
@login_required
def collection_export_xlsx(col_id):
    """Excel-экспорт «кто сдал/не сдал» — только владельцу сбора."""
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        flash("Модуль openpyxl недоступен.", "danger")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    addressee_ids = _collection_addressee_ids(c)
    subs = (
        FileCollectionSubmission.query.options(joinedload(FileCollectionSubmission.user))
        .filter_by(collection_id=c.id)
        .order_by(FileCollectionSubmission.created_at.asc())
        .all()
    )
    subs_by_user: dict[int, list] = {}
    for s in subs:
        subs_by_user.setdefault(s.user_id, []).append(s)

    rows = []
    for uid in addressee_ids:
        u = User.query.get(uid)
        if not u:
            continue
        u_subs = subs_by_user.get(uid, [])
        rows.append({"user": u, "subs": u_subs})
    rows.sort(key=lambda r: (0 if r["subs"] else 1, _user_label(r["user"]).lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Сдачи"
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E7F1FF")
    headers = ["ФИО", "Логин", "E-mail", "Статус", "Кол-во файлов", "Последняя сдача", "Имена файлов"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    submitted = 0
    for r in rows:
        u = r["user"]
        u_subs = r["subs"]
        status = "Сдал" if u_subs else "Не сдал"
        if u_subs:
            submitted += 1
        last_at = u_subs[-1].created_at.strftime("%d.%m.%Y %H:%M") if u_subs else ""
        file_names = "; ".join(s.file_name for s in u_subs)
        ws.append([
            _user_label(u),
            getattr(u, "username", "") or "",
            (getattr(u, "email", "") or "").strip(),
            status,
            len(u_subs),
            last_at,
            file_names,
        ])

    # Авто-ширина столбцов (грубо).
    widths = [32, 20, 28, 10, 16, 18, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Сводка
    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Название сбора", c.title])
    ws2.append(["Создал", _user_label(c.owner) if c.owner else ""])
    ws2.append(["Дедлайн", c.deadline_at.strftime("%d.%m.%Y %H:%M") if c.deadline_at else ""])
    ws2.append(["Статус", "Открыт" if c.status == "open" else "Закрыт"])
    ws2.append(["Адресатов всего", len(rows)])
    ws2.append(["Сдали", submitted])
    ws2.append(["Не сдали", len(rows) - submitted])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = "".join(ch if ch.isalnum() or ch in " -_." else "_" for ch in (c.title or f"col_{c.id}")).strip("_ ")[:60]
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Сбор_{c.id}_{safe_title}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@drive_bp.route("/collections/<int:col_id>/remind", methods=["POST"])
@login_required
def collection_remind(col_id):
    """Повторно уведомить тех, кто ещё не сдал — владельцу сбора."""
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    if c.status != "open":
        flash("Сбор закрыт — напоминания не нужны.", "warning")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    addressee_ids = _collection_addressee_ids(c)
    submitted_ids = {
        s.user_id for s in
        FileCollectionSubmission.query
        .filter_by(collection_id=c.id)
        .with_entities(FileCollectionSubmission.user_id).distinct()
    }
    pending_ids = addressee_ids - submitted_ids
    if not pending_ids:
        flash("Все адресаты уже сдали — некого напоминать.", "info")
        return redirect(url_for("drive.collection_detail", col_id=c.id))

    users = User.query.filter(User.id.in_(pending_ids)).all()
    owner_name = _user_label(c.owner) if c.owner else ""
    deadline = c.deadline_at.strftime('%d.%m.%Y %H:%M') if c.deadline_at else ""
    base_url = (current_app.config.get("APP_BASE_URL") or "").rstrip("/")
    link = f"{base_url}/drive/collections/{c.id}" if base_url else ""
    text_lines = [f"🔔 Напоминание: вы ещё не сдали по сбору «{c.title}»."]
    if owner_name:
        text_lines.append(f"От: {owner_name}")
    if deadline:
        text_lines.append(f"Дедлайн: {deadline}")
    if link:
        text_lines.append(f"Открыть: {link}")
    text = "\n".join(text_lines)
    subject = f"[Диск] Напоминание по сбору: {c.title}"

    try:
        from app.children import _send_max_notification
    except Exception:
        _send_max_notification = None
    try:
        from app.services.mail_settings_service import get_mail_config, send_mail_via_config
    except Exception:
        get_mail_config = send_mail_via_config = None

    _app = current_app._get_current_object()

    def _bg():
        with _app.app_context():
            cfg = None
            try:
                if get_mail_config:
                    cfg = get_mail_config()
            except Exception:
                cfg = None
            for u in users:
                if _send_max_notification:
                    try:
                        _send_max_notification(u, text)
                    except Exception:
                        _app.logger.exception("remind MAX failed uid=%s", u.id)
                email_to = (getattr(u, "email", "") or "").strip()
                if cfg and send_mail_via_config and email_to and cfg.get("smtp_host"):
                    try:
                        send_mail_via_config(recipient=email_to, subject=subject, body=text)
                    except Exception:
                        _app.logger.exception("remind email failed uid=%s", u.id)

    import threading
    threading.Thread(target=_bg, name="drive-col-remind", daemon=True).start()
    flash(f"Напоминание отправлено {len(users)} пользователям.", "success")
    return redirect(url_for("drive.collection_detail", col_id=c.id))


@drive_bp.route("/collections/<int:col_id>/close", methods=["POST"])
@login_required
def collection_close(col_id):
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    c.status = "closed"
    c.closed_at = datetime.utcnow()
    db.session.commit()
    flash("Сбор закрыт.", "success")
    return redirect(url_for("drive.collection_detail", col_id=c.id))


@drive_bp.route("/collections/<int:col_id>/reopen", methods=["POST"])
@login_required
def collection_reopen(col_id):
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    c.status = "open"
    c.closed_at = None
    db.session.commit()
    flash("Сбор снова открыт.", "success")
    return redirect(url_for("drive.collection_detail", col_id=c.id))


@drive_bp.route("/collections/<int:col_id>/delete", methods=["POST"])
@login_required
def collection_delete(col_id):
    c = FileCollection.query.get_or_404(col_id)
    if c.owner_user_id != current_user.id:
        abort(403)
    # удаляем физические файлы сборщин
    base = _upload_root() / "collections" / str(c.id)
    try:
        if base.exists():
            for root, dirs, files in os.walk(base, topdown=False):
                for name in files:
                    try:
                        os.remove(os.path.join(root, name))
                    except Exception:
                        pass
                for name in dirs:
                    try:
                        os.rmdir(os.path.join(root, name))
                    except Exception:
                        pass
            try:
                base.rmdir()
            except Exception:
                pass
    except Exception:
        current_app.logger.exception("Failed to clean collection files")
    db.session.delete(c)
    db.session.commit()
    flash("Сбор удалён.", "success")
    return redirect(url_for("drive.collections_list", tab="owned"))
