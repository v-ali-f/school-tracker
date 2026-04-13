from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


@dataclass
class AttendanceRow:
    school_class: Optional[str]
    fio: Optional[str]
    attendance_date: date
    first_in: Optional[str]
    last_out: Optional[str]
    presence_duration: Optional[str]
    events: Optional[str]
    total_presence: Optional[str]


CLASS_RE = re.compile(r"^\d{1,2}\-[A-Za-zА-Яа-яЁё0-9]+$")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\n", " ").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def build_merged_lookup(ws) -> dict[tuple[int, int], tuple[int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = top_left
    return lookup


def get_merged_value(ws, merged_lookup: dict[tuple[int, int], tuple[int, int]], row: int, col: int) -> Any:
    source_row, source_col = merged_lookup.get((row, col), (row, col))
    return ws.cell(row=source_row, column=source_col).value


def is_class_name(value: Any) -> bool:
    text = clean_text(value)
    if not text:
        return False
    if normalize_text(text) in {"ф.и.о.", "дата", "вход", "выход", "присутствие", "входы/выходы", "всего"}:
        return False
    return bool(CLASS_RE.match(text))


def parse_excel_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None


def parse_excel_time(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%H:%M")

    if isinstance(value, time):
        return value.strftime("%H:%M")

    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"

    text = clean_text(value)
    if not text:
        return None

    text = text.replace(".", ":")
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", text)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2))
        return f"{hh:02d}:{mm:02d}"

    return None


def time_to_minutes(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    try:
        hh, mm = value.split(":")
        return int(hh) * 60 + int(mm)
    except Exception:
        return None


def find_header_row(ws, merged_lookup: dict[tuple[int, int], tuple[int, int]]) -> int:
    for row in range(1, min(ws.max_row, 20) + 1):
        row_values = []
        next_row_values = []
        for col in range(1, min(ws.max_column, 12) + 1):
            row_values.append(normalize_text(get_merged_value(ws, merged_lookup, row, col)))
            next_row_values.append(normalize_text(get_merged_value(ws, merged_lookup, row + 1, col)))

        combined = row_values + next_row_values
        if "ф.и.о." in combined and "дата" in combined and "вход" in combined and "выход" in combined:
            return row

    raise ValueError("Не удалось найти строку заголовка с колонками Ф.И.О., Дата, Вход и Выход.")


def detect_columns(ws, merged_lookup: dict[tuple[int, int], tuple[int, int]], header_row: int) -> dict[str, Optional[int]]:
    columns = {
        "fio": None,
        "date": None,
        "in": None,
        "out": None,
        "presence": None,
        "events": None,
        "total_presence": None,
    }

    headers_top: dict[int, str] = {}
    headers_bottom: dict[int, str] = {}
    headers_combined: dict[int, str] = {}

    for col in range(1, ws.max_column + 1):
        top = normalize_text(get_merged_value(ws, merged_lookup, header_row, col))
        bottom = normalize_text(get_merged_value(ws, merged_lookup, header_row + 1, col))
        combined = " ".join(part for part in [top, bottom] if part).strip()

        headers_top[col] = top
        headers_bottom[col] = bottom
        headers_combined[col] = combined

        if top == "ф.и.о." or bottom == "ф.и.о." or combined == "ф.и.о.":
            columns["fio"] = col
        elif top == "дата" or bottom == "дата" or combined == "дата":
            columns["date"] = col
        elif bottom == "вход" or top == "вход" or combined.endswith(" вход") or combined == "вход":
            columns["in"] = col
        elif bottom == "выход" or top == "выход" or combined.endswith(" выход") or combined == "выход":
            columns["out"] = col
        elif "входы/выходы" in combined:
            columns["events"] = col

    presence_cols = [
        col for col, combined in headers_combined.items()
        if combined.endswith(" присутствие") or combined == "присутствие"
    ]
    if len(presence_cols) >= 1:
        columns["presence"] = presence_cols[0]
    if len(presence_cols) >= 2:
        columns["total_presence"] = presence_cols[1]

    if not columns["fio"] or not columns["date"]:
        raise ValueError(
            f"Не удалось определить колонки Ф.И.О. и Дата в файле. "
            f"Найдено: ФИО={columns['fio']}, Дата={columns['date']}"
        )

    if not columns["in"] or not columns["out"]:
        header_debug = {
            get_column_letter(col): headers_combined[col]
            for col in range(1, ws.max_column + 1)
        }
        raise ValueError(
            f"Не удалось определить колонки Вход и Выход в файле. "
            f"Найдено: Вход={columns['in']}, Выход={columns['out']}. Заголовки: {header_debug}"
        )

    return columns


def parse_attendance_xlsx(file_path: str) -> list[AttendanceRow]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    merged_lookup = build_merged_lookup(ws)

    header_row = find_header_row(ws, merged_lookup)
    columns = detect_columns(ws, merged_lookup, header_row)

    current_class: Optional[str] = None
    current_fio: Optional[str] = None

    # ключ: (class, fio, date)
    aggregated: dict[tuple[str, str, date], dict[str, Any]] = {}

    for row in range(header_row + 1, ws.max_row + 1):
        raw_fio = get_merged_value(ws, merged_lookup, row, columns["fio"])
        raw_date = get_merged_value(ws, merged_lookup, row, columns["date"])

        # строка с названием класса
        if is_class_name(raw_fio):
            current_class = clean_text(raw_fio)
            current_fio = None
            continue

        fio_text = clean_text(raw_fio)
        if fio_text and not is_class_name(fio_text):
            low = normalize_text(fio_text)
            if low not in {"ф.и.о.", "дата", "вход", "выход", "присутствие", "входы/выходы", "всего"}:
                current_fio = fio_text

        attendance_date = parse_excel_date(raw_date)
        if not attendance_date:
            continue

        if not current_fio:
            continue

        key = (
            current_class or "Без класса",
            current_fio,
            attendance_date,
        )

        first_in = parse_excel_time(get_merged_value(ws, merged_lookup, row, columns["in"])) if columns["in"] else None
        last_out = parse_excel_time(get_merged_value(ws, merged_lookup, row, columns["out"])) if columns["out"] else None
        presence_duration = clean_text(get_merged_value(ws, merged_lookup, row, columns["presence"])) if columns["presence"] else None
        events = clean_text(get_merged_value(ws, merged_lookup, row, columns["events"])) if columns["events"] else None
        total_presence = clean_text(get_merged_value(ws, merged_lookup, row, columns["total_presence"])) if columns["total_presence"] else None

        if key not in aggregated:
            aggregated[key] = {
                "school_class": current_class,
                "fio": current_fio,
                "attendance_date": attendance_date,
                "first_in": first_in,
                "last_out": last_out,
                "presence_duration": presence_duration,
                "events": events,
                "total_presence": total_presence,
            }
        else:
            item = aggregated[key]

            # самый ранний вход
            old_in = time_to_minutes(item["first_in"])
            new_in = time_to_minutes(first_in)
            if new_in is not None and (old_in is None or new_in < old_in):
                item["first_in"] = first_in

            # самый поздний выход
            old_out = time_to_minutes(item["last_out"])
            new_out = time_to_minutes(last_out)
            if new_out is not None and (old_out is None or new_out > old_out):
                item["last_out"] = last_out

            # текст присутствия / итог
            if total_presence:
                item["total_presence"] = total_presence
            elif presence_duration and not item["presence_duration"]:
                item["presence_duration"] = presence_duration

            # события объединяем аккуратно
            if events:
                if item["events"]:
                    if events not in item["events"]:
                        item["events"] = f"{item['events']}; {events}"
                else:
                    item["events"] = events

    records: list[AttendanceRow] = []
    for _, item in sorted(aggregated.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        records.append(
            AttendanceRow(
                school_class=item["school_class"],
                fio=item["fio"],
                attendance_date=item["attendance_date"],
                first_in=item["first_in"],
                last_out=item["last_out"],
                presence_duration=item["presence_duration"],
                events=item["events"],
                total_presence=item["total_presence"],
            )
        )

    return records


def parse_attendance_file(file_path: str) -> list[dict[str, Any]]:
    rows = parse_attendance_xlsx(file_path)

    result: list[dict[str, Any]] = []
    for row in rows:
        result.append(
            {
                "school_class": row.school_class,
                "fio": row.fio,
                "date": row.attendance_date,
                "first_in": row.first_in,
                "last_out": row.last_out,
                "presence_duration": row.presence_duration,
                "events": row.events,
                "total_presence": row.total_presence,
            }
        )
    return result