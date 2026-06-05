from pathlib import Path
from flask import Blueprint, flash, redirect, render_template, request, url_for
from sqlalchemy import func, inspect, or_, text
from werkzeug.utils import secure_filename

from app.core.extensions import db
from app.models_legacy import AcademicYear, Building, User
from .models import PreschoolAttendanceRecord, PreschoolAttendanceUpload, PreschoolChild, PreschoolChildMovement, PreschoolChildrenImport, PreschoolGroup, PreschoolRepresentative


bp = Blueprint(
    "preschool",
    __name__,
    url_prefix="/preschool",
    template_folder="templates",
)

_preschool_schema_checked = False

PRESCHOOL_AGE_LEVELS = [
    "ранний возраст",
    "младшая группа",
    "средняя группа",
    "старшая группа",
    "подготовительная группа",
    "разновозрастная группа",
]


def _get_current_year():
    year = (
        AcademicYear.query
        .filter_by(is_current=True)
        .order_by(AcademicYear.id.desc())
        .first()
    )
    if year:
        return year

    return (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .first()
    )


def _add_column_if_missing(inspector, table_name, column_name, ddl):
    columns = {c["name"] for c in inspector.get_columns(table_name)}
    if column_name not in columns:
        db.session.execute(text(ddl))
        db.session.commit()


def ensure_preschool_tables():
    """Создаёт и мягко обновляет таблицы ДОУ на чистой/тестовой базе."""
    global _preschool_schema_checked

    if _preschool_schema_checked:
        return

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    for model in (PreschoolGroup, PreschoolChildrenImport, PreschoolChild, PreschoolRepresentative, PreschoolChildMovement, PreschoolAttendanceUpload, PreschoolAttendanceRecord):
        if model.__tablename__ not in existing_tables:
            model.__table__.create(db.engine, checkfirst=True)

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    # Если preschool_group уже была создана на первом черновом этапе,
    # добавляем новые поля без пересоздания таблицы.
    if "preschool_group" in existing_tables:
        try:
            _add_column_if_missing(
                inspector,
                "preschool_group",
                "academic_year_id",
                "ALTER TABLE preschool_group ADD COLUMN academic_year_id INTEGER",
            )
            _add_column_if_missing(
                inspector,
                "preschool_group",
                "teacher_user_id",
                "ALTER TABLE preschool_group ADD COLUMN teacher_user_id INTEGER",
            )
            _add_column_if_missing(
                inspector,
                "preschool_group",
                "is_archived",
                "ALTER TABLE preschool_group ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            )
        except Exception:
            db.session.rollback()

    if "preschool_child" in existing_tables:
        try:
            _add_column_if_missing(
                inspector,
                "preschool_child",
                "import_batch_id",
                "ALTER TABLE preschool_child ADD COLUMN import_batch_id INTEGER",
            )
            _add_column_if_missing(
                inspector,
                "preschool_child",
                "reg_address",
                "ALTER TABLE preschool_child ADD COLUMN reg_address VARCHAR(700)",
            )
            _add_column_if_missing(
                inspector,
                "preschool_child",
                "living_address",
                "ALTER TABLE preschool_child ADD COLUMN living_address VARCHAR(700)",
            )
            _add_column_if_missing(
                inspector,
                "preschool_child",
                "actual_address",
                "ALTER TABLE preschool_child ADD COLUMN actual_address VARCHAR(700)",
            )
        except Exception:
            db.session.rollback()

    if "preschool_representative" in existing_tables:
        try:
            _add_column_if_missing(
                inspector,
                "preschool_representative",
                "address",
                "ALTER TABLE preschool_representative ADD COLUMN address VARCHAR(700)",
            )
        except Exception:
            db.session.rollback()

    if "preschool_child_movement" in existing_tables:
        try:
            _add_column_if_missing(
                inspector,
                "preschool_child_movement",
                "from_academic_year_id",
                "ALTER TABLE preschool_child_movement ADD COLUMN from_academic_year_id INTEGER",
            )
            _add_column_if_missing(
                inspector,
                "preschool_child_movement",
                "to_academic_year_id",
                "ALTER TABLE preschool_child_movement ADD COLUMN to_academic_year_id INTEGER",
            )
        except Exception:
            db.session.rollback()

    _preschool_schema_checked = True


@bp.before_app_request
def before_request():
    ensure_preschool_tables()


@bp.route("/")
def index():
    groups_count = PreschoolGroup.query.count()
    children_count = PreschoolChild.query.count()

    return render_template(
        "preschool/index.html",
        groups_count=groups_count,
        children_count=children_count,
    )


@bp.route("/children")
def children():
    year_id = request.args.get("academic_year_id", type=int)
    building_id = request.args.get("building_id", type=int)

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )
    month_options = _academic_year_month_options(year)
    month_options = _academic_year_month_options(year)

    buildings = Building.query.order_by(Building.name.asc()).all()

    groups_query = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .outerjoin(Building, PreschoolGroup.building_id == Building.id)
    )

    if building_id:
        groups_query = groups_query.filter(PreschoolGroup.building_id == building_id)

    groups = (
        groups_query
        .order_by(Building.name.asc(), PreschoolGroup.name.asc())
        .all()
    )

    group_rows = []
    total_children = 0
    age_counts = {}
    building_ids = set()

    for group in groups:
        children_count = PreschoolChild.query.filter(PreschoolChild.group_id == group.id).count()
        total_children += children_count

        if group.building_id:
            building_ids.add(group.building_id)

        age_name = group.age_level or "Не указано"
        age_counts[age_name] = age_counts.get(age_name, 0) + children_count

        group_rows.append({
            "id": group.id,
            "name": group.name,
            "building": group.building.name if group.building else "—",
            "age_level": group.age_level or "—",
            "teacher": (
                f"{group.teacher.last_name or ''} {group.teacher.first_name or ''} {group.teacher.middle_name or ''}".strip()
                if group.teacher else (group.teacher_name or "—")
            ),
            "children_count": children_count,
        })

    stats = {
        "total": total_children,
        "groups": len(groups),
        "buildings": len(building_ids),
        "early": age_counts.get("ранний возраст", 0),
        "junior": age_counts.get("младшая группа", 0),
        "middle": age_counts.get("средняя группа", 0),
        "senior": age_counts.get("старшая группа", 0),
        "prep": age_counts.get("подготовительная группа", 0),
    }

    return render_template(
        "preschool/children.html",
        group_rows=group_rows,
        buildings=buildings,
        all_years=all_years,
        year=year,
        selected_building_id=building_id,
        stats=stats,
    )


@bp.route("/children/registry", methods=["GET", "POST"])
def children_registry():
    year_id = request.args.get("academic_year_id", type=int)
    building_id = request.args.get("building_id", type=int)
    group_id = request.args.get("group_id", type=int)
    q = (request.args.get("q") or "").strip()

    page = max(request.args.get("page", 1, type=int), 1)
    per_page = request.args.get("per_page", 20, type=int)
    if per_page not in (20, 50, 100):
        per_page = 20

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    buildings = Building.query.order_by(Building.name.asc()).all()

    groups_query = PreschoolGroup.query.filter(PreschoolGroup.academic_year_id == year.id)
    if building_id:
        groups_query = groups_query.filter(PreschoolGroup.building_id == building_id)

    groups = groups_query.order_by(PreschoolGroup.name.asc()).all()

    if request.method == "POST":
        last_name = (request.form.get("last_name") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        middle_name = (request.form.get("middle_name") or "").strip()
        birth_date_raw = (request.form.get("birth_date") or "").strip()
        personal_account = (request.form.get("personal_account") or "").strip()
        reg_address = (request.form.get("reg_address") or "").strip()
        living_address = (request.form.get("living_address") or "").strip()
        actual_address = (request.form.get("actual_address") or "").strip()
        note = (request.form.get("note") or "").strip()
        group_id_raw = (request.form.get("group_id") or "").strip()

        if not last_name or not first_name:
            flash("Укажите фамилию и имя воспитанника.", "warning")
            return redirect(url_for(
                "preschool.children_registry",
                academic_year_id=year.id,
                building_id=building_id,
                group_id=group_id,
                per_page=per_page,
            ))

        birth_date = None
        if birth_date_raw:
            from datetime import datetime
            try:
                birth_date = datetime.strptime(birth_date_raw, "%Y-%m-%d").date()
            except ValueError:
                birth_date = None

        child = PreschoolChild(
            group_id=int(group_id_raw) if group_id_raw.isdigit() else None,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name or None,
            birth_date=birth_date,
            status="active",
            note=note or None,
        )

        db.session.add(child)
        db.session.commit()

        flash("Воспитанник добавлен в реестр ДОУ.", "success")
        return redirect(url_for(
            "preschool.children_registry",
            academic_year_id=year.id,
            building_id=building_id,
            group_id=group_id,
            per_page=per_page,
        ))

    base_query = (
        PreschoolChild.query
        .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
        .filter(PreschoolGroup.academic_year_id == year.id)
    )

    if building_id:
        base_query = base_query.filter(PreschoolGroup.building_id == building_id)

    if group_id:
        base_query = base_query.filter(PreschoolChild.group_id == group_id)

    if q:
        like = f"%{q}%"
        base_query = base_query.filter(
            or_(
                PreschoolChild.last_name.ilike(like),
                PreschoolChild.first_name.ilike(like),
                PreschoolChild.middle_name.ilike(like),
            )
        )

    ordered_query = (
        base_query
        .order_by(
            PreschoolGroup.name.asc(),
            PreschoolChild.last_name.asc(),
            PreschoolChild.first_name.asc(),
        )
    )

    pagination = ordered_query.paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    items = pagination.items

    stats_query = (
        db.session.query(PreschoolGroup.age_level, func.count(PreschoolChild.id))
        .join(PreschoolChild, PreschoolChild.group_id == PreschoolGroup.id)
        .filter(PreschoolGroup.academic_year_id == year.id)
    )

    if building_id:
        stats_query = stats_query.filter(PreschoolGroup.building_id == building_id)

    if group_id:
        stats_query = stats_query.filter(PreschoolGroup.id == group_id)

    age_counts = {
        name or "Не указано": count
        for name, count in stats_query.group_by(PreschoolGroup.age_level).all()
    }

    groups_total_query = PreschoolGroup.query.filter(PreschoolGroup.academic_year_id == year.id)
    if building_id:
        groups_total_query = groups_total_query.filter(PreschoolGroup.building_id == building_id)

    buildings_used_query = (
        db.session.query(func.count(func.distinct(PreschoolGroup.building_id)))
        .filter(PreschoolGroup.academic_year_id == year.id)
        .filter(PreschoolGroup.building_id.isnot(None))
    )
    if building_id:
        buildings_used_query = buildings_used_query.filter(PreschoolGroup.building_id == building_id)

    stats = {
        "total": pagination.total,
        "groups": groups_total_query.count(),
        "buildings": buildings_used_query.scalar() or 0,
        "early": age_counts.get("ранний возраст", 0),
        "junior": age_counts.get("младшая группа", 0),
        "middle": age_counts.get("средняя группа", 0),
        "senior": age_counts.get("старшая группа", 0),
        "prep": age_counts.get("подготовительная группа", 0),
    }

    return render_template(
        "preschool/children_registry.html",
        items=items,
        groups=groups,
        buildings=buildings,
        all_years=all_years,
        year=year,
        selected_building_id=building_id,
        selected_group_id=group_id,
        q=q,
        pagination=pagination,
        page=page,
        per_page=per_page,
        stats=stats,
    )


@bp.route("/buildings")
def buildings_redirect():
    # Отдельного справочника корпусов ДОУ больше нет:
    # используем общий реестр зданий проекта.
    return redirect(url_for("children.buildings_registry"))


@bp.route("/groups", methods=["GET", "POST"])
def groups():
    year_id = request.args.get("academic_year_id", type=int)
    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )
    buildings = Building.query.order_by(Building.name.asc()).all()
    teachers = (
        User.query
        .order_by(User.last_name.asc(), User.first_name.asc(), User.username.asc())
        .all()
    )

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        age_level = (request.form.get("age_level") or "").strip()
        teacher_user_id_raw = (request.form.get("teacher_user_id") or "").strip()
        teacher_name = (request.form.get("teacher_name") or "").strip()
        building_id_raw = (request.form.get("building_id") or "").strip()
        academic_year_id_raw = (request.form.get("academic_year_id") or "").strip()

        if not name:
            flash("Укажите название группы.", "warning")
            return redirect(url_for("preschool.groups", academic_year_id=year.id))

        academic_year_id = int(academic_year_id_raw) if academic_year_id_raw.isdigit() else year.id
        building_id = int(building_id_raw) if building_id_raw.isdigit() else None
        teacher_user_id = int(teacher_user_id_raw) if teacher_user_id_raw.isdigit() else None

        group = PreschoolGroup(
            academic_year_id=academic_year_id,
            building_id=building_id,
            name=name,
            age_level=age_level or None,
            teacher_user_id=teacher_user_id,
            teacher_name=teacher_name or None,
            is_active=True,
            is_archived=False,
        )
        db.session.add(group)
        db.session.commit()

        flash("Группа ДОУ учебного года добавлена.", "success")
        return redirect(url_for("preschool.groups", academic_year_id=academic_year_id))

    items = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .outerjoin(Building, PreschoolGroup.building_id == Building.id)
        .order_by(Building.name.asc(), PreschoolGroup.name.asc())
        .all()
    )

    return render_template(
        "preschool/groups.html",
        items=items,
        buildings=buildings,
        teachers=teachers,
        age_levels=PRESCHOOL_AGE_LEVELS,
        all_years=all_years,
        year=year,
    )

@bp.route("/groups/<int:group_id>/edit", methods=["GET", "POST"])
def edit_group(group_id):
    group = PreschoolGroup.query.get_or_404(group_id)

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )
    buildings = Building.query.order_by(Building.name.asc()).all()
    teachers = (
        User.query
        .order_by(User.last_name.asc(), User.first_name.asc(), User.username.asc())
        .all()
    )

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        age_level = (request.form.get("age_level") or "").strip()
        teacher_user_id_raw = (request.form.get("teacher_user_id") or "").strip()
        teacher_name = (request.form.get("teacher_name") or "").strip()
        building_id_raw = (request.form.get("building_id") or "").strip()
        academic_year_id_raw = (request.form.get("academic_year_id") or "").strip()

        if not name:
            flash("Укажите название группы.", "warning")
            return redirect(url_for("preschool.edit_group", group_id=group.id))

        group.name = name
        group.age_level = age_level or None
        group.teacher_user_id = int(teacher_user_id_raw) if teacher_user_id_raw.isdigit() else None
        group.teacher_name = teacher_name or None
        group.building_id = int(building_id_raw) if building_id_raw.isdigit() else None
        group.academic_year_id = int(academic_year_id_raw) if academic_year_id_raw.isdigit() else group.academic_year_id

        db.session.commit()

        flash("Группа ДОУ обновлена.", "success")
        return redirect(url_for("preschool.groups", academic_year_id=group.academic_year_id))

    return render_template(
        "preschool/group_form.html",
        group=group,
        buildings=buildings,
        teachers=teachers,
        age_levels=PRESCHOOL_AGE_LEVELS,
        all_years=all_years,
    )


@bp.route("/groups/<int:group_id>/delete", methods=["POST"])
def delete_group(group_id):
    group = PreschoolGroup.query.get_or_404(group_id)
    academic_year_id = group.academic_year_id

    if group.children:
        flash("Нельзя удалить группу, если к ней уже прикреплены воспитанники. Позже добавим архивирование.", "warning")
        return redirect(url_for("preschool.groups", academic_year_id=academic_year_id))

    db.session.delete(group)
    db.session.commit()

    flash("Группа ДОУ удалена.", "success")
    return redirect(url_for("preschool.groups", academic_year_id=academic_year_id))

def _normalize_header(value):
    value = str(value or "").strip().lower()
    value = value.replace("ё", "е")
    value = value.replace(".", "")
    value = value.replace(" ", "")
    value = value.replace("_", "")
    return value


def _split_full_name(full_name):
    parts = [p for p in str(full_name or "").strip().split() if p]
    if not parts:
        return "", "", None

    last_name = parts[0] if len(parts) >= 1 else ""
    first_name = parts[1] if len(parts) >= 2 else ""
    middle_name = " ".join(parts[2:]) if len(parts) >= 3 else None
    return last_name, first_name, middle_name


def _parse_excel_date(value):
    if not value:
        return None

    from datetime import datetime, date

    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw:
        return None

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass

    return None


@bp.route("/children/import", methods=["GET", "POST"])
def import_children():
    year_id = request.args.get("academic_year_id", type=int) or request.form.get("academic_year_id", type=int)
    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    groups = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .order_by(PreschoolGroup.name.asc())
        .all()
    )
    groups_by_name = {g.name.strip().lower(): g for g in groups}
    buildings = Building.query.order_by(Building.name.asc()).all()
    buildings_by_name = {b.name.strip().lower(): b for b in buildings}

    imports = (
        PreschoolChildrenImport.query
        .filter(PreschoolChildrenImport.academic_year_id == year.id)
        .order_by(PreschoolChildrenImport.created_at.desc(), PreschoolChildrenImport.id.desc())
        .all()
    )

    result = None

    if request.method == "POST":
        uploaded = request.files.get("file")

        if not uploaded or not uploaded.filename:
            flash("Выберите Excel-файл с контингентом ДОУ.", "warning")
            return redirect(url_for("preschool.import_children", academic_year_id=year.id))

        if not uploaded.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            flash("Поддерживаются файлы .xlsx, .xlsm или выгрузки Excel с расширением .xls.", "warning")
            return redirect(url_for("preschool.import_children", academic_year_id=year.id))

        try:
            from openpyxl import load_workbook
        except Exception:
            flash("Не установлен openpyxl. Установите зависимости из requirements.txt.", "danger")
            return redirect(url_for("preschool.import_children", academic_year_id=year.id))

        workbook = load_workbook(uploaded, data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            flash("Файл пустой.", "warning")
            return redirect(url_for("preschool.import_children", academic_year_id=year.id))

        header_row_index = None
        header_map = {}

        aliases = {
            "case_number": {"личноедело№", "личноедело", "номерличногодела"},
            "fio": {"фио", "фиоребенка", "фиовоспитанника", "фамилияимяотчество", "воспитанник", "ребенок"},
            "group": {"группа", "группадоу", "наименованиегруппы"},
            "birth_date": {"датарождения", "др", "рождение", "родился", "родилась"},
            "reg_address": {"адресрегистрации", "регистрация", "адреспрописки", "прописка", "регистрацияпоместужительства", "адресрегистрациипоместужительства"},
            "living_address": {"адреспроживания", "проживание", "адресместажительства", "местожительства", "регистрацияпоместупребывания", "адресрегистрациипоместупребывания"},
            "actual_address": {"фактическийадрес", "адресфактический", "фактическоепроживание", "адресфактическогопроживания", "фактическоеместопроживания"},
            "note": {"примечание", "комментарий", "комментарии", "дополнительныесведениякод", "дополнительныесведения"},
        }

        for idx, row in enumerate(rows[:15]):
            normalized = [_normalize_header(cell) for cell in row]
            found = {}
            for col_index, name in enumerate(normalized):
                for key, variants in aliases.items():
                    if name in variants:
                        found[key] = col_index

            if "fio" in found and "group" in found:
                header_row_index = idx
                header_map = found
                break

        if header_row_index is None:
            flash("Не удалось найти заголовки. Нужны минимум колонки: ФИО и Группа.", "danger")
            return redirect(url_for("preschool.import_children", academic_year_id=year.id))

        added = 0
        skipped = 0
        created_groups = 0
        errors = []

        import_batch = PreschoolChildrenImport(
            academic_year_id=year.id,
            filename=uploaded.filename,
            added_count=0,
            skipped_count=0,
            created_groups_count=0,
        )
        db.session.add(import_batch)
        db.session.flush()

        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            case_number = row[header_map["case_number"]] if header_map.get("case_number") is not None and header_map["case_number"] < len(row) else None
            fio = row[header_map["fio"]] if header_map.get("fio") is not None and header_map["fio"] < len(row) else None
            group_name = row[header_map["group"]] if header_map.get("group") is not None and header_map["group"] < len(row) else None
            birth_raw = row[header_map["birth_date"]] if header_map.get("birth_date") is not None and header_map["birth_date"] < len(row) else None
            reg_address = row[header_map["reg_address"]] if header_map.get("reg_address") is not None and header_map["reg_address"] < len(row) else None
            living_address = row[header_map["living_address"]] if header_map.get("living_address") is not None and header_map["living_address"] < len(row) else None
            actual_address = row[header_map["actual_address"]] if header_map.get("actual_address") is not None and header_map["actual_address"] < len(row) else None
            note = row[header_map["note"]] if header_map.get("note") is not None and header_map["note"] < len(row) else None

            if not fio and not group_name:
                continue

            last_name, first_name, middle_name = _split_full_name(fio)
            if not last_name or not first_name:
                skipped += 1
                errors.append(f"Строка {row_number}: не удалось разобрать ФИО «{fio}»")
                continue

            group_key = str(group_name or "").strip().lower()
            group = groups_by_name.get(group_key)

            if not group:
                clean_group_name = str(group_name or "").strip()

                # Пытаемся определить здание по первой части названия группы:
                # например, ДК2-7 -> ДК2.
                building = None
                building_key = None

                if "-" in clean_group_name:
                    building_key = clean_group_name.split("-", 1)[0].strip().lower()
                    building = buildings_by_name.get(building_key)

                group = PreschoolGroup(
                    academic_year_id=year.id,
                    building_id=building.id if building else None,
                    name=clean_group_name,
                    age_level=None,
                    teacher_user_id=None,
                    teacher_name=None,
                    is_active=True,
                    is_archived=False,
                )
                db.session.add(group)
                db.session.flush()

                groups_by_name[group_key] = group
                created_groups += 1

            parsed_birth_date = _parse_excel_date(birth_raw)

            duplicate_query = (
                PreschoolChild.query
                .filter(PreschoolChild.group_id == group.id)
                .filter(func.lower(PreschoolChild.last_name) == last_name.lower())
                .filter(func.lower(PreschoolChild.first_name) == first_name.lower())
                .filter(func.lower(PreschoolChild.middle_name) == (middle_name or "").lower())
            )

            if parsed_birth_date:
                duplicate_query = duplicate_query.filter(PreschoolChild.birth_date == parsed_birth_date)

            duplicate = duplicate_query.first()

            if duplicate:
                skipped += 1
                continue

            child = PreschoolChild(
                group_id=group.id,
                import_batch_id=import_batch.id,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                birth_date=parsed_birth_date,
                personal_account=str(case_number).strip() if case_number else None,
                reg_address=str(reg_address).strip() if reg_address else None,
                living_address=str(living_address).strip() if living_address else None,
                actual_address=str(actual_address).strip() if actual_address else None,
                status="active",
                note=str(note).strip() if note else None,
            )

            db.session.add(child)
            added += 1

        import_batch.added_count = added
        import_batch.skipped_count = skipped
        import_batch.created_groups_count = created_groups

        db.session.commit()

        imports = (
            PreschoolChildrenImport.query
            .filter(PreschoolChildrenImport.academic_year_id == year.id)
            .order_by(PreschoolChildrenImport.created_at.desc(), PreschoolChildrenImport.id.desc())
            .all()
        )

        result = {
            "added": added,
            "skipped": skipped,
            "created_groups": created_groups,
            "errors": errors[:30],
            "errors_total": len(errors),
        }

        flash(f"Импорт завершён: добавлено {added}, создано групп {created_groups}, пропущено {skipped}.", "success")

    return render_template(
        "preschool/import_children.html",
        year=year,
        all_years=all_years,
        groups=groups,
        imports=imports,
        result=result,
    )


@bp.route("/children/imports/<int:import_id>/delete", methods=["POST"])
def delete_children_import(import_id):
    import_batch = PreschoolChildrenImport.query.get_or_404(import_id)
    year_id = import_batch.academic_year_id

    deleted_count = PreschoolChild.query.filter(
        PreschoolChild.import_batch_id == import_batch.id
    ).delete(synchronize_session=False)

    db.session.delete(import_batch)
    db.session.commit()

    flash(f"Импорт удалён. Удалено воспитанников: {deleted_count}.", "success")
    return redirect(url_for("preschool.import_children", academic_year_id=year_id))


@bp.route("/children/clear-year", methods=["POST"])
def clear_children_year():
    year_id = request.form.get("academic_year_id", type=int)
    year = AcademicYear.query.get_or_404(year_id)

    child_ids = [
        child_id for (child_id,) in (
            db.session.query(PreschoolChild.id)
            .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
            .filter(PreschoolGroup.academic_year_id == year.id)
            .all()
        )
    ]

    deleted_count = 0
    if child_ids:
        deleted_count = PreschoolChild.query.filter(
            PreschoolChild.id.in_(child_ids)
        ).delete(synchronize_session=False)

    PreschoolChildrenImport.query.filter(
        PreschoolChildrenImport.academic_year_id == year.id
    ).delete(synchronize_session=False)

    db.session.commit()

    flash(f"Контингент ДОУ за {year.name} очищен. Удалено воспитанников: {deleted_count}.", "success")
    return redirect(url_for("preschool.import_children", academic_year_id=year.id))


def _normalize_phone(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    return raw


@bp.route("/representatives/import", methods=["GET", "POST"])
def import_representatives():
    year_id = request.args.get("academic_year_id", type=int) or request.form.get("academic_year_id", type=int)
    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    result = None

    if request.method == "POST":
        uploaded = request.files.get("file")

        if not uploaded or not uploaded.filename:
            flash("Выберите Excel-файл с представителями.", "warning")
            return redirect(url_for("preschool.import_representatives", academic_year_id=year.id))

        if not uploaded.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            flash("Поддерживаются файлы .xlsx, .xlsm или выгрузки Excel с расширением .xls.", "warning")
            return redirect(url_for("preschool.import_representatives", academic_year_id=year.id))

        try:
            from openpyxl import load_workbook
        except Exception:
            flash("Не установлен openpyxl. Установите зависимости из requirements.txt.", "danger")
            return redirect(url_for("preschool.import_representatives", academic_year_id=year.id))

        workbook = load_workbook(uploaded, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            flash("Файл пустой.", "warning")
            return redirect(url_for("preschool.import_representatives", academic_year_id=year.id))

        aliases = {
            "case_number": {"личноедело№", "личноедело", "номерличногодела", "лд"},
            "child_fio": {"фиоребенка", "ребенок", "воспитанник", "фиовоспитанника"},
            "relation": {"родство", "степеньродства", "типпредставителя", "представитель", "законныйпредставитель"},
            "full_name": {"фио", "фиопредставителя", "фиородителя", "родитель", "представительфио"},
            "phone": {"телефон", "мобильныйтелефон", "контактныйтелефон", "тел"},
            "email": {"email", "почта", "электроннаяпочта", "e-mail"},
            "address": {"адрес", "адресрегистрации", "адреспроживания", "фактическийадрес"},
            "note": {"примечание", "комментарий", "комментарии"},
        }

        header_row_index = None
        header_map = {}

        for idx, row in enumerate(rows[:20]):
            normalized = [_normalize_header(cell) for cell in row]
            found = {}
            for col_index, name in enumerate(normalized):
                for key, variants in aliases.items():
                    if name in variants:
                        found[key] = col_index

            if ("case_number" in found or "child_fio" in found) and "full_name" in found:
                header_row_index = idx
                header_map = found
                break

        if header_row_index is None:
            flash("Не удалось найти заголовки. Нужны минимум: Личное дело № или ФИО ребёнка, и ФИО представителя.", "danger")
            return redirect(url_for("preschool.import_representatives", academic_year_id=year.id))

        added = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            case_number = row[header_map["case_number"]] if header_map.get("case_number") is not None and header_map["case_number"] < len(row) else None
            child_fio = row[header_map["child_fio"]] if header_map.get("child_fio") is not None and header_map["child_fio"] < len(row) else None
            relation = row[header_map["relation"]] if header_map.get("relation") is not None and header_map["relation"] < len(row) else None
            full_name = row[header_map["full_name"]] if header_map.get("full_name") is not None and header_map["full_name"] < len(row) else None
            phone = row[header_map["phone"]] if header_map.get("phone") is not None and header_map["phone"] < len(row) else None
            email = row[header_map["email"]] if header_map.get("email") is not None and header_map["email"] < len(row) else None
            address = row[header_map["address"]] if header_map.get("address") is not None and header_map["address"] < len(row) else None
            note = row[header_map["note"]] if header_map.get("note") is not None and header_map["note"] < len(row) else None

            if not full_name:
                continue

            child = None

            if case_number:
                child = (
                    PreschoolChild.query
                    .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
                    .filter(PreschoolGroup.academic_year_id == year.id)
                    .filter(PreschoolChild.personal_account == str(case_number).strip())
                    .first()
                )

            if not child and child_fio:
                last_name, first_name, middle_name = _split_full_name(child_fio)
                if last_name and first_name:
                    query = (
                        PreschoolChild.query
                        .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
                        .filter(PreschoolGroup.academic_year_id == year.id)
                        .filter(func.lower(PreschoolChild.last_name) == last_name.lower())
                        .filter(func.lower(PreschoolChild.first_name) == first_name.lower())
                    )
                    if middle_name:
                        query = query.filter(func.lower(PreschoolChild.middle_name) == middle_name.lower())
                    child = query.first()

            if not child:
                skipped += 1
                errors.append(f"Строка {row_number}: ребёнок не найден")
                continue

            duplicate = (
                PreschoolRepresentative.query
                .filter(PreschoolRepresentative.child_id == child.id)
                .filter(func.lower(PreschoolRepresentative.full_name) == str(full_name).strip().lower())
                .first()
            )

            if duplicate:
                skipped += 1
                continue

            representative = PreschoolRepresentative(
                child_id=child.id,
                relation=str(relation).strip() if relation else None,
                full_name=str(full_name).strip(),
                phone=_normalize_phone(phone),
                email=str(email).strip() if email else None,
                address=str(address).strip() if address else None,
                note=str(note).strip() if note else None,
            )

            db.session.add(representative)
            added += 1

        db.session.commit()

        result = {
            "added": added,
            "skipped": skipped,
            "errors": errors[:30],
            "errors_total": len(errors),
        }

        flash(f"Импорт представителей завершён: добавлено {added}, пропущено {skipped}.", "success")

    return render_template(
        "preschool/import_representatives.html",
        year=year,
        all_years=all_years,
        result=result,
    )


@bp.route("/representatives/<int:representative_id>/delete", methods=["POST"])
def delete_representative(representative_id):
    representative = PreschoolRepresentative.query.get_or_404(representative_id)
    child_id = representative.child_id

    db.session.delete(representative)
    db.session.commit()

    flash("Представитель удалён.", "success")
    return redirect(url_for("preschool.child_card", child_id=child_id))


def _decode_zip_name(name):
    try:
        return name.encode("cp437").decode("utf-8")
    except Exception:
        return name


def _extract_preschool_group_name(filename):
    import re

    clean = _decode_zip_name(filename)
    base = Path(clean).name
    match = re.search(r"(ДК\s*\d+\s*[-–]\s*\d+)", base, re.IGNORECASE)
    if not match:
        return None

    group_name = match.group(1)
    group_name = group_name.replace(" ", "")
    group_name = group_name.replace("–", "-")
    return group_name.upper()


def _int_from_cell(value):
    if value is None:
        return 0

    raw = str(value).strip()
    if not raw:
        return 0

    raw = raw.replace(",", ".")
    try:
        return int(float(raw))
    except Exception:
        return 0


def _split_attendance_name(full_name):
    parts = [p for p in str(full_name or "").strip().split() if p]
    if len(parts) < 2:
        return None, None, None

    last_name = parts[0]
    first_name = parts[1]
    middle_name = " ".join(parts[2:]) if len(parts) > 2 else None
    return last_name, first_name, middle_name


def _is_attendance_service_row(value):
    """Пропускаем даты, итоги и технические строки внутри табеля."""
    if value is None:
        return True

    raw = str(value).strip()
    if not raw:
        return True

    lowered = raw.lower().replace("ё", "е")

    service_words = (
        "итого",
        "всего",
        "дата",
        "дни",
        "месяц",
        "табель",
        "группа",
        "воспитатель",
    )

    if any(lowered.startswith(word) for word in service_words):
        return True

    # Примеры из табеля: "31 мая 2026 г.", "1 май 2026 г."
    month_words = (
        "январ", "феврал", "март", "апрел", "мая", "май",
        "июн", "июл", "август", "сентябр", "октябр", "ноябр", "декабр",
    )

    if any(month in lowered for month in month_words) and any(ch.isdigit() for ch in lowered):
        return True

    # Даты вида 31.05.2026 или 2026-05-31
    import re
    if re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", lowered):
        return True

    if re.fullmatch(r"\d{4}[./-]\d{1,2}[./-]\d{1,2}", lowered):
        return True

    return False


def _find_or_create_preschool_group(year, group_name):
    if not group_name:
        return None

    group = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .filter(func.lower(PreschoolGroup.name) == group_name.lower())
        .first()
    )

    if group:
        return group

    building = None
    if "-" in group_name:
        building_key = group_name.split("-", 1)[0].strip().lower()
        building = (
            Building.query
            .filter(func.lower(Building.name) == building_key)
            .first()
        )

    group = PreschoolGroup(
        academic_year_id=year.id,
        building_id=building.id if building else None,
        name=group_name,
        is_active=True,
        is_archived=False,
    )
    db.session.add(group)
    db.session.flush()

    return group


def _find_preschool_child(year, group, child_name, account_number):
    if account_number:
        child = (
            PreschoolChild.query
            .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
            .filter(PreschoolGroup.academic_year_id == year.id)
            .filter(PreschoolChild.personal_account == str(account_number).strip())
            .first()
        )
        if child:
            return child

    last_name, first_name, middle_name = _split_attendance_name(child_name)
    if not last_name or not first_name:
        return None

    query = (
        PreschoolChild.query
        .join(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
        .filter(PreschoolGroup.academic_year_id == year.id)
        .filter(func.lower(PreschoolChild.last_name) == last_name.lower())
        .filter(func.lower(PreschoolChild.first_name) == first_name.lower())
    )

    if middle_name:
        query = query.filter(func.lower(PreschoolChild.middle_name) == middle_name.lower())

    if group:
        query = query.filter(PreschoolChild.group_id == group.id)

    return query.first()


def _parse_attendance_workbook(year, upload, filename, file_bytes):
    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    group_name = _extract_preschool_group_name(filename)
    group = _find_or_create_preschool_group(year, group_name)

    created = 0
    errors = []

    # В табелях из архива реальные колонки такие:
    # C = ФИО, AB = номер счета, DI = пропущено всего,
    # DN = засчитываемые пропуски, DT = дни к оплате, EC = причины.
    for row in range(14, sheet.max_row + 1):
        child_name = sheet.cell(row=row, column=3).value

        if not child_name:
            continue

        child_name = str(child_name).strip()

        # Пропускаем даты, итоги и технические строки.
        if _is_attendance_service_row(child_name):
            continue

        # ФИО ребёнка должно состоять минимум из фамилии и имени.
        last_name_check, first_name_check, _ = _split_attendance_name(child_name)
        if not last_name_check or not first_name_check:
            continue

        account_number = sheet.cell(row=row, column=28).value
        missed_total = _int_from_cell(sheet.cell(row=row, column=113).value)
        credited_days = _int_from_cell(sheet.cell(row=row, column=118).value)
        payment_days = _int_from_cell(sheet.cell(row=row, column=124).value)
        reasons = sheet.cell(row=row, column=133).value

        child_days = max(payment_days - credited_days, 0)

        child = _find_preschool_child(year, group, child_name, account_number)

        record = PreschoolAttendanceRecord(
            upload_id=upload.id,
            group_id=group.id if group else None,
            child_id=child.id if child else None,
            source_filename=_decode_zip_name(filename),
            child_name=child_name,
            account_number=str(account_number).strip() if account_number else None,
            missed_total=missed_total,
            credited_days=credited_days,
            payment_days=payment_days,
            child_days=child_days,
            absence_reasons=str(reasons).strip() if reasons else None,
        )

        db.session.add(record)
        created += 1

        if not child:
            errors.append(f"{child_name}: не найден в контингенте")

    return {
        "created": created,
        "group_name": group.name if group else group_name,
        "errors": errors,
    }


def _process_attendance_zip(year, upload):
    import zipfile

    processed_files = 0
    created_records = 0
    errors = []

    with zipfile.ZipFile(upload.stored_filename, "r") as archive:
        for member in archive.namelist():
            decoded_name = _decode_zip_name(member)

            if member.startswith("__MACOSX/") or decoded_name.startswith("__MACOSX/"):
                continue

            if not decoded_name.lower().endswith((".xlsx", ".xlsm")):
                continue

            file_bytes = archive.read(member)
            processed_files += 1

            try:
                result = _parse_attendance_workbook(year, upload, decoded_name, file_bytes)
                created_records += result["created"]
                errors.extend(result["errors"][:20])
            except Exception as exc:
                errors.append(f"{decoded_name}: {exc}")

    upload.status = "processed"
    upload.comment = (
        f"Обработано файлов: {processed_files}. "
        f"Создано строк табеля: {created_records}. "
        f"Ошибок сопоставления: {len(errors)}."
    )

    db.session.commit()

    return {
        "processed_files": processed_files,
        "created_records": created_records,
        "errors": errors,
    }


MONTH_LABELS_RU = {
    "01": "Январь",
    "02": "Февраль",
    "03": "Март",
    "04": "Апрель",
    "05": "Май",
    "06": "Июнь",
    "07": "Июль",
    "08": "Август",
    "09": "Сентябрь",
    "10": "Октябрь",
    "11": "Ноябрь",
    "12": "Декабрь",
}


def _academic_year_month_options(year):
    """Список месяцев для выбранного учебного года: сентябрь–август."""
    if not year or not year.name or "/" not in year.name:
        return []

    try:
        start_year = int(str(year.name).split("/", 1)[0])
    except Exception:
        return []

    months = []
    for month_num in range(9, 13):
        value = f"{start_year}-{month_num:02d}"
        months.append({"value": value, "label": f"{MONTH_LABELS_RU[f'{month_num:02d}']} {start_year}"})

    next_year = start_year + 1
    for month_num in range(1, 9):
        value = f"{next_year}-{month_num:02d}"
        months.append({"value": value, "label": f"{MONTH_LABELS_RU[f'{month_num:02d}']} {next_year}"})

    return months


def _month_label(value):
    if not value or "-" not in str(value):
        return "—"

    year_part, month_part = str(value).split("-", 1)
    month_part = month_part[:2]

    return f"{MONTH_LABELS_RU.get(month_part, month_part)} {year_part}"


def _detect_month_from_zip(stored_filename, year):
    """Пытаемся определить месяц по названиям файлов внутри архива: ДК1-1 май.xlsx."""
    import zipfile

    if not stored_filename or not year:
        return None

    month_words = {
        "январ": "01",
        "феврал": "02",
        "март": "03",
        "апрел": "04",
        "май": "05",
        "мая": "05",
        "июн": "06",
        "июл": "07",
        "август": "08",
        "сентябр": "09",
        "октябр": "10",
        "ноябр": "11",
        "декабр": "12",
    }

    try:
        start_year = int(str(year.name).split("/", 1)[0])
    except Exception:
        start_year = None

    try:
        with zipfile.ZipFile(stored_filename, "r") as archive:
            names = [_decode_zip_name(name).lower() for name in archive.namelist()]
    except Exception:
        return None

    for name in names:
        for word, month_num in month_words.items():
            if word in name:
                if not start_year:
                    return None

                # Учебный год: сентябрь-декабрь относятся к первому году,
                # январь-август — ко второму.
                year_num = start_year if int(month_num) >= 9 else start_year + 1
                return f"{year_num}-{month_num}"

    return None


@bp.route("/attendance", methods=["GET", "POST"])
def attendance():
    year_id = request.args.get("academic_year_id", type=int) or request.form.get("academic_year_id", type=int)
    month = (request.args.get("month") or request.form.get("month") or "").strip()

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )
    month_options = _academic_year_month_options(year)

    if request.method == "POST":
        uploaded = request.files.get("file")

        if not uploaded or not uploaded.filename:
            flash("Выберите ZIP-архив с табелями ДОУ.", "warning")
            return redirect(url_for("preschool.attendance", academic_year_id=year.id, month=month))

        if not uploaded.filename.lower().endswith(".zip"):
            flash("Для загрузки табелей нужен ZIP-архив.", "warning")
            return redirect(url_for("preschool.attendance", academic_year_id=year.id, month=month))

        from flask import current_app
        import uuid

        upload_root = Path(current_app.config.get("UPLOAD_FOLDER", "uploads"))
        folder = upload_root / "preschool" / "attendance"
        folder.mkdir(parents=True, exist_ok=True)

        safe_name = secure_filename(uploaded.filename)
        stored_name = f"{uuid.uuid4().hex}_{safe_name}"
        stored_path = folder / stored_name
        uploaded.save(stored_path)

        detected_month = month or _detect_month_from_zip(str(stored_path), year)

        item = PreschoolAttendanceUpload(
            academic_year_id=year.id,
            month=detected_month or None,
            original_filename=uploaded.filename,
            stored_filename=str(stored_path),
            status="uploaded",
            comment="Архив загружен.",
        )

        db.session.add(item)
        db.session.commit()

        try:
            result = _process_attendance_zip(year, item)
            flash(
                f"ZIP загружен и обработан: файлов {result['processed_files']}, строк {result['created_records']}.",
                "success",
            )
            month = item.month or month
        except Exception as exc:
            item.status = "error"
            item.comment = f"Ошибка обработки архива: {exc}"
            db.session.commit()
            flash(f"ZIP загружен, но обработка завершилась ошибкой: {exc}", "danger")

        return redirect(url_for("preschool.attendance", academic_year_id=year.id, month=month))

    query = PreschoolAttendanceUpload.query.filter(PreschoolAttendanceUpload.academic_year_id == year.id)

    if month:
        query = query.filter(PreschoolAttendanceUpload.month == month)

    uploads = query.order_by(
        PreschoolAttendanceUpload.created_at.desc(),
        PreschoolAttendanceUpload.id.desc()
    ).all()

    upload_stats = {}
    for item in uploads:
        records = PreschoolAttendanceRecord.query.filter(
            PreschoolAttendanceRecord.upload_id == item.id
        )
        upload_stats[item.id] = {
            "records": records.count(),
            "child_days": db.session.query(func.coalesce(func.sum(PreschoolAttendanceRecord.child_days), 0))
                .filter(PreschoolAttendanceRecord.upload_id == item.id)
                .scalar() or 0,
            "payment_days": db.session.query(func.coalesce(func.sum(PreschoolAttendanceRecord.payment_days), 0))
                .filter(PreschoolAttendanceRecord.upload_id == item.id)
                .scalar() or 0,
            "credited_days": db.session.query(func.coalesce(func.sum(PreschoolAttendanceRecord.credited_days), 0))
                .filter(PreschoolAttendanceRecord.upload_id == item.id)
                .scalar() or 0,
        }

    return render_template(
        "preschool/attendance.html",
        year=year,
        all_years=all_years,
        month=month,
        uploads=uploads,
        upload_stats=upload_stats,
        month_options=month_options,
        month_label=_month_label,
    )


@bp.route("/attendance/analytics")
def attendance_analytics():
    year_id = request.args.get("academic_year_id", type=int)
    month = (request.args.get("month") or "").strip()
    upload_id = request.args.get("upload_id", type=int)

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )
    month_options = _academic_year_month_options(year)

    uploads_query = PreschoolAttendanceUpload.query.filter(
        PreschoolAttendanceUpload.academic_year_id == year.id
    )

    if month:
        uploads_query = uploads_query.filter(PreschoolAttendanceUpload.month == month)

    uploads = uploads_query.order_by(
        PreschoolAttendanceUpload.created_at.desc(),
        PreschoolAttendanceUpload.id.desc()
    ).all()

    records_query = (
        PreschoolAttendanceRecord.query
        .join(PreschoolAttendanceUpload, PreschoolAttendanceRecord.upload_id == PreschoolAttendanceUpload.id)
        .outerjoin(PreschoolGroup, PreschoolAttendanceRecord.group_id == PreschoolGroup.id)
        .outerjoin(Building, PreschoolGroup.building_id == Building.id)
        .filter(PreschoolAttendanceUpload.academic_year_id == year.id)
    )

    if month:
        records_query = records_query.filter(PreschoolAttendanceUpload.month == month)

    if upload_id:
        records_query = records_query.filter(PreschoolAttendanceRecord.upload_id == upload_id)

    records = records_query.all()

    total = {
        "records": len(records),
        "matched": sum(1 for item in records if item.child_id),
        "unmatched": sum(1 for item in records if not item.child_id),
        "payment_days": sum(item.payment_days or 0 for item in records),
        "credited_days": sum(item.credited_days or 0 for item in records),
        "child_days": sum(item.child_days or 0 for item in records),
    }

    building_map = {}
    group_map = {}

    for record in records:
        group = record.group
        building_name = group.building.name if group and group.building else "Без корпуса"
        group_name = group.name if group else "Без группы"

        if building_name not in building_map:
            building_map[building_name] = {
                "name": building_name,
                "records": 0,
                "matched": 0,
                "unmatched": 0,
                "payment_days": 0,
                "credited_days": 0,
                "child_days": 0,
                "groups": set(),
            }

        building_row = building_map[building_name]
        building_row["records"] += 1
        building_row["payment_days"] += record.payment_days or 0
        building_row["credited_days"] += record.credited_days or 0
        building_row["child_days"] += record.child_days or 0
        building_row["groups"].add(group_name)

        if record.child_id:
            building_row["matched"] += 1
        else:
            building_row["unmatched"] += 1

        if group_name not in group_map:
            group_map[group_name] = {
                "name": group_name,
                "building": building_name,
                "records": 0,
                "matched": 0,
                "unmatched": 0,
                "payment_days": 0,
                "credited_days": 0,
                "child_days": 0,
            }

        group_row = group_map[group_name]
        group_row["records"] += 1
        group_row["payment_days"] += record.payment_days or 0
        group_row["credited_days"] += record.credited_days or 0
        group_row["child_days"] += record.child_days or 0

        if record.child_id:
            group_row["matched"] += 1
        else:
            group_row["unmatched"] += 1

    building_rows = []
    for row in building_map.values():
        row = dict(row)
        row["groups_count"] = len(row["groups"])
        row.pop("groups", None)
        building_rows.append(row)

    building_rows = sorted(building_rows, key=lambda x: x["name"])
    group_rows = sorted(group_map.values(), key=lambda x: (x["building"], x["name"]))

    # Динамика по месяцам строится по всем загрузкам выбранного учебного года.
    month_records_query = (
        PreschoolAttendanceRecord.query
        .join(PreschoolAttendanceUpload, PreschoolAttendanceRecord.upload_id == PreschoolAttendanceUpload.id)
        .filter(PreschoolAttendanceUpload.academic_year_id == year.id)
    )

    if upload_id:
        month_records_query = month_records_query.filter(PreschoolAttendanceRecord.upload_id == upload_id)

    month_records = month_records_query.all()

    month_map = {}
    for record in month_records:
        upload = record.upload
        month_key = upload.month or "Не указан"

        if month_key not in month_map:
            month_map[month_key] = {
                "month": month_key,
                "month_label": _month_label(month_key) if month_key != "Не указан" else "Не указан",
                "records": 0,
                "payment_days": 0,
                "credited_days": 0,
                "child_days": 0,
            }

        row = month_map[month_key]
        row["records"] += 1
        row["payment_days"] += record.payment_days or 0
        row["credited_days"] += record.credited_days or 0
        row["child_days"] += record.child_days or 0

    def _month_sort_key(row):
        value = row["month"]
        if value == "Не указан":
            return "9999-99"
        return value

    month_rows = sorted(month_map.values(), key=_month_sort_key)

    max_building_child_days = max([row["child_days"] for row in building_rows] or [1])
    max_group_child_days = max([row["child_days"] for row in group_rows] or [1])
    max_month_child_days = max([row["child_days"] for row in month_rows] or [1])

    return render_template(
        "preschool/attendance_analytics.html",
        year=year,
        all_years=all_years,
        month=month,
        upload_id=upload_id,
        uploads=uploads,
        month_options=month_options,
        total=total,
        building_rows=building_rows,
        group_rows=group_rows,
        month_rows=month_rows,
        max_building_child_days=max_building_child_days,
        max_group_child_days=max_group_child_days,
        max_month_child_days=max_month_child_days,
        month_label=_month_label,
    )


@bp.route("/attendance/uploads/<int:upload_id>")
def attendance_upload_detail(upload_id):
    item = PreschoolAttendanceUpload.query.get_or_404(upload_id)

    records_query = (
        PreschoolAttendanceRecord.query
        .filter(PreschoolAttendanceRecord.upload_id == item.id)
        .outerjoin(PreschoolGroup, PreschoolAttendanceRecord.group_id == PreschoolGroup.id)
        .order_by(
            PreschoolGroup.name.asc(),
            PreschoolAttendanceRecord.child_name.asc(),
        )
    )

    records = records_query.all()

    group_rows = []
    group_map = {}

    for record in records:
        key = record.group.name if record.group else "Без группы"

        if key not in group_map:
            group_map[key] = {
                "group_name": key,
                "records": 0,
                "matched": 0,
                "unmatched": 0,
                "payment_days": 0,
                "credited_days": 0,
                "child_days": 0,
            }

        row = group_map[key]
        row["records"] += 1
        row["payment_days"] += record.payment_days or 0
        row["credited_days"] += record.credited_days or 0
        row["child_days"] += record.child_days or 0

        if record.child_id:
            row["matched"] += 1
        else:
            row["unmatched"] += 1

    group_rows = sorted(group_map.values(), key=lambda x: x["group_name"])

    total = {
        "records": len(records),
        "matched": sum(1 for record in records if record.child_id),
        "unmatched": sum(1 for record in records if not record.child_id),
        "payment_days": sum(record.payment_days or 0 for record in records),
        "credited_days": sum(record.credited_days or 0 for record in records),
        "child_days": sum(record.child_days or 0 for record in records),
    }

    unmatched_records = [record for record in records if not record.child_id]

    return render_template(
        "preschool/attendance_upload_detail.html",
        item=item,
        records=records,
        group_rows=group_rows,
        total=total,
        unmatched_records=unmatched_records,
    )


@bp.route("/attendance/uploads/<int:upload_id>/process", methods=["POST"])
def process_attendance_upload(upload_id):
    item = PreschoolAttendanceUpload.query.get_or_404(upload_id)
    year = item.academic_year

    if not year:
        flash("У загрузки не указан учебный год.", "warning")
        return redirect(url_for("preschool.attendance"))

    # Удаляем старые строки этой загрузки, если уже пробовали обрабатывать.
    PreschoolAttendanceRecord.query.filter(
        PreschoolAttendanceRecord.upload_id == item.id
    ).delete(synchronize_session=False)
    db.session.commit()

    try:
        result = _process_attendance_zip(year, item)
        flash(
            f"Архив обработан: файлов {result['processed_files']}, строк {result['created_records']}.",
            "success",
        )
    except Exception as exc:
        item.status = "error"
        item.comment = f"Ошибка обработки архива: {exc}"
        db.session.commit()
        flash(f"Ошибка обработки архива: {exc}", "danger")

    return redirect(url_for(
        "preschool.attendance",
        academic_year_id=item.academic_year_id,
        month=item.month,
    ))


@bp.route("/attendance/uploads/<int:upload_id>/delete", methods=["POST"])
def delete_attendance_upload(upload_id):
    item = PreschoolAttendanceUpload.query.get_or_404(upload_id)
    year_id = item.academic_year_id
    month = item.month

    try:
        if item.stored_filename:
            path = Path(item.stored_filename)
            if path.exists():
                path.unlink()
    except Exception:
        pass

    PreschoolAttendanceRecord.query.filter(
        PreschoolAttendanceRecord.upload_id == item.id
    ).delete(synchronize_session=False)

    db.session.delete(item)
    db.session.commit()

    flash("Загрузка табелей удалена.", "success")
    return redirect(url_for("preschool.attendance", academic_year_id=year_id, month=month))


@bp.route("/movements")
def movements():
    year_id = request.args.get("academic_year_id", type=int)
    movement_type = (request.args.get("movement_type") or "").strip()
    q = (request.args.get("q") or "").strip()

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    query = (
        PreschoolChildMovement.query
        .join(PreschoolChild, PreschoolChildMovement.child_id == PreschoolChild.id)
        .outerjoin(PreschoolGroup, PreschoolChild.group_id == PreschoolGroup.id)
        .filter(PreschoolGroup.academic_year_id == year.id)
    )

    if movement_type:
        query = query.filter(PreschoolChildMovement.movement_type == movement_type)

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                PreschoolChild.last_name.ilike(like),
                PreschoolChild.first_name.ilike(like),
                PreschoolChild.middle_name.ilike(like),
            )
        )

    items = (
        query
        .order_by(
            PreschoolChildMovement.movement_date.desc().nullslast(),
            PreschoolChildMovement.created_at.desc(),
        )
        .all()
    )

    movement_counts = {
        "total": len(items),
        "admission": sum(1 for item in items if item.movement_type == "admission"),
        "transfer_group": sum(1 for item in items if item.movement_type == "transfer_group"),
        "leaving": sum(1 for item in items if item.movement_type == "leaving"),
        "transfer_school": sum(1 for item in items if item.movement_type == "transfer_school"),
        "year_transfer": sum(1 for item in items if item.movement_type == "year_transfer"),
    }

    return render_template(
        "preschool/movements.html",
        year=year,
        all_years=all_years,
        items=items,
        movement_type=movement_type,
        q=q,
        movement_counts=movement_counts,
    )


@bp.route("/movements/bulk", methods=["GET", "POST"])
def bulk_movements():
    year_id = request.args.get("academic_year_id", type=int) or request.form.get("academic_year_id", type=int)
    target_year_id = request.args.get("target_academic_year_id", type=int) or request.form.get("target_academic_year_id", type=int)
    source_group_id = request.args.get("source_group_id", type=int) or request.form.get("source_group_id", type=int)

    year = AcademicYear.query.get(year_id) if year_id else _get_current_year()
    target_year = AcademicYear.query.get(target_year_id) if target_year_id else None

    if not year:
        flash("Сначала создайте учебный год в служебном разделе.", "warning")
        return redirect(url_for("children.academic_years_registry"))

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    groups = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .order_by(PreschoolGroup.name.asc())
        .all()
    )

    target_groups = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == target_year.id)
        .order_by(PreschoolGroup.name.asc())
        .all()
        if target_year else []
    )

    children = []
    if source_group_id:
        children = (
            PreschoolChild.query
            .filter(PreschoolChild.group_id == source_group_id)
            .order_by(PreschoolChild.last_name.asc(), PreschoolChild.first_name.asc())
            .all()
        )

    if request.method == "POST":
        movement_type = (request.form.get("movement_type") or "").strip()
        movement_date_raw = (request.form.get("movement_date") or "").strip()
        target_group_id = request.form.get("target_group_id", type=int)
        basis = (request.form.get("basis") or "").strip()
        comment = (request.form.get("comment") or "").strip()

        child_ids = request.form.getlist("child_ids")

        if movement_type not in ("transfer_group", "leaving", "year_transfer"):
            flash("Выберите тип массового движения.", "warning")
            return redirect(url_for(
                "preschool.bulk_movements",
                academic_year_id=year.id,
                target_academic_year_id=target_year.id if target_year else None,
                source_group_id=source_group_id,
            ))

        if not child_ids:
            flash("Выберите хотя бы одного воспитанника.", "warning")
            return redirect(url_for(
                "preschool.bulk_movements",
                academic_year_id=year.id,
                target_academic_year_id=target_year.id if target_year else None,
                source_group_id=source_group_id,
            ))

        if movement_type in ("transfer_group", "year_transfer") and not target_group_id:
            flash("Для перевода выберите группу назначения.", "warning")
            return redirect(url_for(
                "preschool.bulk_movements",
                academic_year_id=year.id,
                target_academic_year_id=target_year.id if target_year else None,
                source_group_id=source_group_id,
            ))

        target_group = PreschoolGroup.query.get(target_group_id) if target_group_id else None

        if movement_type == "year_transfer":
            if not target_year:
                flash("Для перевода в новый учебный год выберите учебный год назначения.", "warning")
                return redirect(url_for(
                    "preschool.bulk_movements",
                    academic_year_id=year.id,
                    source_group_id=source_group_id,
                ))

            if not target_group or target_group.academic_year_id != target_year.id:
                flash("Группа назначения должна относиться к выбранному новому учебному году.", "warning")
                return redirect(url_for(
                    "preschool.bulk_movements",
                    academic_year_id=year.id,
                    target_academic_year_id=target_year.id,
                    source_group_id=source_group_id,
                ))

        if movement_type == "transfer_group":
            if not target_group or target_group.academic_year_id != year.id:
                flash("Для перевода внутри года выберите группу текущего учебного года.", "warning")
                return redirect(url_for(
                    "preschool.bulk_movements",
                    academic_year_id=year.id,
                    source_group_id=source_group_id,
                ))

        movement_date = None
        if movement_date_raw:
            from datetime import datetime
            try:
                movement_date = datetime.strptime(movement_date_raw, "%Y-%m-%d").date()
            except ValueError:
                movement_date = None

        selected_children = PreschoolChild.query.filter(PreschoolChild.id.in_(child_ids)).all()

        created = 0

        for child in selected_children:
            from_group = child.group
            from_group_id = child.group_id
            from_year_id = from_group.academic_year_id if from_group else year.id

            if movement_type == "transfer_group":
                child.group_id = target_group_id
                child.status = "active"
                to_group_id = target_group_id
                to_year_id = year.id

            elif movement_type == "year_transfer":
                child.group_id = target_group_id
                child.status = "active"
                to_group_id = target_group_id
                to_year_id = target_year.id

            else:
                child.status = "left"
                to_group_id = None
                to_year_id = from_year_id

            movement = PreschoolChildMovement(
                child_id=child.id,
                movement_date=movement_date,
                movement_type=movement_type,
                from_academic_year_id=from_year_id,
                to_academic_year_id=to_year_id,
                from_group_id=from_group_id,
                to_group_id=to_group_id,
                basis=basis or None,
                comment=comment or None,
            )

            db.session.add(movement)
            created += 1

        db.session.commit()

        flash(f"Массовое движение сохранено. Обработано воспитанников: {created}.", "success")
        return redirect(url_for("preschool.movements", academic_year_id=to_year_id, movement_type=movement_type))

    return render_template(
        "preschool/bulk_movements.html",
        year=year,
        target_year=target_year,
        all_years=all_years,
        groups=groups,
        target_groups=target_groups,
        children=children,
        source_group_id=source_group_id,
        target_year_id=target_year_id,
    )


@bp.route("/children/<int:child_id>")
def child_card(child_id):
    child = PreschoolChild.query.get_or_404(child_id)
    group = child.group
    year = group.academic_year if group and group.academic_year else _get_current_year()

    representatives = (
        PreschoolRepresentative.query
        .filter(PreschoolRepresentative.child_id == child.id)
        .order_by(PreschoolRepresentative.relation.asc(), PreschoolRepresentative.full_name.asc())
        .all()
    )

    groups = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .order_by(PreschoolGroup.name.asc())
        .all()
        if year else []
    )

    movements = (
        PreschoolChildMovement.query
        .filter(PreschoolChildMovement.child_id == child.id)
        .order_by(PreschoolChildMovement.movement_date.desc().nullslast(), PreschoolChildMovement.created_at.desc())
        .all()
    )

    return render_template(
        "preschool/child_card.html",
        child=child,
        group=group,
        year=year,
        representatives=representatives,
        groups=groups,
        movements=movements,
    )


@bp.route("/children/<int:child_id>/movement", methods=["POST"])
def add_child_movement(child_id):
    child = PreschoolChild.query.get_or_404(child_id)

    movement_type = (request.form.get("movement_type") or "").strip()
    movement_date_raw = (request.form.get("movement_date") or "").strip()
    to_group_id = request.form.get("to_group_id", type=int)
    basis = (request.form.get("basis") or "").strip()
    comment = (request.form.get("comment") or "").strip()

    if movement_type not in ("admission", "transfer_group", "leaving", "transfer_school"):
        flash("Выберите тип движения.", "warning")
        return redirect(url_for("preschool.child_card", child_id=child.id))

    movement_date = None
    if movement_date_raw:
        from datetime import datetime
        try:
            movement_date = datetime.strptime(movement_date_raw, "%Y-%m-%d").date()
        except ValueError:
            movement_date = None

    from_group_id = child.group_id

    if movement_type == "transfer_group":
        if not to_group_id:
            flash("Для перевода выберите новую группу.", "warning")
            return redirect(url_for("preschool.child_card", child_id=child.id))
        child.group_id = to_group_id
        child.status = "active"

    elif movement_type == "leaving":
        child.status = "left"

    elif movement_type == "admission":
        child.status = "active"
        if to_group_id:
            child.group_id = to_group_id

    elif movement_type == "transfer_school":
        child.status = "transfer_school"

    movement = PreschoolChildMovement(
        child_id=child.id,
        movement_date=movement_date,
        movement_type=movement_type,
        from_academic_year_id=from_academic_year_id,
        to_academic_year_id=child.group.academic_year_id if child.group else from_academic_year_id,
        from_group_id=from_group_id,
        to_group_id=child.group_id if movement_type in ("admission", "transfer_group") else to_group_id,
        basis=basis or None,
        comment=comment or None,
    )

    db.session.add(movement)
    db.session.commit()

    flash("Движение воспитанника сохранено.", "success")
    return redirect(url_for("preschool.child_card", child_id=child.id))


@bp.route("/movements/<int:movement_id>/delete", methods=["POST"])
def delete_child_movement(movement_id):
    movement = PreschoolChildMovement.query.get_or_404(movement_id)
    child_id = movement.child_id

    db.session.delete(movement)
    db.session.commit()

    flash("Запись движения удалена.", "success")
    return redirect(url_for("preschool.child_card", child_id=child_id))


@bp.route("/children/<int:child_id>/edit", methods=["GET", "POST"])
def edit_child(child_id):
    child = PreschoolChild.query.get_or_404(child_id)

    current_group = child.group
    year = current_group.academic_year if current_group and current_group.academic_year else _get_current_year()

    all_years = (
        AcademicYear.query
        .order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc())
        .all()
    )

    groups = (
        PreschoolGroup.query
        .filter(PreschoolGroup.academic_year_id == year.id)
        .order_by(PreschoolGroup.name.asc())
        .all()
        if year else []
    )

    if request.method == "POST":
        last_name = (request.form.get("last_name") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        middle_name = (request.form.get("middle_name") or "").strip()
        birth_date_raw = (request.form.get("birth_date") or "").strip()
        note = (request.form.get("note") or "").strip()
        group_id_raw = (request.form.get("group_id") or "").strip()

        if not last_name or not first_name:
            flash("Укажите фамилию и имя воспитанника.", "warning")
            return redirect(url_for("preschool.edit_child", child_id=child.id))

        birth_date = None
        if birth_date_raw:
            from datetime import datetime
            try:
                birth_date = datetime.strptime(birth_date_raw, "%Y-%m-%d").date()
            except ValueError:
                birth_date = None

        child.group_id = int(group_id_raw) if group_id_raw.isdigit() else None
        child.last_name = last_name
        child.first_name = first_name
        child.middle_name = middle_name or None
        child.birth_date = birth_date
        child.personal_account = personal_account or None
        child.reg_address = reg_address or None
        child.living_address = living_address or None
        child.actual_address = actual_address or None
        child.note = note or None

        db.session.commit()

        year_id = child.group.academic_year_id if child.group else None
        flash("Карточка воспитанника обновлена.", "success")
        return redirect(url_for("preschool.children_registry", academic_year_id=year_id))

    representatives = (
        PreschoolRepresentative.query
        .filter(PreschoolRepresentative.child_id == child.id)
        .order_by(PreschoolRepresentative.full_name.asc())
        .all()
    )

    return render_template(
        "preschool/child_form.html",
        child=child,
        groups=groups,
        all_years=all_years,
        year=year,
        representatives=representatives,
    )


@bp.route("/children/<int:child_id>/delete", methods=["POST"])
def delete_child(child_id):
    child = PreschoolChild.query.get_or_404(child_id)
    year_id = child.group.academic_year_id if child.group else None

    db.session.delete(child)
    db.session.commit()

    flash("Воспитанник удалён из тестового контингента ДОУ.", "success")
    return redirect(url_for("preschool.children_registry", academic_year_id=year_id))

