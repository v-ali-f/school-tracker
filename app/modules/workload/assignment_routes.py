from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from flask import (
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.orm.attributes import set_committed_value

from app.core.cache import cache, make_key
from app.core.extensions import db
from app.core.feature_flags import WORKLOAD_WRITE, is_feature_enabled
from app.models import (
    Building,
    Department,
    EducationActivity,
    EducationPlan,
    EducationPlanBinding,
    EducationPlanLine,
    OrganizationSettings,
    PopulationSnapshotClass,
    SchoolClass,
    TariffCycle,
    TariffCalculationRun,
    TariffLine,
    TariffVersion,
    TeacherLoad,
    TeachingGroup,
    TeachingGroupClass,
    TeachingMetagroupSource,
    User,
    WORKLOAD_ASSIGNMENT_KINDS,
    WORKLOAD_ASSIGNMENT_KIND_LABELS,
    WORKLOAD_ASSIGNMENT_STATUS_LABELS,
    WORKLOAD_APPROVAL_STATUS_LABELS,
    WORKLOAD_NEED_STATUS_LABELS,
    WorkloadAssignment,
    WorkloadNeed,
)
from app.services.workload_distribution_service import (
    WorkloadDistributionError,
    add_assignment_change,
    assignment_snapshot,
    calculate_assignment_annual_hours,
    cancel_assignment,
    decimal_hours,
    generate_plan_needs,
    resolve_line_hours,
    refresh_need_status,
    require_assignment_editable,
    teacher_totals,
    validate_assignment,
)
from app.services.workload_snapshot_service import (
    relink_assignments_to_population_snapshot,
)
from app.services.class_plan_matrix_service import (
    EDUCATION_LEVEL_GRADES,
    EDUCATION_LEVEL_LABELS,
    preload_class_plan_matrix_data,
)
from app.services.workload_assignment_matrix_service import (
    LIST_PLAN_KIND_LABELS,
    PLAN_KIND_LABELS,
    PLAN_KIND_ORDER,
    build_workload_assignment_list,
    build_workload_assignment_matrix,
    need_education_level,
    need_grades,
    need_matches_department,
    need_population_snapshot_ids,
    need_plan_kind,
)
from app.services.teaching_group_matrix_service import (
    build_teaching_group_matrix,
    materialize_default_teaching_groups,
)
from app.services.teaching_group_service import (
    current_population_snapshot,
    ensure_population_snapshot,
)
from app.services.teacher_mcko_service import mcko_overviews_for_teachers
from app.utils.user_matching import normalize_fio
from app.services.workload_editing_workflow_service import (
    WorkloadEditingWorkflowError,
    change_workload_approval_status,
    require_workload_editable,
)

from .access import can_use_workload_permission, require_workload_write
from .scopes import resolve_workload_scope


ZERO = Decimal("0")
DEPARTMENT_NEUTRAL_CURRICULUM_ACTIVITY_NAMES = {
    "индивидуальный проект",
}
WORKSPACE_HOLDER_PAGE_SIZES = (5, 10, 20)
WORKSPACE_DEFAULT_HOLDER_PAGE_SIZE = 5
WORKSPACE_FILTER_FIELDS = (
    "version_id",
    "view",
    "department_id",
    "building_id",
    "education_level",
    "grade",
    "subject_id",
    "teacher_query",
    "presentation",
    "holder_page_size",
)


def _current_organization_id():
    organization = (
        OrganizationSettings.query
        .filter_by(is_active=True)
        .order_by(OrganizationSettings.id.asc())
        .first()
    )
    return organization.id if organization else None


def _require_assignments_read():
    if not can_use_workload_permission("workload.read", current_user):
        abort(403)


def _require_assignments_update():
    require_workload_write()
    if not can_use_workload_permission(
        "workload.assignments.update",
        current_user,
    ):
        abort(403)


def _require_version_workload_editable(version):
    try:
        require_workload_editable(version)
    except WorkloadEditingWorkflowError as exc:
        abort(409, description=str(exc))


def _can_review_workload():
    return bool(
        {"ADMIN", "DIRECTOR"}.intersection(
            {
                str(code).upper()
                for code in getattr(current_user, "role_codes", ())
            }
        )
    )


def _parse_date(value, label):
    text = (value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise WorkloadDistributionError(
            f"Укажите корректную дату: {label}."
        ) from exc


def _draft_versions_query():
    organization_id = _current_organization_id()
    query = TariffVersion.query.join(TariffCycle).filter(
        TariffVersion.status == "DRAFT",
    )
    if organization_id is None:
        query = query.filter(TariffCycle.organization_id.is_(None))
    else:
        query = query.filter(TariffCycle.organization_id == organization_id)
    return query.order_by(TariffVersion.id.desc())


def _scoped_need_query():
    query = WorkloadNeed.query.options(
        joinedload(WorkloadNeed.education_activity).selectinload(
            EducationActivity.department_links
        ),
        joinedload(WorkloadNeed.teaching_group)
        .selectinload(TeachingGroup.source_classes)
        .joinedload(TeachingGroupClass.population_snapshot_class)
        .joinedload(PopulationSnapshotClass.building),
        joinedload(WorkloadNeed.teaching_group)
        .joinedload(TeachingGroup.building),
        joinedload(WorkloadNeed.teaching_group)
        .joinedload(TeachingGroup.metagroup_membership)
        .joinedload(TeachingMetagroupSource.metagroup),
        joinedload(WorkloadNeed.teaching_group)
        .joinedload(TeachingGroup.source_plan_line)
        .joinedload(EducationPlanLine.education_plan)
        .joinedload(EducationPlan.root_plan),
        joinedload(WorkloadNeed.teaching_group)
        .selectinload(TeachingGroup.metagroup_sources)
        .joinedload(TeachingMetagroupSource.source_group)
        .selectinload(TeachingGroup.source_classes)
        .joinedload(TeachingGroupClass.population_snapshot_class)
        .joinedload(PopulationSnapshotClass.building),
        joinedload(WorkloadNeed.teaching_group)
        .selectinload(TeachingGroup.metagroup_sources)
        .joinedload(TeachingMetagroupSource.source_group)
        .joinedload(TeachingGroup.metagroup_membership)
        .joinedload(TeachingMetagroupSource.metagroup),
        joinedload(WorkloadNeed.teaching_group)
        .selectinload(TeachingGroup.metagroup_sources)
        .joinedload(TeachingMetagroupSource.source_group)
        .joinedload(TeachingGroup.source_plan_line)
        .joinedload(EducationPlanLine.education_plan)
        .joinedload(EducationPlan.root_plan),
    )
    organization_id = _current_organization_id()
    if organization_id is None:
        query = query.filter(WorkloadNeed.organization_id.is_(None))
    else:
        query = query.filter(
            WorkloadNeed.organization_id == organization_id
        )
    scope = resolve_workload_scope(current_user)
    if not scope.unrestricted:
        if scope.department_ids:
            query = query.filter(
                WorkloadNeed.department_id.in_(scope.department_ids)
            )
        if scope.building_ids:
            query = query.filter(
                WorkloadNeed.building_id.in_(scope.building_ids)
            )
        if not scope.department_ids and not scope.building_ids:
            return query.filter(db.false())
    return query


def _get_need(need_id, *, for_update=False):
    if for_update:
        _require_assignments_update()
    else:
        _require_assignments_read()
    need = _scoped_need_query().filter(WorkloadNeed.id == need_id).first_or_404()
    if for_update:
        _require_version_workload_editable(need.tariff_version)
    return need


def _get_assignment(assignment_id, *, for_update=False):
    assignment = WorkloadAssignment.query.get_or_404(assignment_id)
    need = _get_need(assignment.workload_need_id, for_update=for_update)
    if assignment.workload_need_id != need.id:
        abort(404)
    return assignment


def _active_employees():
    return (
        User.query
        .filter(
            User.is_active_user.is_(True),
            User.employment_status == "ACTIVE",
            User.archived_at.is_(None),
        )
        .order_by(
            User.last_name.asc(),
            User.first_name.asc(),
            User.middle_name.asc(),
        )
        .all()
    )


def _scope_options():
    scope = resolve_workload_scope(current_user)
    departments = Department.query.order_by(Department.name.asc()).all()
    buildings = Building.query.order_by(Building.name.asc()).all()
    if not scope.unrestricted:
        if scope.department_ids:
            departments = [
                item for item in departments
                if item.id in scope.department_ids
            ]
        if scope.building_ids:
            buildings = [
                item for item in buildings
                if item.id in scope.building_ids
            ]
    return departments, buildings


def _workspace_redirect(**overrides):
    values = {}
    for field in WORKSPACE_FILTER_FIELDS:
        field_values = [
            value.strip()
            for value in request.form.getlist(field)
            if value and value.strip()
        ]
        if field_values:
            values[field] = (
                field_values
                if field in {"education_level", "grade", "subject_id"}
                else field_values[-1]
            )
    values.update(overrides)
    return redirect(url_for("workload.assignment_workspace", **values))


def _workspace_selected_levels(source):
    return {
        value.strip().upper()
        for value in source.getlist("education_level")
        if value and value.strip().upper() in {"NOO", "OOO", "SOO"}
    }


def _workspace_selected_grades(source):
    result = set()
    for value in source.getlist("grade"):
        try:
            grade = int(value)
        except (TypeError, ValueError):
            continue
        if grade in range(1, 12):
            result.add(grade)
    return result


def _workspace_selected_subject_ids(source):
    result = set()
    for value in source.getlist("subject_id"):
        try:
            subject_id = int(value)
        except (TypeError, ValueError):
            continue
        if subject_id > 0:
            result.add(subject_id)
    return result


def _workspace_state(version_id):
    key = f"workload_matrix_state_{version_id}"
    state = session.get(key)
    if not isinstance(state, dict):
        state = {"teacher_ids": [], "vacancies": [], "rows": []}
    state.setdefault("teacher_ids", [])
    state.setdefault("vacancies", [])
    state.setdefault("rows", [])
    return key, state


def _save_workspace_state(key, state):
    session[key] = state
    session.modified = True


def _normalized_vacancy_note(value):
    note = " ".join((value or "").split())
    if note.startswith("(") and note.endswith(")"):
        note = note[1:-1].strip()
    return note[:100].strip()


def _vacancy_label(number, note=None):
    label = f"Вакансия {number}"
    normalized_note = _normalized_vacancy_note(note)
    return f"{label} ({normalized_note})" if normalized_note else label


def _next_vacancy(version_id, state, *, note=None):
    existing_codes = {
        str(item.get("key"))
        for item in state.get("vacancies", [])
        if item.get("key")
    }
    existing_codes.update(
        code for (code,) in db.session.query(
            WorkloadAssignment.position_code,
        ).filter(
            WorkloadAssignment.tariff_version_id == version_id,
            WorkloadAssignment.assignment_kind == "VACANCY",
            WorkloadAssignment.status != "CANCELLED",
        ).distinct().all()
        if code
    )
    number = 1
    while f"VACANCY_{number}" in existing_codes:
        number += 1
    return {
        "key": f"VACANCY_{number}",
        "label": _vacancy_label(number, note),
    }


def _filter_workspace_needs(
    needs,
    *,
    department_id=None,
    building_id=None,
    education_level=None,
    grade=None,
    education_levels=None,
    grades=None,
    subject_ids=None,
    population_snapshot_id=None,
    bound_plan_ids_by_class=None,
):
    selected_levels = set(education_levels or ())
    selected_grades = set(grades or ())
    selected_subject_ids = set(subject_ids or ())
    if education_level:
        selected_levels.add(education_level)
    if grade:
        selected_grades.add(grade)
    result = []
    for need in needs:
        if (
            selected_subject_ids
            and need.education_activity_id not in selected_subject_ids
        ):
            continue
        snapshot_ids = need_population_snapshot_ids(need)
        if (
            population_snapshot_id is not None
            and snapshot_ids
            and snapshot_ids != {population_snapshot_id}
        ):
            continue
        if (
            bound_plan_ids_by_class is not None
            and not _need_matches_current_plan_bindings(
                need,
                bound_plan_ids_by_class,
            )
        ):
            continue
        if building_id and need.building_id != building_id:
            continue
        if (
            selected_levels
            and need_education_level(need) not in selected_levels
        ):
            continue
        if selected_grades and not selected_grades.intersection(
            need_grades(need)
        ):
            continue
        if not need_matches_department(need, department_id):
            continue
        result.append(need)
    return result


def _workspace_bound_plan_ids(snapshot):
    if snapshot is None:
        return {}
    rows = (
        db.session.query(
            EducationPlanBinding.population_snapshot_class_id,
            EducationPlanBinding.education_plan_id,
        )
        .join(
            PopulationSnapshotClass,
            PopulationSnapshotClass.id
            == EducationPlanBinding.population_snapshot_class_id,
        )
        .filter(
            PopulationSnapshotClass.population_snapshot_id == snapshot.id,
        )
        .all()
    )
    result = defaultdict(set)
    for class_id, plan_id in rows:
        result[class_id].add(plan_id)
    return result


def _need_matches_current_plan_bindings(need, bound_plan_ids_by_class):
    group = need.teaching_group
    if group is None:
        return True
    source_groups = (
        [link.source_group for link in group.metagroup_sources]
        if group.group_type == "METAGROUP"
        else [group]
    )
    for source_group in source_groups:
        line = source_group.source_plan_line if source_group else None
        plan = line.education_plan if line else None
        if plan is None:
            continue
        root_plan_id = plan.root_plan_id or plan.id
        for source_class in source_group.source_classes:
            if root_plan_id not in bound_plan_ids_by_class.get(
                source_class.population_snapshot_class_id,
                (),
            ):
                return False
    return True


def _workspace_form_needs(version, *, activity_id=None):
    view_mode = (request.form.get("view") or "all").strip().lower()
    department_id = (
        request.form.get("department_id", type=int)
        if view_mode == "department" else None
    )
    education_levels = _workspace_selected_levels(request.form)
    grades = _workspace_selected_grades(request.form)
    subject_ids = _workspace_selected_subject_ids(request.form)
    query = _scoped_need_query().filter(
        WorkloadNeed.tariff_version_id == version.id,
        WorkloadNeed.status.in_(("OPEN", "PARTIAL", "COVERED")),
    )
    if activity_id is not None:
        query = query.filter(
            WorkloadNeed.education_activity_id == activity_id
        )
    elif subject_ids:
        query = query.filter(
            WorkloadNeed.education_activity_id.in_(subject_ids)
        )
    needs = query.all()
    snapshot = current_population_snapshot(version.id)
    bound_plan_ids_by_class = _workspace_bound_plan_ids(snapshot)
    return _filter_workspace_needs(
        needs,
        department_id=department_id,
        building_id=request.form.get("building_id", type=int),
        education_levels=education_levels,
        grades=grades,
        subject_ids=subject_ids,
        population_snapshot_id=(snapshot.id if snapshot else None),
        bound_plan_ids_by_class=bound_plan_ids_by_class,
    )


def _state_row_matches_holder(
    row,
    holder_type,
    teacher_id,
    vacancy_key,
):
    row_holder_type = row.get("holder_type", "teacher")
    if row_holder_type != holder_type:
        return False
    if holder_type == "vacancy":
        return row.get("vacancy_key") == vacancy_key
    return row.get("teacher_id") == teacher_id


def _workspace_teacher_metadata(teachers, selected_version):
    teacher_ids = [teacher.id for teacher in teachers]
    if not teacher_ids or selected_version is None:
        return {}
    academic_year_id = selected_version.tariff_cycle.academic_year_id
    class_names = defaultdict(list)
    class_rows = (
        SchoolClass.query
        .options(joinedload(SchoolClass.teacher_user))
        .filter(
            SchoolClass.academic_year_id == academic_year_id,
            SchoolClass.teacher_user_id.isnot(None),
            SchoolClass.is_active.is_(True),
            SchoolClass.is_archived.is_(False),
        )
        .order_by(SchoolClass.grade.asc(), SchoolClass.name.asc())
        .all()
    )
    teacher_by_id = {teacher.id: teacher for teacher in teachers}
    for item in class_rows:
        if item.teacher_user_id in teacher_by_id:
            class_names[item.teacher_user_id].append(item.name)

    # Imported staff can occasionally exist as two User rows: the class
    # registry points to one row while workload assignments point to another.
    # Link such rows only when the normalized FIO is unambiguous on both sides.
    unmatched_teachers_by_fio = defaultdict(list)
    for teacher in teachers:
        if class_names.get(teacher.id):
            continue
        fio_key = normalize_fio(
            teacher.last_name,
            teacher.first_name,
            teacher.middle_name,
        )
        if fio_key:
            unmatched_teachers_by_fio[fio_key].append(teacher)

    assigned_users_by_fio = defaultdict(set)
    assigned_classes_by_fio = defaultdict(list)
    for item in class_rows:
        assigned_user = item.teacher_user
        if assigned_user is None:
            continue
        fio_key = normalize_fio(
            assigned_user.last_name,
            assigned_user.first_name,
            assigned_user.middle_name,
        )
        if not fio_key:
            continue
        assigned_users_by_fio[fio_key].add(assigned_user.id)
        assigned_classes_by_fio[fio_key].append(item.name)

    for fio_key, matching_teachers in unmatched_teachers_by_fio.items():
        if len(matching_teachers) != 1:
            continue
        if len(assigned_users_by_fio.get(fio_key, ())) != 1:
            continue
        class_names[matching_teachers[0].id].extend(
            assigned_classes_by_fio[fio_key]
        )
    mcko_by_teacher = mcko_overviews_for_teachers(teacher_ids)
    return {
        teacher_id: {
            "class_teacher": ", ".join(class_names.get(teacher_id, [])) or "—",
            "mcko_overview": mcko_by_teacher.get(teacher_id),
            "mcko_items": list(mcko_by_teacher[teacher_id].valid_results)
            if teacher_id in mcko_by_teacher else [],
        }
        for teacher_id in teacher_ids
    }


def _workspace_teacher_choices(
    employees,
    selected_version,
    department_id,
    assignments,
    excluded_teacher_ids=None,
):
    excluded_teacher_ids = set(excluded_teacher_ids or ())
    available_employees = [
        item for item in employees if item.id not in excluded_teacher_ids
    ]
    suggested_ids = {
        item.employee_user_id
        for item in assignments
        if item.employee_user_id is not None
        and (not department_id or item.department_id == department_id)
    }
    if selected_version is not None and department_id:
        academic_year_id = selected_version.tariff_cycle.academic_year_id
        suggested_ids.update(
            teacher_id
            for (teacher_id,) in (
                TeacherLoad.query
                .with_entities(TeacherLoad.teacher_id)
                .filter(
                    TeacherLoad.department_id == department_id,
                    TeacherLoad.is_archived.is_(False),
                    db.or_(
                        TeacherLoad.academic_year_id == academic_year_id,
                        TeacherLoad.academic_year_id.is_(None),
                    ),
                )
                .distinct()
                .all()
            )
        )
    return {
        "suggested": [
            item for item in available_employees if item.id in suggested_ids
        ],
        "other": [
            item for item in available_employees if item.id not in suggested_ids
        ],
    }


def _workspace_assignment_holder_key(assignment):
    if assignment.assignment_kind == "VACANCY":
        return f"vacancy:{assignment.position_code}"
    return f"teacher:{assignment.employee_user_id}"


def _workspace_department_holder_keys(
    assignments,
    needs_by_id,
    department_id,
):
    """Return holders belonging to a department by curriculum workload.

    Extracurricular and additional-education rows are intentionally ignored
    when department membership is established.  Once a holder qualifies, the
    workspace can display all of that holder's workload.
    """
    if not department_id:
        return None
    holder_keys = set()
    for assignment in assignments:
        need = needs_by_id.get(assignment.workload_need_id)
        if (
            need is None
            or need_plan_kind(need) != "CURRICULUM"
            or _workspace_search_text(need.education_activity.name)
            in DEPARTMENT_NEUTRAL_CURRICULUM_ACTIVITY_NAMES
            or not need_matches_department(need, department_id)
        ):
            continue
        holder_keys.add(_workspace_assignment_holder_key(assignment))
    return holder_keys


def _workspace_state_department_holder_keys(
    rows,
    activities_by_id,
    department_id,
):
    """Apply the same department rule to unsaved workspace holder rows."""
    if not department_id:
        return set()
    holder_keys = set()
    for row in rows:
        if row.get("plan_kind") != "CURRICULUM":
            continue
        activity = activities_by_id.get(row.get("activity_id"))
        if (
            activity is None
            or _workspace_search_text(activity.name)
            in DEPARTMENT_NEUTRAL_CURRICULUM_ACTIVITY_NAMES
            or not any(
                link.is_active and link.department_id == department_id
                for link in activity.department_links
            )
        ):
            continue
        if row.get("holder_type", "teacher") == "teacher":
            teacher_id = row.get("teacher_id")
            if teacher_id is not None:
                holder_keys.add(f"teacher:{teacher_id}")
        elif row.get("holder_type") == "vacancy":
            vacancy_key = row.get("vacancy_key")
            if vacancy_key:
                holder_keys.add(f"vacancy:{vacancy_key}")
    return holder_keys


def _workspace_assignment_holder_label(assignment):
    if assignment.assignment_kind == "VACANCY":
        return assignment.position_title or "Вакансия"
    if assignment.employee is not None:
        return assignment.employee.fio
    return "Преподаватель"


def _workspace_search_text(value):
    return " ".join((value or "").casefold().replace("ё", "е").split())


def _workspace_subject_options(needs):
    options = {}
    for need in needs:
        plan_kind = (
            need.teaching_group.source_plan_line.education_plan.plan_kind
            if need.teaching_group
            and need.teaching_group.source_plan_line
            and need.teaching_group.source_plan_line.education_plan
            else "CURRICULUM"
        )
        options[(need.education_activity_id, plan_kind)] = {
            "activity": need.education_activity,
            "plan_kind": plan_kind,
            "plan_kind_label": PLAN_KIND_LABELS.get(plan_kind, plan_kind),
        }
    return sorted(
        options.values(),
        key=lambda item: (
            PLAN_KIND_ORDER.get(item["plan_kind"], 99),
            item["activity"].name.casefold(),
        ),
    )


def _workspace_subject_filter_options(version):
    if version is None:
        return []
    query = (
        EducationActivity.query
        .join(
            WorkloadNeed,
            WorkloadNeed.education_activity_id == EducationActivity.id,
        )
        .filter(
            WorkloadNeed.tariff_version_id == version.id,
            WorkloadNeed.status != "CANCELLED",
        )
    )
    organization_id = _current_organization_id()
    query = (
        query.filter(WorkloadNeed.organization_id.is_(None))
        if organization_id is None
        else query.filter(WorkloadNeed.organization_id == organization_id)
    )
    scope = resolve_workload_scope(current_user)
    if not scope.unrestricted:
        if scope.department_ids:
            query = query.filter(
                WorkloadNeed.department_id.in_(scope.department_ids)
            )
        if scope.building_ids:
            query = query.filter(
                WorkloadNeed.building_id.in_(scope.building_ids)
            )
        if not scope.department_ids and not scope.building_ids:
            return []
    return query.order_by(EducationActivity.name.asc()).distinct().all()


def _workspace_plan_subject_options(plan_matrices, department_id=None):
    options = {}
    for matrix in plan_matrices:
        for section in matrix.get("sections", []):
            for row in section["rows"]:
                activity = row["activity"]
                links = [
                    link for link in activity.department_links
                    if link.is_active
                ]
                if (
                    department_id
                    and links
                    and not any(
                        link.department_id == department_id
                        for link in links
                    )
                ):
                    continue
                options[(activity.id, row["plan_kind"])] = {
                    "activity": activity,
                    "plan_kind": row["plan_kind"],
                    "plan_kind_label": PLAN_KIND_LABELS.get(
                        row["plan_kind"],
                        row["plan_kind"],
                    ),
                }
    return sorted(
        options.values(),
        key=lambda item: (
            PLAN_KIND_ORDER.get(item["plan_kind"], 99),
            item["activity"].name.casefold(),
        ),
    )


def _workspace_plan_context(
    version,
    *,
    education_level=None,
    grade=None,
    education_levels=None,
    grades=None,
    building_id=None,
    include_matrices=True,
):
    if version is None:
        return None, [], []
    snapshot = ensure_population_snapshot(
        version,
        user_id=current_user.id,
    )[0]
    plans = (
        EducationPlan.query
        .filter_by(
            tariff_version_id=version.id,
            plan_kind="CURRICULUM",
            root_plan_id=None,
        )
        .order_by(
            EducationPlan.education_level.asc(),
            EducationPlan.name.asc(),
        )
        .all()
    )
    scope = resolve_workload_scope(current_user)
    allowed_building_ids = (
        None if scope.unrestricted else set(scope.building_ids)
    )
    selected_levels = set(education_levels or ())
    selected_grades = set(grades or ())
    if education_level:
        selected_levels.add(education_level)
    if grade:
        selected_grades.add(grade)
    matrix_specs = _workspace_matrix_specs(
        selected_levels,
        selected_grades,
    )
    matrix_data = (
        preload_class_plan_matrix_data(
            snapshot,
            plans,
            compact_enrollments=True,
        )
        if snapshot is not None and include_matrices else None
    )
    matrices = [
        build_teaching_group_matrix(
            snapshot,
            plans,
            level,
            version.id,
            grade=matrix_grade,
            building_id=building_id,
            allowed_building_ids=allowed_building_ids,
            matrix_data=matrix_data,
        )
        for level, matrix_grade in matrix_specs
    ] if snapshot is not None and include_matrices else []
    return snapshot, plans, matrices


def _workspace_matrix_specs(education_levels, grades):
    selected_levels = set(education_levels or ())
    selected_grades = set(grades or ())
    if selected_grades:
        specs = []
        for grade in sorted(selected_grades):
            level = (
                "NOO" if grade in range(1, 5)
                else "OOO" if grade in range(5, 10)
                else "SOO"
            )
            if not selected_levels or level in selected_levels:
                specs.append((level, grade))
        return specs
    levels = selected_levels or {"NOO", "OOO", "SOO"}
    return [
        (level, None)
        for level in ("NOO", "OOO", "SOO")
        if level in levels
    ]


def _workspace_matrix_group_ids(plan_matrices):
    """Return target teaching groups represented by the visible matrices."""
    group_ids = set()
    for plan_matrix in plan_matrices or ():
        for section in plan_matrix.get("sections", ()):
            for row in section.get("rows", ()):
                for cell in row.get("cells", {}).values():
                    for source_group in cell.get("groups", ()):
                        membership = source_group.metagroup_membership
                        target_group = (
                            membership.metagroup
                            if membership is not None
                            else source_group
                        )
                        if target_group is not None and target_group.id:
                            group_ids.add(target_group.id)
    return group_ids


def _workspace_need_group_coverage(version):
    merged_source_ids = {
        item.source_group_id
        for item in (
            TeachingMetagroupSource.query
            .join(
                TeachingGroup,
                TeachingGroup.id == TeachingMetagroupSource.metagroup_id,
            )
            .filter(
                TeachingGroup.tariff_version_id == version.id,
                TeachingGroup.status != "CLOSED",
            )
            .all()
        )
    }
    expected_group_ids = set()
    expected_hours_by_group = {}
    for group in (
        TeachingGroup.query
        .options(
            joinedload(TeachingGroup.source_plan_line).selectinload(
                EducationPlanLine.periods
            ),
            selectinload(TeachingGroup.metagroup_sources)
            .joinedload(TeachingMetagroupSource.source_group)
            .joinedload(TeachingGroup.source_plan_line)
            .selectinload(EducationPlanLine.periods),
        )
        .filter(
            TeachingGroup.tariff_version_id == version.id,
            TeachingGroup.status != "CLOSED",
        )
        .all()
    ):
        if group.id in merged_source_ids:
            continue
        source_line = (
            group.metagroup_sources[0].source_group.source_plan_line
            if group.group_type == "METAGROUP" and group.metagroup_sources
            else group.source_plan_line
        )
        if source_line is None:
            continue
        weekly, annual = resolve_line_hours(
            source_line,
            group.valid_from,
            group.valid_to,
        )
        if weekly > ZERO:
            expected_group_ids.add(group.id)
            expected_hours_by_group[group.id] = (weekly, annual)
    current_needs = WorkloadNeed.query.filter(
            WorkloadNeed.tariff_version_id == version.id,
            WorkloadNeed.status != "CANCELLED",
            WorkloadNeed.need_kind == "PLAN",
            WorkloadNeed.teaching_group_id.isnot(None),
        ).all()
    current_group_ids = {
        item.teaching_group_id for item in current_needs
    }
    current_hours_by_group = defaultdict(set)
    for need in current_needs:
        current_hours_by_group[need.teaching_group_id].add((
            Decimal(need.weekly_hours or ZERO).quantize(Decimal("0.001")),
            Decimal(need.annual_hours or ZERO).quantize(Decimal("0.001")),
        ))
    hours_match = all(
        expected_hours in current_hours_by_group.get(group_id, set())
        for group_id, expected_hours in expected_hours_by_group.items()
    )
    return expected_group_ids, current_group_ids, hours_match


def _ensure_workspace_plan_needs(
    version,
    snapshot,
    plans,
    plan_matrices=None,
):
    if version is None or snapshot is None or version.status != "DRAFT":
        return False
    if db.engine.dialect.name == "postgresql":
        # Several Gunicorn workers can open the workspace at the same time
        # immediately after a plan binding changes.  Serialise materialisation
        # per tariff version so they cannot create the same automatic group.
        db.session.execute(
            text(
                "SELECT pg_advisory_xact_lock("
                "hashtext('workload-plan-needs'), :version_id)"
            ),
            {"version_id": version.id},
        )
    plan_state = (
        db.session.query(
            db.func.count(EducationPlanLine.id),
            db.func.max(EducationPlanLine.updated_at),
            db.func.sum(EducationPlan.revision),
            db.func.max(EducationPlan.updated_at),
        )
        .join(
            EducationPlan,
            EducationPlan.id == EducationPlanLine.education_plan_id,
        )
        .filter(EducationPlan.tariff_version_id == version.id)
        .one()
    )
    group_state = (
        db.session.query(
            db.func.count(TeachingGroup.id),
            db.func.max(TeachingGroup.updated_at),
            db.func.sum(TeachingGroup.revision),
        )
        .filter(
            TeachingGroup.tariff_version_id == version.id,
            TeachingGroup.status != "CLOSED",
        )
        .one()
    )
    binding_state = (
        db.session.query(
            db.func.count(EducationPlanBinding.id),
            db.func.max(EducationPlanBinding.updated_at),
            db.func.sum(EducationPlanBinding.revision),
        )
        .join(
            EducationPlan,
            EducationPlan.id == EducationPlanBinding.education_plan_id,
        )
        .filter(EducationPlan.tariff_version_id == version.id)
        .one()
    )
    need_state = (
        db.session.query(
            db.func.count(WorkloadNeed.id),
            db.func.max(WorkloadNeed.id),
            db.func.max(WorkloadNeed.updated_at),
        )
        .filter(
            WorkloadNeed.tariff_version_id == version.id,
            WorkloadNeed.status != "CANCELLED",
        )
        .one()
    )
    sync_key = make_key(
        "workload-plan-needs",
        version.id,
        snapshot.checksum,
        *plan_state,
        *group_state,
        *binding_state,
        *need_state,
    )
    if cache.get(sync_key):
        return False
    latest_need_update = need_state[2]
    plan_and_binding_updates = tuple(filter(None, (
        plan_state[1],
        plan_state[3],
        binding_state[1],
    )))
    if (
        need_state[0]
        and latest_need_update is not None
        and all(
            updated_at <= latest_need_update
            for updated_at in plan_and_binding_updates
        )
    ):
        expected_group_ids, current_group_ids, hours_match = (
            _workspace_need_group_coverage(version)
        )
        if expected_group_ids == current_group_ids and hours_match:
            cache.set(sync_key, True, timeout=3600)
            return False
    created_groups = materialize_default_teaching_groups(
        version=version,
        snapshot=snapshot,
        plans=plans,
        user_id=current_user.id,
        matrices=plan_matrices,
    )
    db.session.flush()
    expected_group_ids, current_group_ids, hours_match = (
        _workspace_need_group_coverage(version)
    )
    if (
        not created_groups
        and expected_group_ids == current_group_ids
        and hours_match
    ):
        cache.set(sync_key, True, timeout=3600)
        return False
    generate_plan_needs(version, user_id=current_user.id)
    db.session.commit()
    cache.set(sync_key, True, timeout=3600)
    return True


def _assignment_from_form(need, assignment=None):
    kind = (request.form.get("assignment_kind") or "").strip().upper()
    employee_id = request.form.get("employee_user_id", type=int)
    if kind == "VACANCY":
        employee_id = None
    weekly = decimal_hours(
        request.form.get("weekly_hours"),
        "часов в неделю",
    )
    annual_input = decimal_hours(
        request.form.get("annual_hours"),
        "часов за период",
        required=False,
    )
    annual = calculate_assignment_annual_hours(
        need,
        weekly,
        annual_input,
    )
    item = assignment or WorkloadAssignment(
        organization_id=need.organization_id,
        tariff_version_id=need.tariff_version_id,
        workload_need_id=need.id,
        created_by_user_id=current_user.id,
    )
    item.employee_user_id = employee_id
    item.position_code = (
        " ".join((request.form.get("position_code") or "").split()).upper()
        or "TEACHER"
    )
    item.position_title = (
        " ".join((request.form.get("position_title") or "").split())
        or ("Вакансия" if kind == "VACANCY" else "Учитель")
    )
    item.department_id = (
        request.form.get("department_id", type=int)
        or need.department_id
    )
    item.building_id = (
        request.form.get("building_id", type=int)
        or need.building_id
    )
    item.assignment_kind = kind
    item.date_from = _parse_date(
        request.form.get("date_from"),
        "начало назначения",
    )
    item.date_to = _parse_date(
        request.form.get("date_to"),
        "окончание назначения",
    )
    item.weekly_hours = weekly
    item.annual_hours = annual
    item.updated_by_user_id = current_user.id
    validate_assignment(
        need,
        item,
        exclude_assignment_id=item.id,
    )
    return item


def _render_assignment_form(need, assignment=None):
    departments, buildings = _scope_options()
    return render_template(
        "workload/assignment_form.html",
        need=need,
        assignment=assignment,
        employees=_active_employees(),
        departments=departments,
        buildings=buildings,
        assignment_kinds=WORKLOAD_ASSIGNMENT_KINDS,
        assignment_kind_labels=WORKLOAD_ASSIGNMENT_KIND_LABELS,
    )


def _teacher_rows(assignments):
    by_teacher = defaultdict(list)
    for assignment in assignments:
        if assignment.employee_user_id is not None:
            by_teacher[assignment.employee_user_id].append(assignment)
    rows = []
    for employee_id, items in by_teacher.items():
        totals = teacher_totals(items)
        employee = items[0].employee
        departments = sorted({
            item.department.name
            for item in items
            if item.department is not None
        })
        rows.append({
            "employee": employee,
            "assignments": items,
            "totals": totals,
            "departments": departments,
        })
    return sorted(rows, key=lambda row: row["employee"].fio.lower())


def _build_workload_list_workbook(workload_list, *, title):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нагрузка списком"
    sheet.sheet_view.showGridLines = False
    columns = [
        "№",
        "Предмет",
        "Класс",
        "Часы",
        "Группа / весь класс",
        "Здание",
    ]
    header_fill = PatternFill("solid", fgColor="2F6FED")
    section_fill = PatternFill("solid", fgColor="DCE8FA")
    total_fill = PatternFill("solid", fgColor="EAF3EA")

    sheet.append([title, "", "", "", "", ""])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    sheet.cell(1, 1).fill = header_fill
    sheet.row_dimensions[1].height = 24

    for block in workload_list["blocks"]:
        teacher_row = sheet.max_row + 2
        sheet.cell(teacher_row, 1, block["label"])
        sheet.merge_cells(
            start_row=teacher_row,
            start_column=1,
            end_row=teacher_row,
            end_column=6,
        )
        sheet.cell(teacher_row, 1).font = Font(
            bold=True,
            color="FFFFFF",
        )
        sheet.cell(teacher_row, 1).fill = header_fill
        number = 1
        for section in block["sections"]:
            section_row = sheet.max_row + 1
            sheet.cell(section_row, 1, section["label"])
            sheet.merge_cells(
                start_row=section_row,
                start_column=1,
                end_row=section_row,
                end_column=6,
            )
            sheet.cell(section_row, 1).font = Font(bold=True)
            sheet.cell(section_row, 1).fill = section_fill
            sheet.append(columns)
            for cell in sheet[sheet.max_row]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            for row in section["rows"]:
                sheet.append([
                    number,
                    row["subject_name"],
                    row["class_label"],
                    float(row["hours"]),
                    row["group_label"],
                    row["building_label"],
                ])
                number += 1
            total_row = sheet.max_row + 1
            sheet.cell(total_row, 1, f"Итого: {section['label']}")
            sheet.merge_cells(
                start_row=total_row,
                start_column=1,
                end_row=total_row,
                end_column=3,
            )
            sheet.cell(total_row, 4, float(section["total"]))
            for cell in sheet[total_row]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
        total_row = sheet.max_row + 1
        sheet.cell(total_row, 1, "Всего у преподавателя")
        sheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=3,
        )
        sheet.cell(total_row, 4, float(block["total"]))
        for cell in sheet[total_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    summary_row = sheet.max_row + 2
    sheet.cell(summary_row, 1, "Итоги по всей выборке")
    sheet.merge_cells(
        start_row=summary_row,
        start_column=1,
        end_row=summary_row,
        end_column=6,
    )
    sheet.cell(summary_row, 1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(summary_row, 1).fill = header_fill
    for plan_kind in sorted(
        PLAN_KIND_ORDER,
        key=lambda item: PLAN_KIND_ORDER[item],
    ):
        row_index = sheet.max_row + 1
        sheet.cell(
            row_index,
            1,
            LIST_PLAN_KIND_LABELS.get(plan_kind, plan_kind),
        )
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=3,
        )
        sheet.cell(
            row_index,
            4,
            float(workload_list["totals"].get(plan_kind, ZERO)),
        )
    grand_total_row = sheet.max_row + 1
    sheet.cell(grand_total_row, 1, "Всего")
    sheet.merge_cells(
        start_row=grand_total_row,
        start_column=1,
        end_row=grand_total_row,
        end_column=3,
    )
    sheet.cell(grand_total_row, 4, float(workload_list["total"]))
    for cell in sheet[grand_total_row]:
        cell.font = Font(bold=True)
        cell.fill = total_fill

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    for column_letter, width in {
        "A": 8,
        "B": 34,
        "C": 18,
        "D": 12,
        "E": 24,
        "F": 28,
    }.items():
        sheet.column_dimensions[column_letter].width = width
    return workbook


def register_assignment_routes(workload_bp):
    @workload_bp.get("/assignments/")
    @login_required
    def assignments():
        _require_assignments_read()
        query = _scoped_need_query()
        version_id = request.args.get("version_id", type=int)
        status = (request.args.get("status") or "").strip().upper()
        department_id = request.args.get("department_id", type=int)
        building_id = request.args.get("building_id", type=int)
        vacancies_only = request.args.get("vacancies") == "1"
        if version_id:
            query = query.filter(
                WorkloadNeed.tariff_version_id == version_id
            )
        if status in WORKLOAD_NEED_STATUS_LABELS:
            query = query.filter(WorkloadNeed.status == status)
        if department_id:
            query = query.filter(WorkloadNeed.department_id == department_id)
        if building_id:
            query = query.filter(WorkloadNeed.building_id == building_id)
        needs = (
            query
            .order_by(
                WorkloadNeed.status.asc(),
                WorkloadNeed.education_activity_id.asc(),
                WorkloadNeed.id.asc(),
            )
            .all()
        )
        if vacancies_only:
            needs = [
                need for need in needs
                if need.status in {"OPEN", "PARTIAL"}
            ]
        totals = {
            "weekly": sum(
                (Decimal(item.weekly_hours or ZERO) for item in needs),
                ZERO,
            ),
            "allocated": sum(
                (item.allocated_weekly_hours for item in needs),
                ZERO,
            ),
        }
        totals["remaining"] = totals["weekly"] - totals["allocated"]
        departments, buildings = _scope_options()
        selected_version = next(
            (
                version
                for version in _draft_versions_query().all()
                if version.id == version_id
            ),
            needs[0].tariff_version if needs else None,
        )
        return render_template(
            "workload/assignments.html",
            needs=needs,
            versions=_draft_versions_query().all(),
            selected_version=selected_version,
            totals=totals,
            departments=departments,
            buildings=buildings,
            status_labels=WORKLOAD_NEED_STATUS_LABELS,
            selected_version_id=version_id,
            selected_status=status,
            selected_department_id=department_id,
            selected_building_id=building_id,
            vacancies_only=vacancies_only,
            can_update=(
                is_feature_enabled(WORKLOAD_WRITE)
                and can_use_workload_permission(
                    "workload.assignments.update",
                    current_user,
                )
                and selected_version is not None
                and selected_version.workload_approval_status == "EDITING"
            ),
        )

    @workload_bp.post("/assignments/generate")
    @login_required
    def assignments_generate():
        _require_assignments_update()
        version_id = request.form.get("version_id", type=int)
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        _require_version_workload_editable(version)
        try:
            result = generate_plan_needs(
                version,
                user_id=current_user.id,
            )
            db.session.commit()
        except WorkloadDistributionError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            flash(
                "Потребность пересчитана: "
                f"создано {result['created']}, "
                f"обновлено {result['updated']}, "
                f"отменено {result['cancelled']}, "
                f"пустых строк пропущено {result['skipped_empty']}.",
                "success",
            )
        if request.form.get("return_to") == "workspace":
            return redirect(
                url_for(
                    "workload.assignment_workspace",
                    version_id=version.id,
                )
            )
        return redirect(url_for("workload.assignments", version_id=version.id))

    @workload_bp.get("/needs/<int:need_id>")
    @login_required
    def need_detail(need_id):
        need = _get_need(need_id)
        return render_template(
            "workload/need_detail.html",
            need=need,
            need_status_labels=WORKLOAD_NEED_STATUS_LABELS,
            assignment_kind_labels=WORKLOAD_ASSIGNMENT_KIND_LABELS,
            assignment_status_labels=WORKLOAD_ASSIGNMENT_STATUS_LABELS,
            can_update=(
                is_feature_enabled(WORKLOAD_WRITE)
                and can_use_workload_permission(
                    "workload.assignments.update",
                    current_user,
                )
                and need.tariff_version.status == "DRAFT"
                and need.tariff_version.workload_approval_status == "EDITING"
                and need.status != "CANCELLED"
            ),
        )

    @workload_bp.route(
        "/needs/<int:need_id>/assignments/new",
        methods=["GET", "POST"],
    )
    @login_required
    def assignment_new(need_id):
        need = _get_need(need_id, for_update=True)
        if request.method == "POST":
            try:
                assignment = _assignment_from_form(need)
                db.session.add(assignment)
                db.session.flush()
                add_assignment_change(
                    assignment,
                    "CREATE",
                    user_id=current_user.id,
                )
                refresh_need_status(need)
                db.session.commit()
            except (WorkloadDistributionError, IntegrityError) as exc:
                db.session.rollback()
                message = (
                    str(exc)
                    if isinstance(exc, WorkloadDistributionError)
                    else "Не удалось сохранить назначение."
                )
                flash(message, "danger")
            else:
                flash("Назначение добавлено.", "success")
                return redirect(
                    url_for("workload.need_detail", need_id=need.id)
                )
        return _render_assignment_form(need)

    @workload_bp.route(
        "/assignments/<int:assignment_id>/edit",
        methods=["GET", "POST"],
    )
    @login_required
    def assignment_edit(assignment_id):
        assignment = _get_assignment(assignment_id, for_update=True)
        need = assignment.workload_need
        if request.method == "POST":
            try:
                require_assignment_editable(
                    assignment,
                    expected_revision=request.form.get(
                        "revision",
                        type=int,
                    ),
                )
                before = assignment_snapshot(assignment)
                reason = " ".join(
                    (request.form.get("reason") or "").split()
                )
                if not reason:
                    raise WorkloadDistributionError(
                        "Укажите основание изменения назначения."
                    )
                _assignment_from_form(need, assignment)
                assignment.revision += 1
                add_assignment_change(
                    assignment,
                    "UPDATE",
                    user_id=current_user.id,
                    before_data=before,
                    reason=reason,
                )
                refresh_need_status(need)
                db.session.commit()
            except (WorkloadDistributionError, IntegrityError) as exc:
                db.session.rollback()
                message = (
                    str(exc)
                    if isinstance(exc, WorkloadDistributionError)
                    else "Не удалось обновить назначение."
                )
                flash(message, "danger")
            else:
                flash("Назначение обновлено.", "success")
                return redirect(
                    url_for("workload.need_detail", need_id=need.id)
                )
        return _render_assignment_form(need, assignment)

    @workload_bp.post("/assignments/<int:assignment_id>/cancel")
    @login_required
    def assignment_cancel(assignment_id):
        assignment = _get_assignment(assignment_id, for_update=True)
        try:
            cancel_assignment(
                assignment,
                user_id=current_user.id,
                expected_revision=request.form.get("revision", type=int),
                reason=request.form.get("reason"),
            )
            db.session.commit()
        except WorkloadDistributionError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            flash("Назначение отменено.", "success")
        return redirect(
            url_for(
                "workload.need_detail",
                need_id=assignment.workload_need_id,
            )
        )

    @workload_bp.get("/teachers/")
    @login_required
    def workload_teachers():
        _require_assignments_read()
        need_ids = [
            item.id
            for item in _scoped_need_query()
            .filter(WorkloadNeed.status != "CANCELLED")
            .all()
        ]
        assignments = (
            WorkloadAssignment.query
            .filter(
                WorkloadAssignment.workload_need_id.in_(need_ids),
                WorkloadAssignment.status != "CANCELLED",
            )
            .all()
            if need_ids else []
        )
        return render_template(
            "workload/teachers.html",
            rows=_teacher_rows(assignments),
        )

    @workload_bp.get("/assignments/workspace")
    @login_required
    def assignment_workspace():
        _require_assignments_read()
        view_mode = (request.args.get("view") or "all").strip().lower()
        if view_mode not in {"all", "department", "vacancies"}:
            view_mode = "all"
        department_id = (
            request.args.get("department_id", type=int)
            if view_mode == "department"
            else None
        )
        building_id = request.args.get("building_id", type=int)
        education_levels = _workspace_selected_levels(request.args)
        grades = _workspace_selected_grades(request.args)
        selected_subject_ids = _workspace_selected_subject_ids(request.args)
        teacher_query = " ".join(
            (request.args.get("teacher_query") or "").split()
        )[:160]
        presentation_mode = (
            "list"
            if (request.args.get("presentation") or "").strip().lower()
            == "list"
            else "matrix"
        )
        fragment_holder_key = (
            request.args.get("fragment_holder_key") or ""
        ).strip()[:80]
        holder_page = max(1, request.args.get("holder_page", type=int) or 1)
        holder_page_size = request.args.get("holder_page_size", type=int)
        if holder_page_size not in WORKSPACE_HOLDER_PAGE_SIZES:
            holder_page_size = WORKSPACE_DEFAULT_HOLDER_PAGE_SIZE
        fragment_type, separator, fragment_value = (
            fragment_holder_key.partition(":")
        )
        if (
            separator != ":"
            or fragment_type not in {"teacher", "vacancy"}
            or not fragment_value
            or not fragment_value.replace("_", "").isalnum()
        ):
            fragment_holder_key = ""
        if fragment_holder_key:
            presentation_mode = "matrix"
        page_fragment = (
            request.args.get("page_fragment") == "1"
            and not fragment_holder_key
            and presentation_mode == "matrix"
        )
        versions = _draft_versions_query().all()
        version_id = request.args.get("version_id", type=int)
        if version_id is None and versions:
            version_id = versions[0].id
        selected_version = next(
            (item for item in versions if item.id == version_id),
            None,
        )
        can_update = (
            is_feature_enabled(WORKLOAD_WRITE)
            and can_use_workload_permission(
                "workload.assignments.update",
                current_user,
            )
            and selected_version is not None
            and selected_version.workload_approval_status == "EDITING"
        )
        snapshot, plans, plan_matrices = _workspace_plan_context(
            selected_version,
            education_levels=education_levels,
            grades=grades,
            building_id=building_id,
            include_matrices=False,
        )
        if can_update and selected_version is not None:
            try:
                _ensure_workspace_plan_needs(
                    selected_version,
                    snapshot,
                    plans,
                    None,
                )
                relinked_assignments = (
                    relink_assignments_to_population_snapshot(
                        selected_version,
                        snapshot,
                        user_id=current_user.id,
                    )
                )
                if (
                    relinked_assignments.transferred
                    or relinked_assignments.cancelled
                ):
                    db.session.commit()
                    message_parts = []
                    if relinked_assignments.transferred:
                        message_parts.append(
                            "перенесено назначений: "
                            f"{relinked_assignments.transferred}"
                        )
                    if relinked_assignments.cancelled:
                        message_parts.append(
                            "снято в изменённых ячейках: "
                            f"{relinked_assignments.cancelled}"
                        )
                    flash(
                        "Нагрузка синхронизирована; "
                        + "; ".join(message_parts)
                        + (
                            ". Снятые часы необходимо распределить заново."
                            if relinked_assignments.cancelled else "."
                        ),
                        (
                            "warning"
                            if relinked_assignments.cancelled
                            else "success"
                        ),
                    )
            except WorkloadDistributionError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
        state_key, state = _workspace_state(version_id or 0)
        fragment_activity_ids = set()
        if fragment_holder_key and selected_version is not None:
            fragment_activity_ids.update(
                row.get("activity_id")
                for row in state["rows"]
                if row.get("activity_id")
                and (
                    (
                        fragment_type == "teacher"
                        and row.get("holder_type", "teacher") == "teacher"
                        and str(row.get("teacher_id") or "") == fragment_value
                    )
                    or (
                        fragment_type == "vacancy"
                        and row.get("holder_type") == "vacancy"
                        and str(row.get("vacancy_key") or "") == fragment_value
                    )
                )
            )
            fragment_assignment_activities = (
                db.session.query(WorkloadNeed.education_activity_id)
                .join(
                    WorkloadAssignment,
                    WorkloadAssignment.workload_need_id == WorkloadNeed.id,
                )
                .filter(
                    WorkloadNeed.tariff_version_id == selected_version.id,
                    WorkloadAssignment.status != "CANCELLED",
                )
            )
            if fragment_type == "teacher" and fragment_value.isdigit():
                fragment_assignment_activities = (
                    fragment_assignment_activities.filter(
                        WorkloadAssignment.employee_user_id
                        == int(fragment_value),
                        WorkloadAssignment.assignment_kind != "VACANCY",
                    )
                )
            else:
                fragment_assignment_activities = (
                    fragment_assignment_activities.filter(
                        WorkloadAssignment.assignment_kind == "VACANCY",
                        WorkloadAssignment.position_code == fragment_value,
                    )
                )
            fragment_activity_ids.update(
                activity_id
                for (activity_id,) in (
                    fragment_assignment_activities.distinct().all()
                )
            )
        query = _scoped_need_query().filter(
            WorkloadNeed.status.in_(
                ("OPEN", "PARTIAL", "COVERED", "OVERALLOCATED")
            ),
        )
        if selected_version is not None:
            query = query.filter(
                WorkloadNeed.tariff_version_id == selected_version.id
            )
        effective_subject_ids = set(selected_subject_ids)
        if fragment_holder_key:
            effective_subject_ids = (
                effective_subject_ids.intersection(fragment_activity_ids)
                if effective_subject_ids else fragment_activity_ids
            )
        # Department view first determines holder membership from every
        # curriculum assignment in the version.  Subject filtering is applied
        # only afterwards to the displayed workload.
        if effective_subject_ids and department_id is None:
            query = query.filter(
                WorkloadNeed.education_activity_id.in_(
                    effective_subject_ids
                )
            )
        # A newly added teacher has no subject rows yet. Keep the needs for
        # this fragment so the class columns and the empty holder row can be
        # rendered. Filtering to false made the fragment markers disappear
        # and caused a misleading 404 after a successful teacher addition.
        if plan_matrices and (
            education_levels or grades or building_id is not None
        ):
            visible_group_ids = _workspace_matrix_group_ids(plan_matrices)
            query = (
                query.filter(
                    WorkloadNeed.teaching_group_id.in_(visible_group_ids)
                )
                if visible_group_ids else query.filter(db.false())
            )
        version_needs = query.order_by(
            WorkloadNeed.status.asc(),
            WorkloadNeed.education_activity_id.asc(),
        ).all()
        version_needs_by_id = {item.id: item for item in version_needs}
        needs = _filter_workspace_needs(
            version_needs,
            department_id=None,
            building_id=building_id,
            education_levels=education_levels,
            grades=grades,
            subject_ids=selected_subject_ids,
            population_snapshot_id=(snapshot.id if snapshot else None),
            bound_plan_ids_by_class=_workspace_bound_plan_ids(snapshot),
        )
        all_assignments_query = (
            WorkloadAssignment.query
            .join(
                WorkloadNeed,
                WorkloadNeed.id == WorkloadAssignment.workload_need_id,
            )
            .options(
                joinedload(WorkloadAssignment.employee),
            )
            .filter(
                WorkloadNeed.tariff_version_id == (
                    selected_version.id if selected_version is not None else -1
                ),
                WorkloadAssignment.status != "CANCELLED",
            )
        )
        assignment_scope = resolve_workload_scope(current_user)
        if not assignment_scope.unrestricted:
            if assignment_scope.department_ids:
                all_assignments_query = all_assignments_query.filter(
                    WorkloadNeed.department_id.in_(
                        assignment_scope.department_ids
                    )
                )
            if assignment_scope.building_ids:
                all_assignments_query = all_assignments_query.filter(
                    WorkloadNeed.building_id.in_(
                        assignment_scope.building_ids
                    )
                )
            if (
                not assignment_scope.department_ids
                and not assignment_scope.building_ids
            ):
                all_assignments_query = all_assignments_query.filter(
                    db.false()
                )
        all_assignments = (
            all_assignments_query.all()
            if selected_version is not None else []
        )
        activities = {
            item.id: item
            for item in EducationActivity.query.options(
                selectinload(EducationActivity.department_links)
            ).filter(
                EducationActivity.id.in_({
                    row.get("activity_id")
                    for row in state["rows"]
                    if row.get("activity_id")
                })
            ).all()
        } if state["rows"] else {}
        department_holder_keys = _workspace_department_holder_keys(
            all_assignments,
            version_needs_by_id,
            department_id,
        )
        if department_holder_keys is not None:
            department_holder_keys.update(
                _workspace_state_department_holder_keys(
                    state["rows"],
                    activities,
                    department_id,
                )
            )
            persisted_holder_keys = {
                _workspace_assignment_holder_key(assignment)
                for assignment in all_assignments
            }
            draft_holder_keys = {
                (
                    f"vacancy:{row.get('vacancy_key')}"
                    if row.get("holder_type") == "vacancy"
                    else f"teacher:{row.get('teacher_id')}"
                )
                for row in state["rows"]
            }
            # Keep a just-added empty holder visible long enough to choose its
            # first curriculum subject.  Holders that already have persisted
            # or draft rows must satisfy the curriculum membership rule.
            department_holder_keys.update(
                f"teacher:{teacher_id}"
                for teacher_id in state["teacher_ids"]
                if f"teacher:{teacher_id}" not in persisted_holder_keys
                and f"teacher:{teacher_id}" not in draft_holder_keys
            )
            department_holder_keys.update(
                f"vacancy:{vacancy['key']}"
                for vacancy in state["vacancies"]
                if vacancy.get("key")
                and f"vacancy:{vacancy['key']}"
                not in persisted_holder_keys
                and f"vacancy:{vacancy['key']}" not in draft_holder_keys
            )
        need_ids = {item.id for item in needs}
        needs_by_id = {item.id: item for item in needs}
        assignments = [
            item
            for item in all_assignments
            if item.workload_need_id in need_ids
            and (
                department_holder_keys is None
                or _workspace_assignment_holder_key(item)
                in department_holder_keys
            )
        ]
        for assignment in assignments:
            set_committed_value(
                assignment,
                "workload_need",
                needs_by_id[assignment.workload_need_id],
            )
        holder_totals = defaultdict(Decimal)
        for assignment in all_assignments:
            holder_key = (
                ("vacancy", assignment.position_code)
                if assignment.assignment_kind == "VACANCY"
                else ("teacher", assignment.employee_user_id)
            )
            holder_totals[holder_key] += Decimal(
                assignment.weekly_hours or ZERO
            )
        employees = _active_employees()
        employee_by_id = {item.id: item for item in employees}
        extra_teacher_ids = list(dict.fromkeys(state["teacher_ids"]))
        vacancies_by_key = {
            item["key"]: item
            for item in state["vacancies"]
            if item.get("key")
        }
        for assignment in assignments:
            if assignment.assignment_kind != "VACANCY":
                continue
            vacancies_by_key.setdefault(assignment.position_code, {
                "key": assignment.position_code,
                "label": assignment.position_title or "Вакансия",
            })
        if len(vacancies_by_key) != len(state["vacancies"]):
            state["vacancies"] = list(vacancies_by_key.values())
            _save_workspace_state(state_key, state)
        extra_vacancies = list(vacancies_by_key.values())
        subject_options = _workspace_plan_subject_options(
            plan_matrices,
            department_id,
        ) or _workspace_subject_options(needs)
        allowed_subject_keys = {
            (item["activity"].id, item["plan_kind"])
            for item in subject_options
        }
        extra_teachers = [
            employee_by_id[teacher_id]
            for teacher_id in extra_teacher_ids
            if teacher_id in employee_by_id
            and (
                department_holder_keys is None
                or f"teacher:{teacher_id}" in department_holder_keys
            )
        ]
        extra_vacancies = [
            vacancy
            for vacancy in extra_vacancies
            if department_holder_keys is None
            or f"vacancy:{vacancy['key']}" in department_holder_keys
        ]
        draft_rows = [
            (
                employee_by_id[row["teacher_id"]],
                activities[row["activity_id"]],
                row["plan_kind"],
            )
            for row in state["rows"]
            if row.get("holder_type", "teacher") == "teacher"
            and row.get("teacher_id") in employee_by_id
            and row.get("activity_id") in activities
            and (
                row.get("activity_id"),
                row.get("plan_kind"),
            ) in allowed_subject_keys
            and (
                not selected_subject_ids
                or row.get("activity_id") in selected_subject_ids
            )
            and (
                department_holder_keys is None
                or f"teacher:{row['teacher_id']}"
                in department_holder_keys
            )
        ]
        draft_vacancy_rows = [
            (
                row["vacancy_key"],
                activities[row["activity_id"]],
                row["plan_kind"],
            )
            for row in state["rows"]
            if row.get("holder_type") == "vacancy"
            and row.get("vacancy_key") in vacancies_by_key
            and row.get("activity_id") in activities
            and (
                row.get("activity_id"),
                row.get("plan_kind"),
            ) in allowed_subject_keys
            and (
                not selected_subject_ids
                or row.get("activity_id") in selected_subject_ids
            )
            and (
                department_holder_keys is None
                or f"vacancy:{row['vacancy_key']}"
                in department_holder_keys
            )
        ]
        holder_labels = {}
        for assignment in assignments:
            if assignment.assignment_kind == "VACANCY":
                holder_key = f"vacancy:{assignment.position_code}"
                holder_labels[holder_key] = (
                    assignment.position_title or "Вакансия"
                )
            elif assignment.employee is not None:
                holder_key = f"teacher:{assignment.employee_user_id}"
                holder_labels[holder_key] = assignment.employee.fio
            else:
                continue
        for teacher, _, _ in draft_rows:
            holder_key = f"teacher:{teacher.id}"
            holder_labels[holder_key] = teacher.fio
        for vacancy_key, _, _ in draft_vacancy_rows:
            holder_key = f"vacancy:{vacancy_key}"
            holder_labels[holder_key] = vacancies_by_key.get(
                vacancy_key,
                {"label": "Вакансия"},
            )["label"]
        if not selected_subject_ids:
            for teacher in extra_teachers:
                holder_labels.setdefault(
                    f"teacher:{teacher.id}",
                    teacher.fio,
                )
            for vacancy in extra_vacancies:
                holder_labels.setdefault(
                    f"vacancy:{vacancy['key']}",
                    vacancy["label"],
                )
        if presentation_mode == "list":
            assigned_holder_keys = {
                (
                    f"vacancy:{assignment.position_code}"
                    if assignment.assignment_kind == "VACANCY"
                    else f"teacher:{assignment.employee_user_id}"
                )
                for assignment in assignments
            }
            holder_labels = {
                key: label
                for key, label in holder_labels.items()
                if key in assigned_holder_keys
            }
        if view_mode == "vacancies":
            holder_labels = {
                key: label
                for key, label in holder_labels.items()
                if key.startswith("vacancy:")
            }
        holder_filter_options = [
            {
                "key": key,
                "label": label,
                "is_vacancy": key.startswith("vacancy:"),
            }
            for key, label in sorted(
                holder_labels.items(),
                key=lambda item: item[1].casefold(),
            )
        ]
        normalized_teacher_query = _workspace_search_text(teacher_query)
        if normalized_teacher_query:
            holder_labels = {
                key: label
                for key, label in holder_labels.items()
                if normalized_teacher_query in _workspace_search_text(label)
            }
        matching_holder_keys = set(holder_labels)
        ordered_holder_keys = [
            key for key, _ in sorted(
                holder_labels.items(),
                key=lambda item: item[1].casefold(),
            )
        ]
        holder_total_count = len(ordered_holder_keys)
        holder_page_count = max(
            1,
            (
                holder_total_count + holder_page_size - 1
            ) // holder_page_size,
        )
        holder_page = min(holder_page, holder_page_count)
        holder_page_start = (holder_page - 1) * holder_page_size
        page_holder_keys = set(ordered_holder_keys[
            holder_page_start:
            holder_page_start + holder_page_size
        ])
        matrix_teachers = {
            assignment.employee_user_id: assignment.employee
            for assignment in assignments
            if assignment.employee_user_id is not None
        }
        matrix_teachers.update({
            item.id: item
            for item in extra_teachers
        })
        if fragment_holder_key:
            visible_teacher_id = (
                int(fragment_value)
                if fragment_type == "teacher" and fragment_value.isdigit()
                else None
            )
            matrix_teachers = {
                teacher_id: teacher
                for teacher_id, teacher in matrix_teachers.items()
                if teacher_id == visible_teacher_id
            }
        else:
            visible_teacher_ids = {
                int(key.partition(":")[2])
                for key in page_holder_keys
                if key.startswith("teacher:")
                and key.partition(":")[2].isdigit()
            }
            matrix_teachers = {
                teacher_id: teacher
                for teacher_id, teacher in matrix_teachers.items()
                if teacher_id in visible_teacher_ids
            }
        matrix = build_workload_assignment_matrix(
            needs,
            assignments,
            plan_matrices=plan_matrices,
            extra_teachers=extra_teachers,
            extra_vacancies=extra_vacancies,
            draft_rows=draft_rows,
            draft_vacancy_rows=draft_vacancy_rows,
            teacher_metadata=_workspace_teacher_metadata(
                list(matrix_teachers.values()),
                selected_version,
            ),
            total_assignments=assignments,
            holder_totals=holder_totals,
            visible_holder_key=fragment_holder_key or None,
            visible_holder_keys=(
                None if fragment_holder_key else page_holder_keys
            ),
        )
        if department_id is not None:
            # Department plan indicators and the unassigned-hours dialog stay
            # scoped to curriculum subjects of the selected department, even
            # though every workload row of matching holders is displayed.
            department_summary_need_ids = {
                need.id
                for need in needs
                if need_plan_kind(need) == "CURRICULUM"
                and _workspace_search_text(need.education_activity.name)
                not in DEPARTMENT_NEUTRAL_CURRICULUM_ACTIVITY_NAMES
                and need_matches_department(need, department_id)
            }
            department_summary_needs = [
                need
                for need in needs
                if need.id in department_summary_need_ids
            ]
            matrix["total_weekly"] = sum(
                (
                    Decimal(need.weekly_hours or ZERO)
                    for need in department_summary_needs
                ),
                ZERO,
            )
            matrix["total_allocated"] = sum(
                (
                    Decimal(assignment.weekly_hours or ZERO)
                    for assignment in all_assignments
                    if assignment.workload_need_id
                    in department_summary_need_ids
                ),
                ZERO,
            )
            matrix["total_remaining"] = (
                matrix["total_weekly"] - matrix["total_allocated"]
            )
            matrix["unassigned_items"] = [
                item
                for item in matrix["unassigned_items"]
                if item["need"].id in department_summary_need_ids
            ]
        if selected_subject_ids:
            matrix["blocks"] = [
                block for block in matrix["blocks"]
                if any(
                    row.get("activity") is not None
                    and row["activity"].id in selected_subject_ids
                    for row in block["rows"]
                )
            ]
            matrix["teacher_count"] = len(matrix["blocks"])
        matrix["blocks"] = [
            block for block in matrix["blocks"]
            if (
                f"vacancy:{block['vacancy_key']}"
                if block["is_vacancy"]
                else f"teacher:{block['teacher_id']}"
            ) in matching_holder_keys
        ]
        if not fragment_holder_key:
            matrix["teacher_count"] = holder_total_count
        list_assignments = assignments
        if view_mode == "vacancies":
            list_assignments = [
                assignment for assignment in assignments
                if assignment.assignment_kind == "VACANCY"
            ]
        if normalized_teacher_query:
            list_assignments = [
                assignment for assignment in list_assignments
                if _workspace_assignment_holder_key(assignment)
                in matching_holder_keys
            ]
        workload_list = build_workload_assignment_list(
            list_assignments,
            visible_holder_keys=page_holder_keys,
        )
        workload_list_summary = build_workload_assignment_list(
            list_assignments
        )
        departments, buildings = _scope_options()
        filter_classes = list(snapshot.classes) if snapshot else []
        workspace_scope = resolve_workload_scope(current_user)
        if not workspace_scope.unrestricted:
            allowed_building_ids = set(workspace_scope.building_ids)
            filter_classes = [
                item for item in filter_classes
                if item.building_id in allowed_building_ids
            ]
        if building_id is not None:
            filter_classes = [
                item for item in filter_classes
                if item.building_id == building_id
            ]
        level_counts = {
            level: sum(
                1
                for item in filter_classes
                if item.grade_snapshot in grades
            )
            for level, grades in EDUCATION_LEVEL_GRADES.items()
        }
        level_grades = sorted(
            set().union(*(
                EDUCATION_LEVEL_GRADES[level]
                for level in education_levels
            ))
            if education_levels else set(range(1, 12))
        )
        totals = {
            "weekly": matrix["total_weekly"],
            "allocated": matrix["total_allocated"],
        }
        totals["remaining"] = totals["weekly"] - totals["allocated"]
        rendered = render_template(
            "workload/assignment_workspace.html",
            needs=needs,
            selected_version=selected_version,
            versions=versions,
            selected_version_id=version_id,
            matrix=matrix,
            totals=totals,
            departments=departments,
            buildings=buildings,
            view_mode=view_mode,
            selected_department_id=department_id,
            selected_building_id=building_id,
            selected_education_levels=sorted(education_levels),
            selected_grades=sorted(grades),
            selected_teacher_query=teacher_query,
            presentation_mode=presentation_mode,
            selected_subject_ids=sorted(selected_subject_ids),
            subject_filter_options=_workspace_subject_filter_options(
                selected_version
            ),
            holder_page=holder_page,
            holder_page_size=holder_page_size,
            holder_page_sizes=WORKSPACE_HOLDER_PAGE_SIZES,
            holder_page_count=holder_page_count,
            holder_total_count=holder_total_count,
            holder_page_from=(
                holder_page_start + 1 if holder_total_count else 0
            ),
            holder_page_to=min(
                holder_page_start + holder_page_size,
                holder_total_count,
            ),
            workload_list=workload_list,
            workload_list_summary=workload_list_summary,
            list_plan_kind_labels=LIST_PLAN_KIND_LABELS,
            holder_filter_options=holder_filter_options,
            level_counts=level_counts,
            level_labels=EDUCATION_LEVEL_LABELS,
            level_grades=level_grades,
            teacher_choices=_workspace_teacher_choices(
                employees,
                selected_version,
                department_id,
                assignments,
                excluded_teacher_ids={
                    assignment.employee_user_id
                    for assignment in all_assignments
                    if assignment.assignment_kind != "VACANCY"
                    and assignment.employee_user_id is not None
                } | set(state["teacher_ids"]) | {
                    row.get("teacher_id")
                    for row in state["rows"]
                    if row.get("holder_type", "teacher") == "teacher"
                    and row.get("teacher_id") is not None
                },
            ),
            replacement_choices=employees,
            copy_teacher_choices=[
                item
                for item in employees
                if item.id not in {
                    assignment.employee_user_id
                    for assignment in all_assignments
                    if assignment.assignment_kind != "VACANCY"
                    and assignment.employee_user_id is not None
                }
                and item.id not in {
                    row.get("teacher_id")
                    for row in state["rows"]
                    if row.get("holder_type", "teacher") == "teacher"
                    and row.get("teacher_id") is not None
                }
            ],
            subject_options=subject_options,
            need_status_labels=WORKLOAD_NEED_STATUS_LABELS,
            assignment_kind_labels=WORKLOAD_ASSIGNMENT_KIND_LABELS,
            can_update=can_update,
            can_manage_editing=(
                selected_version is not None
                and is_feature_enabled(WORKLOAD_WRITE)
                and can_use_workload_permission(
                    "workload.assignments.update",
                    current_user,
                )
            ),
            can_review_workload=_can_review_workload(),
            workload_approval_status_labels=(
                WORKLOAD_APPROVAL_STATUS_LABELS
            ),
            fragment_holder_key=fragment_holder_key,
        )
        if page_fragment:
            start_marker = "<!-- workload-page-region-start -->"
            end_marker = "<!-- workload-page-region-end -->"
            start = rendered.find(start_marker)
            end = rendered.find(end_marker)
            if start < 0 or end < start:
                abort(404)
            return rendered[start + len(start_marker):end]
        if fragment_holder_key:
            start_marker = (
                f"<!-- workload-holder-start:{fragment_holder_key} -->"
            )
            end_marker = (
                f"<!-- workload-holder-end:{fragment_holder_key} -->"
            )
            start = rendered.find(start_marker)
            end = rendered.find(end_marker)
            if start < 0 or end < start:
                abort(404)
            rows = rendered[start + len(start_marker):end]
            return (
                "<table><tbody data-workload-holder-fragment>"
                f"{rows}</tbody></table>"
            )
        return rendered

    @workload_bp.post("/assignments/workspace/status")
    @login_required
    def assignment_workspace_change_status():
        action = (request.form.get("action") or "").strip().upper()
        version = _draft_versions_query().filter(
            TariffVersion.id == request.form.get("version_id", type=int)
        ).first_or_404()
        if action in {"APPROVE", "REQUEST_CHANGES"}:
            if not _can_review_workload():
                abort(403)
        else:
            _require_assignments_update()
        try:
            change_workload_approval_status(
                version,
                action,
                user_id=current_user.id,
                comment=request.form.get("comment"),
            )
            db.session.commit()
        except WorkloadEditingWorkflowError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            messages = {
                "EDITING": "Редактирование нагрузки открыто.",
                "SAVED": (
                    "Изменения нагрузки сохранены. "
                    "Редактирование закрыто."
                ),
                "PENDING_APPROVAL": (
                    "Нагрузка отправлена директору на согласование."
                ),
                "APPROVED": "Нагрузка согласована директором.",
                "CHANGES_REQUESTED": (
                    "Нагрузка возвращена ответственным на исправление."
                ),
            }
            flash(
                messages[version.workload_approval_status],
                "success",
            )
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/teachers")
    @login_required
    def assignment_workspace_teacher_add():
        _require_assignments_update()
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        version_id = request.form.get("version_id", type=int)
        holder_type = (request.form.get("holder_type") or "teacher").strip()
        teacher_id = request.form.get("teacher_id", type=int)
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        _require_version_workload_editable(version)
        key, state = _workspace_state(version.id)
        if holder_type == "vacancy":
            vacancy = _next_vacancy(
                version.id,
                state,
                note=request.form.get("vacancy_note"),
            )
            state["vacancies"].append(vacancy)
            _save_workspace_state(key, state)
            if is_ajax:
                return jsonify({
                    "ok": True,
                    "holder_key": f"vacancy:{vacancy['key']}",
                    "holder_name": vacancy["label"],
                })
            return _workspace_redirect()
        teacher = db.session.get(User, teacher_id) if teacher_id else None
        if (
            teacher is None
            or not teacher.is_active_user
            or teacher.employment_status != "ACTIVE"
        ):
            return fail("Выберите работающего преподавателя.")
        if teacher.id not in state["teacher_ids"]:
            state["teacher_ids"].append(teacher.id)
            _save_workspace_state(key, state)
        if is_ajax:
            return jsonify({
                "ok": True,
                "holder_key": f"teacher:{teacher.id}",
                "teacher_id": teacher.id,
                "teacher_name": teacher.fio,
            })
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/vacancies/label")
    @login_required
    def assignment_workspace_vacancy_label():
        _require_assignments_update()
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        version_id = request.form.get("version_id", type=int)
        vacancy_key = (request.form.get("vacancy_key") or "").strip()
        prefix = "VACANCY_"
        number_text = (
            vacancy_key[len(prefix):]
            if vacancy_key.startswith(prefix) else ""
        )
        if not number_text.isdigit():
            return fail("Вакансия не найдена.")
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        _require_version_workload_editable(version)
        key, state = _workspace_state(version.id)
        state_vacancy = next(
            (
                item for item in state["vacancies"]
                if item.get("key") == vacancy_key
            ),
            None,
        )
        assignments = WorkloadAssignment.query.filter(
            WorkloadAssignment.tariff_version_id == version.id,
            WorkloadAssignment.assignment_kind == "VACANCY",
            WorkloadAssignment.position_code == vacancy_key,
            WorkloadAssignment.status != "CANCELLED",
        ).all()
        if state_vacancy is None and not assignments:
            return fail("Вакансия не найдена.")
        label = _vacancy_label(
            int(number_text),
            request.form.get("vacancy_note"),
        )
        try:
            for assignment in assignments:
                before = assignment_snapshot(assignment)
                assignment.position_title = label
                assignment.updated_by_user_id = current_user.id
                assignment.revision += 1
                add_assignment_change(
                    assignment,
                    "UPDATE",
                    user_id=current_user.id,
                    before_data=before,
                    reason="Изменена подпись вакансии",
                )
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return fail("Не удалось сохранить подпись вакансии.")
        if state_vacancy is None:
            state["vacancies"].append({
                "key": vacancy_key,
                "label": label,
            })
        else:
            state_vacancy["label"] = label
        _save_workspace_state(key, state)
        if is_ajax:
            return jsonify({
                "ok": True,
                "holder_key": f"vacancy:{vacancy_key}",
                "holder_name": label,
            })
        flash("Подпись вакансии сохранена.", "success")
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/subjects")
    @login_required
    def assignment_workspace_subject_add():
        _require_assignments_update()
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        version_id = request.form.get("version_id", type=int)
        holder_type = (request.form.get("holder_type") or "teacher").strip()
        teacher_id = request.form.get("teacher_id", type=int)
        vacancy_key = (request.form.get("vacancy_key") or "").strip()
        activity_id = request.form.get("activity_id", type=int)
        plan_kind = (request.form.get("plan_kind") or "").strip().upper()
        activity_plan_kind = (
            request.form.get("activity_plan_kind") or ""
        ).strip()
        if activity_plan_kind:
            activity_value, separator, kind_value = (
                activity_plan_kind.partition(":")
            )
            if separator:
                try:
                    activity_id = int(activity_value)
                except (TypeError, ValueError):
                    activity_id = None
                plan_kind = kind_value.strip().upper()
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        _require_version_workload_editable(version)
        teacher = db.session.get(User, teacher_id) if teacher_id else None
        activity = (
            db.session.get(EducationActivity, activity_id)
            if activity_id else None
        )
        key, state = _workspace_state(version.id)
        vacancy = next(
            (
                item for item in state["vacancies"]
                if item.get("key") == vacancy_key
            ),
            None,
        )
        if holder_type == "vacancy" and vacancy is None and vacancy_key:
            existing_vacancy = WorkloadAssignment.query.filter(
                WorkloadAssignment.tariff_version_id == version.id,
                WorkloadAssignment.assignment_kind == "VACANCY",
                WorkloadAssignment.position_code == vacancy_key,
                WorkloadAssignment.status != "CANCELLED",
            ).first()
            if existing_vacancy is not None:
                vacancy = {
                    "key": vacancy_key,
                    "label": existing_vacancy.position_title or "Вакансия",
                }
                state["vacancies"].append(vacancy)
        if activity is None or (
            holder_type == "vacancy" and vacancy is None
        ) or (
            holder_type != "vacancy" and teacher is None
        ):
            return fail("Выберите строку нагрузки и предмет.")
        if plan_kind not in PLAN_KIND_LABELS:
            return fail("Выберите допустимую часть учебного плана.")
        matching_needs = [
            need
            for need in _workspace_form_needs(
                version,
                activity_id=activity.id,
            )
            if (
                need.education_activity_id == activity.id
                and need_plan_kind(need) == plan_kind
            )
        ]
        has_need = next(
            iter(matching_needs),
            None,
        )
        if has_need is None:
            return fail("В выбранной версии нет часов по этому предмету.")
        if teacher is not None and teacher.id not in state["teacher_ids"]:
            state["teacher_ids"].append(teacher.id)
        row = {
            "holder_type": holder_type,
            "teacher_id": teacher.id if teacher is not None else None,
            "vacancy_key": vacancy_key if holder_type == "vacancy" else None,
            "activity_id": activity.id,
            "plan_kind": plan_kind,
        }
        if row not in state["rows"]:
            state["rows"].append(row)
        _save_workspace_state(key, state)
        if is_ajax:
            holder_key = (
                f"vacancy:{vacancy_key}"
                if holder_type == "vacancy"
                else f"teacher:{teacher.id}"
            )
            return jsonify({"ok": True, "holder_key": holder_key})
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/holder/copy-subjects")
    @login_required
    def assignment_workspace_holder_copy_subjects():
        _require_assignments_update()
        version_id = request.form.get("version_id", type=int)
        source_teacher_id = request.form.get("source_teacher_id", type=int)
        target_teacher_id = request.form.get("target_teacher_id", type=int)
        version = (
            _draft_versions_query()
            .filter(TariffVersion.id == version_id)
            .first_or_404()
        )
        _require_version_workload_editable(version)
        source_teacher = (
            db.session.get(User, source_teacher_id)
            if source_teacher_id else None
        )
        target_teacher = (
            db.session.get(User, target_teacher_id)
            if target_teacher_id else None
        )
        if source_teacher is None:
            flash("Исходный преподаватель не найден.", "danger")
            return _workspace_redirect()
        if (
            target_teacher is None
            or not target_teacher.is_active_user
            or target_teacher.employment_status != "ACTIVE"
        ):
            flash("Выберите работающего преподавателя.", "danger")
            return _workspace_redirect()
        if source_teacher.id == target_teacher.id:
            flash("Выбран тот же преподаватель.", "warning")
            return _workspace_redirect()

        key, state = _workspace_state(version.id)
        target_has_assignments = WorkloadAssignment.query.filter(
            WorkloadAssignment.tariff_version_id == version.id,
            WorkloadAssignment.status != "CANCELLED",
            WorkloadAssignment.assignment_kind != "VACANCY",
            WorkloadAssignment.employee_user_id == target_teacher.id,
        ).first()
        target_has_rows = any(
            row.get("holder_type", "teacher") == "teacher"
            and row.get("teacher_id") == target_teacher.id
            for row in state["rows"]
        )
        if target_has_assignments is not None or target_has_rows:
            flash(
                f"Копирование не выполнено: у {target_teacher.fio} уже "
                "есть нагрузка в выбранном учебном году.",
                "danger",
            )
            return _workspace_redirect()

        visible_needs = _workspace_form_needs(version)
        visible_need_ids = {need.id for need in visible_needs}
        source_keys = {
            (
                assignment.workload_need.education_activity_id,
                need_plan_kind(assignment.workload_need),
            )
            for assignment in WorkloadAssignment.query.filter(
                WorkloadAssignment.tariff_version_id == version.id,
                WorkloadAssignment.workload_need_id.in_(visible_need_ids),
                WorkloadAssignment.status != "CANCELLED",
                WorkloadAssignment.assignment_kind != "VACANCY",
                WorkloadAssignment.employee_user_id == source_teacher.id,
            ).all()
        } if visible_need_ids else set()
        allowed_keys = {
            (need.education_activity_id, need_plan_kind(need))
            for need in visible_needs
        }
        source_keys.update(
            (row.get("activity_id"), row.get("plan_kind"))
            for row in state["rows"]
            if row.get("holder_type", "teacher") == "teacher"
            and row.get("teacher_id") == source_teacher.id
            and (row.get("activity_id"), row.get("plan_kind"))
            in allowed_keys
        )
        source_keys.discard((None, None))
        if not source_keys:
            flash(
                f"У {source_teacher.fio} нет предметов для копирования "
                "в текущем отображении.",
                "warning",
            )
            return _workspace_redirect()

        if target_teacher.id not in state["teacher_ids"]:
            state["teacher_ids"].append(target_teacher.id)
        for activity_id, plan_kind in sorted(source_keys):
            row = {
                "holder_type": "teacher",
                "teacher_id": target_teacher.id,
                "vacancy_key": None,
                "activity_id": activity_id,
                "plan_kind": plan_kind,
            }
            if row not in state["rows"]:
                state["rows"].append(row)
        _save_workspace_state(key, state)
        flash(
            f"Набор из {len(source_keys)} предметов скопирован: "
            f"{source_teacher.fio} → {target_teacher.fio}. "
            "Часы можно назначить в ячейках новой строки.",
            "success",
        )
        return _workspace_redirect(
            teacher_query=target_teacher.fio,
            focus_holder=f"teacher:{target_teacher.id}",
        )

    @workload_bp.post("/assignments/workspace/subjects/delete")
    @login_required
    def assignment_workspace_subject_delete():
        _require_assignments_update()
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        version_id = request.form.get("version_id", type=int)
        holder_type = (
            request.form.get("holder_type") or "teacher"
        ).strip()
        teacher_id = request.form.get("teacher_id", type=int)
        vacancy_key = (
            request.form.get("vacancy_key") or ""
        ).strip()
        activity_id = request.form.get("activity_id", type=int)
        plan_kind = (
            request.form.get("plan_kind") or ""
        ).strip().upper()
        version = (
            _draft_versions_query()
            .filter(TariffVersion.id == version_id)
            .first_or_404()
        )
        _require_version_workload_editable(version)
        activity = db.session.get(EducationActivity, activity_id)
        if (
            holder_type not in {"teacher", "vacancy"}
            or activity is None
            or plan_kind not in PLAN_KIND_LABELS
            or (holder_type == "teacher" and teacher_id is None)
            or (holder_type == "vacancy" and not vacancy_key)
        ):
            return fail("Строка предмета не найдена.")

        needs = [
            need
            for need in _workspace_form_needs(
                version,
                activity_id=activity.id,
            )
            if (
                need.education_activity_id == activity.id
                and need_plan_kind(need) == plan_kind
            )
        ]
        need_ids = [need.id for need in needs]
        query = WorkloadAssignment.query.filter(
            WorkloadAssignment.tariff_version_id == version.id,
            WorkloadAssignment.workload_need_id.in_(need_ids),
            WorkloadAssignment.status != "CANCELLED",
        )
        if holder_type == "vacancy":
            query = query.filter(
                WorkloadAssignment.assignment_kind == "VACANCY",
                WorkloadAssignment.position_code == vacancy_key,
            )
        else:
            query = query.filter(
                WorkloadAssignment.assignment_kind != "VACANCY",
                WorkloadAssignment.employee_user_id == teacher_id,
            )
        assignments = query.all() if need_ids else []
        key, state = _workspace_state(version.id)
        state = {
            **state,
            "teacher_ids": list(state["teacher_ids"]),
            "vacancies": list(state["vacancies"]),
            "rows": list(state["rows"]),
        }
        rows_before = len(state["rows"])
        state["rows"] = [
            row
            for row in state["rows"]
            if not (
                _state_row_matches_holder(
                    row,
                    holder_type,
                    teacher_id,
                    vacancy_key,
                )
                and row.get("activity_id") == activity.id
                and row.get("plan_kind") == plan_kind
            )
        ]
        removed_draft_rows = rows_before - len(state["rows"])
        try:
            for assignment in assignments:
                cancel_assignment(
                    assignment,
                    user_id=current_user.id,
                    expected_revision=assignment.revision,
                    reason=(
                        "Удалена предметная строка из матрицы "
                        "распределения нагрузки"
                    ),
                )
            db.session.commit()
        except WorkloadDistributionError as exc:
            db.session.rollback()
            return fail(str(exc))
        else:
            _save_workspace_state(key, state)
            holder_key = (
                f"vacancy:{vacancy_key}"
                if holder_type == "vacancy"
                else f"teacher:{teacher_id}"
            )
            if assignments or removed_draft_rows:
                message = (
                    f"{activity.name}: строка удалена, "
                    f"освобождено назначений — {len(assignments)}."
                )
            else:
                message = "Строка предмета уже удалена."
            if is_ajax:
                return jsonify({
                    "ok": True,
                    "holder_key": holder_key,
                    "message": message,
                })
            flash(message, "success" if assignments or removed_draft_rows else "info")
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/holders/delete")
    @login_required
    def assignment_workspace_holder_delete():
        _require_assignments_update()
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        version_id = request.form.get("version_id", type=int)
        holder_type = (
            request.form.get("holder_type") or "teacher"
        ).strip()
        teacher_id = request.form.get("teacher_id", type=int)
        vacancy_key = (
            request.form.get("vacancy_key") or ""
        ).strip()
        version = (
            _draft_versions_query()
            .filter(TariffVersion.id == version_id)
            .first_or_404()
        )
        _require_version_workload_editable(version)
        if (
            holder_type not in {"teacher", "vacancy"}
            or (holder_type == "teacher" and teacher_id is None)
            or (holder_type == "vacancy" and not vacancy_key)
        ):
            return fail("Строка преподавателя не найдена.")

        need_ids = [
            need.id for need in _workspace_form_needs(version)
        ]
        query = WorkloadAssignment.query.filter(
            WorkloadAssignment.tariff_version_id == version.id,
            WorkloadAssignment.workload_need_id.in_(need_ids),
            WorkloadAssignment.status != "CANCELLED",
        )
        if holder_type == "vacancy":
            query = query.filter(
                WorkloadAssignment.assignment_kind == "VACANCY",
                WorkloadAssignment.position_code == vacancy_key,
            )
            holder_label = next(
                (
                    item.position_title for item in query.all()
                    if item.position_title
                ),
                "Вакансия",
            )
            assignments = query.all() if need_ids else []
        else:
            query = query.filter(
                WorkloadAssignment.assignment_kind != "VACANCY",
                WorkloadAssignment.employee_user_id == teacher_id,
            )
            assignments = query.all() if need_ids else []
            teacher = db.session.get(User, teacher_id)
            holder_label = (
                teacher.fio if teacher is not None else "Преподаватель"
            )

        released_weekly_hours = sum(
            (Decimal(item.weekly_hours or ZERO) for item in assignments),
            ZERO,
        )

        key, state = _workspace_state(version.id)
        state = {
            **state,
            "teacher_ids": list(state["teacher_ids"]),
            "vacancies": list(state["vacancies"]),
            "rows": list(state["rows"]),
        }
        state["rows"] = [
            row
            for row in state["rows"]
            if not _state_row_matches_holder(
                row,
                holder_type,
                teacher_id,
                vacancy_key,
            )
        ]
        if holder_type == "vacancy":
            state["vacancies"] = [
                item for item in state["vacancies"]
                if item.get("key") != vacancy_key
            ]
        else:
            state["teacher_ids"] = [
                item for item in state["teacher_ids"]
                if item != teacher_id
            ]
        try:
            for assignment in assignments:
                cancel_assignment(
                    assignment,
                    user_id=current_user.id,
                    expected_revision=assignment.revision,
                    reason=(
                        "Удалена строка преподавателя из матрицы "
                        "распределения нагрузки"
                    ),
                )
            db.session.commit()
        except WorkloadDistributionError as exc:
            db.session.rollback()
            return fail(str(exc))
        else:
            _save_workspace_state(key, state)
            message = (
                f"{holder_label}: строка удалена, "
                f"освобождено назначений — {len(assignments)}."
            )
            if is_ajax:
                payload = {
                    "ok": True,
                    "holder_key": (
                        f"vacancy:{vacancy_key}"
                        if holder_type == "vacancy"
                        else f"teacher:{teacher_id}"
                    ),
                    "holder_name": holder_label,
                    "released_weekly_hours": float(released_weekly_hours),
                    "message": message,
                }
                if (
                    holder_type == "teacher"
                    and teacher is not None
                    and teacher.is_active_user
                    and teacher.employment_status == "ACTIVE"
                ):
                    payload["teacher"] = {
                        "id": teacher.id,
                        "name": teacher.fio,
                    }
                return jsonify(payload)
            flash(message, "success")
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/assign")
    @login_required
    def assignment_workspace_assign():
        need_id = request.form.get("need_id", type=int)
        teacher_id = request.form.get("teacher_id", type=int)
        need = _get_need(need_id, for_update=True)
        teacher = db.session.get(User, teacher_id) if teacher_id else None
        if teacher is None:
            flash("Выберите преподавателя.", "danger")
            return _workspace_redirect()
        weekly = Decimal(need.remaining_weekly_hours or ZERO)
        annual = Decimal(need.remaining_annual_hours or ZERO)
        if weekly <= ZERO:
            flash("Эти часы уже распределены.", "warning")
            return _workspace_redirect()
        assignment = WorkloadAssignment(
            organization_id=need.organization_id,
            tariff_version_id=need.tariff_version_id,
            workload_need_id=need.id,
            employee_user_id=teacher.id,
            position_code="TEACHER",
            position_title="Учитель",
            department_id=(
                request.form.get("department_id", type=int)
                or need.department_id
            ),
            building_id=need.building_id,
            assignment_kind="MAIN",
            date_from=need.date_from,
            date_to=need.date_to,
            weekly_hours=weekly,
            annual_hours=annual,
            status="DRAFT",
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
        )
        try:
            validate_assignment(need, assignment)
            db.session.add(assignment)
            db.session.flush()
            add_assignment_change(
                assignment,
                "CREATE",
                user_id=current_user.id,
            )
            refresh_need_status(need)
            db.session.commit()
        except (WorkloadDistributionError, IntegrityError) as exc:
            db.session.rollback()
            flash(
                str(exc)
                if isinstance(exc, WorkloadDistributionError)
                else "Не удалось назначить преподавателя.",
                "danger",
            )
        else:
            flash(
                f"{need.education_activity.name}: "
                f"{teacher.fio} назначен на полный объём.",
                "success",
            )
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/cell")
    @login_required
    def assignment_workspace_cell_update():
        _require_assignments_update()
        need_id = request.form.get("need_id", type=int)
        holder_type = (request.form.get("holder_type") or "teacher").strip()
        teacher_id = request.form.get("teacher_id", type=int)
        vacancy_key = (request.form.get("vacancy_key") or "").strip()
        need = _get_need(need_id, for_update=True)
        teacher = db.session.get(User, teacher_id) if teacher_id else None
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

        def fail(message):
            if is_ajax:
                return jsonify({"ok": False, "message": message}), 422
            flash(message, "danger")
            return _workspace_redirect()

        if holder_type == "vacancy":
            state_key, state = _workspace_state(need.tariff_version_id)
            vacancy = next(
                (
                    item for item in state["vacancies"]
                    if item.get("key") == vacancy_key
                ),
                None,
            )
            if vacancy is None:
                existing = WorkloadAssignment.query.filter(
                    WorkloadAssignment.tariff_version_id == need.tariff_version_id,
                    WorkloadAssignment.assignment_kind == "VACANCY",
                    WorkloadAssignment.position_code == vacancy_key,
                    WorkloadAssignment.status != "CANCELLED",
                ).first()
                vacancy = {
                    "key": vacancy_key,
                    "label": existing.position_title,
                } if existing is not None else None
            if vacancy is None:
                return fail("Строка вакансии не найдена.")
        elif teacher is None or not teacher.is_active_user:
            return fail("Преподаватель не найден.")
        raw_hours = (request.form.get("hours") or "").strip()
        active_assignments = [
            item for item in need.active_assignments
        ]
        if holder_type == "vacancy":
            own_assignments = [
                item for item in active_assignments
                if item.assignment_kind == "VACANCY"
                and item.position_code == vacancy_key
            ]
        else:
            own_assignments = [
                item for item in active_assignments
                if item.assignment_kind != "VACANCY"
                and item.employee_user_id == teacher.id
            ]
        other_assignments = [
            item for item in active_assignments
            if item not in own_assignments
        ]
        previous_value = sum(
            (
                Decimal(item.weekly_hours or ZERO)
                for item in own_assignments
            ),
            ZERO,
        )
        try:
            if not raw_hours or raw_hours.replace(",", ".") in {"0", "0.0"}:
                for assignment in own_assignments:
                    cancel_assignment(
                        assignment,
                        user_id=current_user.id,
                        expected_revision=assignment.revision,
                        reason="Снято из матрицы распределения нагрузки",
                    )
                refresh_need_status(need)
                value = None
            else:
                hours = decimal_hours(raw_hours, "часов в неделю")
                planned = Decimal(need.weekly_hours or ZERO)
                if hours != planned:
                    raise WorkloadDistributionError(
                        "Для этой группы необходимо назначить полный объём: "
                        f"{planned.normalize()} ч/нед."
                    )
                if other_assignments:
                    raise WorkloadDistributionError(
                        "Эта группа уже назначена другому преподавателю."
                    )
                if own_assignments:
                    assignment = own_assignments[0]
                    assignment.weekly_hours = planned
                    assignment.annual_hours = calculate_assignment_annual_hours(
                        need,
                        planned,
                        None,
                    )
                    assignment.updated_by_user_id = current_user.id
                else:
                    assignment = WorkloadAssignment(
                        organization_id=need.organization_id,
                        tariff_version_id=need.tariff_version_id,
                        workload_need_id=need.id,
                        employee_user_id=(
                            None if holder_type == "vacancy" else teacher.id
                        ),
                        position_code=(
                            vacancy_key
                            if holder_type == "vacancy" else "TEACHER"
                        ),
                        position_title=(
                            vacancy["label"]
                            if holder_type == "vacancy" else "Учитель"
                        ),
                        department_id=need.department_id,
                        building_id=need.building_id,
                        assignment_kind=(
                            "VACANCY" if holder_type == "vacancy" else "MAIN"
                        ),
                        date_from=need.date_from,
                        date_to=need.date_to,
                        weekly_hours=planned,
                        annual_hours=calculate_assignment_annual_hours(
                            need,
                            planned,
                            None,
                        ),
                        status="DRAFT",
                        created_by_user_id=current_user.id,
                        updated_by_user_id=current_user.id,
                    )
                    validate_assignment(need, assignment)
                    db.session.add(assignment)
                    db.session.flush()
                    add_assignment_change(
                        assignment,
                        "CREATE",
                        user_id=current_user.id,
                    )
                refresh_need_status(need)
                value = planned
            db.session.commit()
        except (WorkloadDistributionError, IntegrityError) as exc:
            db.session.rollback()
            return fail(
                str(exc)
                if isinstance(exc, WorkloadDistributionError)
                else "Не удалось сохранить нагрузку."
            )

        if holder_type == "vacancy":
            holder_key = f"vacancy:{vacancy_key}"
        else:
            holder_key = f"teacher:{teacher.id}"
        plan_kind = need_plan_kind(need)
        holder_delta = Decimal(value or ZERO) - previous_value
        allocated_delta = holder_delta
        if is_ajax:
            return jsonify({
                "ok": True,
                "need_id": need.id,
                "holder_key": holder_key,
                "value": float(value) if value is not None else None,
                "activity_id": need.education_activity_id,
                "plan_kind": plan_kind,
                "holder_delta": float(holder_delta),
                "allocated_delta": float(allocated_delta),
            })
        flash("Нагрузка сохранена.", "success")
        return _workspace_redirect()

    @workload_bp.post(
        "/assignments/workspace/<int:assignment_id>/cancel"
    )
    @login_required
    def assignment_workspace_cancel(assignment_id):
        assignment = _get_assignment(assignment_id, for_update=True)
        try:
            cancel_assignment(
                assignment,
                user_id=current_user.id,
                expected_revision=assignment.revision,
                reason="Снято в матрице распределения нагрузки",
            )
            db.session.commit()
        except WorkloadDistributionError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            flash("Назначение снято. Часы снова доступны.", "success")
        return _workspace_redirect()

    @workload_bp.post("/assignments/workspace/holder/replace")
    @login_required
    def assignment_workspace_holder_replace():
        _require_assignments_update()
        version_id = request.form.get("version_id", type=int)
        source_type = (request.form.get("source_type") or "teacher").strip()
        source_teacher_id = request.form.get("source_teacher_id", type=int)
        source_vacancy_key = (
            request.form.get("source_vacancy_key") or ""
        ).strip()
        target_value = (request.form.get("target_holder") or "").strip()
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        _require_version_workload_editable(version)

        target_teacher = None
        target_vacancy = None
        if target_value.startswith("teacher:"):
            try:
                target_teacher_id = int(target_value.split(":", 1)[1])
            except (TypeError, ValueError):
                target_teacher_id = None
            target_teacher = (
                db.session.get(User, target_teacher_id)
                if target_teacher_id else None
            )
            if (
                target_teacher is None
                or not target_teacher.is_active_user
                or target_teacher.employment_status != "ACTIVE"
            ):
                flash("Выберите работающего преподавателя.", "danger")
                return _workspace_redirect()
        elif target_value == "vacancy":
            state_key, state = _workspace_state(version.id)
            target_vacancy = _next_vacancy(
                version.id,
                state,
                note=request.form.get("target_vacancy_note"),
            )
        else:
            flash("Выберите нового преподавателя или вакансию.", "danger")
            return _workspace_redirect()

        source_query = WorkloadAssignment.query.filter(
            WorkloadAssignment.tariff_version_id == version.id,
            WorkloadAssignment.status != "CANCELLED",
        )
        if source_type == "vacancy":
            source_assignments = source_query.filter(
                WorkloadAssignment.assignment_kind == "VACANCY",
                WorkloadAssignment.position_code == source_vacancy_key,
            ).all()
            source_label = next(
                (
                    item.position_title for item in source_assignments
                    if item.position_title
                ),
                "Вакансия",
            )
        else:
            source_assignments = source_query.filter(
                WorkloadAssignment.assignment_kind != "VACANCY",
                WorkloadAssignment.employee_user_id == source_teacher_id,
            ).all()
            source_teacher = db.session.get(User, source_teacher_id)
            source_label = source_teacher.fio if source_teacher else "Преподаватель"

        if target_teacher is not None:
            if source_type != "vacancy" and source_teacher_id == target_teacher.id:
                flash("Выбран тот же преподаватель.", "warning")
                return _workspace_redirect()
            target_has_workload = source_query.filter(
                WorkloadAssignment.assignment_kind != "VACANCY",
                WorkloadAssignment.employee_user_id == target_teacher.id,
            ).first()
            if target_has_workload is not None:
                flash(
                    f"Замена не выполнена: у {target_teacher.fio} уже есть "
                    "нагрузка в выбранном учебном году.",
                    "danger",
                )
                return _workspace_redirect()

        key, state = _workspace_state(version.id)
        try:
            for assignment in source_assignments:
                before = assignment_snapshot(assignment)
                if target_teacher is not None:
                    assignment.employee_user_id = target_teacher.id
                    assignment.assignment_kind = "MAIN"
                    assignment.position_code = "TEACHER"
                    assignment.position_title = "Учитель"
                else:
                    assignment.employee_user_id = None
                    assignment.assignment_kind = "VACANCY"
                    assignment.position_code = target_vacancy["key"]
                    assignment.position_title = target_vacancy["label"]
                assignment.updated_by_user_id = current_user.id
                assignment.revision += 1
                validate_assignment(
                    assignment.workload_need,
                    assignment,
                    exclude_assignment_id=assignment.id,
                )
                add_assignment_change(
                    assignment,
                    "UPDATE",
                    user_id=current_user.id,
                    before_data=before,
                    reason="Замена владельца строки нагрузки",
                )
                refresh_need_status(assignment.workload_need)

            if source_type == "vacancy":
                state["vacancies"] = [
                    item for item in state["vacancies"]
                    if item.get("key") != source_vacancy_key
                ]
            else:
                state["teacher_ids"] = [
                    item for item in state["teacher_ids"]
                    if item != source_teacher_id
                ]

            if target_teacher is not None:
                if target_teacher.id not in state["teacher_ids"]:
                    state["teacher_ids"].append(target_teacher.id)
            else:
                state["vacancies"].append(target_vacancy)

            for row in state["rows"]:
                row_matches_source = (
                    source_type == "vacancy"
                    and row.get("holder_type") == "vacancy"
                    and row.get("vacancy_key") == source_vacancy_key
                ) or (
                    source_type != "vacancy"
                    and row.get("holder_type", "teacher") == "teacher"
                    and row.get("teacher_id") == source_teacher_id
                )
                if not row_matches_source:
                    continue
                if target_teacher is not None:
                    row.update({
                        "holder_type": "teacher",
                        "teacher_id": target_teacher.id,
                        "vacancy_key": None,
                    })
                else:
                    row.update({
                        "holder_type": "vacancy",
                        "teacher_id": None,
                        "vacancy_key": target_vacancy["key"],
                    })
            _save_workspace_state(key, state)
            db.session.commit()
        except (WorkloadDistributionError, IntegrityError) as exc:
            db.session.rollback()
            flash(
                str(exc)
                if isinstance(exc, WorkloadDistributionError)
                else "Не удалось заменить владельца нагрузки.",
                "danger",
            )
            return _workspace_redirect()

        target_label = (
            target_teacher.fio
            if target_teacher is not None else target_vacancy["label"]
        )
        flash(
            f"Вся нагрузка «{source_label}» передана: {target_label}.",
            "success",
        )
        return _workspace_redirect()

    @workload_bp.get("/assignments/workspace/export.xlsx")
    @login_required
    def assignment_workspace_export():
        _require_assignments_read()
        version_id = request.args.get("version_id", type=int)
        version = _draft_versions_query().filter(
            TariffVersion.id == version_id
        ).first_or_404()
        view_mode = (request.args.get("view") or "all").strip().lower()
        department_id = (
            request.args.get("department_id", type=int)
            if view_mode == "department"
            else None
        )
        education_levels = _workspace_selected_levels(request.args)
        grades = _workspace_selected_grades(request.args)
        selected_subject_ids = _workspace_selected_subject_ids(request.args)
        building_id = request.args.get("building_id", type=int)
        teacher_query = " ".join(
            (request.args.get("teacher_query") or "").split()
        )[:160]
        presentation_mode = (
            "list"
            if (request.args.get("presentation") or "").strip().lower()
            == "list"
            else "matrix"
        )
        snapshot, _, plan_matrices = _workspace_plan_context(
            version,
            education_levels=education_levels,
            grades=grades,
            building_id=building_id,
        )
        needs_query = _scoped_need_query().filter(
                WorkloadNeed.tariff_version_id == version.id,
                WorkloadNeed.status.in_(("OPEN", "PARTIAL", "COVERED")),
            )
        if selected_subject_ids:
            needs_query = needs_query.filter(
                WorkloadNeed.education_activity_id.in_(
                    selected_subject_ids
                )
            )
        if education_levels or grades or building_id is not None:
            visible_group_ids = _workspace_matrix_group_ids(plan_matrices)
            needs_query = (
                needs_query.filter(
                    WorkloadNeed.teaching_group_id.in_(visible_group_ids)
                )
                if visible_group_ids else needs_query.filter(db.false())
            )
        needs = _filter_workspace_needs(
            needs_query.all(),
            department_id=department_id,
            building_id=building_id,
            education_levels=education_levels,
            grades=grades,
            subject_ids=selected_subject_ids,
            population_snapshot_id=(snapshot.id if snapshot else None),
        )
        need_ids = [item.id for item in needs]
        assignments = (
            WorkloadAssignment.query
            .options(joinedload(WorkloadAssignment.employee))
            .filter(
                WorkloadAssignment.workload_need_id.in_(need_ids),
                WorkloadAssignment.status != "CANCELLED",
            )
            .all()
            if need_ids else []
        )
        needs_by_id = {item.id: item for item in needs}
        for assignment in assignments:
            set_committed_value(
                assignment,
                "workload_need",
                needs_by_id[assignment.workload_need_id],
            )
        if view_mode == "vacancies":
            assignments = [
                assignment for assignment in assignments
                if assignment.assignment_kind == "VACANCY"
            ]
        normalized_teacher_query = _workspace_search_text(teacher_query)
        if normalized_teacher_query:
            assignments = [
                assignment for assignment in assignments
                if normalized_teacher_query in _workspace_search_text(
                    _workspace_assignment_holder_label(assignment)
                )
            ]
        if presentation_mode == "list":
            workload_list = build_workload_assignment_list(assignments)
            workbook = _build_workload_list_workbook(
                workload_list,
                title=(
                    "Нагрузка педагогов · "
                    f"{version.tariff_cycle.academic_year.name}"
                ),
            )
            stream = BytesIO()
            workbook.save(stream)
            stream.seek(0)
            return send_file(
                stream,
                as_attachment=True,
                download_name=(
                    "Altair_workload_list_"
                    f"{version.tariff_cycle.academic_year.name.replace('/', '-')}"
                    ".xlsx"
                ),
                mimetype=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )
        matrix = build_workload_assignment_matrix(
            needs,
            assignments,
            plan_matrices=plan_matrices,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Нагрузка"
        sheet.append([
            "ФИО преподавателя",
            "Предмет",
            "Всего",
            "По предмету",
            *[
                class_group["label"]
                for class_group in matrix["class_groups"]
                for _ in class_group["columns"]
            ],
        ])
        sheet.append([
            "",
            "",
            "",
            "",
            *[
                column["subheader_label"]
                for column in matrix["columns"]
            ],
        ])
        for column_index in range(1, 5):
            sheet.merge_cells(
                start_row=1,
                start_column=column_index,
                end_row=2,
                end_column=column_index,
            )
        class_column_index = 5
        for class_group in matrix["class_groups"]:
            group_width = len(class_group["columns"])
            if group_width > 1:
                sheet.merge_cells(
                    start_row=1,
                    start_column=class_column_index,
                    end_row=1,
                    end_column=class_column_index + group_width - 1,
                )
            class_column_index += group_width
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F6FED")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
        for block in matrix["blocks"]:
            block_start_row = sheet.max_row + 1
            for row in block["rows"]:
                if row.get("placeholder"):
                    continue
                values = [
                    block["label"],
                    row["activity"].name,
                    float(block["total"]),
                    float(row["total"]),
                ]
                for column in matrix["columns"]:
                    values.append(float(sum(
                        (
                            Decimal(segment["hours"] or ZERO)
                            for segment in row["cells"].get(
                                column["key"],
                                [],
                            )
                        ),
                        ZERO,
                    )))
                sheet.append(values)
            block_end_row = sheet.max_row
            if block_end_row > block_start_row:
                sheet.merge_cells(
                    start_row=block_start_row,
                    start_column=1,
                    end_row=block_end_row,
                    end_column=1,
                )
                sheet.merge_cells(
                    start_row=block_start_row,
                    start_column=3,
                    end_row=block_end_row,
                    end_column=3,
                )
                sheet.cell(block_start_row, 1).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                sheet.cell(block_start_row, 3).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
        sheet.freeze_panes = "E3"
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 26
        sheet.column_dimensions["C"].width = 9
        sheet.column_dimensions["D"].width = 12
        for column_cells in sheet.iter_cols(
            min_col=5,
            max_col=sheet.max_column,
        ):
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = 12
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name=(
                "Altair_workload_"
                f"{version.tariff_cycle.academic_year.name.replace('/', '-')}"
                ".xlsx"
            ),
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    @workload_bp.get("/teachers/<int:user_id>")
    @login_required
    def workload_teacher_detail(user_id):
        scope = resolve_workload_scope(current_user)
        if scope.own_employee_only and current_user.id != user_id:
            abort(403)
        if not (
            can_use_workload_permission("workload.read", current_user)
            or (
                can_use_workload_permission(
                    "workload.self.read",
                    current_user,
                )
                and current_user.id == user_id
            )
        ):
            abort(403)
        employee = User.query.get_or_404(user_id)
        query = WorkloadAssignment.query.filter(
            WorkloadAssignment.employee_user_id == user_id,
            WorkloadAssignment.status != "CANCELLED",
        )
        if not scope.unrestricted:
            need_ids = [
                item.id for item in _scoped_need_query().all()
            ]
            query = query.filter(
                WorkloadAssignment.workload_need_id.in_(need_ids)
            )
        assignments = query.order_by(
            WorkloadAssignment.date_from.asc(),
            WorkloadAssignment.id.asc(),
        ).all()
        assignment_ids = [item.id for item in assignments]
        tariff_lines = {}
        if assignment_ids:
            calculated_lines = (
                TariffLine.query
                .join(TariffCalculationRun)
                .filter(
                    TariffLine.workload_assignment_id.in_(assignment_ids),
                    TariffCalculationRun.status == "SUCCEEDED",
                )
                .order_by(
                    TariffCalculationRun.id.desc(),
                    TariffLine.id.desc(),
                )
                .all()
            )
            for line in calculated_lines:
                tariff_lines.setdefault(
                    line.workload_assignment_id,
                    line,
                )
        calculated_fte = sum(
            (
                Decimal(line.fte_value or ZERO)
                for line in tariff_lines.values()
            ),
            ZERO,
        )
        calculated_amount = sum(
            (
                Decimal(line.total_amount or ZERO)
                for line in tariff_lines.values()
            ),
            ZERO,
        )
        can_view_finance = can_use_workload_permission(
            "workload.finance.read",
            current_user,
        )
        return render_template(
            "workload/teacher_detail.html",
            employee=employee,
            assignments=assignments,
            totals=teacher_totals(assignments),
            tariff_lines=tariff_lines,
            calculated_fte=calculated_fte,
            calculated_amount=calculated_amount,
            can_view_finance=can_view_finance,
            assignment_kind_labels=WORKLOAD_ASSIGNMENT_KIND_LABELS,
        )

    @workload_bp.get("/departments/")
    @login_required
    def workload_departments():
        _require_assignments_read()
        needs = (
            _scoped_need_query()
            .filter(WorkloadNeed.status != "CANCELLED")
            .all()
        )
        rows = defaultdict(lambda: {
            "department": None,
            "needs": [],
            "weekly": ZERO,
            "allocated": ZERO,
            "teachers": set(),
        })
        for need in needs:
            key = need.department_id or 0
            row = rows[key]
            row["department"] = need.department
            row["needs"].append(need)
            row["weekly"] += Decimal(need.weekly_hours or ZERO)
            row["allocated"] += need.allocated_weekly_hours
            row["teachers"].update(
                item.employee_user_id
                for item in need.active_assignments
                if item.employee_user_id is not None
                and item.assignment_kind != "VACANCY"
            )
        result = []
        for row in rows.values():
            row["remaining"] = row["weekly"] - row["allocated"]
            row["teacher_count"] = len(row["teachers"])
            result.append(row)
        result.sort(
            key=lambda item: (
                item["department"].name.lower()
                if item["department"] else "я"
            )
        )
        return render_template(
            "workload/departments.html",
            rows=result,
        )
