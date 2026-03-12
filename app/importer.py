from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required
from openpyxl import load_workbook
from datetime import datetime, date
import re

from .models import Child
from app.core.extensions import db

import_bp = Blueprint("importer", __name__)

# твой формат Excel (английские заголовки)
REQUIRED_HEADERS = [
    "last_name", "first_name", "middle_name", "birth_date",
    "grade", "class_letter",
    "mother_fio", "mother_phone",
    "father_fio", "father_phone",
    "reg_address"
]

def _to_str(x):
    return (str(x).strip() if x is not None else "")

def _parse_birth(x):
    if not x:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x

    s = str(x).strip()
    # поддержим DD.MM.YYYY
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))

    # YYYY-MM-DD
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _parse_grade(x):
    if x is None or x == "":
        return None
    if isinstance(x, (int, float)):
        g = int(x)
        return g
    s = str(x).strip()
    return int(s) if s.isdigit() else None

@import_bp.route("/import/children", methods=["GET", "POST"])
@login_required
def import_children():
    # если нужно только ADMIN — замени login_required на require_roles("ADMIN")
    from .models import AcademicYear, SchoolClass, ChildEnrollment

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Выберите Excel-файл", "danger")
            return redirect(url_for("importer.import_children"))

        year = AcademicYear.query.filter_by(is_current=True).first()
        if not year:
            flash("Не найден текущий учебный год (AcademicYear.is_current=True)", "danger")
            return redirect(url_for("importer.import_children"))

        wb = load_workbook(f, data_only=True)
        ws = wb.active

        header_row = [_to_str(c.value) for c in ws[1]]
        header_to_idx = {h: i for i, h in enumerate(header_row)}

        missing = [h for h in REQUIRED_HEADERS if h not in header_to_idx]
        if missing:
            flash(f"В Excel не хватает колонок: {', '.join(missing)}", "danger")
            return redirect(url_for("importer.import_children"))

        def cell(row_values, name):
            return row_values[header_to_idx[name]]

        created, skipped = 0, 0

        for r in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

            last_name = _to_str(cell(row_values, "last_name"))
            first_name = _to_str(cell(row_values, "first_name"))
            middle_name = _to_str(cell(row_values, "middle_name")) or None
            birth_date = _parse_birth(cell(row_values, "birth_date"))

            grade = _parse_grade(cell(row_values, "grade"))
            class_letter = _to_str(cell(row_values, "class_letter")) or None

            mother_fio = _to_str(cell(row_values, "mother_fio")) or None
            mother_phone = _to_str(cell(row_values, "mother_phone")) or None
            father_fio = _to_str(cell(row_values, "father_fio")) or None
            father_phone = _to_str(cell(row_values, "father_phone")) or None
            reg_address = _to_str(cell(row_values, "reg_address")) or None

            if not last_name or not first_name:
                skipped += 1
                continue

            # ограничим параллель
            if grade is not None and not (1 <= grade <= 11):
                # если хочешь ДО — делаем grade=None, а class_letter="ДО" или наоборот
                skipped += 1
                continue

            class_name = f"{grade}{class_letter or ''}" if grade else None

            child = Child(
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                birth_date=birth_date,

                grade=grade,
                class_letter=class_letter,
                class_name=class_name,  # совместимость

                mother_fio=mother_fio,
                mother_phone=mother_phone,
                father_fio=father_fio,
                father_phone=father_phone,
                reg_address=reg_address,
            )

            db.session.add(child)
            db.session.flush()

            # привязка к реестру классов + enrollment
            if class_name:
                sc = (SchoolClass.query
                      .filter(SchoolClass.academic_year_id == year.id,
                              SchoolClass.name == class_name)
                      .first())
                if not sc:
                    sc = SchoolClass(
                        academic_year_id=year.id,
                        name=class_name,
                        max_students=25
                    )
                    db.session.add(sc)
                    db.session.flush()

                en = ChildEnrollment(
                    child_id=child.id,
                    academic_year_id=year.id,
                    school_class_id=sc.id,
                    status="ACTIVE"
                )
                db.session.add(en)

            created += 1

        db.session.commit()
        flash(f"Импорт завершён. Добавлено детей: {created}, пропущено строк: {skipped}", "success")
        return redirect(url_for("children.list_children"))

    return render_template("children_import.html")