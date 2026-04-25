"""Кубок школы: скачиваем Google Sheets → агрегируем → локальный кеш.

Источник истины — открытая Google-таблица с колонками
(Класс, Дата события, Тип события, Баллы).

Использование из кода:
    from app.kubok import get_rating, refresh_cache
    info = get_rating("7-ПМ")     # dict: place, total_points, activities, ...

Обновление кеша:
    - лениво (автоматом при первом обращении, если данных > CACHE_TTL);
    - вручную: flask --app run refresh-kubok
"""
from __future__ import annotations

import json
import os
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from flask import current_app, g, has_request_context

from app.core.extensions import db
from app.models import ClassRatingSnapshot


DEFAULT_SHEET_ID = "1GMCuc_4v0tbXG6f_OIg0MXB1rAV__maqBtZ4UJ7hUTI"
CACHE_TTL_SECONDS = 24 * 3600  # 24 часа


def _sheet_id() -> str:
    return os.getenv("KUBOK_SHEET_ID", DEFAULT_SHEET_ID)


def _year_label() -> str:
    return os.getenv("KUBOK_YEAR_LABEL", "25-26")


def _download_xlsx() -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{_sheet_id()}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "school-tracker-kubok/1.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read()


def _aggregate(xlsx_bytes: bytes) -> dict[str, dict]:
    """Возвращает {class_name: {place, total_classes, total_points, activities[]}}."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "Журнал рейтинга" not in wb.sheetnames:
        raise RuntimeError(f"Ожидаемый лист «Журнал рейтинга» не найден. Есть: {wb.sheetnames}")
    ws = wb["Журнал рейтинга"]

    totals: dict[str, float] = defaultdict(float)
    activities: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for row in ws.iter_rows(min_row=2, values_only=True):
        cls = row[0] if len(row) > 0 else None
        event = row[2] if len(row) > 2 else None
        pts = row[3] if len(row) > 3 else None
        if not cls or pts is None:
            continue
        try:
            p = float(pts)
        except (TypeError, ValueError):
            continue
        key = str(cls).strip()
        totals[key] += p
        activities[key][(event or "—").strip() or "—"] += p

    ranked = sorted(totals.items(), key=lambda x: -x[1])
    total_classes = len(ranked)
    result: dict[str, dict] = {}
    for place, (cls, s) in enumerate(ranked, 1):
        acts_sorted = sorted(activities[cls].items(), key=lambda x: -x[1])
        result[cls] = {
            "place": place,
            "total_classes": total_classes,
            "total_points": round(s, 2),
            "activities": [
                {"name": n, "points": round(p, 2)}
                for n, p in acts_sorted
            ],
        }
    return result


def refresh_cache() -> tuple[int, datetime]:
    """Скачивает Sheets, агрегирует и перезаписывает кеш. Возвращает (кол-во классов, время)."""
    xlsx = _download_xlsx()
    data = _aggregate(xlsx)
    year = _year_label()

    ClassRatingSnapshot.query.filter_by(year_label=year).delete()
    db.session.flush()

    now = datetime.utcnow()
    for cls, info in data.items():
        db.session.add(ClassRatingSnapshot(
            class_name=cls,
            year_label=year,
            place=info["place"],
            total_classes=info["total_classes"],
            total_points=info["total_points"],
            activities_json=json.dumps(info["activities"], ensure_ascii=False),
            updated_at=now,
        ))
    db.session.commit()
    return len(data), now


def _cache_age() -> timedelta | None:
    newest = (
        db.session.query(db.func.max(ClassRatingSnapshot.updated_at))
        .filter_by(year_label=_year_label())
        .scalar()
    )
    if not newest:
        return None
    return datetime.utcnow() - newest


def _needs_refresh() -> bool:
    age = _cache_age()
    return age is None or age.total_seconds() > CACHE_TTL_SECONDS


def _name_variants(class_name: str) -> list[str]:
    """Варианты написания имени класса для матчинга кеша.

    В school_class имена пишутся без дефиса («7НИ», «11В», «6АГ»),
    а в Google Sheets и соответственно в class_rating_snapshot — с дефисом
    («7-НИ», «11-В», «6-АГ»). Делаем поиск по обоим вариантам.
    """
    if not class_name:
        return []
    raw = str(class_name).strip()
    out = [raw]
    upper = raw.upper()
    if upper != raw:
        out.append(upper)
    # dashed variant: вставить "-" после ведущих цифр
    for src in (raw, upper):
        i = 0
        while i < len(src) and src[i].isdigit():
            i += 1
        if 0 < i < len(src) and src[i] != "-":
            cand = src[:i] + "-" + src[i:]
            if cand not in out:
                out.append(cand)
        # de-dashed variant
        if "-" in src:
            cand = src.replace("-", "")
            if cand not in out:
                out.append(cand)
    return out


def _row_to_dict(row: "ClassRatingSnapshot") -> dict:
    try:
        activities = json.loads(row.activities_json or "[]")
    except ValueError:
        activities = []
    return {
        "class_name": row.class_name,
        "place": row.place,
        "total_classes": row.total_classes,
        "total_points": row.total_points,
        "activities": activities,
        "updated_at": row.updated_at,
    }


def _load_snapshots_by_name() -> dict[str, dict]:
    """Единый batch-load всех снимков года — {class_name: rating_dict}.

    Обогащает каждый rating_dict полями места в параллели и в здании через
    join с SchoolClass по имени (с учётом вариантов — «7-НИ» ↔ «7НИ»).
    Если SchoolClass не найден — оставляет None.
    """
    from app.models import SchoolClass, Building, AcademicYear

    year_label = _year_label()
    rows = ClassRatingSnapshot.query.filter_by(year_label=year_label).all()

    # Собираем rating_dict для каждого snapshot_name + ссылок по всем вариантам имени
    by_name: dict[str, dict] = {}
    for row in rows:
        rd = _row_to_dict(row)
        by_name[row.class_name] = rd

    # Подтягиваем текущий учебный год и классы
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    building_name_by_id: dict[int, str] = {b.id: b.name for b in Building.query.all()}

    # Для каждого snapshot пытаемся найти соответствующий SchoolClass через варианты
    school_classes: list = []
    if current_year is not None:
        school_classes = SchoolClass.query.filter_by(academic_year_id=current_year.id).all()
    class_by_variant: dict[str, "SchoolClass"] = {}
    for sc in school_classes:
        for variant in _name_variants(sc.name):
            class_by_variant.setdefault(variant, sc)

    # Обогащаем snapshot информацией о параллели/здании
    enriched: dict[str, dict] = {}
    for snap_name, rd in by_name.items():
        sc = None
        for variant in _name_variants(snap_name):
            if variant in class_by_variant:
                sc = class_by_variant[variant]
                break
        rd["grade"] = sc.grade if sc else None
        rd["building_id"] = sc.building_id if sc else None
        rd["building_name"] = building_name_by_id.get(sc.building_id) if sc else None
        enriched[snap_name] = rd

    # Ранжирование внутри параллели и внутри здания.
    # Сортируем по total_points DESC; place = 1-based.
    def _rank_group(group_snaps: list[dict], place_field: str, total_field: str) -> None:
        group_snaps.sort(key=lambda r: -r.get("total_points", 0.0))
        for i, r in enumerate(group_snaps, 1):
            r[place_field] = i
            r[total_field] = len(group_snaps)

    # Группировка по grade
    by_grade: dict[int, list[dict]] = defaultdict(list)
    for rd in enriched.values():
        if rd.get("grade") is not None:
            by_grade[rd["grade"]].append(rd)
    for grade, group in by_grade.items():
        _rank_group(group, "place_in_parallel", "total_in_parallel")

    # Группировка по building_id (None складываем в отдельную группу)
    by_building: dict[int | None, list[dict]] = defaultdict(list)
    for rd in enriched.values():
        by_building[rd.get("building_id")].append(rd)
    for bid, group in by_building.items():
        _rank_group(group, "place_in_building", "total_in_building")

    # Заполняем отсутствующие ключи None (классы без grade / building)
    for rd in enriched.values():
        rd.setdefault("place_in_parallel", None)
        rd.setdefault("total_in_parallel", None)
        rd.setdefault("place_in_building", None)
        rd.setdefault("total_in_building", None)

    # Для вариантов имён (с/без дефиса) делаем дополнительные alias-ключи,
    # чтобы get_rating("7НИ") находил snapshot «7-НИ» через _name_variants(class_name).
    return enriched


def _request_cache() -> dict | None:
    """Кеш на текущий HTTP-запрос: {year_label, needs_refresh_checked, snapshots}.
    При первом вызове в запросе — загружает все snapshots одним SELECT;
    далее все вызовы get_rating() — O(1) dict lookup без SQL.
    """
    if not has_request_context():
        return None
    cache = getattr(g, "_kubok_request_cache", None)
    if cache is None:
        cache = {
            "year": _year_label(),
            "refresh_checked": False,
            "snapshots": None,  # lazy-load на первый get_rating
        }
        g._kubok_request_cache = cache
    return cache


def get_rating(class_name: str, auto_refresh: bool = True) -> dict | None:
    """Отдаёт рейтинг класса. Внутри одного HTTP-запроса кеширует batch-load
    всех snapshots в flask.g — минимизирует N запросов SQL до одного.
    """
    cache = _request_cache()

    # Проверка staleness — максимум 1 раз на запрос
    if auto_refresh:
        if cache is None:
            # нет request-контекста: старое поведение
            if _needs_refresh():
                try:
                    refresh_cache()
                except Exception as exc:  # noqa: BLE001
                    current_app.logger.warning("Kubok cache refresh failed: %s", exc)
        elif not cache["refresh_checked"]:
            cache["refresh_checked"] = True
            if _needs_refresh():
                try:
                    refresh_cache()
                    cache["snapshots"] = None  # перечитать после refresh
                except Exception as exc:  # noqa: BLE001
                    current_app.logger.warning("Kubok cache refresh failed: %s", exc)

    # Lookup
    if cache is not None:
        if cache["snapshots"] is None:
            cache["snapshots"] = _load_snapshots_by_name()
        snapshots = cache["snapshots"]
        for variant in _name_variants(class_name):
            if variant in snapshots:
                return snapshots[variant]
        return None

    # Вне request-контекста (CLI / scheduler) — старое поведение
    year = _year_label()
    for variant in _name_variants(class_name):
        row = (
            ClassRatingSnapshot.query
            .filter_by(class_name=variant, year_label=year)
            .first()
        )
        if row:
            return _row_to_dict(row)
    return None


def cache_status() -> dict:
    """Краткая сводка для UI: сколько классов в кеше, когда обновлялось."""
    q = ClassRatingSnapshot.query.filter_by(year_label=_year_label())
    count = q.count()
    newest = (
        db.session.query(db.func.max(ClassRatingSnapshot.updated_at))
        .filter_by(year_label=_year_label())
        .scalar()
    )
    return {
        "classes": count,
        "updated_at": newest,
        "year": _year_label(),
        "stale": _needs_refresh(),
    }
