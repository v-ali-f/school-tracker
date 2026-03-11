import os

from dotenv import load_dotenv
from flask import Flask, current_app
from .core.database import db, migrate
from .core.security import login_manager

load_dotenv(override=True)



def _ensure_runtime_schema(app):
    from sqlalchemy import inspect, text
    from sqlalchemy.schema import CreateColumn

    with app.app_context():
        inspector = inspect(db.engine)
        try:
            cols = {c["name"] for c in inspector.get_columns("child")}
        except Exception:
            cols = set()
        needed = {
            "ovz_doc_number": "ALTER TABLE child ADD COLUMN ovz_doc_number VARCHAR(100)",
            "ovz_doc_date": "ALTER TABLE child ADD COLUMN ovz_doc_date DATE",
            "disability_ipra": "ALTER TABLE child ADD COLUMN disability_ipra VARCHAR(255)",
            "status": "ALTER TABLE child ADD COLUMN status VARCHAR(30) DEFAULT 'ACTIVE'",
            "archived_at": "ALTER TABLE child ADD COLUMN archived_at TIMESTAMP",
        }
        for name, sql in needed.items():
            if name not in cols:
                try:
                    db.session.execute(text(sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        try:
            cw_cols = {c["name"] for c in inspector.get_columns("control_work")}
        except Exception:
            cw_cols = set()
        cw_needed = {
            "grade5_percent": "ALTER TABLE control_work ADD COLUMN grade5_percent INTEGER DEFAULT 85",
            "grade4_percent": "ALTER TABLE control_work ADD COLUMN grade4_percent INTEGER DEFAULT 65",
            "grade3_percent": "ALTER TABLE control_work ADD COLUMN grade3_percent INTEGER DEFAULT 45",
        }
        for name, sql in cw_needed.items():
            if name not in cw_cols:
                try:
                    db.session.execute(text(sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        try:
            task_cols = {c["name"] for c in inspector.get_columns("control_work_task")}
        except Exception:
            task_cols = set()
        task_needed = {
            "description": "ALTER TABLE control_work_task ADD COLUMN description VARCHAR(255)",
            "topic": "ALTER TABLE control_work_task ADD COLUMN topic VARCHAR(255)",
        }
        for name, sql in task_needed.items():
            if name not in task_cols:
                try:
                    db.session.execute(text(sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        try:
            result_cols = {c["name"] for c in inspector.get_columns("control_work_result")}
        except Exception:
            result_cols = set()
        result_needed = {
            "assignment_id": "ALTER TABLE control_work_result ADD COLUMN assignment_id INTEGER",
            "grade5_percent": "ALTER TABLE control_work_result ADD COLUMN grade5_percent INTEGER DEFAULT 85",
            "grade4_percent": "ALTER TABLE control_work_result ADD COLUMN grade4_percent INTEGER DEFAULT 65",
            "grade3_percent": "ALTER TABLE control_work_result ADD COLUMN grade3_percent INTEGER DEFAULT 45",
            "academic_year_id": "ALTER TABLE control_work_result ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE control_work_result ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE control_work_result ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        }
        for name, sql in result_needed.items():
            if name not in result_cols:
                try:
                    db.session.execute(text(sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        try:
            refreshed_result_cols = {c["name"] for c in inspect(db.engine).get_columns("control_work_result")}
        except Exception:
            refreshed_result_cols = set()
        if "assignment_id" in refreshed_result_cols:
            try:
                db.session.execute(text("""
                    UPDATE control_work_result r
                    SET assignment_id = a.id
                    FROM control_work_assignment a
                    WHERE r.assignment_id IS NULL
                      AND a.control_work_id = r.control_work_id
                      AND a.school_class_id = r.school_class_id
                """))
                db.session.commit()
            except Exception:
                db.session.rollback()

        for table_name, additions in {
            "academic_year": {
                "is_closed": "ALTER TABLE academic_year ADD COLUMN is_closed BOOLEAN DEFAULT FALSE",
                "is_archived": "ALTER TABLE academic_year ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
                "updated_at": "ALTER TABLE academic_year ADD COLUMN updated_at TIMESTAMP",
            },
            "school_class": {
                "is_active": "ALTER TABLE school_class ADD COLUMN is_active BOOLEAN DEFAULT TRUE",
                "is_archived": "ALTER TABLE school_class ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            },
            "child_enrollment": {
                "transfer_order_number": "ALTER TABLE child_enrollment ADD COLUMN transfer_order_number VARCHAR(100)",
                "transfer_order_date": "ALTER TABLE child_enrollment ADD COLUMN transfer_order_date DATE",
            },
            "child_parent": {
                "transfer_order_number": "ALTER TABLE child_parent ADD COLUMN transfer_order_number VARCHAR(100)",
                "transfer_order_date": "ALTER TABLE child_parent ADD COLUMN transfer_order_date DATE",
            },
            "parent": {
                "retention_until": "ALTER TABLE parent ADD COLUMN retention_until DATE",
                "is_archived": "ALTER TABLE parent ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
                "created_at": "ALTER TABLE parent ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            },
            "document": {
                "academic_year_id": "ALTER TABLE document ADD COLUMN academic_year_id INTEGER",
                "retention_until": "ALTER TABLE document ADD COLUMN retention_until DATE",
                "is_archived": "ALTER TABLE document ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
                "is_hidden_by_retention": "ALTER TABLE document ADD COLUMN is_hidden_by_retention BOOLEAN DEFAULT FALSE",
                "is_deleted_soft": "ALTER TABLE document ADD COLUMN is_deleted_soft BOOLEAN DEFAULT FALSE",
                "deleted_at": "ALTER TABLE document ADD COLUMN deleted_at TIMESTAMP",
                "deleted_by": "ALTER TABLE document ADD COLUMN deleted_by INTEGER",
            },
            "control_work": {
                "academic_year_id": "ALTER TABLE control_work ADD COLUMN academic_year_id INTEGER",
                "retention_until": "ALTER TABLE control_work ADD COLUMN retention_until DATE",
            },
            "teacher_load": {
                "academic_year_id": "ALTER TABLE teacher_load ADD COLUMN academic_year_id INTEGER",
                "retention_until": "ALTER TABLE teacher_load ADD COLUMN retention_until DATE",
                "is_archived": "ALTER TABLE teacher_load ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            },
            "teacher_mcko_result": {
                "academic_year_id": "ALTER TABLE teacher_mcko_result ADD COLUMN academic_year_id INTEGER",
                "retention_until": "ALTER TABLE teacher_mcko_result ADD COLUMN retention_until DATE",
                "is_archived": "ALTER TABLE teacher_mcko_result ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            },
            "teacher_course": {
                "academic_year_id": "ALTER TABLE teacher_course ADD COLUMN academic_year_id INTEGER",
                "retention_until": "ALTER TABLE teacher_course ADD COLUMN retention_until DATE",
                "is_archived": "ALTER TABLE teacher_course ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            },
            "user": {
                "employment_status": "ALTER TABLE \"user\" ADD COLUMN employment_status VARCHAR(30) DEFAULT 'ACTIVE'",
                "dismissal_date": "ALTER TABLE \"user\" ADD COLUMN dismissal_date DATE",
                "archived_at": "ALTER TABLE \"user\" ADD COLUMN archived_at TIMESTAMP",
            },
        }.items():
            try:
                existing = {c["name"] for c in inspector.get_columns(table_name)}
            except Exception:
                existing = set()
            for name, sql in additions.items():
                if name not in existing:
                    try:
                        db.session.execute(text(sql))
                        db.session.commit()
                    except Exception:
                        db.session.rollback()

        # Generic fallback: if a model already contains new nullable/runtime columns,
        # but the physical PostgreSQL table is older, add the missing columns automatically.
        # This helps after SQLite -> PostgreSQL migrations when metadata changed faster than schema.
        inspector = inspect(db.engine)
        for table in db.metadata.sorted_tables:
            try:
                existing_cols = {c["name"] for c in inspector.get_columns(table.name)}
            except Exception:
                continue
            for column in table.columns:
                if column.name in existing_cols:
                    continue
                if getattr(column, "primary_key", False):
                    continue
                try:
                    col_sql = str(CreateColumn(column).compile(dialect=db.engine.dialect))
                    db.session.execute(text(f'ALTER TABLE "{table.name}" ADD COLUMN {col_sql}'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                    # Best-effort runtime upgrade: skip columns that require a dedicated migration.
                    continue

        db.session.commit()



def _seed_olympiad_subject_mappings(app):
    try:
        from openpyxl import load_workbook
        from .models import Department, DepartmentSubject, OlympiadSubjectMapping, Subject
    except Exception:
        return

    with app.app_context():
        try:
            if OlympiadSubjectMapping.query.count() > 0:
                return
        except Exception:
            return

        seed_path = os.path.join(app.root_path, '..', 'data', 'olympiad_subjects_vsoh.xlsx')
        seed_path = os.path.abspath(seed_path)
        if not os.path.exists(seed_path):
            return

        try:
            wb = load_workbook(seed_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
        except Exception:
            return

        def norm(v):
            return ' '.join(str(v or '').replace('ё', 'е').replace('Ё', 'Е').split()).strip().lower()

        subjects = Subject.query.all()
        subjects_by_name = {norm(s.name): s for s in subjects}

        def match_subject(raw_school_subjects: str):
            variants = []
            for part in str(raw_school_subjects or '').replace(';', ',').split(','):
                item = norm(part)
                if item:
                    variants.append(item)
            for item in variants:
                if item in subjects_by_name:
                    return subjects_by_name[item]
            for item in variants:
                for subj in subjects:
                    subj_norm = norm(subj.name)
                    if subj_norm == item or subj_norm in item or item in subj_norm:
                        return subj
            return None

        created = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue
            olympiad_name = str(row[0] or '').strip()
            school_subjects = str(row[1] or '').strip()
            if not olympiad_name or not school_subjects:
                continue
            subject = match_subject(school_subjects)
            if not subject:
                continue
            dep_link = DepartmentSubject.query.filter_by(subject_id=subject.id).first()
            mapping = OlympiadSubjectMapping.query.filter_by(olympiad_subject_name=olympiad_name).first()
            if mapping:
                continue
            mapping = OlympiadSubjectMapping(
                olympiad_subject_name=olympiad_name,
                subject_id=subject.id,
                department_id=dep_link.department_id if dep_link else None,
                comment=f'Базовая загрузка из перечня ВСОШ: {school_subjects}',
                is_active=True,
            )
            db.session.add(mapping)
            created += 1
        if created:
            db.session.commit()


def get_current_year():
    from .models import AcademicYear
    return AcademicYear.query.filter_by(is_current=True).first()




def create_app():
    app = Flask(__name__)

    from .core import configure_app, init_extensions, register_context_processors, login_manager as core_login_manager
    configure_app(app)
    init_extensions(app)

    from .permissions import has_permission, build_menu_flags
    from .models import User  # noqa

    @core_login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    register_context_processors(app, has_permission, build_menu_flags)

    from .modules import register_blueprints
    register_blueprints(app)

    with app.app_context():
        db.create_all()
    _ensure_runtime_schema(app)
    _seed_olympiad_subject_mappings(app)

    from .retention import apply_retention_policies
    with app.app_context():
        apply_retention_policies()

    print("DATABASE =", app.config["SQLALCHEMY_DATABASE_URI"])
    print("UPLOAD_FOLDER =", app.config["UPLOAD_FOLDER"])

    return app
