from __future__ import annotations
from pathlib import Path

from typing import Optional

import re
from collections import defaultdict
from datetime import date, datetime

from flask import Blueprint, flash, g, redirect, render_template, request, url_for, abort
from flask_login import current_user, login_required
from openpyxl import load_workbook
from sqlalchemy import func, or_

from app.core.extensions import db
from app.services.education_activity_service import (
    assign_subject_activity,
    get_or_create_subject_activity,
    get_subject_activity,
    list_subject_activities,
)
from app.services.workload_integration_service import (
    current_department_load_rows,
    department_teacher_ids,
)
from app.services.teacher_mcko_service import (
    MCKO_LEVEL_LABELS,
    mcko_expires_at,
    mcko_results_for_teachers,
    normalize_mcko_level,
)
from app.services.teacher_attestation_service import (
    ATTESTATION_CATEGORY_LABELS,
    position_compliance_due_at,
)
from app.services.teacher_professional_service import professional_entry_source
from app.services.teacher_professional_audit_service import (
    CHANGE_ARCHIVED,
    CHANGE_CREATED,
    record_professional_change,
)
from .models import (
    ACTIVITY_KIND_LABELS,
    Building,
    ControlWork,
    ControlWorkAssignment,
    ControlWorkResult,
    AcademicYear,
    Department,
    DepartmentLeader,
    DepartmentSubject,
    Debt,
    EducationActivity,
    EducationActivityDepartment,
    Incident,
    TeacherAttestation,
    TeacherCourse,
    TeacherLoad,
    TeacherMckoResult,
    User,
    WORKLOAD_APPROVAL_STATUS_LABELS,
)
from .models.diagnostics import DiagnosticResult, DiagnosticSession
from .permissions import (
    CLASS_TEACHER,
    DEPARTMENT_HEAD,
    METHODIST,
    TEACHER,
    has_permission,
    has_role,
    is_admin,
)
from .services.olympiad_stats_service import department_stats as olympiad_department_stats, dashboard_stats as olympiad_dashboard_stats
from app.users import register_unmatched_staff
from app.utils.user_matching import find_existing_user


departments_bp = Blueprint("departments", __name__, url_prefix="/departments")


VALID_MARK_VALUES = {"2", "3", "4", "5"}

def _safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _diagnostic_mark_label(row) -> str | None:
    raw = str(getattr(row, "mark", "") or "").strip()
    return raw if raw in VALID_MARK_VALUES else None


def _diagnostic_level_label(row) -> str | None:
    raw = ((getattr(row, "level", None) or "").strip()).lower()
    if raw in {"высокий", "высокий уровень"}:
        return "Высокий"
    if raw in {"повышенный", "повышенный уровень"}:
        return "Повышенный"
    if raw in {"базовый", "базовый уровень"}:
        return "Базовый"
    if raw in {"низкий", "ниже базового", "нижебазового"}:
        return "Ниже базового"
    return None


def _detect_diagnostic_result_mode(rows: list[DiagnosticResult]) -> str:
    has_level = any(_diagnostic_level_label(row) for row in rows)
    has_mark = any(_diagnostic_mark_label(row) for row in rows)
    has_percent = any(_safe_float(getattr(row, "percent", None)) is not None for row in rows)
    has_score = any(_safe_float(getattr(row, "total_score", None)) is not None for row in rows)
    if (has_level and has_mark) or (has_level and (has_percent or has_score) and has_mark):
        return "mixed"
    if has_level:
        return "level"
    if has_mark:
        return "mark"
    if has_percent:
        return "percent"
    return "score" if has_score else "percent"


def _result_mode_label(mode: str) -> str:
    return {
        "level": "по уровням",
        "mark": "по отметкам",
        "mixed": "смешанный",
        "percent": "по баллам / проценту",
        "score": "по баллам / проценту",
    }.get(mode or "", "по баллам / проценту")


DEFAULT_DEPARTMENTS = [
    {
        "name": "Кафедра учителей начальных классов",
        "code": "primary",
        "subject_names": [],
        "description": "Автоподбор по нагрузке 1–4 классов, кроме физической культуры и музыки.",
    },
    {
        "name": "Кафедра учителей физической культуры и спорта",
        "code": "sport",
        "subject_names": ["Физическая культура", "Физкультура", "Спорт"],
    },
    {
        "name": "Кафедра учителей естественно-научного цикла",
        "code": "science",
        "subject_names": ["География", "Физика", "Химия", "Биология"],
    },
    {
        "name": "Кафедра математического образования",
        "code": "math",
        "subject_names": [
            "Математика",
            "Алгебра",
            "Геометрия",
            "Вероятность и статистика",
            "Алгебра и начала математического анализа",
            "Информатика",
        ],
    },
    {
        "name": "Кафедра словесности",
        "code": "philology",
        "subject_names": ["Русский язык", "Литература"],
        "description": "Для 5–11 классов.",
    },
    {
        "name": "Кафедра иностранного языка",
        "code": "foreign_language",
        "subject_names": [
            "Английский язык",
            "Немецкий язык",
            "Французский язык",
            "Испанский язык",
            "Китайский язык",
            "Иностранный язык",
        ],
    },
    {
        "name": "Кафедра эстетического образования",
        "code": "art",
        "subject_names": ["Изобразительное искусство", "ИЗО", "Труд", "Технология", "Музыка"],
    },
    {
        "name": "Кафедра общественных наук",
        "code": "social_science",
        "subject_names": ["История", "Обществознание"],
    },
]

PRIMARY_EXCLUDED = {"физическая культура", "физкультура", "музыка"}


def _normalize_subject_name(value: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).lower()


def _extract_grade_from_class_text(class_text: Optional[str]):
    if not class_text:
        return None
    m = re.search(r"(?<!\d)(1[01]|[1-9])(?=\D|$)", str(class_text))
    return int(m.group(1)) if m else None


def _ensure_default_departments():
    if getattr(g, '_departments_ensured', False):
        return
    g._departments_ensured = True
    changed = False
    for item in DEFAULT_DEPARTMENTS:
        dep = Department.query.filter_by(code=item["code"]).first()
        if not dep:
            dep = Department(name=item["name"], code=item["code"], description=item.get("description"))
            db.session.add(dep)
            db.session.flush()
            changed = True
        else:
            if not dep.name:
                dep.name = item["name"]
                changed = True
            if item.get("description") and not dep.description:
                dep.description = item.get("description")
                changed = True
        for subject_name in item.get("subject_names", []):
            activity, activity_created = get_or_create_subject_activity(subject_name)
            if activity_created:
                changed = True
            legacy_subject_id = (
                activity.legacy_subject.id
                if activity.legacy_subject
                else None
            )
            exists = (
                DepartmentSubject.query
                .filter(DepartmentSubject.department_id == dep.id)
                .filter(or_(
                    DepartmentSubject.education_activity_id == activity.id,
                    DepartmentSubject.subject_id == legacy_subject_id,
                ))
                .first()
            )
            if not exists:
                link = DepartmentSubject(department_id=dep.id)
                assign_subject_activity(link, activity)
                db.session.add(link)
                changed = True
            elif exists.education_activity_id != activity.id:
                assign_subject_activity(exists, activity)
                changed = True
    if changed:
        db.session.commit()


def _department_allowed(dep: Department) -> bool:
    if is_admin(current_user):
        return True
    if has_role(METHODIST):
        return True
    if has_role(DEPARTMENT_HEAD):
        return DepartmentLeader.query.filter_by(department_id=dep.id, user_id=current_user.id).first() is not None
    if has_role(TEACHER) or has_role(CLASS_TEACHER):
        return True
    return False


def _load_departments_for_user():
    _ensure_default_departments()
    deps = Department.query.order_by(Department.name.asc()).all()
    if is_admin(current_user) or has_role(METHODIST):
        return deps
    if has_role(DEPARTMENT_HEAD):
        dep_ids = [x.department_id for x in DepartmentLeader.query.filter_by(user_id=current_user.id).all()]
        return [d for d in deps if d.id in dep_ids]
    return deps


def _managed_department_ids(user=None) -> set[int]:
    user = user or current_user
    return {
        row.department_id
        for row in DepartmentLeader.query.filter_by(user_id=user.id).all()
    }


def _can_manage_teacher_professional(
    teacher_id: int,
    *,
    academic_year_id: int | None = None,
) -> bool:
    if teacher_id == current_user.id:
        return True
    if is_admin(current_user) or has_role(METHODIST):
        return True
    if not has_role(DEPARTMENT_HEAD):
        return False
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    year_id = academic_year_id or (current_year.id if current_year else None)
    if not year_id:
        return False
    return any(
        teacher_id in department_teacher_ids(year_id, department_id)
        for department_id in _managed_department_ids()
    )


def _professional_target_from_form() -> User | None:
    teacher_id = request.form.get("teacher_id", type=int)
    if not teacher_id:
        flash("Сначала выберите преподавателя.", "danger")
        return None
    teacher = db.session.get(User, teacher_id)
    if teacher is None:
        flash("Преподаватель не найден.", "danger")
        return None
    if not _can_manage_teacher_professional(teacher.id):
        abort(403)
    return teacher


def _professional_redirect(teacher_id: int | None = None):
    return redirect(
        request.referrer
        or url_for(
            "departments.teacher_profile",
            teacher_id=teacher_id or current_user.id,
        )
    )


def _subject_activity_ids_for_department(dep: Department):
    return [
        link.education_activity_id
        for link in dep.subject_links
        if link.education_activity_id
    ]


DEPARTMENT_LEADER_ROLE_CODES = {
    TEACHER,
    CLASS_TEACHER,
    METHODIST,
    DEPARTMENT_HEAD,
}


def _department_leader_candidates():
    users = (
        User.query
        .filter(
            User.is_active_user.is_(True),
            User.employment_status == "ACTIVE",
        )
        .order_by(
            User.last_name.asc(),
            User.first_name.asc(),
            User.middle_name.asc(),
        )
        .all()
    )
    return [
        user
        for user in users
        if any(
            has_role(role_code, user=user)
            for role_code in DEPARTMENT_LEADER_ROLE_CODES
        )
    ]


def _department_activities():
    links = (
        EducationActivityDepartment.query
        .join(
            EducationActivity,
            EducationActivity.id
            == EducationActivityDepartment.education_activity_id,
        )
        .filter(
            EducationActivityDepartment.is_active.is_(True),
            or_(
                EducationActivityDepartment.valid_from.is_(None),
                EducationActivityDepartment.valid_from <= date.today(),
            ),
            or_(
                EducationActivityDepartment.valid_to.is_(None),
                EducationActivityDepartment.valid_to >= date.today(),
            ),
            EducationActivity.is_active.is_(True),
        )
        .order_by(
            EducationActivity.name.asc(),
            EducationActivity.activity_kind.asc(),
            EducationActivity.id.asc(),
        )
        .all()
    )
    result = defaultdict(list)
    seen = defaultdict(set)
    for link in links:
        activity = link.education_activity
        if activity is None or activity.id in seen[link.department_id]:
            continue
        seen[link.department_id].add(activity.id)
        result[link.department_id].append(activity)
    return result


def _department_for_load(subject_name: Optional[str], grade: Optional[int]):
    normalized = _normalize_subject_name(subject_name)
    if grade and 1 <= grade <= 4 and normalized not in PRIMARY_EXCLUDED:
        return Department.query.filter_by(code="primary").first()

    for dep in Department.query.order_by(Department.id.asc()).all():
        if dep.code == "primary":
            continue
        names = {
            _normalize_subject_name(
                link.education_activity.name
                if link.education_activity
                else (link.subject.name if link.subject else "")
            )
            for link in dep.subject_links
        }
        if normalized in names:
            if dep.code == "philology" and grade and grade < 5:
                continue
            return dep
    return None


def _rebind_all_loads_to_departments():
    for load in TeacherLoad.query.all():
        dep = _department_for_load(load.subject_name, load.grade)
        load.department_id = dep.id if dep else None
    db.session.commit()



def _clean_excel_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalize_person_key(value: str) -> str:
    value = _clean_excel_text(value).lower().replace("ё", "е")
    value = re.sub(r"[^а-яa-z\s.-]", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _fio_initials(value: str):
    parts = [p for p in _normalize_person_key(value).replace(".", " ").split() if p]
    if not parts:
        return None, None, None
    last = parts[0]
    first_i = parts[1][0] if len(parts) > 1 and parts[1] else None
    middle_i = parts[2][0] if len(parts) > 2 and parts[2] else None
    return last, first_i, middle_i


def _find_teacher_by_fio_any_format(fio: str, users_cache: dict):
    fio = _clean_excel_text(fio)
    if not fio:
        return None, "empty_fio"
    key = _normalize_person_key(fio)
    if key in users_cache:
        return users_cache[key], "cache_exact"

    user = None
    try:
        user, reason = find_existing_user(fio=fio)
    except Exception as exc:
        reason = f"find_existing_user_error:{exc}"
    if user:
        users_cache[key] = user
        return user, reason or "find_existing_user"

    last, first_i, middle_i = _fio_initials(fio)
    if not last:
        return None, reason or "no_last_name"

    candidates = []
    for u in User.query.all():
        full = _clean_excel_text(getattr(u, "fio", None) or " ".join([getattr(u, "last_name", "") or "", getattr(u, "first_name", "") or "", getattr(u, "middle_name", "") or ""]))
        u_last, u_first_i, u_middle_i = _fio_initials(full)
        if u_last != last:
            continue
        if first_i and u_first_i and first_i != u_first_i:
            continue
        if middle_i and u_middle_i and middle_i != u_middle_i:
            continue
        candidates.append(u)

    if len(candidates) == 1:
        users_cache[key] = candidates[0]
        return candidates[0], "matched_by_lastname_initials"
    if len(candidates) > 1:
        return None, "ambiguous_lastname_initials"
    return None, reason or "not_found"


def _normalize_load_class_name(value):
    text = _clean_excel_text(value)
    if not text:
        return None
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s*,\s*", ",", text)
    text = re.sub(r"\s*;\s*", ";", text)
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s*\+\s*", "+", text)

    def one(item):
        item = item.strip()
        m = re.match(r"^(\d{1,2})\s*[- ]\s*([А-Яа-яA-Za-zЁё0-9]+)$", item)
        if m:
            return f"{m.group(1)}-{m.group(2).upper()}".replace("Ё", "Е")
        return item.upper().replace("Ё", "Е")

    parts = re.split(r"([,;/+])", text)
    return "".join(one(p) if i % 2 == 0 else p for i, p in enumerate(parts))


def _sheet_building_name(ws_title, filename):
    title = _clean_excel_text(ws_title)
    if title.lower().startswith("учителя"):
        name = re.sub(r"^Учителя\s+", "", title, flags=re.I).strip()
        return name or None
    generic = bool(re.search(r"нагру|учител", title.lower())) or bool(re.match(r"^\d+\s", title))
    if filename and generic:
        stem = Path(filename).stem
        stem = re.sub(r"^(УК|Корпус|Площадка)\s+", "", stem, flags=re.I).strip()
        return stem or title
    return title or None


def _find_header_map(row):
    labels = [_clean_excel_text(x).lower().replace("ё", "е") for x in row]
    aliases = {
        "num": {"#", "№", "номер", "n"},
        "class": {"классы", "класс", "класcы"},
        "subject": {"предмет", "дисциплина"},
        "group": {"группа", "подгруппа"},
        "hours": {"всего", "часы", "часов", "нагрузка"},
    }
    found = {}
    for idx, label in enumerate(labels):
        for key, names in aliases.items():
            if label in names:
                found[key] = idx
    if "class" in found and "subject" in found and "hours" in found:
        return found
    return None


def _parse_float_hours(value):
    text = _clean_excel_text(value).replace(",", ".")
    if not text:
        return 0.0
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(m.group(0)) if m else 0.0



def _clean_excel_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalize_person_key(value: str) -> str:
    value = _clean_excel_text(value).lower().replace("ё", "е")
    value = re.sub(r"[^а-яa-z\s.-]", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _fio_initials(value: str):
    parts = [p for p in _normalize_person_key(value).replace(".", " ").split() if p]
    if not parts:
        return None, None, None
    last = parts[0]
    first_i = parts[1][0] if len(parts) > 1 and parts[1] else None
    middle_i = parts[2][0] if len(parts) > 2 and parts[2] else None
    return last, first_i, middle_i


def _find_teacher_by_fio_any_format(fio: str, users_cache: dict):
    fio = _clean_excel_text(fio)
    if not fio:
        return None, "empty_fio"
    key = _normalize_person_key(fio)
    if key in users_cache:
        return users_cache[key], "cache_exact"

    user = None
    try:
        user, reason = find_existing_user(fio=fio)
    except Exception as exc:
        reason = f"find_existing_user_error:{exc}"
    if user:
        users_cache[key] = user
        return user, reason or "find_existing_user"

    last, first_i, middle_i = _fio_initials(fio)
    if not last:
        return None, reason or "no_last_name"

    candidates = []
    for u in User.query.all():
        full = _clean_excel_text(getattr(u, "fio", None) or " ".join([getattr(u, "last_name", "") or "", getattr(u, "first_name", "") or "", getattr(u, "middle_name", "") or ""]))
        u_last, u_first_i, u_middle_i = _fio_initials(full)
        if u_last != last:
            continue
        if first_i and u_first_i and first_i != u_first_i:
            continue
        if middle_i and u_middle_i and middle_i != u_middle_i:
            continue
        candidates.append(u)

    if len(candidates) == 1:
        users_cache[key] = candidates[0]
        return candidates[0], "matched_by_lastname_initials"
    if len(candidates) > 1:
        return None, "ambiguous_lastname_initials"
    return None, reason or "not_found"


def _normalize_load_class_name(value):
    text = _clean_excel_text(value)
    if not text:
        return None
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s*,\s*", ",", text)
    text = re.sub(r"\s*;\s*", ";", text)
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s*\+\s*", "+", text)

    def one(item):
        item = item.strip()
        m = re.match(r"^(\d{1,2})\s*[- ]\s*([А-Яа-яA-Za-zЁё0-9]+)$", item)
        if m:
            return f"{m.group(1)}-{m.group(2).upper()}".replace("Ё", "Е")
        return item.upper().replace("Ё", "Е")

    parts = re.split(r"([,;/+])", text)
    return "".join(one(p) if i % 2 == 0 else p for i, p in enumerate(parts))


def _sheet_building_name(ws_title, filename):
    title = _clean_excel_text(ws_title)
    if title.lower().startswith("учителя"):
        name = re.sub(r"^Учителя\s+", "", title, flags=re.I).strip()
        return name or None
    generic = bool(re.search(r"нагру|учител", title.lower())) or bool(re.match(r"^\d+\s", title))
    if filename and generic:
        stem = Path(filename).stem
        stem = re.sub(r"^(УК|Корпус|Площадка)\s+", "", stem, flags=re.I).strip()
        return stem or title
    return title or None


def _find_header_map(row):
    labels = [_clean_excel_text(x).lower().replace("ё", "е") for x in row]
    aliases = {
        "num": {"#", "№", "номер", "n"},
        "class": {"классы", "класс", "класcы"},
        "subject": {"предмет", "дисциплина"},
        "group": {"группа", "подгруппа"},
        "hours": {"всего", "часы", "часов", "нагрузка"},
    }
    found = {}
    for idx, label in enumerate(labels):
        for key, names in aliases.items():
            if label in names:
                found[key] = idx
    if "class" in found and "subject" in found and "hours" in found:
        return found
    return None


def _parse_float_hours(value):
    text = _clean_excel_text(value).replace(",", ".")
    if not text:
        return 0.0
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(m.group(0)) if m else 0.0


def _parse_excel_loads(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    created = 0
    updated = 0
    skipped = 0
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    retention_until = None
    if current_year and current_year.end_date:
        try:
            retention_until = current_year.end_date.replace(year=current_year.end_date.year + 7)
        except Exception:
            retention_until = None

    filename = getattr(file_storage, "filename", None)
    users_cache = {_normalize_person_key(u.fio): u for u in User.query.all() if getattr(u, "fio", None)}
    activities = {
        activity.name.lower(): activity
        for activity in list_subject_activities(include_inactive=True)
    }
    buildings = {b.name.lower(): b for b in Building.query.all() if b.name}

    parsed_rows = []
    unmatched_seen = set()
    for ws in wb.worksheets:
        building_name = _sheet_building_name(ws.title, filename)
        building = buildings.get((building_name or "").lower())
        current_teacher = None
        teacher_total = None
        header = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            possible_header = _find_header_map(row)
            if possible_header:
                header = possible_header
                continue
            if not header:
                skipped += 1
                continue

            first = row[header.get("num", 0)] if len(row) > header.get("num", 0) else None
            class_raw = row[header["class"]] if len(row) > header["class"] else None
            subject_raw = row[header["subject"]] if len(row) > header["subject"] else None
            group_raw = row[header.get("group", header["subject"])] if len(row) > header.get("group", header["subject"]) else None
            hours_raw = row[header["hours"]] if len(row) > header["hours"] else None

            first_text = _clean_excel_text(first)
            class_text_raw = _clean_excel_text(class_raw)
            subject_text = _clean_excel_text(subject_raw)
            group_text = _clean_excel_text(group_raw)

            if first_text and not subject_text and not class_text_raw:
                current_teacher, match_reason = _find_teacher_by_fio_any_format(first_text, users_cache)
                teacher_total = _parse_float_hours(hours_raw) if hours_raw not in (None, "") else None
                if not current_teacher:
                    key = (first_text, ws.title, row_idx)
                    if key not in unmatched_seen:
                        register_unmatched_staff(
                            source="teacher_load_import",
                            staff_fio=first_text,
                            import_filename=filename,
                            source_session_key=f"teacher_load:{filename}:{building_name}",
                            role_hint="Учитель",
                            details=f"Лист: {ws.title}; строка: {row_idx}; причина: {match_reason}",
                        )
                        unmatched_seen.add(key)
                    skipped += 1
                continue

            if not current_teacher:
                if any(_clean_excel_text(x) for x in row):
                    skipped += 1
                continue
            if not subject_text:
                skipped += 1
                continue

            class_text = _normalize_load_class_name(class_text_raw)
            group_text = group_text or None
            hours = _parse_float_hours(hours_raw)
            grade = _extract_grade_from_class_text(class_text)
            is_meta_group = bool(class_text and any(sep in class_text for sep in [",", ";", "/", "+"]))
            is_whole_class = (group_text or "").strip().lower() == "весь класс"
            parsed_rows.append({
                "teacher": current_teacher,
                "subject_text": subject_text,
                "class_text": class_text,
                "group_text": group_text,
                "hours": hours,
                "grade": grade,
                "building": building,
                "building_name": building_name,
                "source_sheet": ws.title,
                "row_number": row_idx,
                "is_whole_class": is_whole_class,
                "is_meta_group": is_meta_group,
                "teacher_total_hours": teacher_total,
            })

    if not parsed_rows:
        db.session.rollback()
        return 0, 0, skipped

    building_names = {x["building_name"] for x in parsed_rows if x.get("building_name")}
    building_ids = {x["building"].id for x in parsed_rows if x.get("building")}
    delete_q = TeacherLoad.query
    if building_ids:
        delete_q = delete_q.filter(TeacherLoad.building_id.in_(building_ids))
        delete_q.delete(synchronize_session=False)
    elif building_names:
        delete_q = delete_q.filter(TeacherLoad.building_name.in_(building_names))
        delete_q.delete(synchronize_session=False)
    else:
        TeacherLoad.query.delete()
    db.session.flush()

    for item in parsed_rows:
        subject_text = item["subject_text"]
        activity = activities.get(subject_text.lower())
        if not activity:
            activity, _created = get_or_create_subject_activity(
                subject_text,
                created_by_user_id=getattr(current_user, "id", None),
            )
            activities[subject_text.lower()] = activity
            updated += 1
        load = TeacherLoad(
            teacher_id=item["teacher"].id,
            academic_year_id=current_year.id if current_year else None,
            class_name=item["class_text"],
            group_name=item["group_text"],
            hours=item["hours"],
            grade=item["grade"],
            building_id=item["building"].id if item.get("building") else None,
            building_name=item["building_name"],
            source_sheet=item["source_sheet"],
            row_number=item["row_number"],
            is_whole_class=item["is_whole_class"],
            is_meta_group=item["is_meta_group"],
            teacher_total_hours=item["teacher_total_hours"],
            retention_until=retention_until,
        )
        assign_subject_activity(load, activity)
        db.session.add(load)
        created += 1

    db.session.commit()
    _rebind_all_loads_to_departments()
    return created, updated, skipped

def _control_work_stats(dep: Department, teacher_id=None, academic_year_id=None):
    activity_ids = _subject_activity_ids_for_department(dep)
    if not activity_ids:
        return {
            "total_results": 0,
            "avg_percent": None,
            "by_subject": [],
            "by_teacher": [],
        }

    results_q = (
        db.session.query(
            ControlWorkResult,
            ControlWork,
            EducationActivity,
            ControlWorkAssignment,
            User,
        )
        .join(ControlWork, ControlWork.id == ControlWorkResult.control_work_id)
        .join(
            EducationActivity,
            EducationActivity.id == ControlWork.education_activity_id,
        )
        .outerjoin(
            ControlWorkAssignment,
            db.and_(
                ControlWorkAssignment.control_work_id == ControlWorkResult.control_work_id,
                ControlWorkAssignment.school_class_id == ControlWorkResult.school_class_id,
            ),
        )
        .outerjoin(User, User.id == ControlWorkAssignment.teacher_id)
        .filter(ControlWork.education_activity_id.in_(activity_ids))
    )
    if teacher_id:
        results_q = results_q.filter(ControlWorkAssignment.teacher_id == teacher_id)
    if academic_year_id:
        results_q = results_q.filter(ControlWorkResult.academic_year_id == academic_year_id)

    rows = results_q.all()
    percents = [r.ControlWorkResult.percent for r in rows if r.ControlWorkResult.percent is not None]
    by_subject = defaultdict(list)
    by_teacher = defaultdict(list)

    for row in rows:
        percent = row.ControlWorkResult.percent
        if percent is None:
            continue
        subj_name = (
            row.EducationActivity.name
            if row.EducationActivity
            else "—"
        )
        by_subject[subj_name].append(percent)
        teacher_name = row.User.fio if row.User else "Не указан"
        by_teacher[teacher_name].append(percent)

    subject_stats = [
        {"name": name, "avg": round(sum(vals) / len(vals), 1), "count": len(vals)}
        for name, vals in sorted(by_subject.items())
    ]
    teacher_stats = [
        {"name": name, "avg": round(sum(vals) / len(vals), 1), "count": len(vals)}
        for name, vals in sorted(by_teacher.items())
    ]
    return {
        "total_results": len(percents),
        "avg_percent": round(sum(percents) / len(percents), 1) if percents else None,
        "by_subject": subject_stats,
        "by_teacher": teacher_stats,
    }


def _teacher_scope_user_id():
    teacher_id = request.args.get("teacher_id", type=int)
    if (
        is_admin(current_user)
        or has_role(METHODIST)
        or has_role(DEPARTMENT_HEAD)
    ):
        return teacher_id
    return current_user.id


@departments_bp.before_request
def _bootstrap_departments():
    try:
        _ensure_default_departments()
    except Exception:
        db.session.rollback()


@departments_bp.route("/")
@login_required
def index():
    return redirect(url_for("hub.departments"))


@departments_bp.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if not is_admin(current_user):
        abort(403)

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        code = re.sub(r"[^a-z0-9_]+", "_", (request.form.get("code") or name).strip().lower()).strip("_")
        if not name:
            flash("Укажите название кафедры.", "danger")
            return redirect(url_for("departments.settings"))
        if Department.query.filter(db.func.lower(Department.name) == name.lower()).first():
            flash("Такая кафедра уже существует.", "warning")
            return redirect(url_for("departments.settings"))
        dep = Department(name=name, code=code or None, description=(request.form.get("description") or "").strip() or None)
        db.session.add(dep)
        db.session.commit()
        flash("Кафедра создана.", "success")
        return redirect(url_for("departments.settings"))

    departments = Department.query.order_by(Department.name.asc()).all()
    users = _department_leader_candidates()
    buildings = Building.query.order_by(Building.name.asc()).all()
    return render_template(
        "departments/settings.html",
        departments=departments,
        users=users,
        buildings=buildings,
        department_activities=_department_activities(),
        activity_kind_labels=ACTIVITY_KIND_LABELS,
    )


@departments_bp.route("/settings/<int:department_id>/update", methods=["POST"])
@login_required
def settings_update(department_id):
    if not is_admin(current_user):
        abort(403)
    dep = Department.query.get_or_404(department_id)
    dep.name = (request.form.get("name") or dep.name).strip()
    dep.description = (request.form.get("description") or "").strip() or None
    db.session.commit()
    flash("Настройки кафедры сохранены.", "success")
    return redirect(url_for("departments.settings"))


@departments_bp.route("/settings/<int:department_id>/leader/add", methods=["POST"])
@login_required
def add_leader(department_id):
    if not is_admin(current_user):
        abort(403)
    dep = Department.query.get_or_404(department_id)
    selected_user_ids = []
    raw_user_ids = request.form.getlist("user_ids")
    if request.form.get("user_id"):
        raw_user_ids.append(request.form.get("user_id"))
    candidate_ids = {
        user.id
        for user in _department_leader_candidates()
    }
    for raw in raw_user_ids:
        try:
            value = int(raw)
        except Exception:
            continue
        if value in candidate_ids and value not in selected_user_ids:
            selected_user_ids.append(value)
    building_id = request.form.get("building_id", type=int)
    if not selected_user_ids:
        flash("Выберите учителя для руководства кафедрой.", "danger")
        return redirect(url_for("departments.settings"))

    added = 0
    skipped = 0
    for user_id in selected_user_ids:
        exists = DepartmentLeader.query.filter_by(department_id=dep.id, user_id=user_id, building_id=building_id).first()
        if exists:
            skipped += 1
            continue
        db.session.add(DepartmentLeader(department_id=dep.id, user_id=user_id, building_id=building_id))
        user = User.query.get(user_id)
        if user and user.role != "ADMIN":
            user.role = METHODIST
        added += 1

    db.session.commit()
    if added and skipped:
        flash("Руководители кафедры назначены. Часть уже была добавлена ранее.", "success")
    elif added:
        flash("Руководители кафедры назначены.", "success")
    else:
        flash("Все выбранные руководители уже были добавлены.", "warning")
    return redirect(url_for("departments.settings"))


@departments_bp.route("/settings/leader/<int:leader_id>/delete", methods=["POST"])
@login_required
def delete_leader(leader_id):
    if not is_admin(current_user):
        abort(403)
    leader = DepartmentLeader.query.get_or_404(leader_id)
    db.session.delete(leader)
    db.session.commit()
    flash("Сотрудник исключён из руководства кафедрой.", "success")
    return redirect(url_for("departments.settings"))


@departments_bp.route("/loads", methods=["GET", "POST"])
@login_required
def loads():
    if request.method == "POST":
        flash(
            "Импорт Excel отключён. Нагрузка формируется в разделе "
            "«Учебные планы и нагрузка».",
            "warning",
        )
        return redirect(url_for("departments.loads"))

    q = (request.args.get("q") or "").strip().lower()
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    academic_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    department_id = request.args.get("department_id", type=int)
    subject_id = request.args.get("subject_id", type=int)
    teacher_id = request.args.get("teacher_id", type=int)

    teacher_view_only = current_user.role in {TEACHER, CLASS_TEACHER}
    if teacher_view_only:
        teacher_id = current_user.id
    rows, workload_version = (
        current_department_load_rows(
            academic_year_id,
            department_id=department_id,
            subject_id=subject_id,
            teacher_id=teacher_id,
            query_text=q,
        )
        if academic_year_id
        else ([], None)
    )
    if teacher_view_only:
        available_teachers = [current_user]
    else:
        all_workload_rows, _ = (
            current_department_load_rows(academic_year_id)
            if academic_year_id
            else ([], None)
        )
        available_teacher_ids = sorted({
            row.teacher.id
            for row in all_workload_rows
            if row.teacher is not None
        })
        available_teachers = (
            User.query
            .filter(User.id.in_(available_teacher_ids))
            .order_by(User.last_name.asc(), User.first_name.asc())
            .all()
            if available_teacher_ids
            else []
        )

    teacher_hours = {}
    for item in rows:
        if item.teacher:
            summary = teacher_hours.setdefault(item.teacher.id, {
                "teacher": item.teacher,
                "hours": 0.0,
            })
            summary["hours"] += float(item.hours or 0)

    olympiad_stats = {"total_results": 0, "unique_children": 0, "winners": 0, "prizers": 0, "by_subject": [], "by_teacher": []}

    diagnostics_stats = {"total_results": 0, "sessions_count": 0, "below_basic_count": 0, "avg_percent": None, "by_subject": [], "by_teacher": []}

    return render_template(
        "departments/loads.html",
        rows=rows,
        departments=Department.query.order_by(Department.name.asc()).all(),
        subjects=list_subject_activities(),
        teachers=available_teachers,
        teacher_view_only=teacher_view_only,
        teacher_hours=sorted(
            teacher_hours.values(),
            key=lambda item: (item["teacher"].fio or "").lower(),
        ),
        department_id=department_id,
        subject_id=subject_id,
        teacher_id=teacher_id,
        q=q,
        buildings=Building.query.order_by(Building.name.asc()).all(),
        years=years,
        academic_year_id=academic_year_id,
        olympiad_stats=olympiad_stats,
        diagnostics_stats=diagnostics_stats,
        workload_version=workload_version,
    )


@departments_bp.route("/loads/new", methods=["POST"])
@login_required
def load_new():
    flash("Изменяйте нагрузку только в матрице распределения.", "warning")
    return redirect(url_for("departments.loads"))


@departments_bp.route("/loads/<int:load_id>/update", methods=["POST"])
@login_required
def load_update(load_id):
    flash("Изменяйте нагрузку только в матрице распределения.", "warning")
    return redirect(url_for("departments.loads"))


@departments_bp.route("/loads/<int:load_id>/delete", methods=["POST"])
@login_required
def load_delete(load_id):
    flash("Изменяйте нагрузку только в матрице распределения.", "warning")
    return redirect(url_for("departments.loads"))


@departments_bp.post("/loads/archive/<int:load_id>/delete")
@login_required
def archived_load_delete(load_id):
    if not (
        is_admin(current_user)
        or has_role(METHODIST)
        or has_role(DEPARTMENT_HEAD)
    ):
        abort(403)
    load = TeacherLoad.query.filter_by(
        id=load_id,
        is_archived=True,
    ).first_or_404()
    if load.department_id is not None:
        department = db.session.get(Department, load.department_id)
        if department is not None and not _department_allowed(department):
            abort(403)
    elif not (is_admin(current_user) or has_role(METHODIST)):
        abort(403)

    department_id = load.department_id
    academic_year_id = load.academic_year_id
    db.session.delete(load)
    db.session.commit()
    flash("Строка архивной нагрузки удалена.", "success")
    return redirect(
        request.referrer
        or url_for(
            "departments.summary",
            department_id=department_id,
            academic_year_id=academic_year_id,
        )
    )




def _diagnostics_department_stats(dep: Department, teacher_ids=None, academic_year_id=None, selected_teacher_id=None):
    teacher_ids = teacher_ids or []
    subject_names = {
        _normalize_subject_name(
            link.education_activity.name
            if link.education_activity
            else (link.subject.name if link.subject else "")
        )
        for link in getattr(dep, "subject_links", [])
        if (
            getattr(link, "education_activity", None)
            or getattr(link, "subject", None)
        )
    }

    results = (
        DiagnosticResult.query
        .join(DiagnosticSession, DiagnosticSession.id == DiagnosticResult.session_id)
        .filter(DiagnosticResult.is_final.is_(True))
        .all()
    )

    filtered = []
    for row in results:
        session = row.session
        if academic_year_id and getattr(session, "academic_year_id", None) not in {academic_year_id, None}:
            continue
        subject_name = _normalize_subject_name(getattr(session, "subject", None) or "")
        binding = getattr(row, "teacher_binding", None)
        teacher_id = getattr(binding, "teacher_id", None) if binding else None

        matched_by_subject = bool(subject_name and subject_name in subject_names)
        matched_by_teacher = bool(teacher_id and teacher_id in teacher_ids)
        if not (matched_by_subject or matched_by_teacher):
            continue
        if selected_teacher_id and teacher_id != selected_teacher_id:
            continue
        filtered.append(row)

    def pct(values):
        nums = [float(v) for v in values if v is not None]
        return round(sum(nums) / len(nums), 1) if nums else None

    def score_avg(rows):
        nums = [_safe_float(getattr(row, "total_score", None)) for row in rows]
        nums = [v for v in nums if v is not None]
        return round(sum(nums) / len(nums), 2) if nums else None

    result_mode = _detect_diagnostic_result_mode(filtered)
    below_label = "% двоек" if result_mode == "mark" else "% ниже базового"
    issue_label = "Двоек" if result_mode == "mark" else "Ниже базового"

    by_subject = defaultdict(lambda: {"percents": [], "scores": [], "count": 0, "below": 0, "marks": defaultdict(int)})
    by_teacher = defaultdict(lambda: {"percents": [], "scores": [], "count": 0, "below": 0, "marks": defaultdict(int), "teacher": None})
    level_counts = defaultdict(int)
    mark_counts = defaultdict(int)

    for row in filtered:
        subject_name = (getattr(row.session, "subject", None) or "—").strip() or "—"
        level_label = _diagnostic_level_label(row)
        mark_label = _diagnostic_mark_label(row)
        percent = _safe_float(getattr(row, "percent", None))
        score = _safe_float(getattr(row, "total_score", None))

        below_flag = False
        if result_mode == "mark":
            below_flag = mark_label == "2"
        elif result_mode in {"level", "mixed"}:
            below_flag = level_label == "Ниже базового"

        if level_label:
            level_counts[level_label] += 1
        if mark_label:
            mark_counts[mark_label] += 1

        bucket = by_subject[subject_name]
        bucket["count"] += 1
        if percent is not None:
            bucket["percents"].append(percent)
        if score is not None:
            bucket["scores"].append(score)
        if below_flag:
            bucket["below"] += 1
        if mark_label:
            bucket["marks"][mark_label] += 1

        teacher = getattr(getattr(row, "teacher_binding", None), "teacher", None)
        teacher_key = getattr(teacher, "id", None) or 0
        t_bucket = by_teacher[teacher_key]
        t_bucket["teacher"] = teacher
        t_bucket["count"] += 1
        if percent is not None:
            t_bucket["percents"].append(percent)
        if score is not None:
            t_bucket["scores"].append(score)
        if below_flag:
            t_bucket["below"] += 1
        if mark_label:
            t_bucket["marks"][mark_label] += 1

    def format_bucket(name, bucket, teacher=False):
        teacher_obj = bucket.get("teacher") if teacher else None
        return {
            "name": (teacher_obj.fio if teacher_obj else (name or "Не привязан")) if teacher else name,
            "avg": pct(bucket["percents"]),
            "avg_score": round(sum(bucket["scores"]) / len(bucket["scores"]), 2) if bucket["scores"] else None,
            "count": bucket["count"],
            "below_basic_percent": round(bucket["below"] * 100 / bucket["count"], 1) if bucket["count"] else 0,
            "below_count": bucket["below"],
            "mark2_percent": round(bucket["marks"].get("2", 0) * 100 / bucket["count"], 1) if bucket["count"] else 0,
            "mark2_count": bucket["marks"].get("2", 0),
        }

    by_subject_rows = [format_bucket(name, bucket) for name, bucket in by_subject.items()]
    by_subject_rows.sort(key=lambda x: (((x["avg"] or 0)), x["count"], (x["avg_score"] or 0)), reverse=True)

    by_teacher_rows = [format_bucket(key, bucket, teacher=True) for key, bucket in by_teacher.items()]
    by_teacher_rows.sort(key=lambda x: (((x["avg"] or 0)), x["count"], (x["avg_score"] or 0)), reverse=True)

    percents = [_safe_float(row.percent) for row in filtered if _safe_float(row.percent) is not None]
    scores = [_safe_float(getattr(row, "total_score", None)) for row in filtered if _safe_float(getattr(row, "total_score", None)) is not None]
    unique_subjects = len({row["name"] for row in by_subject_rows if row["name"]})
    unique_teachers = len({row["name"] for row in by_teacher_rows if row["name"] and row["name"] != "Не привязан"})
    has_real_teacher_split = unique_teachers > 1
    show_subject_table = unique_subjects > 1
    show_teacher_table = has_real_teacher_split
    if not show_subject_table and not show_teacher_table and unique_subjects > 1:
        show_subject_table = True
    if not show_subject_table and not show_teacher_table and unique_teachers > 1:
        show_teacher_table = True

    return {
        "total_results": len(filtered),
        "avg_percent": pct(percents),
        "avg_score": round(sum(scores) / len(scores), 2) if scores else None,
        "below_basic_count": level_counts.get("Ниже базового", 0),
        "below_basic_percent": round(level_counts.get("Ниже базового", 0) * 100 / len(filtered), 1) if filtered else 0,
        "mark2_count": mark_counts.get("2", 0),
        "mark2_percent": round(mark_counts.get("2", 0) * 100 / len(filtered), 1) if filtered else 0,
        "levels": dict(level_counts),
        "marks": dict(mark_counts),
        "by_subject": by_subject_rows,
        "by_teacher": by_teacher_rows,
        "result_mode": result_mode,
        "result_mode_label": _result_mode_label(result_mode),
        "issue_count": mark_counts.get("2", 0) if result_mode == "mark" else level_counts.get("Ниже базового", 0),
        "issue_percent": round((mark_counts.get("2", 0) if result_mode == "mark" else level_counts.get("Ниже базового", 0)) * 100 / len(filtered), 1) if filtered else 0,
        "issue_label": issue_label,
        "issue_percent_label": below_label,
        "show_subject_table": show_subject_table,
        "show_teacher_table": show_teacher_table,
        "unique_subjects": unique_subjects,
        "unique_teachers": unique_teachers,
        "has_real_teacher_split": has_real_teacher_split,
    }


@departments_bp.route("/summary")
@login_required
def summary():
    deps = _load_departments_for_user()
    can_manage_department = (
        is_admin(current_user)
        or has_role(METHODIST)
        or has_role(DEPARTMENT_HEAD)
    )
    selected_dep_id = request.args.get("department_id", type=int)
    selected_teacher_id = _teacher_scope_user_id()
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    academic_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    if not can_manage_department and (
        has_role(TEACHER) or has_role(CLASS_TEACHER)
    ):
        deps = [
            department for department in deps
            if current_user.id in department_teacher_ids(
                academic_year_id,
                department.id,
            )
        ]
    dep = None
    if selected_dep_id:
        dep = Department.query.get_or_404(selected_dep_id)
        if not _department_allowed(dep):
            abort(403)
    elif deps:
        dep = deps[0]

    teacher_rows = []
    stats = {"total_results": 0, "avg_percent": None, "by_subject": [], "by_teacher": []}
    olympiad_stats = {"total_results": 0, "unique_children": 0, "winners": 0, "prizers": 0, "by_subject": [], "by_teacher": []}

    diagnostics_stats = {"total_results": 0, "sessions_count": 0, "below_basic_count": 0, "avg_percent": None, "by_subject": [], "by_teacher": []}
    diagnostics_stats = {"total_results": 0, "avg_percent": None, "below_basic_count": 0, "below_basic_percent": 0, "levels": {}, "by_subject": [], "by_teacher": []}
    teacher_ids = []
    mcko_rows = []
    course_rows = []
    department_load_rows = []
    teacher_load_summaries = []
    legacy_load_rows = []
    legacy_load_total = 0
    workload_version = None
    building_id = request.args.get("building_id", type=int)

    if dep:
        if can_manage_department:
            legacy_load_query = TeacherLoad.query.filter(
                TeacherLoad.department_id == dep.id,
            )
            if academic_year_id:
                legacy_load_query = legacy_load_query.filter(or_(
                    TeacherLoad.academic_year_id == academic_year_id,
                    TeacherLoad.academic_year_id.is_(None),
                ))
            if building_id:
                legacy_load_query = legacy_load_query.filter(
                    TeacherLoad.building_id == building_id,
                )
            legacy_load_total = legacy_load_query.count()
            legacy_load_rows = (
                legacy_load_query
                .order_by(
                    TeacherLoad.is_archived.desc(),
                    TeacherLoad.subject_name.asc().nullslast(),
                    TeacherLoad.class_name.asc().nullslast(),
                    TeacherLoad.id.asc(),
                )
                .limit(100)
                .all()
            )
        department_load_rows, workload_version = current_department_load_rows(
            academic_year_id,
            department_id=dep.id,
            building_id=building_id,
        )
        teacher_ids = sorted({
            row.teacher.id
            for row in department_load_rows
            if row.teacher is not None
        })
        if not can_manage_department and (
            has_role(TEACHER) or has_role(CLASS_TEACHER)
        ):
            teacher_ids = [x for x in teacher_ids if x == current_user.id]
            department_load_rows = [
                row for row in department_load_rows
                if row.teacher and row.teacher.id == current_user.id
            ]
        teacher_rows = User.query.filter(User.id.in_(teacher_ids)).order_by(User.last_name.asc(), User.first_name.asc()).all() if teacher_ids else []
        for teacher in teacher_rows:
            teacher_loads = [
                row for row in department_load_rows
                if row.teacher and row.teacher.id == teacher.id
            ]
            teacher_load_summaries.append({
                "teacher": teacher,
                "total": sum(row.hours for row in teacher_loads),
                "subjects": len({row.subject_name for row in teacher_loads}),
                "groups": len({row.group_name for row in teacher_loads if row.group_name}),
            })
        stats = _control_work_stats(dep, teacher_id=selected_teacher_id, academic_year_id=academic_year_id)
        mcko_rows = mcko_results_for_teachers(
            teacher_ids,
            teacher_id=selected_teacher_id,
            academic_year_id=academic_year_id,
        )
        course_q = TeacherCourse.query.filter(
            TeacherCourse.teacher_id.in_(teacher_ids),
            TeacherCourse.is_archived.is_(False),
        ) if teacher_ids else TeacherCourse.query.filter(db.text("0=1"))
        if academic_year_id:
            course_q = course_q.filter(db.or_(TeacherCourse.academic_year_id == academic_year_id, TeacherCourse.academic_year_id.is_(None)))
        if selected_teacher_id:
            course_q = course_q.filter_by(teacher_id=selected_teacher_id)
        course_rows = course_q.order_by(TeacherCourse.start_date.desc().nullslast(), TeacherCourse.created_at.desc()).all()
        attestation_q = TeacherAttestation.query.filter(
            TeacherAttestation.teacher_id.in_(teacher_ids),
            TeacherAttestation.is_archived.is_(False),
        ) if teacher_ids else TeacherAttestation.query.filter(db.text("0=1"))
        if selected_teacher_id:
            attestation_q = attestation_q.filter_by(
                teacher_id=selected_teacher_id,
            )
        attestation_rows = attestation_q.order_by(
            TeacherAttestation.decision_date.desc(),
            TeacherAttestation.created_at.desc(),
        ).all()
        olympiad_stats = olympiad_department_stats(academic_year_id=academic_year_id, department_id=dep.id)
        diagnostics_stats = _diagnostics_department_stats(dep, teacher_ids=teacher_ids, academic_year_id=academic_year_id, selected_teacher_id=selected_teacher_id)
    else:
        attestation_rows = []

    return render_template(
        "departments/summary.html",
        departments=deps,
        dep=dep,
        teachers=teacher_rows,
        all_teachers=User.query.order_by(User.last_name.asc(), User.first_name.asc()).all(),
        selected_teacher_id=selected_teacher_id,
        stats=stats,
        mcko_rows=mcko_rows,
        course_rows=course_rows,
        buildings=Building.query.order_by(Building.name.asc()).all(),
        building_id=building_id,
        years=years,
        academic_year_id=academic_year_id,
        olympiad_stats=olympiad_stats,
        diagnostics_stats=diagnostics_stats,
        workload_version=workload_version,
        teacher_load_summaries=teacher_load_summaries,
        department_load_rows=department_load_rows,
        legacy_load_rows=legacy_load_rows,
        legacy_load_total=legacy_load_total,
        mcko_level_labels=MCKO_LEVEL_LABELS,
        mcko_subjects=list_subject_activities(),
        attestation_rows=attestation_rows,
        attestation_category_labels=ATTESTATION_CATEGORY_LABELS,
        can_manage_professional=bool(
            dep
            and can_manage_department
        ),
    )


@departments_bp.get("/teacher/profile")
@login_required
def my_teacher_profile():
    return redirect(url_for(
        "departments.teacher_profile",
        teacher_id=current_user.id,
        academic_year_id=request.args.get("academic_year_id", type=int),
    ))


@departments_bp.get("/teacher/cabinet")
@login_required
def teacher_cabinet():
    return redirect(url_for(
        "departments.my_teacher_profile",
        academic_year_id=request.args.get("academic_year_id", type=int),
    ))


@departments_bp.get("/teachers")
@login_required
def teacher_registry():
    return redirect(url_for(
        "departments.summary",
        academic_year_id=request.args.get("academic_year_id", type=int),
    ))


@departments_bp.get("/teachers/<int:teacher_id>")
@login_required
def teacher_profile(teacher_id):
    if not _can_manage_teacher_professional(teacher_id):
        abort(403)
    teacher = User.query.get_or_404(teacher_id)
    years = AcademicYear.query.order_by(
        AcademicYear.start_date.desc().nullslast(),
        AcademicYear.name.desc(),
    ).all()
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    academic_year_id = request.args.get("academic_year_id", type=int) or (
        current_year.id if current_year else None
    )
    load_rows, workload_version = (
        current_department_load_rows(
            academic_year_id,
            teacher_id=teacher.id,
        )
        if academic_year_id
        else ([], None)
    )
    totals = defaultdict(float)
    for row in load_rows:
        totals[row.plan_kind] += row.hours
        totals["TOTAL"] += row.hours
    mcko_rows = mcko_results_for_teachers(
        [teacher.id],
        teacher_id=teacher.id,
        academic_year_id=academic_year_id,
    )
    course_rows = (
        TeacherCourse.query
        .filter_by(teacher_id=teacher.id, is_archived=False)
        .order_by(
            TeacherCourse.end_date.desc().nullslast(),
            TeacherCourse.start_date.desc().nullslast(),
            TeacherCourse.created_at.desc(),
        )
        .all()
    )
    attestation_rows = (
        TeacherAttestation.query
        .filter_by(teacher_id=teacher.id, is_archived=False)
        .order_by(
            TeacherAttestation.decision_date.desc(),
            TeacherAttestation.created_at.desc(),
        )
        .all()
    )
    debt_rows = (
        Debt.query
        .filter_by(created_by_user_id=teacher.id)
        .order_by(
            Debt.detected_date.desc(),
            Debt.created_at.desc(),
        )
        .limit(100)
        .all()
    )
    incident_rows = (
        Incident.query
        .filter_by(author_id=teacher.id)
        .order_by(
            Incident.occurred_at.desc(),
            Incident.id.desc(),
        )
        .limit(100)
        .all()
    )
    return render_template(
        "departments/teacher_profile.html",
        teacher=teacher,
        years=years,
        academic_year_id=academic_year_id,
        workload_version=workload_version,
        workload_status_label=(
            WORKLOAD_APPROVAL_STATUS_LABELS.get(
                workload_version.workload_approval_status,
                workload_version.workload_approval_status,
            )
            if workload_version else None
        ),
        workload_is_approved=bool(
            workload_version
            and workload_version.workload_approval_status == "APPROVED"
        ),
        load_rows=load_rows,
        totals=totals,
        mcko_rows=mcko_rows,
        mcko_level_labels=MCKO_LEVEL_LABELS,
        mcko_subjects=list_subject_activities(),
        course_rows=course_rows,
        attestation_rows=attestation_rows,
        debt_rows=debt_rows,
        incident_rows=incident_rows,
        attestation_category_labels=ATTESTATION_CATEGORY_LABELS,
        can_edit_professional=_can_manage_teacher_professional(
            teacher.id,
            academic_year_id=academic_year_id,
        ),
        is_self_profile=teacher.id == current_user.id,
        can_add_incident=(
            teacher.id == current_user.id
            and has_permission("incident_add")
        ),
        today=date.today(),
    )


@departments_bp.route("/teacher/mcko/add", methods=["POST"])
@login_required
def add_mcko():
    teacher = _professional_target_from_form()
    if teacher is None:
        return _professional_redirect()
    passed_at_raw = request.form.get("passed_at") or None
    passed_at = datetime.strptime(passed_at_raw, "%Y-%m-%d").date() if passed_at_raw else None
    subject_id = request.form.get("subject_id", type=int)
    level = normalize_mcko_level(request.form.get("level"))
    certificate_number = (
        request.form.get("certificate_number") or ""
    ).strip() or None
    result_text = (request.form.get("result_text") or "").strip() or None
    if not subject_id or not passed_at or not level:
        flash("Укажите предмет, дату диагностики и уровень МЦКО.", "danger")
        return redirect(request.referrer or url_for("departments.summary"))
    activity = get_subject_activity(subject_id)
    if activity is None:
        flash("Выберите предмет из единого реестра.", "danger")
        return redirect(request.referrer or url_for("departments.summary"))
    selected_year_id = request.form.get("academic_year_id", type=int)
    current_year = (
        AcademicYear.query.get(selected_year_id)
        if selected_year_id
        else AcademicYear.query.filter_by(is_current=True).first()
    )
    retention_until = None
    if current_year and current_year.end_date:
        try:
            retention_until = current_year.end_date.replace(year=current_year.end_date.year + 7)
        except Exception:
            retention_until = None
    result = TeacherMckoResult(
        teacher_id=teacher.id,
        academic_year_id=current_year.id if current_year else None,
        passed_at=passed_at,
        expires_at=mcko_expires_at(passed_at),
        level=level,
        certificate_number=certificate_number,
        result_text=result_text,
        entry_source=professional_entry_source(
            teacher_id=teacher.id,
            actor_id=current_user.id,
        ),
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
        retention_until=retention_until,
    )
    assign_subject_activity(result, activity)
    db.session.add(result)
    record_professional_change(
        result,
        change_kind=CHANGE_CREATED,
        actor_id=current_user.id,
    )
    db.session.commit()
    flash("Результат МЦКО сохранён.", "success")
    return _professional_redirect(teacher.id)


@departments_bp.route("/teacher/course/add", methods=["POST"])
@login_required
def add_course():
    teacher = _professional_target_from_form()
    if teacher is None:
        return _professional_redirect()
    start_date_raw = request.form.get("start_date") or None
    end_date_raw = request.form.get("end_date") or None
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("Укажите название курса.", "danger")
        return _professional_redirect(teacher.id)
    current_year = AcademicYear.query.filter_by(is_current=True).first()
    retention_until = None
    if current_year and current_year.end_date:
        try:
            retention_until = current_year.end_date.replace(year=current_year.end_date.year + 7)
        except Exception:
            retention_until = None
    db.session.add(TeacherCourse(
        teacher_id=teacher.id,
        academic_year_id=current_year.id if current_year else None,
        title=title,
        provider=(request.form.get("provider") or "").strip() or None,
        hours=request.form.get("hours", type=float),
        start_date=datetime.strptime(start_date_raw, "%Y-%m-%d").date() if start_date_raw else None,
        end_date=datetime.strptime(end_date_raw, "%Y-%m-%d").date() if end_date_raw else None,
        notes=(request.form.get("notes") or "").strip() or None,
        retention_until=retention_until,
    ))
    db.session.commit()
    flash("Курс повышения квалификации сохранён.", "success")
    return _professional_redirect(teacher.id)


@departments_bp.post("/teacher/attestation/add")
@login_required
def add_attestation():
    teacher = _professional_target_from_form()
    if teacher is None:
        return _professional_redirect()
    category = (request.form.get("category") or "").strip().upper()
    decision_date_raw = (request.form.get("decision_date") or "").strip()
    valid_until_raw = (request.form.get("valid_until") or "").strip()
    is_indefinite = (
        request.form.get("is_indefinite") or ""
    ).strip().lower() in {"1", "true", "yes", "on"}
    if category not in ATTESTATION_CATEGORY_LABELS or not decision_date_raw:
        flash("Укажите вид аттестации и дату решения.", "danger")
        return _professional_redirect(teacher.id)
    decision_date = datetime.strptime(decision_date_raw, "%Y-%m-%d").date()
    if category == "POSITION_COMPLIANCE":
        is_indefinite = False
    valid_until = (
        datetime.strptime(valid_until_raw, "%Y-%m-%d").date()
        if valid_until_raw
        else None
    )
    if is_indefinite:
        valid_until = None
    elif category == "POSITION_COMPLIANCE" and valid_until is None:
        valid_until = position_compliance_due_at(
            None,
            last_decision_date=decision_date,
        )
    elif category in {"FIRST", "HIGHEST"} and valid_until is None:
        flash("Для срочной категории укажите срок или отметьте её бессрочной.", "danger")
        return _professional_redirect(teacher.id)
    if valid_until and valid_until < decision_date:
        flash("Срок действия не может завершаться раньше даты решения.", "danger")
        return _professional_redirect(teacher.id)
    record = TeacherAttestation(
        teacher_id=teacher.id,
        category=category,
        position_title=(request.form.get("position_title") or "").strip() or None,
        decision_date=decision_date,
        valid_until=valid_until,
        is_indefinite=is_indefinite,
        order_number=(request.form.get("order_number") or "").strip() or None,
        notes=(request.form.get("notes") or "").strip() or None,
        entry_source=professional_entry_source(
            teacher_id=teacher.id,
            actor_id=current_user.id,
        ),
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    db.session.add(record)
    record_professional_change(
        record,
        change_kind=CHANGE_CREATED,
        actor_id=current_user.id,
    )
    db.session.commit()
    flash("Сведения об аттестации сохранены.", "success")
    return _professional_redirect(teacher.id)


def _archive_professional_record(record, *, label: str):
    if not _can_manage_teacher_professional(record.teacher_id):
        abort(403)
    record.is_archived = True
    if hasattr(record, "updated_by_user_id"):
        record.updated_by_user_id = current_user.id
    if hasattr(record, "updated_at"):
        record.updated_at = datetime.utcnow()
    if isinstance(record, (TeacherMckoResult, TeacherAttestation)):
        record_professional_change(
            record,
            change_kind=CHANGE_ARCHIVED,
            actor_id=current_user.id,
        )
    db.session.commit()
    flash(f"{label} удалён из действующих сведений.", "success")
    return _professional_redirect(record.teacher_id)


@departments_bp.post("/teacher/mcko/<int:record_id>/archive")
@login_required
def archive_mcko(record_id):
    return _archive_professional_record(
        TeacherMckoResult.query.get_or_404(record_id),
        label="Результат МЦКО",
    )


@departments_bp.post("/teacher/course/<int:record_id>/archive")
@login_required
def archive_course(record_id):
    return _archive_professional_record(
        TeacherCourse.query.get_or_404(record_id),
        label="Курс",
    )


@departments_bp.post("/teacher/attestation/<int:record_id>/archive")
@login_required
def archive_attestation(record_id):
    return _archive_professional_record(
        TeacherAttestation.query.get_or_404(record_id),
        label="Запись об аттестации",
    )

