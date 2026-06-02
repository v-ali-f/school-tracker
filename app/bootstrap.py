import logging
import os
from sqlalchemy import inspect, text
from sqlalchemy.schema import CreateColumn
from openpyxl import load_workbook

from app.core.extensions import db

logger = logging.getLogger(__name__)


def ensure_runtime_schema():
    inspector = inspect(db.engine)

    # Важно для старых баз: поле is_annulled в олимпиадах в модели NOT NULL.
    # SQLite не поддерживает ALTER COLUMN, поэтому для локальной тестовой базы
    # выполняем только безопасные ADD COLUMN / UPDATE, а PostgreSQL оставляем
    # с полноценной настройкой DEFAULT / NOT NULL.
    try:
        olymp_cols = {c["name"] for c in inspector.get_columns("olympiad_result")}
        is_sqlite = db.engine.dialect.name == "sqlite"

        if "is_annulled" not in olymp_cols:
            db.session.execute(text("ALTER TABLE olympiad_result ADD COLUMN is_annulled BOOLEAN DEFAULT FALSE"))
            db.session.execute(text("UPDATE olympiad_result SET is_annulled = FALSE WHERE is_annulled IS NULL"))

            if not is_sqlite:
                db.session.execute(text("ALTER TABLE olympiad_result ALTER COLUMN is_annulled SET DEFAULT FALSE"))
                db.session.execute(text("ALTER TABLE olympiad_result ALTER COLUMN is_annulled SET NOT NULL"))

            db.session.commit()
        else:
            db.session.execute(text("UPDATE olympiad_result SET is_annulled = FALSE WHERE is_annulled IS NULL"))

            if not is_sqlite:
                db.session.execute(text("ALTER TABLE olympiad_result ALTER COLUMN is_annulled SET DEFAULT FALSE"))
                db.session.execute(text("ALTER TABLE olympiad_result ALTER COLUMN is_annulled SET NOT NULL"))

            db.session.commit()
    except Exception:
        db.session.rollback()
        logger.exception("bootstrap: olympiad_result.is_annulled safe migration failed")
    try:
        cols = {c["name"] for c in inspector.get_columns("child")}
    except Exception:
        logger.exception("bootstrap: failed to inspect columns of 'child'")
        cols = set()

    needed = {
        "ovz_doc_number": "ALTER TABLE child ADD COLUMN ovz_doc_number VARCHAR(100)",
        "ovz_doc_date": "ALTER TABLE child ADD COLUMN ovz_doc_date DATE",
        "disability_ipra": "ALTER TABLE child ADD COLUMN disability_ipra VARCHAR(255)",
        "status": "ALTER TABLE child ADD COLUMN status VARCHAR(30) DEFAULT 'ACTIVE'",
        "archived_at": "ALTER TABLE child ADD COLUMN archived_at TIMESTAMP",
    }

    for name, sql in needed.items():
        if name not in cols:
            try:
                db.session.execute(text(sql))
                db.session.commit()
            except Exception:
                db.session.rollback()
                logger.exception("bootstrap: ALTER failed for child.%s (%s)", name, sql)

    for table_name, additions in {
        "control_work": {
            "work_kind": "ALTER TABLE control_work ADD COLUMN work_kind VARCHAR(50) DEFAULT 'control'",
            "grade5_percent": "ALTER TABLE control_work ADD COLUMN grade5_percent INTEGER DEFAULT 85",
            "grade4_percent": "ALTER TABLE control_work ADD COLUMN grade4_percent INTEGER DEFAULT 65",
            "grade3_percent": "ALTER TABLE control_work ADD COLUMN grade3_percent INTEGER DEFAULT 45",
            "academic_year_id": "ALTER TABLE control_work ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE control_work ADD COLUMN retention_until DATE",
            "dictation_grade5_spelling_max": "ALTER TABLE control_work ADD COLUMN dictation_grade5_spelling_max INTEGER DEFAULT 0",
            "dictation_grade5_punctuation_max": "ALTER TABLE control_work ADD COLUMN dictation_grade5_punctuation_max INTEGER DEFAULT 0",
            "dictation_grade4_spelling_max": "ALTER TABLE control_work ADD COLUMN dictation_grade4_spelling_max INTEGER DEFAULT 2",
            "dictation_grade4_punctuation_max": "ALTER TABLE control_work ADD COLUMN dictation_grade4_punctuation_max INTEGER DEFAULT 2",
            "dictation_grade3_spelling_max": "ALTER TABLE control_work ADD COLUMN dictation_grade3_spelling_max INTEGER DEFAULT 4",
            "dictation_grade3_punctuation_max": "ALTER TABLE control_work ADD COLUMN dictation_grade3_punctuation_max INTEGER DEFAULT 4",
            "dictation_use_grammar_errors": "ALTER TABLE control_work ADD COLUMN dictation_use_grammar_errors BOOLEAN DEFAULT FALSE",
            "dictation_use_corrections": "ALTER TABLE control_work ADD COLUMN dictation_use_corrections BOOLEAN DEFAULT FALSE",
            "updated_at": "ALTER TABLE control_work ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "updated_by": "ALTER TABLE control_work ADD COLUMN updated_by INTEGER",
            "manual_status": "ALTER TABLE control_work ADD COLUMN manual_status VARCHAR(30)",
            "is_archived": "ALTER TABLE control_work ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        },
        "control_work_log": {
            "user_id": "ALTER TABLE control_work_log ADD COLUMN user_id INTEGER",
            "event_type": "ALTER TABLE control_work_log ADD COLUMN event_type VARCHAR(50)",
            "title": "ALTER TABLE control_work_log ADD COLUMN title VARCHAR(255)",
            "old_value": "ALTER TABLE control_work_log ADD COLUMN old_value TEXT",
            "new_value": "ALTER TABLE control_work_log ADD COLUMN new_value TEXT",
            "details": "ALTER TABLE control_work_log ADD COLUMN details TEXT",
            "created_at": "ALTER TABLE control_work_log ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "control_work_task": {
            "description": "ALTER TABLE control_work_task ADD COLUMN description VARCHAR(255)",
            "topic": "ALTER TABLE control_work_task ADD COLUMN topic VARCHAR(255)",
        },
        "control_work_result": {
            "assignment_id": "ALTER TABLE control_work_result ADD COLUMN assignment_id INTEGER",
            "grade5_percent": "ALTER TABLE control_work_result ADD COLUMN grade5_percent INTEGER DEFAULT 85",
            "grade4_percent": "ALTER TABLE control_work_result ADD COLUMN grade4_percent INTEGER DEFAULT 65",
            "grade3_percent": "ALTER TABLE control_work_result ADD COLUMN grade3_percent INTEGER DEFAULT 45",
            "academic_year_id": "ALTER TABLE control_work_result ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE control_work_result ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE control_work_result ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            "result_status": "ALTER TABLE control_work_result ADD COLUMN result_status VARCHAR(20) DEFAULT 'present'",
            "is_absent": "ALTER TABLE control_work_result ADD COLUMN is_absent BOOLEAN DEFAULT FALSE",
            "dictation_mark": "ALTER TABLE control_work_result ADD COLUMN dictation_mark INTEGER",
            "grammar_mark": "ALTER TABLE control_work_result ADD COLUMN grammar_mark INTEGER",
            "final_mark": "ALTER TABLE control_work_result ADD COLUMN final_mark INTEGER",
            "spelling_errors": "ALTER TABLE control_work_result ADD COLUMN spelling_errors INTEGER",
            "punctuation_errors": "ALTER TABLE control_work_result ADD COLUMN punctuation_errors INTEGER",
            "grammar_errors": "ALTER TABLE control_work_result ADD COLUMN grammar_errors INTEGER",
            "corrections_count": "ALTER TABLE control_work_result ADD COLUMN corrections_count INTEGER",
            "teacher_comment": "ALTER TABLE control_work_result ADD COLUMN teacher_comment TEXT",
        },
        "academic_year": {
            "is_closed": "ALTER TABLE academic_year ADD COLUMN is_closed BOOLEAN DEFAULT FALSE",
            "is_archived": "ALTER TABLE academic_year ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            "updated_at": "ALTER TABLE academic_year ADD COLUMN updated_at TIMESTAMP",
        },
        "school_class": {
            "is_active": "ALTER TABLE school_class ADD COLUMN is_active BOOLEAN DEFAULT TRUE",
            "is_archived": "ALTER TABLE school_class ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        },
        "child_enrollment": {
            "transfer_order_number": "ALTER TABLE child_enrollment ADD COLUMN transfer_order_number VARCHAR(100)",
            "transfer_order_date": "ALTER TABLE child_enrollment ADD COLUMN transfer_order_date DATE",
        },
        "child_parent": {
            "transfer_order_number": "ALTER TABLE child_parent ADD COLUMN transfer_order_number VARCHAR(100)",
            "transfer_order_date": "ALTER TABLE child_parent ADD COLUMN transfer_order_date DATE",
        },
        "child_social": {
            "is_single_mother": "ALTER TABLE child_social ADD COLUMN is_single_mother BOOLEAN DEFAULT FALSE",
            "is_single_father": "ALTER TABLE child_social ADD COLUMN is_single_father BOOLEAN DEFAULT FALSE",
            "is_repeat_year": "ALTER TABLE child_social ADD COLUMN is_repeat_year BOOLEAN DEFAULT FALSE",
            "is_svo_family": "ALTER TABLE child_social ADD COLUMN is_svo_family BOOLEAN DEFAULT FALSE",
        },
        "parent": {
            "retention_until": "ALTER TABLE parent ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE parent ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            "created_at": "ALTER TABLE parent ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "document": {
            "academic_year_id": "ALTER TABLE document ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE document ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE document ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
            "is_hidden_by_retention": "ALTER TABLE document ADD COLUMN is_hidden_by_retention BOOLEAN DEFAULT FALSE",
            "is_deleted_soft": "ALTER TABLE document ADD COLUMN is_deleted_soft BOOLEAN DEFAULT FALSE",
            "deleted_at": "ALTER TABLE document ADD COLUMN deleted_at TIMESTAMP",
            "deleted_by": "ALTER TABLE document ADD COLUMN deleted_by INTEGER",
        },
        "teacher_load": {
            "academic_year_id": "ALTER TABLE teacher_load ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE teacher_load ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE teacher_load ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        },
        "teacher_mcko_result": {
            "academic_year_id": "ALTER TABLE teacher_mcko_result ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE teacher_mcko_result ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE teacher_mcko_result ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        },
        "teacher_course": {
            "academic_year_id": "ALTER TABLE teacher_course ADD COLUMN academic_year_id INTEGER",
            "retention_until": "ALTER TABLE teacher_course ADD COLUMN retention_until DATE",
            "is_archived": "ALTER TABLE teacher_course ADD COLUMN is_archived BOOLEAN DEFAULT FALSE",
        },
        "olympiad_subject_mapping": {
            "olympiad_name": "ALTER TABLE olympiad_subject_mapping ADD COLUMN olympiad_name VARCHAR(255)",
            "linked_subject_ids": "ALTER TABLE olympiad_subject_mapping ADD COLUMN linked_subject_ids TEXT",
            "grade_from": "ALTER TABLE olympiad_subject_mapping ADD COLUMN grade_from INTEGER",
            "grade_to": "ALTER TABLE olympiad_subject_mapping ADD COLUMN grade_to INTEGER",
            "priority": "ALTER TABLE olympiad_subject_mapping ADD COLUMN priority INTEGER DEFAULT 100",
        },
        "olympiad_result": {
            "teacher_binding_status": "ALTER TABLE olympiad_result ADD COLUMN teacher_binding_status VARCHAR(30)",
            "teacher_binding_source": "ALTER TABLE olympiad_result ADD COLUMN teacher_binding_source VARCHAR(30)",
            "teacher_binding_reason": "ALTER TABLE olympiad_result ADD COLUMN teacher_binding_reason TEXT",
        },
        "user": {
            "employment_status": 'ALTER TABLE "user" ADD COLUMN employment_status VARCHAR(30) DEFAULT \'ACTIVE\'',
            "dismissal_date": 'ALTER TABLE "user" ADD COLUMN dismissal_date DATE',
            "archived_at": 'ALTER TABLE "user" ADD COLUMN archived_at TIMESTAMP',
            "last_login_at": 'ALTER TABLE "user" ADD COLUMN last_login_at TIMESTAMP',
            "last_seen_at": 'ALTER TABLE "user" ADD COLUMN last_seen_at TIMESTAMP',
            "active_days_count": 'ALTER TABLE "user" ADD COLUMN active_days_count INTEGER NOT NULL DEFAULT 0',
            "notify_incident_mode": 'ALTER TABLE "user" ADD COLUMN notify_incident_mode VARCHAR(20) NOT NULL DEFAULT \'all\'',
            "notify_task_mode": 'ALTER TABLE "user" ADD COLUMN notify_task_mode VARCHAR(20) NOT NULL DEFAULT \'all\'',
            "task_notifications_enabled": 'ALTER TABLE "user" ADD COLUMN task_notifications_enabled BOOLEAN NOT NULL DEFAULT TRUE',
            "task_email_enabled": 'ALTER TABLE "user" ADD COLUMN task_email_enabled BOOLEAN NOT NULL DEFAULT TRUE',
            "task_notify_only_important": 'ALTER TABLE "user" ADD COLUMN task_notify_only_important BOOLEAN NOT NULL DEFAULT FALSE',
        },
    }.items():
        try:
            existing = {c["name"] for c in inspector.get_columns(table_name)}
        except Exception:
            logger.exception("bootstrap: failed to inspect columns of %s", table_name)
            existing = set()

        for name, sql in additions.items():
            if name not in existing:
                try:
                    db.session.execute(text(sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                    logger.exception("bootstrap: ALTER failed for %s.%s (%s)", table_name, name, sql)

    inspector = inspect(db.engine)

    # olympiad_subject_mapping used to have a UNIQUE constraint on olympiad_subject_name.
    # The new logic allows several rows with the same olympiad subject for different
    # grade ranges and linked load subjects, so we must drop the legacy constraint.
    try:
        unique_constraints = {c.get("name") for c in inspector.get_unique_constraints("olympiad_subject_mapping")}
    except Exception:
        logger.exception("bootstrap: failed to inspect unique_constraints of olympiad_subject_mapping")
        unique_constraints = set()
    if "olympiad_subject_mapping_olympiad_subject_name_key" in unique_constraints:
        try:
            db.session.execute(text(
                "ALTER TABLE olympiad_subject_mapping DROP CONSTRAINT IF EXISTS olympiad_subject_mapping_olympiad_subject_name_key"
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
            logger.exception("bootstrap: DROP CONSTRAINT failed for olympiad_subject_mapping")

    # Recreate regular indexes if needed after dropping the legacy unique constraint.
    for sql in [
        "CREATE INDEX IF NOT EXISTS ix_olympiad_subject_mapping_olympiad_subject_name ON olympiad_subject_mapping (olympiad_subject_name)",
        "CREATE INDEX IF NOT EXISTS ix_olympiad_subject_mapping_olympiad_name ON olympiad_subject_mapping (olympiad_name)",
        "CREATE INDEX IF NOT EXISTS ix_olympiad_subject_mapping_subject_id ON olympiad_subject_mapping (subject_id)",
        "CREATE INDEX IF NOT EXISTS ix_olympiad_subject_mapping_department_id ON olympiad_subject_mapping (department_id)",
        "CREATE INDEX IF NOT EXISTS ix_incident_status ON incident (status)",
        "CREATE INDEX IF NOT EXISTS ix_incident_assignee_id ON incident (assignee_id)",
        "CREATE INDEX IF NOT EXISTS ix_incident_author_id ON incident (author_id)",
        "CREATE INDEX IF NOT EXISTS ix_incident_occurred_at ON incident (occurred_at)",
        "CREATE INDEX IF NOT EXISTS ix_incident_category ON incident (category)",
        "CREATE INDEX IF NOT EXISTS ix_incident_note_incident_id ON incident_note (incident_id)",
        "CREATE INDEX IF NOT EXISTS ix_incident_note_author_id ON incident_note (author_id)",
        "CREATE INDEX IF NOT EXISTS ix_incident_status_occurred ON incident (status, occurred_at DESC)",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_saved_view_user_scope_name ON saved_view (user_id, scope, name)",
        # s85: индексы под колокольчик/задачи/контингент/события
        "CREATE INDEX IF NOT EXISTS ix_incident_notification_user_unread ON incident_notification (user_id, is_read, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_task_notification_user_unread ON task_notification (user_id, is_read, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS ix_task_responsible_status ON task (responsible_user_id, status)",
        "CREATE INDEX IF NOT EXISTS ix_task_status_deadline ON task (status, deadline_at)",
        'CREATE INDEX IF NOT EXISTS ix_user_employment_status ON "user" (employment_status)',
        "CREATE INDEX IF NOT EXISTS ix_school_class_year_archived ON school_class (academic_year_id, is_archived)",
        "CREATE INDEX IF NOT EXISTS ix_child_enrollment_status ON child_enrollment (status)",
        "CREATE INDEX IF NOT EXISTS ix_child_enrollment_class_status ON child_enrollment (school_class_id, status)",
        "CREATE INDEX IF NOT EXISTS ix_child_status ON child (status)",
        "CREATE INDEX IF NOT EXISTS ix_child_events_from_class ON child_events (from_class)",
        "CREATE INDEX IF NOT EXISTS ix_child_events_event_type_from_class ON child_events (event_type, from_class)",
        # Password reset tokens
        """CREATE TABLE IF NOT EXISTS password_reset_token (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES "user"(id) ON DELETE CASCADE,
            token_hash VARCHAR(64) NOT NULL UNIQUE,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL,
            used_at TIMESTAMP,
            request_ip VARCHAR(64)
        )""",
        "CREATE INDEX IF NOT EXISTS ix_password_reset_token_user_id ON password_reset_token (user_id)",
        "CREATE INDEX IF NOT EXISTS ix_password_reset_token_token_hash ON password_reset_token (token_hash)",
        # s96: множественные исполнители инцидента
        """CREATE TABLE IF NOT EXISTS incident_assignee (
            incident_id INTEGER NOT NULL REFERENCES incident(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES "user"(id) ON DELETE CASCADE,
            added_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            added_by_id INTEGER REFERENCES "user"(id),
            PRIMARY KEY (incident_id, user_id)
        )""",
        "CREATE INDEX IF NOT EXISTS ix_incident_assignee_user ON incident_assignee (user_id)",
        # backfill: переносим текущих assignee_id в junction-таблицу (идемпотентно).
        """INSERT INTO incident_assignee (incident_id, user_id, added_at)
           SELECT i.id, i.assignee_id, CURRENT_TIMESTAMP FROM incident i
           WHERE i.assignee_id IS NOT NULL AND NOT EXISTS (
             SELECT 1 FROM incident_assignee ia
             WHERE ia.incident_id = i.id AND ia.user_id = i.assignee_id
           )""",
    ]:
        try:
            db.session.execute(text(sql))
            db.session.commit()
        except Exception:
            db.session.rollback()
            logger.exception("bootstrap: CREATE INDEX failed (%s)", sql)

    for table in db.metadata.sorted_tables:
        try:
            existing_cols = {c["name"] for c in inspector.get_columns(table.name)}
        except Exception:
            logger.exception("bootstrap: failed to inspect columns of %s (metadata sweep)", table.name)
            continue

        for column in table.columns:
            if column.name in existing_cols or getattr(column, "primary_key", False):
                continue

            try:
                col_sql = str(CreateColumn(column).compile(dialect=db.engine.dialect))
                db.session.execute(text(f'ALTER TABLE "{table.name}" ADD COLUMN {col_sql}'))
                db.session.commit()
            except Exception:
                db.session.rollback()
                logger.exception("bootstrap: metadata-sweep ADD COLUMN failed for %s.%s", table.name, column.name)


def seed_olympiad_subject_mappings(app):
    try:
        from app.models import DepartmentSubject, OlympiadSubjectMapping, Subject
    except Exception:
        logger.exception("bootstrap: failed to import olympiad-mapping models for seeding")
        return 0

    with app.app_context():
        try:
            if OlympiadSubjectMapping.query.count() > 0:
                return 0
        except Exception:
            logger.exception("bootstrap: failed to read OlympiadSubjectMapping count")
            return 0

        candidate_paths = [
            os.path.join(app.root_path, "..", "data", "olympiad_subjects_vsoh.xlsx"),
            os.path.join(app.root_path, "..", "data_seed", "olympiad_subjects_vsoh.xlsx"),
        ]
        seed_path = next(
            (os.path.abspath(p) for p in candidate_paths if os.path.exists(os.path.abspath(p))),
            None,
        )
        if not seed_path:
            return 0

        try:
            wb = load_workbook(seed_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
        except Exception:
            logger.exception("bootstrap: failed to read seed file %s", seed_path)
            return 0

        def norm(v):
            return " ".join(str(v or "").replace("ё", "е").replace("Ё", "Е").split()).strip().lower()

        subjects = Subject.query.all()
        subjects_by_name = {norm(s.name): s for s in subjects}

        def match_subject(raw_school_subjects: str):
            variants = [
                norm(part)
                for part in str(raw_school_subjects or "").replace(";", ",").split(",")
                if norm(part)
            ]
            for item in variants:
                if item in subjects_by_name:
                    return subjects_by_name[item]

            for item in variants:
                for subj in subjects:
                    subj_norm = norm(subj.name)
                    if subj_norm == item or subj_norm in item or item in subj_norm:
                        return subj
            return None

        created = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue

            olympiad_name = str(row[0] or "").strip()
            school_subjects = str(row[1] or "").strip()

            if not olympiad_name or not school_subjects:
                continue

            subject = match_subject(school_subjects)
            if not subject:
                continue

            dep_link = DepartmentSubject.query.filter_by(subject_id=subject.id).first()
            mapping = OlympiadSubjectMapping.query.filter_by(
                olympiad_subject_name=olympiad_name
            ).first()
            if mapping:
                continue

            mapping = OlympiadSubjectMapping(
                olympiad_subject_name=olympiad_name,
                subject_id=subject.id,
                department_id=dep_link.department_id if dep_link else None,
                comment=f"Базовая загрузка из перечня ВСОШ: {school_subjects}",
                is_active=True,
            )
            db.session.add(mapping)
            created += 1

        if created:
            db.session.commit()

        return created
