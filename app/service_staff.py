
from collections import defaultdict
from math import ceil
from datetime import date, datetime
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
from xml.sax.saxutils import escape
from zipfile import ZipFile, ZIP_DEFLATED

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for, current_app
from flask_login import current_user, login_required
from openpyxl import Workbook
from sqlalchemy.orm import joinedload, subqueryload
from sqlalchemy import or_

from app.core.extensions import db
from app.core.pagination import paginate_list, resolve_pagination
from app.models import AcademicYear, Building, Child, SchoolClass, User, Incident, IncidentChild
from app.models.service_staff import (
    ServiceAssignment,
    ServiceAssignmentHistory,
    ServiceResponsible,
    ServiceSpecialist,
    ServiceSpecialistBuilding,
    ServiceSpecialistSpecialization,
    ServiceSpecialization,
    ServiceActivityType,
    ServiceCyclegram,
    ServiceCyclegramEntry,
    ServiceCyclegramHistory,
    ServicePresentation,
    ServicePresentationBlock,
    ServicePresentationHistory,
    ServiceRateNorm,
)
from app.permissions import can_view_child_basic, has_role, is_admin


service_staff_bp = Blueprint("service_staff", __name__, url_prefix="/service-staff")

SERVICE_ROLE_LABELS = {
    "METHODIST": "Методист",
    "OLIGOPHRENOPEDAGOG": "Олигофренопедагог",
    "SOCIAL_PEDAGOGUE": "Социальный педагог",
    "LOGOPEDIST": "Учитель-логопед",
    "PSYCHOLOGIST": "Педагог-психолог",
    "DEFECTOLOGIST": "Учитель-дефектолог",
    "TUTOR": "Тьютор",
    "ASSISTANT": "Ассистент",
}
SERVICE_ROLE_CODES = set(SERVICE_ROLE_LABELS)

def _user_role_label(code):
    return SERVICE_ROLE_LABELS.get((code or '').upper(), code or '—')

def _ensure_specialist_for_user(user, commit=False):
    if not user:
        return None
    specialist = ServiceSpecialist.query.filter_by(user_id=user.id).first()
    if not specialist:
        specialist = ServiceSpecialist(user_id=user.id, is_active=True)
        db.session.add(specialist)
    specialist.last_name = user.last_name or specialist.last_name or ''
    specialist.first_name = user.first_name or specialist.first_name or ''
    specialist.middle_name = user.middle_name or None
    specialist.phone = user.phone or specialist.phone
    specialist.email = user.email or specialist.email
    specialist.position_title = specialist.position_title or _user_role_label(getattr(user, 'role', None))
    if specialist.is_active is None:
        specialist.is_active = True
    if commit:
        db.session.commit()
    return specialist

def _sync_service_specialists_from_users():
    changed = False
    for user in User.query.order_by(User.last_name.asc(), User.first_name.asc()).all():
        if (getattr(user, 'role', None) or '').upper() in SERVICE_ROLE_CODES:
            specialist = ServiceSpecialist.query.filter_by(user_id=user.id).first()
            if not specialist:
                _ensure_specialist_for_user(user)
                changed = True
    if changed:
        db.session.flush()

SPECIALIZATION_SEED = [
    ("psychologist", "Психолог", 10),
    ("pedagog_psychologist", "Педагог-психолог", 20),
    ("social_pedagogue", "Социальный педагог", 30),
    ("speech_therapist", "Учитель-логопед", 40),
    ("defectologist", "Учитель-дефектолог", 50),
]
ASSIGNMENT_STATUS_CHOICES = [
    ("ACTIVE", "Активно"),
    ("FINISHED", "Завершено"),
    ("ARCHIVED", "Архив"),
]


ASSIGNMENT_ROLE_CHOICES = [
    ("social_pedagogue", "Социальный педагог", "SOCIAL_PEDAGOGUE"),
    ("pedagog_psychologist", "Педагог-психолог", "PSYCHOLOGIST"),
    ("speech_therapist", "Учитель-логопед", "LOGOPEDIST"),
    ("defectologist", "Учитель-дефектолог", "DEFECTOLOGIST"),
    ("tutor", "Тьютор", "TUTOR"),
    ("assistant", "Ассистент", "ASSISTANT"),
]
ASSIGNMENT_ROLE_LABELS = {key: label for key, label, _ in ASSIGNMENT_ROLE_CHOICES}
ASSIGNMENT_ROLE_CODES = {key: code for key, _, code in ASSIGNMENT_ROLE_CHOICES}


PRESENTATION_STATUS_CHOICES = [
    ("DRAFT", "Черновик"),
    ("IN_PROGRESS", "В работе"),
    ("REVIEW", "На проверке"),
    ("APPROVED", "Согласовано"),
    ("EXPORTED", "Выгружено"),
    ("ARCHIVED", "Архив"),
]
PRESENTATION_BLOCK_STATUS_CHOICES = [
    ("NOT_STARTED", "Не начат"),
    ("IN_PROGRESS", "В работе"),
    ("FILLED", "Заполнен"),
    ("CHECKED", "Проверен"),
]
PRESENTATION_BASIS_CHOICES = [
    ("PPK", "ППк"),
    ("CPMPK", "ЦПМПК"),
    ("SUPPORT", "Сопровождение"),
    ("REPEAT", "Повторное представление"),
    ("OTHER", "Иное"),
]
PRESENTATION_DEFAULT_BLOCKS = [
    ("general", "Общие сведения об обучающемся", 10, "AUTO", True),
    ("learning", "Сведения об условиях и результатах обучения", 20, "MIXED", True),
    ("development_start_now", "Характеристика развития на момент поступления и на момент подготовки представления", 30, "MIXED", True),
    ("dynamic", "Динамика развития и динамика освоения программы", 40, "MANUAL", True),
    ("individual", "Индивидуальные особенности, влияющие на обучение", 50, "MANUAL", True),
    ("family", "Отношение семьи к трудностям обучающегося", 60, "MANUAL", True),
    ("support", "Организация коррекционно-развивающей и психолого-педагогической помощи", 70, "MIXED", True),
    ("prevention", "Характеристики взросления, профилактическая работа и дополнительная информация", 80, "MANUAL", False),
    ("conclusion", "Выводы, рекомендации и подписи участников", 90, "MANUAL", True),
]
PRESENTATION_HINTS = {
    "general": "Проверьте автоподстановку ФИО, даты рождения, класса, здания, даты зачисления, программы и формы обучения.",
    "learning": "Опишите успеваемость, учебную мотивацию, поведение в классе, условия обучения и значимые трудности.",
    "development_start_now": "Сравните стартовые данные и текущее состояние. Укажите выраженность положительной/отрицательной динамики.",
    "dynamic": "Опишите изменения в освоении программы, развитии речи, внимания, памяти, коммуникации и саморегуляции.",
    "individual": "Укажите особенности, влияющие на обучение: темп деятельности, истощаемость, особенности поведения, контакта, восприятия.",
    "family": "Отразите позицию семьи, участие родителей, выполнение рекомендаций, взаимодействие со школой.",
    "support": "Укажите специалистов сопровождения, виды помощи, периодичность, результаты и текущие ограничения.",
    "prevention": "Опишите профилактическую работу, особенности социализации, взросления, дополнительные сведения.",
    "conclusion": "Сформулируйте выводы, рекомендации, образовательный маршрут и перечень участников представления.",
}


def _seed_specializations():
    changed = False
    for code, name, sort_order in SPECIALIZATION_SEED:
        row = ServiceSpecialization.query.filter_by(code=code).first()
        if not row:
            db.session.add(ServiceSpecialization(code=code, name=name, sort_order=sort_order, is_active=True))
            changed = True
        else:
            if row.name != name:
                row.name = name
                changed = True
            if row.sort_order != sort_order:
                row.sort_order = sort_order
                changed = True
            if not row.is_active:
                row.is_active = True
                changed = True
    if changed:
        db.session.commit()


def _parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def _status_label(value):
    return dict(ASSIGNMENT_STATUS_CHOICES).get((value or "").upper(), value or "—")


def _is_service_responsible(user=None):
    user = user or current_user
    if not getattr(user, "is_authenticated", False):
        return False
    if has_role("ADMIN", user=user):
        return True
    specialist = ServiceSpecialist.query.filter_by(user_id=user.id).first()
    if not specialist:
        return False
    return ServiceResponsible.query.filter_by(specialist_id=specialist.id, is_active=True).first() is not None


def _is_service_methodist(user=None):
    user = user or current_user
    linked = _linked_specialist(user)
    return bool(linked and getattr(user, "is_authenticated", False) and has_role("METHODIST", user=user))


def _linked_specialist(user=None):
    user = user or current_user
    if not getattr(user, "is_authenticated", False):
        return None
    return ServiceSpecialist.query.filter_by(user_id=user.id, is_active=True).first()


def _can_view_module(user=None):
    user = user or current_user
    return bool(
        getattr(user, "is_authenticated", False)
        and (
            has_role("ADMIN", user=user)
            or _is_service_methodist(user)
            or _is_service_responsible(user)
            or _linked_specialist(user) is not None
        )
    )


def _can_edit_module(user=None):
    user = user or current_user
    return bool(
        getattr(user, "is_authenticated", False)
        and (has_role("ADMIN", user=user) or _is_service_responsible(user))
    )


def _can_view_specialist(specialist: ServiceSpecialist, user=None):
    user = user or current_user
    if _can_edit_module(user) or _is_service_methodist(user):
        return True
    own = _linked_specialist(user)
    return bool(own and own.id == specialist.id)


def can_view_child_service_block(child, user=None):
    user = user or current_user
    if not getattr(user, "is_authenticated", False):
        return False
    if has_role("ADMIN", user=user) or _is_service_methodist(user) or _is_service_responsible(user):
        return True
    if _linked_specialist(user) is not None:
        return True
    return can_view_child_basic(child)


def _require_view():
    _seed_specializations()
    _sync_service_specialists_from_users()
    if not _can_view_module():
        abort(403)


def _require_edit():
    _seed_specializations()
    _sync_service_specialists_from_users()
    if not _can_edit_module():
        abort(403)


def _visible_assignments_query(user=None):
    user = user or current_user
    q = ServiceAssignment.query.options(
        joinedload(ServiceAssignment.child).joinedload(Child.enrollments),
        joinedload(ServiceAssignment.specialist).joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceAssignment.specialist).joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceAssignment.building),
        joinedload(ServiceAssignment.created_by),
    )
    if _can_edit_module(user) or _is_service_methodist(user):
        return q
    linked = _linked_specialist(user)
    if linked:
        return q.filter(ServiceAssignment.specialist_id == linked.id)
    return q.filter(ServiceAssignment.id == -1)


def _specialist_dashboard_stats():
    _seed_specializations()
    specialists = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceSpecialist.building_links).joinedload(ServiceSpecialistBuilding.building),
    ).all()
    active_specialists = [row for row in specialists if row.is_active]
    spec_counts = defaultdict(int)
    building_ids = set()
    for specialist in active_specialists:
        for link in specialist.specialization_links:
            if link.specialization:
                spec_counts[link.specialization.name] += 1
        for link in specialist.building_links:
            if link.building_id:
                building_ids.add(link.building_id)
    assignment_rows = ServiceAssignment.query.all()
    active_assignments = [x for x in assignment_rows if (x.status or "").upper() == "ACTIVE"]
    return {
        "total": len(active_specialists),
        "responsible_count": ServiceResponsible.query.filter_by(is_active=True).count(),
        "buildings_count": len(building_ids),
        "spec_counts": sorted(spec_counts.items(), key=lambda x: x[0]),
        "active_assignments": len(active_assignments),
        "children_count": len({x.child_id for x in active_assignments}),
        "finished_count": len([x for x in assignment_rows if (x.status or "").upper() == "FINISHED"]),
    }


def _user_choices():
    return User.query.order_by(User.last_name.asc(), User.first_name.asc(), User.middle_name.asc()).all()


def _building_choices():
    return Building.query.order_by(Building.name.asc()).all()


def _specialization_choices():
    _seed_specializations()
    return ServiceSpecialization.query.filter_by(is_active=True).order_by(ServiceSpecialization.sort_order.asc(), ServiceSpecialization.name.asc()).all()


def _specialist_choices():
    rows = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization)
    ).order_by(ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc(), ServiceSpecialist.middle_name.asc()).all()
    rows = [x for x in rows if x.is_active]
    if _can_edit_module() or _is_service_methodist():
        return rows
    linked = _linked_specialist()
    return [x for x in rows if linked and x.id == linked.id]


def _class_choices():
    return SchoolClass.query.order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc()).all()


def _extract_parallel_from_class_name(value):
    value = (value or '').strip()
    digits = []
    for ch in value:
        if ch.isdigit():
            digits.append(ch)
        elif digits:
            break
    return int(''.join(digits)) if digits else None


def _child_choices():
    rows = Child.query.order_by(Child.last_name.asc(), Child.first_name.asc(), Child.middle_name.asc()).all()
    if _can_edit_module() or _is_service_methodist() or _linked_specialist() is not None:
        return rows
    visible = []
    for child in rows:
        if can_view_child_basic(child):
            visible.append(child)
    return visible


def _school_class_choices():
    rows = SchoolClass.query.order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc()).all()
    result = []
    seen = set()
    for row in rows:
        key = (row.name or '').strip()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _parallel_choices():
    values = []
    seen = set()
    for row in _school_class_choices():
        parallel = _extract_parallel_from_class_name(getattr(row, 'name', None))
        if parallel is None or parallel in seen:
            continue
        seen.add(parallel)
        values.append(parallel)
    return values


def _apply_specialist_form(specialist: ServiceSpecialist):
    specialist.user_id = int(request.form.get("user_id")) if request.form.get("user_id") else None
    if specialist.user_id:
        linked_user = User.query.get(specialist.user_id)
        specialist.last_name = (request.form.get("last_name") or "").strip() or (linked_user.last_name if linked_user else "")
        specialist.first_name = (request.form.get("first_name") or "").strip() or (linked_user.first_name if linked_user else "")
        specialist.middle_name = (request.form.get("middle_name") or "").strip() or (linked_user.middle_name if linked_user else None)
        specialist.position_title = (request.form.get("position_title") or "").strip() or _user_role_label(getattr(linked_user, 'role', None))
        specialist.phone = (request.form.get("phone") or "").strip() or (linked_user.phone if linked_user else None)
        specialist.email = (request.form.get("email") or "").strip() or (linked_user.email if linked_user else None)
    else:
        specialist.last_name = (request.form.get("last_name") or "").strip()
        specialist.first_name = (request.form.get("first_name") or "").strip()
        specialist.middle_name = (request.form.get("middle_name") or "").strip() or None
        specialist.position_title = (request.form.get("position_title") or "").strip() or None
        specialist.phone = (request.form.get("phone") or "").strip() or None
        specialist.email = (request.form.get("email") or "").strip() or None
    rate_raw = (request.form.get("rate_value") or "").strip().replace(",", ".")
    specialist.rate_value = float(rate_raw) if rate_raw else None
    specialist.is_active = request.form.get("is_active") == "1"
    specialist.admin_comment = (request.form.get("admin_comment") or "").strip() or None

    building_ids =     building_ids = [int(x) for x in request.form.getlist("building_ids") if str(x).isdigit()]
    main_building_id = int(request.form.get("main_building_id")) if request.form.get("main_building_id") else None
    specialist.main_building_id = main_building_id if main_building_id in building_ids or not building_ids else building_ids[0]

    selected_spec_ids = [int(x) for x in request.form.getlist("specialization_ids") if str(x).isdigit()]
    if not specialist.last_name or not specialist.first_name:
        raise ValueError("Укажите фамилию и имя специалиста.")
    if not selected_spec_ids:
        raise ValueError("Выберите хотя бы одну специализацию.")
    if not building_ids:
        raise ValueError("Выберите хотя бы одно здание закрепления.")

    specialist.specialization_links[:] = [
        ServiceSpecialistSpecialization(specialization_id=spec_id) for spec_id in dict.fromkeys(selected_spec_ids)
    ]
    specialist.building_links[:] = [
        ServiceSpecialistBuilding(building_id=building_id, is_main=(building_id == specialist.main_building_id))
        for building_id in dict.fromkeys(building_ids)
    ]


def _registry_rows():
    q = (request.args.get("q") or "").strip()
    building_id = request.args.get("building_id") or ""
    specialization_id = request.args.get("specialization_id") or ""
    is_active = request.args.get("is_active") or ""
    elevated = request.args.get("elevated") or ""

    rows = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.user),
        joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceSpecialist.building_links).joinedload(ServiceSpecialistBuilding.building),
        joinedload(ServiceSpecialist.responsible_links),
    ).order_by(ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc(), ServiceSpecialist.middle_name.asc()).all()

    if not (_can_edit_module() or _is_service_methodist()):
        linked = _linked_specialist()
        rows = [x for x in rows if linked and x.id == linked.id]

    def matches(row):
        if q:
            hay = " ".join([
                row.fio.lower(),
                (row.position_title or "").lower(),
                (row.specializations_text or "").lower(),
            ])
            if q.lower() not in hay:
                return False
        if building_id and not any(link.building_id == int(building_id) for link in row.building_links):
            return False
        if specialization_id and not any(link.specialization_id == int(specialization_id) for link in row.specialization_links):
            return False
        if is_active == "1" and not row.is_active:
            return False
        if is_active == "0" and row.is_active:
            return False
        if elevated == "1" and not row.is_responsible:
            return False
        if elevated == "0" and row.is_responsible:
            return False
        return True

    filtered = [row for row in rows if matches(row)]
    page, per_page = resolve_pagination()
    items, pagination = paginate_list(filtered, page=page, per_page=per_page)
    return items, pagination


def _normalize_text(value):
    return (value or "").strip().lower().replace("ё", "е")


def _specialist_matches_assignment_role(specialist: ServiceSpecialist, role_key: str):
    role_code = ASSIGNMENT_ROLE_CODES.get(role_key)
    if not specialist or not specialist.is_active or not role_code:
        return False
    user_role = (getattr(getattr(specialist, "user", None), "role", None) or "").upper()
    if user_role == role_code:
        return True
    expected = _normalize_text(ASSIGNMENT_ROLE_LABELS.get(role_key))
    position_text = _normalize_text(getattr(specialist, "position_title", None))
    if expected and expected in position_text:
        return True
    spec_names = [_normalize_text(link.specialization.name) for link in getattr(specialist, 'specialization_links', []) if getattr(link, 'specialization', None)]
    if role_key == 'pedagog_psychologist' and any('психолог' in name for name in spec_names):
        return True
    if role_key == 'social_pedagogue' and any('социаль' in name for name in spec_names):
        return True
    if role_key == 'speech_therapist' and any('логопед' in name for name in spec_names):
        return True
    if role_key == 'defectologist' and any('дефектолог' in name for name in spec_names):
        return True
    return False


def _assignment_role_specialists():
    rows = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.user),
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
    ).order_by(ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc(), ServiceSpecialist.middle_name.asc()).all()
    rows = [x for x in rows if x.is_active]

    pedagogical_roles = {"TEACHER", "CLASS_TEACHER", "CURATOR", "TUTOR", "ASSISTANT", "METHODIST"}
    pedagogue_rows = []
    for user in User.query.order_by(User.last_name.asc(), User.first_name.asc(), User.middle_name.asc()).all():
        if not getattr(user, 'is_active_user', True):
            continue
        if (getattr(user, 'role', None) or '').upper() not in pedagogical_roles:
            continue
        specialist = _ensure_specialist_for_user(user)
        if specialist:
            pedagogue_rows.append(specialist)
    if pedagogue_rows:
        db.session.flush()

    result = {}
    for role_key, _, _ in ASSIGNMENT_ROLE_CHOICES:
        if role_key in {'tutor', 'assistant'}:
            pool = pedagogue_rows or rows
        else:
            pool = rows
        unique = []
        seen = set()
        for row in pool:
            if not row or not row.is_active or row.id in seen:
                continue
            seen.add(row.id)
            unique.append(row)
        result[role_key] = unique
    return result


def _multi_assignment_payload_from_request():
    child_id = request.form.get("child_id", type=int)
    building_id = request.form.get("building_id", type=int)
    basis = (request.form.get("basis") or "").strip() or None
    comment = (request.form.get("comment") or "").strip() or None
    status = (request.form.get("status") or "ACTIVE").upper()
    start_date = _parse_date(request.form.get("start_date"))
    end_date = _parse_date(request.form.get("end_date"))
    incident_id = request.form.get("incident_id", type=int) or None
    enabled_roles = [x for x in request.form.getlist("enabled_roles") if x in ASSIGNMENT_ROLE_LABELS]

    if not child_id:
        raise ValueError("Выберите ребенка.")
    if status not in {x[0] for x in ASSIGNMENT_STATUS_CHOICES}:
        raise ValueError("Некорректный статус сопровождения.")
    if not enabled_roles:
        raise ValueError("Выберите хотя бы одного специалиста сопровождения.")

    specialists_by_role = _assignment_role_specialists()
    assignments = []
    for role_key in enabled_roles:
        specialist_id = request.form.get(f"role_{role_key}_specialist_id", type=int)
        if not specialist_id:
            raise ValueError(f'По роли «{ASSIGNMENT_ROLE_LABELS.get(role_key, role_key)}» выберите специалиста.')
        allowed_ids = {row.id for row in specialists_by_role.get(role_key, [])}
        if specialist_id not in allowed_ids:
            raise ValueError(f'Специалист по роли «{ASSIGNMENT_ROLE_LABELS.get(role_key, role_key)}» выбран некорректно.')
        assignments.append(ServiceAssignment(
            child_id=child_id,
            specialist_id=specialist_id,
            building_id=building_id,
            role_title=ASSIGNMENT_ROLE_LABELS.get(role_key),
            basis=basis,
            comment=comment,
            status=status,
            start_date=start_date,
            end_date=end_date,
            incident_id=incident_id,
            created_by_user_id=current_user.id if getattr(current_user, 'is_authenticated', False) else None,
        ))
    return assignments


def _form_state_from_request():
    return {
        'child_id': request.form.get('child_id', ''),
        'building_id': request.form.get('building_id', ''),
        'status': (request.form.get('status') or 'ACTIVE').upper(),
        'start_date': request.form.get('start_date', ''),
        'end_date': request.form.get('end_date', ''),
        'basis': request.form.get('basis', ''),
        'comment': request.form.get('comment', ''),
        'incident_id': request.form.get('incident_id', ''),
        'enabled_roles': [x for x in request.form.getlist('enabled_roles') if x in ASSIGNMENT_ROLE_LABELS],
        'role_specialists': {key: request.form.get(f'role_{key}_specialist_id', '') for key in ASSIGNMENT_ROLE_LABELS},
    }


def _multi_form_state_from_assignments(assignments):
    assignments = list(assignments or [])
    primary = assignments[0] if assignments else None
    role_specialists = {key: '' for key in ASSIGNMENT_ROLE_LABELS}
    enabled_roles = []
    label_to_key = {label: key for key, label in ASSIGNMENT_ROLE_LABELS.items()}
    for row in assignments:
        role_key = label_to_key.get((row.role_title or '').strip())
        if not role_key:
            continue
        enabled_roles.append(role_key)
        role_specialists[role_key] = str(row.specialist_id or '')
    enabled_roles = list(dict.fromkeys(enabled_roles))
    return {
        'child_id': str(primary.child_id) if primary and primary.child_id else '',
        'building_id': str(primary.building_id) if primary and primary.building_id else '',
        'status': (primary.status if primary and primary.status else 'ACTIVE'),
        'start_date': primary.start_date.isoformat() if primary and primary.start_date else '',
        'end_date': primary.end_date.isoformat() if primary and primary.end_date else '',
        'basis': primary.basis if primary and primary.basis else '',
        'comment': primary.comment if primary and primary.comment else '',
        'incident_id': str(primary.incident_id) if primary and primary.incident_id else '',
        'enabled_roles': enabled_roles,
        'role_specialists': role_specialists,
    }


def _save_multi_assignment_group(existing_rows=None):
    existing_rows = list(existing_rows or [])
    child_id = request.form.get('child_id', type=int)
    if not child_id:
        raise ValueError('Выберите ребенка.')

    status = (request.form.get('status') or 'ACTIVE').upper()
    if status not in {x[0] for x in ASSIGNMENT_STATUS_CHOICES}:
        raise ValueError('Некорректный статус сопровождения.')

    building_id = request.form.get('building_id', type=int)
    basis = (request.form.get('basis') or '').strip() or None
    comment = (request.form.get('comment') or '').strip() or None
    incident_id = request.form.get('incident_id', type=int) or None
    start_date = _parse_date(request.form.get('start_date'))
    end_date = _parse_date(request.form.get('end_date'))
    enabled_roles = [x for x in request.form.getlist('enabled_roles') if x in ASSIGNMENT_ROLE_LABELS]
    if not enabled_roles:
        raise ValueError('Выберите хотя бы одного специалиста сопровождения.')

    specialists_by_role = _assignment_role_specialists()
    existing_by_role = {}
    label_to_key = {label: key for key, label in ASSIGNMENT_ROLE_LABELS.items()}
    for row in existing_rows:
        role_key = label_to_key.get((row.role_title or '').strip())
        if role_key and role_key not in existing_by_role:
            existing_by_role[role_key] = row

    kept_ids = set()
    for role_key in enabled_roles:
        specialist_id = request.form.get(f'role_{role_key}_specialist_id', type=int)
        if not specialist_id:
            raise ValueError(f'По роли «{ASSIGNMENT_ROLE_LABELS.get(role_key, role_key)}» выберите специалиста.')
        allowed_ids = {row.id for row in specialists_by_role.get(role_key, [])}
        if specialist_id not in allowed_ids:
            raise ValueError(f'Специалист по роли «{ASSIGNMENT_ROLE_LABELS.get(role_key, role_key)}» выбран некорректно.')

        row = existing_by_role.get(role_key)
        old_status = row.status if row else None
        if not row:
            row = ServiceAssignment(
                child_id=child_id,
                created_by_user_id=current_user.id if getattr(current_user, 'is_authenticated', False) else None,
            )
            db.session.add(row)
        row.child_id = child_id
        row.specialist_id = specialist_id
        row.building_id = building_id
        row.role_title = ASSIGNMENT_ROLE_LABELS.get(role_key)
        row.basis = basis
        row.comment = comment
        row.status = status
        row.start_date = start_date
        row.end_date = end_date
        row.incident_id = incident_id
        if not row.created_by_user_id:
            row.created_by_user_id = current_user.id if getattr(current_user, 'is_authenticated', False) else None
        db.session.flush()
        kept_ids.add(row.id)
        _log_assignment_history(row, old_status, row.status, 'Обновление карточки сопровождения' if old_status is not None else 'Создание назначения')

    for row in existing_rows:
        if row.id in kept_ids:
            continue
        db.session.delete(row)

    return len(kept_ids)


def _assignment_display_group_data():
    query = _visible_assignments_query().order_by(ServiceAssignment.updated_at.desc(), ServiceAssignment.id.desc())
    rows = query.all()

    q = (request.args.get("q") or "").strip().lower()
    specialist_id = request.args.get("specialist_id", type=int)
    building_id = request.args.get("building_id", type=int)
    role_title = (request.args.get("role_title") or "").strip().lower()
    class_id = request.args.get("class_id", type=int)
    grade = request.args.get("grade", type=int)
    category = (request.args.get("category") or "").strip().lower()
    status = (request.args.get("status") or "").strip().upper()
    active_only = request.args.get("active_only") == "1"

    def _category_matches(child, val):
        mapping = {
            "ovz": bool(getattr(child, "is_ovz", False)),
            "vshu": bool(getattr(child, "is_vshu", False)),
            "low": bool(getattr(child, "is_low", False)),
            "az": bool(getattr(child, "is_az", False)),
            "disabled": bool(getattr(child, "is_disabled", False)),
        }
        return mapping.get(val, True)

    filtered = []
    for row in rows:
        child = row.child
        specialist = row.specialist
        school_class = child.current_class if child else None
        if q:
            hay = " ".join([
                child.fio.lower() if child else "",
                specialist.fio.lower() if specialist else "",
                (row.role_title or "").lower(),
            ])
            if q not in hay:
                continue
        if specialist_id and row.specialist_id != specialist_id:
            continue
        if building_id and row.building_id != building_id and not (child and child.current_building and child.current_building.id == building_id):
            continue
        if role_title and role_title not in (row.role_title or "").lower():
            continue
        if class_id and not (school_class and school_class.id == class_id):
            continue
        if grade and not (school_class and school_class.grade == grade):
            continue
        if category and child and not _category_matches(child, category):
            continue
        if status and (row.status or "").upper() != status:
            continue
        if active_only and (row.status or "").upper() != "ACTIVE":
            continue
        filtered.append(row)

    groups = {}
    order = []
    for row in filtered:
        if not row.child_id:
            continue
        if row.child_id not in groups:
            order.append(row.child_id)
            groups[row.child_id] = {
                'child': row.child,
                'class_name': row.child.current_class_name if row.child else '—',
                'class_teacher': ((row.child.current_class.teacher_user.fio if getattr(row.child.current_class, 'teacher_user', None) else (row.child.current_class.teacher_name if getattr(row.child.current_class, 'teacher_name', None) else '—')) if row.child and row.child.current_class else '—'),
                'building_name': (row.building.short_name or row.building.name) if row.building else (row.child_building_name or '—'),
                'start_date': row.start_date,
                'status': row.status,
                'status_label': row.status_label,
                'edit_id': row.id,
                'delete_id': row.id,
                'roles': {label: '—' for label in ASSIGNMENT_ROLE_LABELS.values()},
                'updated_at': row.updated_at or datetime.min,
            }
        group = groups[row.child_id]
        role_label = row.role_title or '—'
        if role_label in group['roles']:
            group['roles'][role_label] = row.specialist.fio if row.specialist else '—'
        if (row.updated_at or datetime.min) > (group.get('updated_at') or datetime.min):
            group['edit_id'] = row.id
            group['start_date'] = row.start_date
            group['status'] = row.status
            group['status_label'] = row.status_label
            group['updated_at'] = row.updated_at or datetime.min

    grouped_rows = [groups[key] for key in order]
    page, per_page = resolve_pagination()
    items, pagination = paginate_list(grouped_rows, page=page, per_page=per_page)
    stats = {
        'children_count': len(grouped_rows),
        'active_count': len([x for x in filtered if (x.status or '').upper() == 'ACTIVE']),
        'finished_count': len([x for x in filtered if (x.status or '').upper() == 'FINISHED']),
        'specialists_count': len({x.specialist_id for x in filtered if x.specialist_id}),
    }
    return items, pagination, stats


def _apply_assignment_form(assignment: ServiceAssignment):
    child_id = request.form.get("child_id", type=int)
    specialist_id = request.form.get("specialist_id", type=int)
    building_id = request.form.get("building_id", type=int)
    role_title = (request.form.get("role_title") or "").strip() or None
    basis = (request.form.get("basis") or "").strip() or None
    comment = (request.form.get("comment") or "").strip() or None
    status = (request.form.get("status") or "ACTIVE").upper()
    start_date = _parse_date(request.form.get("start_date"))
    end_date = _parse_date(request.form.get("end_date"))

    if not child_id:
        raise ValueError("Выберите ребенка.")
    if not specialist_id:
        raise ValueError("Выберите специалиста.")
    if status not in {x[0] for x in ASSIGNMENT_STATUS_CHOICES}:
        raise ValueError("Некорректный статус сопровождения.")

    assignment.child_id = child_id
    assignment.specialist_id = specialist_id
    assignment.building_id = building_id
    assignment.role_title = role_title
    assignment.basis = basis
    assignment.comment = comment
    assignment.status = status
    assignment.start_date = start_date
    assignment.end_date = end_date
    if not assignment.created_by_user_id:
        assignment.created_by_user_id = current_user.id


def _log_assignment_history(assignment: ServiceAssignment, old_status: str, new_status: str, comment: str = None):
    db.session.add(ServiceAssignmentHistory(
        assignment=assignment,
        changed_by_user_id=current_user.id if getattr(current_user, "is_authenticated", False) else None,
        old_status=old_status,
        new_status=new_status,
        comment=comment,
    ))


def _assignment_registry_data():
    query = _visible_assignments_query().order_by(ServiceAssignment.updated_at.desc(), ServiceAssignment.id.desc())
    rows = query.all()

    q = (request.args.get("q") or "").strip().lower()
    specialist_id = request.args.get("specialist_id", type=int)
    building_id = request.args.get("building_id", type=int)
    role_title = (request.args.get("role_title") or "").strip().lower()
    class_id = request.args.get("class_id", type=int)
    grade = request.args.get("grade", type=int)
    category = (request.args.get("category") or "").strip().lower()
    status = (request.args.get("status") or "").strip().upper()
    active_only = request.args.get("active_only") == "1"

    def _category_matches(child, val):
        mapping = {
            "ovz": bool(getattr(child, "is_ovz", False)),
            "vshu": bool(getattr(child, "is_vshu", False)),
            "low": bool(getattr(child, "is_low", False)),
            "az": bool(getattr(child, "is_az", False)),
            "disabled": bool(getattr(child, "is_disabled", False)),
        }
        return mapping.get(val, True)

    filtered = []
    for row in rows:
        child = row.child
        specialist = row.specialist
        school_class = child.current_class if child else None
        if q:
            hay = " ".join([
                child.fio.lower() if child else "",
                specialist.fio.lower() if specialist else "",
                (row.role_title or "").lower(),
            ])
            if q not in hay:
                continue
        if specialist_id and row.specialist_id != specialist_id:
            continue
        if building_id and row.building_id != building_id and not (child and child.current_building and child.current_building.id == building_id):
            continue
        if role_title and role_title not in (row.role_title or "").lower():
            continue
        if class_id and not (school_class and school_class.id == class_id):
            continue
        if grade and not (school_class and school_class.grade == grade):
            continue
        if category and child and not _category_matches(child, category):
            continue
        if status and (row.status or "").upper() != status:
            continue
        if active_only and (row.status or "").upper() != "ACTIVE":
            continue
        filtered.append(row)

    page, per_page = resolve_pagination()
    items, pagination = paginate_list(filtered, page=page, per_page=per_page)
    stats = {
        "children_count": len({x.child_id for x in filtered if x.child_id}),
        "active_count": len([x for x in filtered if (x.status or "").upper() == "ACTIVE"]),
        "finished_count": len([x for x in filtered if (x.status or "").upper() == "FINISHED"]),
        "specialists_count": len({x.specialist_id for x in filtered if x.specialist_id}),
    }
    return items, pagination, stats


def _children_for_specialist_rows():
    rows = _visible_assignments_query().all()
    grouped = defaultdict(list)
    for row in rows:
        grouped[row.specialist].append(row)
    data = []
    for specialist, items in grouped.items():
        items = sorted(items, key=lambda x: (x.child.fio if x.child else "", x.start_date or date.min), reverse=False)
        data.append((specialist, items))
    return sorted(data, key=lambda x: x[0].fio if x[0] else "")


def _building_summary_rows():
    rows = _visible_assignments_query().all()
    grouped = defaultdict(list)
    for row in rows:
        building = row.building or (row.child.current_building if row.child else None)
        grouped[building].append(row)
    data = []
    for building, items in grouped.items():
        data.append((building, sorted(items, key=lambda x: x.child.fio if x.child else "")))
    return sorted(data, key=lambda x: (x[0].name if x[0] else "Без здания"))


@service_staff_bp.route("/")
@login_required
def index():
    _require_view()
    stats = _specialist_dashboard_stats()
    top_specs = [{"name": name, "count": count} for name, count in stats["spec_counts"]]
    return render_template(
        "service_staff/index.html",
        stats=stats,
        top_specs=top_specs,
        can_edit=_can_edit_module(),
        can_view_registry=_can_view_module(),
    )


@service_staff_bp.route("/specialists")
@login_required
def registry():
    _require_view()
    items, pagination = _registry_rows()
    return render_template(
        "service_staff/registry.html",
        rows=items,
        pagination=pagination,
        buildings=_building_choices(),
        specializations=_specialization_choices(),
        can_edit=_can_edit_module(),
    )


@service_staff_bp.route("/specialists/new", methods=["GET", "POST"])
@login_required
def specialist_new():
    _require_edit()
    specialist = ServiceSpecialist(is_active=True)
    if request.method == "POST":
        try:
            _apply_specialist_form(specialist)
            db.session.add(specialist)
            db.session.commit()
            flash("Карточка специалиста создана.", "success")
            return redirect(url_for("service_staff.specialist_card", specialist_id=specialist.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "service_staff/form.html",
        specialist=specialist,
        users=_user_choices(),
        buildings=_building_choices(),
        specializations=_specialization_choices(),
        selected_buildings=[],
        selected_specs=[],
        form_title="Новый специалист службы",
    )


@service_staff_bp.route("/specialists/<int:specialist_id>")
@login_required
def specialist_card(specialist_id: int):
    _require_view()
    specialist = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.user),
        joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceSpecialist.building_links).joinedload(ServiceSpecialistBuilding.building),
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceSpecialist.responsible_links),
    ).get_or_404(specialist_id)
    if not _can_view_specialist(specialist):
        abort(403)
    specialist_assignments = _visible_assignments_query().filter(ServiceAssignment.specialist_id == specialist.id).order_by(ServiceAssignment.updated_at.desc()).limit(100).all()
    return render_template("service_staff/card.html", specialist=specialist, can_edit=_can_edit_module(), specialist_assignments=specialist_assignments)


@service_staff_bp.route("/specialists/<int:specialist_id>/edit", methods=["GET", "POST"])
@login_required
def specialist_edit(specialist_id: int):
    _require_edit()
    specialist = ServiceSpecialist.query.get_or_404(specialist_id)
    if request.method == "POST":
        try:
            _apply_specialist_form(specialist)
            db.session.commit()
            flash("Карточка специалиста обновлена.", "success")
            return redirect(url_for("service_staff.specialist_card", specialist_id=specialist.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "service_staff/form.html",
        specialist=specialist,
        users=_user_choices(),
        buildings=_building_choices(),
        specializations=_specialization_choices(),
        selected_buildings=[link.building_id for link in specialist.building_links],
        selected_specs=[link.specialization_id for link in specialist.specialization_links],
        form_title="Редактирование карточки специалиста",
    )




@service_staff_bp.route("/specialists/<int:specialist_id>/toggle", methods=["POST"])
@login_required
def specialist_toggle(specialist_id: int):
    _require_edit()
    specialist = ServiceSpecialist.query.get_or_404(specialist_id)
    specialist.is_active = not bool(specialist.is_active)
    if not specialist.is_active:
        for row in specialist.responsible_links:
            row.is_active = False
    db.session.commit()
    flash("Состав службы обновлен.", "success")
    return redirect(request.referrer or url_for("service_staff.specialist_card", specialist_id=specialist.id))

@service_staff_bp.route("/structure")
@login_required
def structure():
    _require_view()
    specialists = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceSpecialist.building_links).joinedload(ServiceSpecialistBuilding.building),
    ).order_by(ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc()).all()
    grouped = defaultdict(list)
    for specialist in specialists:
        if not specialist.is_active:
            continue
        if specialist.specialization_links:
            for link in specialist.specialization_links:
                name = link.specialization.name if link.specialization else "Без специализации"
                grouped[name].append(specialist)
        else:
            grouped["Без специализации"].append(specialist)
    data = sorted(grouped.items(), key=lambda x: x[0])
    return render_template("service_staff/structure.html", grouped=data, can_edit=_can_edit_module())


@service_staff_bp.route("/buildings")
@login_required
def buildings_summary():
    _require_view()
    buildings = Building.query.order_by(Building.name.asc()).all()
    # Batch load all building-specialist links with eager-loaded specialist + specializations
    all_links = (
        ServiceSpecialistBuilding.query
        .options(
            joinedload(ServiceSpecialistBuilding.specialist)
            .subqueryload(ServiceSpecialist.specialization_links)
            .joinedload(ServiceSpecialistSpecialization.specialization)
        )
        .all()
    )
    links_by_building = defaultdict(list)
    for lk in all_links:
        links_by_building[lk.building_id].append(lk)

    rows = []
    for building in buildings:
        links = links_by_building.get(building.id, [])
        specialists = [lk.specialist for lk in links if lk.specialist and lk.specialist.is_active]
        spec_counts = defaultdict(int)
        for specialist in specialists:
            for spec_link in specialist.specialization_links:
                if spec_link.specialization:
                    spec_counts[spec_link.specialization.name] += 1
        rows.append({
            "building": building,
            "specialists": sorted({sp.id: sp for sp in specialists}.values(), key=lambda x: x.fio),
            "spec_counts": sorted(spec_counts.items(), key=lambda x: x[0]),
        })
    return render_template("service_staff/buildings.html", rows=rows, can_edit=_can_edit_module())


@service_staff_bp.route("/responsibles", methods=["GET", "POST"])
@login_required
def responsibles():
    _require_view()
    if request.method == "POST":
        _require_edit()
        action = request.form.get("action")
        specialist_id = request.form.get("specialist_id", type=int)
        user_id = request.form.get("user_id", type=int)
        specialist = None
        if specialist_id:
            specialist = ServiceSpecialist.query.get_or_404(specialist_id)
        elif user_id:
            specialist = _ensure_specialist_for_user(User.query.get_or_404(user_id))
            db.session.flush()
        else:
            flash("Выберите пользователя или специалиста.", "danger")
            return redirect(url_for("service_staff.responsibles"))
        row = ServiceResponsible.query.filter_by(specialist_id=specialist.id).first()
        if not row:
            row = ServiceResponsible(specialist_id=specialist.id, assigned_by_user_id=current_user.id, is_active=(action == "assign"))
            db.session.add(row)
        else:
            row.is_active = action == "assign"
            row.assigned_by_user_id = current_user.id
        db.session.commit()
        flash("Статус ответственного обновлен.", "success")
        return redirect(url_for("service_staff.responsibles"))

    responsibles_rows = ServiceResponsible.query.options(
        joinedload(ServiceResponsible.specialist).joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceResponsible.specialist).joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceResponsible.assigned_by),
    ).order_by(ServiceResponsible.updated_at.desc()).all()
    candidates = User.query.filter(User.is_active_user.is_(True)).order_by(User.last_name.asc(), User.first_name.asc(), User.middle_name.asc()).all()
    return render_template(
        "service_staff/responsibles.html",
        responsibles=responsibles_rows,
        candidates=candidates,
        can_edit=_can_edit_module(),
    )


@service_staff_bp.route("/assignments")
@login_required
def assignments_registry():
    _require_view()
    items, pagination, stats = _assignment_display_group_data()
    return render_template(
        "service_staff/assignments_registry.html",
        rows=items,
        pagination=pagination,
        stats=stats,
        specialists=_specialist_choices(),
        buildings=_building_choices(),
        classes=_class_choices(),
        can_edit=_can_edit_module(),
        status_choices=ASSIGNMENT_STATUS_CHOICES,
    )


@service_staff_bp.route("/assignments/new", methods=["GET", "POST"])
@login_required
def assignment_new():
    _require_edit()
    assignment = ServiceAssignment(status="ACTIVE", start_date=date.today())
    form_state = {
        'child_id': '',
        'building_id': '',
        'status': assignment.status,
        'start_date': assignment.start_date.isoformat() if assignment.start_date else '',
        'end_date': '',
        'basis': '',
        'comment': '',
        'enabled_roles': [],
        'role_specialists': {key: '' for key in ASSIGNMENT_ROLE_LABELS},
    }
    if request.method == "POST":
        form_state = _form_state_from_request()
        try:
            new_rows = _multi_assignment_payload_from_request()
            for row in new_rows:
                db.session.add(row)
                db.session.flush()
                _log_assignment_history(row, None, row.status, "Создание назначения")
            db.session.commit()
            flash(f"Карточка сопровождения сохранена. Назначено специалистов: {len(new_rows)}.", "success")
            return redirect(url_for("service_staff.assignments_registry"))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    incidents_list = []
    if is_admin():
        incidents_list = (
            Incident.query
            .order_by(Incident.occurred_at.desc())
            .limit(200)
            .all()
        )
    return render_template(
        "service_staff/assignment_form.html",
        assignment=assignment,
        form_title="Новое назначение ребенка специалисту",
        children=_child_choices(),
        specialists=_specialist_choices(),
        role_specialists=_assignment_role_specialists(),
        assignment_role_choices=ASSIGNMENT_ROLE_CHOICES,
        buildings=_building_choices(),
        classes=_class_choices(),
        status_choices=ASSIGNMENT_STATUS_CHOICES,
        form_state=form_state,
        multi_mode=True,
        incidents_list=incidents_list,
        is_admin_user=is_admin(),
    )


@service_staff_bp.route("/assignments/<int:assignment_id>/edit", methods=["GET", "POST"])
@login_required
def assignment_edit(assignment_id: int):
    _require_view()
    anchor_assignment = _visible_assignments_query().filter(ServiceAssignment.id == assignment_id).first_or_404()
    assignments = _visible_assignments_query().filter(ServiceAssignment.child_id == anchor_assignment.child_id).order_by(ServiceAssignment.id.asc()).all()
    assignment = assignments[0] if assignments else anchor_assignment
    form_state = _multi_form_state_from_assignments(assignments)
    if request.method == "POST":
        _require_edit()
        form_state = _form_state_from_request()
        try:
            saved_count = _save_multi_assignment_group(assignments)
            db.session.commit()
            refreshed = _visible_assignments_query().filter(ServiceAssignment.child_id == request.form.get('child_id', type=int)).order_by(ServiceAssignment.id.asc()).first()
            flash(f"Карточка сопровождения обновлена. Специалистов: {saved_count}.", "success")
            return redirect(url_for("service_staff.assignment_edit", assignment_id=(refreshed.id if refreshed else assignment.id)))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    assignment_ids = [row.id for row in assignments]
    history_rows = ServiceAssignmentHistory.query.options(joinedload(ServiceAssignmentHistory.changed_by)).filter(ServiceAssignmentHistory.assignment_id.in_(assignment_ids)).order_by(ServiceAssignmentHistory.created_at.desc()).all() if assignment_ids else []
    incidents_list = []
    if is_admin():
        incidents_list = (
            Incident.query
            .order_by(Incident.occurred_at.desc())
            .limit(200)
            .all()
        )
    return render_template(
        "service_staff/assignment_form.html",
        assignment=assignment,
        form_title="Редактирование назначения",
        children=_child_choices(),
        specialists=_specialist_choices(),
        role_specialists=_assignment_role_specialists(),
        assignment_role_choices=ASSIGNMENT_ROLE_CHOICES,
        buildings=_building_choices(),
        classes=_class_choices(),
        status_choices=ASSIGNMENT_STATUS_CHOICES,
        history_rows=history_rows,
        form_state=form_state,
        multi_mode=True,
        incidents_list=incidents_list,
        is_admin_user=is_admin(),
    )


@service_staff_bp.route("/assignments/<int:assignment_id>/status", methods=["POST"])
@login_required
def assignment_change_status(assignment_id: int):
    _require_edit()
    assignment = ServiceAssignment.query.get_or_404(assignment_id)
    new_status = (request.form.get("status") or "").upper()
    if new_status not in {x[0] for x in ASSIGNMENT_STATUS_CHOICES}:
        abort(400)
    old_status = assignment.status
    assignment.status = new_status
    if new_status == "FINISHED" and not assignment.end_date:
        assignment.end_date = date.today()
    if new_status == "ACTIVE":
        assignment.end_date = None
    _log_assignment_history(assignment, old_status, new_status, request.form.get("comment") or None)
    db.session.commit()
    flash("Статус сопровождения изменен.", "success")
    return redirect(request.referrer or url_for("service_staff.assignments_registry"))


@service_staff_bp.route("/assignments/<int:assignment_id>/delete", methods=["POST"])
@login_required
def assignment_delete(assignment_id: int):
    _require_edit()
    anchor_assignment = _visible_assignments_query().filter(ServiceAssignment.id == assignment_id).first_or_404()
    rows = _visible_assignments_query().filter(ServiceAssignment.child_id == anchor_assignment.child_id).all()
    deleted_count = len(rows)
    child_name = anchor_assignment.child.fio if anchor_assignment.child else 'карточки'
    for row in rows:
        db.session.delete(row)
    db.session.commit()
    flash(f"Карточка сопровождения удалена: {child_name}. Удалено назначений: {deleted_count}.", "success")
    return redirect(url_for("service_staff.assignments_registry"))


@service_staff_bp.route("/children-summary")
@login_required
def children_summary():
    _require_view()
    grouped = _children_for_specialist_rows()
    return render_template("service_staff/children_summary.html", grouped=grouped)


@service_staff_bp.route("/buildings-children")
@login_required
def buildings_children_summary():
    _require_view()
    grouped = _building_summary_rows()
    return render_template("service_staff/buildings_children_summary.html", grouped=grouped)


@service_staff_bp.route("/assignments/export")
@login_required
def assignments_export():
    _require_view()
    rows, _, _ = _assignment_registry_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "Сопровождение"
    ws.append(["Ребенок", "Класс", "Специалист", "Роль", "Здание", "Дата начала", "Дата окончания", "Статус", "Основание", "Комментарий"])
    for row in rows:
        ws.append([
            row.child.fio if row.child else "",
            row.child.current_class_name if row.child else "",
            row.specialist.fio if row.specialist else "",
            row.role_title or "",
            (row.building.short_name or row.building.name) if row.building else (row.child_building_name or ""),
            row.start_date.strftime("%d.%m.%Y") if row.start_date else "",
            row.end_date.strftime("%d.%m.%Y") if row.end_date else "",
            _status_label(row.status),
            row.basis or "",
            row.comment or "",
        ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="service-assignments.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def child_support_assignments(child, user=None):
    user = user or current_user
    if not can_view_child_service_block(child, user=user):
        return []
    q = ServiceAssignment.query.options(
        joinedload(ServiceAssignment.specialist).joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceAssignment.building),
    ).filter(ServiceAssignment.child_id == child.id).order_by(ServiceAssignment.start_date.desc().nullslast(), ServiceAssignment.updated_at.desc())
    if _can_edit_module(user) or _is_service_methodist(user):
        return q.all()
    linked = _linked_specialist(user)
    if linked:
        return q.filter((ServiceAssignment.specialist_id == linked.id) | (ServiceAssignment.status == "ACTIVE")).all()
    return q.all()


try:
    from weasyprint import HTML
except Exception:
    HTML = None
from flask import current_app

CYCLEGRAM_STATUS_CHOICES = [
    ("DRAFT", "Черновик"),
    ("REVIEW", "На проверке"),
    ("APPROVED", "Утверждена"),
    ("ARCHIVED", "Архив"),
]
WEEKDAYS = [
    (1, "Понедельник"),
    (2, "Вторник"),
    (3, "Среда"),
    (4, "Четверг"),
    (5, "Пятница"),
    (6, "Суббота"),
]
ACTIVITY_TYPE_SEED = [
    ("individual_diagnostics", "Индивидуальная диагностика обучающегося", "PRACTICAL", "психолог, логопед, дефектолог", "Диагностика с [ФИО], [класс]", True, False, False, 10),
    ("group_diagnostics", "Групповая диагностика", "PRACTICAL", "психолог, логопед, дефектолог", "Диагностика в [класс/группа]", False, True, True, 20),
    ("student_consultation", "Индивидуальное консультирование обучающегося", "PRACTICAL", "психолог, соц. педагог", "Консультирование [ФИО]", True, False, False, 30),
    ("parent_consultation", "Консультирование родителя", "PRACTICAL", "все специалисты", "Консультация родителя [ФИО ребенка]", True, False, False, 40),
    ("teacher_consultation", "Консультирование педагога / классного руководителя", "PRACTICAL", "все специалисты", "Консультация педагога по [ФИО ребенка/класс]", False, False, False, 50),
    ("correction_individual", "Коррекционно-развивающее индивидуальное занятие", "PRACTICAL", "психолог, дефектолог", "Коррекционно-развивающее занятие с [ФИО]", True, False, False, 60),
    ("correction_group", "Коррекционно-развивающее групповое занятие", "PRACTICAL", "психолог, дефектолог", "Коррекционно-развивающее занятие, группа [название]", False, True, True, 70),
    ("speech_individual", "Логопедическое индивидуальное занятие", "PRACTICAL", "логопед", "Логопедическое занятие с [ФИО]", True, False, False, 80),
    ("speech_group", "Логопедическое подгрупповое занятие", "PRACTICAL", "логопед", "Логопедическое занятие, группа [название]", False, True, True, 90),
    ("defect_individual", "Дефектологическое индивидуальное занятие", "PRACTICAL", "дефектолог", "Дефектологическое занятие с [ФИО]", True, False, False, 100),
    ("defect_group", "Дефектологическое подгрупповое занятие", "PRACTICAL", "дефектолог", "Дефектологическое занятие, группа [название]", False, True, True, 110),
    ("lesson_observation", "Наблюдение на уроке / занятии", "PRACTICAL", "все специалисты", "Наблюдение на уроке у [класс/предмет/ФИО]", False, False, False, 120),
    ("ppk", "Участие в консилиуме / ППк", "PRACTICAL", "все специалисты", "Участие в консилиуме по [ФИО/класс/случай]", False, False, False, 130),
    ("preventive", "Профилактическое занятие / мероприятие", "PRACTICAL", "психолог, соц. педагог", "Профилактическая работа в [класс/группа]", False, True, True, 140),
    ("psychoeducation", "Психологическое просвещение", "PRACTICAL", "психолог", "Просветительское занятие для [аудитория]", False, True, True, 150),
    ("family_support", "Социально-педагогическое сопровождение семьи / ребенка", "PRACTICAL", "соц. педагог", "Сопровождение семьи [ФИО ребенка]", True, False, False, 160),
    ("interagency", "Межведомственное взаимодействие", "PRACTICAL", "соц. педагог, психолог", "Взаимодействие по [ФИО ребенка/семье]", False, False, False, 170),
    ("docs", "Заполнение документации и журналов", "METHODICAL", "все специалисты", "Оформление рабочей документации", False, False, False, 210),
    ("report_presentation", "Подготовка заключения / представления / характеристики", "METHODICAL", "все специалисты", "Подготовка представления на [ФИО]", False, False, False, 220),
    ("materials", "Подготовка материалов к занятию", "METHODICAL", "все специалисты", "Подготовка материалов к занятиям", False, False, False, 230),
    ("diagnostics_analysis", "Анализ результатов диагностики", "METHODICAL", "все специалисты", "Анализ диагностики [ФИО/класс]", False, False, False, 240),
    ("methodical_meeting", "Методическое совещание / кафедра / объединение", "METHODICAL", "все специалисты", "Методическое совещание службы", False, False, False, 250),
    ("self_education", "Самообразование / изучение методических материалов", "METHODICAL", "все специалисты", "Самообразование по теме [тема]", False, False, False, 260),
    ("planning", "Планирование работы специалиста", "METHODICAL", "все специалисты", "Планирование работы на период", False, False, False, 270),
    ("reports", "Подготовка отчетов и справок", "METHODICAL", "все специалисты", "Подготовка отчета по зданию/направлению", False, False, False, 280),
]


def _seed_activity_types():
    changed = False
    for code, name, category, scope, template_text, requires_child, requires_group, is_group, sort_order in ACTIVITY_TYPE_SEED:
        row = ServiceActivityType.query.filter_by(code=code).first()
        if not row:
            row = ServiceActivityType(code=code)
            db.session.add(row)
            changed = True
        for field, value in {
            'name': name,
            'work_category': category,
            'specialist_scope': scope,
            'template_text': template_text,
            'requires_child': requires_child,
            'requires_group': requires_group,
            'is_group_activity': is_group,
            'sort_order': sort_order,
            'is_active': True,
        }.items():
            if getattr(row, field) != value:
                setattr(row, field, value)
                changed = True
    if changed:
        db.session.commit()


def _time_to_minutes(value):
    value = (value or '').strip()
    if not value:
        return None
    try:
        hh, mm = value.split(':')
        hh = int(hh)
        mm = int(mm)
        if hh < 0 or hh > 23 or mm < 0 or mm > 59:
            return None
        return hh * 60 + mm
    except Exception:
        return None


def _cyclegram_status_label(value):
    return dict(CYCLEGRAM_STATUS_CHOICES).get((value or '').upper(), value or '—')


def _can_manage_cyclegram(cyclegram, user=None):
    user = user or current_user
    if _can_edit_module(user) or _is_service_methodist(user):
        return True
    linked = _linked_specialist(user)
    return bool(linked and linked.id == cyclegram.specialist_id and (cyclegram.status or '').upper() in {'DRAFT', 'REVIEW'})


def _visible_cyclegrams_query(user=None):
    user = user or current_user
    q = ServiceCyclegram.query.options(
        joinedload(ServiceCyclegram.specialist).joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceCyclegram.history_entries).joinedload(ServiceCyclegramHistory.changed_by),
    )
    if _can_edit_module(user) or _is_service_methodist(user):
        return q
    linked = _linked_specialist(user)
    if linked:
        return q.filter(ServiceCyclegram.specialist_id == linked.id)
    return q.filter(ServiceCyclegram.id == -1)


def _activity_type_choices():
    _seed_activity_types()
    return ServiceActivityType.query.filter_by(is_active=True).order_by(ServiceActivityType.sort_order.asc(), ServiceActivityType.name.asc()).all()


def _activity_type_map():
    return {row.id: row for row in _activity_type_choices()}


def _log_cyclegram_history(cyclegram, action, old_status=None, new_status=None, comment=None):
    db.session.add(ServiceCyclegramHistory(
        cyclegram_id=cyclegram.id,
        changed_by_user_id=getattr(current_user, 'id', None),
        action=action,
        old_status=old_status,
        new_status=new_status,
        comment=comment,
    ))


def _specialist_display_buildings(specialist):
    if not specialist:
        return ''
    if specialist.buildings_text:
        return specialist.buildings_text
    if specialist.main_building:
        return specialist.main_building.short_name or specialist.main_building.name
    return ''


def _academic_year_choices():
    return AcademicYear.query.order_by(AcademicYear.is_current.desc(), AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()


def _default_academic_year_name():
    current = AcademicYear.query.filter_by(is_current=True).order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.id.desc()).first()
    if current and current.name:
        return current.name
    latest = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).first()
    if latest and latest.name:
        return latest.name
    return f'{date.today().year}/{date.today().year + 1}'


def _cyclegram_position_choices():
    values = []
    seen = set()
    for label in SERVICE_ROLE_LABELS.values():
        if label not in seen:
            seen.add(label)
            values.append(label)
    for specialist in ServiceSpecialist.query.order_by(ServiceSpecialist.position_title.asc().nullslast(), ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc()).all():
        value = (specialist.position_title or '').strip()
        if value and value not in seen:
            seen.add(value)
            values.append(value)
    return values


def _cyclegram_selected_building_ids(cyclegram):
    if not cyclegram:
        return []

    values = []
    text = (getattr(cyclegram, 'buildings_text', None) or '').strip().lower()
    if text:
        for building in _building_choices():
            names = [(building.name or '').strip().lower(), (getattr(building, 'short_name', '') or '').strip().lower()]
            if any(name and name in text for name in names):
                values.append(building.id)

    if getattr(cyclegram, 'specialist', None):
        for link in getattr(cyclegram.specialist, 'building_links', []) or []:
            if getattr(link, 'building_id', None):
                values.append(link.building_id)
        if getattr(cyclegram.specialist, 'main_building_id', None):
            values.append(cyclegram.specialist.main_building_id)

    result = []
    seen = set()
    for value in values:
        try:
            value = int(value)
        except (TypeError, ValueError):
            continue
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _apply_cyclegram_form(cyclegram):
    specialist_id = request.form.get('specialist_id')
    academic_year = (request.form.get('academic_year') or '').strip()
    if not specialist_id:
        raise ValueError('Выберите специалиста.')
    if not academic_year:
        raise ValueError('Укажите учебный год.')
    specialist = ServiceSpecialist.query.get(int(specialist_id))
    if not specialist:
        raise ValueError('Специалист не найден.')
    cyclegram.specialist_id = specialist.id
    cyclegram.academic_year = academic_year
    cyclegram.title = (request.form.get('title') or '').strip() or f'Циклограмма {specialist.fio}'
    cyclegram.position_title = (request.form.get('position_title') or '').strip() or specialist.position_title
    rate_raw = (request.form.get('rate_value') or '').strip().replace(',', '.')
    cyclegram.rate_value = float(rate_raw) if rate_raw else specialist.rate_value
    building_ids = [int(x) for x in request.form.getlist('building_ids') if str(x).isdigit()]
    if building_ids:
        selected_buildings = Building.query.filter(Building.id.in_(building_ids)).order_by(Building.name.asc()).all()
        cyclegram.buildings_text = ', '.join([x.name for x in selected_buildings])
    else:
        cyclegram.buildings_text = (request.form.get('buildings_text') or '').strip() or _specialist_display_buildings(specialist)
    cyclegram.updated_by_user_id = getattr(current_user, 'id', None)
    if not cyclegram.created_by_user_id:
        cyclegram.created_by_user_id = getattr(current_user, 'id', None)


def _activity_type_default_text(activity_type, child=None, group_text=None):
    text = activity_type.template_text or activity_type.name
    if child:
        text = text.replace('[ФИО]', child.fio).replace('[ФИО ребенка]', child.fio).replace('[класс]', child.current_class_name or '')
    if group_text:
        text = text.replace('[класс/группа]', group_text).replace('[название]', group_text).replace('[аудитория]', group_text)
    return text


def _apply_cyclegram_entry_form(entry, cyclegram):
    weekday = int(request.form.get('weekday') or 0)
    start_time = (request.form.get('start_time') or '').strip()
    end_time = (request.form.get('end_time') or '').strip()
    activity_type_id = int(request.form.get('activity_type_id') or 0)
    activity_type = ServiceActivityType.query.get(activity_type_id)
    if weekday not in dict(WEEKDAYS):
        raise ValueError('Выберите день недели.')
    if not activity_type:
        raise ValueError('Выберите вид деятельности.')
    start_min = _time_to_minutes(start_time)
    end_min = _time_to_minutes(end_time)
    if start_min is None or end_min is None:
        raise ValueError('Укажите корректный временной интервал в формате ЧЧ:ММ.')
    if end_min <= start_min:
        raise ValueError('Окончание интервала должно быть позже начала.')
    child = None
    child_id = request.form.get('child_id')
    if child_id:
        child = Child.query.get(int(child_id))
    parallel_value = (request.form.get('parallel_value') or '').strip()
    class_name = (request.form.get('class_name') or '').strip()
    group_name = (request.form.get('group_name') or '').strip()
    group_text = (request.form.get('group_text') or '').strip()
    if not group_text:
        if group_name:
            group_text = group_name
        elif child:
            group_text = ''
        elif class_name:
            group_text = class_name
        elif parallel_value:
            group_text = f'{parallel_value} классы'
    group_text = group_text or None
    if activity_type.requires_child and not child and not group_text:
        raise ValueError('Для выбранного вида деятельности требуется указать параллель, класс, группу или ребенка.')
    if activity_type.requires_group and not group_text and not child:
        raise ValueError('Для выбранного вида деятельности требуется указать группу, параллель, класс или ребенка.')
    if child and class_name and (child.current_class_name or '').strip() != class_name:
        raise ValueError('Выбранный ребенок не относится к указанному классу.')
    if class_name and parallel_value:
        class_parallel = _extract_parallel_from_class_name(class_name)
        if class_parallel is not None and str(class_parallel) != str(parallel_value):
            raise ValueError('Выбранный класс не относится к указанной параллели.')
    minutes = end_min - start_min
    if request.form.get('adjust_minutes') == '1' and (_can_edit_module() or _is_service_methodist()):
        minutes_raw = (request.form.get('minutes') or '').strip()
        if minutes_raw:
            minutes = int(minutes_raw)
            entry.minutes_adjusted = True
            entry.adjustment_reason = (request.form.get('adjustment_reason') or '').strip() or None
    else:
        entry.minutes_adjusted = False
        entry.adjustment_reason = None

    building_id = request.form.get('building_id')
    sort_order = request.form.get('sort_order')

    # overlap check
    other_rows = ServiceCyclegramEntry.query.filter_by(cyclegram_id=cyclegram.id, weekday=weekday).all()
    for row in other_rows:
        if entry.id and row.id == entry.id:
            continue
        row_start = _time_to_minutes(row.start_time)
        row_end = _time_to_minutes(row.end_time)
        if row_start is None or row_end is None:
            continue
        if max(start_min, row_start) < min(end_min, row_end):
            raise ValueError('Внутри одного дня интервалы не должны пересекаться.')

    entry.cyclegram_id = cyclegram.id
    entry.weekday = weekday
    entry.start_time = start_time
    entry.end_time = end_time
    entry.activity_type_id = activity_type.id
    entry.work_category = activity_type.work_category
    entry.description = (request.form.get('description') or '').strip() or _activity_type_default_text(activity_type, child=child, group_text=group_text)
    entry.child_id = child.id if child else None
    entry.group_text = group_text
    entry.minutes = minutes
    entry.building_id = int(building_id) if building_id else None
    entry.comment = (request.form.get('comment') or '').strip() or None
    entry.sort_order = int(sort_order) if sort_order and str(sort_order).isdigit() else (start_min or 0)


def _cyclegram_registry_data():
    q = _visible_cyclegrams_query()
    academic_year = (request.args.get('academic_year') or '').strip()
    specialist_id = request.args.get('specialist_id')
    status = (request.args.get('status') or '').strip().upper()
    building_id = request.args.get('building_id')
    q_text = (request.args.get('q') or '').strip().lower()
    rows = q.order_by(ServiceCyclegram.updated_at.desc(), ServiceCyclegram.created_at.desc()).all()
    filtered = []
    for row in rows:
        if academic_year and row.academic_year != academic_year:
            continue
        if specialist_id and str(row.specialist_id) != str(specialist_id):
            continue
        if status and (row.status or '').upper() != status:
            continue
        if building_id:
            bid = str(building_id)
            specialist_buildings = [str(x.building_id) for x in row.specialist.building_links]
            if bid not in specialist_buildings and str(getattr(row.specialist, 'main_building_id', '') or '') != bid:
                continue
        if q_text:
            hay = ' '.join([
                row.academic_year or '', row.title or '', row.specialist.fio if row.specialist else '', row.position_title or '', row.buildings_text or ''
            ]).lower()
            if q_text not in hay:
                continue
        filtered.append(row)
    page, per_page = resolve_pagination()
    items, pagination = paginate_list(filtered, page=page, per_page=per_page)
    stats = {
        'total': len(filtered),
        'draft_count': len([x for x in filtered if (x.status or '').upper() == 'DRAFT']),
        'review_count': len([x for x in filtered if (x.status or '').upper() == 'REVIEW']),
        'approved_count': len([x for x in filtered if (x.status or '').upper() == 'APPROVED']),
    }
    return items, pagination, stats


def _cyclegram_day_blocks(cyclegram):
    rows = sorted(cyclegram.entries, key=lambda x: (x.weekday, x.sort_order, x.start_time))
    grouped = defaultdict(list)
    for row in rows:
        grouped[row.weekday].append(row)
    result = []
    practical_total = 0
    methodical_total = 0
    for weekday_code, weekday_name in WEEKDAYS:
        day_rows = grouped.get(weekday_code, [])
        p = sum(x.minutes for x in day_rows if (x.work_category or '').upper() == 'PRACTICAL')
        m = sum(x.minutes for x in day_rows if (x.work_category or '').upper() == 'METHODICAL')
        practical_total += p
        methodical_total += m
        result.append({'code': weekday_code, 'name': weekday_name, 'rows': day_rows, 'practical_minutes': p, 'methodical_minutes': m, 'total_minutes': p + m})
    week_total = practical_total + methodical_total
    practical_pct = round(practical_total * 100 / week_total, 1) if week_total else 0
    methodical_pct = round(methodical_total * 100 / week_total, 1) if week_total else 0
    return {
        'days': result,
        'practical_total': practical_total,
        'methodical_total': methodical_total,
        'week_total': week_total,
        'practical_pct': practical_pct,
        'methodical_pct': methodical_pct,
        'norm_ok': practical_pct >= 50 and methodical_pct <= 50 if week_total else False,
    }


def _cyclegram_can_view(cyclegram):
    return _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram.id).first() is not None


def _render_cyclegram_pdf(cyclegram, summary):
    html = render_template('service_staff/cyclegram_print.html', cyclegram=cyclegram, summary=summary, print_mode=True)
    pdf = HTML(string=html, base_url=current_app.root_path).write_pdf()
    bio = BytesIO(pdf)
    bio.seek(0)
    return bio


@service_staff_bp.route('/cyclegrams')
@login_required
def cyclegrams_registry():
    _require_view()
    _seed_activity_types()
    rows, pagination, stats = _cyclegram_registry_data()
    academic_years = sorted({x.academic_year for x in ServiceCyclegram.query.all() if x.academic_year})
    if not academic_years:
        academic_years = [f'{date.today().year}/{date.today().year + 1}']
    return render_template(
        'service_staff/cyclegrams_registry.html',
        rows=rows,
        pagination=pagination,
        stats=stats,
        specialists=_specialist_choices(),
        buildings=_building_choices(),
        academic_years=academic_years,
        status_choices=CYCLEGRAM_STATUS_CHOICES,
        can_edit=_can_edit_module() or _linked_specialist() is not None,
    )


@service_staff_bp.route('/cyclegrams/new', methods=['GET', 'POST'])
@login_required
def cyclegram_new():
    _require_view()
    if not (_can_edit_module() or _linked_specialist() is not None):
        abort(403)
    cyclegram = ServiceCyclegram(status='DRAFT', academic_year=_default_academic_year_name())
    linked = _linked_specialist()
    if linked and not (_can_edit_module() or _is_service_methodist()):
        cyclegram.specialist_id = linked.id
        cyclegram.position_title = linked.position_title
        cyclegram.rate_value = linked.rate_value
        cyclegram.buildings_text = _specialist_display_buildings(linked)
    copy_from_id = request.args.get('copy_from')
    if copy_from_id and str(copy_from_id).isdigit():
        source = _visible_cyclegrams_query().filter(ServiceCyclegram.id == int(copy_from_id)).first()
        if source:
            cyclegram.specialist_id = source.specialist_id
            cyclegram.academic_year = source.academic_year
            cyclegram.title = source.title
            cyclegram.position_title = source.position_title
            cyclegram.rate_value = source.rate_value
            cyclegram.buildings_text = source.buildings_text
            cyclegram.copied_from_cyclegram_id = source.id
    if request.method == 'POST':
        try:
            _apply_cyclegram_form(cyclegram)
            db.session.add(cyclegram)
            db.session.flush()
            if copy_from_id and str(copy_from_id).isdigit():
                source = ServiceCyclegram.query.get(int(copy_from_id))
                if source:
                    for row in source.entries:
                        db.session.add(ServiceCyclegramEntry(
                            cyclegram_id=cyclegram.id,
                            weekday=row.weekday,
                            start_time=row.start_time,
                            end_time=row.end_time,
                            activity_type_id=row.activity_type_id,
                            description=row.description,
                            child_id=row.child_id,
                            group_text=row.group_text,
                            work_category=row.work_category,
                            minutes=row.minutes,
                            building_id=row.building_id,
                            comment=row.comment,
                            minutes_adjusted=row.minutes_adjusted,
                            adjustment_reason=row.adjustment_reason,
                            sort_order=row.sort_order,
                        ))
            _log_cyclegram_history(cyclegram, 'create', None, cyclegram.status, 'Создание циклограммы')
            db.session.commit()
            flash('Циклограмма создана.', 'success')
            return redirect(url_for('service_staff.cyclegram_card', cyclegram_id=cyclegram.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('service_staff/cyclegram_form.html', cyclegram=cyclegram, specialists=_specialist_choices(), buildings=_building_choices(), academic_years=_academic_year_choices(), position_choices=_cyclegram_position_choices(), selected_building_ids=_cyclegram_selected_building_ids(cyclegram), form_title='Новая циклограмма')


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>')
@login_required
def cyclegram_card(cyclegram_id: int):
    _require_view()
    _seed_activity_types()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    summary = _cyclegram_day_blocks(cyclegram)
    return render_template(
        'service_staff/cyclegram_card.html',
        cyclegram=cyclegram,
        summary=summary,
        activity_types=_activity_type_choices(),
        children=_child_choices(),
        school_classes=_school_class_choices(),
        parallel_choices=_parallel_choices(),
        buildings=_building_choices(),
        can_manage=_can_manage_cyclegram(cyclegram),
        can_review=_can_edit_module() or _is_service_methodist(),
        status_choices=CYCLEGRAM_STATUS_CHOICES,
        extract_parallel_from_class_name=_extract_parallel_from_class_name,
    )


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/edit', methods=['GET', 'POST'])
@login_required
def cyclegram_edit(cyclegram_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    if not _can_manage_cyclegram(cyclegram):
        abort(403)
    if request.method == 'POST':
        try:
            old_status = cyclegram.status
            _apply_cyclegram_form(cyclegram)
            _log_cyclegram_history(cyclegram, 'edit', old_status, cyclegram.status, 'Обновление карточки циклограммы')
            db.session.commit()
            flash('Карточка циклограммы обновлена.', 'success')
            return redirect(url_for('service_staff.cyclegram_card', cyclegram_id=cyclegram.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('service_staff/cyclegram_form.html', cyclegram=cyclegram, specialists=_specialist_choices(), buildings=_building_choices(), academic_years=_academic_year_choices(), position_choices=_cyclegram_position_choices(), selected_building_ids=_cyclegram_selected_building_ids(cyclegram), form_title='Редактирование циклограммы')


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/entries/new', methods=['POST'])
@login_required
def cyclegram_entry_new(cyclegram_id: int):
    _require_view()
    _seed_activity_types()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    if not _can_manage_cyclegram(cyclegram):
        abort(403)
    entry = ServiceCyclegramEntry()
    try:
        _apply_cyclegram_entry_form(entry, cyclegram)
        db.session.add(entry)
        _log_cyclegram_history(cyclegram, 'entry_add', cyclegram.status, cyclegram.status, 'Добавлена запись циклограммы')
        db.session.commit()
        flash('Запись добавлена.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(str(exc), 'danger')
    return redirect(url_for('service_staff.cyclegram_card', cyclegram_id=cyclegram.id))


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/entries/<int:entry_id>/delete', methods=['POST'])
@login_required
def cyclegram_entry_delete(cyclegram_id: int, entry_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    if not _can_manage_cyclegram(cyclegram):
        abort(403)
    entry = ServiceCyclegramEntry.query.filter_by(id=entry_id, cyclegram_id=cyclegram.id).first_or_404()
    db.session.delete(entry)
    _log_cyclegram_history(cyclegram, 'entry_delete', cyclegram.status, cyclegram.status, 'Удалена запись циклограммы')
    db.session.commit()
    flash('Запись удалена.', 'success')
    return redirect(url_for('service_staff.cyclegram_card', cyclegram_id=cyclegram.id))


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/status', methods=['POST'])
@login_required
def cyclegram_change_status(cyclegram_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    new_status = (request.form.get('status') or '').upper()
    if new_status not in {x[0] for x in CYCLEGRAM_STATUS_CHOICES}:
        abort(400)
    if new_status in {'APPROVED', 'ARCHIVED'} and not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    if new_status == 'REVIEW' and not (_can_manage_cyclegram(cyclegram) or _can_edit_module() or _is_service_methodist()):
        abort(403)
    old_status = cyclegram.status
    cyclegram.status = new_status
    cyclegram.reviewer_comment = (request.form.get('reviewer_comment') or '').strip() or cyclegram.reviewer_comment
    cyclegram.updated_by_user_id = getattr(current_user, 'id', None)
    _log_cyclegram_history(cyclegram, 'status_change', old_status, new_status, request.form.get('comment') or cyclegram.reviewer_comment)
    db.session.commit()
    flash('Статус циклограммы обновлен.', 'success')
    return redirect(url_for('service_staff.cyclegram_card', cyclegram_id=cyclegram.id))


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/export.xlsx')
@login_required
def cyclegram_export_excel(cyclegram_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    summary = _cyclegram_day_blocks(cyclegram)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Циклограмма'
    ws.append(['Циклограмма рабочего времени'])
    ws.append([cyclegram.specialist.fio if cyclegram.specialist else ''])
    ws.append([f'Учебный год: {cyclegram.academic_year}'])
    ws.append([f'Должность: {cyclegram.position_title or "—"}'])
    ws.append([f'Здания: {cyclegram.buildings_text or "—"}'])
    ws.append([])
    ws.append(['День недели', 'Интервал', 'Вид деятельности', 'Описание', 'Ребенок', 'Группа/класс', 'Категория', 'Минуты', 'Здание', 'Комментарий'])
    for day in summary['days']:
        for row in day['rows']:
            ws.append([
                day['name'], row.interval_text, row.activity_type.name if row.activity_type else '', row.description or '',
                row.child.fio if row.child else '', row.group_text or '',
                'Практическая' if (row.work_category or '').upper() == 'PRACTICAL' else 'Организационно-методическая',
                row.minutes,
                (row.building.short_name or row.building.name) if row.building else '',
                row.comment or '',
            ])
        ws.append([day['name'], '', 'Итого за день', '', '', '', '', day['total_minutes'], '', ''])
    ws.append([])
    ws.append(['Практическая работа, мин.', summary['practical_total']])
    ws.append(['Организационно-методическая работа, мин.', summary['methodical_total']])
    ws.append(['Всего за неделю, мин.', summary['week_total']])
    ws.append(['Практическая, %', summary['practical_pct']])
    ws.append(['Организационно-методическая, %', summary['methodical_pct']])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f'cyclegram-{cyclegram.id}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/export.pdf')
@login_required
def cyclegram_export_pdf(cyclegram_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    summary = _cyclegram_day_blocks(cyclegram)
    bio = _render_cyclegram_pdf(cyclegram, summary)
    return send_file(bio, as_attachment=True, download_name=f'cyclegram-{cyclegram.id}.pdf', mimetype='application/pdf')


@service_staff_bp.route('/cyclegrams/<int:cyclegram_id>/print')
@login_required
def cyclegram_print(cyclegram_id: int):
    _require_view()
    cyclegram = _visible_cyclegrams_query().filter(ServiceCyclegram.id == cyclegram_id).first_or_404()
    summary = _cyclegram_day_blocks(cyclegram)
    return render_template('service_staff/cyclegram_print.html', cyclegram=cyclegram, summary=summary, print_mode=False)


# ===== Этап 4. Электронные представления =====

def _presentation_autofill_text(child):
    birth = child.date_of_birth.strftime('%d.%m.%Y') if child and getattr(child, 'date_of_birth', None) else '—'
    enroll = child.enrollment_date.strftime('%d.%m.%Y') if child and getattr(child, 'enrollment_date', None) else '—'
    building = (child.current_building.short_name or child.current_building.name) if child and getattr(child, 'current_building', None) else '—'
    return "\n".join([
        f"ФИО: {child.fio if child else '—'}",
        f"Дата рождения: {birth}",
        f"Класс: {child.current_class_name if child else '—'}",
        f"Дата зачисления: {enroll}",
        f"Программа: {getattr(child, 'education_program', None) or '—'}",
        f"Форма обучения: {getattr(child, 'education_form', None) or '—'}",
        f"Здание: {building}",
    ])


def _presentation_support_text(child):
    rows = ServiceAssignment.query.options(joinedload(ServiceAssignment.specialist)).filter_by(child_id=child.id).all() if child else []
    if not rows:
        return ''
    parts = []
    for row in rows:
        specialist = row.specialist.fio if row.specialist else '—'
        position = row.specialist.position_title if row.specialist else ''
        status = row.status_label
        period = []
        if row.start_date:
            period.append('с ' + row.start_date.strftime('%d.%m.%Y'))
        if row.end_date:
            period.append('по ' + row.end_date.strftime('%d.%m.%Y'))
        suffix = f"; {' '.join(period)}" if period else ''
        parts.append(f"{specialist} ({position}) — {status}{suffix}")
    return "\n".join(parts)


def _presentation_child_choices():
    return Child.query.order_by(Child.last_name.asc(), Child.first_name.asc(), Child.middle_name.asc()).all()


def _presentation_methodist_choices():
    return User.query.order_by(User.last_name.asc(), User.first_name.asc(), User.middle_name.asc()).all()


def _presentation_executor_can_edit(block, presentation):
    if _can_edit_module() or _is_service_methodist():
        return True
    linked = _linked_specialist()
    if linked and block.executor_specialist_id == linked.id:
        return True
    if block.executor_user_id and block.executor_user_id == getattr(current_user, 'id', None):
        return True
    return False


def _presentation_visible_query(user=None):
    user = user or current_user
    q = ServicePresentation.query.options(
        joinedload(ServicePresentation.child).joinedload(Child.enrollments),
        joinedload(ServicePresentation.building),
        joinedload(ServicePresentation.school_class),
        joinedload(ServicePresentation.methodist),
    )
    if _can_edit_module(user) or _is_service_methodist(user):
        return q
    linked = _linked_specialist(user)
    if linked:
        return q.join(ServicePresentation.blocks).filter(ServicePresentationBlock.executor_specialist_id == linked.id).distinct()
    return q.filter(ServicePresentation.id == -1)


def _log_presentation_history(presentation, action, block=None, comment=None):
    db.session.add(ServicePresentationHistory(
        presentation_id=presentation.id,
        block_id=block.id if block else None,
        user_id=getattr(current_user, 'id', None),
        action=action,
        comment=comment,
    ))


def _presentation_refresh_ready_percent(presentation):
    total = len(presentation.blocks)
    checked = len([b for b in presentation.blocks if (b.status or '').upper() == 'CHECKED'])
    filled = len([b for b in presentation.blocks if (b.status or '').upper() in {'FILLED', 'CHECKED'}])
    presentation.ready_percent = round(((checked if checked else filled) * 100 / total), 0) if total else 0


def _seed_presentation_blocks(presentation):
    child = presentation.child
    support_text = _presentation_support_text(child)
    for code, title, sort_order, fill_mode, is_required in PRESENTATION_DEFAULT_BLOCKS:
        block = ServicePresentationBlock(
            presentation_id=presentation.id,
            block_code=code,
            title=title,
            sort_order=sort_order,
            fill_mode=fill_mode,
            is_required=is_required,
            hint_text=PRESENTATION_HINTS.get(code, ''),
            recommended_text=PRESENTATION_HINTS.get(code, ''),
            content_text='',
            status='NOT_STARTED',
        )
        if code == 'general':
            block.content_text = _presentation_autofill_text(child)
            block.source_name = 'child'
            block.source_updated_at = datetime.utcnow()
            block.status = 'FILLED'
        elif code == 'support' and support_text:
            block.content_text = support_text
            block.source_name = 'service_assignment'
            block.source_updated_at = datetime.utcnow()
            block.status = 'FILLED'
        db.session.add(block)
    db.session.flush()
    _presentation_refresh_ready_percent(presentation)


def _presentation_stats():
    rows = ServicePresentation.query.all()
    return {
        'total': len(rows),
        'approved': len([x for x in rows if (x.status or '').upper() == 'APPROVED']),
        'review': len([x for x in rows if (x.status or '').upper() == 'REVIEW']),
        'draft': len([x for x in rows if (x.status or '').upper() in {'DRAFT', 'IN_PROGRESS'}]),
    }


def _presentation_registry_data():
    q = _presentation_visible_query()
    academic_year = (request.args.get('academic_year') or '').strip()
    status = (request.args.get('status') or '').strip().upper()
    child_id = request.args.get('child_id', type=int)
    class_id = request.args.get('class_id', type=int)
    building_id = request.args.get('building_id', type=int)
    methodist_id = request.args.get('methodist_id', type=int)
    if academic_year:
        q = q.filter(ServicePresentation.academic_year == academic_year)
    if status:
        q = q.filter(ServicePresentation.status == status)
    if child_id:
        q = q.filter(ServicePresentation.child_id == child_id)
    if class_id:
        q = q.filter(ServicePresentation.school_class_id == class_id)
    if building_id:
        q = q.filter(ServicePresentation.building_id == building_id)
    if methodist_id:
        q = q.filter(ServicePresentation.methodist_user_id == methodist_id)
    rows = q.order_by(ServicePresentation.updated_at.desc()).all()
    page, per_page = resolve_pagination(default_per_page=20)
    items, pagination = paginate_list(rows, page=page, per_page=per_page)
    return items, pagination, _presentation_stats()


def _presentation_full_title(presentation):
    child_name = presentation.child.fio if presentation.child else 'Обучающийся'
    return presentation.title or f'Представление на {child_name}'


def _render_presentation_docx(presentation):
    paragraphs = [
        'Представление на обучающегося',
        _presentation_full_title(presentation),
        f'Статус: {presentation.status_label}',
        f'Учебный год: {presentation.academic_year}',
        '',
    ]
    for block in sorted(presentation.blocks, key=lambda x: x.sort_order):
        paragraphs.append(block.title)
        for line in (block.content_text or '').splitlines() or ['']:
            paragraphs.append(line)
        paragraphs.append('')

    def w_p(text):
        safe = escape(text or '')
        return f'<w:p><w:r><w:t xml:space="preserve">{safe}</w:t></w:r></w:p>'

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body>' + ''.join(w_p(p) for p in paragraphs) + '<w:sectPr/>'
        '</w:body></w:document>'
    )
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""
    core = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<dc:title>{escape(_presentation_full_title(presentation))}</dc:title>
<dc:creator>OpenAI</dc:creator>
</cp:coreProperties>"""
    app_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
<Application>Microsoft Office Word</Application>
</Properties>"""
    bio = BytesIO()
    with ZipFile(bio, 'w', ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', rels)
        zf.writestr('docProps/core.xml', core)
        zf.writestr('docProps/app.xml', app_xml)
        zf.writestr('word/document.xml', document_xml)
    bio.seek(0)
    return bio


def _render_presentation_pdf(presentation):
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    width, height = A4
    x = 40
    y = height - 40
    c.setFont('Helvetica-Bold', 14)
    c.drawString(x, y, 'Представление на обучающегося')
    y -= 22
    c.setFont('Helvetica', 10)
    for line in [_presentation_full_title(presentation), f'Статус: {presentation.status_label}', f'Учебный год: {presentation.academic_year}']:
        c.drawString(x, y, line[:120])
        y -= 14
    y -= 6
    for block in sorted(presentation.blocks, key=lambda x: x.sort_order):
        if y < 80:
            c.showPage(); y = height - 40
        c.setFont('Helvetica-Bold', 11)
        c.drawString(x, y, block.title[:110])
        y -= 14
        c.setFont('Helvetica', 9)
        for raw in (block.content_text or '—').splitlines() or ['—']:
            chunks = [raw[i:i+120] for i in range(0, len(raw), 120)] or [' ']
            for chunk in chunks:
                if y < 60:
                    c.showPage(); y = height - 40; c.setFont('Helvetica', 9)
                c.drawString(x + 6, y, chunk)
                y -= 12
        y -= 6
    c.save()
    bio.seek(0)
    return bio


def _apply_presentation_form(presentation):
    child_id = request.form.get('child_id', type=int)
    if not child_id:
        raise ValueError('Выберите обучающегося.')
    child = Child.query.get(child_id)
    if not child:
        raise ValueError('Обучающийся не найден.')
    presentation.child_id = child.id
    presentation.academic_year = (request.form.get('academic_year') or '').strip() or f"{date.today().year}/{date.today().year + 1}"
    presentation.basis = (request.form.get('basis') or '').strip() or 'SUPPORT'
    presentation.methodist_user_id = request.form.get('methodist_user_id', type=int) or None
    presentation.title = (request.form.get('title') or '').strip() or f'Представление на {child.fio}'
    presentation.school_class_id = child.current_class.id if getattr(child, 'current_class', None) else None
    presentation.building_id = child.current_building.id if getattr(child, 'current_building', None) else None
    presentation.last_changed_by_user_id = getattr(current_user, 'id', None)
    if not presentation.id:
        presentation.initiator_user_id = getattr(current_user, 'id', None)


def _assign_default_presentation_executors(presentation):
    specialists = [a.specialist for a in ServiceAssignment.query.options(joinedload(ServiceAssignment.specialist)).filter_by(child_id=presentation.child_id).all() if a.specialist]
    first = specialists[0] if specialists else None
    for block in presentation.blocks:
        if block.block_code in {'support', 'dynamic'} and first:
            block.executor_specialist_id = first.id
            block.executor_user_id = first.user_id
        elif block.block_code == 'conclusion' and presentation.methodist_user_id:
            block.executor_user_id = presentation.methodist_user_id


@service_staff_bp.route('/presentations')
@login_required
def presentations_registry():
    _require_view()
    rows, pagination, stats = _presentation_registry_data()
    academic_years = sorted({x.academic_year for x in ServicePresentation.query.all() if x.academic_year}, reverse=True)
    if not academic_years:
        academic_years = [f'{date.today().year}/{date.today().year + 1}']
    return render_template(
        'service_staff/presentations_registry.html',
        rows=rows,
        pagination=pagination,
        stats=stats,
        academic_years=academic_years,
        basis_choices=PRESENTATION_BASIS_CHOICES,
        status_choices=PRESENTATION_STATUS_CHOICES,
        children=_presentation_child_choices(),
        classes=SchoolClass.query.order_by(SchoolClass.name.asc()).all(),
        buildings=_building_choices(),
        methodists=_presentation_methodist_choices(),
        can_edit=_can_edit_module() or _is_service_methodist(),
    )


@service_staff_bp.route('/presentations/new', methods=['GET', 'POST'])
@login_required
def presentation_new():
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    presentation = ServicePresentation(status='DRAFT', academic_year=f'{date.today().year}/{date.today().year + 1}')
    if request.method == 'POST':
        try:
            _apply_presentation_form(presentation)
            db.session.add(presentation)
            db.session.flush()
            _seed_presentation_blocks(presentation)
            _assign_default_presentation_executors(presentation)
            _presentation_refresh_ready_percent(presentation)
            _log_presentation_history(presentation, 'create', comment='Создание представления')
            db.session.commit()
            flash('Представление создано.', 'success')
            return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('service_staff/presentation_form.html', presentation=presentation, children=_presentation_child_choices(), presentation_classes=_presentation_class_choices(), grade_choices=_presentation_grade_choices(), methodists=_presentation_methodist_choices(), basis_choices=PRESENTATION_BASIS_CHOICES, form_title='Новое электронное представление')


@service_staff_bp.route('/presentations/<int:presentation_id>')
@login_required
def presentation_card(presentation_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    selected_block_id = request.args.get('block_id', type=int)
    blocks = sorted(presentation.blocks, key=lambda x: x.sort_order)
    selected_block = next((x for x in blocks if x.id == selected_block_id), None) or (blocks[0] if blocks else None)
    _presentation_refresh_ready_percent(presentation)
    db.session.commit()
    return render_template(
        'service_staff/presentation_card.html',
        presentation=presentation,
        blocks=blocks,
        selected_block=selected_block,
        users=_user_choices(),
        specialists=_specialist_choices(),
        block_status_choices=PRESENTATION_BLOCK_STATUS_CHOICES,
        presentation_status_choices=PRESENTATION_STATUS_CHOICES,
        can_manage=_can_edit_module() or _is_service_methodist(),
        can_edit_selected=_presentation_executor_can_edit(selected_block, presentation) if selected_block else False,
    )


@service_staff_bp.route('/presentations/<int:presentation_id>/blocks/<int:block_id>/save', methods=['POST'])
@login_required
def presentation_block_save(presentation_id: int, block_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    block = ServicePresentationBlock.query.filter_by(id=block_id, presentation_id=presentation.id).first_or_404()
    if not _presentation_executor_can_edit(block, presentation):
        abort(403)
    block.content_text = (request.form.get('content_text') or '').strip()
    block.status = (request.form.get('status') or block.status or 'IN_PROGRESS').upper()
    if block.status not in {x[0] for x in PRESENTATION_BLOCK_STATUS_CHOICES}:
        block.status = 'IN_PROGRESS'
    block.updated_by_user_id = getattr(current_user, 'id', None)
    presentation.last_changed_by_user_id = getattr(current_user, 'id', None)
    presentation.status = 'IN_PROGRESS' if presentation.status == 'DRAFT' else presentation.status
    _presentation_refresh_ready_percent(presentation)
    _log_presentation_history(presentation, 'block_save', block=block, comment='Сохранение блока')
    db.session.commit()
    flash('Блок сохранен.', 'success')
    return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id, block_id=block.id))


@service_staff_bp.route('/presentations/<int:presentation_id>/blocks/<int:block_id>/review', methods=['POST'])
@login_required
def presentation_block_review(presentation_id: int, block_id: int):
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    block = ServicePresentationBlock.query.filter_by(id=block_id, presentation_id=presentation.id).first_or_404()
    action = request.form.get('review_action') or 'check'
    comment = (request.form.get('reviewer_comment') or '').strip()
    if action == 'return':
        block.status = 'IN_PROGRESS'
        block.reviewer_comment = comment or 'Возвращено на доработку.'
        history_action = 'block_return'
        flash('Блок возвращен на доработку.', 'warning')
    else:
        block.status = 'CHECKED'
        block.reviewer_comment = comment
        history_action = 'block_check'
        flash('Блок проверен.', 'success')
    block.updated_by_user_id = getattr(current_user, 'id', None)
    _presentation_refresh_ready_percent(presentation)
    _log_presentation_history(presentation, history_action, block=block, comment=comment)
    db.session.commit()
    return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id, block_id=block.id))


@service_staff_bp.route('/presentations/<int:presentation_id>/assign', methods=['POST'])
@login_required
def presentation_assign_block(presentation_id: int):
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    block = ServicePresentationBlock.query.filter_by(id=request.form.get('block_id', type=int), presentation_id=presentation.id).first_or_404()
    block.executor_user_id = request.form.get('executor_user_id', type=int) or None
    block.executor_specialist_id = request.form.get('executor_specialist_id', type=int) or None
    _log_presentation_history(presentation, 'assign_executor', block=block, comment='Перераспределение блока')
    db.session.commit()
    flash('Исполнитель обновлен.', 'success')
    return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id, block_id=block.id))


@service_staff_bp.route('/presentations/<int:presentation_id>/status', methods=['POST'])
@login_required
def presentation_change_status(presentation_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    new_status = (request.form.get('status') or '').upper()
    if new_status not in {x[0] for x in PRESENTATION_STATUS_CHOICES}:
        abort(400)
    if new_status == 'APPROVED':
        missing = [b.title for b in presentation.blocks if b.is_required and (b.status or '').upper() != 'CHECKED']
        if missing:
            flash('Нельзя согласовать документ: не проверены обязательные блоки.', 'danger')
            return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id))
    presentation.status = new_status
    presentation.last_changed_by_user_id = getattr(current_user, 'id', None)
    _log_presentation_history(presentation, 'status_change', comment=f'Новый статус: {presentation.status_label}')
    db.session.commit()
    flash('Статус документа обновлен.', 'success')
    return redirect(url_for('service_staff.presentation_card', presentation_id=presentation.id))



@service_staff_bp.route('/presentations/<int:presentation_id>/delete', methods=['POST'])
@login_required
def presentation_delete(presentation_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    _log_presentation_history(presentation, 'delete', comment='Удаление представления')
    db.session.delete(presentation)
    db.session.commit()
    flash('Представление удалено.', 'success')
    return redirect(url_for('service_staff.presentations_registry'))


@service_staff_bp.route('/presentations/<int:presentation_id>/export.docx')
@login_required
def presentation_export_docx(presentation_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    bio = _render_presentation_docx(presentation)
    if presentation.status == 'APPROVED':
        presentation.status = 'EXPORTED'
    _log_presentation_history(presentation, 'export_docx', comment='Выгрузка DOCX')
    db.session.commit()
    return send_file(bio, as_attachment=True, download_name=f'presentation-{presentation.id}.docx', mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')


@service_staff_bp.route('/presentations/<int:presentation_id>/export.pdf')
@login_required
def presentation_export_pdf(presentation_id: int):
    _require_view()
    presentation = _presentation_visible_query().filter(ServicePresentation.id == presentation_id).first_or_404()
    bio = _render_presentation_pdf(presentation)
    if presentation.status == 'APPROVED':
        presentation.status = 'EXPORTED'
    _log_presentation_history(presentation, 'export_pdf', comment='Выгрузка PDF')
    db.session.commit()
    return send_file(bio, as_attachment=True, download_name=f'presentation-{presentation.id}.pdf', mimetype='application/pdf')


ROUNDING_RULE_CHOICES = [
    ("UP", "Округление вверх"),
    ("HALF_UP", "До 0,5"),
    ("INT", "До целого"),
]


def _presentation_grade_choices():
    rows = SchoolClass.query.filter(SchoolClass.grade.isnot(None)).order_by(SchoolClass.grade.asc(), SchoolClass.name.asc()).all()
    result = []
    seen = set()
    for row in rows:
        if row.grade not in seen:
            seen.add(row.grade)
            result.append(row.grade)
    return result


def _presentation_class_choices():
    return SchoolClass.query.filter_by(is_active=True, is_archived=False).order_by(SchoolClass.grade.asc(), SchoolClass.name.asc()).all()


def _current_children_query():
    return Child.query.options(joinedload(Child.enrollments).joinedload('school_class').joinedload('building')).filter(Child.status == 'ACTIVE')


def _child_category_labels(child):
    labels = []
    if getattr(child, 'is_ovz', False):
        labels.append('ОВЗ')
    if getattr(child, 'is_disabled', False):
        labels.append('Инвалидность')
    if getattr(child, 'is_vshu', False):
        labels.append('ВШУ')
    if getattr(child, 'is_low', False):
        labels.append('Низкие результаты')
    if getattr(child, 'is_az', False):
        labels.append('Адаптация')
    return labels or ['Без категории']


def _seed_rate_norms():
    _seed_specializations()
    changed = False
    for spec in ServiceSpecialization.query.order_by(ServiceSpecialization.sort_order.asc(), ServiceSpecialization.name.asc()).all():
        exists = ServiceRateNorm.query.filter_by(specialization_id=spec.id, building_id=None).first()
        if not exists:
            db.session.add(ServiceRateNorm(
                specialization_id=spec.id,
                building_id=None,
                effective_from=date.today(),
                children_per_rate=25.0,
                category_coefficient=1.0,
                weekly_hours_norm=36.0,
                complexity_coefficient=1.0,
                rounding_rule='UP',
                is_active=True,
            ))
            changed = True
    if changed:
        db.session.commit()


def _norm_visible_rows():
    _seed_rate_norms()
    q = ServiceRateNorm.query.options(joinedload(ServiceRateNorm.specialization), joinedload(ServiceRateNorm.building))
    specialization_id = request.args.get('specialization_id', type=int)
    building_id = request.args.get('building_id', type=int)
    if specialization_id:
        q = q.filter(ServiceRateNorm.specialization_id == specialization_id)
    if building_id:
        q = q.filter(ServiceRateNorm.building_id == building_id)
    q = q.order_by(ServiceRateNorm.is_active.desc(), ServiceRateNorm.effective_from.desc(), ServiceRateNorm.id.desc())
    page, per_page = resolve_pagination(default_per_page=20)
    items, pagination = paginate_list(q.all(), page=page, per_page=per_page)
    return items, pagination


def _parse_float(name, default=None):
    raw = (request.form.get(name) or '').strip().replace(',', '.')
    if not raw:
        return default
    try:
        return float(raw)
    except Exception:
        raise ValueError(f'Поле «{name}» должно быть числом.')


def _apply_norm_form(norm):
    norm.specialization_id = request.form.get('specialization_id', type=int)
    if not norm.specialization_id:
        raise ValueError('Выберите специализацию.')
    norm.building_id = request.form.get('building_id', type=int) or None
    norm.effective_from = _parse_date(request.form.get('effective_from')) or date.today()
    norm.children_per_rate = _parse_float('children_per_rate', 25.0)
    norm.category_coefficient = _parse_float('category_coefficient', 1.0)
    norm.weekly_hours_norm = _parse_float('weekly_hours_norm', 36.0)
    norm.complexity_coefficient = _parse_float('complexity_coefficient', 1.0)
    norm.rounding_rule = (request.form.get('rounding_rule') or 'UP').upper()
    if norm.rounding_rule not in {x[0] for x in ROUNDING_RULE_CHOICES}:
        norm.rounding_rule = 'UP'
    norm.is_active = bool(request.form.get('is_active'))
    norm.comment = (request.form.get('comment') or '').strip() or None


def _round_rate(value, rule='UP'):
    if value is None:
        return 0.0
    rule = (rule or 'UP').upper()
    if rule == 'INT':
        return float(round(value))
    if rule == 'HALF_UP':
        return ceil(value * 2) / 2
    return float(ceil(value))


def _pick_norm_for_specialist(spec_link_ids, building_id=None, target_date=None):
    target_date = target_date or date.today()
    if not spec_link_ids:
        return None
    q = ServiceRateNorm.query.filter(ServiceRateNorm.is_active.is_(True), ServiceRateNorm.specialization_id.in_(spec_link_ids), ServiceRateNorm.effective_from <= target_date)
    building_norm = None
    if building_id:
        building_norm = q.filter(ServiceRateNorm.building_id == building_id).order_by(ServiceRateNorm.effective_from.desc()).first()
    if building_norm:
        return building_norm
    return q.filter(ServiceRateNorm.building_id.is_(None)).order_by(ServiceRateNorm.effective_from.desc()).first()


def _fact_weekly_minutes(cyclegram):
    return sum((entry.minutes or 0) for entry in (cyclegram.entries or []))


def _analytics_filters():
    return {
        'building_id': request.args.get('building_id', type=int),
        'specialization_id': request.args.get('specialization_id', type=int),
        'specialist_id': request.args.get('specialist_id', type=int),
        'grade': request.args.get('grade', type=int),
        'class_id': request.args.get('class_id', type=int),
        'query': (request.args.get('q') or '').strip(),
    }


def _specialist_analytics_rows():
    _seed_rate_norms()
    filters = _analytics_filters()
    specialists = ServiceSpecialist.query.options(
        joinedload(ServiceSpecialist.main_building),
        joinedload(ServiceSpecialist.building_links).joinedload(ServiceSpecialistBuilding.building),
        joinedload(ServiceSpecialist.specialization_links).joinedload(ServiceSpecialistSpecialization.specialization),
        joinedload(ServiceSpecialist.child_assignments).joinedload(ServiceAssignment.child),
    ).order_by(ServiceSpecialist.last_name.asc(), ServiceSpecialist.first_name.asc()).all()
    if not (_can_edit_module() or _is_service_methodist()):
        linked = _linked_specialist()
        specialists = [x for x in specialists if linked and x.id == linked.id]
    rows = []
    warning_rows = []
    active_cyclegrams = {c.specialist_id: c for c in ServiceCyclegram.query.options(joinedload(ServiceCyclegram.entries)).filter(ServiceCyclegram.status.in_(['DRAFT', 'REVIEW', 'APPROVED'])).all()}
    class_obj = SchoolClass.query.get(filters['class_id']) if filters['class_id'] else None
    for specialist in specialists:
        if not specialist.is_active:
            continue
        spec_ids = [link.specialization_id for link in specialist.specialization_links if link.specialization_id]
        if filters['specialization_id'] and filters['specialization_id'] not in spec_ids:
            continue
        if filters['specialist_id'] and filters['specialist_id'] != specialist.id:
            continue
        if filters['building_id']:
            building_ids = {link.building_id for link in specialist.building_links}
            if filters['building_id'] not in building_ids and specialist.main_building_id != filters['building_id']:
                continue
        assignments = [a for a in (specialist.child_assignments or []) if (a.status or '').upper() == 'ACTIVE' and a.child]
        if filters['grade']:
            assignments = [a for a in assignments if getattr(getattr(a.child, 'current_class', None), 'grade', None) == filters['grade']]
        if class_obj:
            assignments = [a for a in assignments if getattr(getattr(a.child, 'current_class', None), 'id', None) == class_obj.id]
        if filters['query']:
            q = filters['query'].lower()
            assignments = [a for a in assignments if q in (a.child.fio or '').lower() or q in (specialist.fio or '').lower()]
            if not assignments and q not in (specialist.fio or '').lower():
                continue
        children_count = len({a.child_id for a in assignments})
        weighted = 0.0
        categories = defaultdict(int)
        for a in assignments:
            factor = 1.0
            labels = _child_category_labels(a.child)
            for label in labels:
                categories[label] += 1
            if getattr(a.child, 'is_ovz', False):
                factor = max(factor, 1.2)
            if getattr(a.child, 'is_disabled', False):
                factor = max(factor, 1.3)
            weighted += factor
        norm = _pick_norm_for_specialist(spec_ids, specialist.main_building_id)
        if not norm:
            warning_rows.append({'type': 'danger', 'message': f'Нет норматива для специалиста {specialist.fio}.'})
        children_per_rate = norm.children_per_rate if norm else 0
        recommended = _round_rate((weighted / children_per_rate) if children_per_rate else 0, norm.rounding_rule if norm else 'UP') if children_per_rate else 0
        actual_rate = float(specialist.rate_value or 0)
        deviation = round(actual_rate - recommended, 2)
        cyclegram = active_cyclegrams.get(specialist.id)
        weekly_minutes = _fact_weekly_minutes(cyclegram) if cyclegram else 0
        weekly_hours = round(weekly_minutes / 60, 2)
        weekly_norm = float(norm.weekly_hours_norm if norm else 36.0)
        rows.append({
            'specialist': specialist,
            'children_count': children_count,
            'weighted_children': round(weighted, 2),
            'categories': dict(categories),
            'norm': norm,
            'children_per_rate': children_per_rate,
            'recommended_rate': recommended,
            'actual_rate': actual_rate,
            'deviation': deviation,
            'weekly_hours': weekly_hours,
            'weekly_norm': weekly_norm,
            'cyclegram_status': cyclegram.status_label if cyclegram else 'Нет циклограммы',
            'presentations_in_work': ServicePresentation.query.filter(ServicePresentation.status.in_(['DRAFT','IN_PROGRESS','REVIEW']), ServicePresentation.child_id.in_([a.child_id for a in assignments] or [-1])).count(),
        })
        if not spec_ids:
            warning_rows.append({'type': 'warning', 'message': f'У специалиста {specialist.fio} не заполнена специализация.'})
        if specialist.main_building_id is None:
            warning_rows.append({'type': 'warning', 'message': f'У специалиста {specialist.fio} не заполнено основное здание.'})
    page, per_page = resolve_pagination(default_per_page=20)
    items, pagination = paginate_list(rows, page=page, per_page=per_page)
    return items, pagination, warning_rows


def _building_analytics_rows():
    filters = _analytics_filters()
    buildings = Building.query.order_by(Building.name.asc()).all()
    assignments = ServiceAssignment.query.options(joinedload(ServiceAssignment.child), joinedload(ServiceAssignment.specialist)).filter(ServiceAssignment.status == 'ACTIVE').all()
    specialists = ServiceSpecialist.query.options(joinedload(ServiceSpecialist.main_building)).filter_by(is_active=True).all()
    rows = []
    for building in buildings:
        if filters['building_id'] and building.id != filters['building_id']:
            continue
        b_assignments = [a for a in assignments if (a.building_id == building.id) or (getattr(getattr(a.child, 'current_building', None), 'id', None) == building.id)]
        if filters['grade']:
            b_assignments = [a for a in b_assignments if getattr(getattr(a.child, 'current_class', None), 'grade', None) == filters['grade']]
        if filters['class_id']:
            b_assignments = [a for a in b_assignments if getattr(getattr(a.child, 'current_class', None), 'id', None) == filters['class_id']]
        if filters['query']:
            q = filters['query'].lower()
            b_assignments = [a for a in b_assignments if q in (a.child.fio or '').lower() or q in (a.specialist.fio or '').lower()]
        b_children = len({a.child_id for a in b_assignments})
        b_specs = [s for s in specialists if s.main_building_id == building.id or any(link.building_id == building.id for link in s.building_links)]
        deficit = len([s for s in b_specs if float(s.rate_value or 0) < 1.0])
        rows.append({
            'building': building,
            'children_count': b_children,
            'specialists_count': len({s.id for s in b_specs}),
            'active_assignments': len(b_assignments),
            'deficit_count': deficit,
            'presentations_in_work': ServicePresentation.query.filter(ServicePresentation.building_id == building.id, ServicePresentation.status.in_(['DRAFT','IN_PROGRESS','REVIEW'])).count(),
        })
    page, per_page = resolve_pagination(default_per_page=20)
    items, pagination = paginate_list(rows, page=page, per_page=per_page)
    return items, pagination


def _service_dashboard_metrics():
    specialist_rows, _p, warnings = _specialist_analytics_rows()
    building_rows, _bp = _building_analytics_rows()
    all_spec_rows = specialist_rows
    # recalc without pagination for headline metrics
    all_spec_rows = _specialist_analytics_rows()[0]
    active_presentations = ServicePresentation.query.filter(ServicePresentation.status.in_(['DRAFT','IN_PROGRESS','REVIEW'])).count()
    missing_cyclegrams = len([r for r in all_spec_rows if r['cyclegram_status'] == 'Нет циклограммы'])
    overloaded = len([r for r in all_spec_rows if r['deviation'] < 0])
    return {
        'children_support': len({a.child_id for a in ServiceAssignment.query.filter_by(status='ACTIVE').all()}),
        'specialists_total': ServiceSpecialist.query.filter_by(is_active=True).count(),
        'active_presentations': active_presentations,
        'missing_cyclegrams': missing_cyclegrams,
        'deficit_directions': overloaded,
        'warnings_count': len(warnings),
    }


def _render_specialists_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сводка по специалистам'
    ws.append(['Специалист', 'Специализации', 'Здание', 'Детей', 'Приведенных детей', 'Норматив детей на ставку', 'Рекомендуемые ставки', 'Фактическая ставка', 'Отклонение', 'Часы по циклограмме', 'Норма часов', 'Представления в работе'])
    for row in rows:
        spec = row['specialist']
        ws.append([
            spec.fio,
            spec.specializations_text,
            (spec.main_building.short_name or spec.main_building.name) if spec.main_building else '—',
            row['children_count'],
            row['weighted_children'],
            row['children_per_rate'] or '—',
            row['recommended_rate'],
            row['actual_rate'],
            row['deviation'],
            row['weekly_hours'],
            row['weekly_norm'],
            row['presentations_in_work'],
        ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _render_buildings_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сводка по зданиям'
    ws.append(['Здание', 'Детей на сопровождении', 'Специалистов', 'Активных назначений', 'Дефицитных направлений', 'Представлений в работе'])
    for row in rows:
        building = row['building']
        ws.append([
            building.short_name or building.name,
            row['children_count'],
            row['specialists_count'],
            row['active_assignments'],
            row['deficit_count'],
            row['presentations_in_work'],
        ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _render_service_report_pdf(title, lines):
    bio = BytesIO()
    pdf = canvas.Canvas(bio, pagesize=A4)
    width, height = A4
    y = height - 40
    pdf.setFont('Helvetica-Bold', 12)
    pdf.drawString(40, y, title)
    y -= 24
    pdf.setFont('Helvetica', 10)
    for line in lines:
        if y < 50:
            pdf.showPage()
            y = height - 40
            pdf.setFont('Helvetica', 10)
        pdf.drawString(40, y, str(line)[:140])
        y -= 16
    pdf.save()
    bio.seek(0)
    return bio


@service_staff_bp.route('/norms')
@login_required
def norms_registry():
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    rows, pagination = _norm_visible_rows()
    return render_template('service_staff/norms_registry.html', rows=rows, pagination=pagination, specializations=_specialization_choices(), buildings=_building_choices(), rounding_rules=ROUNDING_RULE_CHOICES, can_edit=True)


@service_staff_bp.route('/norms/new', methods=['GET', 'POST'])
@login_required
def norm_new():
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    norm = ServiceRateNorm(effective_from=date.today(), is_active=True, children_per_rate=25, category_coefficient=1.0, weekly_hours_norm=36, complexity_coefficient=1.0, rounding_rule='UP')
    if request.method == 'POST':
        try:
            _apply_norm_form(norm)
            db.session.add(norm)
            db.session.commit()
            flash('Норматив сохранен.', 'success')
            return redirect(url_for('service_staff.norms_registry'))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('service_staff/norm_form.html', norm=norm, specializations=_specialization_choices(), buildings=_building_choices(), rounding_rules=ROUNDING_RULE_CHOICES)


@service_staff_bp.route('/norms/<int:norm_id>/edit', methods=['GET', 'POST'])
@login_required
def norm_edit(norm_id: int):
    _require_view()
    if not (_can_edit_module() or _is_service_methodist()):
        abort(403)
    norm = ServiceRateNorm.query.get_or_404(norm_id)
    if request.method == 'POST':
        try:
            _apply_norm_form(norm)
            db.session.commit()
            flash('Норматив обновлен.', 'success')
            return redirect(url_for('service_staff.norms_registry'))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('service_staff/norm_form.html', norm=norm, specializations=_specialization_choices(), buildings=_building_choices(), rounding_rules=ROUNDING_RULE_CHOICES)


@service_staff_bp.route('/analytics')
@login_required
def analytics_dashboard():
    _require_view()
    metrics = _service_dashboard_metrics()
    specialist_rows, _, warnings = _specialist_analytics_rows()
    building_rows, _ = _building_analytics_rows()
    return render_template('service_staff/analytics_dashboard.html', metrics=metrics, specialist_rows=specialist_rows[:10], building_rows=building_rows[:8], warnings=warnings[:10], buildings=_building_choices(), specializations=_specialization_choices(), specialists=_specialist_choices(), grade_choices=_presentation_grade_choices(), classes=_presentation_class_choices(), filters=_analytics_filters())


@service_staff_bp.route('/analytics/specialists')
@login_required
def analytics_specialists():
    _require_view()
    rows, pagination, warnings = _specialist_analytics_rows()
    return render_template('service_staff/analytics_specialists.html', rows=rows, pagination=pagination, warnings=warnings, buildings=_building_choices(), specializations=_specialization_choices(), specialists=_specialist_choices(), grade_choices=_presentation_grade_choices(), classes=_presentation_class_choices(), filters=_analytics_filters())


@service_staff_bp.route('/analytics/buildings')
@login_required
def analytics_buildings():
    _require_view()
    rows, pagination = _building_analytics_rows()
    return render_template('service_staff/analytics_buildings.html', rows=rows, pagination=pagination, buildings=_building_choices(), specializations=_specialization_choices(), specialists=_specialist_choices(), grade_choices=_presentation_grade_choices(), classes=_presentation_class_choices(), filters=_analytics_filters())


@service_staff_bp.route('/analytics/data-quality')
@login_required
def analytics_data_quality():
    _require_view()
    _rows, _pagination, warnings = _specialist_analytics_rows()
    return render_template('service_staff/analytics_data_quality.html', warnings=warnings)


@service_staff_bp.route('/analytics/specialists/export.xlsx')
@login_required
def analytics_specialists_export_xlsx():
    _require_view()
    rows, _pagination, _warnings = _specialist_analytics_rows()
    bio = _render_specialists_xlsx(rows)
    return send_file(bio, as_attachment=True, download_name='service-specialists-analytics.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@service_staff_bp.route('/analytics/buildings/export.xlsx')
@login_required
def analytics_buildings_export_xlsx():
    _require_view()
    rows, _pagination = _building_analytics_rows()
    bio = _render_buildings_xlsx(rows)
    return send_file(bio, as_attachment=True, download_name='service-buildings-analytics.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@service_staff_bp.route('/analytics/report.pdf')
@login_required
def analytics_report_pdf():
    _require_view()
    metrics = _service_dashboard_metrics()
    rows, _pagination, warnings = _specialist_analytics_rows()
    lines = [
        f"Детей на сопровождении: {metrics['children_support']}",
        f"Специалистов: {metrics['specialists_total']}",
        f"Активных представлений: {metrics['active_presentations']}",
        f"Незаполненных циклограмм: {metrics['missing_cyclegrams']}",
        f"Дефицитных направлений: {metrics['deficit_directions']}",
        '',
        'Сводка по специалистам:',
    ]
    for row in rows[:30]:
        lines.append(f"{row['specialist'].fio}: детей {row['children_count']}, рекомендовано ставок {row['recommended_rate']}, фактически {row['actual_rate']}, отклонение {row['deviation']}")
    if warnings:
        lines.append('')
        lines.append('Предупреждения:')
        for warning in warnings[:20]:
            lines.append(warning['message'])
    bio = _render_service_report_pdf('Аналитическая справка социально-психологической службы', lines)
    return send_file(bio, as_attachment=True, download_name='service-analytics-report.pdf', mimetype='application/pdf')
