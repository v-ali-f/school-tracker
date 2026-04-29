from datetime import datetime
from io import BytesIO
from collections import defaultdict
from xml.sax.saxutils import escape

from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file, jsonify
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy import func as sa_func

from app.core.extensions import db
from .models import (
    ControlWork, ControlWorkTask, ControlWorkAssignment, ControlWorkResult, ControlWorkLog,
    SchoolClass, User, ChildEnrollment, Child, AcademicYear, Subject, TeacherLoad, DepartmentLeader
)
from .permissions import has_permission
from app.services.org_settings_service import get_organization_header_lines, get_organization_signature_block

control_bp = Blueprint("control_works", __name__, url_prefix="/control-works")


def _parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def _current_year():
    return AcademicYear.query.filter_by(is_current=True).first()


def _fmt_date(value):
    return value.strftime("%d.%m.%Y") if value else "—"

WORK_KIND_CHOICES = ControlWork.WORK_KIND_CHOICES
WORK_KIND_DEFAULT = ControlWork.WORK_KIND_CONTROL
DICTATION_KIND = ControlWork.WORK_KIND_DICTATION


def _is_dictation(work):
    return getattr(work, 'work_kind', None) == DICTATION_KIND


def _parse_int_field(name, label, min_value=0, max_value=None):
    raw = (request.form.get(name) or '').strip()
    if raw == '':
        return None
    try:
        value = int(raw)
    except ValueError:
        raise ValueError(f"{label}: нужно указать целое число.")
    if value < min_value:
        raise ValueError(f"{label}: допустимо значение не меньше {min_value}.")
    if max_value is not None and value > max_value:
        raise ValueError(f"{label}: допустимо значение не больше {max_value}.")
    return value


def _dictation_mark_from_errors(spelling_errors, punctuation_errors, work=None, grammar_errors=0, corrections_count=0):
    s = int(spelling_errors or 0)
    p = int(punctuation_errors or 0)
    g = int(grammar_errors or 0)
    c = int(corrections_count or 0)
    cfg = {
        "g5s": getattr(work, "dictation_grade5_spelling_max", 0) if work else 0,
        "g5p": getattr(work, "dictation_grade5_punctuation_max", 0) if work else 0,
        "g4s": getattr(work, "dictation_grade4_spelling_max", 2) if work else 2,
        "g4p": getattr(work, "dictation_grade4_punctuation_max", 2) if work else 2,
        "g3s": getattr(work, "dictation_grade3_spelling_max", 4) if work else 4,
        "g3p": getattr(work, "dictation_grade3_punctuation_max", 4) if work else 4,
        "use_g": bool(getattr(work, "dictation_use_grammar_errors", False)) if work else False,
        "use_c": bool(getattr(work, "dictation_use_corrections", False)) if work else False,
    }
    if cfg["use_g"]:
        s += g
    if cfg["use_c"]:
        p += c
    if s <= cfg["g5s"] and p <= cfg["g5p"]:
        return 5
    if s <= cfg["g4s"] and p <= cfg["g4p"]:
        return 4
    if s <= cfg["g3s"] and p <= cfg["g3p"]:
        return 3
    return 2


def _dictation_rules_text(work):
    return (
        f"5 — до {getattr(work, 'dictation_grade5_spelling_max', 0)} орф. и до {getattr(work, 'dictation_grade5_punctuation_max', 0)} пункт.; "
        f"4 — до {getattr(work, 'dictation_grade4_spelling_max', 2)} орф. и до {getattr(work, 'dictation_grade4_punctuation_max', 2)} пункт.; "
        f"3 — до {getattr(work, 'dictation_grade3_spelling_max', 4)} орф. и до {getattr(work, 'dictation_grade3_punctuation_max', 4)} пункт.; "
        f"2 — выше порога отметки 3"
    )


def _dictation_settings_from_form(work=None):
    return {
        "dictation_grade5_spelling_max": _parse_int_field("dictation_grade5_spelling_max", "Диктант: орфографические ошибки для отметки 5", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade5_spelling_max", 0),
        "dictation_grade5_punctuation_max": _parse_int_field("dictation_grade5_punctuation_max", "Диктант: пунктуационные ошибки для отметки 5", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade5_punctuation_max", 0),
        "dictation_grade4_spelling_max": _parse_int_field("dictation_grade4_spelling_max", "Диктант: орфографические ошибки для отметки 4", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade4_spelling_max", 2),
        "dictation_grade4_punctuation_max": _parse_int_field("dictation_grade4_punctuation_max", "Диктант: пунктуационные ошибки для отметки 4", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade4_punctuation_max", 2),
        "dictation_grade3_spelling_max": _parse_int_field("dictation_grade3_spelling_max", "Диктант: орфографические ошибки для отметки 3", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade3_spelling_max", 4),
        "dictation_grade3_punctuation_max": _parse_int_field("dictation_grade3_punctuation_max", "Диктант: пунктуационные ошибки для отметки 3", min_value=0) if request.method == "POST" else getattr(work, "dictation_grade3_punctuation_max", 4),
        "dictation_use_grammar_errors": request.form.get("dictation_use_grammar_errors") == "1" if request.method == "POST" else bool(getattr(work, "dictation_use_grammar_errors", False)),
        "dictation_use_corrections": request.form.get("dictation_use_corrections") == "1" if request.method == "POST" else bool(getattr(work, "dictation_use_corrections", False)),
    }


def _selected_work_kind(work=None):
    value = (request.form.get("work_kind") or (getattr(work, "work_kind", None) if work else None) or WORK_KIND_DEFAULT).strip()
    allowed = {key for key, _ in WORK_KIND_CHOICES}
    return value if value in allowed else WORK_KIND_DEFAULT


def _safe_avg(values, digits=2):
    vals = [v for v in values if v is not None]
    return round(sum(vals) / len(vals), digits) if vals else None


def _score_status(percent):
    if percent is None:
        return "Нет данных"
    if percent < 50:
        return "Критично"
    if percent < 80:
        return "Стабильно"
    return "Высокий результат"


def _bar_class(percent):
    if percent is None:
        return "bg-secondary"
    if percent < 50:
        return "bg-danger"
    if percent < 65:
        return "bg-warning"
    if percent < 80:
        return "bg-primary"
    return "bg-success"






def _work_status_badge(status):
    mapping = {
        "Черновик": "secondary",
        "Заполнение": "warning",
        "Заполнено": "success",
        "Проверено": "primary",
        "Закрыто": "dark",
        "Просрочено": "danger",
    }
    return mapping.get(status, "secondary")



MANUAL_STATUS_CHOICES = ["Черновик", "Заполнение", "Заполнено", "Проверено", "Закрыто", "Архив"]


def _work_base_status(work, total_students, processed, has_assignments):
    if total_students == 0 and not has_assignments:
        return "Черновик"
    if processed == 0:
        return "Просрочено" if work.deadline_date and work.deadline_date < datetime.utcnow().date() else "Черновик"
    if total_students and processed < total_students:
        return "Просрочено" if work.deadline_date and work.deadline_date < datetime.utcnow().date() else "Заполнение"
    return "Заполнено"


def _work_effective_status(work, total_students, processed, has_assignments):
    manual = (getattr(work, "manual_status", None) or "").strip()
    if manual == "Архив" or getattr(work, "is_archived", False):
        return "Архив"
    if manual in {"Проверено", "Закрыто"}:
        return manual
    return _work_base_status(work, total_students, processed, has_assignments)


def _work_attention_reasons(work, total_students, processed, absent, has_assignments):
    reasons = []
    base_status = _work_base_status(work, total_students, processed, has_assignments)
    today = datetime.utcnow().date()
    if has_assignments and processed == 0:
        reasons.append("Работа создана, но заполнение не начато")
    if total_students and processed < total_students:
        reasons.append("Есть незаполненные результаты")
    if total_students and processed > total_students:
        reasons.append("Результатов больше, чем учеников по списку")
    if absent > 0:
        reasons.append("Есть отсутствующие")
    if work.deadline_date and work.deadline_date < today and base_status != "Заполнено":
        reasons.append("Истек срок заполнения")
    if base_status == "Черновик" and work.created_at and (datetime.utcnow() - work.created_at).days >= 7:
        reasons.append("Долго находится в статусе «Черновик»")
    if base_status == "Заполнение" and getattr(work, "updated_at", None) and (datetime.utcnow() - work.updated_at).days >= 7:
        reasons.append("Долго находится в статусе «Заполнение»")
    uniq = []
    for reason in reasons:
        if reason not in uniq:
            uniq.append(reason)
    return uniq


def _log_control_work_event(work, event_type, title, old_value=None, new_value=None, details=None, user_id=None):
    if not work or not getattr(work, "id", None):
        return
    db.session.add(ControlWorkLog(
        control_work_id=work.id,
        user_id=user_id or getattr(current_user, "id", None),
        event_type=event_type,
        title=title,
        old_value=(str(old_value) if old_value not in (None, "") else None),
        new_value=(str(new_value) if new_value not in (None, "") else None),
        details=details,
    ))


def _collect_work_change_logs(work, previous):
    current = {
        "theme": work.theme or "",
        "work_kind": work.work_kind_label,
        "work_date": _fmt_date(work.work_date),
        "deadline_date": _fmt_date(work.deadline_date),
        "subject": str(work.subject_name or ""),
        "manual_status": getattr(work, "manual_status", None) or "",
    }
    labels = {
        "theme": "Изменено название работы",
        "work_kind": "Изменен вид работы",
        "work_date": "Изменена дата проведения",
        "deadline_date": "Изменен срок заполнения",
        "subject": "Изменен предмет",
        "manual_status": "Изменен управленческий статус",
    }
    for key, label in labels.items():
        old = previous.get(key, "") if previous else ""
        new = current.get(key, "")
        if old != new:
            _log_control_work_event(work, key, label, old_value=old or "—", new_value=new or "—")


def _query_args_without_page(extra=None):
    data = request.args.to_dict(flat=True)
    data.pop("page", None)
    if extra:
        for k, v in extra.items():
            if v is None or v == "":
                data.pop(k, None)
            else:
                data[k] = v
    return data


def _build_work_registry_row(work, precomputed_results=None, precomputed_total_students=None):
    """s85: precomputed_* подаются из batch-загрузки в _build_registry_dataset.
    Если None — fallback на одиночные запросы (для обратной совместимости).
    """
    assignments = list(work.assignments or [])
    class_ids = [a.school_class_id for a in assignments if a.school_class_id]
    year_id = work.academic_year_id or (_current_year().id if _current_year() else None)

    if precomputed_total_students is not None:
        total_students = precomputed_total_students
    else:
        total_students = 0
        if class_ids and year_id:
            total_students = (
                ChildEnrollment.query
                .filter(
                    ChildEnrollment.academic_year_id == year_id,
                    ChildEnrollment.school_class_id.in_(class_ids),
                    ChildEnrollment.ended_at.is_(None),
                )
                .count()
            )

    if precomputed_results is not None:
        results = precomputed_results
    else:
        results = ControlWorkResult.query.filter_by(control_work_id=work.id).all()
    processed = 0
    absent = 0
    participants = 0
    dictation_marks = []
    grammar_marks = []
    final_marks = []
    error_totals = []

    for row in results:
        if getattr(row, "is_absent", False):
            absent += 1
            processed += 1
            continue
        if _is_dictation(work):
            has_value = any([
                row.dictation_mark is not None,
                row.grammar_mark is not None,
                row.final_mark is not None,
                row.spelling_errors is not None,
                row.punctuation_errors is not None,
                row.grammar_errors is not None,
                row.corrections_count is not None,
            ])
            if has_value:
                processed += 1
                participants += 1
                if row.dictation_mark is not None:
                    dictation_marks.append(row.dictation_mark)
                if row.grammar_mark is not None:
                    grammar_marks.append(row.grammar_mark)
                eff_final = row.final_mark if row.final_mark is not None else row.dictation_mark
                if eff_final is not None:
                    final_marks.append(eff_final)
                errs = [v for v in [row.spelling_errors, row.punctuation_errors, row.grammar_errors] if v is not None]
                if errs:
                    error_totals.append(sum(errs))
        else:
            if row.mark is not None or row.percent is not None or row.total_score is not None:
                processed += 1
                participants += 1

    unfilled = max(total_students - processed, 0) if total_students else 0
    participation_percent = round((participants / total_students) * 100, 1) if total_students else None
    completion_percent = round((processed / total_students) * 100, 1) if total_students else None

    unique_class_names = []
    seen_classes = set()
    teacher_names = []
    seen_teachers = set()
    for a in assignments:
        class_name = a.school_class.name if a.school_class else "—"
        if class_name not in seen_classes:
            seen_classes.add(class_name)
            unique_class_names.append(class_name)
        teacher_name = a.teacher.fio if a.teacher else "Не назначен"
        if teacher_name not in seen_teachers:
            seen_teachers.add(teacher_name)
            teacher_names.append(teacher_name)

    base_status = _work_base_status(work, total_students, processed, bool(assignments))
    status = _work_effective_status(work, total_students, processed, bool(assignments))
    attention_reasons = _work_attention_reasons(work, total_students, processed, absent, bool(assignments))

    return {
        "work": work,
        "work_id": work.id,
        "theme": work.theme or "—",
        "work_kind": work.work_kind,
        "work_kind_label": work.work_kind_label,
        "subject_id": work.subject_id,
        "subject_name": work.subject_name or "—",
        "academic_year_id": work.academic_year_id,
        "parallel": getattr(work, "parallel", None),
        "class_names": ", ".join(unique_class_names) if unique_class_names else "—",
        "class_ids": class_ids,
        "teacher_names": ", ".join(teacher_names) if teacher_names else "—",
        "teacher_ids": [a.teacher_id for a in assignments if a.teacher_id],
        "work_date": work.work_date,
        "deadline_date": work.deadline_date,
        "created_at": work.created_at,
        "created_by_name": work.creator.fio if getattr(work, "creator", None) else "—",
        "updated_at": getattr(work, "updated_at", None),
        "updated_by_name": work.updater.fio if getattr(work, "updater", None) else (work.creator.fio if getattr(work, "creator", None) else "—"),
        "assignments_count": len(assignments),
        "total_students": total_students,
        "participants": participants,
        "absent": absent,
        "processed": processed,
        "unfilled": unfilled,
        "participation_percent": participation_percent,
        "completion_percent": completion_percent,
        "base_status": base_status,
        "status": status,
        "status_badge": _work_status_badge(status),
        "manual_status": getattr(work, "manual_status", None),
        "is_archived": bool(getattr(work, "is_archived", False)),
        "is_overdue": bool(work.deadline_date and work.deadline_date < datetime.utcnow().date() and status not in ["Заполнено", "Проверено", "Закрыто", "Архив"]),
        "days_left": (work.deadline_date - datetime.utcnow().date()).days if work.deadline_date else None,
        "has_absent": absent > 0,
        "requires_attention": bool(attention_reasons),
        "attention_reasons": attention_reasons,
        "attention_reason": "; ".join(attention_reasons[:2]) if attention_reasons else "",
        "dictation_has_grammar": _is_dictation(work),
        "dictation_avg": _safe_avg(dictation_marks, 2),
        "grammar_avg": _safe_avg(grammar_marks, 2),
        "final_avg": _safe_avg(final_marks, 2),
        "errors_avg": _safe_avg(error_totals, 2),
    }


def _build_registry_dataset(selected_year_id=None, selected_subject_id=None, selected_teacher_id=None, selected_class_id=None, selected_grade=None, selected_work_kind=None, selected_status=None, search_text=None, overdue_only=False, has_absent=None, only_attention=False, include_archived=False):
    query = _assigned_works_query().order_by(ControlWork.work_date.desc().nullslast(), ControlWork.created_at.desc())
    if selected_year_id:
        query = query.filter(ControlWork.academic_year_id == selected_year_id)
    if selected_subject_id:
        query = query.filter(ControlWork.subject_id == selected_subject_id)
    if selected_grade:
        query = query.filter(ControlWork.parallel == selected_grade)
    if selected_work_kind:
        query = query.filter(ControlWork.work_kind == selected_work_kind)
    if selected_class_id:
        query = query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).filter(ControlWorkAssignment.school_class_id == selected_class_id)
    if selected_teacher_id:
        query = query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).filter(ControlWorkAssignment.teacher_id == selected_teacher_id)

    # s85: eager-load assignments + связанные классы/учителя/авторы — снимает lazy loads в _build_work_registry_row
    query = query.options(
        selectinload(ControlWork.assignments).joinedload(ControlWorkAssignment.school_class),
        selectinload(ControlWork.assignments).joinedload(ControlWorkAssignment.teacher),
        joinedload(ControlWork.creator),
        joinedload(ControlWork.updater),
        joinedload(ControlWork.subject_ref),
    )
    works = query.distinct().all()

    # s85: один SELECT на все ControlWorkResult, группировка в Python — заменяет N запросов.
    work_ids = [w.id for w in works]
    results_by_work = defaultdict(list)
    if work_ids:
        for r in ControlWorkResult.query.filter(ControlWorkResult.control_work_id.in_(work_ids)).all():
            results_by_work[r.control_work_id].append(r)

    # s85: один GROUP BY на все enrollment-counts вместо N COUNT-запросов.
    # Каждая работа имеет свой year_id и набор class_ids — собираем (year_id, class_id) пары,
    # одним запросом считаем enrollments по всем таким парам, потом суммируем для каждой работы.
    pair_keys = set()  # set of (year_id, class_id)
    for w in works:
        y = w.academic_year_id or (_current_year().id if _current_year() else None)
        if not y:
            continue
        for a in (w.assignments or []):
            if a.school_class_id:
                pair_keys.add((y, a.school_class_id))
    enrollment_count_by_pair = {}
    if pair_keys:
        years_set = {p[0] for p in pair_keys}
        classes_set = {p[1] for p in pair_keys}
        rows_cnt = (
            db.session.query(
                ChildEnrollment.academic_year_id,
                ChildEnrollment.school_class_id,
                sa_func.count(ChildEnrollment.id),
            )
            .filter(
                ChildEnrollment.academic_year_id.in_(years_set),
                ChildEnrollment.school_class_id.in_(classes_set),
                ChildEnrollment.ended_at.is_(None),
            )
            .group_by(ChildEnrollment.academic_year_id, ChildEnrollment.school_class_id)
            .all()
        )
        for yid, cid, cnt in rows_cnt:
            enrollment_count_by_pair[(yid, cid)] = cnt

    def _total_students_for(w):
        y = w.academic_year_id or (_current_year().id if _current_year() else None)
        if not y:
            return 0
        return sum(
            enrollment_count_by_pair.get((y, a.school_class_id), 0)
            for a in (w.assignments or [])
            if a.school_class_id
        )

    rows = [
        _build_work_registry_row(
            work,
            precomputed_results=results_by_work.get(work.id, []),
            precomputed_total_students=_total_students_for(work),
        )
        for work in works
    ]

    if selected_status:
        rows = [r for r in rows if r["status"] == selected_status]
    if only_attention:
        rows = [r for r in rows if r["requires_attention"]]
    if not include_archived:
        rows = [r for r in rows if not r["is_archived"]]
    if overdue_only:
        rows = [r for r in rows if r["is_overdue"]]
    if has_absent is True:
        rows = [r for r in rows if r["has_absent"]]
    elif has_absent is False:
        rows = [r for r in rows if not r["has_absent"]]
    if search_text:
        token = search_text.strip().lower()
        rows = [r for r in rows if token in (f"{r['theme']} {r['subject_name']} {r['class_names']} {r['teacher_names']}".lower())]

    summary = {
        "works_count": len(rows),
        "filled_count": sum(1 for r in rows if r["status"] == "Заполнено"),
        "overdue_count": sum(1 for r in rows if r["is_overdue"]),
        "participants": sum(r["participants"] for r in rows),
        "absent": sum(r["absent"] for r in rows),
        "unfilled": sum(r["unfilled"] for r in rows),
        "completion_avg": _safe_avg([r["completion_percent"] for r in rows if r["completion_percent"] is not None], 1),
        "attention_count": sum(1 for r in rows if r["requires_attention"]),
        "draft_count": sum(1 for r in rows if r["status"] == "Черновик"),
    }
    return {"rows": rows, "summary": summary}


def _make_registry_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"
    headers = [
        "Дата", "Название", "Вид работы", "Предмет", "Параллель", "Классы", "Учитель", "Срок сдачи",
        "Статус", "Требует внимания", "Причина", "По списку", "Писали", "Отсутствовали", "Не заполнено", "Участие %", "Заполнение %",
        "Создано", "Создал", "Изменено", "Изменил", "Ср. диктант", "Ср. грамм.", "Ср. итог", "Ср. ошибок"
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            _fmt_date(row["work_date"]), row["theme"], row["work_kind_label"], row["subject_name"], row["parallel"] or "—", row["class_names"],
            row["teacher_names"], _fmt_date(row["deadline_date"]), row["status"], "Да" if row["requires_attention"] else "Нет", row["attention_reason"], row["total_students"], row["participants"], row["absent"],
            row["unfilled"], row["participation_percent"], row["completion_percent"], _fmt_date(row["created_at"].date()) if row["created_at"] else "—", row["created_by_name"], _fmt_date(row["updated_at"].date()) if row["updated_at"] else "—", row["updated_by_name"],
            row["dictation_avg"], row["grammar_avg"], row["final_avg"], row["errors_avg"]
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
def _department_ids_for_user(user=None):
    user = user or current_user
    if has_permission("control_works_edit", user=user) or getattr(user, "role", None) == "ADMIN":
        return None
    if getattr(user, "role", None) == "METHODIST":
        dep_ids = sorted({row.department_id for row in DepartmentLeader.query.filter_by(user_id=user.id).all() if row.department_id})
        return dep_ids or None
    return []


def _normalize_class_token(value):
    raw = (value or '').strip().upper()
    if not raw:
        return ''
    translate_map = str.maketrans({
        'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н', 'K': 'К', 'M': 'М',
        'O': 'О', 'P': 'Р', 'T': 'Т', 'X': 'Х', 'Y': 'У'
    })
    raw = raw.translate(translate_map)
    for ch in (' ', '-', '–', '—', '.', '/', '\\'):
        raw = raw.replace(ch, '')
    return raw


def _split_class_tokens(value):
    raw = (value or '')
    if not raw:
        return []
    prepared = raw
    for sep in ('\n', ';', ',', '|'):
        prepared = prepared.replace(sep, ' ')
    parts = [part for part in prepared.split() if part]
    compact = _normalize_class_token(raw)
    tokens = {_normalize_class_token(part) for part in parts if _normalize_class_token(part)}
    if compact:
        tokens.add(compact)
    return list(tokens)


def _teacher_load_candidates(subject_id=None, grade=None, class_name=None, academic_year_id=None, department_ids=None):
    query = TeacherLoad.query.filter(TeacherLoad.is_archived.is_(False))
    if academic_year_id:
        query = query.filter((TeacherLoad.academic_year_id == academic_year_id) | (TeacherLoad.academic_year_id.is_(None)))
    if subject_id:
        query = query.filter(TeacherLoad.subject_id == subject_id)
    if department_ids:
        query = query.filter(TeacherLoad.department_id.in_(department_ids))
    if grade is not None:
        query = query.filter((TeacherLoad.grade == grade) | (TeacherLoad.grade.is_(None)))

    rows = query.all()
    class_token = _normalize_class_token(class_name)
    matched = []
    for row in rows:
        if not row.teacher_id:
            continue

        row_tokens = _split_class_tokens(row.class_name)
        exact_match = False
        partial_match = False

        if class_token:
            if row_tokens:
                exact_match = class_token in row_tokens
                if not exact_match:
                    partial_match = any(class_token in token or token in class_token for token in row_tokens)
                else:
                    partial_match = True
                if not partial_match and row.grade is not None and grade is not None and row.grade != grade:
                    continue
            elif row.grade is not None and grade is not None and row.grade != grade:
                continue

        matched.append((row, exact_match, partial_match))

    unique = {}
    for row, exact_match, partial_match in matched:
        current = unique.get(row.teacher_id)
        rank = 2 if exact_match else (1 if partial_match else 0)
        if current is None or rank > current[1] or (rank == current[1] and float(row.hours or 0) > float(current[0].hours or 0)):
            unique[row.teacher_id] = (row, rank)

    ranked_rows = []
    preferred_teacher_ids = [row.teacher_id for row, rank in unique.values() if rank == 2]
    single_preferred_teacher_id = preferred_teacher_ids[0] if len(preferred_teacher_ids) == 1 else None

    for row, rank in unique.values():
        ranked_rows.append({
            'id': row.teacher_id,
            'name': row.teacher.fio if row.teacher else (row.subject_name or 'Учитель'),
            'hours': float(row.hours or 0),
            'class_name': row.class_name or '',
            'department_id': row.department_id,
            'preferred': bool(single_preferred_teacher_id and row.teacher_id == single_preferred_teacher_id),
            'match_rank': rank,
        })

    if not ranked_rows:
        return ranked_rows

    if not single_preferred_teacher_id and len(ranked_rows) == 1:
        ranked_rows[0]['preferred'] = True

    ranked_rows.sort(key=lambda x: (-x['match_rank'], -x['hours'], x['name']))
    return ranked_rows


def _teacher_options_map(classes, subject_id=None, academic_year_id=None, department_ids=None):
    data = {}
    for c in classes:
        data[c.id] = _teacher_load_candidates(
            subject_id=subject_id,
            grade=c.grade,
            class_name=c.name,
            academic_year_id=academic_year_id,
            department_ids=department_ids,
        )
    return data


def _auto_teacher_by_class(classes, teacher_options_map):
    auto_map = {}
    for c in classes:
        options = teacher_options_map.get(c.id, []) or []
        preferred = next((item for item in options if item.get('preferred')), None)
        if preferred:
            auto_map[c.id] = preferred['id']
        elif len(options) == 1:
            auto_map[c.id] = options[0]['id']
    return auto_map


def _get_archive_filters():
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    teachers = User.query.order_by(User.last_name.asc(), User.first_name.asc()).all()

    selected_year_id = request.args.get("academic_year_id", type=int)
    selected_subject_id = request.args.get("subject_id", type=int)
    selected_teacher_id = request.args.get("teacher_id", type=int)
    if not has_permission("control_works_edit"):
        selected_teacher_id = current_user.id

    return {
        "years": years,
        "subjects": subjects,
        "teachers": teachers,
        "selected_year_id": selected_year_id,
        "selected_subject_id": selected_subject_id,
        "selected_teacher_id": selected_teacher_id,
    }


def _archive_works_query(filters):
    query = ControlWork.query
    if filters["selected_year_id"]:
        query = query.filter(ControlWork.academic_year_id == filters["selected_year_id"])
    if filters["selected_subject_id"]:
        query = query.filter(ControlWork.subject_id == filters["selected_subject_id"])
    if filters["selected_teacher_id"]:
        query = query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).filter(ControlWorkAssignment.teacher_id == filters["selected_teacher_id"])
    return query.distinct().order_by(ControlWork.work_date.desc().nullslast(), ControlWork.created_at.desc())


def _build_control_work_report(work, teacher_id=None):
    tasks = sorted((work.tasks or []), key=lambda x: x.task_number or 0)
    class_rows = []
    all_results = []
    present_results = []
    absent_results = []
    mark_counts = {2: 0, 3: 0, 4: 0, 5: 0}
    topic_stats = {}
    dictation_class_rows = []
    dictation_child_rows = []
    is_dictation = _is_dictation(work)

    # Batch-load ALL results for this work at once (instead of per-assignment)
    _all_work_results = (
        ControlWorkResult.query
        .filter_by(control_work_id=work.id)
        .options(joinedload(ControlWorkResult.child))
        .all()
    )
    _results_by_key = {}
    for r in _all_work_results:
        key = (r.assignment_id, r.school_class_id)
        _results_by_key.setdefault(key, []).append(r)

    for assignment in work.assignments or []:
        if teacher_id and assignment.teacher_id != teacher_id:
            continue
        if not has_permission("control_works_edit") and teacher_id is None and assignment.teacher_id != current_user.id:
            continue

        results = _results_by_key.get((assignment.id, assignment.school_class_id), [])
        all_results.extend(results)
        class_present = [r for r in results if not getattr(r, "is_absent", False)]
        class_absent = [r for r in results if getattr(r, "is_absent", False)]
        present_results.extend(class_present)
        absent_results.extend(class_absent)
        marks_source = [r.final_mark for r in class_present if is_dictation and r.final_mark is not None]
        if not marks_source:
            marks_source = [r.mark for r in class_present if r.mark is not None]
        avg_mark = _safe_avg(marks_source)
        avg_percent = _safe_avg([r.percent for r in class_present if r.percent is not None])
        for m in marks_source:
            if m in mark_counts:
                mark_counts[m] += 1
        class_row = {
            "assignment": assignment,
            "results_count": len(results),
            "participants_count": len(class_present),
            "absent_count": len(class_absent),
            "avg_mark": avg_mark,
            "avg_percent": avg_percent,
            "marks": {m: sum(1 for x in marks_source if x == m) for m in [2, 3, 4, 5]},
            "status": _score_status(avg_percent),
            "bar_class": _bar_class(avg_percent),
        }
        if is_dictation:
            class_row.update({
                "avg_dictation_mark": _safe_avg([r.dictation_mark for r in class_present if r.dictation_mark is not None]),
                "avg_grammar_mark": _safe_avg([r.grammar_mark for r in class_present if r.grammar_mark is not None]),
                "avg_final_mark": _safe_avg([r.final_mark for r in class_present if r.final_mark is not None]),
                "avg_spelling_errors": _safe_avg([r.spelling_errors for r in class_present if r.spelling_errors is not None]),
                "avg_punctuation_errors": _safe_avg([r.punctuation_errors for r in class_present if r.punctuation_errors is not None]),
                "avg_grammar_errors": _safe_avg([r.grammar_errors for r in class_present if r.grammar_errors is not None]),
                "avg_corrections_count": _safe_avg([r.corrections_count for r in class_present if r.corrections_count is not None]),
            })
            dictation_class_rows.append(class_row.copy())
            for r in results:
                child = getattr(r, 'child', None)
                dictation_child_rows.append({
                    "child_name": child.fio if child else f"ID {r.child_id}",
                    "class_name": assignment.school_class.name if assignment.school_class else '—',
                    "teacher_name": assignment.teacher.fio if assignment.teacher else '—',
                    "is_absent": bool(getattr(r, 'is_absent', False)),
                    "dictation_mark": None if getattr(r, 'is_absent', False) else r.dictation_mark,
                    "grammar_mark": None if getattr(r, 'is_absent', False) else r.grammar_mark,
                    "final_mark": None if getattr(r, 'is_absent', False) else (r.final_mark if r.final_mark is not None else r.mark),
                    "spelling_errors": None if getattr(r, 'is_absent', False) else r.spelling_errors,
                    "punctuation_errors": None if getattr(r, 'is_absent', False) else r.punctuation_errors,
                    "grammar_errors": None if getattr(r, 'is_absent', False) else r.grammar_errors,
                    "corrections_count": None if getattr(r, 'is_absent', False) else r.corrections_count,
                    "teacher_comment": r.teacher_comment,
                })
        class_rows.append(class_row)

    task_rows = []
    results_with_scores = [r for r in present_results if r.total_score is not None]
    max_total = sum(t.max_score or 0 for t in tasks)
    if not is_dictation:
        for task in tasks:
            avg_score = None
            percent = None
            if results_with_scores and max_total > 0 and (task.max_score or 0) > 0:
                est_scores = []
                for r in results_with_scores:
                    est_scores.append((r.total_score or 0) * (task.max_score or 0) / max_total)
                avg_score = round(sum(est_scores) / len(est_scores), 2) if est_scores else None
                percent = round((avg_score / (task.max_score or 1)) * 100, 2) if avg_score is not None else None
            task_rows.append({
                "task": task,
                "avg_score": avg_score,
                "percent": percent,
                "status": _score_status(percent),
                "bar_class": _bar_class(percent),
            })
            topic_key = (task.topic or "—").strip()
            if topic_key not in topic_stats:
                topic_stats[topic_key] = {"sum": 0, "count": 0}
            if percent is not None:
                topic_stats[topic_key]["sum"] += percent
                topic_stats[topic_key]["count"] += 1

    topic_rows = []
    for topic, data in topic_stats.items():
        avg_percent = round(data["sum"] / data["count"], 2) if data["count"] else None
        topic_rows.append({
            "topic": topic,
            "percent": avg_percent,
            "status": _score_status(avg_percent),
            "bar_class": _bar_class(avg_percent),
        })
    topic_rows.sort(key=lambda x: (x["percent"] is None, x["percent"] or 0))

    marks = [r.final_mark for r in present_results if is_dictation and r.final_mark is not None]
    if not marks:
        marks = [r.mark for r in present_results if r.mark is not None]
    percents = [r.percent for r in present_results if r.percent is not None]
    class_rows.sort(key=lambda x: (x["avg_percent"] is None, -(x["avg_percent"] or 0)))
    dictation_class_rows.sort(key=lambda x: (x.get("avg_final_mark") is None, -(x.get("avg_final_mark") or 0), x["assignment"].school_class.name if x.get("assignment") and x["assignment"].school_class else ''))
    dictation_child_rows.sort(key=lambda x: (x["class_name"], x["child_name"]))

    report = {
        "classes": len(class_rows),
        "results": len(all_results),
        "participants": len(present_results),
        "absent_count": len(absent_results),
        "participation_percent": round((len(present_results) / len(all_results)) * 100, 1) if all_results else None,
        "avg_mark": _safe_avg(marks),
        "avg_percent": _safe_avg(percents),
        "status": _score_status(_safe_avg(percents)),
        "bar_class": _bar_class(_safe_avg(percents)),
        "avg_dictation_mark": _safe_avg([r.dictation_mark for r in present_results if r.dictation_mark is not None]) if is_dictation else None,
        "avg_grammar_mark": _safe_avg([r.grammar_mark for r in present_results if r.grammar_mark is not None]) if is_dictation else None,
        "avg_final_mark": _safe_avg([r.final_mark for r in present_results if r.final_mark is not None]) if is_dictation else None,
        "avg_spelling_errors": _safe_avg([r.spelling_errors for r in present_results if r.spelling_errors is not None]) if is_dictation else None,
        "avg_punctuation_errors": _safe_avg([r.punctuation_errors for r in present_results if r.punctuation_errors is not None]) if is_dictation else None,
        "avg_grammar_errors": _safe_avg([r.grammar_errors for r in present_results if r.grammar_errors is not None]) if is_dictation else None,
        "avg_corrections_count": _safe_avg([r.corrections_count for r in present_results if r.corrections_count is not None]) if is_dictation else None,
        "dictation_quality": round((sum(1 for r in present_results if (r.final_mark if r.final_mark is not None else r.dictation_mark) in [4, 5]) / len(present_results)) * 100, 1) if is_dictation and present_results else None,
        "dictation_low": sum(1 for r in present_results if (r.final_mark if r.final_mark is not None else r.dictation_mark) == 2) if is_dictation else None,
    }
    problem_topics = [r for r in topic_rows if (r.get("percent") is not None and r.get("percent") < 65)]
    return {
        "work": work,
        "class_rows": class_rows,
        "report": report,
        "task_rows": task_rows,
        "topic_rows": topic_rows,
        "mark_counts": mark_counts,
        "problem_topics": problem_topics,
        "is_dictation": is_dictation,
        "dictation_class_rows": dictation_class_rows,
        "dictation_child_rows": dictation_child_rows,
    }




def _build_assignment_report(work, assignment):
    tasks = sorted((work.tasks or []), key=lambda x: x.task_number or 0)
    results = ControlWorkResult.query.filter_by(control_work_id=work.id, assignment_id=assignment.id, school_class_id=assignment.school_class_id).all()
    present_results = [r for r in results if not getattr(r, "is_absent", False)]
    absent_results = [r for r in results if getattr(r, "is_absent", False)]
    is_dictation = _is_dictation(work)

    marks_source = [r.final_mark for r in present_results if is_dictation and r.final_mark is not None]
    if not marks_source:
        marks_source = [r.mark for r in present_results if r.mark is not None]
    mark_counts = {m: sum(1 for x in marks_source if x == m) for m in [2, 3, 4, 5]}
    avg_mark = _safe_avg(marks_source)
    avg_percent = _safe_avg([r.percent for r in present_results if r.percent is not None])

    task_rows = []
    topic_stats = {}
    results_with_scores = [r for r in present_results if r.total_score is not None]
    max_total = sum(t.max_score or 0 for t in tasks)
    if not is_dictation:
        for task in tasks:
            avg_score = None
            percent = None
            if results_with_scores and max_total > 0 and (task.max_score or 0) > 0:
                est_scores = [((r.total_score or 0) * (task.max_score or 0) / max_total) for r in results_with_scores]
                avg_score = round(sum(est_scores) / len(est_scores), 2) if est_scores else None
                percent = round((avg_score / (task.max_score or 1)) * 100, 2) if avg_score is not None else None
            task_rows.append({
                "task": task,
                "avg_score": avg_score,
                "percent": percent,
                "status": _score_status(percent),
                "bar_class": _bar_class(percent),
            })
            topic_key = (task.topic or "—").strip()
            if topic_key not in topic_stats:
                topic_stats[topic_key] = {"sum": 0, "count": 0}
            if percent is not None:
                topic_stats[topic_key]["sum"] += percent
                topic_stats[topic_key]["count"] += 1

    topic_rows = []
    for topic, data in topic_stats.items():
        avg_topic_percent = round(data["sum"] / data["count"], 2) if data["count"] else None
        topic_rows.append({
            "topic": topic,
            "percent": avg_topic_percent,
            "status": _score_status(avg_topic_percent),
            "bar_class": _bar_class(avg_topic_percent),
        })
    topic_rows.sort(key=lambda x: (x["percent"] is None, x["percent"] or 0))

    child_rows = []
    for r in sorted(results, key=lambda x: ((x.child.last_name if x.child else ''), (x.child.first_name if x.child else ''), (x.child.middle_name if x.child else ''))):
        child = getattr(r, 'child', None)
        child_rows.append({
            "child_name": child.fio if child else f"ID {r.child_id}",
            "is_absent": bool(getattr(r, 'is_absent', False)),
            "percent": None if getattr(r, 'is_absent', False) else r.percent,
            "mark": None if getattr(r, 'is_absent', False) else (r.final_mark if is_dictation and r.final_mark is not None else r.mark),
            "dictation_mark": None if getattr(r, 'is_absent', False) else r.dictation_mark,
            "grammar_mark": None if getattr(r, 'is_absent', False) else r.grammar_mark,
            "status": "Отсутствовал" if getattr(r, 'is_absent', False) else _score_status(r.percent),
        })

    report = {
        "results": len(results),
        "participants": len(present_results),
        "absent_count": len(absent_results),
        "participation_percent": round((len(present_results) / len(results)) * 100, 1) if results else None,
        "avg_mark": avg_mark,
        "avg_percent": avg_percent,
        "status": _score_status(avg_percent),
        "bar_class": _bar_class(avg_percent),
        "teacher_name": assignment.teacher.fio if assignment.teacher else '—',
        "class_name": assignment.school_class.name if assignment.school_class else '—',
        "avg_dictation_mark": _safe_avg([r.dictation_mark for r in present_results if r.dictation_mark is not None]) if is_dictation else None,
        "avg_grammar_mark": _safe_avg([r.grammar_mark for r in present_results if r.grammar_mark is not None]) if is_dictation else None,
        "avg_final_mark": _safe_avg([r.final_mark for r in present_results if r.final_mark is not None]) if is_dictation else None,
    }
    problem_topics = [r for r in topic_rows if (r.get("percent") is not None and r.get("percent") < 65)]
    return {
        "work": work,
        "assignment": assignment,
        "report": report,
        "task_rows": task_rows,
        "topic_rows": topic_rows,
        "problem_topics": problem_topics,
        "mark_counts": mark_counts,
        "child_rows": child_rows,
        "is_dictation": is_dictation,
    }




def _pdf_value(value, suffix=""):
    if value is None or value == "":
        return "—"
    return f"{value}{suffix}"


def _safe_filename_part(value, fallback):
    raw = str(value or '').strip()
    if not raw:
        return fallback
    allowed = []
    for ch in raw:
        if ch.isalnum() or ch in ('_', '-', '.'): 
            allowed.append(ch)
        elif ch.isspace():
            allowed.append('_')
    cleaned = ''.join(allowed).strip('_.-')
    return cleaned or fallback




def _register_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    regular_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    bundled_fonts_dir = os.path.join(base_dir, 'fonts')
    font_candidates = [
        ('DejaVuSans', 'DejaVuSans-Bold', [
            (os.path.join(bundled_fonts_dir, 'DejaVuSans.ttf'), os.path.join(bundled_fonts_dir, 'DejaVuSans-Bold.ttf')),
            ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
            ('/usr/local/share/fonts/DejaVuSans.ttf', '/usr/local/share/fonts/DejaVuSans-Bold.ttf'),
            ('/Library/Fonts/Arial Unicode.ttf', '/Library/Fonts/Arial Bold.ttf'),
            ('/System/Library/Fonts/Supplemental/Arial Unicode.ttf', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
            ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf'),
        ]),
        ('NotoSans', 'NotoSans-Bold', [
            (os.path.join(bundled_fonts_dir, 'NotoSans-Regular.ttf'), os.path.join(bundled_fonts_dir, 'NotoSans-Bold.ttf')),
            ('/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf', '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf'),
            ('/usr/share/fonts/opentype/noto/NotoSans-Regular.ttf', '/usr/share/fonts/opentype/noto/NotoSans-Bold.ttf'),
        ]),
    ]

    for regular_name, bold_name, candidates in font_candidates:
        for regular_path, bold_path in candidates:
            try:
                if not os.path.exists(regular_path):
                    continue
                if regular_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(regular_name, regular_path))
                if os.path.exists(bold_path):
                    if bold_name not in pdfmetrics.getRegisteredFontNames():
                        pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                else:
                    bold_name = regular_name
                return regular_name, bold_name
            except Exception:
                continue
    return regular_font, bold_font

def _build_pdf_doc(title, orientation='portrait'):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception as exc:
        raise RuntimeError('Для PDF-экспорта нужен пакет reportlab') from exc

    regular_font, bold_font = _register_pdf_fonts()

    out = BytesIO()
    pagesize = landscape(A4) if orientation == 'landscape' else A4
    doc = SimpleDocTemplate(out, pagesize=pagesize, leftMargin=18, rightMargin=18, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    styles['Title'].fontName = bold_font
    styles['Title'].fontSize = 16
    styles['Title'].leading = 19
    styles['Heading1'].fontName = bold_font
    styles['Heading2'].fontName = bold_font
    styles['Heading3'].fontName = bold_font
    styles['Normal'].fontName = regular_font
    styles['BodyText'].fontName = regular_font
    styles.add(ParagraphStyle(name='Cell', parent=styles['BodyText'], fontName=regular_font, fontSize=8, leading=10, alignment=TA_LEFT, spaceAfter=0))
    styles.add(ParagraphStyle(name='CellBold', parent=styles['BodyText'], fontName=bold_font, fontSize=8, leading=10, alignment=TA_LEFT, spaceAfter=0))
    styles.add(ParagraphStyle(name='Small', parent=styles['BodyText'], fontName=regular_font, fontSize=8, leading=10, spaceAfter=0))
    styles.add(ParagraphStyle(name='Section', parent=styles['Heading2'], fontName=bold_font, fontSize=11, leading=13, spaceBefore=6, spaceAfter=6))
    ctx = {
        'out': out, 'doc': doc, 'styles': styles, 'colors': colors,
        'Paragraph': Paragraph, 'Spacer': Spacer, 'Table': Table, 'TableStyle': TableStyle, 'KeepTogether': KeepTogether,
        'regular_font': regular_font, 'bold_font': bold_font,
    }
    return ctx


def _p(text, styles, style='Cell'):
    from reportlab.platypus import Paragraph
    return Paragraph(escape(str(text or '—')).replace('\n', '<br/>'), styles[style])


def _make_pdf_table(ctx, headers, rows, col_widths=None, repeat_rows=1, compact=False):
    table_data = [[_p(h, ctx['styles'], 'CellBold') for h in headers]]
    for row in rows:
        table_data.append([_p(cell, ctx['styles'], 'Small' if compact else 'Cell') for cell in row])
    table = ctx['Table'](table_data, colWidths=col_widths, repeatRows=repeat_rows)
    table.setStyle(ctx['TableStyle']([
        ('BACKGROUND', (0, 0), (-1, 0), ctx['colors'].HexColor('#D9EAF7')),
        ('FONTNAME', (0, 0), (-1, 0), ctx['bold_font']),
        ('GRID', (0, 0), (-1, -1), 0.4, ctx['colors'].HexColor('#B8C7D6')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    return table


def _render_control_work_summary_pdf(work, payload):
    ctx = _build_pdf_doc('Сводный отчет по контрольной работе', orientation='landscape')
    styles = ctx['styles']
    items = []
    items.append(ctx['Paragraph']('Сводный отчет по контрольной работе', styles['Title']))
    org_header_lines = get_organization_header_lines()
    if org_header_lines:
        items.append(ctx['Paragraph']('<br/>'.join(escape(x) for x in org_header_lines), styles['Normal']))
        items.append(ctx['Spacer'](1, 8))
    header_lines = [
        f"{work.subject_name or '—'} · {work.work_kind_label or '—'}",
        f"Дата проведения: {_fmt_date(work.work_date)} · Срок сдачи: {_fmt_date(work.deadline_date)}",
    ]
    if work.theme:
        header_lines.append(f"Наименование работы: {work.theme}")
    if getattr(work, 'parallel', None):
        header_lines.append(f"Параллель: {work.parallel}")
    items.append(ctx['Paragraph']('<br/>'.join(escape(x) for x in header_lines), styles['Normal']))
    items.append(ctx['Spacer'](1, 8))

    report = payload['report']
    info_rows = [
        ['Количество учащихся по списку', _pdf_value(report.get('results'))],
        ['Писали', _pdf_value(report.get('participants'))],
        ['Отсутствовали', _pdf_value(report.get('absent_count'))],
        ['Процент участия', _pdf_value(report.get('participation_percent'), '%')],
        ['Средний результат', _pdf_value(report.get('avg_percent'), '%')],
        ['Средняя отметка', _pdf_value(report.get('avg_mark'))],
    ]
    items.append(ctx['Paragraph']('Общие показатели', styles['Section']))
    items.append(_make_pdf_table(ctx, ['Показатель', 'Значение'], info_rows, col_widths=[250, 120]))
    items.append(ctx['Spacer'](1, 8))

    if payload.get('is_dictation'):
        dict_rows = [
            ['Отметка за диктант', _pdf_value(report.get('avg_dictation_mark'))],
            ['Отметка за грамматическое задание', _pdf_value(report.get('avg_grammar_mark'))],
            ['Итоговая отметка', _pdf_value(report.get('avg_final_mark'))],
            ['Орфографические ошибки', _pdf_value(report.get('avg_spelling_errors'))],
            ['Пунктуационные ошибки', _pdf_value(report.get('avg_punctuation_errors'))],
            ['Грамматические ошибки', _pdf_value(report.get('avg_grammar_errors'))],
            ['Исправления', _pdf_value(report.get('avg_corrections_count'))],
        ]
        items.append(ctx['Paragraph']('Сводка по диктанту', styles['Section']))
        items.append(_make_pdf_table(ctx, ['Показатель', 'Значение'], dict_rows, col_widths=[250, 120]))
        items.append(ctx['Spacer'](1, 8))
    else:
        items.append(ctx['Paragraph']('Распределение отметок', styles['Section']))
        grade_rows = [[str(mark), payload['mark_counts'].get(mark, 0)] for mark in [2,3,4,5]]
        items.append(_make_pdf_table(ctx, ['Отметка', 'Количество'], grade_rows, col_widths=[120,120]))
        items.append(ctx['Spacer'](1, 8))

    if not payload.get('is_dictation'):
        task_rows = []
        for row in payload.get('task_rows', []):
            task_rows.append([
                row['task'].task_number, row['task'].topic or '—', row['task'].max_score,
                _pdf_value(row.get('avg_score')), _pdf_value(row.get('percent'), '%')
            ])
        items.append(ctx['Paragraph']('Анализ заданий', styles['Section']))
        if task_rows:
            items.append(_make_pdf_table(ctx, ['Задание', 'Тема', 'Макс. балл', 'Средний балл', '% выполнения', 'Статус'], task_rows, col_widths=[55, 180, 60, 75, 75, 80], compact=True))
        else:
            items.append(ctx['Paragraph']('Нет данных по заданиям.', styles['Normal']))
        items.append(ctx['Spacer'](1, 8))

    class_headers = ['Класс', 'Учитель', 'По списку', 'Писали', 'Отсутств.', 'Участие %']
    class_widths = [55, 105, 50, 45, 55, 55]
    if payload.get('is_dictation'):
        class_headers += ['Диктант', 'Грамм.', 'Итог', 'Ошибки']
        class_widths += [45, 45, 45, 50]
        class_rows = []
        for row in payload.get('dictation_class_rows', []):
            err_parts = [row.get('avg_spelling_errors'), row.get('avg_punctuation_errors'), row.get('avg_grammar_errors')]
            err_txt = '/'.join(str(x) for x in err_parts if x is not None) or '—'
            class_rows.append([
                row['assignment'].school_class.name if row['assignment'].school_class else '—',
                row['assignment'].teacher.fio if row['assignment'].teacher else '—',
                row['results_count'], row['participants_count'], row['absent_count'],
                round((row['participants_count']/row['results_count'])*100,1) if row['results_count'] else '—',
                _pdf_value(row.get('avg_dictation_mark')), _pdf_value(row.get('avg_grammar_mark')), _pdf_value(row.get('avg_final_mark')), err_txt
            ])
    else:
        class_headers += ['Ср.%', 'Ср. отметка', '2', '3', '4', '5', 'Статус']
        class_widths += [50, 58, 25, 25, 25, 25, 55]
        class_rows = []
        for row in payload.get('class_rows', []):
            class_rows.append([
                row['assignment'].school_class.name if row['assignment'].school_class else '—',
                row['assignment'].teacher.fio if row['assignment'].teacher else '—',
                row['results_count'], row['participants_count'], row['absent_count'],
                round((row['participants_count']/row['results_count'])*100,1) if row['results_count'] else '—',
                _pdf_value(row.get('avg_percent')), _pdf_value(row.get('avg_mark')),
                row['marks'][2], row['marks'][3], row['marks'][4], row['marks'][5]
            ])
    items.append(ctx['Paragraph']('Аналитика по классам', styles['Section']))
    if class_rows:
        items.append(_make_pdf_table(ctx, class_headers, class_rows, col_widths=class_widths, compact=True))
    else:
        items.append(ctx['Paragraph']('Нет данных по классам.', styles['Normal']))
    items.append(ctx['Spacer'](1, 8))

    topic_rows = payload.get('problem_topics') or payload.get('topic_rows') or []
    if topic_rows and not payload.get('is_dictation'):
        items.append(ctx['Paragraph']('Темы по уровню результата', styles['Section']))
        items.append(_make_pdf_table(ctx, ['Тема', 'Средний результат'], [[r['topic'], _pdf_value(r.get('percent'), '%'), r.get('status') or '—'] for r in topic_rows], col_widths=[300,120,120]))

    signature_block = get_organization_signature_block()
    if signature_block:
        items.append(ctx['Spacer'](1, 14))
        items.append(ctx['Paragraph'](escape(signature_block), styles['Normal']))

    ctx['doc'].build(items)
    ctx['out'].seek(0)
    return ctx['out']


def _render_control_work_class_pdf(work, payload):
    ctx = _build_pdf_doc('Отчет по классу')
    styles = ctx['styles']
    items = []
    report = payload['report']
    items.append(ctx['Paragraph']('Отчет по классу', styles['Title']))
    org_header_lines = get_organization_header_lines()
    if org_header_lines:
        items.append(ctx['Paragraph']('<br/>'.join(escape(x) for x in org_header_lines), styles['Normal']))
        items.append(ctx['Spacer'](1, 8))
    header_lines = [
        f"{work.subject_name or '—'} · {work.work_kind_label or '—'}",
        f"Класс: {report.get('class_name') or '—'} · Учитель: {report.get('teacher_name') or '—'}",
        f"Дата проведения: {_fmt_date(work.work_date)} · Срок сдачи: {_fmt_date(work.deadline_date)}",
    ]
    if work.theme:
        header_lines.append(f"Наименование работы: {work.theme}")
    items.append(ctx['Paragraph']('<br/>'.join(escape(x) for x in header_lines), styles['Normal']))
    items.append(ctx['Spacer'](1, 8))

    info_rows = [
        ['Количество учащихся по списку', _pdf_value(report.get('results'))],
        ['Писали', _pdf_value(report.get('participants'))],
        ['Отсутствовали', _pdf_value(report.get('absent_count'))],
        ['Процент участия', _pdf_value(report.get('participation_percent'), '%')],
        ['Средний результат', _pdf_value(report.get('avg_percent'), '%')],
        ['Средняя отметка', _pdf_value(report.get('avg_mark'))],
    ]
    items.append(ctx['Paragraph']('Общие показатели', styles['Section']))
    items.append(_make_pdf_table(ctx, ['Показатель', 'Значение'], info_rows, col_widths=[260,120]))
    items.append(ctx['Spacer'](1, 8))

    if payload.get('is_dictation'):
        dict_rows = [
            ['Отметка за диктант', _pdf_value(report.get('avg_dictation_mark'))],
            ['Отметка за грамматическое задание', _pdf_value(report.get('avg_grammar_mark'))],
            ['Итоговая отметка', _pdf_value(report.get('avg_final_mark'))],
        ]
        items.append(ctx['Paragraph']('Сводка по диктанту', styles['Section']))
        items.append(_make_pdf_table(ctx, ['Показатель', 'Значение'], dict_rows, col_widths=[260,120]))
    else:
        items.append(ctx['Paragraph']('Распределение отметок', styles['Section']))
        grade_rows = [[str(mark), payload['mark_counts'].get(mark, 0)] for mark in [2,3,4,5]]
        items.append(_make_pdf_table(ctx, ['Отметка', 'Количество'], grade_rows, col_widths=[120,120]))
    items.append(ctx['Spacer'](1, 8))

    if not payload.get('is_dictation'):
        items.append(ctx['Paragraph']('Анализ заданий', styles['Section']))
        task_rows = [[r['task'].task_number, r['task'].topic or '—', r['task'].max_score, _pdf_value(r.get('avg_score')), _pdf_value(r.get('percent'), '%')] for r in payload.get('task_rows', [])]
        if task_rows:
            items.append(_make_pdf_table(ctx, ['Задание', 'Тема', 'Макс. балл', 'Средний балл', '% выполнения', 'Статус'], task_rows, col_widths=[55, 180, 60, 75, 75, 90], compact=True))
        else:
            items.append(ctx['Paragraph']('Нет данных по заданиям.', styles['Normal']))
        items.append(ctx['Spacer'](1, 8))

        topic_rows = payload.get('problem_topics') or payload.get('topic_rows') or []
        items.append(ctx['Paragraph']('Проблемные темы', styles['Section']))
        if topic_rows:
            items.append(_make_pdf_table(ctx, ['Тема', 'Средний результат'], [[r['topic'], _pdf_value(r.get('percent'), '%'), r.get('status') or '—'] for r in topic_rows], col_widths=[280,120,120]))
        else:
            items.append(ctx['Paragraph']('Проблемные темы пока не выявлены.', styles['Normal']))
        items.append(ctx['Spacer'](1, 8))

    items.append(ctx['Paragraph']('Список учащихся и результаты', styles['Section']))
    if payload.get('is_dictation'):
        child_rows = [[r['child_name'], r.get('status') or ('Отсутствовал' if r.get('is_absent') else 'Писал'), _pdf_value(r.get('dictation_mark')), _pdf_value(r.get('grammar_mark')), _pdf_value(r.get('mark'))] for r in payload.get('child_rows', [])]
        headers = ['Ученик', 'Статус', 'Диктант', 'Грамм.', 'Итог']
        widths = [220, 110, 60, 60, 60]
    else:
        child_rows = [[r['child_name'], 'Отсутствовал' if r.get('is_absent') else 'Присутствовал', _pdf_value(r.get('percent'), '%'), _pdf_value(r.get('mark')), r.get('status') or '—'] for r in payload.get('child_rows', [])]
        headers = ['Ученик', 'Статус', 'Процент', 'Отметка', 'Комментарий']
        widths = [220, 110, 70, 60, 90]
    if child_rows:
        items.append(_make_pdf_table(ctx, headers, child_rows, col_widths=widths, compact=True))
    else:
        items.append(ctx['Paragraph']('Результаты по классу пока не заполнены.', styles['Normal']))

    ctx['doc'].build(items)
    ctx['out'].seek(0)
    return ctx['out']

def _autosize_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(c.value or '')) for c in col_cells) if col_cells else 10
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)


def _make_work_report_excel(work, payload):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сводный отчет'
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    bold = Font(bold=True)

    info = [
        ('Название', f"{work.subject_name} — {work.work_kind_label}" + (f" · {work.theme}" if work.theme else '')),
        ('Вид работы', work.work_kind_label),
        ('Предмет', work.subject_name),
        ('Дата проведения', _fmt_date(work.work_date)),
        ('Срок сдачи', _fmt_date(work.deadline_date)),
        ('Классов', payload['report']['classes']),
        ('По списку', payload['report']['results']),
        ('Писали', payload['report']['participants']),
        ('Отсутствовали', payload['report']['absent_count']),
        ('Процент участия', payload['report']['participation_percent']),
        ('Средний результат', payload['report']['avg_percent']),
        ('Средняя отметка', payload['report']['avg_mark']),
    ]
    ws.append(['Общая информация', 'Значение'])
    for c in ws[1]:
        c.font = bold; c.fill = header_fill
    for row in info:
        ws.append(list(row))

    ws2 = wb.create_sheet('Классы')
    ws2.append(['Класс', 'Учитель', 'По списку', 'Писали', 'Отсутствовали', 'Процент участия', 'Средний результат', 'Средняя отметка', '2', '3', '4', '5', 'Статус'])
    for c in ws2[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['class_rows']:
        ws2.append([
            row['assignment'].school_class.name if row['assignment'].school_class else '—',
            row['assignment'].teacher.fio if row['assignment'].teacher else '—',
            row['results_count'], row['participants_count'], row['absent_count'],
            round((row['participants_count'] / row['results_count']) * 100, 1) if row['results_count'] else None,
            row['avg_percent'], row['avg_mark'], row['marks'][2], row['marks'][3], row['marks'][4], row['marks'][5]
        ])

    ws3 = wb.create_sheet('Задания')
    ws3.append(['Задание', 'Описание', 'Тема', 'Макс. балл', 'Средний балл', '% выполнения', 'Статус'])
    for c in ws3[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['task_rows']:
        ws3.append([row['task'].task_number, row['task'].description or '—', row['task'].topic or '—', row['task'].max_score, row['avg_score'], row['percent']])

    ws4 = wb.create_sheet('Темы')
    ws4.append(['Тема', 'Средний результат'])
    for c in ws4[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['topic_rows']:
        ws4.append([row['topic'], row['percent']])

    for sheet in wb.worksheets:
        _autosize_columns(sheet)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _make_assignment_report_excel(work, payload):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчет по классу'
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    bold = Font(bold=True)
    report = payload['report']
    ws.append(['Общая информация', 'Значение'])
    for c in ws[1]:
        c.font = bold; c.fill = header_fill
    info = [
        ('Название', f"{work.subject_name} — {work.work_kind_label}" + (f" · {work.theme}" if work.theme else '')),
        ('Класс', report['class_name']),
        ('Учитель', report['teacher_name']),
        ('Дата проведения', _fmt_date(work.work_date)),
        ('Срок сдачи', _fmt_date(work.deadline_date)),
        ('По списку', report['results']),
        ('Писали', report['participants']),
        ('Отсутствовали', report['absent_count']),
        ('Процент участия', report['participation_percent']),
        ('Средний результат', report['avg_percent']),
        ('Средняя отметка', report['avg_mark']),
    ]
    for row in info:
        ws.append(list(row))

    ws2 = wb.create_sheet('Задания')
    ws2.append(['Задание', 'Описание', 'Тема', 'Макс. балл', 'Средний балл', '% выполнения', 'Статус'])
    for c in ws2[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['task_rows']:
        ws2.append([row['task'].task_number, row['task'].description or '—', row['task'].topic or '—', row['task'].max_score, row['avg_score'], row['percent']])

    ws3 = wb.create_sheet('Ученики')
    ws3.append(['Ученик', 'Статус', 'Процент', 'Отметка'])
    for c in ws3[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['child_rows']:
        ws3.append([row['child_name'], 'Отсутствовал' if row['is_absent'] else 'Присутствовал', row['percent'], row['mark']])

    ws4 = wb.create_sheet('Темы')
    ws4.append(['Тема', 'Средний результат'])
    for c in ws4[1]:
        c.font = bold; c.fill = header_fill
    for row in payload['topic_rows']:
        ws4.append([row['topic'], row['percent']])

    for sheet in wb.worksheets:
        _autosize_columns(sheet)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _build_archive_dataset(filters):
    works = _archive_works_query(filters).all()
    teacher_groups = defaultdict(lambda: {"teacher": None, "works": [], "avg_percent": None, "results": 0})
    subject_groups = defaultdict(lambda: {"subject": None, "works": [], "avg_percent": None, "results": 0})
    year_groups = defaultdict(lambda: {"year": None, "works": [], "avg_percent": None, "results": 0, "avg_mark": None})
    work_rows = []

    for work in works:
        report_pack = _build_control_work_report(work, teacher_id=filters["selected_teacher_id"] if not has_permission("control_works_edit") else filters["selected_teacher_id"])
        report = report_pack["report"]
        teacher_names = sorted({a.teacher.fio for a in (work.assignments or []) if a.teacher and (not filters["selected_teacher_id"] or a.teacher_id == filters["selected_teacher_id"])})
        subject_name = work.subject_name
        year_name = work.academic_year.name if work.academic_year else "Без года"
        row = {
            "id": work.id,
            "work_date": work.work_date,
            "work_date_text": _fmt_date(work.work_date),
            "year_name": year_name,
            "subject_name": subject_name,
            "work_kind_label": work.work_kind_label,
            "theme": work.theme,
            "teachers_text": ", ".join(teacher_names) if teacher_names else "—",
            "avg_percent": report["avg_percent"],
            "avg_mark": report["avg_mark"],
            "results": report["results"],
            "classes": report["classes"],
            "status": report["status"],
            "bar_class": report["bar_class"],
        }
        work_rows.append(row)

        for a in (work.assignments or []):
            if not a.teacher:
                continue
            if filters["selected_teacher_id"] and a.teacher_id != filters["selected_teacher_id"]:
                continue
            g = teacher_groups[a.teacher_id]
            g["teacher"] = a.teacher
            g["works"].append(row)

        sg = subject_groups[work.subject_id]
        sg["subject"] = work.subject_ref
        sg["works"].append(row)

        ykey = work.academic_year_id or 0
        yg = year_groups[ykey]
        yg["year"] = work.academic_year
        yg["works"].append(row)

    for group in teacher_groups.values():
        group["avg_percent"] = _safe_avg([w["avg_percent"] for w in group["works"]])
        group["results"] = sum(w["results"] or 0 for w in group["works"])
        group["bar_class"] = _bar_class(group["avg_percent"])
        group["status"] = _score_status(group["avg_percent"])

    for group in subject_groups.values():
        group["avg_percent"] = _safe_avg([w["avg_percent"] for w in group["works"]])
        group["results"] = sum(w["results"] or 0 for w in group["works"])
        group["bar_class"] = _bar_class(group["avg_percent"])
        group["status"] = _score_status(group["avg_percent"])

    for group in year_groups.values():
        group["avg_percent"] = _safe_avg([w["avg_percent"] for w in group["works"]])
        group["avg_mark"] = _safe_avg([w["avg_mark"] for w in group["works"]])
        group["results"] = sum(w["results"] or 0 for w in group["works"])
        group["bar_class"] = _bar_class(group["avg_percent"])
        group["status"] = _score_status(group["avg_percent"])

    summary = {
        "works_count": len(work_rows),
        "results_count": sum(row["results"] or 0 for row in work_rows),
        "avg_percent": _safe_avg([row["avg_percent"] for row in work_rows]),
        "avg_mark": _safe_avg([row["avg_mark"] for row in work_rows]),
        "bar_class": _bar_class(_safe_avg([row["avg_percent"] for row in work_rows])),
        "status": _score_status(_safe_avg([row["avg_percent"] for row in work_rows])),
    }

    return {
        "works": work_rows,
        "teacher_groups": sorted(teacher_groups.values(), key=lambda x: ((x["teacher"].fio if x["teacher"] else ""))),
        "subject_groups": sorted(subject_groups.values(), key=lambda x: ((x["subject"].name if x["subject"] else ""))),
        "year_groups": sorted(year_groups.values(), key=lambda x: ((x["year"].start_date if x["year"] and x["year"].start_date else datetime.min.date())), reverse=True),
        "summary": summary,
    }


def _make_archive_excel(filters, dataset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Архив контрольных"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    ws.append(["Учебный год", "Дата", "Предмет", "Вид работы", "Тема", "Учителя", "Классов", "Результатов", "Средний %", "Средняя отметка", "Статус"])
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
    for row in dataset["works"]:
        ws.append([row["year_name"], row["work_date_text"], row["subject_name"], row["work_kind_label"], row["theme"], row["teachers_text"], row["classes"], row["results"], row["avg_percent"], row["avg_mark"], row["status"]])
    for col in ["A","B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 28

    ws2 = wb.create_sheet("По учителям")
    ws2.append(["Учитель", "Контрольных", "Результатов", "Средний %", "Статус"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill
    for g in dataset["teacher_groups"]:
        ws2.append([g["teacher"].fio if g["teacher"] else "—", len(g["works"]), g["results"], g["avg_percent"], g["status"]])

    ws3 = wb.create_sheet("По предметам")
    ws3.append(["Предмет", "Контрольных", "Результатов", "Средний %", "Статус"])
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
    for g in dataset["subject_groups"]:
        ws3.append([g["subject"].name if g["subject"] else "—", len(g["works"]), g["results"], g["avg_percent"], g["status"]])

    ws4 = wb.create_sheet("По годам")
    ws4.append(["Учебный год", "Контрольных", "Результатов", "Средний %", "Средняя отметка", "Статус"])
    for cell in ws4[1]:
        cell.font = bold
        cell.fill = header_fill
    for g in dataset["year_groups"]:
        ws4.append([g["year"].name if g["year"] else "Без года", len(g["works"]), g["results"], g["avg_percent"], g["avg_mark"], g["status"]])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _make_archive_pdf(dataset, filters):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception as exc:
        raise RuntimeError("Для PDF-экспорта нужен пакет reportlab") from exc

    regular_font, bold_font = _register_pdf_fonts()

    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles['Title'].fontName = bold_font
    styles['Normal'].fontName = regular_font
    items = []
    items.append(Paragraph("Архив контрольных работ", styles["Title"]))
    items.append(Paragraph(f"Контрольных: {dataset['summary']['works_count']} · Результатов: {dataset['summary']['results_count']} · Средний %: {dataset['summary']['avg_percent'] if dataset['summary']['avg_percent'] is not None else '—'}", styles["Normal"]))
    items.append(Spacer(1, 12))

    table_data = [["Год", "Дата", "Предмет", "Вид", "Тема", "Учителя", "Результ.", "Ср.%", "Ср.отметка"]]
    for row in dataset["works"][:80]:
        table_data.append([row["year_name"], row["work_date_text"], row["subject_name"], row["work_kind_label"], row["theme"], row["teachers_text"], row["results"], row["avg_percent"] if row["avg_percent"] is not None else "—", row["avg_mark"] if row["avg_mark"] is not None else "—"])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    items.append(table)
    items.append(Spacer(1, 12))

    teacher_table = [["Учитель", "Контрольных", "Результатов", "Ср.%", "Статус"]]
    for g in dataset["teacher_groups"][:40]:
        teacher_table.append([g["teacher"].fio if g["teacher"] else "—", len(g["works"]), g["results"], g["avg_percent"] if g["avg_percent"] is not None else "—", g["status"]])
    t2 = Table(teacher_table, repeatRows=1)
    t2.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EEF5FB")), ("GRID", (0,0), (-1,-1), 0.5, colors.grey)]))
    items.append(Paragraph("Сводка по учителям", styles["Heading2"]))
    items.append(t2)

    doc.build(items)
    out.seek(0)
    return out


def _mark_from_percent(percent, work=None):
    if percent is None:
        return None
    g5 = getattr(work, "grade5_percent", 85) or 85
    g4 = getattr(work, "grade4_percent", 65) or 65
    g3 = getattr(work, "grade3_percent", 45) or 45
    if percent >= g5:
        return 5
    if percent >= g4:
        return 4
    if percent >= g3:
        return 3
    return 2


def _can_view_assignment(assignment):
    if has_permission("control_works_edit"):
        return True
    return assignment.teacher_id == current_user.id


def _assigned_works_query():
    if has_permission("control_works_edit"):
        return ControlWork.query.order_by(ControlWork.work_date.desc().nullslast(), ControlWork.created_at.desc())
    return (
        ControlWork.query
        .join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id)
        .filter(ControlWorkAssignment.teacher_id == current_user.id)
        .distinct()
        .order_by(ControlWork.work_date.desc().nullslast(), ControlWork.created_at.desc())
    )


def _form_context(selected_subject_id=None):
    year = _current_year()
    classes = (
        SchoolClass.query.filter_by(academic_year_id=year.id)
        .order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc())
        .all() if year else []
    )
    teachers = User.query.order_by(User.last_name.asc(), User.first_name.asc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    parallels = sorted({c.grade for c in classes if c.grade is not None})
    department_ids = _department_ids_for_user()
    teacher_options_map = _teacher_options_map(classes, subject_id=selected_subject_id, academic_year_id=(year.id if year else None), department_ids=department_ids)
    auto_teacher_by_class = _auto_teacher_by_class(classes, teacher_options_map)
    return year, classes, teachers, subjects, parallels, teacher_options_map, auto_teacher_by_class


def _task_count_from_request_or_work(work=None):
    selected_kind = _selected_work_kind(work)
    if selected_kind == DICTATION_KIND:
        return 0
    if request.method == "POST":
        return max(1, min(int(request.form.get("task_count") or 1), 20))
    if work and work.tasks:
        return len(work.tasks)
    return 5


def _save_work_from_form(work=None):
    subject_id = request.form.get("subject_id", type=int)
    theme = (request.form.get("theme") or "").strip()
    work_kind = _selected_work_kind(work)
    work_date = _parse_date(request.form.get("work_date"))
    deadline_date = _parse_date(request.form.get("deadline_date"))
    task_count = _task_count_from_request_or_work(work)
    grade5_percent = request.form.get("grade5_percent", type=int) or 85
    grade4_percent = request.form.get("grade4_percent", type=int) or 65
    grade3_percent = request.form.get("grade3_percent", type=int) or 45
    current_year = _current_year()

    if not subject_id or not theme:
        raise ValueError("Выберите предмет из реестра и укажите тему контрольной работы.")

    selected_class_ids = request.form.getlist("class_ids")
    if not selected_class_ids:
        raise ValueError("Нужно выбрать хотя бы один класс.")

    previous = None
    is_new = work is None
    if work is None:
        work = ControlWork(created_by=current_user.id, updated_by=current_user.id)
        db.session.add(work)
    else:
        previous = {
            "theme": work.theme or "",
            "work_kind": work.work_kind_label,
            "work_date": _fmt_date(work.work_date),
            "deadline_date": _fmt_date(work.deadline_date),
            "subject": str(work.subject_name or ""),
            "manual_status": getattr(work, "manual_status", None) or "",
        }

    work.subject_id = subject_id
    work.work_kind = work_kind
    work.academic_year_id = current_year.id if current_year else getattr(work, "academic_year_id", None)
    if current_year and current_year.end_date:
        try:
            work.retention_until = current_year.end_date.replace(year=current_year.end_date.year + 7)
        except Exception:
            pass
    work.theme = theme
    work.work_date = work_date
    work.deadline_date = deadline_date
    work.grade5_percent = grade5_percent
    work.grade4_percent = grade4_percent
    work.grade3_percent = grade3_percent
    work.updated_by = current_user.id
    work.updated_at = datetime.utcnow()
    db.session.flush()

    ControlWorkTask.query.filter_by(control_work_id=work.id).delete()
    for i in range(1, task_count + 1):
        max_score = int(request.form.get(f"max_score_{i}") or 0)
        description = (request.form.get(f"description_{i}") or "").strip() or None
        topic = (request.form.get(f"topic_{i}") or "").strip() or None
        db.session.add(ControlWorkTask(
            control_work_id=work.id,
            task_number=i,
            max_score=max_score,
            description=description,
            topic=topic,
        ))

    ControlWorkAssignment.query.filter_by(control_work_id=work.id).delete()
    selected_classes = {c.id: c for c in SchoolClass.query.filter(SchoolClass.id.in_([int(x) for x in selected_class_ids])).all()}
    department_ids = _department_ids_for_user()
    for class_id in selected_class_ids:
        class_id_int = int(class_id)
        teacher_id = request.form.get(f"teacher_for_{class_id}", type=int)
        if not teacher_id:
            school_class = selected_classes.get(class_id_int)
            candidates = _teacher_load_candidates(
                subject_id=subject_id,
                grade=(school_class.grade if school_class else None),
                class_name=(school_class.name if school_class else None),
                academic_year_id=(current_year.id if current_year else None),
                department_ids=department_ids,
            )
            preferred = next((item for item in candidates if item.get("preferred")), None)
            if preferred:
                teacher_id = preferred["id"]
            elif len(candidates) == 1:
                teacher_id = candidates[0]["id"]
        db.session.add(ControlWorkAssignment(
            control_work_id=work.id,
            school_class_id=class_id_int,
            teacher_id=teacher_id,
            status="ASSIGNED",
        ))

    if is_new:
        _log_control_work_event(work, "created", "Создание контрольной работы", new_value=f"{work.subject_name} — {work.theme}")
    else:
        _collect_work_change_logs(work, previous)
        _log_control_work_event(work, "assignments", "Обновлен состав классов и назначений", details=f"Классов: {len(selected_class_ids)}")
    return work


@control_bp.route("/api/teachers-by-load")
@login_required
def teachers_by_load_api():
    if not has_permission("control_works_view"):
        abort(403)
    subject_id = request.args.get("subject_id", type=int)
    year = _current_year()
    department_ids = _department_ids_for_user()
    payload = {}
    class_ids = [int(x) for x in request.args.getlist("class_id") if str(x).isdigit()]
    if class_ids:
        classes = SchoolClass.query.filter(SchoolClass.id.in_(class_ids)).all()
        for school_class in classes:
            payload[str(school_class.id)] = _teacher_load_candidates(
                subject_id=subject_id,
                grade=school_class.grade,
                class_name=school_class.name,
                academic_year_id=(year.id if year else None),
                department_ids=department_ids,
            )
    else:
        parallel = request.args.get("parallel", type=int)
        query = SchoolClass.query
        if year:
            query = query.filter_by(academic_year_id=year.id)
        if parallel:
            query = query.filter_by(grade=parallel)
        for school_class in query.all():
            payload[str(school_class.id)] = _teacher_load_candidates(
                subject_id=subject_id,
                grade=school_class.grade,
                class_name=school_class.name,
                academic_year_id=(year.id if year else None),
                department_ids=department_ids,
            )
    return jsonify(payload)


def _control_summary_dataset(selected_year_id=None, selected_subject_id=None, selected_teacher_id=None, selected_grade=None, selected_class_id=None):
    current_year = _current_year()
    selected_year_id = selected_year_id or (current_year.id if current_year else None)
    department_ids = _department_ids_for_user()

    work_query = ControlWork.query
    result_query = (
        db.session.query(ControlWorkResult, ControlWork, ControlWorkAssignment, SchoolClass, Child, User)
        .join(ControlWork, ControlWork.id == ControlWorkResult.control_work_id)
        .join(ControlWorkAssignment, db.and_(
            ControlWorkAssignment.control_work_id == ControlWorkResult.control_work_id,
            ControlWorkAssignment.school_class_id == ControlWorkResult.school_class_id,
        ))
        .join(SchoolClass, SchoolClass.id == ControlWorkResult.school_class_id)
        .join(Child, Child.id == ControlWorkResult.child_id)
        .outerjoin(User, User.id == ControlWorkAssignment.teacher_id)
    )

    if selected_year_id:
        work_query = work_query.filter(ControlWork.academic_year_id == selected_year_id)
        result_query = result_query.filter(ControlWork.academic_year_id == selected_year_id)
    if selected_subject_id:
        work_query = work_query.filter(ControlWork.subject_id == selected_subject_id)
        result_query = result_query.filter(ControlWork.subject_id == selected_subject_id)
    if selected_grade:
        result_query = result_query.filter(SchoolClass.grade == selected_grade)
    if selected_class_id:
        result_query = result_query.filter(SchoolClass.id == selected_class_id)
    if selected_teacher_id:
        result_query = result_query.filter(ControlWorkAssignment.teacher_id == selected_teacher_id)
        work_query = work_query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).filter(ControlWorkAssignment.teacher_id == selected_teacher_id)
    elif has_permission("control_works_view") and not has_permission("control_works_edit") and getattr(current_user, "role", None) != "METHODIST":
        result_query = result_query.filter(ControlWorkAssignment.teacher_id == current_user.id)
        work_query = work_query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).filter(ControlWorkAssignment.teacher_id == current_user.id)
    elif department_ids:
        result_query = result_query.join(TeacherLoad, db.and_(
            TeacherLoad.teacher_id == ControlWorkAssignment.teacher_id,
            TeacherLoad.subject_id == ControlWork.subject_id,
            TeacherLoad.is_archived.is_(False),
        )).filter(TeacherLoad.department_id.in_(department_ids))
        work_query = work_query.join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id).join(TeacherLoad, db.and_(
            TeacherLoad.teacher_id == ControlWorkAssignment.teacher_id,
            TeacherLoad.subject_id == ControlWork.subject_id,
            TeacherLoad.is_archived.is_(False),
        )).filter(TeacherLoad.department_id.in_(department_ids))

    work_count = work_query.order_by(None).distinct().count()
    rows_raw = result_query.all()
    seen_pairs = set()
    rows = []
    for item in rows_raw:
        result, work, assignment, school_class, child, teacher = item
        key = (result.id, assignment.id)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        rows.append(item)

    percents = []
    mark_counts = {2: 0, 3: 0, 4: 0, 5: 0}
    teacher_stats = defaultdict(lambda: {"teacher_id": None, "teacher_name": "—", "subject_names": set(), "results": 0, "participants": 0, "absent": 0, "percents": [], "marks": []})
    class_stats = defaultdict(lambda: {"class_id": None, "class_name": "—", "results": 0, "participants": 0, "absent": 0, "percents": [], "marks": [], "subject_names": set()})
    child_stats = defaultdict(lambda: {"child_name": "—", "class_name": "—", "results": 0, "participants": 0, "absent": 0, "percents": [], "marks": [], "subject_names": set()})
    dictation_teacher_stats = defaultdict(lambda: {"teacher_id": None, "teacher_name": "—", "participants": 0, "absent": 0, "dictation_marks": [], "grammar_marks": [], "final_marks": [], "spelling": [], "punctuation": [], "grammar_errors": []})
    dictation_class_stats = defaultdict(lambda: {"class_id": None, "class_name": "—", "participants": 0, "absent": 0, "dictation_marks": [], "grammar_marks": [], "final_marks": [], "spelling": [], "punctuation": [], "grammar_errors": []})
    dictation_child_stats = defaultdict(lambda: {"child_name": "—", "class_name": "—", "participants": 0, "absent": 0, "dictation_marks": [], "grammar_marks": [], "final_marks": [], "spelling": [], "punctuation": [], "grammar_errors": []})
    work_stats = defaultdict(lambda: {"work_id": None, "subject_name": "—", "work_kind_label": "—", "theme": "—", "work_date": None, "class_name": "—", "class_id": None, "teacher_name": "—", "results": 0, "participants": 0, "absent": 0, "percents": [], "marks": [], "is_dictation": False, "dictation_marks": [], "grammar_marks": [], "final_marks": []})

    for result, work, assignment, school_class, child, teacher in rows:
        is_absent = bool(getattr(result, "is_absent", False))
        is_dictation = _is_dictation(work)
        percent = float(result.percent) if result.percent is not None and not is_absent else None
        effective_mark = None
        if not is_absent:
            effective_mark = result.final_mark if is_dictation and result.final_mark is not None else result.mark
            if effective_mark is None and is_dictation:
                effective_mark = result.dictation_mark
        if percent is not None:
            percents.append(percent)
        if effective_mark in mark_counts:
            mark_counts[effective_mark] += 1

        teacher_key = assignment.teacher_id or 0
        teacher_stats[teacher_key]["teacher_id"] = assignment.teacher_id
        teacher_stats[teacher_key]["teacher_name"] = teacher.fio if teacher else "Не назначен"
        teacher_stats[teacher_key]["subject_names"].add(work.subject_name)
        teacher_stats[teacher_key]["results"] += 1
        if is_absent:
            teacher_stats[teacher_key]["absent"] += 1
        else:
            teacher_stats[teacher_key]["participants"] += 1
        if percent is not None:
            teacher_stats[teacher_key]["percents"].append(percent)
        if effective_mark is not None:
            teacher_stats[teacher_key]["marks"].append(effective_mark)

        class_stats[school_class.id]["class_id"] = school_class.id
        class_stats[school_class.id]["class_name"] = school_class.name
        class_stats[school_class.id]["subject_names"].add(work.subject_name)
        class_stats[school_class.id]["results"] += 1
        if is_absent:
            class_stats[school_class.id]["absent"] += 1
        else:
            class_stats[school_class.id]["participants"] += 1
        if percent is not None:
            class_stats[school_class.id]["percents"].append(percent)
        if effective_mark is not None:
            class_stats[school_class.id]["marks"].append(effective_mark)

        child_stats[child.id]["child_name"] = child.fio
        child_stats[child.id]["class_name"] = school_class.name
        child_stats[child.id]["subject_names"].add(work.subject_name)
        child_stats[child.id]["results"] += 1
        if is_absent:
            child_stats[child.id]["absent"] += 1
        else:
            child_stats[child.id]["participants"] += 1
        if percent is not None:
            child_stats[child.id]["percents"].append(percent)
        if effective_mark is not None:
            child_stats[child.id]["marks"].append(effective_mark)

        work_stats[work.id]["work_id"] = work.id
        work_stats[work.id]["subject_name"] = work.subject_name or "—"
        work_stats[work.id]["work_kind_label"] = work.work_kind_label
        work_stats[work.id]["theme"] = work.theme or "—"
        work_stats[work.id]["work_date"] = work.work_date
        work_stats[work.id]["class_name"] = school_class.name if school_class else "—"
        work_stats[work.id]["class_id"] = school_class.id if school_class else None
        work_stats[work.id]["teacher_name"] = teacher.fio if teacher else "Не назначен"
        work_stats[work.id]["results"] += 1
        work_stats[work.id]["is_dictation"] = is_dictation
        if is_absent:
            work_stats[work.id]["absent"] += 1
        else:
            work_stats[work.id]["participants"] += 1
        if percent is not None:
            work_stats[work.id]["percents"].append(percent)
        if effective_mark is not None:
            work_stats[work.id]["marks"].append(effective_mark)

        if is_dictation:
            dictation_teacher_stats[teacher_key]["teacher_name"] = teacher.fio if teacher else "Не назначен"
            dictation_class_stats[school_class.id]["class_id"] = school_class.id
            dictation_class_stats[school_class.id]["class_name"] = school_class.name
            dictation_child_stats[child.id]["child_name"] = child.fio
            dictation_child_stats[child.id]["class_name"] = school_class.name
            target_sets = [dictation_teacher_stats[teacher_key], dictation_class_stats[school_class.id], dictation_child_stats[child.id], work_stats[work.id]]
            for bucket in target_sets:
                if bucket is not work_stats[work.id]:
                    if is_absent:
                        bucket["absent"] += 1
                    else:
                        bucket["participants"] += 1
                if not is_absent:
                    if result.dictation_mark is not None:
                        bucket["dictation_marks"].append(result.dictation_mark)
                    if result.grammar_mark is not None:
                        bucket["grammar_marks"].append(result.grammar_mark)
                    if result.final_mark is not None:
                        bucket["final_marks"].append(result.final_mark)
                    if result.spelling_errors is not None and "spelling" in bucket:
                        bucket["spelling"].append(result.spelling_errors)
                    if result.punctuation_errors is not None and "punctuation" in bucket:
                        bucket["punctuation"].append(result.punctuation_errors)
                    if result.grammar_errors is not None and "grammar_errors" in bucket:
                        bucket["grammar_errors"].append(result.grammar_errors)

    def finalize(items, key_name):
        out = []
        for data in items.values():
            avg_percent = _safe_avg(data["percents"], digits=1)
            marks = data["marks"]
            quality = round(((sum(1 for m in marks if m in [4, 5]) / len(marks)) * 100), 1) if marks else None
            success = round(((sum(1 for m in marks if m in [3, 4, 5]) / len(marks)) * 100), 1) if marks else None
            out.append({
                key_name: data[key_name],
                "teacher_id": data.get("teacher_id"),
                "class_id": data.get("class_id"),
                "subjects_text": ", ".join(sorted(x for x in data["subject_names"] if x)) or "—",
                "results": data["results"],
                "participants": data["participants"],
                "absent": data["absent"],
                "avg_percent": avg_percent,
                "quality": quality,
                "success": success,
                "avg_mark": _safe_avg(marks, digits=2),
            })
        return sorted(out, key=lambda x: ((x["avg_percent"] is None), -(x["avg_percent"] or 0), x[key_name]))

    def finalize_dictation(items, key_name):
        out = []
        for data in items.values():
            participants = data["participants"]
            absent = data["absent"]
            total = participants + absent
            final_marks = data["final_marks"] or data["dictation_marks"]
            quality = round((sum(1 for m in final_marks if m in [4,5]) / len(final_marks)) * 100, 1) if final_marks else None
            low = sum(1 for m in final_marks if m == 2)
            out.append({
                key_name: data[key_name],
                "teacher_id": data.get("teacher_id"),
                "class_id": data.get("class_id"),
                "participants": participants,
                "absent": absent,
                "participation_percent": round((participants / total) * 100, 1) if total else None,
                "avg_dictation_mark": _safe_avg(data["dictation_marks"], digits=2),
                "avg_grammar_mark": _safe_avg(data["grammar_marks"], digits=2),
                "avg_final_mark": _safe_avg(final_marks, digits=2),
                "avg_spelling_errors": _safe_avg(data["spelling"], digits=2),
                "avg_punctuation_errors": _safe_avg(data["punctuation"], digits=2),
                "avg_grammar_errors": _safe_avg(data["grammar_errors"], digits=2),
                "quality": quality,
                "low_count": low,
            })
        return sorted(out, key=lambda x: ((x["avg_final_mark"] is None), -(x["avg_final_mark"] or 0), x[key_name]))

    work_rows = []
    for data in work_stats.values():
        final_marks = data["final_marks"] or data["dictation_marks"]
        avg_mark = _safe_avg(final_marks if data["is_dictation"] else data["marks"], digits=2)
        work_rows.append({
            "work_id": data["work_id"],
            "subject_name": data["subject_name"],
            "work_kind_label": data["work_kind_label"],
            "theme": data["theme"],
            "work_date": _fmt_date(data["work_date"]),
            "class_name": data["class_name"],
            "class_id": data["class_id"],
            "teacher_name": data["teacher_name"],
            "results": data["results"],
            "participants": data["participants"],
            "absent": data["absent"],
            "participation_percent": round((data["participants"] / data["results"]) * 100, 1) if data["results"] else None,
            "avg_percent": _safe_avg(data["percents"], digits=1),
            "avg_mark": avg_mark,
            "avg_dictation_mark": _safe_avg(data["dictation_marks"], digits=2) if data["is_dictation"] else None,
            "avg_grammar_mark": _safe_avg(data["grammar_marks"], digits=2) if data["is_dictation"] else None,
        })
    work_rows = sorted(work_rows, key=lambda x: (x["class_name"], x["work_date"], x["subject_name"], x["theme"]))

    overall_quality = round(((mark_counts[4] + mark_counts[5]) / sum(mark_counts.values())) * 100, 1) if sum(mark_counts.values()) else None
    overall_success = round(((mark_counts[3] + mark_counts[4] + mark_counts[5]) / sum(mark_counts.values())) * 100, 1) if sum(mark_counts.values()) else None
    dictation_rows = [item for item in rows if _is_dictation(item[1])]
    dictation_present = [r for r, *_ in dictation_rows if not getattr(r, "is_absent", False)]
    dictation_absent = [r for r, *_ in dictation_rows if getattr(r, "is_absent", False)]
    dictation_final_marks = [r.final_mark if r.final_mark is not None else r.dictation_mark for r in dictation_present if (r.final_mark is not None or r.dictation_mark is not None)]

    return {
        "summary": {
            "works_count": work_count,
            "results_count": len(rows),
            "participants_count": sum(1 for result, *_ in rows if not getattr(result, "is_absent", False)),
            "absent_count": sum(1 for result, *_ in rows if getattr(result, "is_absent", False)),
            "participation_percent": round((sum(1 for result, *_ in rows if not getattr(result, "is_absent", False)) / len(rows)) * 100, 1) if rows else None,
            "avg_percent": _safe_avg(percents, digits=1),
            "quality": overall_quality,
            "success": overall_success,
            "mark_counts": mark_counts,
        },
        "teacher_rows": finalize(teacher_stats, "teacher_name")[:50],
        "class_rows": finalize(class_stats, "class_name")[:50],
        "child_rows": finalize(child_stats, "child_name")[:100],
        "work_rows": work_rows[:100],
        "dictation": {
            "results_count": len(dictation_rows),
            "participants_count": len(dictation_present),
            "absent_count": len(dictation_absent),
            "participation_percent": round((len(dictation_present) / len(dictation_rows)) * 100, 1) if dictation_rows else None,
            "avg_dictation_mark": _safe_avg([r.dictation_mark for r in dictation_present if r.dictation_mark is not None], digits=2),
            "avg_grammar_mark": _safe_avg([r.grammar_mark for r in dictation_present if r.grammar_mark is not None], digits=2),
            "avg_final_mark": _safe_avg(dictation_final_marks, digits=2),
            "avg_spelling_errors": _safe_avg([r.spelling_errors for r in dictation_present if r.spelling_errors is not None], digits=2),
            "avg_punctuation_errors": _safe_avg([r.punctuation_errors for r in dictation_present if r.punctuation_errors is not None], digits=2),
            "avg_grammar_errors": _safe_avg([r.grammar_errors for r in dictation_present if r.grammar_errors is not None], digits=2),
            "quality": round((sum(1 for m in dictation_final_marks if m in [4,5]) / len(dictation_final_marks)) * 100, 1) if dictation_final_marks else None,
            "low_count": sum(1 for m in dictation_final_marks if m == 2),
            "teacher_rows": finalize_dictation(dictation_teacher_stats, "teacher_name")[:50],
            "class_rows": finalize_dictation(dictation_class_stats, "class_name")[:50],
            "child_rows": finalize_dictation(dictation_child_stats, "child_name")[:100],
        },
    }


def _make_summary_excel(dataset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Свод"
    ws.append(["Показатель", "Значение"])
    summary = dataset["summary"]
    ws.append(["Контрольных работ", summary["works_count"]])
    ws.append(["Записей", summary["results_count"]])
    ws.append(["Писали", summary["participants_count"]])
    ws.append(["Отсутствовали", summary["absent_count"]])
    ws.append(["Процент участия", summary["participation_percent"]])
    ws.append(["Средний процент", summary["avg_percent"]])
    ws.append(["Качество", summary["quality"]])
    ws.append(["Успеваемость", summary["success"]])
    ws.append(["Оценка 5", summary["mark_counts"][5]])
    ws.append(["Оценка 4", summary["mark_counts"][4]])
    ws.append(["Оценка 3", summary["mark_counts"][3]])
    ws.append(["Оценка 2", summary["mark_counts"][2]])

    def add_sheet(title, headers, rows):
        sh = wb.create_sheet(title)
        sh.append(headers)
        for row in rows:
            sh.append(row)
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        sh.freeze_panes = "A2"

    add_sheet("По учителям", ["Учитель", "Предметы", "Записей", "Писали", "Отсутствовали", "Средний %", "Средний балл", "Качество", "Успеваемость"], [
        [r["teacher_name"], r["subjects_text"], r["results"], r["participants"], r["absent"], r["avg_percent"], r["avg_mark"], r["quality"], r["success"]] for r in dataset["teacher_rows"]
    ])
    add_sheet("По классам", ["Класс", "Предметы", "Записей", "Писали", "Отсутствовали", "Средний %", "Качество", "Успеваемость"], [
        [r["class_name"], r["subjects_text"], r["results"], r["participants"], r["absent"], r["avg_percent"], r["quality"], r["success"]] for r in dataset["class_rows"]
    ])
    add_sheet("Работы", ["Дата", "Класс", "Предмет", "Вид работы", "Название", "Учитель", "Записей", "Писали", "Отсутствовали", "Процент участия", "Средний %", "Средний балл", "Ср. диктант", "Ср. грамматическое"], [
        [r["work_date"], r["class_name"], r["subject_name"], r["work_kind_label"], r["theme"], r["teacher_name"], r["results"], r["participants"], r["absent"], r["participation_percent"], r["avg_percent"], r["avg_mark"], r["avg_dictation_mark"], r["avg_grammar_mark"]] for r in dataset["work_rows"]
    ])
    if dataset.get("dictation") and dataset["dictation"].get("results_count"):
        add_sheet("Диктанты по классам", ["Класс", "Писали", "Отсутствовали", "Ср. диктант", "Ср. грамм.", "Ср. итог", "Орф.", "Пункт.", "Грамм.", "Качество"], [
            [r["class_name"], r["participants"], r["absent"], r["avg_dictation_mark"], r["avg_grammar_mark"], r["avg_final_mark"], r["avg_spelling_errors"], r["avg_punctuation_errors"], r["avg_grammar_errors"], r["quality"]] for r in dataset["dictation"]["class_rows"]
        ])
    for sh in wb.worksheets:
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@control_bp.route("/summary/export.xlsx")
@login_required
def control_works_summary_xlsx():
    if not has_permission("control_works_view"):
        abort(403)
    current_year = _current_year()
    selected_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    dataset = _control_summary_dataset(
        selected_year_id=selected_year_id,
        selected_subject_id=request.args.get("subject_id", type=int),
        selected_teacher_id=request.args.get("teacher_id", type=int),
        selected_grade=request.args.get("grade", type=int),
        selected_class_id=request.args.get("class_id", type=int),
    )
    output = _make_summary_excel(dataset)
    return send_file(output, as_attachment=True, download_name="control_works_summary.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@control_bp.route("/summary")
@login_required
def control_works_summary():
    if not has_permission("control_works_view"):
        abort(403)
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    current_year = _current_year()
    selected_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    selected_subject_id = request.args.get("subject_id", type=int)
    selected_teacher_id = request.args.get("teacher_id", type=int)
    selected_grade = request.args.get("grade", type=int)
    selected_class_id = request.args.get("class_id", type=int)

    teacher_query = User.query.order_by(User.last_name.asc(), User.first_name.asc())
    department_ids = _department_ids_for_user()
    if department_ids:
        teacher_ids = sorted({row.teacher_id for row in TeacherLoad.query.filter(TeacherLoad.department_id.in_(department_ids), TeacherLoad.is_archived.is_(False)).all() if row.teacher_id})
        teacher_query = teacher_query.filter(User.id.in_(teacher_ids or [0]))
    if not has_permission("control_works_edit") and getattr(current_user, "role", None) != "METHODIST":
        teacher_query = teacher_query.filter(User.id == current_user.id)
        selected_teacher_id = current_user.id
    teachers = teacher_query.all()

    classes_query = SchoolClass.query.order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc())
    if selected_year_id:
        classes_query = classes_query.filter(SchoolClass.academic_year_id == selected_year_id)
    classes = classes_query.all()
    parallels = sorted({c.grade for c in classes if c.grade is not None})

    dataset = _control_summary_dataset(
        selected_year_id=selected_year_id,
        selected_subject_id=selected_subject_id,
        selected_teacher_id=selected_teacher_id,
        selected_grade=selected_grade,
        selected_class_id=selected_class_id,
    )
    return render_template("control_works/summary.html", years=years, subjects=subjects, teachers=teachers, classes=classes, parallels=parallels, selected_year_id=selected_year_id, selected_subject_id=selected_subject_id, selected_teacher_id=selected_teacher_id, selected_grade=selected_grade, selected_class_id=selected_class_id, dataset=dataset)




@control_bp.route("/registry/export.xlsx")
@login_required
def control_works_registry_xlsx():
    if not has_permission("control_works_view"):
        abort(403)
    current_year = _current_year()
    selected_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    dataset = _build_registry_dataset(
        selected_year_id=selected_year_id,
        selected_subject_id=request.args.get("subject_id", type=int),
        selected_teacher_id=request.args.get("teacher_id", type=int),
        selected_class_id=request.args.get("class_id", type=int),
        selected_grade=request.args.get("grade", type=int),
        selected_work_kind=(request.args.get("work_kind") or "").strip() or None,
        selected_status=(request.args.get("status") or "").strip() or None,
        search_text=(request.args.get("q") or "").strip() or None,
        overdue_only=request.args.get("overdue") == "1",
        has_absent=True if request.args.get("has_absent") == "1" else (False if request.args.get("has_absent") == "0" else None),
        only_attention=request.args.get("attention") == "1",
        include_archived=request.args.get("include_archived") == "1",
    )
    output = _make_registry_excel(dataset["rows"])
    return send_file(output, as_attachment=True, download_name="control_works_registry.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@control_bp.route("/registry")
@login_required
def control_works_registry():
    if not has_permission("control_works_view"):
        abort(403)
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    current_year = _current_year()
    selected_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    selected_subject_id = request.args.get("subject_id", type=int)
    selected_teacher_id = request.args.get("teacher_id", type=int)
    selected_class_id = request.args.get("class_id", type=int)
    selected_grade = request.args.get("grade", type=int)
    selected_work_kind = (request.args.get("work_kind") or "").strip() or None
    selected_status = (request.args.get("status") or "").strip() or None
    search_text = (request.args.get("q") or "").strip()
    overdue_only = request.args.get("overdue") == "1"
    has_absent = True if request.args.get("has_absent") == "1" else (False if request.args.get("has_absent") == "0" else None)
    attention_only = request.args.get("attention") == "1"
    include_archived = request.args.get("include_archived") == "1"

    teacher_query = User.query.order_by(User.last_name.asc(), User.first_name.asc())
    department_ids = _department_ids_for_user()
    if department_ids:
        teacher_ids = sorted({row.teacher_id for row in TeacherLoad.query.filter(TeacherLoad.department_id.in_(department_ids), TeacherLoad.is_archived.is_(False)).all() if row.teacher_id})
        teacher_query = teacher_query.filter(User.id.in_(teacher_ids or [0]))
    if not has_permission("control_works_edit") and getattr(current_user, "role", None) != "METHODIST":
        teacher_query = teacher_query.filter(User.id == current_user.id)
        selected_teacher_id = current_user.id
    teachers = teacher_query.all()

    classes_query = SchoolClass.query.order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc())
    if selected_year_id:
        classes_query = classes_query.filter(SchoolClass.academic_year_id == selected_year_id)
    classes = classes_query.all()
    parallels = sorted({c.grade for c in classes if c.grade is not None})

    dataset = _build_registry_dataset(
        selected_year_id=selected_year_id,
        selected_subject_id=selected_subject_id,
        selected_teacher_id=selected_teacher_id,
        selected_class_id=selected_class_id,
        selected_grade=selected_grade,
        selected_work_kind=selected_work_kind,
        selected_status=selected_status,
        search_text=search_text,
        overdue_only=overdue_only,
        has_absent=has_absent,
        only_attention=attention_only,
        include_archived=include_archived,
    )
    return render_template(
        "control_works/registry.html",
        years=years, subjects=subjects, teachers=teachers, classes=classes, parallels=parallels,
        selected_year_id=selected_year_id, selected_subject_id=selected_subject_id, selected_teacher_id=selected_teacher_id,
        selected_class_id=selected_class_id, selected_grade=selected_grade, selected_work_kind=selected_work_kind,
        selected_status=selected_status, search_text=search_text, overdue_only=overdue_only, has_absent=has_absent, attention_only=attention_only, include_archived=include_archived,
        work_kind_choices=WORK_KIND_CHOICES, status_choices=["Черновик", "Заполнение", "Заполнено", "Просрочено", "Проверено", "Закрыто", "Архив"], dataset=dataset
    )
@control_bp.route("/")
@login_required
def list_control_works():
    if not has_permission("control_works_view"):
        abort(403)

    if not has_permission("control_works_edit"):
        return redirect(url_for("control_works.my_control_works"))

    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    selected_year_id = request.args.get("academic_year_id", type=int) or (_current_year().id if _current_year() else None)
    query = _assigned_works_query()
    if selected_year_id:
        query = query.filter(ControlWork.academic_year_id == selected_year_id)
    works = query.all()
    stats = {}
    registry_stats = {}
    for work in works:
        assignments = work.assignments or []
        stats[work.id] = {
            "total": len(assignments),
            "filled": sum(1 for a in assignments if a.status == "FILLED")
        }
        registry_stats[work.id] = _build_work_registry_row(work)
    return render_template("control_works/list.html", works=works, stats=stats, registry_stats=registry_stats, my_only=False, years=years, selected_year_id=selected_year_id)


@control_bp.route("/my")
@login_required
def my_control_works():
    if not has_permission("control_works_view"):
        abort(403)
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    selected_year_id = request.args.get("academic_year_id", type=int) or (_current_year().id if _current_year() else None)
    works_q = (
        ControlWork.query
        .join(ControlWorkAssignment, ControlWorkAssignment.control_work_id == ControlWork.id)
        .filter(ControlWorkAssignment.teacher_id == current_user.id)
        .distinct()
        .order_by(ControlWork.work_date.desc().nullslast(), ControlWork.created_at.desc())
        
    )
    if selected_year_id:
        works_q = works_q.filter(ControlWork.academic_year_id == selected_year_id)
    works = works_q.all()
    stats = {}
    registry_stats = {}
    for work in works:
        assignments = [a for a in (work.assignments or []) if a.teacher_id == current_user.id]
        stats[work.id] = {"total": len(assignments), "filled": sum(1 for a in assignments if a.status == "FILLED")}
        registry_stats[work.id] = _build_work_registry_row(work)
    return render_template("control_works/list.html", works=works, stats=stats, registry_stats=registry_stats, my_only=True, years=years, selected_year_id=selected_year_id)


@control_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_control_work():
    if not has_permission("control_works_edit"):
        abort(403)
    selected_subject_id = request.form.get("subject_id", type=int) or (work.subject_id if "work" in locals() and work else None)
    year, classes, teachers, subjects, parallels, teacher_options_map, auto_teacher_by_class = _form_context(selected_subject_id=selected_subject_id)
    task_count = _task_count_from_request_or_work()

    if request.method == "POST":
        try:
            work = _save_work_from_form()
            db.session.commit()
            flash("Контрольная работа создана.", "success")
            return redirect(url_for("control_works.view_control_work", work_id=work.id))
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")

    return render_template("control_works/form.html", classes=classes, teachers=teachers, subjects=subjects, task_count=task_count, parallels=parallels, teacher_options_map=teacher_options_map, auto_teacher_by_class=auto_teacher_by_class, work_kind_choices=WORK_KIND_CHOICES, work=None)


@control_bp.route("/<int:work_id>/edit", methods=["GET", "POST"])
@login_required
def edit_control_work(work_id):
    if not has_permission("control_works_edit"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    selected_subject_id = request.form.get("subject_id", type=int) or (work.subject_id if "work" in locals() and work else None)
    year, classes, teachers, subjects, parallels, teacher_options_map, auto_teacher_by_class = _form_context(selected_subject_id=selected_subject_id)
    task_count = _task_count_from_request_or_work(work)

    if request.method == "POST":
        try:
            _save_work_from_form(work)
            db.session.commit()
            flash("Настройки контрольной обновлены.", "success")
            return redirect(url_for("control_works.view_control_work", work_id=work.id))
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")

    selected_class_ids = {a.school_class_id for a in (work.assignments or [])}
    teacher_by_class = {a.school_class_id: a.teacher_id for a in (work.assignments or [])}
    return render_template(
        "control_works/form.html",
        classes=classes,
        teachers=teachers,
        subjects=subjects,
        task_count=task_count,
        parallels=parallels,
        work=work,
        work_kind_choices=WORK_KIND_CHOICES,
        selected_class_ids=selected_class_ids,
        teacher_by_class=teacher_by_class,
        teacher_options_map=teacher_options_map,
        auto_teacher_by_class=auto_teacher_by_class,
    )


@control_bp.route("/<int:work_id>/delete", methods=["POST"])
@login_required
def delete_control_work(work_id):
    if not has_permission("control_works_edit"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    db.session.delete(work)
    db.session.commit()
    flash("Контрольная работа удалена.", "success")
    return redirect(url_for("control_works.list_control_works"))


@control_bp.route("/<int:work_id>")
@login_required
def view_control_work(work_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    if not has_permission("control_works_edit"):
        allowed = any(a.teacher_id == current_user.id for a in (work.assignments or []))
        if not allowed:
            abort(403)

    report_pack = _build_control_work_report(work)
    assignments = [a for a in (work.assignments or []) if has_permission("control_works_edit") or a.teacher_id == current_user.id]
    filled_count = sum(1 for a in assignments if a.status == 'FILLED')
    assignment_status = {
        "total": len(assignments),
        "filled": filled_count,
        "pending": max(len(assignments) - filled_count, 0),
    }

    return render_template(
        "control_works/detail.html",
        work=work,
        is_dictation=_is_dictation(work),
        dictation_rules_text=_dictation_rules_text(work) if _is_dictation(work) else None,
        report_pack=report_pack,
        report=report_pack["report"],
        assignment_status=assignment_status,
        registry_row=_build_work_registry_row(work),
        history_rows=ControlWorkLog.query.filter_by(control_work_id=work.id).order_by(ControlWorkLog.created_at.desc()).limit(20).all(),
        return_args=_query_args_without_page(),
    )


@control_bp.route("/<int:work_id>/assignment/<int:assignment_id>", methods=["GET", "POST"])
@login_required
def assignment_results(work_id, assignment_id):
    if not has_permission("control_works_view"):
        abort(403)

    work = ControlWork.query.get_or_404(work_id)
    assignment = ControlWorkAssignment.query.get_or_404(assignment_id)
    if assignment.control_work_id != work.id:
        abort(404)
    if not _can_view_assignment(assignment):
        abort(403)

    school_class = assignment.school_class
    year = _current_year()
    enrollments = []
    if year and school_class:
        enrollments = (
            ChildEnrollment.query
            .join(Child, Child.id == ChildEnrollment.child_id)
            .filter(
                ChildEnrollment.academic_year_id == year.id,
                ChildEnrollment.school_class_id == school_class.id,
                ChildEnrollment.ended_at.is_(None),
            )
            .order_by(Child.last_name.asc(), Child.first_name.asc(), Child.middle_name.asc())
            .all()
        )
    tasks = sorted((work.tasks or []), key=lambda x: x.task_number or 0)
    max_total = sum(task.max_score or 0 for task in tasks)

    existing = {r.child_id: r for r in ControlWorkResult.query.filter_by(control_work_id=work.id, assignment_id=assignment.id, school_class_id=school_class.id).all()}
    posted_scores = {}

    if request.method == "POST":
        has_errors = False
        for en in enrollments:
            is_absent = request.form.get(f"absent_{en.child_id}") == "1"
            total_score = 0
            any_value = False
            dictation_payload = {}

            if _is_dictation(work):
                try:
                    dictation_payload = {
                        "grammar_mark": _parse_int_field(f"grammar_mark_{en.child_id}", f"{en.child.fio}: отметка за грамматическое задание", min_value=2, max_value=5),
                        "spelling_errors": _parse_int_field(f"spelling_errors_{en.child_id}", f"{en.child.fio}: орфографические ошибки", min_value=0),
                        "punctuation_errors": _parse_int_field(f"punctuation_errors_{en.child_id}", f"{en.child.fio}: пунктуационные ошибки", min_value=0),
                        "grammar_errors": _parse_int_field(f"grammar_errors_{en.child_id}", f"{en.child.fio}: грамматические ошибки", min_value=0),
                        "corrections_count": _parse_int_field(f"corrections_count_{en.child_id}", f"{en.child.fio}: исправления", min_value=0),
                        "teacher_comment": (request.form.get(f"teacher_comment_{en.child_id}") or "").strip() or None,
                    }
                except ValueError as e:
                    flash(str(e), "danger")
                    has_errors = True
                    continue
                has_dictation_data = any(dictation_payload.get(key) is not None for key in ["spelling_errors", "punctuation_errors", "grammar_errors", "corrections_count"])
                if not is_absent and has_dictation_data:
                    dictation_payload["dictation_mark"] = _dictation_mark_from_errors(
                        dictation_payload.get("spelling_errors") or 0,
                        dictation_payload.get("punctuation_errors") or 0,
                        work=work,
                        grammar_errors=dictation_payload.get("grammar_errors") or 0,
                        corrections_count=dictation_payload.get("corrections_count") or 0,
                    )
                else:
                    dictation_payload["dictation_mark"] = None
                dictation_payload["final_mark"] = None
            else:
                for task in tasks:
                    raw_value = (request.form.get(f"task_{task.id}_{en.child_id}") or "").strip()
                    posted_scores[(en.child_id, task.id)] = raw_value
                    if is_absent:
                        continue
                    if raw_value == "":
                        continue
                    any_value = True
                    try:
                        value = int(raw_value)
                    except ValueError:
                        flash(f"{en.child.fio}: в задании {task.task_number} должно быть целое число.", "danger")
                        has_errors = True
                        continue
                    max_score = task.max_score or 0
                    if value < 0 or value > max_score:
                        flash(f"{en.child.fio}: в задании {task.task_number} допустимо только от 0 до {max_score}.", "danger")
                        has_errors = True
                        continue
                    total_score += value

            if has_errors:
                continue

            row = existing.get(en.child_id)
            if row is None:
                row = ControlWorkResult(
                    control_work_id=work.id,
                    assignment_id=assignment.id,
                    school_class_id=school_class.id,
                    academic_year_id=(work.academic_year_id or (year.id if year else None)),
                    child_id=en.child_id,
                    created_by=current_user.id,
                    grade5_percent=work.grade5_percent,
                    grade4_percent=work.grade4_percent,
                    grade3_percent=work.grade3_percent,
                    retention_until=work.retention_until,
                )
                db.session.add(row)

            row.assignment_id = assignment.id
            row.school_class_id = school_class.id
            row.academic_year_id = work.academic_year_id or (year.id if year else row.academic_year_id)
            row.grade5_percent = work.grade5_percent
            row.grade4_percent = work.grade4_percent
            row.grade3_percent = work.grade3_percent
            row.retention_until = work.retention_until
            row.result_status = ControlWorkResult.RESULT_STATUS_ABSENT if is_absent else ControlWorkResult.RESULT_STATUS_PRESENT
            row.is_absent = is_absent

            if _is_dictation(work):
                if is_absent:
                    row.total_score = None
                    row.percent = None
                    row.mark = None
                    row.dictation_mark = None
                    row.grammar_mark = None
                    row.final_mark = None
                    row.spelling_errors = None
                    row.punctuation_errors = None
                    row.grammar_errors = None
                    row.corrections_count = None
                    row.teacher_comment = None
                else:
                    row.dictation_mark = dictation_payload.get("dictation_mark")
                    row.grammar_mark = dictation_payload.get("grammar_mark")
                    row.final_mark = dictation_payload.get("final_mark")
                    row.spelling_errors = dictation_payload.get("spelling_errors")
                    row.punctuation_errors = dictation_payload.get("punctuation_errors")
                    row.grammar_errors = dictation_payload.get("grammar_errors")
                    row.corrections_count = dictation_payload.get("corrections_count")
                    row.teacher_comment = dictation_payload.get("teacher_comment")
                    row.mark = row.final_mark
                    row.total_score = None
                    row.percent = None
            elif is_absent:
                row.total_score = None
                row.percent = None
                row.mark = None
                row.dictation_mark = None
                row.grammar_mark = None
                row.final_mark = None
                row.spelling_errors = None
                row.punctuation_errors = None
                row.grammar_errors = None
                row.corrections_count = None
                row.teacher_comment = None
            elif any_value:
                percent = round((total_score / max_total) * 100, 2) if max_total > 0 else None
                row.total_score = total_score
                row.percent = percent
                row.mark = _mark_from_percent(percent, work)
                row.dictation_mark = None
                row.grammar_mark = None
                row.final_mark = None
                row.spelling_errors = None
                row.punctuation_errors = None
                row.grammar_errors = None
                row.corrections_count = None
                row.teacher_comment = None
            else:
                row.total_score = None
                row.percent = None
                row.mark = None
                row.dictation_mark = None
                row.grammar_mark = None
                row.final_mark = None
                row.spelling_errors = None
                row.punctuation_errors = None
                row.grammar_errors = None
                row.corrections_count = None
                row.teacher_comment = None

            child = en.child
            if (not is_absent) and row.mark == 2:
                child.is_low = True
                existing_subjects = [x.strip() for x in (child.low_subjects or "").split(",") if x.strip()]
                if work.subject_name not in existing_subjects:
                    existing_subjects.append(work.subject_name)
                    child.low_subjects = ", ".join(existing_subjects)
                child.low_notes = f"Контрольная: {work.subject_name} — {work.theme}"

        if has_errors:
            db.session.rollback()
            return render_template("control_works/results.html", work=work, assignment=assignment, enrollments=enrollments, existing=existing, max_total=max_total, tasks=tasks, posted_scores=posted_scores, posted_absent=request.form, is_dictation=_is_dictation(work), dictation_rules_text=_dictation_rules_text(work) if _is_dictation(work) else None)

        assignment.status = "FILLED"
        work.updated_by = current_user.id
        work.updated_at = datetime.utcnow()
        _log_control_work_event(work, "results", "Сохранены результаты", details="Сохранены результаты по назначению")
        db.session.commit()
        flash("Результаты сохранены.", "success")
        return redirect(url_for("control_works.view_control_work", work_id=work.id))

    return render_template("control_works/results.html", work=work, assignment=assignment, enrollments=enrollments, existing=existing, max_total=max_total, tasks=tasks, posted_scores=posted_scores, posted_absent={}, is_dictation=_is_dictation(work))


@control_bp.route("/<int:work_id>/report")
@login_required
def control_work_report(work_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    teacher_id = None
    if not has_permission("control_works_edit"):
        if not any(a.teacher_id == current_user.id for a in (work.assignments or [])):
            abort(403)
        teacher_id = current_user.id

    payload = _build_control_work_report(work, teacher_id=teacher_id)
    payload["dictation_rules_text"] = _dictation_rules_text(work) if _is_dictation(work) else None
    payload["return_args"] = _query_args_without_page()
    return render_template("control_works/report.html", **payload)


@control_bp.route("/<int:work_id>/report/export.xlsx")
@login_required
def control_work_report_export(work_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    teacher_id = None
    if not has_permission("control_works_edit"):
        if not any(a.teacher_id == current_user.id for a in (work.assignments or [])):
            abort(403)
        teacher_id = current_user.id
    payload = _build_control_work_report(work, teacher_id=teacher_id)
    output = _make_work_report_excel(work, payload)
    subject = (work.subject_name or 'predmet').replace(' ', '_')
    date_part = work.work_date.strftime('%Y%m%d') if work.work_date else 'no_date'
    return send_file(output, as_attachment=True, download_name=f'kontrolnaya_rabota_svod_{subject}_{date_part}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@control_bp.route("/<int:work_id>/report/class/<int:assignment_id>")
@login_required
def control_work_class_report(work_id, assignment_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    assignment = ControlWorkAssignment.query.get_or_404(assignment_id)
    if assignment.control_work_id != work.id:
        abort(404)
    if not _can_view_assignment(assignment):
        abort(403)
    payload = _build_assignment_report(work, assignment)
    payload['dictation_rules_text'] = _dictation_rules_text(work) if _is_dictation(work) else None
    payload['return_args'] = _query_args_without_page()
    return render_template('control_works/class_report.html', **payload)


@control_bp.route("/<int:work_id>/report/export.pdf")
@login_required
def control_work_report_export_pdf(work_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    teacher_id = None
    if not has_permission("control_works_edit"):
        if not any(a.teacher_id == current_user.id for a in (work.assignments or [])):
            abort(403)
        teacher_id = current_user.id
    payload = _build_control_work_report(work, teacher_id=teacher_id)
    output = _render_control_work_summary_pdf(work, payload)
    subject = _safe_filename_part(work.subject_name, 'predmet')
    date_part = work.work_date.strftime('%Y%m%d') if work.work_date else 'no_date'
    return send_file(output, as_attachment=True, download_name=f'kontrolnaya_rabota_svod_{subject}_{date_part}.pdf', mimetype='application/pdf')


@control_bp.route("/<int:work_id>/report/class/<int:assignment_id>/export.pdf")
@login_required
def control_work_class_report_export_pdf(work_id, assignment_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    assignment = ControlWorkAssignment.query.get_or_404(assignment_id)
    if assignment.control_work_id != work.id:
        abort(404)
    if not _can_view_assignment(assignment):
        abort(403)
    payload = _build_assignment_report(work, assignment)
    output = _render_control_work_class_pdf(work, payload)
    class_name = _safe_filename_part(payload['report']['class_name'], 'klass')
    subject = _safe_filename_part(work.subject_name, 'predmet')
    date_part = work.work_date.strftime('%Y%m%d') if work.work_date else 'no_date'
    return send_file(output, as_attachment=True, download_name=f'kontrolnaya_rabota_{class_name}_{subject}_{date_part}.pdf', mimetype='application/pdf')


@control_bp.route("/<int:work_id>/report/class/<int:assignment_id>/export.xlsx")
@login_required
def control_work_class_report_export(work_id, assignment_id):
    if not has_permission("control_works_view"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    assignment = ControlWorkAssignment.query.get_or_404(assignment_id)
    if assignment.control_work_id != work.id:
        abort(404)
    if not _can_view_assignment(assignment):
        abort(403)
    payload = _build_assignment_report(work, assignment)
    output = _make_assignment_report_excel(work, payload)
    class_name = (payload['report']['class_name'] or 'klass').replace(' ', '_')
    subject = (work.subject_name or 'predmet').replace(' ', '_')
    date_part = work.work_date.strftime('%Y%m%d') if work.work_date else 'no_date'
    return send_file(output, as_attachment=True, download_name=f'kontrolnaya_rabota_{class_name}_{subject}_{date_part}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@control_bp.route("/registry/bulk", methods=["POST"])
@login_required
def control_works_registry_bulk():
    if not has_permission("control_works_edit"):
        abort(403)
    action = (request.form.get("action") or "").strip()
    ids = [int(x) for x in request.form.getlist("work_ids") if str(x).isdigit()]
    if not ids:
        flash("Не выбраны контрольные работы для группового действия.", "warning")
        return redirect(url_for("control_works.control_works_registry", **request.args.to_dict(flat=True)))
    works = ControlWork.query.filter(ControlWork.id.in_(ids)).all()
    status_map = {
        "to_draft": "Черновик",
        "to_fill": "Заполнение",
        "to_checked": "Проверено",
        "to_closed": "Закрыто",
        "to_archive": "Архив",
    }
    if action == "export_selected":
        rows = [_build_work_registry_row(w) for w in works]
        output = _make_registry_excel(rows)
        return send_file(output, as_attachment=True, download_name="control_works_selected.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    changed = 0
    for work in works:
        if action in status_map:
            old = _build_work_registry_row(work)["status"]
            new_status = status_map[action]
            work.manual_status = new_status
            work.is_archived = new_status == "Архив"
            work.updated_by = current_user.id
            work.updated_at = datetime.utcnow()
            _log_control_work_event(work, "bulk_status", "Массовое изменение статуса", old_value=old, new_value=new_status)
            changed += 1
    db.session.commit()
    flash(f"Групповая операция выполнена: {changed} записей.", "success")
    return redirect(url_for("control_works.control_works_registry", **request.args.to_dict(flat=True)))


@control_bp.route("/journal")
@login_required
def control_works_journal():
    if not has_permission("control_works_view"):
        abort(403)
    current_year = _current_year()
    selected_year_id = request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None)
    selected_subject_id = request.args.get("subject_id", type=int)
    selected_teacher_id = request.args.get("teacher_id", type=int)
    selected_class_id = request.args.get("class_id", type=int)
    selected_grade = request.args.get("grade", type=int)
    selected_work_kind = (request.args.get("work_kind") or "").strip() or None
    selected_status = (request.args.get("status") or "").strip() or None
    overdue_only = request.args.get("overdue") == "1"
    only_attention = request.args.get("attention") == "1"
    dataset = _build_registry_dataset(selected_year_id=selected_year_id, selected_subject_id=selected_subject_id, selected_teacher_id=selected_teacher_id, selected_class_id=selected_class_id, selected_grade=selected_grade, selected_work_kind=selected_work_kind, selected_status=selected_status, overdue_only=overdue_only, only_attention=only_attention, include_archived=True)
    years = AcademicYear.query.order_by(AcademicYear.start_date.desc().nullslast(), AcademicYear.name.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    teacher_query = User.query.order_by(User.last_name.asc(), User.first_name.asc())
    if not has_permission("control_works_edit") and getattr(current_user, "role", None) != "METHODIST":
        teacher_query = teacher_query.filter(User.id == current_user.id)
        selected_teacher_id = current_user.id
    teachers = teacher_query.all()
    classes_query = SchoolClass.query.order_by(SchoolClass.grade.asc().nullslast(), SchoolClass.name.asc())
    if selected_year_id:
        classes_query = classes_query.filter(SchoolClass.academic_year_id == selected_year_id)
    classes = classes_query.all()
    parallels = sorted({c.grade for c in classes if c.grade is not None})
    return render_template("control_works/journal.html", years=years, subjects=subjects, teachers=teachers, classes=classes, parallels=parallels, selected_year_id=selected_year_id, selected_subject_id=selected_subject_id, selected_teacher_id=selected_teacher_id, selected_class_id=selected_class_id, selected_grade=selected_grade, selected_work_kind=selected_work_kind, selected_status=selected_status, overdue_only=overdue_only, attention_only=only_attention, work_kind_choices=WORK_KIND_CHOICES, status_choices=["Черновик", "Заполнение", "Заполнено", "Просрочено", "Проверено", "Закрыто", "Архив"], dataset=dataset)


@control_bp.route("/journal/export.xlsx")
@login_required
def control_works_journal_xlsx():
    if not has_permission("control_works_view"):
        abort(403)
    current_year = _current_year()
    dataset = _build_registry_dataset(selected_year_id=request.args.get("academic_year_id", type=int) or (current_year.id if current_year else None), selected_subject_id=request.args.get("subject_id", type=int), selected_teacher_id=request.args.get("teacher_id", type=int), selected_class_id=request.args.get("class_id", type=int), selected_grade=request.args.get("grade", type=int), selected_work_kind=(request.args.get("work_kind") or "").strip() or None, selected_status=(request.args.get("status") or "").strip() or None, overdue_only=request.args.get("overdue") == "1", only_attention=request.args.get("attention") == "1", include_archived=True)
    output = _make_registry_excel(dataset["rows"])
    return send_file(output, as_attachment=True, download_name="control_works_journal.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@control_bp.route("/<int:work_id>/set-status", methods=["POST"])
@login_required
def control_work_set_status(work_id):
    if not has_permission("control_works_edit"):
        abort(403)
    work = ControlWork.query.get_or_404(work_id)
    new_status = (request.form.get("status") or "").strip()
    if new_status not in MANUAL_STATUS_CHOICES:
        flash("Недопустимый статус.", "danger")
        return redirect(url_for("control_works.view_control_work", work_id=work.id, **_query_args_without_page()))
    old_status = _build_work_registry_row(work)["status"]
    work.manual_status = new_status
    work.is_archived = new_status == "Архив"
    work.updated_by = current_user.id
    work.updated_at = datetime.utcnow()
    _log_control_work_event(work, "status", "Изменен статус работы", old_value=old_status, new_value=new_status)
    db.session.commit()
    flash("Статус обновлен.", "success")
    return redirect(url_for("control_works.view_control_work", work_id=work.id, **_query_args_without_page()))


@control_bp.route("/archive")
@login_required
def control_works_archive():
    if not has_permission("control_works_view"):
        abort(403)
    filters = _get_archive_filters()
    dataset = _build_archive_dataset(filters)
    return render_template("control_works/archive.html", **filters, dataset=dataset)


@control_bp.route("/archive/export.xlsx")
@login_required
def control_works_archive_xlsx():
    if not has_permission("control_works_view"):
        abort(403)
    filters = _get_archive_filters()
    dataset = _build_archive_dataset(filters)
    output = _make_archive_excel(filters, dataset)
    year_part = next((y.name for y in filters["years"] if y.id == filters["selected_year_id"]), "all") if filters["selected_year_id"] else "all"
    return send_file(output, as_attachment=True, download_name=f"control_works_archive_{year_part}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@control_bp.route("/archive/export.pdf")
@login_required
def control_works_archive_pdf():
    if not has_permission("control_works_view"):
        abort(403)
    filters = _get_archive_filters()
    dataset = _build_archive_dataset(filters)
    try:
        output = _make_archive_pdf(dataset, filters)
    except RuntimeError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("control_works.control_works_archive", **request.args))
    year_part = next((y.name for y in filters["years"] if y.id == filters["selected_year_id"]), "all") if filters["selected_year_id"] else "all"
    return send_file(output, as_attachment=True, download_name=f"control_works_archive_{year_part}.pdf", mimetype="application/pdf")
