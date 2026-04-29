"""Раздел «Доп. образование» — единый хаб ШСК + ДО + Кубок (свод + подстраницы).

Сессия 89 (27.04.2026). Источники данных:
- ШСК: app/sport_club.py (sportmos/full.json)
- ДО: app/extracurricular.py (data/do_index.json)
- Кубок: app/kubok.py (ClassRatingSnapshot, ADMIN-only)

Доступы:
- ADMIN / SOCIAL_PEDAGOG / METHODIST — весь свод по школе.
- CLASS_TEACHER — только свой класс.
- Кубок-блок внутри свода и в Excel — только ADMIN.
"""
from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from flask import Blueprint, abort, render_template, request, send_file
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font

from app.permissions import has_any_role, has_role, is_admin

extracurricular_hub_bp = Blueprint("extracurricular_hub", __name__, url_prefix="/extracurricular")

_FULL_ACCESS_ROLES = ("ADMIN", "SOCIAL_PEDAGOG", "METHODIST")


def _has_access() -> bool:
    return has_any_role(*_FULL_ACCESS_ROLES) or has_role("CLASS_TEACHER")


def _scope_classes():
    """Возвращает (classes, scope_label, is_full_school).

    Полный доступ — все классы текущего учебного года.
    CLASS_TEACHER — только класс, у которого teacher_user_id == current_user.id.
    """
    from app.models import AcademicYear, SchoolClass

    year = AcademicYear.query.filter_by(is_current=True).first()
    if not year:
        return [], "Учебный год не задан", False

    if has_any_role(*_FULL_ACCESS_ROLES):
        classes = (
            SchoolClass.query.filter_by(academic_year_id=year.id, is_active=True)
            .order_by(
                SchoolClass.grade.asc().nullslast(),
                SchoolClass.letter.asc().nullslast(),
                SchoolClass.name.asc(),
            )
            .all()
        )
        return classes, f"Школа · {year.name}", True

    if has_role("CLASS_TEACHER"):
        cls = (
            SchoolClass.query.filter_by(
                teacher_user_id=current_user.id,
                academic_year_id=year.id,
                is_active=True,
            )
            .order_by(SchoolClass.name.asc())
            .first()
        )
        if cls is None:
            return [], "Класс не найден", False
        return [cls], f"Класс {cls.name}", False

    return [], "Нет доступа", False


def _build_rows(classes):
    """Считает по списку классов: total, in_sport_club, in_do, kubok rating.

    Возвращает (rows, totals).
    """
    from app.extracurricular import count_in_do_for_children
    from app.models import Child, ChildEnrollment
    from app.sport_club import count_in_club_for_children
    from app.core.extensions import db

    if not classes:
        return [], {
            "total": 0,
            "in_sport_club": 0,
            "in_do": 0,
            "by_grade": {},
            "by_building": {},
        }

    class_ids = [c.id for c in classes]
    children_by_class: dict[int, list] = defaultdict(list)
    for ch, cls_id in (
        db.session.query(Child, ChildEnrollment.school_class_id)
        .join(ChildEnrollment, ChildEnrollment.child_id == Child.id)
        .filter(
            ChildEnrollment.ended_at.is_(None),
            ChildEnrollment.school_class_id.in_(class_ids),
        )
        .all()
    ):
        children_by_class[cls_id].append(ch)

    show_kubok = is_admin()
    if show_kubok:
        from app.kubok import get_rating
    else:
        get_rating = lambda _name: None  # noqa: E731

    rows = []
    totals_total = 0
    totals_sc = 0
    totals_do = 0
    by_grade: dict = defaultdict(lambda: {"total": 0, "sc": 0, "do": 0, "classes": 0})
    by_building: dict = defaultdict(lambda: {"total": 0, "sc": 0, "do": 0, "classes": 0})

    for c in classes:
        kids = children_by_class.get(c.id, [])
        total = len(kids)
        sc = count_in_club_for_children(kids)
        do = count_in_do_for_children(kids)

        kr = get_rating(c.name) if show_kubok else None
        building_name = c.building.name if c.building else "Без здания"

        rows.append(
            {
                "class": c,
                "total": total,
                "in_sport_club": sc,
                "in_do": do,
                "kubok": kr,
                "building_name": building_name,
            }
        )

        totals_total += total
        totals_sc += sc
        totals_do += do

        if c.grade is not None:
            g = by_grade[c.grade]
            g["total"] += total
            g["sc"] += sc
            g["do"] += do
            g["classes"] += 1

        b = by_building[building_name]
        b["total"] += total
        b["sc"] += sc
        b["do"] += do
        b["classes"] += 1

    return rows, {
        "total": totals_total,
        "in_sport_club": totals_sc,
        "in_do": totals_do,
        "by_grade": dict(by_grade),
        "by_building": dict(by_building),
        "show_kubok": show_kubok,
    }


@extracurricular_hub_bp.route("/")
@login_required
def index():
    if not _has_access():
        abort(403)
    classes, scope_label, is_full = _scope_classes()
    rows, totals = _build_rows(classes)
    return render_template(
        "extracurricular_hub.html",
        rows=rows,
        totals=totals,
        scope_label=scope_label,
        is_full_school=is_full,
        show_kubok=totals.get("show_kubok", False),
    )


@extracurricular_hub_bp.route("/do")
@login_required
def do():
    if not _has_access():
        abort(403)
    from app.extracurricular import list_programs, count_in_program_for_children
    from app.models import Child, ChildEnrollment
    from app.core.extensions import db

    classes, scope_label, is_full = _scope_classes()
    rows, totals = _build_rows(classes)

    program = (request.args.get("program") or "").strip()
    programs = list_programs()

    if program:
        class_ids = [c.id for c in classes]
        children_by_class: dict = {}
        if class_ids:
            for ch, cls_id in (
                db.session.query(Child, ChildEnrollment.school_class_id)
                .join(ChildEnrollment, ChildEnrollment.child_id == Child.id)
                .filter(
                    ChildEnrollment.ended_at.is_(None),
                    ChildEnrollment.school_class_id.in_(class_ids),
                )
                .all()
            ):
                children_by_class.setdefault(cls_id, []).append(ch)
        prog_total = 0
        for r in rows:
            kids = children_by_class.get(r["class"].id, [])
            r["in_program"] = count_in_program_for_children(kids, program)
            prog_total += r["in_program"]
        totals["in_program"] = prog_total

    return render_template(
        "extracurricular_do.html",
        rows=rows,
        totals=totals,
        scope_label=scope_label,
        is_full_school=is_full,
        programs=programs,
        selected_program=program,
    )


@extracurricular_hub_bp.route("/sport-club")
@login_required
def sport_club():
    if not _has_access():
        abort(403)
    classes, scope_label, is_full = _scope_classes()
    rows, totals = _build_rows(classes)
    return render_template(
        "extracurricular_sport_club.html",
        rows=rows,
        totals=totals,
        scope_label=scope_label,
        is_full_school=is_full,
    )


@extracurricular_hub_bp.route("/export.xlsx")
@login_required
def export_xlsx():
    if not _has_access():
        abort(403)
    section = (request.args.get("section") or "summary").strip().lower()
    program = (request.args.get("program") or "").strip()
    classes, scope_label, _ = _scope_classes()
    rows, totals = _build_rows(classes)
    show_kubok = totals.get("show_kubok", False)

    if section == "do" and program:
        from app.extracurricular import count_in_program_for_children
        from app.models import Child, ChildEnrollment
        from app.core.extensions import db
        class_ids = [c.id for c in classes]
        children_by_class: dict = {}
        if class_ids:
            for ch, cls_id in (
                db.session.query(Child, ChildEnrollment.school_class_id)
                .join(ChildEnrollment, ChildEnrollment.child_id == Child.id)
                .filter(
                    ChildEnrollment.ended_at.is_(None),
                    ChildEnrollment.school_class_id.in_(class_ids),
                )
                .all()
            ):
                children_by_class.setdefault(cls_id, []).append(ch)
        prog_total = 0
        for r in rows:
            kids = children_by_class.get(r["class"].id, [])
            r["in_program"] = count_in_program_for_children(kids, program)
            prog_total += r["in_program"]
        totals["in_program"] = prog_total

    wb = Workbook()
    ws = wb.active
    ws.title = "Доп.образование"

    bold = Font(bold=True)

    title = {
        "summary": "Доп. образование — свод",
        "do": "Дополнительное образование" + (f" · «{program}»" if program else ""),
        "sport-club": "ШСК",
    }.get(section, "Доп. образование")
    ws.append([title])
    ws["A1"].font = bold
    ws.append([scope_label])
    ws.append([])

    do_label = f"На программе «{program}»" if (section == "do" and program) else "ДО"
    headers = ["Класс", "Здание", "Учеников"]
    if section in ("summary", "sport-club"):
        headers.append("ШСК")
    if section in ("summary", "do"):
        headers.append(do_label)
    if section == "summary" and show_kubok:
        headers += ["Кубок (баллы)", "Кубок (место)"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for r in rows:
        line = [r["class"].name, r["building_name"], r["total"]]
        if section in ("summary", "sport-club"):
            line.append(r["in_sport_club"])
        if section in ("summary", "do"):
            line.append(r.get("in_program") if (section == "do" and program) else r["in_do"])
        if section == "summary" and show_kubok:
            kr = r.get("kubok")
            if kr:
                line += [kr.get("total_points"), f"{kr.get('place')} из {kr.get('total_classes')}"]
            else:
                line += ["", ""]
        ws.append(line)

    ws.append([])
    totals_line = ["Итого", "", totals["total"]]
    if section in ("summary", "sport-club"):
        totals_line.append(totals["in_sport_club"])
    if section in ("summary", "do"):
        totals_line.append(totals.get("in_program", 0) if (section == "do" and program) else totals["in_do"])
    ws.append(totals_line)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for col_letter in ("A", "B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[col_letter].width = 18

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"extracurricular_{section}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
