import os
from decimal import Decimal
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _number(value):
    value = Decimal(value or 0)
    if value == value.to_integral():
        return int(value)
    return float(value)


def _display(value):
    value = Decimal(value or 0)
    if value == value.to_integral():
        return str(int(value))
    return format(value.normalize(), "f").replace(".", ",")


def _column_label(column):
    if column["plan"] is None:
        plan_name = "Без УП"
    else:
        plan_name = column["plan"].name
    return (
        f"{plan_name}\n"
        f"{column['student_count']} уч. · {column['period_label']}"
    )


def build_class_plan_matrix_xlsx(matrix, academic_year_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = matrix["education_level"]
    total_columns = len(matrix["columns"]) + 1
    last_column = get_column_letter(max(total_columns, 1))

    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "Свод учебных планов по классам"
    sheet["A1"].font = Font(size=14, bold=True, color="1E2D49")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = (
        f"{academic_year_name} · {matrix['level_label']} · "
        "недельная нагрузка на 1 сентября"
    )
    sheet["A2"].font = Font(size=10, color="5E6B82")
    sheet["A2"].alignment = Alignment(horizontal="center")

    sheet["A4"] = "Предмет или курс"
    sheet.merge_cells("A4:A5")
    column_index = 2
    alternate = False
    for group in matrix["class_groups"]:
        group_width = len(group["columns"])
        group_start = column_index
        group_end = column_index + group_width - 1
        if group_width > 1:
            sheet.merge_cells(
                start_row=4,
                start_column=group_start,
                end_row=4,
                end_column=group_end,
            )
        sheet.cell(4, group_start, group["snapshot_class"].name_snapshot)
        fill = "EAF3FF" if not alternate else "F3F6FA"
        for current in range(group_start, group_end + 1):
            sheet.cell(4, current).fill = PatternFill(
                "solid",
                fgColor=fill,
            )
        for column in group["columns"]:
            sheet.cell(5, column_index, _column_label(column))
            column_index += 1
        alternate = not alternate

    thin = Side(style="thin", color="D8E1ED")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F5F7FA")
    section_fill = PatternFill("solid", fgColor="EAF3FF")
    subtotal_fill = PatternFill("solid", fgColor="F6F8FB")
    total_fill = PatternFill("solid", fgColor="DDEBFC")

    for row in sheet.iter_rows(
        min_row=4,
        max_row=5,
        min_col=1,
        max_col=total_columns,
    ):
        for cell in row:
            cell.font = Font(bold=True, color="4F5D75")
            if cell.fill.fill_type is None:
                cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

    row_index = 6
    for section in matrix["sections"]:
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=total_columns,
        )
        section_cell = sheet.cell(row_index, 1, section["label"])
        section_cell.font = Font(bold=True, color="214D84")
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(vertical="center")
        for current in range(1, total_columns + 1):
            sheet.cell(row_index, current).border = border
            sheet.cell(row_index, current).fill = section_fill
        row_index += 1

        for row in section["rows"]:
            sheet.cell(row_index, 1, row["activity"].name)
            for offset, column in enumerate(matrix["columns"], start=2):
                cell = row["cells"].get(column["key"])
                if cell is not None and cell["hours"] is not None:
                    sheet.cell(row_index, offset, _number(cell["hours"]))
            row_index += 1

        sheet.cell(row_index, 1, "Итого по разделу")
        for offset, column in enumerate(matrix["columns"], start=2):
            sheet.cell(
                row_index,
                offset,
                _number(section["column_totals"].get(column["key"], 0)),
            )
        for cell in sheet[row_index]:
            cell.fill = subtotal_fill
            cell.font = Font(bold=True)
        row_index += 1

    if matrix["sections"]:
        sheet.cell(row_index, 1, "Всего часов в неделю")
        for offset, column in enumerate(matrix["columns"], start=2):
            sheet.cell(
                row_index,
                offset,
                _number(matrix["column_totals"].get(column["key"], 0)),
            )
        for cell in sheet[row_index]:
            cell.fill = total_fill
            cell.font = Font(bold=True, color="174B88")

    for row in sheet.iter_rows(
        min_row=6,
        max_row=max(row_index, 6),
        min_col=1,
        max_col=total_columns,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "center",
                vertical="center",
                wrap_text=True,
            )

    sheet.column_dimensions["A"].width = 38
    for current in range(2, total_columns + 1):
        sheet.column_dimensions[get_column_letter(current)].width = 16
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[5].height = 34
    sheet.freeze_panes = "B6"
    sheet.auto_filter.ref = f"A4:{last_column}5"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "4:5"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _register_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        (
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        ),
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "/usr/local/share/fonts/DejaVuSans.ttf",
            "/usr/local/share/fonts/DejaVuSans-Bold.ttf",
        ),
    ]
    for regular_path, bold_path in candidates:
        if not os.path.exists(regular_path):
            continue
        regular_name = "AltairMatrixRegular"
        bold_name = "AltairMatrixBold"
        if regular_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(regular_name, regular_path))
        if os.path.exists(bold_path):
            if bold_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
        else:
            bold_name = regular_name
        return regular_name, bold_name
    return "Helvetica", "Helvetica-Bold"


def _paragraph(value, style):
    from reportlab.platypus import Paragraph

    return Paragraph(
        escape(str(value or "—")).replace("\n", "<br/>"),
        style,
    )


def _pdf_table(matrix, columns, styles, colors, fonts, page_width):
    from reportlab.platypus import Table, TableStyle

    regular_font, bold_font = fonts
    data = [[
        _paragraph("Предмет или курс", styles["Header"]),
        *[
            _paragraph(
                (
                    f"{column['snapshot_class'].name_snapshot}\n"
                    f"{_column_label(column)}"
                ),
                styles["Header"],
            )
            for column in columns
        ],
    ]]
    section_rows = []
    subtotal_rows = []
    for section in matrix["sections"]:
        section_rows.append(len(data))
        data.append([
            _paragraph(section["label"], styles["Section"]),
            *([""] * len(columns)),
        ])
        for row in section["rows"]:
            data.append([
                _paragraph(row["activity"].name, styles["Subject"]),
                *[
                    _paragraph(
                        (
                            _display(row["cells"][column["key"]]["hours"])
                            if (
                                column["key"] in row["cells"]
                                and row["cells"][column["key"]]["hours"]
                                is not None
                            )
                            else "—"
                        ),
                        styles["Cell"],
                    )
                    for column in columns
                ],
            ])
        subtotal_rows.append(len(data))
        data.append([
            _paragraph("Итого по разделу", styles["Subject"]),
            *[
                _paragraph(
                    _display(
                        section["column_totals"].get(column["key"], 0)
                    ),
                    styles["Cell"],
                )
                for column in columns
            ],
        ])

    total_row = None
    if matrix["sections"]:
        total_row = len(data)
        data.append([
            _paragraph("Всего часов в неделю", styles["Subject"]),
            *[
                _paragraph(
                    _display(
                        matrix["column_totals"].get(column["key"], 0)
                    ),
                    styles["Cell"],
                )
                for column in columns
            ],
        ])

    subject_width = 155
    numeric_width = max(48, (page_width - subject_width) / len(columns))
    table = Table(
        data,
        colWidths=[subject_width] + [numeric_width] * len(columns),
        repeatRows=1,
    )
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CAD5E3")),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for row_index in section_rows:
        commands.extend([
            ("SPAN", (0, row_index), (-1, row_index)),
            (
                "BACKGROUND",
                (0, row_index),
                (-1, row_index),
                colors.HexColor("#EAF3FF"),
            ),
            ("FONTNAME", (0, row_index), (-1, row_index), bold_font),
        ])
    for row_index in subtotal_rows:
        commands.extend([
            (
                "BACKGROUND",
                (0, row_index),
                (-1, row_index),
                colors.HexColor("#F6F8FB"),
            ),
            ("FONTNAME", (0, row_index), (-1, row_index), bold_font),
        ])
    if total_row is not None:
        commands.extend([
            (
                "BACKGROUND",
                (0, total_row),
                (-1, total_row),
                colors.HexColor("#DDEBFC"),
            ),
            ("FONTNAME", (0, total_row), (-1, total_row), bold_font),
        ])
    table.setStyle(TableStyle(commands))
    return table


def build_class_plan_matrix_pdf(matrix, academic_year_name):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
    )

    regular_font, bold_font = _register_pdf_fonts()
    page_size = landscape(A3)
    margin = 12 * mm
    page_width = page_size[0] - (margin * 2)
    styles = {
        "Title": ParagraphStyle(
            "MatrixTitle",
            fontName=bold_font,
            fontSize=14,
            leading=17,
            textColor=colors.HexColor("#1E2D49"),
            alignment=TA_CENTER,
        ),
        "Meta": ParagraphStyle(
            "MatrixMeta",
            fontName=regular_font,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#5E6B82"),
            alignment=TA_CENTER,
        ),
        "Header": ParagraphStyle(
            "MatrixHeader",
            fontName=bold_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#44516A"),
            alignment=TA_CENTER,
        ),
        "Section": ParagraphStyle(
            "MatrixSection",
            fontName=bold_font,
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#214D84"),
            alignment=TA_LEFT,
        ),
        "Subject": ParagraphStyle(
            "MatrixSubject",
            fontName=bold_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#26344D"),
            alignment=TA_LEFT,
        ),
        "Cell": ParagraphStyle(
            "MatrixCell",
            fontName=regular_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#26344D"),
            alignment=TA_CENTER,
        ),
    }

    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    story = []
    chunks = [
        matrix["columns"][start:start + 10]
        for start in range(0, len(matrix["columns"]), 10)
    ] or [[]]
    for index, columns in enumerate(chunks):
        if index:
            story.append(PageBreak())
        story.append(
            Paragraph("Свод учебных планов по классам", styles["Title"])
        )
        class_names = []
        for column in columns:
            name = column["snapshot_class"].name_snapshot
            if name not in class_names:
                class_names.append(name)
        range_label = (
            f" · классы {class_names[0]}–{class_names[-1]}"
            if len(class_names) > 1
            else f" · класс {class_names[0]}"
            if class_names else ""
        )
        story.append(
            Paragraph(
                (
                    f"{escape(academic_year_name)} · "
                    f"{escape(matrix['level_label'])}{escape(range_label)}"
                ),
                styles["Meta"],
            )
        )
        story.append(Spacer(1, 5))
        if columns:
            story.append(
                _pdf_table(
                    matrix,
                    columns,
                    styles,
                    colors,
                    (regular_font, bold_font),
                    page_width,
                )
            )
        else:
            story.append(
                Paragraph(
                    "Для выбранного уровня нет классов.",
                    styles["Meta"],
                )
            )

    document.build(story)
    stream.seek(0)
    return stream


__all__ = [
    "build_class_plan_matrix_pdf",
    "build_class_plan_matrix_xlsx",
]
