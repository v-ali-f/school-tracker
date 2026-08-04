from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import (
    EducationPlan,
    PopulationSnapshotClass,
    TariffCycle,
    TariffVersion,
    TeachingMetagroupSource,
    WorkloadAssignment,
    WorkloadNeed,
)
from app.services.teaching_group_matrix_service import (
    build_group_composition_workspace,
    build_teaching_group_matrix,
)
from app.services.teaching_group_service import current_population_snapshot


def education_level_for_grade(grade):
    if grade in range(1, 5):
        return "NOO"
    if grade in range(5, 10):
        return "OOO"
    if grade in range(10, 12):
        return "SOO"
    return None


def build_classroom_group_context(school_class):
    versions = (
        TariffVersion.query
        .join(TariffCycle)
        .filter(
            TariffCycle.academic_year_id
            == school_class.academic_year_id,
        )
        .order_by(
            (TariffVersion.status == "DRAFT").desc(),
            TariffVersion.version_no.desc(),
        )
        .all()
    )
    version = next(
        (item for item in versions if item.status == "DRAFT"),
        versions[0] if versions else None,
    )
    snapshot = (
        current_population_snapshot(version.id)
        if version is not None else None
    )
    plans = (
        EducationPlan.query
        .filter_by(
            tariff_version_id=version.id,
            plan_kind="CURRICULUM",
            root_plan_id=None,
        )
        .order_by(
            EducationPlan.education_level.asc(),
            EducationPlan.name.asc(),
        )
        .all()
        if version is not None else []
    )
    level = education_level_for_grade(school_class.grade)
    if version is None or snapshot is None or level is None:
        matrix = {
            "columns": [],
            "sections": [],
            "divided_count": 0,
            "incomplete_count": 0,
            "approved_count": 0,
        }
    else:
        matrix = build_teaching_group_matrix(
            snapshot,
            plans,
            level,
            version.id,
            grade=school_class.grade,
        )
        matrix["columns"] = [
            column
            for column in matrix["columns"]
            if (
                column["snapshot_class"].source_school_class_id
                == school_class.id
            )
        ]
    snapshot_class = (
        PopulationSnapshotClass.query
        .filter_by(
            population_snapshot_id=snapshot.id,
            source_school_class_id=school_class.id,
        )
        .first()
        if snapshot is not None else None
    )
    composition = build_group_composition_workspace(matrix)
    return {
        "school_class": school_class,
        "academic_year": school_class.academic_year,
        "version": version,
        "snapshot": snapshot,
        "snapshot_class": snapshot_class,
        "plans": plans,
        "matrix": matrix,
        "composition": composition,
        "education_level": level,
    }


def select_classroom_composition_item(composition, item_key=None):
    requested_key = (item_key or "").strip()
    if requested_key:
        return next(
            (
                item
                for item in composition["items"]
                if item["key"] == requested_key
            ),
            None,
        )
    return next(
        (
            item
            for item in composition["items"]
            if not item["composition_approved"]
        ),
        composition["items"][0]
        if composition["items"] else None,
    )


def build_classroom_curriculum_rows(context):
    """Строки учебного плана класса с преподавателями из нагрузки."""
    matrix = context["matrix"]
    source_group_ids = {
        group.id
        for section in matrix["sections"]
        for row in section["rows"]
        for column in matrix["columns"]
        for cell in [row["cells"].get(column["key"])]
        if cell is not None
        for group in cell.get("groups", ())
    }
    metagroup_ids_by_source = {
        group_id: set()
        for group_id in source_group_ids
    }
    if source_group_ids:
        for link in (
            TeachingMetagroupSource.query
            .filter(
                TeachingMetagroupSource.source_group_id.in_(
                    source_group_ids
                )
            )
            .all()
        ):
            metagroup_ids_by_source.setdefault(
                link.source_group_id,
                set(),
            ).add(link.metagroup_id)

    workload_group_ids = set(source_group_ids)
    workload_group_ids.update(
        metagroup_id
        for metagroup_ids in metagroup_ids_by_source.values()
        for metagroup_id in metagroup_ids
    )
    teacher_names_by_group = {
        group_id: set()
        for group_id in workload_group_ids
    }
    if workload_group_ids:
        assignments = (
            WorkloadAssignment.query
            .join(WorkloadNeed)
            .filter(
                WorkloadNeed.teaching_group_id.in_(
                    workload_group_ids
                ),
                WorkloadAssignment.status != "CANCELLED",
                WorkloadAssignment.employee_user_id.isnot(None),
            )
            .all()
        )
        for assignment in assignments:
            if assignment.employee is None:
                continue
            teacher_names_by_group.setdefault(
                assignment.workload_need.teaching_group_id,
                set(),
            ).add(assignment.employee.fio)

    result = []
    for section in matrix["sections"]:
        for row in section["rows"]:
            for column in matrix["columns"]:
                cell = row["cells"].get(column["key"])
                if cell is None or column["is_unassigned"]:
                    continue
                groups = tuple(cell.get("groups", ()))
                teacher_names = set()
                for group in groups:
                    teacher_names.update(
                        teacher_names_by_group.get(group.id, ())
                    )
                    for metagroup_id in metagroup_ids_by_source.get(
                        group.id,
                        (),
                    ):
                        teacher_names.update(
                            teacher_names_by_group.get(
                                metagroup_id,
                                (),
                            )
                        )
                result.append({
                    "section_label": section["label"],
                    "activity": row["activity"],
                    "plan": column["plan"],
                    "weekly_hours": cell.get("hours"),
                    "group_count": len(groups) or 1,
                    "teacher_names": tuple(sorted(
                        teacher_names,
                        key=str.casefold,
                    )),
                })
    return result


def build_classroom_group_xlsx(context):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Распределение"
    school_class = context["school_class"]
    academic_year = context["academic_year"]
    composition = context["composition"]

    sheet.append(["Распределение учеников по учебным группам"])
    sheet.append([
        "Класс",
        school_class.name,
        "Учебный год",
        academic_year.name,
    ])
    sheet.append([])
    sheet.append([
        "Предмет",
        "Ученик",
        "Группа",
        "Преподаватель",
        "Согласование",
    ])

    header_fill = PatternFill("solid", fgColor="DCEBFA")
    approved_fill = PatternFill("solid", fgColor="E2F4E8")
    for cell in sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
    sheet["A1"].font = Font(bold=True, size=14)

    for item in composition["items"]:
        group_indexes = {
            group.id: index
            for index, group in enumerate(item["groups"], start=1)
        }
        for enrollment in item["enrollments"]:
            group_id = item["assignment_by_member_id"].get(enrollment.id)
            teacher_names = (
                item["teacher_names_by_group"].get(group_id, ())
                if group_id is not None else ()
            )
            sheet.append([
                item["activity"].name,
                enrollment.fio_snapshot,
                (
                    f"Группа {group_indexes[group_id]}"
                    if group_id in group_indexes
                    else "Не распределён"
                ),
                ", ".join(teacher_names) or "Не назначен",
                (
                    "Согласовано классным руководителем"
                    if item["composition_approved"]
                    else "Не согласовано"
                ),
            ])
            if item["composition_approved"]:
                for cell in sheet[sheet.max_row]:
                    cell.fill = approved_fill

    if not composition["items"]:
        sheet.append([
            "В классе пока нет предметов с делением на группы."
        ])

    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = (
        f"A4:E{sheet.max_row}"
        if sheet.max_row >= 4 else None
    )
    for column, width in {
        "A": 32,
        "B": 38,
        "C": 18,
        "D": 34,
        "E": 38,
    }.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


__all__ = [
    "build_classroom_group_context",
    "build_classroom_curriculum_rows",
    "build_classroom_group_xlsx",
    "education_level_for_grade",
    "select_classroom_composition_item",
]
