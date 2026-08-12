"""Excel and PDF exports for the group matrix and metagroup register."""

from decimal import Decimal
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.class_plan_matrix_export_service import _register_pdf_fonts
from app.services.class_plan_matrix_service import effective_line_weekly_hours


def _number(value):
    value = Decimal(value or 0)
    if value == value.to_integral():
        return int(value)
    return float(value)


def _display(value):
    value = Decimal(value or 0)
    if value == value.to_integral():
        return str(int(value))
    return format(value.normalize(), "f").replace(".", ",")


def _column_label(column):
    plan_name = column["plan"].name if column["plan"] else "Без УП"
    return (
        f"{column['class_display_name']}\n"
        f"{plan_name}\n{column['student_count']} уч."
    )


def _base_sheet(title, subtitle, total_columns):
    workbook = Workbook()
    sheet = workbook.active
    last_column = get_column_letter(max(total_columns, 1))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=14, bold=True, color="1E2D49")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color="5E6B82")
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    return workbook, sheet


def _finish_sheet(sheet, *, max_row, max_column, freeze_panes):
    thin = Side(style="thin", color="CAD5E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(
        min_row=4,
        max_row=max(max_row, 4),
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "center",
                vertical="center",
                wrap_text=True,
            )
    sheet.freeze_panes = freeze_panes
    sheet.print_title_rows = "4:4"


def build_teaching_group_matrix_xlsx(matrix, academic_year_name):
    total_columns = len(matrix["columns"]) + 1
    workbook, sheet = _base_sheet(
        "Количество учебных групп",
        f"{academic_year_name} · {matrix['level_label']}",
        total_columns,
    )
    sheet.title = matrix["education_level"]
    sheet.cell(4, 1, "Предмет или курс")
    for offset, column in enumerate(matrix["columns"], start=2):
        sheet.cell(4, offset, _column_label(column))
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.font = Font(bold=True, color="44516A")

    row_index = 5
    section_fill = PatternFill("solid", fgColor="EAF3FF")
    divided_fill = PatternFill("solid", fgColor="FFF3CD")
    for section in matrix["sections"]:
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=total_columns,
        )
        sheet.cell(row_index, 1, section["label"])
        for current in range(1, total_columns + 1):
            sheet.cell(row_index, current).fill = section_fill
            sheet.cell(row_index, current).font = Font(
                bold=True,
                color="214D84",
            )
        row_index += 1
        for row in section["rows"]:
            sheet.cell(row_index, 1, row["activity"].name)
            for offset, column in enumerate(matrix["columns"], start=2):
                cell = row["cells"].get(column["key"])
                if cell is None or column["is_unassigned"]:
                    continue
                count = int(cell["group_count"] or 1)
                sheet.cell(row_index, offset, count)
                if count > 1:
                    sheet.cell(row_index, offset).fill = divided_fill
            row_index += 1

    _finish_sheet(
        sheet,
        max_row=row_index - 1,
        max_column=total_columns,
        freeze_panes="B5",
    )
    sheet.column_dimensions["A"].width = 38
    for current in range(2, total_columns + 1):
        sheet.column_dimensions[get_column_letter(current)].width = 15
    sheet.row_dimensions[4].height = 44
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _metagroup_rows(groups):
    rows = []
    for group in groups:
        class_names = sorted({
            link.population_snapshot_class.name_snapshot
            for link in group.source_classes
        })
        grades = sorted({
            link.population_snapshot_class.grade_snapshot
            for link in group.source_classes
            if link.population_snapshot_class.grade_snapshot is not None
        })
        grade = grades[0] if grades else None
        weekly_hours = effective_line_weekly_hours(
            group.source_plan_line,
            grade,
        )
        rows.append({
            "name": group.name,
            "activity": group.education_activity.name,
            "classes": ", ".join(class_names),
            "sources": "; ".join(
                link.source_group.name
                for link in group.metagroup_sources
            ),
            "weekly_hours": weekly_hours,
            "student_count": group.actual_size,
            "status": (
                "Состав сформирован"
                if group.status == "READY"
                else "Нужно распределить детей"
            ),
        })
    return rows


def build_metagroup_register_xlsx(groups, academic_year_name, level_label):
    headers = (
        "Название",
        "Предмет или курс",
        "Классы",
        "Исходные группы",
        "Часов в неделю",
        "Учеников",
        "Статус",
    )
    workbook, sheet = _base_sheet(
        "Реестр метагрупп",
        f"{academic_year_name} · {level_label}",
        len(headers),
    )
    sheet.title = "Метагруппы"
    for column, label in enumerate(headers, start=1):
        sheet.cell(4, column, label)
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.font = Font(bold=True, color="44516A")
    row_index = 5
    for item in _metagroup_rows(groups):
        values = (
            item["name"],
            item["activity"],
            item["classes"],
            item["sources"],
            _number(item["weekly_hours"]),
            item["student_count"],
            item["status"],
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row_index, column, value)
        row_index += 1
    _finish_sheet(
        sheet,
        max_row=row_index - 1,
        max_column=len(headers),
        freeze_panes="A5",
    )
    for column, width in enumerate((30, 26, 16, 44, 15, 12, 26), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _pdf_styles(colors, regular_font, bold_font):
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import ParagraphStyle

    return {
        "title": ParagraphStyle(
            "GroupExportTitle",
            fontName=bold_font,
            fontSize=14,
            leading=17,
            textColor=colors.HexColor("#1E2D49"),
            alignment=TA_CENTER,
        ),
        "meta": ParagraphStyle(
            "GroupExportMeta",
            fontName=regular_font,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#5E6B82"),
            alignment=TA_CENTER,
        ),
        "header": ParagraphStyle(
            "GroupExportHeader",
            fontName=bold_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#44516A"),
            alignment=TA_CENTER,
        ),
        "cell": ParagraphStyle(
            "GroupExportCell",
            fontName=regular_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#26344D"),
            alignment=TA_CENTER,
        ),
        "left": ParagraphStyle(
            "GroupExportLeft",
            fontName=regular_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#26344D"),
            alignment=TA_LEFT,
        ),
        "section": ParagraphStyle(
            "GroupExportSection",
            fontName=bold_font,
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#214D84"),
            alignment=TA_LEFT,
        ),
    }


def _p(value, style):
    from reportlab.platypus import Paragraph

    return Paragraph(
        escape(str(value or "")).replace("\n", "<br/>"),
        style,
    )


def build_teaching_group_matrix_pdf(matrix, academic_year_name):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    regular_font, bold_font = _register_pdf_fonts()
    styles = _pdf_styles(colors, regular_font, bold_font)
    page_size = landscape(A3)
    margin = 12 * mm
    page_width = page_size[0] - margin * 2
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    story = []
    chunks = [
        matrix["columns"][start:start + 12]
        for start in range(0, len(matrix["columns"]), 12)
    ] or [[]]
    for page_index, columns in enumerate(chunks):
        if page_index:
            story.append(PageBreak())
        story.append(Paragraph("Количество учебных групп", styles["title"]))
        story.append(Paragraph(
            f"{escape(academic_year_name)} · {escape(matrix['level_label'])}",
            styles["meta"],
        ))
        story.append(Spacer(1, 5))
        if not columns:
            story.append(Paragraph("Нет данных для выгрузки.", styles["meta"]))
            continue
        data = [[
            _p("Предмет или курс", styles["header"]),
            *[_p(_column_label(column), styles["header"]) for column in columns],
        ]]
        section_rows = []
        divided_cells = []
        for section in matrix["sections"]:
            section_rows.append(len(data))
            data.append([_p(section["label"], styles["section"]), *([""] * len(columns))])
            for row in section["rows"]:
                row_index = len(data)
                values = [_p(row["activity"].name, styles["left"])]
                for column_index, column in enumerate(columns, start=1):
                    cell = row["cells"].get(column["key"])
                    count = (
                        int(cell["group_count"] or 1)
                        if cell is not None and not column["is_unassigned"]
                        else None
                    )
                    values.append(_p(count, styles["cell"]))
                    if count and count > 1:
                        divided_cells.append((column_index, row_index))
                data.append(values)
        subject_width = 155
        numeric_width = max(45, (page_width - subject_width) / len(columns))
        table = Table(
            data,
            colWidths=[subject_width] + [numeric_width] * len(columns),
            repeatRows=1,
        )
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CAD5E3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for row_index in section_rows:
            commands.extend([
                ("SPAN", (0, row_index), (-1, row_index)),
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#EAF3FF")),
            ])
        for column_index, row_index in divided_cells:
            commands.append((
                "BACKGROUND",
                (column_index, row_index),
                (column_index, row_index),
                colors.HexColor("#FFF3CD"),
            ))
        table.setStyle(TableStyle(commands))
        story.append(table)
    document.build(story)
    stream.seek(0)
    return stream


def build_metagroup_register_pdf(groups, academic_year_name, level_label):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    regular_font, bold_font = _register_pdf_fonts()
    styles = _pdf_styles(colors, regular_font, bold_font)
    page_size = landscape(A4)
    margin = 10 * mm
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    data = [[
        _p("Название", styles["header"]),
        _p("Предмет", styles["header"]),
        _p("Классы", styles["header"]),
        _p("Исходные группы", styles["header"]),
        _p("Ч/нед.", styles["header"]),
        _p("Уч.", styles["header"]),
        _p("Статус", styles["header"]),
    ]]
    for item in _metagroup_rows(groups):
        data.append([
            _p(item["name"], styles["left"]),
            _p(item["activity"], styles["left"]),
            _p(item["classes"], styles["cell"]),
            _p(item["sources"], styles["left"]),
            _p(_display(item["weekly_hours"]), styles["cell"]),
            _p(item["student_count"], styles["cell"]),
            _p(item["status"], styles["left"]),
        ])
    table = Table(
        data,
        colWidths=[100, 90, 58, 180, 45, 40, 100],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CAD5E3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story = [
        Paragraph("Реестр метагрупп", styles["title"]),
        Paragraph(
            f"{escape(academic_year_name)} · {escape(level_label)}",
            styles["meta"],
        ),
        Spacer(1, 6),
        table,
    ]
    document.build(story)
    stream.seek(0)
    return stream


__all__ = [
    "build_metagroup_register_pdf",
    "build_metagroup_register_xlsx",
    "build_teaching_group_matrix_pdf",
    "build_teaching_group_matrix_xlsx",
]
