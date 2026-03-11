from __future__ import annotations

import hashlib
import io
import json
import zipfile
from datetime import datetime
from typing import Iterable, List

from openpyxl import load_workbook

from .. import db
from ..models import OlympiadImportSession, OlympiadResult, OlympiadUnmatchedRow
from .olympiad_matcher import (
    find_child_for_row,
    find_department_for_row,
    find_subject_for_row,
    find_teacher_for_row,
)

COLUMN_ALIASES = {
    "fio": [
        "фио", "фамилия имя отчество", "участник", "фио участника",
        "фамилия и инициалы", "фамилия, инициалы", "фамилия имя"
    ],
    "class_study": ["класс обучения", "класс", "класс обуч", "класс ученика"],
    "class_participation": ["класс участия", "параллель участия"],
    "school_login": ["логин школы", "school_login", "login"],
    "school_ekis": ["екіс", "екис", "код екис", "екис школы", "school_ekis"],
    "school_name": ["школа", "название школы", "образовательная организация", "наименование школы"],
    "score": ["балл", "набранный балл", "результат", "score"],
    "max_score": ["макс балл", "максимальный балл"],
    "percent": ["процент", "%"],
    "status": ["статус", "результат статус"],
    "reason": ["причина", "примечание", "комментарий"],
    "subject": ["предмет", "olympiad subject"],
    "olympiad_date": ["дата проведения"],
    "publication_date": ["дата публикации"],
}


def _norm(value):
    return " ".join(str(value or "").strip().lower().replace("ё", "е").split())


def _to_float(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return None


def _to_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            continue
    return None


def detect_columns(headers: Iterable[str]):
    mapping = {}
    header_list = [str(x or "").strip() for x in headers]
    normalized = [_norm(h) for h in header_list]
    for target, aliases in COLUMN_ALIASES.items():
        for idx, name in enumerate(normalized):
            if name == target or name in aliases:
                mapping[target] = idx
                break
    return mapping


def _rows_from_workbook(file_bytes: bytes, source_file_name: str) -> List[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = []
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        if not values:
            continue
        headers = values[0]
        mapping = detect_columns(headers)
        for idx, row in enumerate(values[1:], start=2):
            raw = list(row)
            item = {
                "sheet_name": ws.title,
                "source_file_name": source_file_name,
                "source_row_number": idx,
            }
            for key, pos in mapping.items():
                item[key] = raw[pos] if pos < len(raw) else None
            if not item.get("fio") and not item.get("score") and not item.get("school_login"):
                continue
            result.append(item)
    return result


def read_excel(file_storage) -> List[dict]:
    return _rows_from_workbook(file_storage.read(), getattr(file_storage, "filename", "import.xlsx"))


def read_zip(file_storage) -> List[dict]:
    rows = []
    payload = file_storage.read()
    with zipfile.ZipFile(io.BytesIO(payload)) as zf:
        for name in zf.namelist():
            lower = name.lower()
            if lower.endswith("/") or not lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                continue
            try:
                rows.extend(_rows_from_workbook(zf.read(name), name.split("/")[-1]))
            except Exception:
                continue
    return rows


def _normalize_school_name(value):
    return _norm(value).replace('"', '')


def filter_school_rows(rows: List[dict], use_login=True, use_ekis=True, use_name=True):
    filtered = []
    for row in rows:
        login = str(row.get("school_login") or "").strip().lower()
        ekis = str(row.get("school_ekis") or "").strip()
        school_name = _normalize_school_name(row.get("school_name") or "")
        login_ok = login == "sch778547"
        ekis_ok = ekis == "2357"
        name_ok = school_name in {"гбоу школа № 547", "гбоу школа no 547", "гбоу школа n 547"}
        if (use_login and login_ok) or (use_ekis and ekis_ok) or (use_name and name_ok):
            filtered.append(row)
    return filtered


def build_row_hash(row: dict, academic_year_id=None, stage=None, subject_id=None):
    payload = {
        "academic_year_id": academic_year_id,
        "stage": stage,
        "subject_id": subject_id,
        "fio": row.get("fio"),
        "class_study": row.get("class_study"),
        "class_participation": row.get("class_participation"),
        "score": row.get("score"),
        "status": row.get("status"),
        "sheet_name": row.get("sheet_name"),
        "source_row_number": row.get("source_row_number"),
        "source_file_name": row.get("source_file_name"),
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def extract_unique_subjects(rows: List[dict]) -> List[str]:
    result = []
    seen = set()
    for row in rows:
        value = str(row.get("subject") or "").strip()
        if value and value.lower() not in seen:
            seen.add(value.lower())
            result.append(value)
    return result


def find_existing_result(child_id, subject_id, academic_year_id, stage=None, score=None, status=None):
    if not child_id or not subject_id or not academic_year_id:
        return None
    q = OlympiadResult.query.filter(
        OlympiadResult.child_id == child_id,
        OlympiadResult.subject_id == subject_id,
        OlympiadResult.academic_year_id == academic_year_id,
        OlympiadResult.is_archived.is_(False),
    )
    if stage:
        q = q.filter(OlympiadResult.stage == stage)
    if score is not None:
        q = q.filter(OlympiadResult.score == score)
    if status:
        q = q.filter(OlympiadResult.status == status)
    return q.first()


def preview_import(rows: List[dict], academic_year_id=None, stage=None, subject_id=None,
                   teacher_binding_mode: str = "auto", selected_teacher_id=None, selected_teacher_ids=None,
                   selected_department_id=None):
    preview = []
    for row in rows[:500]:
        child, child_error = find_child_for_row(row, academic_year_id=academic_year_id)
        subject, mapped_department, subject_error = find_subject_for_row(row, manual_subject_id=subject_id)
        department, _ = find_department_for_row(row, teacher_load=None, subject=subject, subject_department=mapped_department, selected_department_id=selected_department_id)
        load, teacher_error = find_teacher_for_row(
            row,
            child=child,
            subject=subject,
            academic_year_id=academic_year_id,
            department=department,
            teacher_binding_mode=teacher_binding_mode,
            selected_teacher_id=selected_teacher_id,
            selected_teacher_ids=selected_teacher_ids,
        ) if child and subject else (None, None)
        if load and not department:
            department, _ = find_department_for_row(row, teacher_load=load, subject=subject, subject_department=mapped_department, selected_department_id=selected_department_id)
        row_subject_id = subject.id if subject else subject_id
        existing = find_existing_result(
            child.id if child else None,
            row_subject_id,
            academic_year_id,
            stage=stage,
            score=_to_float(row.get("score")),
            status=str(row.get("status") or "").strip() or None,
        )
        duplicate_status = "already_loaded" if existing else None
        reasons = [x for x in [subject_error, child_error, teacher_error] if x]
        preview.append({
            "row": row,
            "child": child,
            "child_error": child_error,
            "subject": subject,
            "subject_error": subject_error,
            "teacher_load": load,
            "teacher_error": teacher_error,
            "department": department,
            "row_hash": build_row_hash(row, academic_year_id, stage, row_subject_id),
            "duplicate_status": duplicate_status,
            "existing_result": existing,
            "unmatched_reason": "; ".join(reasons) if reasons else None,
        })
    return preview


def execute_import(rows: List[dict], *, academic_year_id: int, stage: str, subject_id=None, subject_name=None,
                   imported_by=None, teacher_binding_mode: str = "auto", selected_teacher_id=None,
                   selected_teacher_ids=None, selected_department_id=None):
    status_counter = {"winner": 0, "prizer": 0, "participant": 0}
    session = OlympiadImportSession(
        academic_year_id=academic_year_id,
        stage=stage,
        subject_id=subject_id,
        subject_name=subject_name,
        department_id=selected_department_id,
        source_file_name=rows[0].get("source_file_name") if rows else None,
        imported_by=imported_by,
        total_rows=len(rows),
        school_rows=len(rows),
        status="DONE",
    )
    db.session.add(session)
    db.session.flush()

    matched_rows = unmatched_rows = created_rows = duplicate_rows = error_rows = updated_rows = 0

    for row in rows:
        try:
            child, child_error = find_child_for_row(row, academic_year_id=academic_year_id)
            subject, mapped_department, subject_error = find_subject_for_row(row, manual_subject_id=subject_id)
            department, _ = find_department_for_row(row, teacher_load=None, subject=subject, subject_department=mapped_department, selected_department_id=selected_department_id)
            load, teacher_error = find_teacher_for_row(
                row,
                child=child,
                subject=subject,
                academic_year_id=academic_year_id,
                department=department,
                teacher_binding_mode=teacher_binding_mode,
                selected_teacher_id=selected_teacher_id,
                selected_teacher_ids=selected_teacher_ids,
            ) if child and subject else (None, None)
            if load and not department:
                department, _ = find_department_for_row(row, teacher_load=load, subject=subject, subject_department=mapped_department, selected_department_id=selected_department_id)
            resolved_subject_id = subject.id if subject else subject_id
            row_hash = build_row_hash(row, academic_year_id, stage, resolved_subject_id)
            existing_by_hash = OlympiadResult.query.filter_by(source_row_hash=row_hash).first()
            if existing_by_hash:
                duplicate_rows += 1
                continue

            unmatched_reason = "; ".join([x for x in [subject_error, child_error, teacher_error] if x]) or None
            raw_status = str(row.get("status") or "").strip()
            raw_status_lower = raw_status.lower()
            if "побед" in raw_status_lower:
                status_counter["winner"] += 1
            elif "приз" in raw_status_lower:
                status_counter["prizer"] += 1
            else:
                status_counter["participant"] += 1

            if not child or not subject:
                unmatched_rows += 1
                db.session.add(OlympiadUnmatchedRow(
                    import_session_id=session.id,
                    raw_fio=str(row.get("fio") or ""),
                    raw_class_study=str(row.get("class_study") or ""),
                    raw_class_participation=str(row.get("class_participation") or ""),
                    raw_score=str(row.get("score") or ""),
                    raw_status=raw_status,
                    raw_reason=str(row.get("reason") or ""),
                    raw_subject=str(row.get("subject") or subject_name or ""),
                    raw_stage=stage,
                    raw_school_login=str(row.get("school_login") or ""),
                    raw_school_ekis=str(row.get("school_ekis") or ""),
                    raw_payload_json=json.dumps(row, ensure_ascii=False, default=str),
                    unmatched_reason=unmatched_reason,
                    maybe_left_school=("не найден" in (child_error or "").lower()),
                ))
                continue
            matched_rows += 1
            school_class = child.current_class
            for e in child.enrollments or []:
                if e.academic_year_id == academic_year_id and e.school_class:
                    school_class = e.school_class
                    break
            existing = find_existing_result(
                child.id,
                resolved_subject_id,
                academic_year_id,
                stage=stage,
                score=_to_float(row.get("score")),
                status=raw_status or None,
            )
            if existing:
                duplicate_rows += 1
                continue
            result = OlympiadResult(
                academic_year_id=academic_year_id,
                child_id=child.id,
                school_class_id=school_class.id if school_class else None,
                teacher_id=load.teacher_id if load else None,
                department_id=department.id if department else selected_department_id,
                subject_id=resolved_subject_id,
                subject_name=(subject.name if subject else (row.get("subject") or subject_name)),
                stage=stage,
                class_study_text=str(row.get("class_study") or ""),
                class_participation_text=str(row.get("class_participation") or ""),
                score=_to_float(row.get("score")),
                max_score=_to_float(row.get("max_score")),
                percent=_to_float(row.get("percent")),
                status=raw_status or None,
                reason=str(row.get("reason") or "").strip() or None,
                olympiad_date=_to_date(row.get("olympiad_date")),
                publication_date=_to_date(row.get("publication_date")),
                school_login=str(row.get("school_login") or "").strip() or None,
                school_ekis=str(row.get("school_ekis") or "").strip() or None,
                school_name=str(row.get("school_name") or "").strip() or None,
                source_file_name=str(row.get("source_file_name") or "").strip() or None,
                source_sheet_name=str(row.get("sheet_name") or "").strip() or None,
                source_row_number=row.get("source_row_number"),
                source_row_hash=row_hash,
                import_session_id=session.id,
                created_by=imported_by,
            )
            db.session.add(result)
            created_rows += 1
        except Exception:
            error_rows += 1
            continue

    session.matched_rows = matched_rows
    session.unmatched_rows = unmatched_rows
    session.created_rows = created_rows
    session.updated_rows = updated_rows
    session.duplicate_rows = duplicate_rows
    session.error_rows = error_rows
    session.comment = (
        "В файле присутствуют результаты учеников, которые не были сопоставлены с карточками обучающихся. "
        f"Среди них: победителей — {status_counter['winner']}, призеров — {status_counter['prizer']}, участников — {status_counter['participant']}. "
        "Часть результатов может относиться к выбывшим ученикам."
        if unmatched_rows
        else "Все строки файла сопоставлены успешно."
    )
    db.session.commit()
    return session
