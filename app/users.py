from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import current_user
from sqlalchemy import or_, func, desc
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook, Workbook

from .models import User, PageVisit
from app.models_legacy import Role, UserRole
from app.core.extensions import db
from .roles import require_roles
from app.utils.user_matching import find_existing_user, normalize_fio, potential_duplicate_groups

users_bp = Blueprint("users", __name__)

ROLE_OPTIONS = [
    "ADMIN", "DIRECTOR", "CLASS_TEACHER", "CURATOR", "TEACHER", "VIEWER", "METHODIST", "KPP",
    "OLIGOPHRENOPEDAGOG", "SOCIAL_PEDAGOG", "LOGOPEDIST", "PSYCHOLOGIST", "DEFECTOLOGIST", "TUTOR", "ASSISTANT", "SENIOR_EDUCATOR", "EDUCATOR",
]
ROLE_LABELS = {
    "ADMIN": "Администратор",
    "DIRECTOR": "Директор",
    "CLASS_TEACHER": "Классный руководитель",
    "CURATOR": "Куратор",
    "TEACHER": "Учитель",
    "VIEWER": "Просмотр",
    "METHODIST": "Методист",
    "KPP": "КПП",
    "OLIGOPHRENOPEDAGOG": "Олигофренопедагог",
    "SOCIAL_PEDAGOG": "Социальный педагог",
    "LOGOPEDIST": "Учитель-логопед",
    "PSYCHOLOGIST": "Педагог-психолог",
    "DEFECTOLOGIST": "Учитель-дефектолог",
    "TUTOR": "Тьютор",
    "ASSISTANT": "Ассистент",
    "SENIOR_EDUCATOR": "Старший воспитатель",
    "EDUCATOR": "Воспитатель",
}
RUSSIAN_ROLE_MAP = {
    "администратор": "ADMIN",
    "директор": "DIRECTOR",
    "классный руководитель": "CLASS_TEACHER",
    "куратор": "CURATOR",
    "учитель": "TEACHER",
    "просмотр": "VIEWER",
    "методист": "METHODIST",
    "кпп": "KPP",
    "олигофренопедагог": "OLIGOPHRENOPEDAGOG",
    "социальный педагог": "SOCIAL_PEDAGOG",
    "учитель-логопед": "LOGOPEDIST",
    "учитель логопед": "LOGOPEDIST",
    "педагог-психолог": "PSYCHOLOGIST",
    "педагог психолог": "PSYCHOLOGIST",
    "учитель-дефектолог": "DEFECTOLOGIST",
    "учитель дефектолог": "DEFECTOLOGIST",
    "тьютор": "TUTOR",
    "ассистент": "ASSISTANT",
    "старший воспитатель": "SENIOR_EDUCATOR",
    "воспитатель": "EDUCATOR",
}
SERVICE_ROLE_CODES = {"METHODIST", "OLIGOPHRENOPEDAGOG", "SOCIAL_PEDAGOG", "LOGOPEDIST", "PSYCHOLOGIST", "DEFECTOLOGIST", "TUTOR", "ASSISTANT"}




class ServiceImportUnmatchedStaff(db.Model):
    __tablename__ = "service_import_unmatched_staff"

    id = db.Column(db.Integer, primary_key=True)
    source = db.Column(db.String(50), nullable=False, index=True)
    import_filename = db.Column(db.String(255), nullable=True)
    source_session_key = db.Column(db.String(120), nullable=True, index=True)
    imported_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)
    staff_fio = db.Column(db.String(255), nullable=False, index=True)
    normalized_fio = db.Column(db.String(255), nullable=True, index=True)
    username = db.Column(db.String(120), nullable=True)
    phone = db.Column(db.String(32), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    role_hint = db.Column(db.String(120), nullable=True)
    details = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(30), nullable=False, default="NEW", index=True)
    matched_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    matched_by_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    matched_at = db.Column(db.DateTime, nullable=True)

    matched_user = db.relationship("User", foreign_keys=[matched_user_id])
    matched_by = db.relationship("User", foreign_keys=[matched_by_user_id])


def register_unmatched_staff(*, source: str, staff_fio: str, import_filename: str | None = None, source_session_key: str | None = None, username: str | None = None, phone: str | None = None, email: str | None = None, role_hint: str | None = None, details: str | None = None):
    fio = (staff_fio or "").strip()
    if not fio:
        return None
    normalized = normalize_fio(fio=fio)
    row = ServiceImportUnmatchedStaff.query.filter_by(
        source=source,
        source_session_key=source_session_key,
        normalized_fio=normalized,
        status="NEW",
    ).first()
    if row:
        row.imported_at = datetime.utcnow()
        row.import_filename = import_filename or row.import_filename
        row.username = username or row.username
        row.phone = phone or row.phone
        row.email = email or row.email
        row.role_hint = role_hint or row.role_hint
        row.details = details or row.details
        return row
    row = ServiceImportUnmatchedStaff(
        source=source,
        import_filename=import_filename,
        source_session_key=source_session_key,
        staff_fio=fio,
        normalized_fio=normalized,
        username=username,
        phone=phone,
        email=email,
        role_hint=role_hint,
        details=details,
    )
    db.session.add(row)
    return row

class UserImportSession(db.Model):
    __tablename__ = "user_import_session"

    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    imported_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)
    imported_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)
    rows_total = db.Column(db.Integer, nullable=False, default=0)
    created_count = db.Column(db.Integer, nullable=False, default=0)
    updated_count = db.Column(db.Integer, nullable=False, default=0)
    skipped_count = db.Column(db.Integer, nullable=False, default=0)
    duplicate_count = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String(30), nullable=False, default="DONE", index=True)
    notes = db.Column(db.Text, nullable=True)
    reverted_at = db.Column(db.DateTime, nullable=True)
    reverted_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    importer = db.relationship("User", foreign_keys=[imported_by])
    reverter = db.relationship("User", foreign_keys=[reverted_by])


class UserImportRow(db.Model):
    __tablename__ = "user_import_row"

    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey("user_import_session.id"), nullable=False, index=True)
    row_num = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(120), nullable=True, index=True)
    action = db.Column(db.String(30), nullable=False, default="SKIPPED", index=True)
    note = db.Column(db.String(255), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)

    previous_role = db.Column(db.String(30), nullable=True)
    previous_last_name = db.Column(db.String(120), nullable=True)
    previous_first_name = db.Column(db.String(120), nullable=True)
    previous_middle_name = db.Column(db.String(120), nullable=True)
    previous_phone = db.Column(db.String(32), nullable=True)
    previous_email = db.Column(db.String(120), nullable=True)
    previous_employment_status = db.Column(db.String(30), nullable=True)
    previous_is_active_user = db.Column(db.Boolean, nullable=True)
    previous_archived_at = db.Column(db.DateTime, nullable=True)
    previous_dismissal_date = db.Column(db.Date, nullable=True)
    password_was_changed = db.Column(db.Boolean, nullable=False, default=False)

    import_session = db.relationship(
        "UserImportSession",
        backref=db.backref("rows", lazy=True, cascade="all, delete-orphan"),
    )
    user = db.relationship("User", foreign_keys=[user_id])


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()




def normalize_role(value: str) -> str:
    raw = (value or '').strip()
    if not raw:
        return 'VIEWER'
    upper = raw.upper()
    if upper in ROLE_OPTIONS:
        return upper
    return RUSSIAN_ROLE_MAP.get(raw.lower(), 'VIEWER')


def role_label(code: str) -> str:
    return ROLE_LABELS.get((code or '').upper(), code or '—')


def _sync_user_roles_table(user: User):
    """Синхронизирует таблицу user_role с полем user.role.
    Если у пользователя есть записи в user_role (новая схема),
    они перезаписываются в соответствии с user.role.
    """
    if not user or not user.id:
        return
    existing = UserRole.query.filter_by(user_id=user.id).all()
    if not existing:
        return
    role_code = (user.role or "VIEWER").upper()
    role_obj = Role.query.filter_by(code=role_code).first()
    for ur in existing:
        db.session.delete(ur)
    db.session.flush()
    if role_obj:
        db.session.add(UserRole(user_id=user.id, role_id=role_obj.id))


def _sync_service_specialist(user: User):
    if not user:
        return
    try:
        from app.models.service_staff import ServiceSpecialist
    except Exception:
        return
    specialist = ServiceSpecialist.query.filter_by(user_id=user.id).first()
    if (user.role or '').upper() in SERVICE_ROLE_CODES:
        if not specialist:
            specialist = ServiceSpecialist(user_id=user.id)
            db.session.add(specialist)
        specialist.last_name = user.last_name or specialist.last_name or ''
        specialist.first_name = user.first_name or specialist.first_name or ''
        specialist.middle_name = user.middle_name or None
        specialist.phone = user.phone or None
        specialist.email = user.email or None
        specialist.position_title = role_label(user.role)
        if specialist.is_active is None:
            specialist.is_active = True
    elif specialist and not specialist.position_title:
        specialist.position_title = role_label(user.role)

def _snapshot_user(row_log: UserImportRow, user: User):
    row_log.previous_role = user.role
    row_log.previous_last_name = user.last_name
    row_log.previous_first_name = user.first_name
    row_log.previous_middle_name = user.middle_name
    row_log.previous_phone = user.phone
    row_log.previous_email = user.email
    row_log.previous_employment_status = user.employment_status
    row_log.previous_is_active_user = user.is_active_user
    row_log.previous_archived_at = user.archived_at
    row_log.previous_dismissal_date = user.dismissal_date


@users_bp.route("/admin/users")
@require_roles("ADMIN")
def users_list():
    status = (request.args.get("status") or "active").lower()
    q = (request.args.get("q") or "").strip().lower()
    query = User.query

    if status == "archived":
        query = query.filter(User.employment_status.in_(["DISMISSED", "ARCHIVED"]))
    elif status == "all":
        pass
    else:
        query = query.filter(~User.employment_status.in_(["DISMISSED", "ARCHIVED"]))

    if q:
        query = query.filter(
            or_(
                db.func.lower(db.func.coalesce(User.last_name, "")).contains(q),
                db.func.lower(db.func.coalesce(User.first_name, "")).contains(q),
                db.func.lower(db.func.coalesce(User.middle_name, "")).contains(q),
                db.func.lower(User.username).contains(q),
            )
        )

    users = query.order_by(User.last_name.asc(), User.first_name.asc(), User.username.asc()).all()
    recent_imports = UserImportSession.query.order_by(UserImportSession.imported_at.desc()).limit(10).all()
    unmatched_count = ServiceImportUnmatchedStaff.query.filter_by(status="NEW").count()
    duplicate_groups = potential_duplicate_groups()
    return render_template("users_list.html", users=users, status=status, q=q, recent_imports=recent_imports, role_labels=ROLE_LABELS, unmatched_count=unmatched_count, duplicate_groups_count=len(duplicate_groups))


@users_bp.route("/admin/users/activity")
@require_roles("ADMIN")
def users_activity():
    bucket = (request.args.get("bucket") or "all").lower()
    sort = (request.args.get("sort") or "seen_asc").lower()

    query = User.query.filter(~User.employment_status.in_(["DISMISSED", "ARCHIVED"]))
    users = query.all()

    now = datetime.utcnow()

    def compute_bucket(u):
        seen = u.last_seen_at or u.last_login_at
        if seen is None:
            return "never"
        days = (now - seen).days
        if days <= 7:
            return "active"
        if days <= 30:
            return "sleepy"
        return "lost"

    rows = []
    for u in users:
        seen = u.last_seen_at or u.last_login_at
        days_ago = (now - seen).days if seen else None
        rows.append({
            "user": u,
            "last_login_at": u.last_login_at,
            "last_seen_at": u.last_seen_at,
            "effective_seen_at": seen,
            "days_ago": days_ago,
            "active_days_count": u.active_days_count or 0,
            "bucket": compute_bucket(u),
        })

    totals = {
        "total": len(rows),
        "never": sum(1 for r in rows if r["bucket"] == "never"),
        "active": sum(1 for r in rows if r["bucket"] == "active"),
        "sleepy": sum(1 for r in rows if r["bucket"] == "sleepy"),
        "lost": sum(1 for r in rows if r["bucket"] == "lost"),
    }

    if bucket in {"never", "active", "sleepy", "lost"}:
        rows = [r for r in rows if r["bucket"] == bucket]

    if sort == "seen_desc":
        rows.sort(key=lambda r: (r["effective_seen_at"] is None, r["effective_seen_at"] or datetime.min), reverse=True)
    elif sort == "days_desc":
        rows.sort(key=lambda r: r["active_days_count"], reverse=True)
    elif sort == "days_asc":
        rows.sort(key=lambda r: r["active_days_count"])
    elif sort == "fio":
        rows.sort(key=lambda r: ((r["user"].last_name or "").lower(), (r["user"].first_name or "").lower()))
    else:
        rows.sort(key=lambda r: (r["effective_seen_at"] is not None, r["effective_seen_at"] or datetime.max))

    return render_template(
        "users_activity.html",
        rows=rows,
        totals=totals,
        bucket=bucket,
        sort=sort,
        role_labels=ROLE_LABELS,
        now=now,
    )


@users_bp.route("/admin/users/paths")
@require_roles("ADMIN")
def users_paths():
    period = (request.args.get("period") or "7d").lower()
    selected_user_id = request.args.get("user_id", type=int)

    now = datetime.utcnow()
    if period == "today":
        since = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "30d":
        since = now - timedelta(days=30)
    elif period == "all":
        since = None
    else:
        period = "7d"
        since = now - timedelta(days=7)

    base = db.session.query(PageVisit)
    if since is not None:
        base = base.filter(PageVisit.ts >= since)

    total_visits = base.with_entities(func.count(PageVisit.id)).scalar() or 0
    unique_users_count = base.with_entities(
        func.count(func.distinct(PageVisit.user_id))
    ).scalar() or 0

    top_pages_q = (
        base.with_entities(PageVisit.endpoint, func.count(PageVisit.id).label("c"))
        .group_by(PageVisit.endpoint)
        .order_by(desc("c"))
        .limit(20)
    )
    top_pages = [(ep, c) for ep, c in top_pages_q.all()]

    transitions_q = (
        base.with_entities(
            PageVisit.referrer_endpoint,
            PageVisit.endpoint,
            func.count(PageVisit.id).label("c"),
        )
        .filter(PageVisit.referrer_endpoint.isnot(None))
        .filter(PageVisit.referrer_endpoint != PageVisit.endpoint)
        .group_by(PageVisit.referrer_endpoint, PageVisit.endpoint)
        .order_by(desc("c"))
        .limit(30)
    )
    top_transitions = [(a, b, c) for a, b, c in transitions_q.all()]

    active_user_ids_q = (
        base.with_entities(PageVisit.user_id, func.count(PageVisit.id).label("c"))
        .group_by(PageVisit.user_id)
        .order_by(desc("c"))
        .limit(50)
    )
    active_pairs = active_user_ids_q.all()
    active_user_ids = [uid for uid, _ in active_pairs]
    visits_by_user = {uid: c for uid, c in active_pairs}

    user_options = []
    if active_user_ids:
        users_for_picker = (
            User.query.filter(User.id.in_(active_user_ids)).all()
        )
        users_for_picker.sort(
            key=lambda u: visits_by_user.get(u.id, 0),
            reverse=True,
        )
        user_options = [
            {
                "id": u.id,
                "fio": u.fio or u.username,
                "username": u.username,
                "role": ROLE_LABELS.get(u.role, u.role),
                "visits": visits_by_user.get(u.id, 0),
            }
            for u in users_for_picker
        ]

    timeline = []
    selected_user = None
    if selected_user_id:
        selected_user = User.query.get(selected_user_id)
        if selected_user:
            tl_q = (
                db.session.query(PageVisit)
                .filter(PageVisit.user_id == selected_user_id)
            )
            if since is not None:
                tl_q = tl_q.filter(PageVisit.ts >= since)
            timeline = (
                tl_q.order_by(PageVisit.ts.desc()).limit(150).all()
            )

    return render_template(
        "users_paths.html",
        period=period,
        total_visits=total_visits,
        unique_users_count=unique_users_count,
        top_pages=top_pages,
        top_transitions=top_transitions,
        user_options=user_options,
        timeline=timeline,
        selected_user=selected_user,
        selected_user_id=selected_user_id,
        role_labels=ROLE_LABELS,
        now=now,
    )


@users_bp.route("/admin/users/new", methods=["GET", "POST"])
@require_roles("ADMIN")
def users_new():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        last_name = (request.form.get("last_name") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        middle_name = (request.form.get("middle_name") or "").strip() or None
        phone = (request.form.get("phone") or "").strip() or None
        role = normalize_role(request.form.get("role") or "VIEWER")

        if not username or not password:
            flash("Заполните логин и пароль", "danger")
            return render_template("user_new.html", roles=ROLE_OPTIONS, role_labels=ROLE_LABELS)

        if User.query.filter_by(username=username).first():
            flash("Пользователь уже существует", "danger")
            return render_template("user_new.html", roles=ROLE_OPTIONS, role_labels=ROLE_LABELS)

        u = User(username=username, role=role, last_name=last_name, first_name=first_name, middle_name=middle_name, phone=phone)
        u.set_password(password)
        db.session.add(u)
        db.session.flush()
        _sync_service_specialist(u)
        db.session.commit()
        flash("Пользователь создан", "success")
        return redirect(url_for("users.users_list"))

    return render_template("user_new.html", roles=ROLE_OPTIONS, role_labels=ROLE_LABELS)


@users_bp.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@require_roles("ADMIN")
def users_edit(user_id):
    u = User.query.get_or_404(user_id)

    if request.method == "POST":
        u.role = normalize_role(request.form.get("role") or u.role)
        password = (request.form.get("password") or "").strip()
        u.last_name = (request.form.get("last_name") or "").strip()
        u.first_name = (request.form.get("first_name") or "").strip()
        u.middle_name = (request.form.get("middle_name") or "").strip() or None
        u.phone = (request.form.get("phone") or "").strip() or None

        if password:
            u.set_password(password)

        _sync_user_roles_table(u)
        _sync_service_specialist(u)
        db.session.commit()
        flash("Сохранено", "success")
        return redirect(url_for("users.users_list"))

    return render_template("user_edit.html", u=u, roles=ROLE_OPTIONS, role_labels=ROLE_LABELS)


@users_bp.route("/admin/users/<int:user_id>/archive", methods=["POST"])
@require_roles("ADMIN")
def users_archive(user_id):
    u = User.query.get_or_404(user_id)

    if u.username == "admin":
        flash("Нельзя архивировать admin", "danger")
        return redirect(url_for("users.users_list"))

    u.employment_status = "ARCHIVED"
    u.is_active_user = False

    if not u.archived_at:
        u.archived_at = datetime.utcnow()
    if not u.dismissal_date:
        u.dismissal_date = datetime.utcnow().date()

    db.session.commit()
    flash("Сотрудник переведён в архив. Документы и история сохранены.", "success")
    return redirect(url_for("users.users_list", status="archived"))


@users_bp.route("/admin/users/<int:user_id>/restore", methods=["POST"])
@require_roles("ADMIN")
def users_restore(user_id):
    u = User.query.get_or_404(user_id)
    u.employment_status = "ACTIVE"
    u.is_active_user = True
    u.archived_at = None
    u.dismissal_date = None
    db.session.commit()
    flash("Сотрудник возвращён в активный список.", "success")
    return redirect(url_for("users.users_list", status="active"))


@users_bp.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@require_roles("ADMIN")
def users_delete(user_id):
    u = User.query.get_or_404(user_id)

    if u.username == "admin":
        flash("Нельзя удалить admin", "danger")
        return redirect(url_for("users.users_list"))

    u.employment_status = "ARCHIVED"
    u.is_active_user = False
    u.archived_at = datetime.utcnow()

    if not u.dismissal_date:
        u.dismissal_date = datetime.utcnow().date()

    db.session.commit()
    flash("Физическое удаление заменено архивированием. Сотрудник сохранён в истории.", "success")
    return redirect(url_for("users.users_list", status="archived"))


@users_bp.route("/admin/users/import", methods=["GET", "POST"])
@require_roles("ADMIN")
def users_import():
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Выберите Excel файл", "danger")
            return redirect(url_for("users.users_import"))

        try:
            wb = load_workbook(f, data_only=True)
        except Exception as exc:
            flash(f"Не удалось открыть Excel файл: {exc}", "danger")
            return redirect(url_for("users.users_import"))

        ws = wb.active
        raw_headers = [_normalize_cell(cell.value) for cell in ws[1]]
        headers = [h.lower() for h in raw_headers]
        idx = {h: i for i, h in enumerate(headers) if h}
        # aliases for user import headers
        for alias, canonical in {
            "username": "username",
            "логин": "username",
            "user": "username",
            "password": "password",
            "пароль": "password",
            "role": "role",
            "роль": "role",
            "last_name": "last_name",
            "lastname": "last_name",
            "surname": "last_name",
            "фамилия": "last_name",
            "first_name": "first_name",
            "firstname": "first_name",
            "first name": "first_name",
            "имя": "first_name",
            "middle_name": "middle_name",
            "middlename": "middle_name",
            "middle name": "middle_name",
            "отчество": "middle_name",
            "phone": "phone",
            "телефон": "phone",
            "mobile": "phone",
            "email": "email",
            "почта": "email",
            "электронная почта": "email",
        }.items():
            if canonical not in idx and alias in idx:
                idx[canonical] = idx[alias]

        required = ["username", "role", "last_name", "first_name"]
        missing = [c for c in required if c not in idx]
        if missing:
            flash(f"Не хватает колонок: {', '.join(missing)}", "danger")
            return redirect(url_for("users.users_import"))

        session_log = UserImportSession(
            filename=f.filename,
            imported_by=getattr(current_user, "id", None),
            rows_total=max(ws.max_row - 1, 0),
            status="DONE",
        )
        db.session.add(session_log)
        db.session.flush()

        created = 0
        updated = 0
        skipped = 0
        duplicates = 0
        seen_usernames = set()

        for r in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

            username = _normalize_cell(row_values[idx["username"]])
            password = _normalize_cell(row_values[idx["password"]]) if "password" in idx else ""
            role = normalize_role(_normalize_cell(row_values[idx["role"]]))
            last_name = _normalize_cell(row_values[idx["last_name"]])
            first_name = _normalize_cell(row_values[idx["first_name"]])
            phone = _normalize_cell(row_values[idx["phone"]]) or None
            middle_name = _normalize_cell(row_values[idx["middle_name"]]) if "middle_name" in idx else ""
            email = _normalize_cell(row_values[idx["email"]]) if "email" in idx else ""

            row_log = UserImportRow(
                session_id=session_log.id,
                row_num=r,
                username=username or None,
                action="SKIPPED",
            )
            db.session.add(row_log)

            if not username:
                skipped += 1
                row_log.note = "Пустой логин"
                continue

            key = username.lower()
            if key in seen_usernames:
                duplicates += 1
                skipped += 1
                row_log.action = "DUPLICATE"
                row_log.note = "Дубликат логина в файле"
                continue
            seen_usernames.add(key)

            if role == "VIEWER" and (_normalize_cell(row_values[idx["role"]]) or "").strip() and (_normalize_cell(row_values[idx["role"]]).strip().upper() not in ROLE_OPTIONS):
                row_log.note = "Неизвестная роль заменена на VIEWER"

            u, match_reason = find_existing_user(
                username=username,
                email=email or None,
                phone=phone,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name or None,
            )

            if not u and str(match_reason).startswith("conflict:"):
                duplicates += 1
                skipped += 1
                row_log.action = "DUPLICATE"
                row_log.note = f"Конфликт сопоставления: {match_reason.split(':', 1)[1]}"
                continue

            if not u:
                u = User(username=username)
                row_log.action = "CREATED"
                created += 1
                row_log.note = "Создан по импорту пользователей"
            else:
                row_log.action = "UPDATED"
                updated += 1
                _snapshot_user(row_log, u)
                row_log.note = f"Сопоставлен по: {match_reason}"
                if not u.username:
                    u.username = username

            u.role = role
            u.last_name = last_name
            u.first_name = first_name
            u.middle_name = middle_name or None
            u.phone = phone
            u.email = email or None
            u.is_active_user = True

            if u.employment_status in ["DISMISSED", "ARCHIVED"]:
                u.employment_status = "ACTIVE"
                u.archived_at = None
                u.dismissal_date = None
            else:
                u.employment_status = u.employment_status or "ACTIVE"

            if password:
                u.set_password(password)
            else:
                if not getattr(u, "password_hash", None):
                    u.set_password("Temp123")

            if not u.id:
                db.session.add(u)
                db.session.flush()

            _sync_user_roles_table(u)
            _sync_service_specialist(u)
            row_log.user_id = u.id
            row_log.password_was_changed = bool(password)

        session_log.created_count = created
        session_log.updated_count = updated
        session_log.skipped_count = skipped
        session_log.duplicate_count = duplicates

        try:
            db.session.commit()
        except IntegrityError as exc:
            db.session.rollback()
            flash(f"Импорт не завершён из-за конфликта уникальности: {exc}", "danger")
            return redirect(url_for("users.users_import"))
        flash(
            f"Импорт завершён. Создано: {created}, обновлено: {updated}, "
            f"пропущено: {skipped}, дублей в файле: {duplicates}",
            "success",
        )
        return redirect(url_for("users.users_list"))

    return render_template("users_import.html", role_labels=ROLE_LABELS)


@users_bp.route("/admin/users/imports/<int:session_id>/rollback", methods=["POST"])
@require_roles("ADMIN")
def users_import_rollback(session_id):
    session_log = UserImportSession.query.get_or_404(session_id)

    if session_log.reverted_at:
        flash("Этот импорт уже был отменён ранее.", "warning")
        return redirect(url_for("users.users_list"))

    try:
        for row in UserImportRow.query.filter_by(session_id=session_id).order_by(UserImportRow.row_num.desc()).all():
            user = User.query.get(row.user_id) if row.user_id else None

            if row.action == "CREATED":
                if user and user.username != "admin":
                    db.session.delete(user)

            elif row.action == "UPDATED" and user:
                user.role = row.previous_role or user.role
                user.last_name = row.previous_last_name
                user.first_name = row.previous_first_name
                user.middle_name = row.previous_middle_name
                user.phone = row.previous_phone
                user.email = row.previous_email
                user.employment_status = row.previous_employment_status or user.employment_status

                if row.previous_is_active_user is not None:
                    user.is_active_user = row.previous_is_active_user

                user.archived_at = row.previous_archived_at
                user.dismissal_date = row.previous_dismissal_date

        session_log.reverted_at = datetime.utcnow()
        session_log.reverted_by = getattr(current_user, "id", None)
        session_log.status = "ROLLED_BACK"

        db.session.commit()
        flash("Импорт пользователей отменён. Созданные записи удалены, изменённые восстановлены.", "success")

    except Exception as exc:
        db.session.rollback()
        flash(f"Не удалось откатить импорт: {exc}", "danger")

    return redirect(url_for("users.users_list"))


@users_bp.route("/admin/users/unmatched-staff")
@require_roles("ADMIN")
def unmatched_staff_registry():
    status = (request.args.get("status") or "new").upper()
    query = ServiceImportUnmatchedStaff.query
    if status != "ALL":
        query = query.filter_by(status="NEW" if status == "NEW" else status)
    rows = query.order_by(ServiceImportUnmatchedStaff.imported_at.desc(), ServiceImportUnmatchedStaff.id.desc()).all()
    users = User.query.order_by(User.last_name.asc(), User.first_name.asc(), User.username.asc()).all()
    return render_template("unmatched_staff_registry.html", rows=rows, users=users, status=status)


@users_bp.route("/admin/users/unmatched-staff/<int:row_id>/match", methods=["POST"])
@require_roles("ADMIN")
def unmatched_staff_match(row_id):
    row = ServiceImportUnmatchedStaff.query.get_or_404(row_id)
    user_id = request.form.get("user_id", type=int)
    user = User.query.get(user_id) if user_id else None
    if not user:
        flash("Выберите пользователя для сопоставления.", "danger")
        return redirect(url_for("users.unmatched_staff_registry"))
    row.matched_user_id = user.id
    row.matched_by_user_id = getattr(current_user, "id", None)
    row.matched_at = datetime.utcnow()
    row.status = "MATCHED"
    db.session.commit()
    flash("Несопоставленный сотрудник привязан к существующему пользователю.", "success")
    return redirect(url_for("users.unmatched_staff_registry"))


@users_bp.route("/admin/users/unmatched-staff/<int:row_id>/ignore", methods=["POST"])
@require_roles("ADMIN")
def unmatched_staff_ignore(row_id):
    row = ServiceImportUnmatchedStaff.query.get_or_404(row_id)
    row.status = "IGNORED"
    row.matched_by_user_id = getattr(current_user, "id", None)
    row.matched_at = datetime.utcnow()
    db.session.commit()
    flash("Строка помечена как пропущенная.", "success")
    return redirect(url_for("users.unmatched_staff_registry"))


@users_bp.route("/admin/users/duplicates")
@require_roles("ADMIN")
def users_duplicates():
    groups = potential_duplicate_groups()
    return render_template("user_duplicates.html", groups=groups, role_labels=ROLE_LABELS)


@users_bp.route("/admin/users/import-template")
@require_roles("ADMIN")
def users_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["username", "password", "role", "last_name", "first_name", "middle_name", "phone", "email"])
    ws.append(["ivanov_i", "Temp123", "Педагог-психолог", "Иванов", "Иван", "Иванович", "+7 999 000-00-00", "teacher@example.org"])
    ws.append(["petrova_a", "Temp123", "Учитель-логопед", "Петрова", "Анна", "", "+7 999 111-11-11", "curator@example.org"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="users_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )